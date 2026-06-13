#!/usr/bin/env python3
"""
Local/unit-style workbook validator for the Bonelord 469 incremental decode loop.

This intentionally does NOT evaluate Excel formulas (openpyxl doesn't). It validates
runner invariants based on workbook state + DP tokenization in code.
"""

from __future__ import annotations

import pathlib
import sys
from typing import Dict, List, Sequence, Set, Tuple

import openpyxl

STATUS_MODEL_CONVERGED = "MODEL_CONVERGED"
STATUS_PUZZLE_SOLVED_LEGACY = "PUZZLE_SOLVED"


def _die(msg: str) -> int:
    print(f"ERROR: {msg}", file=sys.stderr)
    return 1


def _load_flow_module() -> object:
    # Import the runner module by putting scripts/ on sys.path.
    scripts_dir = pathlib.Path(__file__).resolve().parent
    sys.path.insert(0, str(scripts_dir))
    import bonelord_flow_next_iteration as flow  # type: ignore

    return flow


def _read_flowrunlog_steps(flow: object, wb: openpyxl.Workbook, iter_num: int) -> Tuple[Set[int], Dict[int, Dict[str, object]]]:
    ws = wb["FlowRunLog"]
    header = flow.ws_find_header_row(  # type: ignore[attr-defined]
        ws,
        ["Iteration", "StepID", "UTC", "Result", "Summary", "ChangedBooksCount", "EvidenceAvg", "WeakFrac", "MicroFrac", "Notes"],
        max_scan=3,
    )
    c = flow.ws_headers(ws, header)  # type: ignore[attr-defined]

    steps: Set[int] = set()
    by_step: Dict[int, Dict[str, object]] = {}
    for r in range(header + 1, ws.max_row + 1):
        it = ws.cell(r, c["Iteration"]).value
        if it != iter_num:
            continue
        sid = ws.cell(r, c["StepID"]).value
        if sid is None:
            continue
        try:
            sid_i = int(sid)
        except Exception:
            continue
        steps.add(sid_i)
        row_payload = {
            "Result": ws.cell(r, c["Result"]).value,
            "Summary": ws.cell(r, c["Summary"]).value,
            "ChangedBooksCount": ws.cell(r, c["ChangedBooksCount"]).value,
            "Notes": ws.cell(r, c["Notes"]).value,
        }
        if sid_i not in by_step:
            by_step[sid_i] = row_payload
        else:
            prev_res = str(by_step[sid_i].get("Result") or "")
            next_res = str(row_payload.get("Result") or "")
            if prev_res != "FAILED" and next_res == "FAILED":
                by_step[sid_i] = row_payload
    return steps, by_step


def _parse_int_from_summary(prefix: str, summary: object) -> int:
    s = str(summary or "")
    if prefix not in s:
        return 0
    # Example: "Mechanical promotions approved: 0"
    try:
        tail = s.split(prefix, 1)[1]
        # take first integer-like token
        tok = tail.strip().split()[0].strip().rstrip(";")
        return int(tok)
    except Exception:
        return 0


def main(argv: Sequence[str]) -> int:
    if len(argv) != 2:
        print(f"Usage: {argv[0]} PATH_TO_WORKBOOK.xlsx", file=sys.stderr)
        return 2

    flow = _load_flow_module()
    path = argv[1]
    if not pathlib.Path(path).exists():
        return _die(f"Workbook not found: {path}")

    wb = openpyxl.load_workbook(path)

    store = flow.open_flow_store(wb, workbook_path=path)  # type: ignore[attr-defined]
    state = store["state_map"]
    settings_map = store["settings_map"]
    cur_iter = int(flow.flow_state_get(store, "CurrentIteration", 0) or 0)  # type: ignore[attr-defined]
    status = str(flow.flow_state_get(store, "Status", "") or "")  # type: ignore[attr-defined]
    print(f"Workbook iteration={cur_iter} status={status}")

    # Invariant 1: Coverage must remain strict.
    cov_ok, cov_bad = flow.books_coverage_strictplus_ok(wb)  # type: ignore[attr-defined]
    if not cov_ok:
        return _die(f"Coverage_StrictPlus_v108 not 1 for {cov_bad} books")

    # Invariant 2: GroundTruth live check must pass under current active token set.
    enforced_gt_ids, _enforced_n, _gt_policy_status = flow.resolve_enforced_groundtruth_ids(  # type: ignore[attr-defined]
        wb,
        iter_num=cur_iter,
        settings_map=settings_map,
    )
    _glossary_ws, glossary_map = flow.load_glossary(wb)  # type: ignore[attr-defined]
    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}
    ok_gt, bad_gt, bad_gt_all = flow.groundtruth_live_check(wb, active, enforced_crib_ids=enforced_gt_ids)  # type: ignore[attr-defined]
    if not ok_gt:
        ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad_gt[:12])
        return _die(f"GroundTruth live check failed: {len(bad_gt)} mismatches (CribID(s) {ids})")
    soft_gt = max(0, len(bad_gt_all) - len(bad_gt))

    # Invariant 2b: status semantics for new convergence model.
    if status in ("RESOLVED", STATUS_MODEL_CONVERGED, STATUS_PUZZLE_SOLVED_LEGACY) and soft_gt > 0:
        return _die(f"Status={status} requires zero soft GT mismatches, got {soft_gt}")
    if status == "SOFT_RESOLVED" and soft_gt <= 0:
        return _die("Status=SOFT_RESOLVED requires soft GT mismatches > 0")

    # Invariant 2c: FlowState GT counters should match live values when present.
    fs_bad_enf = flow.flow_state_get(store, "GTBadEnforcedCount", None)  # type: ignore[attr-defined]
    fs_bad_all = flow.flow_state_get(store, "GTBadAllCount", None)  # type: ignore[attr-defined]
    fs_soft = flow.flow_state_get(store, "GTSoftMismatchCount", None)  # type: ignore[attr-defined]
    if fs_bad_enf is not None and int(fs_bad_enf or 0) != len(bad_gt):
        return _die(f"FlowState GTBadEnforcedCount mismatch: state={fs_bad_enf} live={len(bad_gt)}")
    if fs_bad_all is not None and int(fs_bad_all or 0) != len(bad_gt_all):
        return _die(f"FlowState GTBadAllCount mismatch: state={fs_bad_all} live={len(bad_gt_all)}")
    if fs_soft is not None and int(fs_soft or 0) != soft_gt:
        return _die(f"FlowState GTSoftMismatchCount mismatch: state={fs_soft} live={soft_gt}")

    # Invariant 3: FlowRunLog must have the expected steps for the latest iteration.
    expected_steps = {
        10,
        12,
        20,
        25,
        27,
        28,
        29,
        30,
        40,
        50,
        55,
        60,
        70,
        75,
        80,
        82,
        85,
        86,
        87,
        90,
        91,
        109,
        112,
        92,
        93,
        94,
        95,
        96,
        97,
        98,
        99,
        101,
        110,
        102,
        103,
        104,
        111,
        105,
        107,
        108,
        113,
        114,
        106,
    }
    steps, by_step = _read_flowrunlog_steps(flow, wb, cur_iter)
    missing = sorted(expected_steps - steps)
    if missing:
        return _die(f"FlowRunLog missing steps for iter {cur_iter}: {missing}")

    # Invariant 4: Critical guardrail steps should not be FAILED.
    for sid in (12, 28, 29, 40):
        res = str(by_step.get(sid, {}).get("Result") or "")
        if res == "FAILED":
            return _die(f"FlowRunLog step {sid} is FAILED: {by_step.get(sid, {})}")

    # Invariant 5: If no decode-affecting write path fired, StrictPlus books_changed should remain 0/70.
    mech = _parse_int_from_summary("Mechanical promotions approved:", by_step.get(40, {}).get("Summary"))
    retext = _parse_int_from_summary("Semantic glossary retext applied:", by_step.get(96, {}).get("Summary"))
    en_retext = _parse_int_from_summary("English glossary retext applied:", by_step.get(99, {}).get("Summary"))
    anti_unk = int(flow.flow_state_get(store, "AntiHallucinationUNKApplied", 0) or 0)  # type: ignore[attr-defined]
    reverse_phrase = int(flow.flow_state_get(store, "ReversePhraseRetextApplied", 0) or 0)  # type: ignore[attr-defined]
    if mech == 0 and retext == 0 and en_retext == 0 and anti_unk == 0 and reverse_phrase == 0:
        changed = str(by_step.get(60, {}).get("ChangedBooksCount") or "")
        if changed and changed != "0/70":
            return _die(f"Expected ChangedBooksCount=0/70 when mech_promoted==0, got {changed!r}")

    print("OK: invariants satisfied")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
