#!/usr/bin/env python3
"""
Run Bonelord "next iteration" repeatedly until it becomes stale.

Stale definition (default):
- 2 consecutive iterations where:
  - mech_promoted == 0
  - and the "Evolution" deltas are all zero at 6-decimal precision
    (EvAvg/Weak/Micro/Single/Tokens), as recorded by the runner.

This script reads those stats from the workbook after each iteration, so it's stable even if
stdout formatting changes slightly.
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl

import bonelord_flow_next_iteration as flow


_EV_RE = re.compile(
    r"Before Ev=(?P<bev>-?\d+\.\d+), Weak=(?P<bw>-?\d+\.\d+), Micro=(?P<bm>-?\d+\.\d+), Single=(?P<bs>-?\d+\.\d+), Tokens=(?P<bt>\d+); "
    r"After Ev=(?P<aev>-?\d+\.\d+), Weak=(?P<aw>-?\d+\.\d+), Micro=(?P<am>-?\d+\.\d+), Single=(?P<as>-?\d+\.\d+), Tokens=(?P<at>\d+)"
)
_RETEXT_RE = re.compile(r"Semantic glossary retext applied:\s*(?P<n>\d+)")
_REV_EMIT_RE = re.compile(r"emitted_tokens=(?P<n>\d+)")
_REV_HITS_RE = re.compile(r"hits=(?P<n>\d+)")
_REVERT_RE = re.compile(r"Semantic reverts applied:\s*(?P<n>\d+)")
_EN_RE = re.compile(r"English layer: map_rows=(?P<map>\d+), repl=(?P<repl>\d+), books_changed=(?P<bc>\d+), master_changed=(?P<mc>\d+)")
_EN_GLOSS_RETEXT_RE = re.compile(r"English glossary retext applied:\s*(?P<n>\d+)")


def _db_path_for_workbook(path: Path) -> Path:
    return path.resolve().parent / "data" / "bonelord_workbook.sqlite"


def _sql_scalar_to_python(value: object) -> object:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return ""
    upper = text.upper()
    if upper == "TRUE":
        return True
    if upper == "FALSE":
        return False
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except Exception:
            return value
    if re.fullmatch(r"-?(?:\d+\.\d*|\.\d+)", text):
        try:
            return float(text)
        except Exception:
            return value
    return value


def _read_snapshot_kv(db_path: Path, table_name: str) -> Dict[str, object]:
    out: Dict[str, object] = {}
    conn = sqlite3.connect(str(db_path))
    try:
        row = conn.execute(
            "SELECT export_id FROM snapshot_refs WHERE name = 'canonical'"
        ).fetchone()
        if row is None:
            return out
        export_id = int(row[0])
        rows = conn.execute(
            f'SELECT key, value FROM "{table_name}" WHERE __export_id = ? ORDER BY __row_index',
            (export_id,),
        ).fetchall()
        for key, value in rows:
            if not isinstance(key, str) or not key.strip():
                continue
            out[key] = _sql_scalar_to_python(value)
        return out
    finally:
        conn.close()


def _read_iter_summary(wb: openpyxl.Workbook, iter_num: int) -> Dict[str, object]:
    name = f"Iter{iter_num}_Summary"
    if name not in wb.sheetnames:
        return {}
    ws = wb[name]
    header = flow.ws_find_header_row(ws, ["Metric", "Value"], max_scan=3)
    c = flow.ws_headers(ws, header)
    out: Dict[str, object] = {}
    for r in range(header + 1, ws.max_row + 1):
        k = ws.cell(r, c["Metric"]).value
        v = ws.cell(r, c["Value"]).value
        if not isinstance(k, str) or not k.strip():
            continue
        out[k.strip()] = v
    return out


def _read_flowrunlog_row(wb: openpyxl.Workbook, iter_num: int, step_id: int) -> Dict[str, object]:
    ws = wb["FlowRunLog"]
    header = flow.ws_find_header_row(
        ws,
        ["Iteration", "StepID", "UTC", "Result", "Summary", "ChangedBooksCount", "EvidenceAvg", "WeakFrac", "MicroFrac", "Notes"],
        max_scan=3,
    )
    c = flow.ws_headers(ws, header)
    for r in range(ws.max_row, header, -1):
        it = ws.cell(r, c["Iteration"]).value
        sid = ws.cell(r, c["StepID"]).value
        if it != iter_num:
            continue
        if sid is None or int(sid) != int(step_id):
            continue
        return {
            "Result": ws.cell(r, c["Result"]).value,
            "Summary": ws.cell(r, c["Summary"]).value,
            "Notes": ws.cell(r, c["Notes"]).value,
        }
    return {}


@dataclass(frozen=True)
class IterStats:
    iter_num: int
    status: str
    plateau_rung: int
    plateau_rung_changed: int
    mech_promoted: int
    semantic_retext: int
    semantic_revert: int
    english_retext: int
    autophrase_changed: int
    ctx_improve_streak: int
    ctx_avg_score: float
    code_map_rows: int
    code_fp_changed: int
    code_changed_rows: int
    code_overrides_total: int
    seq_matches: int
    seq_fp_changed: int
    sestina_cands: int
    sestina_best_score30: int
    sestina_fp_changed: int
    ev_avg: float
    weak: float
    micro: float
    single: float
    tokens: int
    d_ev: float
    d_weak: float
    d_micro: float
    d_single: float
    d_tokens: int
    reverse_emitted: int
    reverse_hits: int
    english_changed_rows: int


def _parse_step40_notes(notes: object) -> Optional[Tuple[float, float, float, float, int, float, float, float, float, int]]:
    m = _EV_RE.search(str(notes or ""))
    if not m:
        return None
    bev = float(m.group("bev"))
    bw = float(m.group("bw"))
    bm = float(m.group("bm"))
    bs = float(m.group("bs"))
    bt = int(m.group("bt"))
    aev = float(m.group("aev"))
    aw = float(m.group("aw"))
    am = float(m.group("am"))
    a_s = float(m.group("as"))
    at = int(m.group("at"))
    return bev, bw, bm, bs, bt, aev, aw, am, a_s, at


def _is_stale(stat: IterStats) -> bool:
    if stat.status == "BLOCKED":
        return False
    if stat.plateau_rung_changed != 0:
        return False
    if stat.mech_promoted != 0:
        return False
    if stat.semantic_retext != 0:
        return False
    if stat.semantic_revert != 0:
        return False
    if stat.english_retext != 0:
        return False
    if stat.autophrase_changed != 0:
        return False
    if stat.reverse_emitted != 0:
        return False
    if stat.reverse_hits != 0:
        return False
    if stat.english_changed_rows != 0:
        return False
    if stat.ctx_improve_streak != 0:
        return False
    if stat.code_fp_changed != 0:
        return False
    if stat.code_changed_rows != 0:
        return False
    if stat.seq_fp_changed != 0:
        return False
    if stat.sestina_fp_changed != 0:
        return False
    # We treat "stale" at the printed precision. This matches how the human loop is tracked.
    return (
        round(stat.d_ev, 6) == 0.0
        and round(stat.d_weak, 6) == 0.0
        and round(stat.d_micro, 6) == 0.0
        and round(stat.d_single, 6) == 0.0
        and stat.d_tokens == 0
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--max-iters", type=int, default=25)
    ap.add_argument("--stale-consecutive", type=int, default=2)
    args = ap.parse_args()

    path = args.workbook
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    stale_run = 0
    rows: list[IterStats] = []
    prev_plateau_rung: Optional[int] = None

    for _ in range(int(args.max_iters)):
        flow.run_next_iteration(str(path))

        wb = openpyxl.load_workbook(path, data_only=True)
        db_path = _db_path_for_workbook(path)
        state = _read_snapshot_kv(db_path, "sheet__flowstate")
        settings = _read_snapshot_kv(db_path, "sheet__flowsettings")
        it = int(state.get("CurrentIteration") or 0)
        status = str(state.get("Status") or "")
        try:
            plateau_rung = int(settings.get("PlateauLadder_Rung") or 0)
        except Exception:
            plateau_rung = 0
        plateau_rung_changed = 0
        if prev_plateau_rung is not None and int(plateau_rung) != int(prev_plateau_rung):
            plateau_rung_changed = 1
        prev_plateau_rung = int(plateau_rung)
        ctx_streak = int(state.get("ContextEnglishImproveStreak") or 0)
        try:
            ctx_score = float(state.get("ContextEnglishAvgScore") or 0.0)
        except Exception:
            ctx_score = 0.0
        code_map_rows = int(state.get("CodeWordMapCount") or 0)
        code_fp_changed = int(state.get("CodeWordMapFingerprintChanged") or 0)
        code_changed_rows = int(state.get("CodeAwareBooksChangedRows") or 0)
        code_overrides_total = int(state.get("CodeAwareOverridesTotal") or 0)
        seq_matches = int(state.get("SequenceMatchesCount") or 0)
        seq_fp_changed = int(state.get("SequenceMatchesFingerprintChanged") or 0)
        sestina_cands = int(state.get("SestinaCandidatesCount") or 0)
        sestina_best = int(state.get("SestinaBestScore30") or 0)
        sestina_fp_changed = int(state.get("SestinaFingerprintChanged") or 0)

        summ = _read_iter_summary(wb, it)
        mech = int(summ.get("Mechanical promotions accepted") or 0)
        ev = float(summ.get("Token-evidence avg (Books, length-weighted)") or 0.0)
        weak = float(summ.get("WEAK char frac (Books, length-weighted)") or 0.0)
        micro = float(summ.get("MICRO_MEDIUM char frac (Books, length-weighted)") or 0.0)

        step70 = _read_flowrunlog_row(wb, it, 70)
        single = 0.0
        m_sc = re.search(r"SingleCharFrac=(\d+\.\d+)", str(step70.get("Notes") or ""))
        if m_sc:
            single = float(m_sc.group(1))

        step40 = _read_flowrunlog_row(wb, it, 40)
        parsed = _parse_step40_notes(step40.get("Notes"))
        if parsed:
            bev, bw, bm, bs, bt, aev, aw, am, a_s, at = parsed
            tokens = at
            d_ev = aev - bev
            d_weak = aw - bw
            d_micro = am - bm
            d_single = a_s - bs
            d_tokens = at - bt
        else:
            tokens = 0
            d_ev = 0.0
            d_weak = 0.0
            d_micro = 0.0
            d_single = 0.0
            d_tokens = 0

        step96 = _read_flowrunlog_row(wb, it, 96)
        sem_retext = 0
        m_rt = _RETEXT_RE.search(str(step96.get("Summary") or ""))
        if m_rt:
            sem_retext = int(m_rt.group("n"))

        step97 = _read_flowrunlog_row(wb, it, 97)
        sem_revert = 0
        m_rv = _REVERT_RE.search(str(step97.get("Summary") or ""))
        if m_rv:
            sem_revert = int(m_rv.group("n"))

        step28 = _read_flowrunlog_row(wb, it, 28)
        rev_emitted = 0
        m_rev = _REV_EMIT_RE.search(str(step28.get("Summary") or ""))
        if m_rev:
            rev_emitted = int(m_rev.group("n"))
        rev_hits = 0
        m_hits = _REV_HITS_RE.search(str(step28.get("Summary") or ""))
        if m_hits:
            rev_hits = int(m_hits.group("n"))

        step27 = _read_flowrunlog_row(wb, it, 27)
        autophrase_changed = 1 if str(step27.get("Result") or "") == "CHANGED" else 0

        step99 = _read_flowrunlog_row(wb, it, 99)
        english_retext = 0
        m_er = _EN_GLOSS_RETEXT_RE.search(str(step99.get("Summary") or ""))
        if m_er:
            english_retext = int(m_er.group("n"))

        step98 = _read_flowrunlog_row(wb, it, 98)
        english_changed = 0
        m_en = _EN_RE.search(str(step98.get("Summary") or ""))
        if m_en:
            english_changed = int(m_en.group("bc")) + int(m_en.group("mc"))

        stat = IterStats(
            iter_num=it,
            status=status,
            plateau_rung=int(plateau_rung),
            plateau_rung_changed=int(plateau_rung_changed),
            mech_promoted=mech,
            semantic_retext=sem_retext,
            semantic_revert=sem_revert,
            english_retext=english_retext,
            autophrase_changed=autophrase_changed,
            ctx_improve_streak=ctx_streak,
            ctx_avg_score=ctx_score,
            code_map_rows=code_map_rows,
            code_fp_changed=code_fp_changed,
            code_changed_rows=code_changed_rows,
            code_overrides_total=code_overrides_total,
            seq_matches=seq_matches,
            seq_fp_changed=seq_fp_changed,
            sestina_cands=sestina_cands,
            sestina_best_score30=sestina_best,
            sestina_fp_changed=sestina_fp_changed,
            ev_avg=ev,
            weak=weak,
            micro=micro,
            single=single,
            tokens=tokens,
            d_ev=d_ev,
            d_weak=d_weak,
            d_micro=d_micro,
            d_single=d_single,
            d_tokens=d_tokens,
            reverse_emitted=rev_emitted,
            reverse_hits=rev_hits,
            english_changed_rows=english_changed,
        )
        rows.append(stat)

        if _is_stale(stat):
            stale_run += 1
        else:
            stale_run = 0
        if stale_run >= int(args.stale_consecutive):
            break

    # Print a compact table.
    print(
        "iter\tplateau\tpchg\tmech\tsem_retext\tsem_revert\ten_retext\tautophrase\tctx_streak\tctx_score\tcode_map\tcode_fpchg\tcode_changed\tcode_overrides\tseq_matches\tseq_fpchg\tsest_cands\tsest_best30\tsest_fpchg\trev_hits\trev_emit\ten_changed\tEvAvg\tWeak\tMicro\tSingle\tTokens\tdEv\tdWeak\tdMicro\tdSingle\tdTokens\tstatus"
    )
    for s in rows:
        print(
            f"{s.iter_num}\t{s.plateau_rung}\t{s.plateau_rung_changed}\t{s.mech_promoted}\t{s.semantic_retext}\t{s.semantic_revert}\t{s.english_retext}\t{s.autophrase_changed}\t{s.ctx_improve_streak}\t{s.ctx_avg_score:.6f}\t{s.code_map_rows}\t{s.code_fp_changed}\t{s.code_changed_rows}\t{s.code_overrides_total}\t{s.seq_matches}\t{s.seq_fp_changed}\t{s.sestina_cands}\t{s.sestina_best_score30}\t{s.sestina_fp_changed}\t{s.reverse_hits}\t{s.reverse_emitted}\t{s.english_changed_rows}\t{s.ev_avg:.6f}\t{s.weak:.6f}\t{s.micro:.6f}\t{s.single:.6f}\t{s.tokens}"
            f"\t{s.d_ev:+.6f}\t{s.d_weak:+.6f}\t{s.d_micro:+.6f}\t{s.d_single:+.6f}\t{s.d_tokens:+d}\t{s.status}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
