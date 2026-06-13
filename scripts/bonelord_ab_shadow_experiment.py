#!/usr/bin/env python3
"""
Run convergence A/B profiles in shadow workbooks (never mutates canonical workbook).

Profiles:
- control: current settings
- conservative: tighter/noise-reduced hints
- moderate: middle ground between control and conservative
"""

from __future__ import annotations

import argparse
import json
import pathlib
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from typing import Dict, Tuple

import openpyxl


SCRIPTS_DIR = pathlib.Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_DIR))
import bonelord_flow_next_iteration as flow  # type: ignore


PROFILES: Dict[str, Dict[str, object]] = {
    "control": {},
    "conservative": {
        "SequenceHints_Boost": 130,
        "SequenceMatch_MaxCandidates": 2000,
        "SequenceMatch_CandidateMaxBookFreq": 6,
        "SequenceMatch_ContextWindow": 3,
        "ContextEnglishMap_MinTopShare": 0.85,
        "SequenceWordHints_StopwordRatio": 0.60,
        "SequenceMatch_ContextMinOverlap": 1,
    },
    "moderate": {
        "SequenceHints_Boost": 180,
        "SequenceMatch_MaxCandidates": 3000,
        "SequenceMatch_CandidateMaxBookFreq": 9,
        "SequenceMatch_ContextWindow": 4,
        "ContextEnglishMap_MinTopShare": 0.80,
        "SequenceWordHints_StopwordRatio": 0.70,
        "SequenceMatch_ContextMinOverlap": 1,
    },
}


def _read_step80_metrics(wb: openpyxl.Workbook, iter_num: int) -> Tuple[float, float, float]:
    ws = wb["FlowRunLog"]
    header = flow.ws_find_header_row(  # type: ignore[attr-defined]
        ws,
        ["Iteration", "StepID", "EvidenceAvg", "WeakFrac", "MicroFrac"],
        max_scan=3,
    )
    c = flow.ws_headers(ws, header)  # type: ignore[attr-defined]
    for r in range(ws.max_row, header, -1):
        it = ws.cell(r, c["Iteration"]).value
        sid = ws.cell(r, c["StepID"]).value
        if it != iter_num or sid != 80:
            continue
        ev = float(ws.cell(r, c["EvidenceAvg"]).value or 0.0)
        weak = float(ws.cell(r, c["WeakFrac"]).value or 0.0)
        micro = float(ws.cell(r, c["MicroFrac"]).value or 0.0)
        return ev, weak, micro
    return 0.0, 0.0, 0.0


def _collect_metrics(path: pathlib.Path) -> Dict[str, object]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws_state, state = flow.load_flow_state(wb)  # type: ignore[attr-defined]
    del ws_state
    cur_iter = int(state.get("CurrentIteration", (None, None, 0))[2] or 0)
    status = str(state.get("Status", (None, None, ""))[2] or "")
    gt_bad_enf = int(state.get("GTBadEnforcedCount", (None, None, 0))[2] or 0)
    gt_soft = int(state.get("GTSoftMismatchCount", (None, None, 0))[2] or 0)
    promo_skips = int(state.get("PromotionSkipCount", (None, None, 0))[2] or 0)
    ctx_avg = float(state.get("ContextEnglishAvgScore", (None, None, 0.0))[2] or 0.0)
    seq_matches = int(state.get("SequenceMatchesCount", (None, None, 0))[2] or 0)
    cov_ok, cov_bad = flow.books_coverage_strictplus_ok(wb)  # type: ignore[attr-defined]
    ev, weak, micro = _read_step80_metrics(wb, cur_iter)
    wb.close()
    return {
        "iteration": cur_iter,
        "status": status,
        "coverage_ok": bool(cov_ok),
        "coverage_bad_books": int(cov_bad),
        "gt_bad_enforced": gt_bad_enf,
        "gt_soft": gt_soft,
        "promotion_skips": promo_skips,
        "ctx_avg": ctx_avg,
        "seq_matches": seq_matches,
        "evidence_avg": ev,
        "weak": weak,
        "micro": micro,
    }


def _apply_profile(path: pathlib.Path, name: str) -> None:
    if name not in PROFILES:
        raise SystemExit(f"Unknown profile: {name}")
    if not PROFILES[name]:
        return
    wb = openpyxl.load_workbook(path)
    ws_settings, _settings_map = flow.load_flow_settings(wb)  # type: ignore[attr-defined]
    for k, v in PROFILES[name].items():
        flow.set_setting(ws_settings, k, v, note=f"shadow-ab profile={name}")  # type: ignore[attr-defined]
    wb.save(path)
    wb.close()


def _run_iterations(path: pathlib.Path, iterations: int) -> None:
    runner = SCRIPTS_DIR / "bonelord_flow_next_iteration.py"
    for _ in range(iterations):
        subprocess.run([sys.executable, str(runner), str(path)], check=True)


def main() -> int:
    ap = argparse.ArgumentParser(description="Run A/B convergence profiles on shadow workbooks.")
    ap.add_argument("workbook", help="Path to canonical workbook (.xlsx).")
    ap.add_argument("--iterations", type=int, default=3, help="Iterations per profile (default: 3).")
    ap.add_argument(
        "--out-dir",
        default="tmp/spreadsheets",
        help="Directory where shadow workbooks and report are written (default: tmp/spreadsheets).",
    )
    args = ap.parse_args()

    canonical = pathlib.Path(args.workbook).resolve()
    if not canonical.exists():
        raise SystemExit(f"Workbook not found: {canonical}")
    out_dir = pathlib.Path(args.out_dir).resolve()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%SZ")
    run_dir = out_dir / f"shadow_ab_{stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    baseline = _collect_metrics(canonical)
    results: Dict[str, Dict[str, object]] = {}

    for profile in ("control", "conservative", "moderate"):
        shadow = run_dir / f"{profile}.xlsx"
        shutil.copy2(canonical, shadow)
        _apply_profile(shadow, profile)
        _run_iterations(shadow, int(args.iterations))
        metrics = _collect_metrics(shadow)

        hard_gates = bool(metrics["coverage_ok"]) and int(metrics["gt_bad_enforced"]) == 0
        no_regression = (
            float(metrics["weak"]) <= float(baseline["weak"])
            and float(metrics["micro"]) <= float(baseline["micro"])
            and float(metrics["evidence_avg"]) >= float(baseline["evidence_avg"])
        )
        noise_penalty = max(0, int(metrics["gt_soft"]) - int(baseline["gt_soft"])) * 0.1 + int(metrics["promotion_skips"]) * 0.001
        score = float(metrics["ctx_avg"]) + 0.02 * int(metrics["seq_matches"]) - noise_penalty
        eligible = bool(hard_gates and no_regression)
        if not eligible:
            score -= 1000.0

        metrics["hard_gates"] = hard_gates
        metrics["no_regression"] = no_regression
        metrics["eligible"] = eligible
        metrics["score"] = round(score, 6)
        results[profile] = metrics

    ranked = sorted(results.items(), key=lambda kv: float(kv[1]["score"]), reverse=True)
    winner = ranked[0][0] if ranked else "control"
    report = {
        "canonical": str(canonical),
        "iterations_per_profile": int(args.iterations),
        "baseline": baseline,
        "profiles": results,
        "winner": winner,
    }
    report_path = run_dir / "ab_report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(f"Shadow A/B completed. winner={winner}")
    print(f"Report: {report_path}")
    for name, data in ranked:
        print(
            f"- {name}: score={data['score']} eligible={int(bool(data['eligible']))} "
            f"status={data['status']} ctx={data['ctx_avg']:.6f} seq={data['seq_matches']} "
            f"ev={data['evidence_avg']:.6f} weak={data['weak']:.6f} micro={data['micro']:.6f} "
            f"soft={data['gt_soft']} skips={data['promotion_skips']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

