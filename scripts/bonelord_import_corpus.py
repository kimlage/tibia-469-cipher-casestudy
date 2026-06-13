#!/usr/bin/env python3
"""
Import local plaintext lines into the workbook's LoreCorpus_User sheet.

This is the safe/no-network hook for adding official dialogues or other reference texts
without hardcoding them into the runner.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Set, Tuple

import openpyxl


def _ws_find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, required: Set[str], max_scan: int = 10) -> int:
    for r in range(1, max_scan + 1):
        got = {ws.cell(r, c).value for c in range(1, ws.max_column + 1) if isinstance(ws.cell(r, c).value, str)}
        if required.issubset(got):
            return r
    raise SystemExit(f"Could not find header row in sheet {ws.title} with required columns: {sorted(required)}")


def _ws_headers(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            out[v.strip()] = c
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("input", type=Path, help="Plaintext file: one phrase/line per row.")
    ap.add_argument("--corpus-id", required=True)
    ap.add_argument("--lang", default="EN")
    ap.add_argument("--license", default="user-provided")
    ap.add_argument("--source", default="")
    ap.add_argument("--notes", default="")
    args = ap.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")
    if not args.input.exists():
        raise SystemExit(f"Input file not found: {args.input}")

    wb = openpyxl.load_workbook(args.workbook)
    headers = ["CorpusID", "Lang", "License", "Source", "LineID", "Text", "AddedIter", "Notes"]
    if "LoreCorpus_User" not in wb.sheetnames:
        ws = wb.create_sheet("LoreCorpus_User")
        for i, h in enumerate(headers, start=1):
            ws.cell(1, i).value = h
    ws = wb["LoreCorpus_User"]

    hrow = _ws_find_header_row(ws, {"CorpusID", "LineID", "Text"}, max_scan=3)
    c = _ws_headers(ws, hrow)

    # Read current iteration (optional) for AddedIter.
    added_iter = 0
    if "FlowState" in wb.sheetnames:
        ws_state = wb["FlowState"]
        for r in range(2, ws_state.max_row + 1):
            k = ws_state.cell(r, 1).value
            if str(k or "").strip() == "CurrentIteration":
                try:
                    added_iter = int(ws_state.cell(r, 2).value or 0)
                except Exception:
                    added_iter = 0
                break

    cid = str(args.corpus_id).strip()
    if not cid:
        raise SystemExit("--corpus-id cannot be empty")

    # Track existing lines for this corpus (dedupe).
    existing: Set[Tuple[str, str]] = set()
    max_line = 0
    for r in range(hrow + 1, ws.max_row + 1):
        cur_cid = ws.cell(r, c["CorpusID"]).value
        if str(cur_cid or "").strip() != cid:
            continue
        try:
            lid = int(ws.cell(r, c["LineID"]).value or 0)
        except Exception:
            lid = 0
        max_line = max(max_line, lid)
        txt = str(ws.cell(r, c["Text"]).value or "").strip()
        if txt:
            existing.add((cid, txt.lower()))

    # Append new lines.
    appended = 0
    line_id = max_line
    raw = args.input.read_text(encoding="utf-8", errors="replace").splitlines()
    for s in raw:
        txt = str(s or "").strip()
        if not txt:
            continue
        key = (cid, txt.lower())
        if key in existing:
            continue
        line_id += 1
        r = ws.max_row + 1
        ws.cell(r, c["CorpusID"]).value = cid
        ws.cell(r, c.get("Lang", c["CorpusID"] + 1)).value = str(args.lang).strip().upper() or "EN"
        ws.cell(r, c.get("License", c["CorpusID"] + 2)).value = str(args.license).strip()
        ws.cell(r, c.get("Source", c["CorpusID"] + 3)).value = str(args.source).strip()
        ws.cell(r, c["LineID"]).value = line_id
        ws.cell(r, c["Text"]).value = txt
        ws.cell(r, c.get("AddedIter", c["Text"] + 1)).value = added_iter
        ws.cell(r, c.get("Notes", c["Text"] + 2)).value = str(args.notes).strip()
        existing.add(key)
        appended += 1

    wb.save(args.workbook)
    print(f"Appended {appended} lines to LoreCorpus_User (CorpusID={cid!r}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

