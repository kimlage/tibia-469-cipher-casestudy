#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import load_workbook


DEFAULT_ARTIFACT = "./bonelord_469_iter129.xlsx"
DEFAULT_DB = "./data/bonelord_operational.sqlite"
HEADER_SCAN_ROWS = 10
DEFAULT_SHEETS = (
    "FlowState",
    "FlowSettings",
    "Books",
    "Contigs",
    "Glossary",
    "ExternalRefs_v115",
    "ExternalRoundTrip_Auto",
    "CodeStreamCandidates_v119",
    "DigitInsertionDP_v117",
    "BooksDigitModel_v118",
    "DigitOmissionStats_v118",
    "ExternalGroupDecode_v120",
    "ExternalGroupCandidates_v120",
    "ExternalGroundTruthCheck_v120",
    "ExternalValidation_v128",
    "ExternalValidation_v129",
    "DigitCodeMap_Auto",
    "DigitLetterCodes_Auto",
    "DigitCodeContext_Auto",
    "CodeWordMap_Auto",
    "ExternalCommunitySources_v472",
    "ExternalSourceDigitHits_v472",
    "ExternalRefCandidates_v472",
    "GroundTruth",
    "GroundTruthPolicy_Auto",
    "AnchorCribs_Auto",
    "AnchorOccurrences_Auto",
    "SuperAnchors_Auto",
    "CandidatePromotions",
    "FlowRunLog",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a compact operational SQLite DB from the canonical legacy artifact")
    parser.add_argument("--artifact", default=DEFAULT_ARTIFACT)
    parser.add_argument("--db", default=DEFAULT_DB)
    parser.add_argument("--replace", action="store_true")
    parser.add_argument("--sheet", action="append", default=None, help="Sheet to include. Can be repeated.")
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_identifier(text: object, fallback: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "_", str(text or "").strip().lower())
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        value = fallback
    if value[0].isdigit():
        value = f"c_{value}"
    return value


def normalize_headers(values: Sequence[object]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for idx, raw in enumerate(values, start=1):
        name = normalize_identifier(raw, f"col_{idx}")
        seen[name] = seen.get(name, 0) + 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        out.append(name)
    return out


def to_text(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def detect_header(ws) -> tuple[int, List[str], List[str]]:
    best_row = 1
    best_values: List[object] = []
    best_score = -1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, HEADER_SCAN_ROWS), values_only=True), start=1):
        values = list(row)
        score = sum(1 for value in values if value is not None and str(value).strip())
        if score > best_score:
            best_row = row_idx
            best_values = values
            best_score = score
    raw_names = ["" if value is None else str(value).strip() for value in best_values]
    return best_row, raw_names, normalize_headers(raw_names)


def ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;
        PRAGMA synchronous=NORMAL;

        CREATE TABLE IF NOT EXISTS exports (
            export_id INTEGER PRIMARY KEY AUTOINCREMENT,
            artifact_path TEXT NOT NULL,
            artifact_name TEXT NOT NULL,
            artifact_sha256 TEXT NOT NULL,
            artifact_size INTEGER NOT NULL,
            artifact_mtime REAL NOT NULL,
            artifact_kind TEXT NOT NULL,
            exported_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS snapshot_refs (
            name TEXT PRIMARY KEY,
            export_id INTEGER NOT NULL,
            artifact_path TEXT,
            artifact_sha256 TEXT,
            note TEXT,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS sheets (
            export_id INTEGER NOT NULL,
            sheet_index INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            max_row INTEGER NOT NULL,
            max_col INTEGER NOT NULL,
            header_row INTEGER NOT NULL,
            PRIMARY KEY (export_id, sheet_name)
        );

        CREATE TABLE IF NOT EXISTS row_json (
            export_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            row_json TEXT NOT NULL,
            PRIMARY KEY (export_id, sheet_name, row_index)
        );
        """
    )


def ensure_sheet_table(conn: sqlite3.Connection, sheet_name: str, columns: Sequence[str]) -> str:
    table = f"sheet__{normalize_identifier(sheet_name, 'sheet')}"
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS "{table}" (
            __export_id INTEGER NOT NULL,
            __row_index INTEGER NOT NULL,
            __sheet_name TEXT NOT NULL,
            PRIMARY KEY (__export_id, __row_index)
        )
        """
    )
    existing = {row[1] for row in conn.execute(f'PRAGMA table_info("{table}")')}
    for col in columns:
        if col not in existing:
            conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "{col}" TEXT')
    return table


def insert_export(conn: sqlite3.Connection, artifact: Path) -> int:
    stat = artifact.stat()
    digest = sha256_file(artifact)
    cur = conn.execute(
        """
        INSERT INTO exports (
            artifact_path, artifact_name, artifact_sha256, artifact_size,
            artifact_mtime, artifact_kind, exported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(artifact),
            artifact.name,
            digest,
            stat.st_size,
            stat.st_mtime,
            "legacy_xlsx_compact",
            datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        ),
    )
    return int(cur.lastrowid)


def export_sheet(conn: sqlite3.Connection, export_id: int, sheet_index: int, ws) -> int:
    header_row, raw_names, normalized = detect_header(ws)
    conn.execute(
        "INSERT INTO sheets (export_id, sheet_index, sheet_name, max_row, max_col, header_row) VALUES (?, ?, ?, ?, ?, ?)",
        (export_id, sheet_index, ws.title, ws.max_row, ws.max_column, header_row),
    )
    table = ensure_sheet_table(conn, ws.title, normalized)
    rows_written = 0
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_index <= header_row:
            continue
        payload: Dict[str, Optional[str]] = {}
        for idx, col in enumerate(normalized):
            value = row[idx] if idx < len(row) else None
            payload[col] = to_text(value)
        if not any(value is not None and str(value).strip() for value in payload.values()):
            continue
        conn.execute(
            "INSERT INTO row_json (export_id, sheet_name, row_index, row_json) VALUES (?, ?, ?, ?)",
            (export_id, ws.title, row_index, json.dumps(payload, ensure_ascii=True)),
        )
        cols = ['__export_id', '__row_index', '__sheet_name'] + [f'"{col}"' for col in normalized]
        vals = [export_id, row_index, ws.title] + [payload.get(col) for col in normalized]
        conn.execute(
            f'INSERT INTO "{table}" ({", ".join(cols)}) VALUES ({", ".join(["?"] * len(vals))})',
            vals,
        )
        rows_written += 1
    return rows_written


def main() -> int:
    args = parse_args()
    artifact = Path(args.artifact).resolve()
    db_path = Path(args.db).resolve()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if args.replace and db_path.exists():
        db_path.unlink()
    sheets = tuple(args.sheet) if args.sheet else DEFAULT_SHEETS

    wb = load_workbook(artifact, data_only=True, read_only=True)
    conn = sqlite3.connect(str(db_path))
    try:
        ensure_schema(conn)
        export_id = insert_export(conn, artifact)
        rows_by_sheet: Dict[str, int] = {}
        for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
            if sheet_name not in sheets:
                continue
            rows_by_sheet[sheet_name] = export_sheet(conn, export_id, sheet_index, wb[sheet_name])
            conn.commit()
        digest = sha256_file(artifact)
        now = datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
        conn.execute(
            """
            INSERT INTO snapshot_refs (name, export_id, artifact_path, artifact_sha256, note, updated_at)
            VALUES ('canonical', ?, ?, ?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                export_id = excluded.export_id,
                artifact_path = excluded.artifact_path,
                artifact_sha256 = excluded.artifact_sha256,
                note = excluded.note,
                updated_at = excluded.updated_at
            """,
            (export_id, str(artifact), digest, "compact operational snapshot", now),
        )
        conn.commit()
        print(json.dumps({"db": str(db_path), "export_id": export_id, "rows_by_sheet": rows_by_sheet}, ensure_ascii=True, indent=2))
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
