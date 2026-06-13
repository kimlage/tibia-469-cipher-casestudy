#!/usr/bin/env python3
"""
Fetch a Tibia-derived text dataset from the internet and store a *signature word index*
inside the workbook (no full text import).

This is intended to expand semantic alignment safely without embedding large copyrighted corpora
into the XLSX. The runner merges this sheet into LoreAlignment_Auto.
"""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Set, Tuple

import openpyxl
import requests

import bonelord_flow_next_iteration as flow


WORD_RE = re.compile(r"[A-Za-z']+")


def _utc_now() -> str:
    return datetime.now(timezone.utc).strftime(flow.ISO_UTC_FMT)


def _iter_words(text: str) -> Iterable[str]:
    for m in WORD_RE.finditer(str(text or "")):
        w = m.group(0)
        if w:
            yield w


def _target_signatures_from_workbook(wb: openpyxl.Workbook) -> Set[str]:
    ws = wb["Glossary"]
    h = flow.ws_find_header_row(ws, ["Token", "Translation", "Use_StrictPlus_v108", "TokenType"], max_scan=3)
    c = flow.ws_headers(ws, h)
    out: Set[str] = set()

    for r in range(h + 1, ws.max_row + 1):
        tok = ws.cell(r, c["Token"]).value
        if not isinstance(tok, str) or not tok:
            continue
        if "*" in tok:
            continue
        if not ws.cell(r, c["Use_StrictPlus_v108"]).value:
            continue
        tr = ws.cell(r, c["Translation"]).value
        if tr is None or str(tr).strip() == "":
            continue
        ttype = str(ws.cell(r, c["TokenType"]).value or "")
        if ttype in ("marker", "macro"):
            continue
        sig = flow._token_signature(tok)  # type: ignore[attr-defined]
        if sig:
            out.add(sig)
    return out


def _fetch_json(url: str, timeout_s: int) -> object:
    resp = requests.get(url, timeout=timeout_s)
    resp.raise_for_status()
    return resp.json()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--npc-url", default="https://resources.talesoftibia.com/data/npcs/npc_transcript_database.json")
    ap.add_argument("--book-url", default="https://resources.talesoftibia.com/data/books/book_database.json")
    ap.add_argument("--timeout", type=int, default=60)
    ap.add_argument("--max-words-per-sig", type=int, default=80)
    args = ap.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    wb = openpyxl.load_workbook(args.workbook)

    # Canon flags from FlowSettings (keep consistent with runner).
    ws_settings, settings_map = flow.load_flow_settings(wb)
    drop_final_e = flow.parse_bool(flow.get_setting(settings_map, "Lore_Canon_DropFinalE", False), False)
    drop_all_h = flow.parse_bool(flow.get_setting(settings_map, "Lore_Canon_DropAllH", False), False)
    drop_all_o = flow.parse_bool(flow.get_setting(settings_map, "Lore_Canon_DropAllO", False), False)

    target_sigs = _target_signatures_from_workbook(wb)
    if not target_sigs:
        raise SystemExit("No target token signatures found (Glossary active leaf tokens).")

    sig_word_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sig_corpora: Dict[str, Set[str]] = defaultdict(set)

    # NPC transcripts
    npc_obj = _fetch_json(str(args.npc_url), timeout_s=int(args.timeout))
    if isinstance(npc_obj, list):
        for item in npc_obj:
            if not isinstance(item, dict):
                continue
            conv = item.get("conversation")
            if not isinstance(conv, list):
                continue
            for turn in conv:
                if not isinstance(turn, dict):
                    continue
                answers = turn.get("answer")
                if isinstance(answers, str):
                    answers = [answers]
                if not isinstance(answers, list):
                    continue
                for a in answers:
                    if not isinstance(a, str) or not a.strip():
                        continue
                    for raw in _iter_words(a):
                        surface = re.sub(r"[^a-z']", "", raw.lower())
                        if not surface:
                            continue
                        canon = flow._lore_canon_word(  # type: ignore[attr-defined]
                            surface, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o
                        )
                        if not canon:
                            continue
                        sig = flow._lore_signature(canon)  # type: ignore[attr-defined]
                        if sig in target_sigs:
                            sig_word_counts[sig][surface] += 1
                            sig_corpora[sig].add("TIBIA_NPC")

    # Books
    book_obj = _fetch_json(str(args.book_url), timeout_s=int(args.timeout))
    if isinstance(book_obj, list):
        for item in book_obj:
            if not isinstance(item, dict):
                continue
            text = item.get("text")
            if not isinstance(text, str) or not text.strip():
                continue
            for raw in _iter_words(text):
                surface = re.sub(r"[^a-z']", "", raw.lower())
                if not surface:
                    continue
                canon = flow._lore_canon_word(  # type: ignore[attr-defined]
                    surface, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o
                )
                if not canon:
                    continue
                sig = flow._lore_signature(canon)  # type: ignore[attr-defined]
                if sig in target_sigs:
                    sig_word_counts[sig][surface] += 1
                    sig_corpora[sig].add("TIBIA_BOOK")

    # Write sheet
    sheet_name = "LoreSigIndex_Tibia_Auto"
    ws = flow.ensure_sheet(wb, sheet_name, ["Sig", "Word", "Count", "CorpusID", "FetchedUTC", "Source"])  # type: ignore[attr-defined]
    h = flow.ws_find_header_row(ws, ["Sig", "Word", "Count"], max_scan=3)
    c = flow.ws_headers(ws, h)
    if ws.max_row > h:
        ws.delete_rows(h + 1, ws.max_row - h)

    fetched = _utc_now()
    source = "resources.talesoftibia.com (derived word signature index only; no full text stored)"

    rr = h + 1
    max_words_per_sig = int(args.max_words_per_sig)
    for sig in sorted(sig_word_counts.keys()):
        wc = sig_word_counts[sig]
        corp = ",".join(sorted(sig_corpora.get(sig) or []))
        for word, cnt in sorted(wc.items(), key=lambda kv: (-kv[1], kv[0]))[:max_words_per_sig]:
            ws.cell(rr, c["Sig"]).value = sig
            ws.cell(rr, c["Word"]).value = word
            ws.cell(rr, c["Count"]).value = int(cnt)
            if "CorpusID" in c:
                ws.cell(rr, c["CorpusID"]).value = corp
            if "FetchedUTC" in c:
                ws.cell(rr, c["FetchedUTC"]).value = fetched
            if "Source" in c:
                ws.cell(rr, c["Source"]).value = source
            rr += 1

    flow.upsert_sheet_index_entry(wb, sheet_name, "Tibia-derived signature index for semantic alignment (no full text).")  # type: ignore[attr-defined]

    wb.save(args.workbook)
    print(f"Wrote {rr - (h + 1)} rows to {sheet_name}. target_sigs={len(target_sigs)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

