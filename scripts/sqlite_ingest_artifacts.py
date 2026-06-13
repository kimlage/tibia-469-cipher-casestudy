#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sqlite3
from pathlib import Path
from typing import Iterable, List, Set

from openpyxl import load_workbook

from export_workbook_to_sqlite import ensure_schema, export_sheet, existing_export_id, insert_export, sha256_file


DEFAULT_DB = "./data/bonelord_workbook.sqlite"
DEFAULT_ROOT = "."


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest legacy spreadsheet artifacts into SQLite")
    parser.add_argument("path", nargs="*", default=[DEFAULT_ROOT], help="File or directory roots to scan")
    parser.add_argument("--db", default=DEFAULT_DB, help="SQLite DB path")
    parser.add_argument("--pattern", default="*.xlsx", help="Artifact filename pattern")
    parser.add_argument("--replace-db", action="store_true", help="Delete the destination DB first")
    parser.add_argument("--limit", type=int, default=0, help="Optional cap on discovered files")
    return parser.parse_args()


def discover_artifacts(paths: Iterable[str], pattern: str, limit: int) -> List[Path]:
    out: List[Path] = []
    seen: Set[Path] = set()
    for raw in paths:
        root = Path(raw).resolve()
        if root.is_file():
            candidates = [root] if root.match(pattern) else []
        elif root.is_dir():
            candidates = sorted(root.rglob(pattern))
        else:
            candidates = []
        for path in candidates:
            if path in seen:
                continue
            seen.add(path)
            out.append(path)
            if limit > 0 and len(out) >= limit:
                return out
    return out


def ingest_one(conn: sqlite3.Connection, artifact_path: Path) -> dict:
    try:
        digest = sha256_file(artifact_path)
        already = existing_export_id(conn, digest)
        if already is not None:
            return {
                "artifact_path": str(artifact_path),
                "export_id": already,
                "status": "skipped_duplicate",
            }

        export_id = insert_export(conn, artifact_path)
        raw_wb = load_workbook(artifact_path, data_only=False, read_only=True)
        cached_wb = load_workbook(artifact_path, data_only=True, read_only=True)

        total_cells = 0
        total_rows = 0
        for sheet_index, sheet_name in enumerate(raw_wb.sheetnames, start=1):
            raw_ws = raw_wb[sheet_name]
            cached_ws = cached_wb[sheet_name]
            cell_rows, data_rows = export_sheet(conn, export_id, sheet_index, raw_ws, cached_ws)
            total_cells += cell_rows
            total_rows += data_rows
            conn.commit()

        return {
            "artifact_path": str(artifact_path),
            "export_id": export_id,
            "status": "ingested",
            "sheet_count": len(raw_wb.sheetnames),
            "data_rows": total_rows,
            "cell_rows": total_cells,
        }
    except Exception as exc:
        conn.rollback()
        return {
            "artifact_path": str(artifact_path),
            "status": "error",
            "error": f"{type(exc).__name__}: {exc}",
        }


def main() -> int:
    args = parse_args()
    db_path = Path(args.db).resolve()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if args.replace_db and db_path.exists():
        db_path.unlink()

    artifacts = discover_artifacts(args.path, args.pattern, args.limit)
    conn = sqlite3.connect(str(db_path))
    try:
        ensure_schema(conn)
        results = [ingest_one(conn, artifact) for artifact in artifacts]
        print(
            json.dumps(
                {
                    "db_path": str(db_path),
                    "artifact_count": len(artifacts),
                    "ingested": sum(1 for item in results if item["status"] == "ingested"),
                    "skipped_duplicate": sum(1 for item in results if item["status"] == "skipped_duplicate"),
                    "results": results,
                },
                ensure_ascii=True,
                indent=2,
            )
        )
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
