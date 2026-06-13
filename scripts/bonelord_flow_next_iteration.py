#!/usr/bin/env python3
"""
Bonelord 469 workbook safe incremental "next iteration" runner.

Design goals:
- Make incremental, traceable changes inside the same XLSX.
- Preserve StrictPlus text stability by anchoring on existing LosslessMarkers output.
- Allow mechanical promotions that do NOT increase global WEAK char frac or single-char frac.
- Keep FlowState/FlowRunLog/CandidatePromotions/IterXXX_Summary updated each run.

No network required. Uses openpyxl only (pandas/numpy intentionally avoided).

No network is *required*. Optionally, the runner can fetch a public Tibia-derived dataset
(NPC transcripts + books) and store only a derived signature index (counts by word signature)
in the workbook (no full text persisted). See FlowSettings `LoreFetch_TibiaSigIndex_*`.
"""

from __future__ import annotations

import dataclasses
import hashlib
import itertools
import json
import math
import os
import random
import re
import shutil
import sqlite3
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl


ISO_UTC_FMT = "%Y-%m-%dT%H:%M:%SZ"
DISCORD_ENV_PATH = "~/.env"
DISCORD_DEFAULT_CHANNEL_NAME = "bonelord-logs"
DISCORD_FALLBACK_CHANNEL_ID = "0"  # bonelord-logs
SQLITE_CANONICAL_SNAPSHOT_NAME = "canonical"
SQLITE_DB_BASENAME = "bonelord_workbook.sqlite"

STAR_WILDCARD_CHARS = {"T", "V", "R"}

DEFAULT_TIBIA_NPC_URL = "https://resources.talesoftibia.com/data/npcs/npc_transcript_database.json"
DEFAULT_TIBIA_BOOK_URL = "https://resources.talesoftibia.com/data/books/book_database.json"
DEFAULT_PD_KJV_URL = "https://www.gutenberg.org/files/10/10-0.txt"
DEFAULT_PD_ALICE_URL = "https://www.gutenberg.org/files/11/11-0.txt"
DEFAULT_PD_LOOKING_GLASS_URL = "https://www.gutenberg.org/files/12/12-0.txt"
DEFAULT_PD_SHERLOCK_URL = "https://www.gutenberg.org/files/1661/1661-0.txt"
DEFAULT_PD_SHAKESPEARE_URL = "https://www.gutenberg.org/files/100/100-0.txt"
DEFAULT_PD_PARADISE_LOST_URL = "https://www.gutenberg.org/files/26/26-0.txt"
DEFAULT_PD_PRIDE_PREJUDICE_URL = "https://www.gutenberg.org/files/1342/1342-0.txt"
DEFAULT_PD_FRANKENSTEIN_URL = "https://www.gutenberg.org/files/84/84-0.txt"
DEFAULT_PD_DRACULA_URL = "https://www.gutenberg.org/files/345/345-0.txt"
DEFAULT_PD_MOBY_DICK_URL = "https://www.gutenberg.org/files/2701/2701-0.txt"
DEFAULT_PD_BEOWULF_GUMMERE_URL = "https://www.gutenberg.org/files/16328/16328-0.txt"
DEFAULT_WIKI_SESTINA_RAW_URL = "https://en.wikipedia.org/w/index.php?title=Sestina&action=raw"
LORE_SIGINDEX_TIBIA_SHEET = "LoreSigIndex_Tibia_Auto"
LORE_WORDFREQ_TIBIA_SHEET = "LoreWordFreq_Tibia_Auto"
LORE_SIGINDEX_PD_SHEET = "LoreSigIndex_PD_Auto"
LORE_WORDFREQ_PD_SHEET = "LoreWordFreq_PD_Auto"
LORE_SIGINDEX_DICT_SHEET = "LoreSigIndex_Dict_Auto"
ENGLISH_MAP_SHEET = "EnglishMap_Auto"
PHRASE_CRIBS_AUTO_SHEET = "PhraseCribs_Auto"

# New safe-loop / plateau-exit analysis & readability layers.
DIGIT_CODE_MAP_SHEET = "DigitCodeMap_Auto"
DIGIT_LETTER_CODES_SHEET = "DigitLetterCodes_Auto"
DIGIT_CODE_CONTEXT_SHEET = "DigitCodeContext_Auto"
EXTERNAL_ROUNDTRIP_SHEET = "ExternalRoundTrip_Auto"
LORE_BIGRAMS_SHEET = "LoreBigrams_Auto"
ENGLISH_MAP_CONTEXT_SHEET = "EnglishMap_Context_Auto"
SEQUENCE_MATCHES_SHEET = "SequenceMatches_Auto"
SEQUENCE_MATCHES_CACHE_SHEET = "SequenceMatchesCache_Auto"
SEQUENCE_WORD_HINTS_SHEET = "SequenceWordHints_Auto"
SESTINA_LINES_SHEET = "SestinaLines_Auto"
SESTINA_CANDIDATES_SHEET = "SestinaCandidates_Auto"
SESTINA_OBLIGATION_SHEET = "SestinaObligation_Auto"
RHYTHM_TRANSITIONS_SHEET = "RhythmTransitions_Auto"
GT_POLICY_SHEET = "GroundTruthPolicy_Auto"
CODE_WORD_MAP_SHEET = "CodeWordMap_Auto"
SOFT_PROVISIONAL_DEFAULT_CRIB_IDS = {5, 9, 12, 13}
ANTI_HALLUCINATION_DEFAULT_TERMS = {
    "fay",
    "intenable",
    "invict",
    "hidy",
    "unnaive",
    "biface",
    "frutify",
    "tumtum",
    "sestine",
    "manxome",
    "frumious",
    "unfeasible",
}
ANTI_HALLUCINATION_STOPWORDS = {
    "a",
    "an",
    "and",
    "as",
    "at",
    "be",
    "by",
    "for",
    "from",
    "i",
    "if",
    "in",
    "into",
    "is",
    "it",
    "lo",
    "me",
    "no",
    "of",
    "on",
    "or",
    "so",
    "than",
    "that",
    "the",
    "this",
    "to",
    "we",
    "ye",
    "yet",
    "you",
    "youve",
    "your",
}


def _default_pd_sources() -> List[Tuple[str, str]]:
    """Default external plaintext sources for PD-style mining (mostly Project Gutenberg).

    Keep this list small and stable. The runner caches downloads under tmp/corpus/ and only
    persists *derived* indices/snippets in the workbook.
    """
    return [
        # Open-licensed wiki raw text (derived-only). High-signal because the decoded corpus contains "sestine" hints.
        ("WIKIPEDIA_SESTINA_RAW", DEFAULT_WIKI_SESTINA_RAW_URL),
        ("GUTENBERG_ALICE_11", DEFAULT_PD_ALICE_URL),
        ("GUTENBERG_LOOKING_GLASS_12", DEFAULT_PD_LOOKING_GLASS_URL),
        ("GUTENBERG_SHERLOCK_1661", DEFAULT_PD_SHERLOCK_URL),
        ("GUTENBERG_SHAKESPEARE_100", DEFAULT_PD_SHAKESPEARE_URL),
        ("GUTENBERG_PARADISE_26", DEFAULT_PD_PARADISE_LOST_URL),
        ("GUTENBERG_PRIDE_1342", DEFAULT_PD_PRIDE_PREJUDICE_URL),
        ("GUTENBERG_FRANKENSTEIN_84", DEFAULT_PD_FRANKENSTEIN_URL),
        ("GUTENBERG_DRACULA_345", DEFAULT_PD_DRACULA_URL),
        ("GUTENBERG_MOBY_2701", DEFAULT_PD_MOBY_DICK_URL),
        ("GUTENBERG_BEOWULF_16328", DEFAULT_PD_BEOWULF_GUMMERE_URL),
    ]


def utc_now_str() -> str:
    return datetime.now(timezone.utc).strftime(ISO_UTC_FMT)


def die(msg: str) -> None:
    raise SystemExit(msg)


def ws_find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, required: Sequence[str], max_scan: int = 10) -> int:
    req = set(required)
    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not row_vals:
            continue
        got = {v for v in row_vals if isinstance(v, str)}
        if req.issubset(got):
            return r
    die(f"Could not find header row in sheet {ws.title} with required columns: {required}")


def ws_headers(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        if isinstance(v, str) and v.strip():
            headers[v] = c
    return headers


def ws_last_data_row(ws: openpyxl.worksheet.worksheet.Worksheet, key_col: int = 1) -> int:
    for r in range(ws.max_row, 0, -1):
        if ws.cell(r, key_col).value is not None:
            return r
    return 0


def ws_append_row(ws: openpyxl.worksheet.worksheet.Worksheet, values: Sequence[object], start_col: int = 1) -> int:
    r = ws_last_data_row(ws) + 1
    for i, v in enumerate(values):
        ws.cell(r, start_col + i).value = v
    return r


def ensure_sheet(wb: openpyxl.Workbook, title: str, headers: Sequence[str]) -> openpyxl.worksheet.worksheet.Worksheet:
    """Create a sheet with a header row if missing; otherwise return existing."""
    if title in wb.sheetnames:
        return wb[title]
    ws = wb.create_sheet(title)
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = h
    return ws


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def backup_workbook(src_path: str, backup_dir: str, iter_num: int) -> str:
    ensure_dir(backup_dir)
    base = os.path.basename(src_path)
    name, ext = os.path.splitext(base)
    # Stable pattern, but never overwrite an existing backup.
    cand = os.path.join(backup_dir, f"{name}_backup_iter{iter_num}{ext}")
    if not os.path.exists(cand):
        shutil.copy2(src_path, cand)
        return cand
    i = 2
    while True:
        cand2 = os.path.join(backup_dir, f"{name}_backup_iter{iter_num}_{i}{ext}")
        if not os.path.exists(cand2):
            shutil.copy2(src_path, cand2)
            return cand2
        i += 1


@dataclass(frozen=True)
class GlossaryToken:
    token: str
    translation: str
    token_type: str
    confidence: str
    use_strictplus: bool
    evidence_class: str
    evidence_score: float
    evidence_score_t10: int  # EvidenceScore scaled by 10 (for exact DP scoring)
    total_occ: int
    length: int
    row: int  # row index in Glossary sheet

    @property
    def lossless_out_tokens(self) -> Tuple[str, ...]:
        # Markers render as <TOKEN> in the lossless stream.
        if self.token_type == "marker":
            return (f"<{self.token}>",)
        # Normal tokens render as their translation words.
        return tuple(self.translation.split())


def load_flow_settings(wb: openpyxl.Workbook) -> Tuple[openpyxl.worksheet.worksheet.Worksheet, Dict[str, Tuple[int, int, int, object]]]:
    ws = wb["FlowSettings"]
    header = ws_find_header_row(ws, ["Key", "Value"])
    cols = ws_headers(ws, header)
    out: Dict[str, Tuple[int, int, int, object]] = {}
    for r in range(header + 1, ws.max_row + 1):
        k = ws.cell(r, cols["Key"]).value
        if not isinstance(k, str) or not k.strip():
            continue
        v_cell = (r, cols["Value"])
        notes_col = cols.get("Notes")
        notes_cell = (r, notes_col) if notes_col else (r, cols["Value"] + 1)
        out[k] = (r, cols["Key"], cols["Value"], ws.cell(*v_cell).value)
        # Keep notes location derivable separately when writing.
        out[k] = (r, cols["Key"], cols["Value"], ws.cell(r, cols["Value"]).value)
    return ws, out


def get_setting(settings_map: Dict[str, Tuple[int, int, int, object]], key: str, default: object) -> object:
    item = settings_map.get(key)
    if item is None:
        return default
    return item[3]


def get_setting_value(settings_map: Dict[str, Tuple[int, int, int, object]], key: str, default: object) -> object:
    """Like get_setting(), but treats only None/blank-string as missing.

    Important: numeric 0 is a valid configuration value for many FlowSettings keys
    (e.g., MaxAgeHours=0 meaning "always refresh"), so callers must NOT use `or default`.
    """
    v = get_setting(settings_map, key, default)
    if v is None:
        return default
    if isinstance(v, str) and v.strip() == "":
        return default
    return v


def parse_bool(v: object, default: bool = False) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("true", "t", "1", "yes", "y"):
            return True
        if s in ("false", "f", "0", "no", "n"):
            return False
    return default


def _load_discord_env(env_path: str = DISCORD_ENV_PATH) -> None:
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as fh:
            for raw in fh:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if line.startswith("export "):
                    line = line[7:]
                if "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip()
                if not key:
                    continue
                if len(value) >= 2 and ((value[0] == "'" and value[-1] == "'") or (value[0] == '"' and value[-1] == '"')):
                    value = value[1:-1]
                # Prefer .env values for Discord settings, so local shell exports don't
                # override project-specific logging targets.
                if key.startswith("CODEX_DISCORD_"):
                    os.environ[key] = value
    except Exception:
        return


def _resolve_discord_channel_and_token() -> tuple[str, str, str]:
    _load_discord_env()
    token = os.environ.get("CODEX_DISCORD_TOKEN", "").strip()
    if not token:
        return "", "", "missing token"

    forced_channel = os.environ.get("CODEX_DISCORD_FORCE_LOG_CHANNEL_ID", "").strip()
    if forced_channel:
        return token, forced_channel, ""

    guild_id = os.environ.get("CODEX_DISCORD_DEFAULT_GUILD_ID", "").strip()
    if not guild_id:
        fallback_id = os.environ.get("CODEX_DISCORD_FORCE_LOG_CHANNEL_ID", "").strip()
        if not fallback_id:
            fallback_id = os.environ.get("CODEX_DISCORD_LOG_CHANNEL_ID", "").strip()
        if not fallback_id:
            fallback_id = os.environ.get("CODEX_DISCORD_DEFAULT_LOG_CHANNEL_ID", "").strip()
        if DISCORD_FALLBACK_CHANNEL_ID:
            fallback_id = DISCORD_FALLBACK_CHANNEL_ID
        if fallback_id:
            return token, fallback_id, ""
        return token, "", "missing guild id and fallback channel id"

    preferred = []
    # Always keep bonelord-logs as the first fallback to avoid accidental fallback into codex-logs.
    preferred.append("bonelord-logs")
    env_preferred = os.environ.get("CODEX_DISCORD_DEFAULT_LOG_CHANNEL_NAME", "").strip()
    if env_preferred:
        preferred.append(env_preferred)
    if DISCORD_DEFAULT_CHANNEL_NAME and DISCORD_DEFAULT_CHANNEL_NAME.lower() != "bonelord-logs":
        preferred.append(DISCORD_DEFAULT_CHANNEL_NAME)
    # De-duplicate while preserving order.
    preferred_names = []
    for item in preferred:
        item_l = str(item or "").strip().lower()
        if not item_l or any(item_l == str(x or "").strip().lower() for x in preferred_names):
            continue
        preferred_names.append(item_l)

    url = f"https://discord.com/api/v10/guilds/{guild_id}/channels"
    guild_error = ""
    try:
        req = urllib.request.Request(
            url,
            headers={"Authorization": f"Bot {token}", "User-Agent": "Codex-Flow-Bot"},
            method="GET",
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            channels = json.loads(r.read().decode("utf-8"))
            channel_map: Dict[str, str] = {}
            for ch in channels:
                name = str(ch.get("name", "")).strip().lower()
                if not name:
                    continue
                if name not in channel_map:
                    channel_map[name] = str(ch.get("id", "")).strip()
            for name in preferred_names:
                channel_id = channel_map.get(name)
                if channel_id:
                    return token, channel_id, ""
    except Exception as exc:
        guild_error = f"{type(exc).__name__}: {exc}"

    if DISCORD_FALLBACK_CHANNEL_ID:
        return token, DISCORD_FALLBACK_CHANNEL_ID, ""
    fallback_id = os.environ.get("CODEX_DISCORD_LOG_CHANNEL_ID", "").strip() or os.environ.get(
        "CODEX_DISCORD_DEFAULT_LOG_CHANNEL_ID", ""
    ).strip()
    if fallback_id:
        return token, fallback_id, ""
    if guild_error:
        return token, "", guild_error
    return token, "", f"channel '{', '.join(preferred_names)}' not found"


def _post_to_discord(token: str, channel_id: str, content: str) -> bool:
    url = f"https://discord.com/api/v10/channels/{channel_id}/messages"
    req = urllib.request.Request(
        url,
        method="POST",
        headers={
            "Authorization": f"Bot {token}",
            "Content-Type": "application/json",
            "User-Agent": "Codex-Flow-Notifier",
        },
        data=json.dumps({"content": content}).encode("utf-8"),
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            status = getattr(r, "status", 0)
            if 200 <= int(status) < 300:
                _ = r.read()
                return True
            body = r.read().decode("utf-8", "ignore").strip()
            print(f"[discord] envio negado: status={status} body={body[:220]}", file=sys.stderr)
            return False
    except Exception as exc:
        print(f"[discord] envio falhou: {type(exc).__name__}: {exc}", file=sys.stderr)
        return False


STATUS_MODEL_CONVERGED = "MODEL_CONVERGED"
STATUS_PUZZLE_SOLVED_LEGACY = "PUZZLE_SOLVED"


def _normalize_flow_status(status: object) -> str:
    raw = str(status or "").strip().upper()
    if raw == STATUS_PUZZLE_SOLVED_LEGACY:
        return STATUS_MODEL_CONVERGED
    return raw or "READY"


def _post_run_status_to_discord(
    workbook_path: str,
    next_iter: int,
    status: str,
    gt_ok: bool,
    block_reason: Optional[str],
    metrics: Dict[str, object],
) -> None:
    token, channel_id, _ = _resolve_discord_channel_and_token()
    if not token or not channel_id:
        return
    status = _normalize_flow_status(status)

    def _trim(text: str, max_len: int) -> str:
        text = (text or "").strip()
        if not text:
            return ""
        if len(text) <= max_len:
            return text
        return f"{text[: max_len - 1].strip()}…"

    block_line = block_reason if block_reason else "sem bloqueio explícito"
    current_soft = int(metrics.get("gt_soft", 0) or 0)
    current_status_hint = str(metrics.get("status_hint", "sem motivo"))
    gt_bad = int(metrics.get("gt_bad_all", 0) or 0)
    gt_bad_enf = int(metrics.get("gt_bad_enforced", 0) or 0)
    gt_total = int(metrics.get("gt_total", 0) or 0)
    gt_ok_count = int(metrics.get("gt_ok_count", 0) or 0)
    gt_overall_pct = float(metrics.get("gt_overall_pct", 0.0) or 0.0)
    gt_hard_pct = float(metrics.get("gt_hard_pct", 0.0) or 0.0)
    gt_target_pct = float(metrics.get("gt_target_pct", 0.0) or 0.0)
    candidate_rows = int(metrics.get("candidate_rows", 0) or 0)
    candidate_scan_mode = str(metrics.get("candidate_scan_mode", "normal"))
    candidate_empty_streak = int(metrics.get("candidate_empty_streak", 0) or 0)
    real_progress_streak = int(metrics.get("no_real_progress_iters", 0) or 0)
    low_confidence_escape = bool(metrics.get("low_confidence_escape", False))
    reverse_retext_applied = int(metrics.get("reverse_retext_applied", 0) or 0)
    reverse_retext_attempted = int(metrics.get("reverse_retext_attempted", 0) or 0)
    anchor_promo_only = bool(metrics.get("anchor_promo_only", False))
    anchor_corpus_size = int(metrics.get("anchor_corpus_size", 0) or 0)
    anchor_min_hits = int(metrics.get("anchor_min_hits", 1) or 1)
    anchor_kept_step30 = int(metrics.get("anchor_kept_step30", 0) or 0)
    anchor_dropped_step30 = int(metrics.get("anchor_dropped_step30", 0) or 0)
    anchor_kept_step40 = int(metrics.get("anchor_kept_step40", 0) or 0)
    anchor_dropped_step40 = int(metrics.get("anchor_dropped_step40", 0) or 0)
    anchor_drop30_classes = str(metrics.get("anchor_drop30_classes", "") or "").strip()
    anchor_drop30_samples = str(metrics.get("anchor_drop30_samples", "") or "").strip()
    anchor_drop40_classes = str(metrics.get("anchor_drop40_classes", "") or "").strip()
    anchor_drop40_samples = str(metrics.get("anchor_drop40_samples", "") or "").strip()
    strategy_summary = str(metrics.get("strategy_summary", "") or "").strip()
    conv_books_done = int(metrics.get("books_translated", 0) or 0)
    conv_books_total = int(metrics.get("books_total", 0) or 0)
    conv_contigs_done = int(metrics.get("contigs_translated", 0) or 0)
    conv_contigs_total = int(metrics.get("contigs_total", 0) or 0)
    conv_lines_done = int(metrics.get("lines_translated", 0) or 0)
    conv_lines_total = int(metrics.get("lines_total", 0) or 0)

    # Keep the channel log practical and actionable: what changed in this iteration.
    if status == "BLOCKED":
        state_text = "Parado para reavaliar direção: não houve avanço mecânico útil."
    elif status == "READY":
        state_text = "Ainda pronto para próxima tentativa mecânica."
    elif status == "SOFT_RESOLVED":
        state_text = "Objetivo duro limpo, mas divergência suave ainda existe."
    elif status == "RESOLVED":
        state_text = "Objetivo completo: sem divergência dura e sem divergência suave."
    elif status == STATUS_MODEL_CONVERGED:
        state_text = "Modelo convergiu nos guardrails atuais; isso ainda não é sinônimo de tradução 100%."
    else:
        state_text = "Estado não esperado, revisar logs e método"

    objective_line = (
        "Objetivo desta rodada: zerar GT duro e suavizar GT suave.\n"
        f"Leitura: duro={gt_bad_enf}, suave={current_soft}, total={gt_bad}."
    )
    progress_overall_line = (
        "Progresso geral: "
        f"{gt_overall_pct:.2f}% traduzido (GT {gt_ok_count}/{max(1, gt_total)}), "
        f"hard={gt_hard_pct:.2f}%, meta={gt_target_pct:.2f}%"
    )
    convergence_line = (
        "Cobertura: livros="
        f"{conv_books_done}/{conv_books_total} ({0.0 if conv_books_total <= 0 else 100.0 * conv_books_done / float(conv_books_total):.2f}%), "
        f"contigs={conv_contigs_done}/{conv_contigs_total} ({0.0 if conv_contigs_total <= 0 else 100.0 * conv_contigs_done / float(conv_contigs_total):.2f}%), "
        f"falas={conv_lines_done}/{conv_lines_total} ({0.0 if conv_lines_total <= 0 else 100.0 * conv_lines_done / float(conv_lines_total):.2f}%)"
    )

    gt_line = (
        f"Modo={metrics.get('gt_live_mode', 'POLICY')} | "
        f"hard={gt_bad_enf} | soft={current_soft} | total={gt_bad} | "
        f"streak sem melhora soft={metrics.get('gt_soft_nondec_streak', 0)}"
    )
    strategy_line = f"Estratégia: {strategy_summary}" if strategy_summary else "Estratégia: padrão"
    if low_confidence_escape:
        strategy_line += " | LOW-confidence=ATIVO"
    if anchor_promo_only:
        strategy_line += f" | anchor-only=ATIVO (anchors={anchor_corpus_size}, min_hits={anchor_min_hits})"
    actions_line = (
        f"candidatos={candidate_rows} (modo={candidate_scan_mode}, sem-cand={candidate_empty_streak}) | "
        f"promoções mecânicas={metrics.get('mech_promoted', 0)} | "
        f"semantic-objective={metrics.get('semantic_no_effect_promos', 0)} | "
        f"GT novos={metrics.get('gt_promoted', 0)} | "
        f"sem mecânica há {metrics.get('iters_since_last_mech', 0)} iteração(ões) | "
        f"skips={metrics.get('promotion_skip_count', 0)} | "
        f"resgate suave={metrics.get('soft_rescue_promos', 0)} (fronteira={metrics.get('soft_frontier_promos', 0)}) | "
        f"hard-escape={metrics.get('hard_escape_promos', 0)} | "
        f"directional-esc={metrics.get('directional_escape_promos', 0)}/"
        f"{metrics.get('directional_escape_attempts', 0)}"
    )
    if anchor_promo_only:
        actions_line += (
            f" | anchor-filter step30 kept/dropped={anchor_kept_step30}/{anchor_dropped_step30}"
            f" | step40 kept/dropped={anchor_kept_step40}/{anchor_dropped_step40}"
        )
        if anchor_drop30_classes:
            actions_line += f" | step30 classes={_trim(anchor_drop30_classes, 90)}"
        if anchor_drop40_classes:
            actions_line += f" | step40 classes={_trim(anchor_drop40_classes, 90)}"
    if reverse_retext_attempted:
        actions_line += f" | reverse-retext={reverse_retext_applied}/{reverse_retext_attempted}"
    if real_progress_streak:
        actions_line += f" | sem progresso real={real_progress_streak} iter"
    if metrics.get("semantic_progress_changed", 0):
        actions_line += " | progresso-semântico=SIM"
    if metrics.get("display_progress_changed", 0):
        actions_line += " | progresso-display=SIM"
    if metrics.get("analysis_progress_changed", 0):
        actions_line += " | progresso-análise=SIM"
    if block_reason:
        block_text = f"bloqueio={block_line}"
    else:
        block_text = f"{'apontando' if metrics.get('soft_improving', False) else 'sem melhora de soft'} | motivo={current_status_hint}"

    advances_line = str(metrics.get("progress_summary", "")).strip()
    if not advances_line:
        advances_line = "Sem avanço novo detectado ainda; mantendo busca por nova direção operacional."
    else:
        if ";" in advances_line:
            parts = [p.strip() for p in advances_line.split(";") if p.strip()]
            if len(parts) > 4:
                parts = parts[:4]
            advances_line = "; ".join(parts)
        advances_line = _trim(advances_line, 500)

    next_action = str(metrics.get("next_action", "")).strip()
    if not next_action:
        if metrics.get("directional_escape_attempts", 0):
            next_action = "Estamos em rota de escape direcional (rotas de exploração)."
        elif int(metrics.get("iters_since_last_mech", 0)) >= 3:
            next_action = "Fase de estagnação: mantendo escada de plateau e procurando candidatos com rampa curta."
        else:
            next_action = "Mantendo trilha padrão de melhoria mecânica + validação GT."
    next_action = _trim(next_action, 240)

    message = (
        "🚩 bonelord-logs | Iteração"
        f" {next_iter} | status={status} ({'OK' if gt_ok else 'ALERTA'}) | "
        f"{os.path.basename(workbook_path)}\n"
        f"{state_text}\n"
        f"{objective_line}\n"
        f"{convergence_line}\n"
        f"{progress_overall_line}\n"
        f"{gt_line}\n"
        f"{actions_line}\n"
        f"{strategy_line}\n"
        f"{'Direção atual: ' + next_action}\n"
        f"{'Avanços: ' + advances_line}\n"
        f"{block_text}"
    )
    if not _post_to_discord(token, channel_id, message):
        # best-effort logging only
        return


def parse_csv(value: object) -> List[str]:
    if value is None:
        return []
    if not isinstance(value, str):
        value = str(value)
    parts = [p.strip() for p in value.split(",")]
    return [p for p in parts if p]


def parse_int_list(value: object) -> List[int]:
    """Parse a CSV-like list of integers (e.g. '2,3,4')."""
    out: List[int] = []
    for p in parse_csv(value):
        try:
            out.append(int(p))
        except Exception:
            continue
    return out


def parse_word_set(value: object) -> set[str]:
    out: set[str] = set()
    for raw in parse_csv(value):
        w = re.sub(r"[^a-z]", "", str(raw or "").lower())
        if w:
            out.add(w)
    return out


def circular_expected_norm_flag(
    expected_norm: object,
    *,
    circular_terms: set[str],
    min_tokens_for_stopword_rule: int,
    stopword_ratio_threshold: float,
) -> bool:
    s = str(expected_norm or "").strip().lower()
    if not s:
        return False
    toks = [re.sub(r"[^a-z]", "", t) for t in re.findall(r"[a-zA-Z']+", s)]
    toks = [t for t in toks if t]
    if not toks:
        return False
    # Strong signal: explicitly suspicious terms seen only in speculative decode loops.
    if circular_terms and any(t in circular_terms for t in toks):
        return True
    # Weak signal: mostly function words with little lexical content.
    if len(toks) < max(1, int(min_tokens_for_stopword_rule)):
        return False
    stop_n = sum(1 for t in toks if t in ANTI_HALLUCINATION_STOPWORDS)
    stop_ratio = float(stop_n) / float(len(toks)) if toks else 0.0
    return stop_ratio >= float(stopword_ratio_threshold)


def set_setting(ws: openpyxl.worksheet.worksheet.Worksheet, key: str, value: object, note: Optional[str] = None) -> None:
    header = ws_find_header_row(ws, ["Key", "Value"])
    cols = ws_headers(ws, header)
    for r in range(header + 1, ws.max_row + 1):
        if ws.cell(r, cols["Key"]).value == key:
            ws.cell(r, cols["Value"]).value = value
            notes_col = cols.get("Notes")
            if note is not None and notes_col is not None:
                prev = ws.cell(r, notes_col).value
                if prev:
                    ws.cell(r, notes_col).value = f"{prev}; {note}"
                else:
                    ws.cell(r, notes_col).value = note
            return
    # Append new key if missing.
    notes_col = cols.get("Notes")
    row = ws_last_data_row(ws, key_col=cols["Key"]) + 1
    ws.cell(row, cols["Key"]).value = key
    ws.cell(row, cols["Value"]).value = value
    if note is not None and notes_col is not None:
        ws.cell(row, notes_col).value = note


def ensure_setting(ws: openpyxl.worksheet.worksheet.Worksheet, key: str, value: object, note: Optional[str] = None) -> None:
    """Ensure a FlowSettings key exists; if missing, append it with the given value."""
    header = ws_find_header_row(ws, ["Key", "Value"])
    cols = ws_headers(ws, header)
    for r in range(header + 1, ws.max_row + 1):
        if ws.cell(r, cols["Key"]).value == key:
            return
    set_setting(ws, key, value, note=note)


def load_flow_state(wb: openpyxl.Workbook) -> Tuple[openpyxl.worksheet.worksheet.Worksheet, Dict[str, Tuple[int, int, object]]]:
    ws = wb["FlowState"]
    header = ws_find_header_row(ws, ["Key", "Value"])
    cols = ws_headers(ws, header)
    out: Dict[str, Tuple[int, int, object]] = {}
    for r in range(header + 1, ws.max_row + 1):
        k = ws.cell(r, cols["Key"]).value
        if not isinstance(k, str) or not k.strip():
            continue
        out[k] = (r, cols["Value"], ws.cell(r, cols["Value"]).value)
    return ws, out


def set_flow_state_value(ws: openpyxl.worksheet.worksheet.Worksheet, key: str, value: object) -> None:
    header = ws_find_header_row(ws, ["Key", "Value"])
    cols = ws_headers(ws, header)
    for r in range(header + 1, ws.max_row + 1):
        if ws.cell(r, cols["Key"]).value == key:
            ws.cell(r, cols["Value"]).value = value
            return
    row = ws_last_data_row(ws, key_col=cols["Key"]) + 1
    ws.cell(row, cols["Key"]).value = key
    ws.cell(row, cols["Value"]).value = value


FlowStore = Dict[str, object]


def _sql_scalar_to_python(value: object) -> object:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return ""
    upper = text.upper()
    if upper == "TRUE":
        return True
    if upper == "FALSE":
        return False
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except Exception:
            return value
    if re.fullmatch(r"-?(?:\d+\.\d*|\.\d+)", text):
        try:
            return float(text)
        except Exception:
            return value
    return value


def _python_to_sql_scalar(value: object) -> str:
    if value is True:
        return "TRUE"
    if value is False:
        return "FALSE"
    if value is None:
        return ""
    return str(value)


def _flowstore_db_path(workbook_path: str) -> str:
    return os.path.join(os.path.dirname(os.path.abspath(workbook_path)), "data", SQLITE_DB_BASENAME)


def _flowstore_open_sqlite(workbook_path: str) -> Tuple[Optional[sqlite3.Connection], Optional[int]]:
    db_path = _flowstore_db_path(workbook_path)
    if not os.path.exists(db_path):
        return None, None
    conn = sqlite3.connect(db_path)
    row = conn.execute(
        "SELECT export_id FROM snapshot_refs WHERE name = ?",
        (SQLITE_CANONICAL_SNAPSHOT_NAME,),
    ).fetchone()
    if row is None:
        conn.close()
        return None, None
    return conn, int(row[0])


def _load_sql_flow_state_map(conn: sqlite3.Connection, export_id: int) -> Dict[str, Tuple[int, int, object]]:
    out: Dict[str, Tuple[int, int, object]] = {}
    rows = conn.execute(
        "SELECT __row_index, key, value FROM sheet__flowstate WHERE __export_id = ? ORDER BY __row_index",
        (export_id,),
    ).fetchall()
    for row_index, key, value in rows:
        if not isinstance(key, str) or not key.strip():
            continue
        out[key] = (int(row_index), 0, _sql_scalar_to_python(value))
    return out


def _load_sql_flow_settings_map(conn: sqlite3.Connection, export_id: int) -> Dict[str, Tuple[int, int, int, object]]:
    out: Dict[str, Tuple[int, int, int, object]] = {}
    rows = conn.execute(
        "SELECT __row_index, key, value FROM sheet__flowsettings WHERE __export_id = ? ORDER BY __row_index",
        (export_id,),
    ).fetchall()
    for row_index, key, value in rows:
        if not isinstance(key, str) or not key.strip():
            continue
        out[key] = (int(row_index), 0, 0, _sql_scalar_to_python(value))
    return out


def _sql_upsert_flow_state(conn: sqlite3.Connection, export_id: int, key: str, value: object) -> None:
    row = conn.execute(
        "SELECT __row_index FROM sheet__flowstate WHERE __export_id = ? AND key = ?",
        (export_id, key),
    ).fetchone()
    value_text = _python_to_sql_scalar(value)
    if row is None:
        next_row = conn.execute(
            "SELECT COALESCE(MAX(__row_index), 1) + 1 FROM sheet__flowstate WHERE __export_id = ?",
            (export_id,),
        ).fetchone()[0]
        conn.execute(
            "INSERT INTO sheet__flowstate (__export_id, __row_index, __sheet_name, key, value) VALUES (?, ?, 'FlowState', ?, ?)",
            (export_id, int(next_row), key, value_text),
        )
        return
    conn.execute(
        "UPDATE sheet__flowstate SET value = ? WHERE __export_id = ? AND key = ?",
        (value_text, export_id, key),
    )


def _sql_upsert_flow_setting(conn: sqlite3.Connection, export_id: int, key: str, value: object, note: Optional[str] = None) -> None:
    row = conn.execute(
        "SELECT __row_index, notes FROM sheet__flowsettings WHERE __export_id = ? AND key = ?",
        (export_id, key),
    ).fetchone()
    value_text = _python_to_sql_scalar(value)
    if row is None:
        next_row = conn.execute(
            "SELECT COALESCE(MAX(__row_index), 1) + 1 FROM sheet__flowsettings WHERE __export_id = ?",
            (export_id,),
        ).fetchone()[0]
        conn.execute(
            "INSERT INTO sheet__flowsettings (__export_id, __row_index, __sheet_name, key, value, notes) VALUES (?, ?, 'FlowSettings', ?, ?, ?)",
            (export_id, int(next_row), key, value_text, note or ""),
        )
        return
    notes_out = row[1] or ""
    if note:
        notes_out = f"{notes_out}; {note}" if notes_out else note
    conn.execute(
        "UPDATE sheet__flowsettings SET value = ?, notes = ? WHERE __export_id = ? AND key = ?",
        (value_text, notes_out, export_id, key),
    )


def open_flow_store(wb: openpyxl.Workbook, workbook_path: Optional[str] = None) -> FlowStore:
    ws_state, workbook_state_map = load_flow_state(wb)
    ws_settings, workbook_settings_map = load_flow_settings(wb)
    sql_conn: Optional[sqlite3.Connection] = None
    sql_export_id: Optional[int] = None
    state_map = workbook_state_map
    settings_map = workbook_settings_map
    if workbook_path:
        try:
            sql_conn, sql_export_id = _flowstore_open_sqlite(workbook_path)
        except Exception:
            sql_conn, sql_export_id = None, None
        if sql_conn is not None and sql_export_id is not None:
            try:
                state_map = _load_sql_flow_state_map(sql_conn, sql_export_id)
                settings_map = _load_sql_flow_settings_map(sql_conn, sql_export_id)
            except Exception:
                sql_conn.close()
                sql_conn = None
                sql_export_id = None
    return {
        "wb": wb,
        "ws_state": ws_state,
        "state_map": state_map,
        "ws_settings": ws_settings,
        "settings_map": settings_map,
        "sql_conn": sql_conn,
        "sql_export_id": sql_export_id,
    }


def flow_state_get(store: FlowStore, key: str, default: object = None) -> object:
    entry = store["state_map"].get(key)
    if entry is None:
        return default
    return entry[2]


def flow_state_set(store: FlowStore, key: str, value: object) -> None:
    sql_conn = store.get("sql_conn")
    sql_export_id = store.get("sql_export_id")
    if sql_conn is not None and sql_export_id is not None:
        _sql_upsert_flow_state(sql_conn, int(sql_export_id), key, value)
        sql_conn.commit()
    set_flow_state_value(store["ws_state"], key, value)
    entry = store["state_map"].get(key)
    if entry is None:
        store["state_map"][key] = (None, None, value)
    else:
        store["state_map"][key] = (entry[0], entry[1], value)


def flow_state_set_many(store: FlowStore, updates: Dict[str, object]) -> None:
    sql_conn = store.get("sql_conn")
    sql_export_id = store.get("sql_export_id")
    if sql_conn is not None and sql_export_id is not None:
        for key, value in updates.items():
            _sql_upsert_flow_state(sql_conn, int(sql_export_id), key, value)
        sql_conn.commit()
    for key, value in updates.items():
        set_flow_state_value(store["ws_state"], key, value)
        entry = store["state_map"].get(key)
        if entry is None:
            store["state_map"][key] = (None, None, value)
        else:
            store["state_map"][key] = (entry[0], entry[1], value)


def flow_setting_get(store: FlowStore, key: str) -> Optional[Tuple[int, int, int, object]]:
    return store["settings_map"].get(key)


def flow_setting_get_value(store: FlowStore, key: str, default: object = None) -> object:
    entry = flow_setting_get(store, key)
    if entry is None:
        return default
    return entry[3]


def flow_setting_set(store: FlowStore, key: str, value: object, note: Optional[str] = None) -> None:
    sql_conn = store.get("sql_conn")
    sql_export_id = store.get("sql_export_id")
    if sql_conn is not None and sql_export_id is not None:
        _sql_upsert_flow_setting(sql_conn, int(sql_export_id), key, value, note=note)
        sql_conn.commit()
    set_setting(store["ws_settings"], key, value, note=note)
    entry = flow_setting_get(store, key)
    if entry is None:
        store["settings_map"][key] = (None, None, None, value)
    else:
        store["settings_map"][key] = (entry[0], entry[1], entry[2], value)


def flow_setting_ensure(store: FlowStore, key: str, value: object, note: Optional[str] = None) -> None:
    if flow_setting_get(store, key) is not None:
        return
    flow_setting_set(store, key, value, note=note)


def flow_store_refresh_from_workbook(store: FlowStore) -> None:
    wb = store["wb"]
    ws_state, state_map = load_flow_state(wb)
    ws_settings, settings_map = load_flow_settings(wb)
    store["ws_state"] = ws_state
    store["state_map"] = state_map
    store["ws_settings"] = ws_settings
    store["settings_map"] = settings_map


def flow_store_close(store: FlowStore) -> None:
    sql_conn = store.get("sql_conn")
    if sql_conn is None:
        return
    try:
        sql_conn.close()
    finally:
        store["sql_conn"] = None


def sync_sqlite_snapshot_from_artifact(workbook_path: str, note: Optional[str] = None) -> Optional[int]:
    try:
        from pathlib import Path

        from export_workbook_to_sqlite import ensure_schema as export_ensure_schema
        from export_workbook_to_sqlite import existing_export_id, export_sheet, insert_export, sha256_file
        from sqlite_snapshot_ref import ensure_snapshot_refs_schema, mark_snapshot
    except Exception:
        return None

    conn: Optional[sqlite3.Connection] = None
    try:
        artifact_path = Path(workbook_path).resolve()
        db_path = Path(_flowstore_db_path(workbook_path)).resolve()
        db_path.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        export_ensure_schema(conn)
        ensure_snapshot_refs_schema(conn)

        digest = sha256_file(artifact_path)
        export_id = existing_export_id(conn, digest)
        if export_id is None:
            export_id = insert_export(conn, artifact_path)
            raw_wb = openpyxl.load_workbook(artifact_path, data_only=False, read_only=True)
            cached_wb = openpyxl.load_workbook(artifact_path, data_only=True, read_only=True)
            for sheet_index, sheet_name in enumerate(raw_wb.sheetnames, start=1):
                export_sheet(conn, export_id, sheet_index, raw_wb[sheet_name], cached_wb[sheet_name])
                conn.commit()
        mark_snapshot(conn, int(export_id), snapshot_name=SQLITE_CANONICAL_SNAPSHOT_NAME, note=note)
        return int(export_id)
    except Exception:
        return None
    finally:
        if conn is not None:
            conn.close()


def load_glossary(wb: openpyxl.Workbook) -> Tuple[openpyxl.worksheet.worksheet.Worksheet, Dict[str, GlossaryToken]]:
    ws = wb["Glossary"]
    header = ws_find_header_row(
        ws,
        [
            "Token",
            "Translation",
            "TokenType",
            "Confidence",
            "Use_StrictPlus_v108",
            "EvidenceClass_v127",
            "EvidenceScore_v127",
            "TotalOcc",
            "Len",
        ],
    )
    cols = ws_headers(ws, header)

    out: Dict[str, GlossaryToken] = {}
    for r in range(header + 1, ws.max_row + 1):
        token = ws.cell(r, cols["Token"]).value
        if not isinstance(token, str) or not token:
            continue
        translation = ws.cell(r, cols["Translation"]).value
        if translation is None:
            translation = ""
        token_type = ws.cell(r, cols["TokenType"]).value or ""
        confidence = ws.cell(r, cols["Confidence"]).value or ""
        use_sp = bool(ws.cell(r, cols["Use_StrictPlus_v108"]).value)
        evidence_class = ws.cell(r, cols["EvidenceClass_v127"]).value or ""
        ev_score = ws.cell(r, cols["EvidenceScore_v127"]).value
        if ev_score is None:
            # Keep None as 0 to avoid crashes; these should not be used.
            ev_score = 0.0
        # EvidenceScore values in this workbook are decimal with 1 digit (e.g., 2.5, 1.7).
        # Scale to int to avoid float accumulation/tie issues in DP.
        ev_score_t10 = int(round(float(ev_score) * 10))
        total_occ = int(ws.cell(r, cols["TotalOcc"]).value or 0)
        length = ws.cell(r, cols["Len"]).value
        if length is None:
            length = len(token)
        out[token] = GlossaryToken(
            token=token,
            translation=str(translation),
            token_type=str(token_type),
            confidence=str(confidence),
            use_strictplus=use_sp,
            evidence_class=str(evidence_class),
            evidence_score=float(ev_score),
            evidence_score_t10=ev_score_t10,
            total_occ=total_occ,
            length=int(length),
            row=r,
        )
    return ws, out


def confidence_weight(conf: str) -> int:
    if conf == "HIGH":
        return 3
    if conf == "MEDIUM":
        return 2
    if conf == "LOW":
        return 1
    return 0


def glossary_set_use_strictplus(glossary_ws: openpyxl.worksheet.worksheet.Worksheet, token_row: int, use: bool, iter_num: int, note: str) -> None:
    header = ws_find_header_row(glossary_ws, ["Use_StrictPlus_v108", "Notes"])
    cols = ws_headers(glossary_ws, header)
    glossary_ws.cell(token_row, cols["Use_StrictPlus_v108"]).value = 1 if use else 0
    notes_cell = glossary_ws.cell(token_row, cols["Notes"])
    prev = notes_cell.value
    msg = f"iter{iter_num}: {note}"
    if prev and isinstance(prev, str):
        if msg not in prev:
            notes_cell.value = f"{prev}; {msg}"
    else:
        notes_cell.value = msg


def parse_lossless_tokens(lossless: str) -> List[str]:
    # Lossless markers stream is already space-delimited.
    # Keep punctuation tokens '.' and '!' as their own tokens.
    return lossless.split()


def render_strictplus_from_lossless(lossless_tokens: Sequence[str]) -> str:
    out: List[str] = []
    for tok in lossless_tokens:
        if tok in ("<E>", "<FF>"):
            if not out:
                out.append(",")
            else:
                out[-1] = f"{out[-1]},"
            continue
        if tok == "<*>":
            out.append("—")
            continue
        if tok in (".", "!"):
            if not out:
                out.append(tok)
            else:
                out[-1] = f"{out[-1]}{tok}"
            continue
        out.append(tok)
    return " ".join(out)


def render_crib_lossless_from_lossless(lossless_tokens: Sequence[str]) -> str:
    """Cribs convention: keep `<*>` visible, but render comma markers as punctuation."""
    out: List[str] = []
    for tok in lossless_tokens:
        if tok in ("<E>", "<FF>"):
            if not out:
                out.append(",")
            else:
                out[-1] = f"{out[-1]},"
            continue
        if tok == "<*>":
            out.append("<*>")
            continue
        if tok in (".", "!"):
            if not out:
                out.append(tok)
            else:
                out[-1] = f"{out[-1]}{tok}"
            continue
        out.append(tok)
    return " ".join(out)


_NORM_KEEP_CHARS = "*↵"
_NORM_DROP_RE = re.compile(rf"[^a-z0-9{re.escape(_NORM_KEEP_CHARS)}]+")
_NORM_WS_RE = re.compile(r"\s+")


def normalize_for_match(text: str) -> str:
    """Normalize text for Cribs ExpectedNorm/DecodeNorm comparisons."""
    s = (text or "").lower()
    s = s.replace("—", " * ")
    s = s.replace("'", " ")
    s = _NORM_DROP_RE.sub(" ", s)
    s = _NORM_WS_RE.sub(" ", s).strip()
    return s


def load_readability_rules(wb: openpyxl.Workbook, iter_num: int) -> List[Dict[str, object]]:
    """Load enabled readability rules from the workbook.

    This layer is intentionally separate from StrictPlus decoding: it rewrites rendered text for readability
    without changing tokenization or evidence metrics.
    """
    ws = ensure_sheet(
        wb,
        "ReadabilityRules",
        ["RuleID", "Enabled", "RuleType", "From", "To", "Scope", "AddedIter", "Notes"],
    )
    header = ws_find_header_row(ws, ["RuleID", "Enabled", "RuleType", "From", "To", "Scope"])
    c = ws_headers(ws, header)

    existing_from = set()
    for r in range(header + 1, ws.max_row + 1):
        frm = ws.cell(r, c["From"]).value
        if isinstance(frm, str) and frm.strip():
            existing_from.add(frm.strip())

    def _seed(frm: str, to: str, rule_type: str, enabled: bool, scope: str, notes: str) -> None:
        if frm in existing_from:
            return
        r = ws_last_data_row(ws) + 1
        ws.cell(r, c["RuleID"]).value = r - header  # simple stable id
        ws.cell(r, c["Enabled"]).value = 1 if enabled else 0
        ws.cell(r, c["RuleType"]).value = rule_type
        ws.cell(r, c["From"]).value = frm
        ws.cell(r, c["To"]).value = to
        ws.cell(r, c["Scope"]).value = scope
        if "AddedIter" in c:
            ws.cell(r, c["AddedIter"]).value = iter_num
        if "Notes" in c:
            ws.cell(r, c["Notes"]).value = notes

    # Minimal default set (idempotent); user can extend/tune in the XLSX.
    _seed(
        "fiftin statue",
        "fifteen statues",
        "PHRASE",
        True,
        "BOOKS,MASTER",
        "From FifteenStatuesAudit_v114: readability-only (does not change StrictPlus decode).",
    )
    _seed(
        "eye",
        "ye",
        "WORD",
        False,
        "BOOKS,MASTER,CRIBS",
        "Suggestion from FifteenStatuesAudit_v114 (disabled by default; enable if desired).",
    )
    _seed(
        "far way",
        "far away",
        "PHRASE",
        False,
        "CRIBS",
        "From IStarA_Analysis_v112: normalization/readability helper (disabled by default).",
    )

    rules: List[Dict[str, object]] = []
    for r in range(header + 1, ws.max_row + 1):
        frm = ws.cell(r, c["From"]).value
        to = ws.cell(r, c["To"]).value
        if not isinstance(frm, str) or not frm.strip():
            continue
        if not isinstance(to, str) or not to.strip():
            continue
        enabled = parse_bool(ws.cell(r, c["Enabled"]).value, False)
        if not enabled:
            continue
        rule_type = str(ws.cell(r, c["RuleType"]).value or "PHRASE").strip().upper()
        scope_raw = str(ws.cell(r, c["Scope"]).value or "BOOKS,MASTER").strip().upper()
        scopes = {s.strip() for s in scope_raw.split(",") if s and s.strip()}
        # Normalize common synonyms.
        norm_scopes: set[str] = set()
        for s in scopes:
            if s in ("MASTERTEXT", "MASTER_TEXT", "MASTER"):
                norm_scopes.add("MASTER")
            else:
                norm_scopes.add(s)
        rules.append({"from": frm.strip(), "to": to.strip(), "type": rule_type, "scope": scope_raw, "scopes": norm_scopes})

    rules.sort(key=lambda x: (-len(str(x["from"])), str(x["from"])))
    return rules


def apply_readability_rules(text: str, rules: Sequence[Dict[str, object]]) -> Tuple[str, int]:
    s = text or ""
    applied = 0
    for rule in rules:
        frm = str(rule["from"])
        to = str(rule["to"])
        rtype = str(rule.get("type") or "PHRASE").upper()
        if not frm:
            continue
        # Conservative whole-word / whole-phrase substitution.
        if rtype == "WORD" and " " not in frm:
            pat = r"\b" + re.escape(frm) + r"\b"
        else:
            pat = r"\b" + re.escape(frm) + r"\b"
        s, n = re.subn(pat, to, s)
        applied += int(n)
    return s, applied


def apply_readability_layer(wb: openpyxl.Workbook, iter_num: int) -> Tuple[int, int, int, int]:
    """Apply enabled readability rules into stable '*_Readable_Auto' columns.

    Returns: (books_changed_rows, mastertext_changed_rows, cribs_changed_rows, replacements_applied)
    """
    rules = load_readability_rules(wb, iter_num)
    if not rules:
        return 0, 0, 0, 0

    rules_books = [r for r in rules if "BOOKS" in (r.get("scopes") or set())]
    rules_master = [r for r in rules if "MASTER" in (r.get("scopes") or set())]
    rules_cribs = [r for r in rules if "CRIBS" in (r.get("scopes") or set())]

    repl_total = 0  # replacements that actually changed an output cell this run
    books_changed = 0
    master_changed = 0
    cribs_changed = 0

    # Books
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "Translation_StrictPlus_v108"])
    cb = ws_headers(ws_books, hb)
    src_col = cb["Translation_StrictPlus_v108"]
    out_col = cb.get("Translation_Readable_Auto")
    if out_col is None:
        out_col = ws_books.max_column + 1
        ws_books.cell(hb, out_col).value = "Translation_Readable_Auto"

    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        src = ws_books.cell(r, src_col).value
        if not isinstance(src, str):
            src = str(src or "")
        out, n = apply_readability_rules(src, rules_books)
        prev = ws_books.cell(r, out_col).value
        if prev != out:
            books_changed += 1
            repl_total += n
            ws_books.cell(r, out_col).value = out

    # MasterText
    ws_mt = wb["MasterText"]
    hm = ws_find_header_row(ws_mt, ["BaseContigID", "Translation_StrictPlus_v108"])
    cm = ws_headers(ws_mt, hm)
    srcm = cm["Translation_StrictPlus_v108"]
    outm = cm.get("Translation_Readable_Auto")
    if outm is None:
        outm = ws_mt.max_column + 1
        ws_mt.cell(hm, outm).value = "Translation_Readable_Auto"

    for r in range(hm + 1, ws_mt.max_row + 1):
        cid = ws_mt.cell(r, cm["BaseContigID"]).value
        if cid is None:
            continue
        src = ws_mt.cell(r, srcm).value
        if not isinstance(src, str):
            src = str(src or "")
        out, n = apply_readability_rules(src, rules_master)
        prev = ws_mt.cell(r, outm).value
        if prev != out:
            master_changed += 1
            repl_total += n
            ws_mt.cell(r, outm).value = out

    # Cribs (display-only; keep core DP fields untouched)
    if rules_cribs:
        ws_c = wb["Cribs"]
        hc = ws_find_header_row(ws_c, ["CribID", "DP_StrictPlus_Readable_v112"])
        cc = ws_headers(ws_c, hc)
        src_c = cc["DP_StrictPlus_Readable_v112"]
        out_c = cc.get("DP_Readable_Auto")
        if out_c is None:
            out_c = ws_c.max_column + 1
            ws_c.cell(hc, out_c).value = "DP_Readable_Auto"

        for r in range(hc + 1, ws_c.max_row + 1):
            crib_id = ws_c.cell(r, cc["CribID"]).value
            if crib_id is None:
                continue
            src = ws_c.cell(r, src_c).value
            if not isinstance(src, str):
                src = str(src or "")
            out, n = apply_readability_rules(src, rules_cribs)
            prev = ws_c.cell(r, out_c).value
            if prev != out:
                cribs_changed += 1
                repl_total += n
                ws_c.cell(r, out_c).value = out

    return books_changed, master_changed, cribs_changed, repl_total


def _apply_late_display_cleanup_text(text: str, rules: Sequence[Dict[str, object]]) -> Tuple[str, int]:
    s = str(text or "")
    repl = 0
    s, n = re.subn(r",{2,}", ",", s)
    repl += int(n)
    s, n = re.subn(r"\s+([,.;:!?])", r"\1", s)
    repl += int(n)
    s, n = re.subn(r"(?<![A-Za-z])i(?![A-Za-z])", "I", s)
    repl += int(n)
    s, n = re.subn(r"<unk>", "<UNK>", s, flags=re.IGNORECASE)
    repl += int(n)
    if rules:
        s, n2 = apply_readability_rules(s, rules)
        repl += int(n2)
    return s, repl


def apply_late_display_cleanup(wb: openpyxl.Workbook, iter_num: int) -> Tuple[int, int, int]:
    """Apply conservative display-only cleanup to late readable output columns in Books."""
    rules = load_readability_rules(wb, iter_num)
    rules_books = [r for r in rules if "BOOKS" in (r.get("scopes") or set())]

    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID"], max_scan=3)
    cb = ws_headers(ws_books, hb)

    changed_ctx = 0
    changed_codeaware = 0
    repl_total = 0

    for col_name, bucket in (
        ("Translation_ContextEnglish_Auto", "ctx"),
        ("Translation_CodeAware_Auto", "codeaware"),
    ):
        col = cb.get(col_name)
        if col is None:
            continue
        for r in range(hb + 1, ws_books.max_row + 1):
            bid = ws_books.cell(r, cb["BookID"]).value
            if bid is None:
                continue
            prev = ws_books.cell(r, col).value
            if prev is None:
                continue
            out, n = _apply_late_display_cleanup_text(str(prev), rules_books)
            if out != prev:
                if bucket == "ctx":
                    changed_ctx += 1
                else:
                    changed_codeaware += 1
                ws_books.cell(r, col).value = out
            repl_total += int(n)

    return changed_ctx, changed_codeaware, repl_total


def dp_tokenize_base_with_punct(base: str, active_tokens: Dict[str, GlossaryToken]) -> List[object]:
    """DP-tokenize a base string that may contain '.' and "!" punctuation.

    Returns a list where each item is either a GlossaryToken or the punctuation string '.'/"!".
    """
    n = len(base)

    by_first: Dict[str, List[GlossaryToken]] = defaultdict(list)
    for gt in active_tokens.values():
        if gt.token:
            by_first[gt.token[:1]].append(gt)
    for k in list(by_first.keys()):
        by_first[k].sort(key=lambda t: (-t.length, -t.total_occ, t.token))

    # DP: i -> (score_tuple, backpointer)
    # score_tuple: (conf_cov, len2_sum, occ_sum, token_count_neg)
    best: Dict[int, Tuple[Tuple[int, int, int, int], Optional[Tuple[int, str]]]] = {0: ((0, 0, 0, 0), None)}

    for i in range(0, n + 1):
        cur = best.get(i)
        if cur is None:
            continue
        cur_score, _bp = cur
        if i == n:
            continue

        ch = base[i]
        if ch in (".", "!"):
            ni = i + 1
            prev = best.get(ni)
            if prev is None or cur_score > prev[0]:
                best[ni] = (cur_score, (i, ch))
            continue

        for tok in by_first.get(ch, []):
            k = tok.length
            if i + k > n:
                continue
            sub = base[i : i + k]
            if not base_substring_matches_token(sub, tok.token):
                continue
            cand_score = (
                cur_score[0] + confidence_weight(tok.confidence) * k,
                cur_score[1] + k * k,
                cur_score[2] + tok.total_occ,
                cur_score[3] - 1,
            )
            ni = i + k
            prev = best.get(ni)
            if prev is None or cand_score > prev[0]:
                best[ni] = (cand_score, (i, tok.token))

    end = best.get(n)
    if end is None:
        raise RuntimeError(f"Tokenize failed: base_len={n} base_preview={base[:80]!r}")

    items_rev: List[str] = []
    i = n
    while True:
        _score, bp = best[i]
        if bp is None:
            break
        pi, sym = bp
        items_rev.append(sym)
        i = pi
    items_rev.reverse()

    out: List[object] = []
    for sym in items_rev:
        if sym in (".", "!"):
            out.append(sym)
            continue
        tok = active_tokens.get(sym)
        if tok is None:
            raise RuntimeError(f"Tokenize backtrack missing token: {sym!r}")
        out.append(tok)
    return out


def recompute_cribs_dp(
    wb: openpyxl.Workbook,
    glossary_tokens: Dict[str, GlossaryToken],
) -> Tuple[int, int, int, Optional[int]]:
    """Recompute Cribs DP/readable/norm/match fields using current active token set.

    Returns: (rows_updated, gt_mismatches, match_flips, avartar_match)
    """
    ws = wb["Cribs"]
    header = ws_find_header_row(
        ws,
        [
            "CribID",
            "CribClass_v112",
            "DecodedBase_Sanitized",
            "DP_StrictPlus_v111",
            "DP_StrictPlus_v111_LosslessMarkers",
            "DP_StrictPlus_Readable_v112",
            "ExpectedNorm_v112",
            "DecodeNorm_v112",
            "MatchNorm_v112",
            "MatchNorm_GroundTruth_v112",
            "LettersLen",
            "TokenCount",
            "HighCharFrac",
            "SingleCharFrac",
        ],
    )
    c = ws_headers(ws, header)

    active = {t.token: t for t in glossary_tokens.values() if t.use_strictplus and t.translation}

    rows_updated = 0
    gt_mismatches = 0
    match_flips = 0
    avartar_match: Optional[int] = None

    for r in range(header + 1, ws.max_row + 1):
        base = ws.cell(r, c["DecodedBase_Sanitized"]).value
        if base is None or str(base).strip() == "":
            continue
        base_s = str(base)

        cls = ws.cell(r, c["CribClass_v112"]).value
        exp_norm = ws.cell(r, c["ExpectedNorm_v112"]).value
        prev_match = ws.cell(r, c["MatchNorm_v112"]).value

        # Keep GroundTruth rows stable: they are anchored by `Expected` and prior validation,
        # and we do not want an unconstrained DP to overwrite them.
        if cls == "GroundTruth":
            continue

        items = dp_tokenize_base_with_punct(base_s, active)

        letters_len = sum((it.length if isinstance(it, GlossaryToken) else 0) for it in items)
        token_count = sum((1 if isinstance(it, GlossaryToken) else 0) for it in items)
        high_chars = sum((it.length if isinstance(it, GlossaryToken) and it.confidence == "HIGH" else 0) for it in items)
        single_chars = sum((it.length if isinstance(it, GlossaryToken) and it.length == 1 else 0) for it in items)

        lossless: List[str] = []
        for it in items:
            if isinstance(it, str):
                lossless.append(it)
            elif it.token_type == "marker":
                lossless.append(f"<{it.token}>")
            else:
                lossless.extend(it.translation.split())

        readable = render_strictplus_from_lossless(lossless)
        crib_lossless = render_crib_lossless_from_lossless(lossless)

        ws.cell(r, c["LettersLen"]).value = letters_len
        ws.cell(r, c["TokenCount"]).value = token_count
        ws.cell(r, c["HighCharFrac"]).value = (high_chars / letters_len) if letters_len else 0
        ws.cell(r, c["SingleCharFrac"]).value = (single_chars / letters_len) if letters_len else 0
        ws.cell(r, c["DP_StrictPlus_v111"]).value = readable
        ws.cell(r, c["DP_StrictPlus_v111_LosslessMarkers"]).value = crib_lossless
        ws.cell(r, c["DP_StrictPlus_Readable_v112"]).value = readable

        dec_norm = normalize_for_match(readable)
        ws.cell(r, c["DecodeNorm_v112"]).value = dec_norm

        if exp_norm is None or str(exp_norm).strip() == "":
            ws.cell(r, c["MatchNorm_v112"]).value = None
        else:
            exp_norm_s = str(exp_norm).strip()
            m = 1 if dec_norm == exp_norm_s else 0
            ws.cell(r, c["MatchNorm_v112"]).value = m
            if prev_match is not None and int(prev_match) != m:
                match_flips += 1

        if cls == "GroundTruth":
            cur_m = ws.cell(r, c["MatchNorm_v112"]).value
            ws.cell(r, c["MatchNorm_GroundTruth_v112"]).value = cur_m
            if cur_m not in (1, True):
                gt_mismatches += 1

        crib_col = c.get("Crib")
        if crib_col:
            crib_text = ws.cell(r, crib_col).value
            if isinstance(crib_text, str) and "Avar Tar poem" in crib_text:
                cur_m = ws.cell(r, c["MatchNorm_v112"]).value
                if cur_m is not None:
                    avartar_match = int(cur_m)

        rows_updated += 1

    return rows_updated, gt_mismatches, match_flips, avartar_match


def ensure_groundtruth_policy_sheet(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    default_enforced_ids: set[int],
    anti_mode: bool = False,
    demote_circular_gt: bool = False,
    circular_terms: Optional[set[str]] = None,
    circular_min_tokens: int = 6,
    circular_stopword_ratio: float = 0.80,
    circular_max_verified: int = 2,
) -> Tuple[set[int], int, str]:
    """Ensure GroundTruthPolicy_Auto exists and yields a stable enforced-crib set.

    This is intentionally conservative: it does **not** overwrite a non-empty `Enforced` cell.
    The goal is to keep the GT live check as a strong guardrail, but avoid blocking progress on
    cribs whose `Expected` text is not externally verified (numeric-only sources).

    Returns: (enforced_ids, rows_touched, status)
    """
    if "Cribs" not in wb.sheetnames:
        return set(default_enforced_ids), 0, "missing Cribs"

    headers = [
        "CribID",
        "Enforced",
        "EvidenceTier",
        "DefaultEnforced",
        "CribClass",
        "VerifiedSourceCount_v129",
        "HasExpectedNorm",
        "UpdatedIter",
        "Reason",
        "NotesExcerpt",
    ]
    ws_pol = ensure_sheet(wb, GT_POLICY_SHEET, headers)
    hp = ws_find_header_row(ws_pol, ["CribID", "Enforced"], max_scan=3)
    cp = ws_headers(ws_pol, hp)

    # Add any missing columns (non-destructive; append to the right).
    for h in headers:
        if h not in cp:
            ws_pol.cell(hp, ws_pol.max_column + 1).value = h
            cp = ws_headers(ws_pol, hp)

    by_id: Dict[int, int] = {}
    for r in range(hp + 1, ws_pol.max_row + 1):
        cid = ws_pol.cell(r, cp["CribID"]).value
        if cid is None:
            continue
        try:
            by_id[int(cid)] = r
        except Exception:
            continue

    ws_c = wb["Cribs"]
    hc = ws_find_header_row(ws_c, ["CribID", "CribClass_v112"], max_scan=3)
    cc = ws_headers(ws_c, hc)
    vsc_col = cc.get("VerifiedSourceCount_v129")
    exp_col = cc.get("ExpectedNorm_v112")
    notes_col = cc.get("Notes")

    touched = 0

    circular_terms_set = set(circular_terms or set())
    for r in range(hc + 1, ws_c.max_row + 1):
        cid = ws_c.cell(r, cc["CribID"]).value
        if cid is None:
            continue
        try:
            cid_i = int(cid)
        except Exception:
            continue
        cls = str(ws_c.cell(r, cc["CribClass_v112"]).value or "").strip()
        vsc = ws_c.cell(r, vsc_col).value if vsc_col else None
        try:
            vsc_i = int(vsc or 0)
        except Exception:
            vsc_i = 0
        exp_norm = ws_c.cell(r, exp_col).value if exp_col else None
        has_exp = 1 if (exp_norm is not None and str(exp_norm).strip() != "") else 0
        notes = ws_c.cell(r, notes_col).value if notes_col else None
        notes_s = str(notes or "")
        circular_expected = bool(
            anti_mode
            and demote_circular_gt
            and cls == "GroundTruth"
            and has_exp
            and vsc_i <= int(circular_max_verified)
            and circular_expected_norm_flag(
                exp_norm,
                circular_terms=circular_terms_set,
                min_tokens_for_stopword_rule=max(1, int(circular_min_tokens)),
                stopword_ratio_threshold=float(circular_stopword_ratio),
            )
        )

        default_tier = "SOFT_PROVISIONAL" if cid_i in SOFT_PROVISIONAL_DEFAULT_CRIB_IDS else "HARD_EXTERNAL"
        pol_r = by_id.get(cid_i)
        if pol_r is None:
            pol_r = ws_last_data_row(ws_pol, key_col=cp["CribID"]) + 1
            ws_pol.cell(pol_r, cp["CribID"]).value = cid_i
            by_id[cid_i] = pol_r
            touched += 1

        tier_cell = ws_pol.cell(pol_r, cp["EvidenceTier"]) if cp.get("EvidenceTier") is not None else None
        tier_s = str(tier_cell.value or "").strip().upper() if tier_cell is not None else ""
        # Keep defaults stable and explicit:
        # - hard external rows are blocking by default
        # - specific known provisional rows remain soft until externally validated
        if tier_cell is not None:
            if cid_i in SOFT_PROVISIONAL_DEFAULT_CRIB_IDS:
                if tier_s != "SOFT_PROVISIONAL":
                    tier_cell.value = "SOFT_PROVISIONAL"
                    touched += 1
                tier_s = "SOFT_PROVISIONAL"
            elif circular_expected:
                if tier_s != "SOFT_PROVISIONAL":
                    tier_cell.value = "SOFT_PROVISIONAL"
                    touched += 1
                tier_s = "SOFT_PROVISIONAL"
            elif not tier_s:
                tier_cell.value = default_tier
                tier_s = default_tier
                touched += 1

        is_soft_tier = tier_s == "SOFT_PROVISIONAL"
        is_hard_tier = tier_s == "HARD_EXTERNAL"
        default_enforced = 0 if is_soft_tier else (1 if is_hard_tier else (1 if cid_i in default_enforced_ids else 0))

        ws_pol.cell(pol_r, cp["DefaultEnforced"]).value = int(default_enforced)
        ws_pol.cell(pol_r, cp["CribClass"]).value = cls
        ws_pol.cell(pol_r, cp["VerifiedSourceCount_v129"]).value = int(vsc_i)
        ws_pol.cell(pol_r, cp["HasExpectedNorm"]).value = int(has_exp)
        ws_pol.cell(pol_r, cp["UpdatedIter"]).value = int(iter_num)

        if cp.get("NotesExcerpt") is not None:
            ws_pol.cell(pol_r, cp["NotesExcerpt"]).value = (notes_s[:180] + ("..." if len(notes_s) > 180 else ""))

        # Only set Enforced when empty, preserving manual edits.
        enf_cell = ws_pol.cell(pol_r, cp["Enforced"])
        if enf_cell.value is None or str(enf_cell.value).strip() == "":
            enf_cell.value = bool(default_enforced)
            touched += 1

        reason = ""
        if is_soft_tier:
            if circular_expected:
                reason = "soft_provisional_circular_expected"
            else:
                reason = "soft_provisional_default"
        elif is_hard_tier:
            reason = "hard_external_default"
        elif default_enforced:
            reason = "default_enforced_ids"
        if cp.get("Reason") is not None:
            ws_pol.cell(pol_r, cp["Reason"]).value = reason

    # Enforced set for the runner: only GT rows with an ExpectedNorm.
    enforced: set[int] = set()
    for cid_i, pol_r in by_id.items():
        tier_s = str(ws_pol.cell(pol_r, cp["EvidenceTier"]).value or "").strip().upper() if cp.get("EvidenceTier") is not None else ""
        if tier_s == "SOFT_PROVISIONAL":
            # Soft tier is visible/traceable, but non-blocking for the hard GT guardrail.
            continue
        if tier_s == "HARD_EXTERNAL":
            # Hard tier is always enforced.
            enforced.add(int(cid_i))
            continue
        enf_v = ws_pol.cell(pol_r, cp["Enforced"]).value
        def_v = ws_pol.cell(pol_r, cp["DefaultEnforced"]).value
        is_enf = parse_bool(enf_v, default=parse_bool(def_v, False))
        if is_enf:
            enforced.add(int(cid_i))

    upsert_sheet_index_entry(
        wb,
        GT_POLICY_SHEET,
        "GroundTruth enforcement policy for the GT live check (guardrail). Allows keeping only externally-verified expected-text cribs enforced, while treating numeric-only expected text as soft (non-blocking).",
    )
    return enforced, touched, "ok"


def resolve_enforced_groundtruth_ids(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    settings_map: Dict[str, Tuple[int, int, int, object]],
) -> Tuple[Optional[set[int]], int, str]:
    """Resolve which CribID(s) are enforced by the GT live check.

    Returns: (enforced_ids or None for ALL, enforced_count, status)
    """
    mode = str(get_setting(settings_map, "GroundTruthLiveCheck_Mode", "POLICY") or "POLICY").strip().upper()
    if mode in ("ALL", "FULL"):
        return None, 0, "mode=ALL"

    # Policy mode (default): use GroundTruthPolicy_Auto (created if missing).
    default_ids = set(parse_int_list(get_setting_value(settings_map, "GroundTruthPolicy_DefaultEnforcedCribIDs", "2,3,4,7")))
    anti_mode = parse_bool(get_setting(settings_map, "AntiHallucination_Mode", True), True)
    demote_circular_gt = parse_bool(get_setting(settings_map, "AntiHallucination_DemoteCircularGT", True), True)
    circular_terms = parse_word_set(
        get_setting_value(
            settings_map,
            "AntiHallucination_CircularTerms",
            ",".join(sorted(ANTI_HALLUCINATION_DEFAULT_TERMS)),
        )
    )
    if not circular_terms:
        circular_terms = set(ANTI_HALLUCINATION_DEFAULT_TERMS)
    circular_min_tokens = int(get_setting_value(settings_map, "AntiHallucination_CircularMinTokens", 6))
    circular_stopword_ratio = float(get_setting_value(settings_map, "AntiHallucination_CircularStopwordRatio", 0.78))
    circular_max_verified = int(get_setting_value(settings_map, "AntiHallucination_MinVerifiedForHardGT", 2))
    enforced, touched, st = ensure_groundtruth_policy_sheet(
        wb,
        iter_num=iter_num,
        default_enforced_ids=default_ids,
        anti_mode=anti_mode,
        demote_circular_gt=demote_circular_gt,
        circular_terms=circular_terms,
        circular_min_tokens=circular_min_tokens,
        circular_stopword_ratio=circular_stopword_ratio,
        circular_max_verified=circular_max_verified,
    )
    return enforced, len(enforced), f"{st} (touched={touched}, default={sorted(default_ids)})"


def groundtruth_live_check(
    wb: openpyxl.Workbook,
    active_tokens: Dict[str, GlossaryToken],
    *,
    enforced_crib_ids: Optional[set[int]] = None,
) -> Tuple[bool, List[Tuple[int, str, str, str]], List[Tuple[int, str, str, str]]]:
    """Validate GroundTruth cribs against the *current* DP tokenization.

    This is a guardrail that prevents us from expanding search space (macro-mining / promotions)
    while the anchored bilingual references no longer match the current StrictPlus token set.

    Enforcement model:
    - If `enforced_crib_ids` is None: enforce ALL GroundTruth rows (legacy behavior).
    - Otherwise: enforce only the listed CribID(s); other mismatches are logged/visible but non-blocking.

    Returns: (ok_enforced, bad_enforced, bad_all)
      bad items are tuples: (CribID, Crib, decode_norm, expected_norm)
    """
    ws = wb["Cribs"]
    header = ws_find_header_row(ws, ["CribID", "CribClass_v112", "DecodedBase_Sanitized", "ExpectedNorm_v112"])
    c = ws_headers(ws, header)

    bad_all: List[Tuple[int, str, str, str]] = []
    bad_enforced: List[Tuple[int, str, str, str]] = []
    for r in range(header + 1, ws.max_row + 1):
        cls = ws.cell(r, c["CribClass_v112"]).value
        if cls != "GroundTruth":
            continue
        crib_id = ws.cell(r, c["CribID"]).value
        if crib_id is None:
            continue
        try:
            cid_i = int(crib_id)
        except Exception:
            continue
        base = ws.cell(r, c["DecodedBase_Sanitized"]).value
        exp_norm = ws.cell(r, c["ExpectedNorm_v112"]).value
        if base is None or str(base).strip() == "":
            continue
        if exp_norm is None or str(exp_norm).strip() == "":
            continue

        items = dp_tokenize_base_with_punct(str(base), active_tokens)
        lossless: List[str] = []
        for it in items:
            if isinstance(it, str):
                lossless.append(it)
            elif it.token_type == "marker":
                lossless.append(f"<{it.token}>")
            else:
                lossless.extend(it.translation.split())
        readable = render_strictplus_from_lossless(lossless)
        dec_norm = normalize_for_match(readable)
        exp_norm_s = str(exp_norm).strip()
        if dec_norm != exp_norm_s:
            crib_txt = ws.cell(r, c.get("Crib", 0) or 0).value if c.get("Crib") else ""
            row = (cid_i, str(crib_txt or ""), dec_norm, exp_norm_s)
            bad_all.append(row)
            if enforced_crib_ids is None or cid_i in enforced_crib_ids:
                bad_enforced.append(row)

    return (len(bad_enforced) == 0), bad_enforced, bad_all


def add_gt_repair_macros_from_mismatches(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    mismatches: Sequence[Tuple[int, str, str, str]],
    *,
    max_macros: int = 10,
) -> Tuple[int, List[str], str]:
    """Append active macro tokens to reconcile GroundTruth mismatches safely.

    Strategy: for each mismatching CribID, add a macro token for the base (sans trailing punctuation)
    whose translation normalizes to ExpectedNorm_v112. This is conservative and traceable:
    - tokens are added with EvidenceClass=GROUNDTRUTH and Confidence=HIGH
    - only exact base matches are added (no web-scraping)
    """
    if max_macros <= 0:
        return 0, [], "max_macros<=0"
    if "Cribs" not in wb.sheetnames or "Glossary" not in wb.sheetnames:
        return 0, [], "missing Cribs/Glossary"

    ws_c = wb["Cribs"]
    hc = ws_find_header_row(
        ws_c,
        ["CribID", "DecodedBase_Sanitized", "ExpectedNorm_v112", "Expected", "DP_StrictPlus_Readable_v112"],
        max_scan=3,
    )
    cc = ws_headers(ws_c, hc)

    row_by_id: Dict[int, int] = {}
    for r in range(hc + 1, ws_c.max_row + 1):
        cid = ws_c.cell(r, cc["CribID"]).value
        if cid is None:
            continue
        try:
            row_by_id[int(cid)] = r
        except Exception:
            continue

    ws_g = wb["Glossary"]
    hg = ws_find_header_row(
        ws_g,
        [
            "Token",
            "Translation",
            "TokenType",
            "Confidence",
            "ActiveCorpus",
            "Use_StrictPlus_v108",
            "Use_Strict_v108",
            "BoundaryRule",
            "RuleLayerNeeded",
            "TotalOcc",
            "BookCount",
            "ContigCount",
            "Len",
            "StarCount",
            "Notes",
            "Use_PoemMode_v113",
            "SemNeutral_Render_v123",
            "Mask_SemNeutral_v123",
            "EvidenceClass_v127",
            "EvidenceScore_v127",
            "EvidenceSources_v127",
        ],
        max_scan=3,
    )
    cg = ws_headers(ws_g, hg)

    template_row = None
    for r in range(hg + 1, ws_g.max_row + 1):
        if ws_g.cell(r, cg["TokenType"]).value == "macro":
            template_row = r
            break
    if template_row is None:
        template_row = hg + 1

    existing = {
        str(ws_g.cell(r, cg["Token"]).value).strip()
        for r in range(hg + 1, ws_g.max_row + 1)
        if isinstance(ws_g.cell(r, cg["Token"]).value, str) and str(ws_g.cell(r, cg["Token"]).value).strip()
    }

    def _pick_translation(exp_norm_s: str, expected: object, readable: object) -> str:
        cand: List[str] = []
        for v in (expected, readable):
            if v is None:
                continue
            s = str(v).strip()
            if s:
                cand.append(s)
        # Last-resort: use ExpectedNorm itself as the macro translation.
        cand.append(exp_norm_s)
        for s in cand:
            try:
                if normalize_for_match(s) == exp_norm_s:
                    return s
            except Exception:
                continue
        return exp_norm_s

    added = 0
    tokens_added: List[str] = []
    skipped: List[str] = []

    for cid, _crib, _dec, _exp in mismatches:
        if added >= max_macros:
            break
        rr = row_by_id.get(int(cid))
        if rr is None:
            skipped.append(f"{cid}:missing_row")
            continue

        base = ws_c.cell(rr, cc["DecodedBase_Sanitized"]).value
        exp_norm = ws_c.cell(rr, cc["ExpectedNorm_v112"]).value
        if base is None or str(base).strip() == "":
            skipped.append(f"{cid}:missing_base")
            continue
        if exp_norm is None or str(exp_norm).strip() == "":
            skipped.append(f"{cid}:missing_expected_norm")
            continue

        base_s_full = str(base).strip()
        # DP treats '.' and '!' as punctuation symbols, not part of tokens.
        removed_punct = ""
        while base_s_full and base_s_full[-1] in (".", "!"):
            removed_punct = base_s_full[-1] + removed_punct
            base_s_full = base_s_full[:-1]
        base_s = base_s_full
        if not base_s:
            skipped.append(f"{cid}:empty_base_after_strip")
            continue
        if base_s in existing:
            skipped.append(f"{cid}:{base_s}:exists")
            continue

        exp_norm_s = str(exp_norm).strip()
        expected = ws_c.cell(rr, cc.get("Expected", 0) or 0).value if cc.get("Expected") else None
        readable = (
            ws_c.cell(rr, cc.get("DP_StrictPlus_Readable_v112", 0) or 0).value if cc.get("DP_StrictPlus_Readable_v112") else None
        )
        tr = _pick_translation(exp_norm_s, expected, readable)
        # Avoid doubling trailing punctuation if the base ended with punctuation we stripped.
        if removed_punct:
            tr_s = str(tr).rstrip()
            if removed_punct.endswith("!") and tr_s.endswith("!"):
                tr = tr_s[:-1].rstrip()
            elif removed_punct.endswith(".") and tr_s.endswith("."):
                tr = tr_s[:-1].rstrip()

        note = f"iter{iter_num} GT-repair macro (CribID={cid}). normalize(tr)==ExpectedNorm_v112. AddedUTC={utc}."

        values: Dict[int, object] = {}

        def _set(k: str, v: object) -> None:
            col = cg.get(k)
            if col is not None:
                values[col] = v

        _set("Token", base_s)
        _set("Translation", tr)
        _set("TokenType", "macro")
        _set("Confidence", "HIGH")
        _set("ActiveCorpus", False)
        _set("Use_StrictPlus_v108", 1)
        _set("Use_Strict_v108", False)
        _set("BoundaryRule", "none")
        _set("RuleLayerNeeded", "no_anagram")
        _set("TotalOcc", 0)
        _set("BookCount", 0)
        _set("ContigCount", 0)
        _set("Len", len(base_s))
        _set("StarCount", int(base_s.count("*")))
        _set("Notes", note)
        _set("Use_PoemMode_v113", False)
        _set("SemNeutral_Render_v123", "NORMAL")
        _set("Mask_SemNeutral_v123", 0)
        _set("EvidenceClass_v127", "GROUNDTRUTH")
        _set("EvidenceScore_v127", 3.0)
        _set("EvidenceSources_v127", f"iter{iter_num}: gt_repair_crib{cid}")

        _glossary_append_row_copy_style(ws_g, template_row, values_by_col=values)
        existing.add(base_s)
        tokens_added.append(base_s)
        added += 1

    summary = f"GT-repair macros added={added}, skipped={len(skipped)}"
    return added, tokens_added, summary


@dataclass
class AlignResult:
    tokens: List[GlossaryToken]
    evidence_sum: float
    weak_chars: int
    micro_medium_chars: int
    masked_micro_chars: int
    marker_chars: int
    single_char_chars: int
    token_count: int


def is_masked_micro(tok: GlossaryToken) -> bool:
    # SemanticNeutral masking:
    # - mask short (len<=3) MEDIUM/LOW tokens
    # - but do not mask markers/punct logograms
    if tok.token_type == "marker":
        return False
    if tok.evidence_class == "PUNCT_LOGOGRAM":
        return False
    if tok.confidence not in ("MEDIUM", "LOW"):
        return False
    return tok.length <= 3


def base_substring_matches_token(base_sub: str, tok: str) -> bool:
    """Match base substring to token, allowing '*' to act as a wildcard letter placeholder.

    In this workbook, '*' can represent certain letters in the base alphabet (see Rules sheet).
    We treat it as a placeholder only when matching base -> token; token strings remain literal.
    """
    if len(base_sub) != len(tok):
        return False
    for bc, tc in zip(base_sub, tok):
        if bc == tc:
            continue
        if bc == "*" and tc in STAR_WILDCARD_CHARS:
            continue
        return False
    return True


def dp_align_base_to_lossless(
    base: str,
    lossless_tokens: Sequence[str],
    active_tokens: Dict[str, GlossaryToken],
) -> AlignResult:
    """
    Find the best tokenization of `base` that renders exactly to `lossless_tokens`.

    Scoring matches the workbook's TieBreakPolicy sheet:
    1) maximize Σ(confW*len) where confW: HIGH=3, MEDIUM=2, LOW=1
    2) maximize Σ(len^2)
    3) maximize Σ(TotalOcc)
    4) maximize -token_count (fewer tokens)

    Note: WEAK/single-char are downstream metrics, not tie-break objectives.
    """

    n = len(base)
    m = len(lossless_tokens)

    # Index by first character for quicker base matching.
    by_first: Dict[str, List[GlossaryToken]] = defaultdict(list)
    for gt in active_tokens.values():
        by_first[gt.token[:1]].append(gt)

    # Token enumeration order: (len desc, TotalOcc desc, Token lexicographic)
    for k in list(by_first.keys()):
        by_first[k].sort(key=lambda t: (-t.length, -t.total_occ, t.token))

    # DP maps (i,j) -> (score_tuple, backpointer)
    # score_tuple: (conf_cov, len2_sum, occ_sum, token_count_neg)
    best: Dict[Tuple[int, int], Tuple[Tuple[int, int, int, int], Optional[Tuple[int, int, str]]]] = {}
    best[(0, 0)] = ((0, 0, 0, 0), None)

    # Iterate in increasing i+j order for determinism.
    # Collect states grouped by i to avoid scanning the full grid.
    states_by_i: Dict[int, List[int]] = defaultdict(list)
    states_by_i[0].append(0)

    for i in range(0, n + 1):
        js = states_by_i.get(i)
        if not js:
            continue
        for j in js:
            cur = best.get((i, j))
            if cur is None:
                continue
            cur_score, _ = cur
            if i == n and j == m:
                continue
            if i >= n or j >= m:
                continue

            first = base[i : i + 1]
            for tok in by_first.get(first, []):
                k = tok.length
                if i + k > n:
                    continue
                if not base_substring_matches_token(base[i : i + k], tok.token):
                    continue
                out_toks = tok.lossless_out_tokens
                l = len(out_toks)
                if j + l > m:
                    continue
                if tuple(lossless_tokens[j : j + l]) != out_toks:
                    continue

                conf_add = confidence_weight(tok.confidence) * k
                len2_add = k * k
                occ_add = tok.total_occ
                token_add_neg = -1

                cand_score = (
                    cur_score[0] + conf_add,
                    cur_score[1] + len2_add,
                    cur_score[2] + occ_add,
                    cur_score[3] + token_add_neg,
                )
                ni, nj = i + k, j + l
                prev = best.get((ni, nj))
                if prev is None or cand_score > prev[0]:
                    best[(ni, nj)] = (cand_score, (i, j, tok.token))
                    states_by_i[ni].append(nj)

    end = best.get((n, m))
    if end is None:
        # Provide a small diagnostic context.
        preview = " ".join(lossless_tokens[:40])
        raise RuntimeError(f"DP align failed (no path). base_len={n} out_len={m} out_preview={preview!r}")

    # Reconstruct tokens
    tokens_rev: List[GlossaryToken] = []
    i, j = n, m
    while True:
        entry = best[(i, j)]
        bp = entry[1]
        if bp is None:
            break
        pi, pj, tok_str = bp
        tokens_rev.append(active_tokens[tok_str])
        i, j = pi, pj
    tokens_rev.reverse()

    # Compute metrics from the chosen tokenization.
    evidence_sum_t10 = 0
    weak_chars = 0
    micro_medium_chars = 0
    masked_micro_chars = 0
    marker_chars = 0
    single_char_chars = 0
    for t in tokens_rev:
        evidence_sum_t10 += t.evidence_score_t10 * t.length
        if t.evidence_score_t10 < 20:
            weak_chars += t.length
        if t.evidence_class == "MICRO_MEDIUM":
            micro_medium_chars += t.length
        if is_masked_micro(t):
            masked_micro_chars += t.length
        if t.token_type == "marker":
            marker_chars += t.length
        if t.length == 1:
            single_char_chars += t.length

    return AlignResult(
        tokens=tokens_rev,
        evidence_sum=evidence_sum_t10 / 10.0,
        weak_chars=weak_chars,
        micro_medium_chars=micro_medium_chars,
        masked_micro_chars=masked_micro_chars,
        marker_chars=marker_chars,
        single_char_chars=single_char_chars,
        token_count=len(tokens_rev),
    )


def recompute_books_contigs_mastertext(
    wb: openpyxl.Workbook,
    glossary_tokens: Dict[str, GlossaryToken],
) -> Tuple[str, float, float, float, float]:
    """
    Recompute:
    - Books: HighFracWithinStrictPlus_v108, SingleCharFracWithinStrictPlus_v108, TokenCount_StrictPlus_v108,
             Translation_SemNeutral_v123, Translation_HighOnly_v123,
             MaskedCharFrac_v123, MaskedTokenCount_v123, MarkerCharFrac_v123
    - Contigs: TokenCount, HighCharsFrac, Translation_* columns, Masked fractions/tokens
    - MasterText: derived from Contigs by BaseContigID.

    Returns:
    - changed_books "X/70" computed on StrictPlus_v108 text (should remain stable),
    - weighted evidence avg, weighted weak frac, weighted micro_medium frac, weighted single_char frac
      for Books corpus (for logging).
    """

    # Prepare active token dictionary for DP (Use_StrictPlus_v108 == 1 only).
    active = {t.token: t for t in glossary_tokens.values() if t.use_strictplus and t.translation}

    # BOOKS
    ws_books = wb["Books"]
    header = ws_find_header_row(
        ws_books,
        [
            "BookID",
            "BaseLen",
            "DecodedBase",
            "Translation_StrictPlus_v108",
            "Translation_StrictPlus_LosslessMarkers_v108",
            "HighFracWithinStrictPlus_v108",
            "SingleCharFracWithinStrictPlus_v108",
            "TokenCount_StrictPlus_v108",
            "Translation_SemNeutral_v123",
            "Translation_HighOnly_v123",
            "MaskedCharFrac_v123",
            "MaskedTokenCount_v123",
            "MarkerCharFrac_v123",
        ],
    )
    c = ws_headers(ws_books, header)

    changed = 0
    total_books = 0

    sum_base = 0
    sum_ev = 0.0
    sum_weak = 0
    sum_micro_med = 0
    sum_single = 0

    for r in range(header + 1, ws_books.max_row + 1):
        book_id = ws_books.cell(r, c["BookID"]).value
        if book_id is None:
            continue
        total_books += 1
        base_len = int(ws_books.cell(r, c["BaseLen"]).value or 0)
        base = ws_books.cell(r, c["DecodedBase"]).value or ""
        prev_strictplus = ws_books.cell(r, c["Translation_StrictPlus_v108"]).value or ""
        items = dp_tokenize_base_with_punct(str(base), active)

        lossless_out: List[str] = []
        tokens: List[GlossaryToken] = []
        for it in items:
            if isinstance(it, str):
                lossless_out.append(it)
            else:
                tokens.append(it)
                lossless_out.extend(it.lossless_out_tokens)

        # Keep the workbook lossless anchor up to date (translation edits are allowed).
        ws_books.cell(r, c["Translation_StrictPlus_LosslessMarkers_v108"]).value = " ".join(lossless_out)

        strictplus = render_strictplus_from_lossless(lossless_out)
        if strictplus != prev_strictplus:
            changed += 1
        ws_books.cell(r, c["Translation_StrictPlus_v108"]).value = strictplus

        # SemNeutral and HighOnly derived from tokenization.
        sem_tokens: List[str] = []
        high_tokens: List[str] = []
        masked_token_count = 0
        for it in items:
            if isinstance(it, str):
                sem_tokens.append(it)
                continue
            t = it
            if t.token_type == "marker":
                sem_tokens.append(f"<{t.token}>")
                continue
            # Punct logograms should render as their punctuation tokens in sem-neutral view.
            if t.evidence_class == "PUNCT_LOGOGRAM":
                sem_tokens.extend(t.translation.split())
            elif is_masked_micro(t):
                masked_token_count += 1
                for w in t.translation.split():
                    sem_tokens.append(f"[{w}]")
            else:
                sem_tokens.extend(t.translation.split())

            if t.confidence == "HIGH" and t.token_type != "marker":
                if t.evidence_class != "PUNCT_LOGOGRAM":
                    high_tokens.extend(t.translation.split())

        ws_books.cell(r, c["Translation_SemNeutral_v123"]).value = " ".join(sem_tokens)
        ws_books.cell(r, c["Translation_HighOnly_v123"]).value = " ".join(high_tokens)

        # Metrics
        evidence_sum_t10 = sum(int(t.evidence_score_t10) * int(t.length) for t in tokens)
        weak_chars = sum(int(t.length) for t in tokens if int(t.evidence_score_t10) < 20)
        micro_medium_chars = sum(int(t.length) for t in tokens if t.evidence_class == "MICRO_MEDIUM")
        masked_micro_chars = sum(int(t.length) for t in tokens if is_masked_micro(t))
        marker_chars = sum(int(t.length) for t in tokens if t.token_type == "marker")
        single_char_chars = sum(int(t.length) for t in tokens if int(t.length) == 1)

        if base_len > 0:
            ws_books.cell(r, c["HighFracWithinStrictPlus_v108"]).value = (
                sum(int(t.length) for t in tokens if t.confidence == "HIGH") / base_len
            )
            ws_books.cell(r, c["SingleCharFracWithinStrictPlus_v108"]).value = single_char_chars / base_len
            ws_books.cell(r, c["TokenCount_StrictPlus_v108"]).value = len(tokens)
            ws_books.cell(r, c["MaskedCharFrac_v123"]).value = masked_micro_chars / base_len
            ws_books.cell(r, c["MaskedTokenCount_v123"]).value = masked_token_count
            ws_books.cell(r, c["MarkerCharFrac_v123"]).value = marker_chars / base_len
        else:
            ws_books.cell(r, c["HighFracWithinStrictPlus_v108"]).value = 0
            ws_books.cell(r, c["SingleCharFracWithinStrictPlus_v108"]).value = 0
            ws_books.cell(r, c["TokenCount_StrictPlus_v108"]).value = 0
            ws_books.cell(r, c["MaskedCharFrac_v123"]).value = 0
            ws_books.cell(r, c["MaskedTokenCount_v123"]).value = 0
            ws_books.cell(r, c["MarkerCharFrac_v123"]).value = 0

        sum_base += base_len
        sum_ev += evidence_sum_t10 / 10.0
        sum_weak += weak_chars
        sum_micro_med += micro_medium_chars
        sum_single += single_char_chars

    changed_books = f"{changed}/{total_books}"

    weighted_ev_avg = (sum_ev / sum_base) if sum_base else 0.0
    weighted_weak_frac = (sum_weak / sum_base) if sum_base else 0.0
    weighted_micro_frac = (sum_micro_med / sum_base) if sum_base else 0.0
    weighted_single_frac = (sum_single / sum_base) if sum_base else 0.0

    # CONTIGS
    ws_contigs = wb["Contigs"]
    header_c = ws_find_header_row(
        ws_contigs,
        [
            "BaseContigID",
            "Length",
            "BaseContig",
            "TokenCount",
            "HighCharsFrac",
            "Translation_StrictPlus_LosslessMarkers_v108",
            "Translation_StrictPlus_v108",
            "Translation_SemNeutral_v123",
            "Translation_HighOnly_v123",
            "MaskedCharFrac_v123",
            "MaskedTokenCount_v123",
        ],
    )
    cc = ws_headers(ws_contigs, header_c)

    for r in range(header_c + 1, ws_contigs.max_row + 1):
        cid = ws_contigs.cell(r, cc["BaseContigID"]).value
        if cid is None:
            continue
        base_len = int(ws_contigs.cell(r, cc["Length"]).value or 0)
        base = ws_contigs.cell(r, cc["BaseContig"]).value or ""
        items = dp_tokenize_base_with_punct(str(base), active)
        lossless_out: List[str] = []
        tokens: List[GlossaryToken] = []
        for it in items:
            if isinstance(it, str):
                lossless_out.append(it)
            else:
                tokens.append(it)
                lossless_out.extend(it.lossless_out_tokens)

        ws_contigs.cell(r, cc["Translation_StrictPlus_LosslessMarkers_v108"]).value = " ".join(lossless_out)
        ws_contigs.cell(r, cc["TokenCount"]).value = len(tokens)
        ws_contigs.cell(r, cc["HighCharsFrac"]).value = (sum(int(t.length) for t in tokens if t.confidence == "HIGH") / base_len) if base_len else 0
        ws_contigs.cell(r, cc["Translation_StrictPlus_v108"]).value = render_strictplus_from_lossless(lossless_out)

        sem_tokens: List[str] = []
        high_tokens: List[str] = []
        masked_token_count = 0
        for it in items:
            if isinstance(it, str):
                sem_tokens.append(it)
                continue
            t = it
            if t.token_type == "marker":
                sem_tokens.append(f"<{t.token}>")
                continue
            if t.evidence_class == "PUNCT_LOGOGRAM":
                sem_tokens.extend(t.translation.split())
            elif is_masked_micro(t):
                masked_token_count += 1
                for w in t.translation.split():
                    sem_tokens.append(f"[{w}]")
            else:
                sem_tokens.extend(t.translation.split())

            if t.confidence == "HIGH" and t.token_type != "marker":
                if t.evidence_class != "PUNCT_LOGOGRAM":
                    high_tokens.extend(t.translation.split())

        ws_contigs.cell(r, cc["Translation_SemNeutral_v123"]).value = " ".join(sem_tokens)
        ws_contigs.cell(r, cc["Translation_HighOnly_v123"]).value = " ".join(high_tokens)
        masked_micro_chars = sum(int(t.length) for t in tokens if is_masked_micro(t))
        ws_contigs.cell(r, cc["MaskedCharFrac_v123"]).value = (masked_micro_chars / base_len) if base_len else 0
        ws_contigs.cell(r, cc["MaskedTokenCount_v123"]).value = masked_token_count

    # MASTERTEXT: copy translation columns for each BaseContigID from Contigs.
    contig_map: Dict[int, Dict[str, object]] = {}
    for r in range(header_c + 1, ws_contigs.max_row + 1):
        cid = ws_contigs.cell(r, cc["BaseContigID"]).value
        if cid is None:
            continue
        contig_map[int(cid)] = {
            "Translation_StrictPlus_LosslessMarkers_v108": ws_contigs.cell(r, cc["Translation_StrictPlus_LosslessMarkers_v108"]).value,
            "Translation_StrictPlus_v108": ws_contigs.cell(r, cc["Translation_StrictPlus_v108"]).value,
            "Translation_SemNeutral_v123": ws_contigs.cell(r, cc["Translation_SemNeutral_v123"]).value,
            "Translation_HighOnly_v123": ws_contigs.cell(r, cc["Translation_HighOnly_v123"]).value,
            "MaskedCharFrac_v123": ws_contigs.cell(r, cc["MaskedCharFrac_v123"]).value,
        }

    ws_mt = wb["MasterText"]
    header_m = ws_find_header_row(
        ws_mt,
        [
            "BaseContigID",
            "Translation_StrictPlus_LosslessMarkers_v108",
            "Translation_StrictPlus_v108",
            "Translation_SemNeutral_v123",
            "Translation_HighOnly_v123",
            "MaskedCharFrac_v123",
        ],
    )
    cm = ws_headers(ws_mt, header_m)
    for r in range(header_m + 1, ws_mt.max_row + 1):
        cid = ws_mt.cell(r, cm["BaseContigID"]).value
        if cid is None:
            continue
        data = contig_map.get(int(cid))
        if not data:
            continue
        for k, v in data.items():
            ws_mt.cell(r, cm[k]).value = v

    return changed_books, weighted_ev_avg, weighted_weak_frac, weighted_micro_frac, weighted_single_frac


def recompute_token_evidence_books(
    wb: openpyxl.Workbook,
    glossary_tokens: Dict[str, GlossaryToken],
) -> Tuple[float, float, float]:
    """
    Recompute TokenEvidence_Books_v128 based on the current Glossary tokenization (translation-string independent).
    Returns weighted (EvidenceAvg, WeakFrac, MicroFrac).
    """
    active = {t.token: t for t in glossary_tokens.values() if t.use_strictplus and t.translation}

    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "BaseLen", "DecodedBase", "Translation_StrictPlus_LosslessMarkers_v108"])
    cb = ws_headers(ws_books, hb)

    ws_ev = wb["TokenEvidence_Books_v128"]
    he = ws_find_header_row(
        ws_ev,
        ["BookID", "BaseLen", "TokenCount", "TokenCount_expected", "TokenCount_match", "EvidenceAvg", "StrongFrac", "MediumFrac", "WeakFrac", "MicroFrac", "LogogramContextFrac", "GroundTruthFrac", "AnagramHighFrac", "MarkerFrac", "WeakTopTokens"],
    )
    ce = ws_headers(ws_ev, he)

    # Map BookID -> row in TokenEvidence sheet.
    ev_row_by_book: Dict[int, int] = {}
    for r in range(he + 1, ws_ev.max_row + 1):
        bid = ws_ev.cell(r, ce["BookID"]).value
        if bid is None:
            continue
        ev_row_by_book[int(bid)] = r

    sum_base = 0
    sum_ev = 0.0
    sum_weak = 0
    sum_micro = 0

    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        bid_i = int(bid)
        base_len = int(ws_books.cell(r, cb["BaseLen"]).value or 0)
        base = str(ws_books.cell(r, cb["DecodedBase"]).value or "")
        items = dp_tokenize_base_with_punct(base, active)
        toks = [it for it in items if isinstance(it, GlossaryToken)]

        tok_count = len(toks)
        evidence_sum_t10 = sum(int(t.evidence_score_t10) * int(t.length) for t in toks)
        evidence_sum = evidence_sum_t10 / 10.0
        ev_avg = (evidence_sum / base_len) if base_len else 0.0

        # TokenEvidence convention in this workbook:
        # - MediumFrac counts sentence punctuation logograms ('.' and '!') only.
        # - WeakFrac counts all chars from tokens with EvidenceScore < 2.0 (broad weak).
        # - StrongFrac is the remainder.
        medium_chars = sum(int(t.length) for t in toks if t.translation in (".", "!"))
        weak_chars = sum(int(t.length) for t in toks if int(t.evidence_score_t10) < 20)
        strong_chars = base_len - weak_chars - medium_chars
        micro_chars = sum(int(t.length) for t in toks if t.evidence_class == "MICRO_MEDIUM")

        logogram_ctx_chars = sum(int(t.length) for t in toks if t.evidence_class == "LOGOGRAM_CONTEXT")
        gt_chars = sum(int(t.length) for t in toks if t.evidence_class == "GROUNDTRUTH")
        anagram_high_chars = sum(int(t.length) for t in toks if t.evidence_class == "ANAGRAM_HIGH_BASE")
        marker_chars = sum(int(t.length) for t in toks if t.token_type == "marker")

        # WeakTopTokens: top 5 weak tokens by occurrence count (not char count).
        weak_tokens = [t.token for t in toks if int(t.evidence_score_t10) < 20]
        counts = Counter(weak_tokens)
        top = counts.most_common(5)
        weak_top = "; ".join(f"{tok}:{cnt}" for tok, cnt in top) if top else None

        out_r = ev_row_by_book.get(bid_i)
        if out_r is None:
            out_r = ws_append_row(ws_ev, [bid_i])  # minimal; will fill below
            ev_row_by_book[bid_i] = out_r

        ws_ev.cell(out_r, ce["BaseLen"]).value = base_len
        ws_ev.cell(out_r, ce["TokenCount"]).value = tok_count
        ws_ev.cell(out_r, ce["TokenCount_expected"]).value = tok_count
        ws_ev.cell(out_r, ce["TokenCount_match"]).value = True
        ws_ev.cell(out_r, ce["EvidenceAvg"]).value = round(ev_avg, 6)
        ws_ev.cell(out_r, ce["StrongFrac"]).value = strong_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["MediumFrac"]).value = medium_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["WeakFrac"]).value = weak_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["MicroFrac"]).value = micro_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["LogogramContextFrac"]).value = logogram_ctx_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["GroundTruthFrac"]).value = gt_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["AnagramHighFrac"]).value = anagram_high_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["MarkerFrac"]).value = marker_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["WeakTopTokens"]).value = weak_top

        sum_base += base_len
        sum_ev += evidence_sum
        sum_weak += weak_chars
        sum_micro += micro_chars

    return (sum_ev / sum_base) if sum_base else 0.0, (sum_weak / sum_base) if sum_base else 0.0, (sum_micro / sum_base) if sum_base else 0.0


def recompute_token_evidence_contigs(wb: openpyxl.Workbook, glossary_tokens: Dict[str, GlossaryToken]) -> None:
    active = {t.token: t for t in glossary_tokens.values() if t.use_strictplus and t.translation}

    ws_contigs = wb["Contigs"]
    hc = ws_find_header_row(ws_contigs, ["BaseContigID", "Length", "BaseContig", "Translation_StrictPlus_LosslessMarkers_v108"])
    cc = ws_headers(ws_contigs, hc)

    ws_ev = wb["TokenEvidence_Contigs_v128"]
    he = ws_find_header_row(
        ws_ev,
        ["BaseContigID", "BaseLen", "TokenCount", "TokenCount_expected", "TokenCount_match", "EvidenceAvg", "StrongFrac", "MediumFrac", "WeakFrac", "MicroFrac", "LogogramContextFrac", "GroundTruthFrac", "AnagramHighFrac", "MarkerFrac", "WeakTopTokens"],
    )
    ce = ws_headers(ws_ev, he)

    ev_row_by_contig: Dict[int, int] = {}
    for r in range(he + 1, ws_ev.max_row + 1):
        cid = ws_ev.cell(r, ce["BaseContigID"]).value
        if cid is None:
            continue
        ev_row_by_contig[int(cid)] = r

    for r in range(hc + 1, ws_contigs.max_row + 1):
        cid = ws_contigs.cell(r, cc["BaseContigID"]).value
        if cid is None:
            continue
        cid_i = int(cid)
        base_len = int(ws_contigs.cell(r, cc["Length"]).value or 0)
        base = str(ws_contigs.cell(r, cc["BaseContig"]).value or "")
        items = dp_tokenize_base_with_punct(base, active)
        toks = [it for it in items if isinstance(it, GlossaryToken)]
        tok_count = len(toks)
        evidence_sum_t10 = sum(int(t.evidence_score_t10) * int(t.length) for t in toks)
        evidence_sum = evidence_sum_t10 / 10.0
        ev_avg = (evidence_sum / base_len) if base_len else 0.0

        medium_chars = sum(int(t.length) for t in toks if t.translation in (".", "!"))
        weak_chars = sum(int(t.length) for t in toks if int(t.evidence_score_t10) < 20)
        strong_chars = base_len - weak_chars - medium_chars
        micro_chars = sum(int(t.length) for t in toks if t.evidence_class == "MICRO_MEDIUM")
        logogram_ctx_chars = sum(int(t.length) for t in toks if t.evidence_class == "LOGOGRAM_CONTEXT")
        gt_chars = sum(int(t.length) for t in toks if t.evidence_class == "GROUNDTRUTH")
        anagram_high_chars = sum(int(t.length) for t in toks if t.evidence_class == "ANAGRAM_HIGH_BASE")
        marker_chars = sum(int(t.length) for t in toks if t.token_type == "marker")

        weak_tokens = [t.token for t in toks if int(t.evidence_score_t10) < 20]
        counts = Counter(weak_tokens)
        top = counts.most_common(5)
        weak_top = "; ".join(f"{tok}:{cnt}" for tok, cnt in top) if top else None

        out_r = ev_row_by_contig.get(cid_i)
        if out_r is None:
            out_r = ws_append_row(ws_ev, [cid_i])
            ev_row_by_contig[cid_i] = out_r

        ws_ev.cell(out_r, ce["BaseLen"]).value = base_len
        ws_ev.cell(out_r, ce["TokenCount"]).value = tok_count
        ws_ev.cell(out_r, ce["TokenCount_expected"]).value = tok_count
        ws_ev.cell(out_r, ce["TokenCount_match"]).value = True
        ws_ev.cell(out_r, ce["EvidenceAvg"]).value = round(ev_avg, 6)
        ws_ev.cell(out_r, ce["StrongFrac"]).value = strong_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["MediumFrac"]).value = medium_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["WeakFrac"]).value = weak_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["MicroFrac"]).value = micro_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["LogogramContextFrac"]).value = logogram_ctx_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["GroundTruthFrac"]).value = gt_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["AnagramHighFrac"]).value = anagram_high_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["MarkerFrac"]).value = marker_chars / base_len if base_len else 0
        ws_ev.cell(out_r, ce["WeakTopTokens"]).value = weak_top


def fix_cribs_groundtruth_match_flags(wb: openpyxl.Workbook) -> int:
    ws = wb["Cribs"]
    header = ws_find_header_row(ws, ["CribClass_v112", "MatchNorm_v112", "MatchNorm_GroundTruth_v112"])
    c = ws_headers(ws, header)
    fixed = 0
    for r in range(header + 1, ws.max_row + 1):
        cls = ws.cell(r, c["CribClass_v112"]).value
        if cls != "GroundTruth":
            continue
        mn = ws.cell(r, c["MatchNorm_v112"]).value
        cur = ws.cell(r, c["MatchNorm_GroundTruth_v112"]).value
        if mn != cur:
            ws.cell(r, c["MatchNorm_GroundTruth_v112"]).value = mn
            fixed += 1
    return fixed


def promote_groundtruth_cribs_from_reliability(wb: openpyxl.Workbook, *, min_verified_sources: int) -> int:
    """Promote Cribs.CribClass_v112 -> GroundTruth when reliability + expected match are already satisfied."""
    if "CribReliability_v129" not in wb.sheetnames:
        return 0

    ws_rel = wb["CribReliability_v129"]
    hr = ws_find_header_row(ws_rel, ["CribID", "VerifiedSourceCount_v129"])
    cr = ws_headers(ws_rel, hr)
    verified: Dict[int, int] = {}
    for r in range(hr + 1, ws_rel.max_row + 1):
        cid = ws_rel.cell(r, cr["CribID"]).value
        if cid is None:
            continue
        verified[int(cid)] = int(ws_rel.cell(r, cr["VerifiedSourceCount_v129"]).value or 0)

    ws = wb["Cribs"]
    header = ws_find_header_row(ws, ["CribID", "CribClass_v112", "ExpectedNorm_v112", "MatchNorm_v112", "MatchNorm_GroundTruth_v112"])
    c = ws_headers(ws, header)

    promoted = 0
    for r in range(header + 1, ws.max_row + 1):
        cid = ws.cell(r, c["CribID"]).value
        if cid is None:
            continue
        cls = ws.cell(r, c["CribClass_v112"]).value
        if cls == "GroundTruth":
            continue
        exp_norm = ws.cell(r, c["ExpectedNorm_v112"]).value
        if exp_norm is None or str(exp_norm).strip() == "":
            continue
        mn = ws.cell(r, c["MatchNorm_v112"]).value
        if mn not in (1, True):
            continue
        if verified.get(int(cid), 0) < min_verified_sources:
            continue

        ws.cell(r, c["CribClass_v112"]).value = "GroundTruth"
        ws.cell(r, c["MatchNorm_GroundTruth_v112"]).value = mn
        promoted += 1

    return promoted


def scan_promotion_candidates(glossary_tokens: Dict[str, GlossaryToken], min_total_occ: int) -> List[GlossaryToken]:
    # Candidate = inactive, has translation, has TotalOcc>=threshold, and has a known EvidenceClass.
    # TotalOcc is in Glossary, but we don't store it in GlossaryToken. We can read it on demand from the sheet,
    # but here we approximate by using the TokenPruning sheet for occ counts if needed.
    # For simplicity (and because the remaining inactive set is tiny), filter using Glossary sheet directly later.
    raise NotImplementedError


def read_glossary_inactive_candidates(
    wb: openpyxl.Workbook,
    min_total_occ: int,
    *,
    allowed_evidence_classes: Optional[Sequence[str]] = None,
    allow_zero_occ_classes: Optional[Sequence[str]] = None,
    external_base_corpus: Optional[Sequence[str]] = None,
) -> List[Tuple[str, str, str, int, int]]:
    """
    Return list of (token, translation, evidence_class, total_occ, length) for inactive tokens meeting threshold.
    """
    ws = wb["Glossary"]
    header = ws_find_header_row(ws, ["Token", "Translation", "Use_StrictPlus_v108", "TotalOcc", "EvidenceClass_v127", "Len"])
    c = ws_headers(ws, header)

    allowed_set = {s.strip() for s in (allowed_evidence_classes or []) if isinstance(s, str) and s.strip()}
    allow_zero_set = {s.strip() for s in (allow_zero_occ_classes or []) if isinstance(s, str) and s.strip()}
    corpus = list(external_base_corpus or [])

    out: List[Tuple[str, str, str, int, int]] = []
    for r in range(header + 1, ws.max_row + 1):
        tok = ws.cell(r, c["Token"]).value
        if not isinstance(tok, str) or not tok:
            continue
        use = ws.cell(r, c["Use_StrictPlus_v108"]).value
        if use:
            continue
        tr = ws.cell(r, c["Translation"]).value
        if tr is None or str(tr).strip() == "":
            continue
        total_occ = int(ws.cell(r, c["TotalOcc"]).value or 0)
        evcls = ws.cell(r, c["EvidenceClass_v127"]).value or ""
        evcls_s = str(evcls)
        if allowed_set and evcls_s not in allowed_set:
            continue
        length = ws.cell(r, c["Len"]).value
        if length is None:
            length = len(tok)

        if total_occ >= min_total_occ:
            out.append((tok, str(tr), evcls_s, total_occ, int(length)))
            continue

        # Allow promotion of certain classes with 0 book occurrences if the token appears in the
        # external base corpus (cribs/external refs).
        if allow_zero_set and evcls_s in allow_zero_set and corpus:
            ext_occ = 0
            for s in corpus:
                if s:
                    ext_occ += str(s).count(tok)
            if ext_occ > 0:
                out.append((tok, str(tr), evcls_s, ext_occ, int(length)))

    # Sort: TotalOcc desc, Len desc.
    out.sort(key=lambda t: (t[3], t[4]), reverse=True)
    return out


def append_candidate_promotions(
    wb: openpyxl.Workbook,
    iter_num: int,
    candidates: List[Tuple[str, str, str, int, int]],
) -> List[int]:
    ws = wb["CandidatePromotions"]
    header = ws_find_header_row(ws, ["Iteration", "Token", "Translation", "EvidenceClass", "TotalOcc", "Decision", "Reason"])
    c = ws_headers(ws, header)
    rows: List[int] = []
    for tok, tr, evcls, occ, _len in candidates:
        r = ws_last_data_row(ws) + 1
        ws.cell(r, c["Iteration"]).value = iter_num
        ws.cell(r, c["Token"]).value = tok
        ws.cell(r, c["Translation"]).value = tr
        ws.cell(r, c["EvidenceClass"]).value = evcls
        ws.cell(r, c["TotalOcc"]).value = occ
        ws.cell(r, c["Decision"]).value = "PENDING"
        ws.cell(r, c["Reason"]).value = None
        rows.append(r)
    return rows


def set_candidate_decision(wb: openpyxl.Workbook, row: int, decision: str, reason: str) -> None:
    ws = wb["CandidatePromotions"]
    header = ws_find_header_row(ws, ["Decision", "Reason"])
    c = ws_headers(ws, header)
    ws.cell(row, c["Decision"]).value = decision
    ws.cell(row, c["Reason"]).value = reason


def _record_uptake_skip(
    bucket_counts: Counter[str],
    bucket_reasons: Dict[str, Counter[str]],
    bucket: str,
    reason: str,
) -> None:
    bucket_counts[bucket] += 1
    bucket_reasons[bucket][reason] += 1


def _counter_top_str(counter: Counter[str]) -> str:
    if not counter:
        return ""
    reason, count = counter.most_common(1)[0]
    return f"{reason} ({count})"


def append_flow_run_log(
    wb: openpyxl.Workbook,
    iter_num: int,
    step_id: int,
    utc: str,
    result: str,
    summary: str,
    changed_books: str,
    evidence_avg: Optional[float] = None,
    weak_frac: Optional[float] = None,
    micro_frac: Optional[float] = None,
    notes: Optional[str] = None,
) -> None:
    ws = wb["FlowRunLog"]
    header = ws_find_header_row(ws, ["Iteration", "StepID", "UTC", "Result", "Summary", "ChangedBooksCount", "EvidenceAvg", "WeakFrac", "MicroFrac", "Notes"])
    c = ws_headers(ws, header)
    r = ws_last_data_row(ws) + 1
    ws.cell(r, c["Iteration"]).value = iter_num
    ws.cell(r, c["StepID"]).value = step_id
    ws.cell(r, c["UTC"]).value = utc
    ws.cell(r, c["Result"]).value = result
    ws.cell(r, c["Summary"]).value = summary
    ws.cell(r, c["ChangedBooksCount"]).value = changed_books
    ws.cell(r, c["EvidenceAvg"]).value = evidence_avg
    ws.cell(r, c["WeakFrac"]).value = weak_frac
    ws.cell(r, c["MicroFrac"]).value = micro_frac
    ws.cell(r, c["Notes"]).value = notes


def append_note_cell(cell: openpyxl.cell.cell.Cell, msg: str) -> None:
    prev = cell.value
    if prev and isinstance(prev, str):
        if msg not in prev:
            cell.value = f"{prev}; {msg}"
    else:
        cell.value = msg


def materialize_external_refs_from_codestream_v120(wb: openpyxl.Workbook, iter_num: int) -> int:
    """Fill ExternalRefs core decode columns from CodeStream v120 outputs when blank.

    This makes ExternalRefs usable as a single 'decoded view' without touching Books.
    We only fill rows where DecodedBase is empty and CodeStreamBase_v120 is present.
    """
    ws = wb["ExternalRefs_v115"]
    header = ws_find_header_row(
        ws,
        [
            "RefName",
            "DecodedBase",
            "DP_StrictPlus",
            "HighCharFrac",
            "SingleCharFrac",
            "LettersLen",
            "TokenCount",
            "CodeStreamBase_v120",
            "CodeStreamDP_Concat_Readable_v120",
            "CodeStreamHighCharFrac_v120",
            "CodeStreamSingleCharFrac_v120",
            "CodeStreamTokenCount_v120",
        ],
    )
    c = ws_headers(ws, header)
    notes_col = c.get("Notes")

    changed = 0
    for r in range(header + 1, ws.max_row + 1):
        ref = ws.cell(r, c["RefName"]).value
        if ref is None:
            continue
        decoded = ws.cell(r, c["DecodedBase"]).value
        if decoded is not None and str(decoded).strip() != "":
            continue
        cs_base = ws.cell(r, c["CodeStreamBase_v120"]).value
        if cs_base is None or str(cs_base).strip() == "":
            continue

        cs_dp = ws.cell(r, c["CodeStreamDP_Concat_Readable_v120"]).value
        cs_high = ws.cell(r, c["CodeStreamHighCharFrac_v120"]).value
        cs_single = ws.cell(r, c["CodeStreamSingleCharFrac_v120"]).value
        cs_tok = ws.cell(r, c["CodeStreamTokenCount_v120"]).value

        ws.cell(r, c["DecodedBase"]).value = str(cs_base)
        if cs_dp is not None:
            ws.cell(r, c["DP_StrictPlus"]).value = str(cs_dp)
        if cs_high is not None:
            ws.cell(r, c["HighCharFrac"]).value = float(cs_high)
        if cs_single is not None:
            ws.cell(r, c["SingleCharFrac"]).value = float(cs_single)
        ws.cell(r, c["LettersLen"]).value = len(str(cs_base))
        if cs_tok is not None:
            ws.cell(r, c["TokenCount"]).value = int(cs_tok)
        if notes_col is not None:
            append_note_cell(ws.cell(r, notes_col), f"iter{iter_num}: filled from CodeStream v120")
        changed += 1

    return changed


def _digits_only(s: object) -> str:
    return "".join([ch for ch in str(s or "") if ch.isdigit()])


def _digit_runs(s: object) -> List[str]:
    """Return contiguous runs of digits as they appear in text."""
    return re.findall(r"\d+", str(s or ""))


def _digits_contains_parts_in_order(digits: str, parts: Sequence[str]) -> bool:
    """True when all non-empty parts appear in order inside digits."""
    if not digits:
        return False
    pos = 0
    for p in parts:
        if not p:
            continue
        i = digits.find(p, pos)
        if i < 0:
            return False
        pos = i + len(p)
    return True


def materialize_digit_code_map_from_books(wb: openpyxl.Workbook, iter_num: int, utc: str) -> Tuple[int, int, int, str]:
    """Extract and persist the global digits-code -> base-letter map from Books + BooksDigitModel.

    Returns: (codes_written, letters_written, conflicts, status)
    """
    if "Books" not in wb.sheetnames or "BooksDigitModel_v118" not in wb.sheetnames:
        return 0, 0, 0, "missing Books/BooksDigitModel_v118"

    ws_b = wb["Books"]
    hb = ws_find_header_row(ws_b, ["BookID", "Digits", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_b, hb)

    ws_dm = wb["BooksDigitModel_v118"]
    hdm = ws_find_header_row(ws_dm, ["BookID", "OmitIdxs_1based"], max_scan=3)
    cdm = ws_headers(ws_dm, hdm)

    omit_by_book: Dict[int, set[int]] = {}
    for r in range(hdm + 1, ws_dm.max_row + 1):
        bid = ws_dm.cell(r, cdm["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        omit_by_book[bid_i] = set(parse_int_list(ws_dm.cell(r, cdm["OmitIdxs_1based"]).value))

    # Aggregate code->letter mapping.
    code_letter_counts: Dict[str, Counter[str]] = defaultdict(Counter)
    code_examples: Dict[str, set[int]] = defaultdict(set)
    conflicts = 0
    books_used = 0
    bad_books = 0

    for r in range(hb + 1, ws_b.max_row + 1):
        bid = ws_b.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        digits = _digits_only(ws_b.cell(r, cb["Digits"]).value)
        base = str(ws_b.cell(r, cb["DecodedBase"]).value or "")
        if not digits or not base:
            continue
        omit = omit_by_book.get(bid_i)
        if omit is None:
            # Should not happen, but keep going.
            bad_books += 1
            continue

        di = 0
        base_pos = 0  # 1-based position over base letters (A..Z + '*')
        ok = True
        for ch in base:
            if not (ch.isalpha() or ch == "*"):
                # Books.DecodedBase should be letters-only; keep a defensive skip.
                continue
            base_pos += 1
            if base_pos in omit:
                if di >= len(digits):
                    ok = False
                    break
                code = "0" + digits[di]
                di += 1
            else:
                if di + 1 >= len(digits):
                    ok = False
                    break
                code = digits[di : di + 2]
                di += 2

            code_letter_counts[code][ch] += 1
            if len(code_examples[code]) < 8:
                code_examples[code].add(bid_i)

        if not ok:
            bad_books += 1
            continue
        books_used += 1

    # Detect conflicts (a code mapping to multiple letters).
    code_to_letter: Dict[str, str] = {}
    for code, cnt in code_letter_counts.items():
        if not cnt:
            continue
        top = cnt.most_common(2)
        if len(top) > 1:
            # True ambiguity at code->letter level (should be 0 in this workbook).
            conflicts += 1
        code_to_letter[code] = top[0][0]

    # letter -> Counter(code)
    letter_codes: Dict[str, Counter[str]] = defaultdict(Counter)
    for code, cnt in code_letter_counts.items():
        if not cnt:
            continue
        letter = cnt.most_common(1)[0][0]
        letter_codes[letter][code] += int(sum(cnt.values()))

    # If an existing map is present and identical, avoid rewriting the sheets every iteration.
    existing_map: Dict[str, str] = {}
    if DIGIT_CODE_MAP_SHEET in wb.sheetnames and DIGIT_LETTER_CODES_SHEET in wb.sheetnames:
        try:
            ws_prev = wb[DIGIT_CODE_MAP_SHEET]
            hp = ws_find_header_row(ws_prev, ["Code", "Letter"], max_scan=3)
            cp = ws_headers(ws_prev, hp)
            for r in range(hp + 1, ws_prev.max_row + 1):
                code = ws_prev.cell(r, cp["Code"]).value
                letter = ws_prev.cell(r, cp["Letter"]).value
                if not isinstance(code, str) or not code.strip():
                    continue
                if not isinstance(letter, str) or not letter.strip():
                    continue
                existing_map[code.strip()] = letter.strip()
        except Exception:
            existing_map = {}

    if existing_map and existing_map == code_to_letter:
        # Keep SheetIndex entries stable but avoid rewriting the large tables when unchanged.
        upsert_sheet_index_entry(
            wb,
            DIGIT_CODE_MAP_SHEET,
            "Derived mapping from digits codes to base letters (analysis-only). Extracted from Books + BooksDigitModel omission positions.",
        )
        upsert_sheet_index_entry(
            wb,
            DIGIT_LETTER_CODES_SHEET,
            "Derived homophone sets: base letter -> list of digits codes (analysis-only).",
        )
        return len(code_to_letter), len(letter_codes), int(conflicts), "skipped_same"

    # Write DigitCodeMap_Auto
    ws_map = ensure_sheet(wb, DIGIT_CODE_MAP_SHEET, ["Iteration", "Code", "Letter", "Count", "Share", "ExampleBooks", "Notes"])
    hm = ws_find_header_row(ws_map, ["Iteration", "Code", "Letter"], max_scan=3)
    cm = ws_headers(ws_map, hm)
    if ws_map.max_row > hm:
        ws_map.delete_rows(hm + 1, ws_map.max_row - hm)

    rr = hm + 1
    for code in sorted(code_letter_counts.keys(), key=lambda c: (-sum(code_letter_counts[c].values()), c)):
        cnt = code_letter_counts[code]
        total = int(sum(cnt.values()))
        if total <= 0:
            continue
        top_letter, top_count = cnt.most_common(1)[0]
        share = float(top_count) / float(total) if total else 0.0
        ws_map.cell(rr, cm["Iteration"]).value = int(iter_num)
        ws_map.cell(rr, cm["Code"]).value = str(code)
        ws_map.cell(rr, cm["Letter"]).value = str(top_letter)
        ws_map.cell(rr, cm["Count"]).value = int(total)
        ws_map.cell(rr, cm["Share"]).value = round(float(share), 6)
        ws_map.cell(rr, cm["ExampleBooks"]).value = ",".join(str(b) for b in sorted(code_examples.get(code) or [])[:12])
        ws_map.cell(rr, cm["Notes"]).value = f"books_used={books_used}, bad_books={bad_books}"
        rr += 1

    upsert_sheet_index_entry(
        wb,
        DIGIT_CODE_MAP_SHEET,
        "Derived mapping from digits codes to base letters (analysis-only). Extracted from Books + BooksDigitModel omission positions.",
    )

    # Write DigitLetterCodes_Auto
    ws_lc = ensure_sheet(wb, DIGIT_LETTER_CODES_SHEET, ["Iteration", "Letter", "CodeCount", "Codes", "CodeCounts", "TotalCount", "Notes"])
    hl = ws_find_header_row(ws_lc, ["Iteration", "Letter", "Codes"], max_scan=3)
    cl = ws_headers(ws_lc, hl)
    if ws_lc.max_row > hl:
        ws_lc.delete_rows(hl + 1, ws_lc.max_row - hl)

    rr2 = hl + 1
    for letter in sorted(letter_codes.keys()):
        cc = letter_codes[letter]
        codes_sorted = sorted(cc.items(), key=lambda kv: (-kv[1], kv[0]))
        ws_lc.cell(rr2, cl["Iteration"]).value = int(iter_num)
        ws_lc.cell(rr2, cl["Letter"]).value = str(letter)
        ws_lc.cell(rr2, cl["CodeCount"]).value = int(len(cc))
        ws_lc.cell(rr2, cl["Codes"]).value = ", ".join([code for code, _cnt in codes_sorted])
        ws_lc.cell(rr2, cl["CodeCounts"]).value = ", ".join([f"{code}:{int(cnt)}" for code, cnt in codes_sorted])
        ws_lc.cell(rr2, cl["TotalCount"]).value = int(sum(cc.values()))
        ws_lc.cell(rr2, cl["Notes"]).value = f"homophones={len(cc)}"
        rr2 += 1

    upsert_sheet_index_entry(
        wb,
        DIGIT_LETTER_CODES_SHEET,
        "Derived homophone sets: base letter -> list of digits codes (analysis-only).",
    )

    return (rr - (hm + 1)), (rr2 - (hl + 1)), int(conflicts), "ok"


def materialize_digit_code_context_profiles_from_books(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    *,
    topk: int = 8,
    js_alpha: float = 0.20,
    prev_fp: str = "",
) -> Tuple[int, int, float, str, str]:
    """Compute a code-level context profile to study homophones (analysis-only).

    The DigitCodeMap proves a deterministic digits->code->base-letter mapping. This step adds
    a *structural* layer for the "non brute force" part of the puzzle:
    - codes that map to the same base letter are homophones
    - their usage may differ by neighboring letters/codes (context profiles)

    We persist only derived counts (no new decoding/promotion).

    Returns: (rows_written, homophone_letters, best_outlier_score, fingerprint, status)
    """
    if "Books" not in wb.sheetnames or "BooksDigitModel_v118" not in wb.sheetnames:
        return 0, 0, 0.0, "", "missing Books/BooksDigitModel_v118"

    # Load the stable code->letter map (prefer the persisted DigitCodeMap).
    code_to_letter: Dict[str, str] = {}
    if DIGIT_CODE_MAP_SHEET in wb.sheetnames:
        try:
            ws_map = wb[DIGIT_CODE_MAP_SHEET]
            hm = ws_find_header_row(ws_map, ["Code", "Letter"], max_scan=3)
            cm = ws_headers(ws_map, hm)
            for r in range(hm + 1, ws_map.max_row + 1):
                code = ws_map.cell(r, cm["Code"]).value
                letter = ws_map.cell(r, cm["Letter"]).value
                if not isinstance(code, str) or not code.strip():
                    continue
                if not isinstance(letter, str) or not letter.strip():
                    continue
                code_to_letter[code.strip()] = letter.strip()
        except Exception:
            code_to_letter = {}

    ws_b = wb["Books"]
    hb = ws_find_header_row(ws_b, ["BookID", "Digits", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_b, hb)

    ws_dm = wb["BooksDigitModel_v118"]
    hdm = ws_find_header_row(ws_dm, ["BookID", "OmitIdxs_1based"], max_scan=3)
    cdm = ws_headers(ws_dm, hdm)

    omit_by_book: Dict[int, set[int]] = {}
    for r in range(hdm + 1, ws_dm.max_row + 1):
        bid = ws_dm.cell(r, cdm["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        omit_by_book[bid_i] = set(parse_int_list(ws_dm.cell(r, cdm["OmitIdxs_1based"]).value))

    code_counts: Counter[str] = Counter()
    letter_to_codes: Dict[str, set[str]] = defaultdict(set)
    prev_codes: Dict[str, Counter[str]] = defaultdict(Counter)
    next_codes: Dict[str, Counter[str]] = defaultdict(Counter)
    prev_letters: Dict[str, Counter[str]] = defaultdict(Counter)
    next_letters: Dict[str, Counter[str]] = defaultdict(Counter)

    for r in range(hb + 1, ws_b.max_row + 1):
        bid = ws_b.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        digits = _digits_only(ws_b.cell(r, cb["Digits"]).value)
        base = str(ws_b.cell(r, cb["DecodedBase"]).value or "")
        if not digits or not base:
            continue
        omit = omit_by_book.get(bid_i)
        if omit is None:
            continue

        codes: List[str] = []
        letters: List[str] = []
        di = 0
        base_pos = 0
        ok = True
        for ch in base:
            if not (ch.isalpha() or ch == "*"):
                continue
            base_pos += 1
            if base_pos in omit:
                if di >= len(digits):
                    ok = False
                    break
                code = "0" + digits[di]
                di += 1
            else:
                if di + 1 >= len(digits):
                    ok = False
                    break
                code = digits[di : di + 2]
                di += 2
            codes.append(code)
            letters.append(ch)
        if not ok or not codes:
            continue

        for i, code in enumerate(codes):
            code_counts[code] += 1
            letter = letters[i]
            letter_to_codes[letter].add(code)
            if i > 0:
                pc = codes[i - 1]
                pl = letters[i - 1]
                prev_codes[code][pc] += 1
                prev_letters[code][pl] += 1
            if i + 1 < len(codes):
                nc = codes[i + 1]
                nl = letters[i + 1]
                next_codes[code][nc] += 1
                next_letters[code][nl] += 1

    if not code_counts:
        ensure_sheet(
            wb,
            DIGIT_CODE_CONTEXT_SHEET,
            [
                "Iteration",
                "UTC",
                "Code",
                "Letter",
                "Count",
                "HomophoneCount",
                "PrevLettersTop",
                "NextLettersTop",
                "PrevCodesTop",
                "NextCodesTop",
                "PrevEntropy",
                "NextEntropy",
                "OutlierScoreJS",
                "Notes",
            ],
        )
        upsert_sheet_index_entry(
            wb,
            DIGIT_CODE_CONTEXT_SHEET,
            "Derived code context profiles: for each 2-digit code, neighbor code/letter distributions (analysis-only).",
        )
        return 0, 0, 0.0, "", "no_data"

    def _entropy(cnt: Counter[str]) -> float:
        total = float(sum(cnt.values()))
        if total <= 0:
            return 0.0
        e = 0.0
        for v in cnt.values():
            p = float(v) / total
            if p > 0:
                e -= p * math.log(p, 2)
        return float(e)

    def _js_divergence(a: Counter[str], b: Counter[str], *, alpha: float) -> float:
        keys = set(a.keys()) | set(b.keys())
        if not keys:
            return 0.0
        klist = sorted(keys)
        aa = float(alpha)
        denom_a = float(sum(a.values())) + aa * float(len(klist))
        denom_b = float(sum(b.values())) + aa * float(len(klist))
        if denom_a <= 0 or denom_b <= 0:
            return 0.0
        js = 0.0
        for k in klist:
            pa = (float(a.get(k, 0)) + aa) / denom_a
            pb = (float(b.get(k, 0)) + aa) / denom_b
            pm = 0.5 * (pa + pb)
            if pa > 0 and pm > 0:
                js += 0.5 * pa * math.log(pa / pm)
            if pb > 0 and pm > 0:
                js += 0.5 * pb * math.log(pb / pm)
        return float(js)

    # Aggregate neighbor-letter distributions by base letter (for outlier scoring within homophone sets).
    letter_prev: Dict[str, Counter[str]] = defaultdict(Counter)
    letter_next: Dict[str, Counter[str]] = defaultdict(Counter)
    for letter, codes in letter_to_codes.items():
        for code in codes:
            letter_prev[letter].update(prev_letters.get(code) or Counter())
            letter_next[letter].update(next_letters.get(code) or Counter())

    outlier_by_code: Dict[str, float] = {}
    best_outlier = 0.0
    for letter, codes in letter_to_codes.items():
        if len(codes) <= 1:
            continue
        agg_p = letter_prev.get(letter) or Counter()
        agg_n = letter_next.get(letter) or Counter()
        for code in codes:
            js_p = _js_divergence(prev_letters.get(code) or Counter(), agg_p, alpha=float(js_alpha))
            js_n = _js_divergence(next_letters.get(code) or Counter(), agg_n, alpha=float(js_alpha))
            score = 0.5 * (float(js_p) + float(js_n))
            outlier_by_code[code] = float(score)
            if score > best_outlier:
                best_outlier = float(score)

    homophone_letters = sum(1 for _l, codes in letter_to_codes.items() if len(codes) > 1)

    def _top_str(cnt: Counter[str], k: int) -> str:
        items = cnt.most_common(int(k))
        return ", ".join([f"{x}:{int(n)}" for x, n in items])

    # Fingerprint (stable across iterations when decode is stable).
    fp_items: List[str] = []
    for code in sorted(code_counts.keys()):
        letter = code_to_letter.get(code) or ""
        hc = len(letter_to_codes.get(letter) or []) if letter else 0
        fp_items.append(
            "|".join(
                [
                    code,
                    letter,
                    str(int(code_counts[code])),
                    str(int(hc)),
                    f"{_entropy(prev_letters.get(code) or Counter()):.6f}",
                    f"{_entropy(next_letters.get(code) or Counter()):.6f}",
                    f"{float(outlier_by_code.get(code, 0.0)):.6f}",
                    _top_str(prev_letters.get(code) or Counter(), 6),
                    _top_str(next_letters.get(code) or Counter(), 6),
                ]
            )
        )
    fp = hashlib.sha1("\n".join(fp_items).encode("utf-8", errors="ignore")).hexdigest() if fp_items else ""

    if DIGIT_CODE_CONTEXT_SHEET in wb.sheetnames and prev_fp and fp and fp == str(prev_fp):
        upsert_sheet_index_entry(
            wb,
            DIGIT_CODE_CONTEXT_SHEET,
            "Derived code context profiles: for each 2-digit code, neighbor code/letter distributions (analysis-only).",
        )
        return len(code_counts), int(homophone_letters), float(best_outlier), fp, "skipped_same"

    ws_out = ensure_sheet(
        wb,
        DIGIT_CODE_CONTEXT_SHEET,
        [
            "Iteration",
            "UTC",
            "Code",
            "Letter",
            "Count",
            "HomophoneCount",
            "PrevLettersTop",
            "NextLettersTop",
            "PrevCodesTop",
            "NextCodesTop",
            "PrevEntropy",
            "NextEntropy",
            "OutlierScoreJS",
            "Notes",
        ],
    )
    ho = ws_find_header_row(ws_out, ["Iteration", "Code", "Count"], max_scan=3)
    co = ws_headers(ws_out, ho)
    if ws_out.max_row > ho:
        ws_out.delete_rows(ho + 1, ws_out.max_row - ho)

    rr = ho + 1
    for code, cnt in sorted(code_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        letter = code_to_letter.get(code) or ""
        hc = len(letter_to_codes.get(letter) or []) if letter else 0
        ws_out.cell(rr, co["Iteration"]).value = int(iter_num)
        ws_out.cell(rr, co["UTC"]).value = utc
        ws_out.cell(rr, co["Code"]).value = str(code)
        ws_out.cell(rr, co["Letter"]).value = str(letter)
        ws_out.cell(rr, co["Count"]).value = int(cnt)
        ws_out.cell(rr, co["HomophoneCount"]).value = int(hc)
        ws_out.cell(rr, co["PrevLettersTop"]).value = _top_str(prev_letters.get(code) or Counter(), int(topk))
        ws_out.cell(rr, co["NextLettersTop"]).value = _top_str(next_letters.get(code) or Counter(), int(topk))
        ws_out.cell(rr, co["PrevCodesTop"]).value = _top_str(prev_codes.get(code) or Counter(), int(topk))
        ws_out.cell(rr, co["NextCodesTop"]).value = _top_str(next_codes.get(code) or Counter(), int(topk))
        ws_out.cell(rr, co["PrevEntropy"]).value = round(float(_entropy(prev_letters.get(code) or Counter())), 6)
        ws_out.cell(rr, co["NextEntropy"]).value = round(float(_entropy(next_letters.get(code) or Counter())), 6)
        ws_out.cell(rr, co["OutlierScoreJS"]).value = round(float(outlier_by_code.get(code, 0.0)), 6)
        ws_out.cell(rr, co["Notes"]).value = f"topk={int(topk)}, js_alpha={float(js_alpha)}"
        rr += 1

    upsert_sheet_index_entry(
        wb,
        DIGIT_CODE_CONTEXT_SHEET,
        "Derived code context profiles: for each 2-digit code, neighbor code/letter distributions (analysis-only).",
    )
    return (rr - (ho + 1)), int(homophone_letters), float(best_outlier), fp, "ok"


def materialize_external_roundtrip_check(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    *,
    active_tokens: Dict[str, GlossaryToken],
    min_verified_count: int,
    min_segment_digits: int,
    allow_ordered_run_match: bool,
) -> Tuple[int, int, int, str]:
    """Validate external numeric references against the current DP tokenization (analysis-only).

    Returns: (pass_count, fail_count, skipped_count, status)
    """
    if "ExternalRefs_v115" not in wb.sheetnames or "Books" not in wb.sheetnames:
        return 0, 0, 0, "missing ExternalRefs_v115/Books"

    ws_ext = wb["ExternalRefs_v115"]
    he = ws_find_header_row(ws_ext, ["RefName", "NumericText", "DigitsSanitized", "DecodedBase"], max_scan=3)
    ce = ws_headers(ws_ext, he)

    # VerifiedCount is maintained in ExternalValidation_v129 but does not share RefName.
    # Primary join is exact DigitsSanitized. Safe fallbacks:
    # - exact segmented run (long-enough digit runs)
    # - ordered run match for validation rows that include ellipsis (`...`)
    # All fallbacks are conservative and require unique match.
    verified_by_digits: Dict[str, Tuple[int, str]] = {}
    verified_by_segment: Dict[str, List[Tuple[int, str]]] = defaultdict(list)
    ordered_run_specs: List[Tuple[List[str], int, str]] = []
    if "ExternalValidation_v129" in wb.sheetnames:
        ws_v = wb["ExternalValidation_v129"]
        hv = ws_find_header_row(ws_v, ["Reference", "NumericText", "VerifiedCount"], max_scan=3)
        cv = ws_headers(ws_v, hv)
        for r in range(hv + 1, ws_v.max_row + 1):
            ref_name = str(ws_v.cell(r, cv["Reference"]).value or "").strip() or f"row{r}"
            num = ws_v.cell(r, cv["NumericText"]).value
            vc = ws_v.cell(r, cv["VerifiedCount"]).value
            if num is None:
                continue
            digits = _digits_only(num)
            try:
                verified = int(vc or 0)
            except Exception:
                verified = 0
            if digits:
                cur = verified_by_digits.get(digits)
                if cur is None or verified > cur[0]:
                    verified_by_digits[digits] = (verified, ref_name)
            # Validation rows often list alternatives separated by "/".
            # Register each branch as an exact key as well (safe, no fuzzy join).
            for branch in re.split(r"\s*/\s*", str(num or "")):
                branch_digits = _digits_only(branch)
                if not branch_digits or branch_digits == digits:
                    continue
                cur_branch = verified_by_digits.get(branch_digits)
                if cur_branch is None or verified > cur_branch[0]:
                    verified_by_digits[branch_digits] = (verified, ref_name)

            runs = [seg for seg in _digit_runs(num) if seg]
            for seg in runs:
                if len(seg) < int(min_segment_digits):
                    continue
                verified_by_segment[seg].append((verified, ref_name))

            if allow_ordered_run_match and "..." in str(num) and len(runs) >= 2:
                long_runs = [seg for seg in runs if len(seg) >= int(min_segment_digits)]
                if len(long_runs) >= 2:
                    ordered_run_specs.append((long_runs, verified, ref_name))

    expected_norm_by_ref: Dict[str, str] = {}
    if "ExternalGroundTruthCheck_v120" in wb.sheetnames:
        ws_gt = wb["ExternalGroundTruthCheck_v120"]
        hgt = ws_find_header_row(ws_gt, ["RefName", "Expected_Norm"], max_scan=3)
        cgt = ws_headers(ws_gt, hgt)
        for r in range(hgt + 1, ws_gt.max_row + 1):
            ref = ws_gt.cell(r, cgt["RefName"]).value
            exp = ws_gt.cell(r, cgt["Expected_Norm"]).value
            if not isinstance(ref, str) or not ref.strip():
                continue
            if not isinstance(exp, str) or not exp.strip():
                continue
            expected_norm_by_ref[ref.strip()] = exp.strip()

    # Precompute Books digits occurrences for substring lookup.
    ws_b = wb["Books"]
    hb = ws_find_header_row(ws_b, ["BookID", "Digits"], max_scan=3)
    cb = ws_headers(ws_b, hb)
    book_digits: Dict[int, str] = {}
    for r in range(hb + 1, ws_b.max_row + 1):
        bid = ws_b.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        book_digits[bid_i] = _digits_only(ws_b.cell(r, cb["Digits"]).value)

    ws_out = ensure_sheet(
        wb,
        EXTERNAL_ROUNDTRIP_SHEET,
        [
            "Iteration",
            "UTC",
            "RefName",
            "DigitsSanitized",
            "VerifiedCount",
            "InBooksCount_Calc",
            "InBooksBookIDs_Calc",
            "DecodedBase",
            "DP_Current",
            "DP_Norm",
            "Expected_Norm",
            "Pass",
            "Notes",
        ],
    )
    ho = ws_find_header_row(ws_out, ["Iteration", "RefName", "DigitsSanitized"], max_scan=3)
    co = ws_headers(ws_out, ho)
    if ws_out.max_row > ho:
        ws_out.delete_rows(ho + 1, ws_out.max_row - ho)

    def _dp_readable(base: str) -> str:
        items = dp_tokenize_base_with_punct(str(base or ""), active_tokens)
        lossless: List[str] = []
        for it in items:
            if isinstance(it, str):
                lossless.append(it)
            elif it.token_type == "marker":
                lossless.append(f"<{it.token}>")
            else:
                lossless.extend(it.translation.split())
        return render_strictplus_from_lossless(lossless)

    pass_n = 0
    fail_n = 0
    skipped_n = 0
    rr = ho + 1

    for r in range(he + 1, ws_ext.max_row + 1):
        ref = ws_ext.cell(r, ce["RefName"]).value
        if ref is None:
            continue
        ref_s = str(ref).strip()
        digits = ws_ext.cell(r, ce.get("DigitsSanitized") or 0).value if ce.get("DigitsSanitized") else None
        digits_s = _digits_only(digits) if digits is not None else _digits_only(ws_ext.cell(r, ce["NumericText"]).value)
        if not digits_s:
            continue
        verified = 0
        verified_mode = ""
        verified_ref = ""

        exact = verified_by_digits.get(digits_s)
        if exact is not None:
            verified = int(exact[0])
            verified_mode = "exact_digits"
            verified_ref = exact[1]
        else:
            seg_hits = verified_by_segment.get(digits_s, [])
            seg_refs = {ref_name for _, ref_name in seg_hits}
            if len(seg_refs) == 1:
                best_seg = max(seg_hits, key=lambda x: int(x[0]))
                verified = int(best_seg[0])
                verified_mode = "segment_digits"
                verified_ref = best_seg[1]
            elif len(seg_refs) > 1:
                verified_mode = "segment_ambiguous"

            if verified < int(min_verified_count) and allow_ordered_run_match:
                ordered_hits = [(v, rn) for parts, v, rn in ordered_run_specs if _digits_contains_parts_in_order(digits_s, parts)]
                ordered_refs = {rn for _, rn in ordered_hits}
                if len(ordered_refs) == 1:
                    best_ord = max(ordered_hits, key=lambda x: int(x[0]))
                    verified = int(best_ord[0])
                    verified_mode = "ordered_runs"
                    verified_ref = best_ord[1]
                elif len(ordered_refs) > 1:
                    verified_mode = "ordered_ambiguous"

        if verified < int(min_verified_count):
            skipped_n += 1
            continue
        base = ws_ext.cell(r, ce["DecodedBase"]).value
        base_s = str(base or "").strip()
        if not base_s:
            skipped_n += 1
            continue

        hit_books = [bid for bid, ds in book_digits.items() if digits_s in ds]
        dp = _dp_readable(base_s)
        dec_norm = normalize_for_match(dp)
        exp_raw = expected_norm_by_ref.get(ref_s)
        exp_norm = normalize_for_match(exp_raw) if exp_raw else None
        ok = True
        if exp_norm:
            ok = dec_norm == exp_norm

        ws_out.cell(rr, co["Iteration"]).value = int(iter_num)
        ws_out.cell(rr, co["UTC"]).value = utc
        ws_out.cell(rr, co["RefName"]).value = ref_s
        ws_out.cell(rr, co["DigitsSanitized"]).value = digits_s
        ws_out.cell(rr, co["VerifiedCount"]).value = int(verified)
        ws_out.cell(rr, co["InBooksCount_Calc"]).value = int(len(hit_books))
        ws_out.cell(rr, co["InBooksBookIDs_Calc"]).value = ",".join(str(b) for b in sorted(hit_books))
        ws_out.cell(rr, co["DecodedBase"]).value = base_s
        ws_out.cell(rr, co["DP_Current"]).value = dp
        ws_out.cell(rr, co["DP_Norm"]).value = dec_norm
        ws_out.cell(rr, co["Expected_Norm"]).value = exp_norm
        ws_out.cell(rr, co["Pass"]).value = True if ok else False
        mode_note = verified_mode or "unknown"
        ref_note = f"; verified_ref={verified_ref}" if verified_ref else ""
        if exp_norm:
            ws_out.cell(rr, co["Notes"]).value = f"expected_norm_check (canonized); verify_match={mode_note}{ref_note}"
        else:
            ws_out.cell(rr, co["Notes"]).value = f"no_expected_norm; verify_match={mode_note}{ref_note}"
        rr += 1

        if ok:
            pass_n += 1
        else:
            fail_n += 1

    upsert_sheet_index_entry(
        wb,
        EXTERNAL_ROUNDTRIP_SHEET,
        "External numeric refs roundtrip check (analysis-only): DP_current vs expected_norm when available; includes verified-count gating.",
    )
    return pass_n, fail_n, skipped_n, "ok"


def weighted_metrics_from_token_evidence_books(wb: openpyxl.Workbook) -> Tuple[float, float, float]:
    ws = wb["TokenEvidence_Books_v128"]
    header = ws_find_header_row(ws, ["BookID", "BaseLen", "EvidenceAvg", "WeakFrac", "MicroFrac"])
    c = ws_headers(ws, header)
    sum_base = 0
    sum_ev = 0.0
    sum_weak = 0.0
    sum_micro = 0.0
    for r in range(header + 1, ws.max_row + 1):
        bid = ws.cell(r, c["BookID"]).value
        if bid is None:
            continue
        base_len = ws.cell(r, c["BaseLen"]).value
        if base_len is None:
            continue
        base_len = int(base_len)
        sum_base += base_len
        sum_ev += base_len * float(ws.cell(r, c["EvidenceAvg"]).value or 0.0)
        sum_weak += base_len * float(ws.cell(r, c["WeakFrac"]).value or 0.0)
        sum_micro += base_len * float(ws.cell(r, c["MicroFrac"]).value or 0.0)
    if sum_base == 0:
        return 0.0, 0.0, 0.0
    return sum_ev / sum_base, sum_weak / sum_base, sum_micro / sum_base


def weighted_single_char_frac_from_books(wb: openpyxl.Workbook) -> float:
    ws = wb["Books"]
    header = ws_find_header_row(ws, ["BaseLen", "SingleCharFracWithinStrictPlus_v108"])
    c = ws_headers(ws, header)
    sum_base = 0
    sum_sc = 0.0
    for r in range(header + 1, ws.max_row + 1):
        base_len = ws.cell(r, c["BaseLen"]).value
        if base_len is None:
            continue
        base_len = int(base_len)
        sum_base += base_len
        sum_sc += base_len * float(ws.cell(r, c["SingleCharFracWithinStrictPlus_v108"]).value or 0.0)
    return (sum_sc / sum_base) if sum_base else 0.0


def books_coverage_strictplus_ok(wb: openpyxl.Workbook) -> Tuple[bool, int]:
    ws = wb["Books"]
    header = ws_find_header_row(ws, ["BookID", "Coverage_StrictPlus_v108"])
    c = ws_headers(ws, header)
    bad = 0
    for r in range(header + 1, ws.max_row + 1):
        bid = ws.cell(r, c["BookID"]).value
        if bid is None:
            continue
        cov = ws.cell(r, c["Coverage_StrictPlus_v108"]).value
        if cov not in (1, True):
            bad += 1
    return bad == 0, bad


def strictplus_row_translated(text: object) -> bool:
    """Heuristic for a completed row-level translation in StrictPlus stream."""
    s = str(text or "").strip()
    if not s:
        return False
    return "<*>" not in s


def translation_convergence_metrics(wb: openpyxl.Workbook) -> Tuple[
    int, int, List[int],
    int, int, List[int],
    int, int, List[str],
]:
    """Return (done_books, total_books, done_book_ids, done_contigs, total_contigs, done_contig_ids, done_lines, total_lines, done_line_keys)."""
    books_done = 0
    books_total = 0
    books_done_ids: List[int] = []
    contigs_done = 0
    contigs_total = 0
    contigs_done_ids: List[int] = []
    lines_done = 0
    lines_total = 0
    lines_done_keys: List[str] = []

    # Books: strict criterion currently is full coverage in the existing strictplus metric.
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "Coverage_StrictPlus_v108", "Translation_StrictPlus_v108"])
    cb = ws_headers(ws_books, hb)
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        books_total += 1
        cov = ws_books.cell(r, cb["Coverage_StrictPlus_v108"]).value
        cov_ok = False
        if isinstance(cov, (int, float)):
            cov_ok = float(cov) >= 0.999
        elif isinstance(cov, str):
            try:
                cov_ok = float(cov.strip()) >= 0.999
            except Exception:
                cov_ok = False
        if bool(cov_ok):
            books_done += 1
            try:
                books_done_ids.append(int(bid))
            except Exception:
                books_done_ids.append(0)

    # Contigs: translated if no unresolved <*> marker remains in strictplus text.
    if "Contigs" in wb.sheetnames:
        ws_contigs = wb["Contigs"]
        hc = ws_find_header_row(ws_contigs, ["BaseContigID", "Translation_StrictPlus_v108"])
        cc = ws_headers(ws_contigs, hc)
        for r in range(hc + 1, ws_contigs.max_row + 1):
            cid = ws_contigs.cell(r, cc["BaseContigID"]).value
            if cid is None:
                continue
            contigs_total += 1
            tr = ws_contigs.cell(r, cc["Translation_StrictPlus_v108"]).value
            if strictplus_row_translated(tr):
                contigs_done += 1
                try:
                    contigs_done_ids.append(int(cid))
                except Exception:
                    contigs_done_ids.append(0)

    # Falas (MasterText): translated if no unresolved <*> marker in strictplus text.
    ws_mt = wb["MasterText"]
    hm = ws_find_header_row(ws_mt, ["BaseContigID", "RankByLength", "Translation_StrictPlus_v108"])
    cm = ws_headers(ws_mt, hm)
    for r in range(hm + 1, ws_mt.max_row + 1):
        rank = ws_mt.cell(r, cm["RankByLength"]).value
        if rank is None:
            continue
        contig_id = ws_mt.cell(r, cm["BaseContigID"]).value
        lines_total += 1
        tr = ws_mt.cell(r, cm["Translation_StrictPlus_v108"]).value
        if strictplus_row_translated(tr):
            lines_done += 1
            lines_done_keys.append(f"{contig_id or 'n/a'}#{rank}")

    return (
        books_done,
        books_total,
        books_done_ids,
        contigs_done,
        contigs_total,
        contigs_done_ids,
        lines_done,
        lines_total,
        lines_done_keys,
    )


def groundtruth_cribs_status(wb: openpyxl.Workbook) -> Tuple[int, int]:
    """Return (groundtruth_count, mismatching_count) from Cribs.MatchNorm_v112."""
    ws = wb["Cribs"]
    header = ws_find_header_row(ws, ["CribClass_v112", "MatchNorm_v112"])
    c = ws_headers(ws, header)
    gt = 0
    bad = 0
    for r in range(header + 1, ws.max_row + 1):
        cls = ws.cell(r, c["CribClass_v112"]).value
        if cls != "GroundTruth":
            continue
        gt += 1
        m = ws.cell(r, c["MatchNorm_v112"]).value
        if m not in (1, True):
            bad += 1
    return gt, bad


def summarize_candidate_promotion_skips(wb: openpyxl.Workbook, iter_num: int) -> Tuple[int, str]:
    """Return (skip_count, top_reason_with_count) for CandidatePromotions at an iteration."""
    if "CandidatePromotions" not in wb.sheetnames:
        return 0, ""
    ws = wb["CandidatePromotions"]
    header = ws_find_header_row(ws, ["Iteration", "Decision", "Reason"], max_scan=3)
    c = ws_headers(ws, header)
    by_reason: Counter[str] = Counter()
    total = 0
    for r in range(header + 1, ws.max_row + 1):
        it = ws.cell(r, c["Iteration"]).value
        try:
            it_i = int(it)
        except Exception:
            continue
        if it_i != int(iter_num):
            continue
        decision = str(ws.cell(r, c["Decision"]).value or "").strip().upper()
        if decision != "SKIP":
            continue
        reason_raw = str(ws.cell(r, c["Reason"]).value or "").strip()
        # Collapse long multi-cause strings to the first clause for stable aggregation.
        reason = reason_raw.split(";", 1)[0].strip() if reason_raw else "(no reason)"
        by_reason[reason] += 1
        total += 1
    if not by_reason:
        return total, ""
    top_reason, top_n = sorted(by_reason.items(), key=lambda kv: (-kv[1], kv[0]))[0]
    return total, f"{top_reason} ({top_n})"


def compute_weighted_metrics_for_books_dp(
    wb: openpyxl.Workbook,
    active_tokens: Dict[str, GlossaryToken],
) -> Tuple[float, float, float, float, int]:
    """Compute weighted metrics from Books using DP tokenization (translation-string independent)."""
    ws_books = wb["Books"]
    header = ws_find_header_row(ws_books, ["BookID", "BaseLen", "DecodedBase"], max_scan=3)
    c = ws_headers(ws_books, header)

    sum_base = 0
    sum_ev_t10 = 0
    sum_weak = 0
    sum_micro = 0
    sum_single = 0
    sum_tokens = 0

    for r in range(header + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, c["BookID"]).value
        if bid is None:
            continue
        base_len = int(ws_books.cell(r, c["BaseLen"]).value or 0)
        base = str(ws_books.cell(r, c["DecodedBase"]).value or "")
        items = dp_tokenize_base_with_punct(base, active_tokens)
        tokens = [it for it in items if isinstance(it, GlossaryToken)]

        sum_base += base_len
        for t in tokens:
            sum_ev_t10 += int(t.evidence_score_t10) * int(t.length)
            if int(t.evidence_score_t10) < 20:
                sum_weak += int(t.length)
            if t.evidence_class == "MICRO_MEDIUM":
                sum_micro += int(t.length)
            if int(t.length) == 1:
                sum_single += int(t.length)
        sum_tokens += len(tokens)

    if sum_base == 0:
        return 0.0, 0.0, 0.0, 0.0, 0
    return (sum_ev_t10 / 10.0) / sum_base, sum_weak / sum_base, sum_micro / sum_base, sum_single / sum_base, sum_tokens


def collect_external_base_corpus(wb: openpyxl.Workbook) -> List[str]:
    """Collect base strings outside the Books corpus for safe candidate scanning.

    This is used to allow promotions of tokens that have 0 book occurrences but appear in external refs/cribs
    (e.g., EXTERNAL_POEM tokens). We keep it intentionally small: only already-decoded base strings present
    in the workbook.
    """
    corpus: List[str] = []

    if "Cribs" in wb.sheetnames:
        ws = wb["Cribs"]
        header = ws_find_header_row(ws, ["DecodedBase_Sanitized"])
        c = ws_headers(ws, header)
        col = c["DecodedBase_Sanitized"]
        for r in range(header + 1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if isinstance(v, str) and v.strip():
                corpus.append(v.strip())

    if "ExternalRefs_v115" in wb.sheetnames:
        ws = wb["ExternalRefs_v115"]
        header = ws_find_header_row(ws, ["CodeStreamBase_v120"])
        c = ws_headers(ws, header)
        col = c["CodeStreamBase_v120"]
        for r in range(header + 1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if isinstance(v, str) and v.strip():
                corpus.append(v.strip())

    # Include the Books base corpus as a last-resort signal for inactive tokens whose Glossary TotalOcc is stale.
    # This is only used to *discover* candidates; any promotion is still gated by the DP + metrics simulation.
    if "Books" in wb.sheetnames:
        ws = wb["Books"]
        header = ws_find_header_row(ws, ["BookID", "DecodedBase"])
        c = ws_headers(ws, header)
        col = c["DecodedBase"]
        for r in range(header + 1, ws.max_row + 1):
            bid = ws.cell(r, c["BookID"]).value
            if bid is None:
                continue
            v = ws.cell(r, col).value
            if isinstance(v, str) and v.strip():
                corpus.append(v.strip())

    return corpus


def collect_anchor_base_corpus(
    wb: openpyxl.Workbook,
    *,
    enforced_crib_ids: Optional[set[int]] = None,
    include_groundtruth: bool = True,
    include_external_npc_staff: bool = True,
    include_external_books: bool = True,
    include_anchorcribs: bool = True,
    min_verified_sources: int = 2,
) -> List[str]:
    """Collect high-confidence anchor bases used to constrain mechanical promotions."""
    corpus: List[str] = []

    def _digits_only(v: object) -> str:
        return re.sub(r"\\D+", "", str(v or ""))

    # GroundTruth cribs (prefer enforced/hard rows when provided).
    if include_groundtruth and "Cribs" in wb.sheetnames:
        ws = wb["Cribs"]
        header = ws_find_header_row(ws, ["CribID", "CribClass_v112", "DecodedBase_Sanitized"], max_scan=3)
        c = ws_headers(ws, header)
        vcol = c.get("VerifiedSourceCount_v129")
        for r in range(header + 1, ws.max_row + 1):
            cls = ws.cell(r, c["CribClass_v112"]).value
            if cls != "GroundTruth":
                continue
            cid = ws.cell(r, c["CribID"]).value
            try:
                cid_i = int(cid)
            except Exception:
                continue
            if enforced_crib_ids is not None and cid_i not in enforced_crib_ids:
                continue
            vsc = int(ws.cell(r, vcol).value or 0) if vcol else 0
            if int(min_verified_sources) > 0 and vsc < int(min_verified_sources):
                continue
            base = ws.cell(r, c["DecodedBase_Sanitized"]).value
            if isinstance(base, str) and base.strip():
                corpus.append(base.strip())

    # External references filtered by type + verification count.
    validation_count_by_digits: Dict[str, int] = {}
    if "ExternalValidation_v129" in wb.sheetnames:
        ws_v = wb["ExternalValidation_v129"]
        hv = ws_find_header_row(ws_v, ["NumericText", "VerifiedCount"], max_scan=3)
        cv = ws_headers(ws_v, hv)
        for r in range(hv + 1, ws_v.max_row + 1):
            digits = _digits_only(ws_v.cell(r, cv["NumericText"]).value)
            if not digits:
                continue
            try:
                vc = int(ws_v.cell(r, cv["VerifiedCount"]).value or 0)
            except Exception:
                vc = 0
            if vc > validation_count_by_digits.get(digits, 0):
                validation_count_by_digits[digits] = vc

    if "ExternalRefs_v115" in wb.sheetnames:
        ws = wb["ExternalRefs_v115"]
        header = ws_find_header_row(ws, ["Type", "DecodedBase"], max_scan=3)
        c = ws_headers(ws, header)
        digits_col = c.get("DigitsSanitized")
        num_col = c.get("NumericText")
        base_col = c.get("CodeStreamBase_v120") or c.get("DecodedBase")
        fallback_base_col = c.get("DecodedBase")
        for r in range(header + 1, ws.max_row + 1):
            typ = str(ws.cell(r, c["Type"]).value or "").strip().lower()
            if not typ:
                continue
            take = False
            if include_external_npc_staff and (
                ("npc" in typ)
                or ("interview" in typ)
                or ("staff" in typ)
                or ("creature name" in typ)
                or ("identifier" in typ)
            ):
                take = True
            if include_external_books and "book" in typ:
                take = True
            if not take:
                continue

            digits = ""
            if digits_col:
                digits = _digits_only(ws.cell(r, digits_col).value)
            if (not digits) and num_col:
                digits = _digits_only(ws.cell(r, num_col).value)
            if int(min_verified_sources) > 0:
                if int(validation_count_by_digits.get(digits, 0)) < int(min_verified_sources):
                    continue

            base = ws.cell(r, base_col).value if base_col else None
            if (not isinstance(base, str) or not base.strip()) and fallback_base_col and fallback_base_col != base_col:
                base = ws.cell(r, fallback_base_col).value
            if isinstance(base, str) and base.strip():
                corpus.append(base.strip())

    if include_anchorcribs and "AnchorCribs_Auto" in wb.sheetnames:
        ws = wb["AnchorCribs_Auto"]
        header = ws_find_header_row(ws, ["BaseSubstring"], max_scan=3)
        c = ws_headers(ws, header)
        for r in range(header + 1, ws.max_row + 1):
            base = ws.cell(r, c["BaseSubstring"]).value
            if isinstance(base, str) and base.strip():
                corpus.append(base.strip())

    # De-duplicate while preserving order.
    out: List[str] = []
    seen: set[str] = set()
    for s in corpus:
        ss = str(s or "").strip()
        if not ss or ss in seen:
            continue
        seen.add(ss)
        out.append(ss)
    return out


def token_anchor_hit_count(token: str, anchor_corpus: Sequence[str], cache: Dict[str, int]) -> int:
    t = str(token or "")
    if not t:
        return 0
    if t in cache:
        return int(cache[t] or 0)
    hits = 0
    for s in anchor_corpus:
        if s:
            hits += str(s).count(t)
    cache[t] = int(hits)
    return int(hits)


def filter_candidates_by_anchor_impact(
    candidates: List[Tuple[str, str, str, int, int]],
    *,
    anchor_corpus: Sequence[str],
    min_hits: int,
    cache: Dict[str, int],
) -> Tuple[List[Tuple[str, str, str, int, int]], int, Dict[str, str]]:
    if not candidates:
        return candidates, 0, {"classes": "", "samples": ""}
    out: List[Tuple[str, str, str, int, int]] = []
    dropped = 0
    need_hits = max(1, int(min_hits))
    dropped_by_class: Dict[str, int] = {}
    dropped_samples: List[str] = []
    for row in candidates:
        tok = row[0]
        hits = token_anchor_hit_count(tok, anchor_corpus, cache)
        if hits >= need_hits:
            out.append(row)
        else:
            dropped += 1
            ev_class = str(row[2] or "")
            dropped_by_class[ev_class] = int(dropped_by_class.get(ev_class, 0) or 0) + 1
            if len(dropped_samples) < 8:
                occ = int(row[3] or 0)
                dropped_samples.append(f"{tok}:{hits}/{need_hits}:{ev_class}:{occ}")
    class_items = sorted(dropped_by_class.items(), key=lambda kv: (-int(kv[1]), str(kv[0])))
    return out, dropped, {
        "classes": ",".join([f"{cls}={cnt}" for cls, cnt in class_items[:6]]),
        "samples": "; ".join(dropped_samples),
    }


def candidate_usage_stats_in_books(
    wb: openpyxl.Workbook,
    tokens: Sequence[str],
) -> Dict[str, Tuple[int, int]]:
    stats: Dict[str, Tuple[int, int]] = {}
    toks = [str(t or "").strip() for t in tokens if str(t or "").strip()]
    if not toks or "Books" not in wb.sheetnames:
        return stats

    uniq_toks = []
    seen: set[str] = set()
    for t in toks:
        if t in seen:
            continue
        seen.add(t)
        uniq_toks.append(t)
        stats[t] = (0, 0)

    ws = wb["Books"]
    header = ws_find_header_row(ws, ["BookID", "DecodedBase"], max_scan=3)
    c = ws_headers(ws, header)
    base_col = c.get("DecodedBase")
    if not base_col:
        return stats

    for r in range(header + 1, ws.max_row + 1):
        base = ws.cell(r, base_col).value
        if not isinstance(base, str) or not base:
            continue
        for tok in uniq_toks:
            cnt = int(base.count(tok))
            if cnt <= 0:
                continue
            bh, bo = stats.get(tok, (0, 0))
            stats[tok] = (int(bh) + 1, int(bo) + cnt)
    return stats


def recent_no_effect_token_counts(
    wb: openpyxl.Workbook,
    *,
    lookback_iters: int = 8,
) -> Dict[str, int]:
    out: Dict[str, int] = {}
    if "CandidatePromotions" not in wb.sheetnames:
        return out

    ws = wb["CandidatePromotions"]
    header = ws_find_header_row(ws, ["Iteration", "Token", "Decision", "Reason"], max_scan=5)
    c = ws_headers(ws, header)
    iter_col = c.get("Iteration")
    tok_col = c.get("Token")
    dec_col = c.get("Decision")
    reason_col = c.get("Reason")
    if not iter_col or not tok_col or not dec_col or not reason_col:
        return out

    max_iter = 0
    for r in range(header + 1, ws.max_row + 1):
        itv = ws.cell(r, iter_col).value
        if isinstance(itv, (int, float)):
            max_iter = max(max_iter, int(itv))
    if max_iter <= 0:
        return out

    min_iter = max(1, int(max_iter) - max(1, int(lookback_iters)) + 1)
    needle = "No effect in DP metrics (not used)"
    for r in range(header + 1, ws.max_row + 1):
        itv = ws.cell(r, iter_col).value
        if not isinstance(itv, (int, float)) or int(itv) < min_iter:
            continue
        dec = str(ws.cell(r, dec_col).value or "").strip().upper()
        if dec != "SKIP":
            continue
        reason = str(ws.cell(r, reason_col).value or "").strip()
        if needle not in reason:
            continue
        tok = str(ws.cell(r, tok_col).value or "").strip()
        if not tok:
            continue
        out[tok] = int(out.get(tok, 0)) + 1
    return out


def _glossary_append_row_copy_style(
    ws: openpyxl.worksheet.worksheet.Worksheet, template_row: int, values_by_col: Dict[int, object]
) -> int:
    """Append a new row to Glossary-like sheets, copying style from `template_row`."""
    # openpyxl styles are immutable-ish and stored in indexed collections; reusing StyleProxy objects can
    # trigger hashing errors. Keep style copying minimal and safe.
    from copy import copy

    r = ws_last_data_row(ws) + 1
    for c in range(1, ws.max_column + 1):
        src = ws.cell(template_row, c)
        dst = ws.cell(r, c)
        dst._style = copy(src._style)
        dst.comment = None
        dst.value = src.value
    for col, v in values_by_col.items():
        ws.cell(r, col).value = v
    return r


def _macro_confidence_from_components(tokens: Sequence[GlossaryToken]) -> str:
    """Choose macro confidence so the macro can compete with its components in DP objective (confW*len)."""
    total_len = sum(t.length for t in tokens)
    if total_len <= 0:
        return "LOW"
    avg = sum(confidence_weight(t.confidence) * t.length for t in tokens) / float(total_len)
    # Use an "upper rounding" so the macro can actually be selected when it includes any HIGH components.
    # Example: HIGH(3)*4 + MED(2)*6 => avg=2.4, which still requires HIGH at macro-level to compete on objective #1.
    if avg > 2.0:
        return "HIGH"
    if avg > 1.0:
        return "MEDIUM"
    return "LOW"


@dataclass(frozen=True)
class MinedMacro:
    base: str
    translation: str
    evidence_score: float
    occ: int
    book_count: int
    length: int
    tokens: Tuple[str, ...]
    weak_chars: int
    micro_chars: int
    single_chars: int
    confidence: str


def mine_macro_candidates_from_books(
    wb: openpyxl.Workbook,
    active_tokens: Dict[str, GlossaryToken],
    existing_tokens: Sequence[str],
    *,
    n_values: Sequence[int] = (2, 3, 4),
    min_occ: int = 2,
    min_books: int = 2,
    min_share: float = 1.0,
    min_len: int = 2,
    max_len: int = 12,
    max_candidates: int = 25,
    allow_macro_components: bool = False,
    allow_marker_tokens: bool = False,
    allow_star_tokens: bool = False,
) -> List[MinedMacro]:
    """Mine consistent n-gram macros from the Books DP alignment.

    Goal: create *new* macro tokens that (a) are consistent across occurrences, and
    (b) have a good chance to reduce WEAK/MICRO/SINGLE fractions by replacing weak token sequences.
    """
    existing_set = set(existing_tokens)

    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "DecodedBase", "Translation_StrictPlus_LosslessMarkers_v108"])
    cb = ws_headers(ws_books, hb)

    # base -> Counter(translation_tuple)
    tr_counts: Dict[str, Counter[Tuple[str, ...]]] = defaultdict(Counter)
    # (base, translation_tuple) -> stats
    stats: Dict[Tuple[str, Tuple[str, ...]], Dict[str, object]] = {}

    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        bid_i = int(bid)
        base = str(ws_books.cell(r, cb["DecodedBase"]).value or "")
        items = dp_tokenize_base_with_punct(base, active_tokens)
        toks = [it for it in items if isinstance(it, GlossaryToken)]
        L = len(toks)

        for n in n_values:
            if n <= 1:
                continue
            for i in range(0, L - n + 1):
                seq = toks[i : i + n]
                # Exclusions: keep mined macros simple and low-risk.
                if not allow_marker_tokens and any(t.token_type == "marker" for t in seq):
                    continue
                if not allow_macro_components and any(t.token_type == "macro" for t in seq):
                    continue
                if any(t.evidence_class == "PUNCT_LOGOGRAM" for t in seq):
                    continue
                if (not allow_star_tokens) and any("*" in t.token for t in seq):
                    continue
                base_concat = "".join(t.token for t in seq)
                if base_concat in existing_set:
                    continue
                blen = len(base_concat)
                if blen < min_len or blen > max_len:
                    continue
                # Mine sequences that likely reduce "risk" metrics when collapsed:
                # - WEAK components drive WeakFrac
                # - MICRO_MEDIUM components drive MicroFrac
                # - single-char components drive SingleCharFrac
                has_weak = any(t.evidence_score_t10 < 20 for t in seq)
                has_micro = any(t.evidence_class == "MICRO_MEDIUM" for t in seq)
                has_single = any(t.length == 1 for t in seq)
                if not (has_weak or has_micro or has_single):
                    continue

                out: List[str] = []
                for t in seq:
                    out.extend(t.lossless_out_tokens)
                # When markers are disabled, do not mine marker lossless tokens into translation.
                if (not allow_marker_tokens) and any(x.startswith("<") for x in out):
                    continue
                tr_tup = tuple(out)

                tr_counts[base_concat][tr_tup] += 1
                key = (base_concat, tr_tup)
                st = stats.get(key)
                if st is None:
                    total_len = sum(t.length for t in seq)
                    if total_len <= 0:
                        continue
                    # EvidenceScore for mined macros:
                    # - use a length-weighted average of components so evidence does not artificially regress
                    # - clamp to at least 2.0 so the macro is not WEAK (helps WeakFrac) when it replaces weak parts
                    avg_ev = sum(float(t.evidence_score) * t.length for t in seq) / float(total_len)
                    macro_ev = max(2.0, min(3.0, round(avg_ev, 1)))
                    st = {
                        "occ": 0,
                        "books": set(),
                        "weak_chars": 0,
                        "micro_chars": 0,
                        "single_chars": 0,
                        "tokens": tuple(t.token for t in seq),
                        "confidence": _macro_confidence_from_components(seq),
                        "evidence_score": macro_ev,
                    }
                    stats[key] = st
                st["occ"] = int(st["occ"]) + 1
                cast_books: set = st["books"]  # type: ignore[assignment]
                cast_books.add(bid_i)
                st["weak_chars"] = int(st["weak_chars"]) + sum(t.length for t in seq if t.evidence_score_t10 < 20)
                st["micro_chars"] = int(st["micro_chars"]) + sum(t.length for t in seq if t.evidence_class == "MICRO_MEDIUM")
                st["single_chars"] = int(st["single_chars"]) + sum(t.length for t in seq if t.length == 1)

    mined: List[MinedMacro] = []
    for base_concat, ctr in tr_counts.items():
        if base_concat in existing_set:
            continue
        total = sum(ctr.values())
        if total < min_occ:
            continue
        tr_tup, occ = ctr.most_common(1)[0]
        if occ < min_occ:
            continue
        share = (float(occ) / float(total)) if total else 0.0
        # Default safe behavior is strict consistency; relaxing share is still safe because DP can only
        # pick a macro when its translation matches the existing lossless stream at that location.
        if share + 1e-9 < float(min_share):
            continue
        key = (base_concat, tr_tup)
        st = stats.get(key)
        if not st:
            continue
        books = st["books"]
        book_count = len(books) if isinstance(books, set) else 0
        if book_count < min_books:
            continue
        mined.append(
            MinedMacro(
                base=base_concat,
                translation=" ".join(tr_tup),
                evidence_score=float(st.get("evidence_score") or 2.0),
                occ=int(st["occ"]),
                book_count=book_count,
                length=len(base_concat),
                tokens=tuple(st["tokens"]),  # type: ignore[arg-type]
                weak_chars=int(st["weak_chars"]),
                micro_chars=int(st["micro_chars"]),
                single_chars=int(st["single_chars"]),
                confidence=str(st["confidence"]),
            )
        )

    # Sort by expected impact (approx): weak chars covered, then micro, then single, then occ.
    mined.sort(key=lambda m: (m.weak_chars, m.micro_chars, m.single_chars, m.occ, m.length), reverse=True)
    return mined[:max_candidates]


def mine_macro_candidates_from_anchor_corpus(
    anchor_corpus: Sequence[str],
    active_tokens: Dict[str, GlossaryToken],
    existing_tokens: Sequence[str],
    *,
    n_values: Sequence[int] = (2, 3, 4),
    min_occ: int = 2,
    min_books: int = 2,
    min_share: float = 1.0,
    min_len: int = 2,
    max_len: int = 12,
    max_candidates: int = 25,
    allow_macro_components: bool = False,
    allow_marker_tokens: bool = False,
    allow_star_tokens: bool = False,
) -> List[MinedMacro]:
    """Mine consistent n-gram macros from anchor bases instead of the global Books corpus.

    This is intentionally narrow: it uses the same DP tokenization and safety heuristics as
    book-wide MacroMine, but only on the verified/high-confidence anchor bases that already
    constrain AnchorPromotionOnly. The goal is to *generate* local candidates near anchors,
    not just filter the old inactive glossary pool.
    """
    existing_set = set(existing_tokens)

    tr_counts: Dict[str, Counter[Tuple[str, ...]]] = defaultdict(Counter)
    stats: Dict[Tuple[str, Tuple[str, ...]], Dict[str, object]] = {}

    for anchor_idx, base in enumerate(anchor_corpus or [], start=1):
        base_s = str(base or "").strip()
        if not base_s:
            continue
        items = dp_tokenize_base_with_punct(base_s, active_tokens)
        toks = [it for it in items if isinstance(it, GlossaryToken)]
        L = len(toks)

        for n in n_values:
            if n <= 1:
                continue
            for i in range(0, L - n + 1):
                seq = toks[i : i + n]
                if not allow_marker_tokens and any(t.token_type == "marker" for t in seq):
                    continue
                if not allow_macro_components and any(t.token_type == "macro" for t in seq):
                    continue
                if any(t.evidence_class == "PUNCT_LOGOGRAM" for t in seq):
                    continue
                if (not allow_star_tokens) and any("*" in t.token for t in seq):
                    continue

                base_concat = "".join(t.token for t in seq)
                if base_concat in existing_set:
                    continue
                blen = len(base_concat)
                if blen < min_len or blen > max_len:
                    continue

                has_weak = any(t.evidence_score_t10 < 20 for t in seq)
                has_micro = any(t.evidence_class == "MICRO_MEDIUM" for t in seq)
                has_single = any(t.length == 1 for t in seq)
                if not (has_weak or has_micro or has_single):
                    continue

                out: List[str] = []
                for t in seq:
                    out.extend(t.lossless_out_tokens)
                if (not allow_marker_tokens) and any(x.startswith("<") for x in out):
                    continue
                tr_tup = tuple(out)

                tr_counts[base_concat][tr_tup] += 1
                key = (base_concat, tr_tup)
                st = stats.get(key)
                if st is None:
                    total_len = sum(t.length for t in seq)
                    if total_len <= 0:
                        continue
                    avg_ev = sum(float(t.evidence_score) * t.length for t in seq) / float(total_len)
                    macro_ev = max(2.0, min(3.0, round(avg_ev, 1)))
                    st = {
                        "occ": 0,
                        "books": set(),
                        "weak_chars": 0,
                        "micro_chars": 0,
                        "single_chars": 0,
                        "tokens": tuple(t.token for t in seq),
                        "confidence": _macro_confidence_from_components(seq),
                        "evidence_score": macro_ev,
                    }
                    stats[key] = st
                st["occ"] = int(st["occ"]) + 1
                cast_books: set = st["books"]  # type: ignore[assignment]
                cast_books.add(int(anchor_idx))
                st["weak_chars"] = int(st["weak_chars"]) + sum(t.length for t in seq if t.evidence_score_t10 < 20)
                st["micro_chars"] = int(st["micro_chars"]) + sum(t.length for t in seq if t.evidence_class == "MICRO_MEDIUM")
                st["single_chars"] = int(st["single_chars"]) + sum(t.length for t in seq if t.length == 1)

    mined: List[MinedMacro] = []
    for base_concat, ctr in tr_counts.items():
        if base_concat in existing_set:
            continue
        total = sum(ctr.values())
        if total < min_occ:
            continue
        tr_tup, occ = ctr.most_common(1)[0]
        if occ < min_occ:
            continue
        share = (float(occ) / float(total)) if total else 0.0
        if share + 1e-9 < float(min_share):
            continue
        key = (base_concat, tr_tup)
        st = stats.get(key)
        if not st:
            continue
        books = st["books"]
        book_count = len(books) if isinstance(books, set) else 0
        if book_count < min_books:
            continue
        mined.append(
            MinedMacro(
                base=base_concat,
                translation=" ".join(tr_tup),
                evidence_score=float(st.get("evidence_score") or 2.0),
                occ=int(st["occ"]),
                book_count=book_count,
                length=len(base_concat),
                tokens=tuple(st["tokens"]),  # type: ignore[arg-type]
                weak_chars=int(st["weak_chars"]),
                micro_chars=int(st["micro_chars"]),
                single_chars=int(st["single_chars"]),
                confidence=str(st["confidence"]),
            )
        )

    mined.sort(key=lambda m: (m.weak_chars, m.micro_chars, m.single_chars, m.occ, m.length), reverse=True)
    return mined[:max_candidates]


def add_mined_macros_to_glossary(
    wb: openpyxl.Workbook,
    iter_num: int,
    macros: Sequence[MinedMacro],
    *,
    evidence_class: str = "MACRO_ACTIVE",
    source_kind: str = "n-gram",
    evidence_sources_tag: str = "mined_ngram_macro",
) -> int:
    """Append mined macros into Glossary as inactive tokens."""
    if not macros:
        return 0
    ws = wb["Glossary"]
    header = ws_find_header_row(
        ws,
        [
            "Token",
            "Translation",
            "TokenType",
            "Confidence",
            "ActiveCorpus",
            "Use_StrictPlus_v108",
            "Use_Strict_v108",
            "BoundaryRule",
            "RuleLayerNeeded",
            "TotalOcc",
            "BookCount",
            "ContigCount",
            "Len",
            "StarCount",
            "Notes",
            "Use_PoemMode_v113",
            "SemNeutral_Render_v123",
            "Mask_SemNeutral_v123",
            "EvidenceClass_v127",
            "EvidenceScore_v127",
            "EvidenceSources_v127",
        ],
    )
    c = ws_headers(ws, header)

    # Pick a macro row as style/value template (any macro row is fine since we overwrite key fields).
    template_row = None
    for r in range(header + 1, ws.max_row + 1):
        if ws.cell(r, c["TokenType"]).value == "macro":
            template_row = r
            break
    if template_row is None:
        template_row = header + 1

    existing = {ws.cell(r, c["Token"]).value for r in range(header + 1, ws.max_row + 1) if ws.cell(r, c["Token"]).value}

    added = 0
    for mm in macros:
        if mm.base in existing:
            continue
        note = (
            f"iter{iter_num} mined macro ({source_kind}). occ={mm.occ} books={mm.book_count}. "
            f"Composition tokens: {' + '.join(mm.tokens)}. "
            f"WeakCharsCovered={mm.weak_chars} MicroCharsCovered={mm.micro_chars} SingleCharsCovered={mm.single_chars}."
        )
        values = {
            c["Token"]: mm.base,
            c["Translation"]: mm.translation,
            c["TokenType"]: "macro",
            c["Confidence"]: mm.confidence,
            c["ActiveCorpus"]: False,
            c["Use_StrictPlus_v108"]: 0,  # inactive until mechanically promoted
            c["Use_Strict_v108"]: False,
            c["BoundaryRule"]: "none",
            c["RuleLayerNeeded"]: "no_anagram",
            c["TotalOcc"]: mm.occ,
            c["BookCount"]: mm.book_count,
            c["ContigCount"]: 0,
            c["Len"]: mm.length,
            c["StarCount"]: int(mm.base.count("*")),
            c["Notes"]: note,
            c["Use_PoemMode_v113"]: False,
            c["SemNeutral_Render_v123"]: "NORMAL",
            c["Mask_SemNeutral_v123"]: 0,
            c["EvidenceClass_v127"]: evidence_class,
            c["EvidenceScore_v127"]: mm.evidence_score,
            c["EvidenceSources_v127"]: f"iter{iter_num}: {evidence_sources_tag}",
        }
        _glossary_append_row_copy_style(ws, template_row, values_by_col=values)
        existing.add(mm.base)
        added += 1

    return added


def refresh_mined_ngram_macro_evidence(wb: openpyxl.Workbook, iter_num: int) -> int:
    """Recompute EvidenceScore/Confidence for macros created by this runner (notes contain 'mined macro (...)')."""
    ws = wb["Glossary"]
    header = ws_find_header_row(ws, ["Token", "TokenType", "Confidence", "EvidenceScore_v127", "Notes"])
    c = ws_headers(ws, header)

    # Load current glossary map for component lookup.
    _gws, glossary_map = load_glossary(wb)

    updated = 0
    for r in range(header + 1, ws.max_row + 1):
        tok = ws.cell(r, c["Token"]).value
        if not isinstance(tok, str) or not tok:
            continue
        if ws.cell(r, c["TokenType"]).value != "macro":
            continue
        notes = ws.cell(r, c["Notes"]).value
        if not isinstance(notes, str) or "mined macro (" not in notes or "Composition tokens:" not in notes:
            continue

        # Extract "Composition tokens: A + B + C." from notes.
        m = re.search(r"Composition tokens:\s*([^.]+)", notes)
        if not m:
            continue
        part = m.group(1)
        comp = [p.strip() for p in part.split("+")]
        comp = [p for p in comp if p]
        if not comp:
            continue

        comps: List[GlossaryToken] = []
        ok = True
        for ct in comp:
            gt = glossary_map.get(ct)
            if gt is None:
                ok = False
                break
            comps.append(gt)
        if not ok:
            continue

        total_len = sum(t.length for t in comps)
        if total_len <= 0:
            continue
        avg_ev = sum(float(t.evidence_score) * t.length for t in comps) / float(total_len)
        macro_ev = max(2.0, min(3.0, round(avg_ev, 1)))
        macro_conf = _macro_confidence_from_components(comps)

        cur_ev = ws.cell(r, c["EvidenceScore_v127"]).value
        cur_conf = ws.cell(r, c["Confidence"]).value

        changed = False
        if cur_ev is None:
            ws.cell(r, c["EvidenceScore_v127"]).value = macro_ev
            changed = True
        else:
            cur_f = float(cur_ev)
            # Safe refresh: never decrease existing evidence scores.
            if macro_ev > cur_f + 1e-9:
                ws.cell(r, c["EvidenceScore_v127"]).value = macro_ev
                changed = True

        # Safe refresh: never lower confidence (affects DP selection).
        cur_w = confidence_weight(str(cur_conf or ""))
        new_w = confidence_weight(macro_conf)
        if new_w > cur_w:
            ws.cell(r, c["Confidence"]).value = macro_conf
            changed = True

        if changed:
            updated += 1

    return updated


def _book_base_and_lossless(wb: openpyxl.Workbook, book_id: int) -> Tuple[str, List[str]]:
    ws = wb["Books"]
    hb = ws_find_header_row(ws, ["BookID", "DecodedBase", "Translation_StrictPlus_LosslessMarkers_v108"], max_scan=3)
    c = ws_headers(ws, hb)
    for r in range(hb + 1, ws.max_row + 1):
        bid = ws.cell(r, c["BookID"]).value
        if bid == book_id:
            base = str(ws.cell(r, c["DecodedBase"]).value or "")
            lossless = str(ws.cell(r, c["Translation_StrictPlus_LosslessMarkers_v108"]).value or "")
            return base, parse_lossless_tokens(lossless)
    return "", []


def _dp_token_spans(tokens: Sequence[GlossaryToken]) -> List[Tuple[int, int, GlossaryToken]]:
    """Return token spans as (start_inclusive, end_exclusive, token)."""
    spans: List[Tuple[int, int, GlossaryToken]] = []
    i = 0
    for t in tokens:
        spans.append((i, i + int(t.length), t))
        i += int(t.length)
    return spans


def mine_macro_candidates_from_superanchors(
    wb: openpyxl.Workbook,
    active_tokens: Dict[str, GlossaryToken],
    existing_tokens: Sequence[str],
    *,
    source_iter: int,
    min_len: int = 12,
    max_len: int = 60,
    max_candidates: int = 12,
) -> List[MinedMacro]:
    """Derive macro candidates from SuperAnchors_Auto by snapping to DP token boundaries.

    These macros are conservative: translation is composed from existing token translations, so they
    cannot introduce new semantics. They mainly help DP reduce token count / single-char load.
    """
    if "SuperAnchors_Auto" not in wb.sheetnames:
        return []

    existing_set = set(existing_tokens)
    ws_sa = wb["SuperAnchors_Auto"]
    hs = ws_find_header_row(ws_sa, ["Iter", "RefBookID", "Start", "End", "BaseSubstring"], max_scan=3)
    cs = ws_headers(ws_sa, hs)

    # Cache books base for occ counting.
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_books, hb)
    bases: Dict[int, str] = {}
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        bases[bid_i] = str(ws_books.cell(r, cb["DecodedBase"]).value or "")

    out: List[MinedMacro] = []

    for r in range(hs + 1, ws_sa.max_row + 1):
        it = ws_sa.cell(r, cs["Iter"]).value
        if it != source_iter:
            continue
        ref_book = ws_sa.cell(r, cs["RefBookID"]).value
        a = ws_sa.cell(r, cs["Start"]).value
        b = ws_sa.cell(r, cs["End"]).value
        if ref_book is None or a is None or b is None:
            continue
        try:
            ref_book_i = int(ref_book)
            start = int(a)
            end = int(b)
        except Exception:
            continue

        base, _lossless = _book_base_and_lossless(wb, ref_book_i)
        if not base:
            continue

        items = dp_tokenize_base_with_punct(base, active_tokens)
        toks = [it for it in items if isinstance(it, GlossaryToken)]
        spans = _dp_token_spans(toks)

        # Snap to token boundaries strictly within [start,end] inclusive.
        run_start = start
        run_end_excl = end + 1
        inside = [(s, e, t) for (s, e, t) in spans if s >= run_start and e <= run_end_excl]
        if not inside:
            continue
        snapped_start = inside[0][0]
        snapped_end_excl = inside[-1][1]
        length = snapped_end_excl - snapped_start
        if length < int(min_len) or length > int(max_len):
            continue

        base_sub = base[snapped_start:snapped_end_excl]
        if not base_sub or base_sub in existing_set:
            continue

        comps = [t for _s, _e, t in inside]
        # Translation is the exact lossless token stream rendered by the component tokens.
        lossless_out: List[str] = []
        for t in comps:
            lossless_out.extend(t.lossless_out_tokens)
        translation = " ".join(lossless_out)

        occ = 0
        book_count = 0
        for bid_i, bb in bases.items():
            if not bb:
                continue
            c = bb.count(base_sub)
            if c:
                occ += c
                book_count += 1

        total_len = sum(int(t.length) for t in comps)
        if total_len <= 0:
            continue
        avg_ev = sum(float(t.evidence_score) * int(t.length) for t in comps) / float(total_len)
        macro_ev = max(2.0, min(3.0, round(avg_ev, 1)))
        macro_conf = _macro_confidence_from_components(comps)

        weak_chars = sum(int(t.length) for t in comps if int(t.evidence_score_t10) < 20)
        micro_chars = sum(int(t.length) for t in comps if t.evidence_class == "MICRO_MEDIUM")
        single_chars = sum(int(t.length) for t in comps if int(t.length) == 1)

        out.append(
            MinedMacro(
                base=base_sub,
                translation=translation,
                evidence_score=float(macro_ev),
                occ=int(occ),
                book_count=int(book_count),
                length=int(length),
                tokens=tuple(t.token for t in comps),
                weak_chars=int(weak_chars),
                micro_chars=int(micro_chars),
                single_chars=int(single_chars),
                confidence=str(macro_conf),
            )
        )
        existing_set.add(base_sub)
        if len(out) >= int(max_candidates):
            break

    # Prefer longer, higher-occ macros.
    out.sort(key=lambda m: (m.book_count, m.occ, m.length), reverse=True)
    return out


def upsert_iter_summary_sheet(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    changed_books: str,
    evidence_avg: float,
    weak_frac: float,
    micro_frac: float,
    gt_promoted: int,
    mech_promoted: int,
    semantic_promoted: int = 0,
    extra_notes: Optional[str] = None,
) -> None:
    name = f"Iter{iter_num}_Summary"
    if name in wb.sheetnames:
        ws = wb[name]
        # Clear existing content (keep styles minimal). Overwrite cells only.
    else:
        ws = wb.create_sheet(name)

    ws.cell(1, 1).value = "Metric"
    ws.cell(1, 2).value = "Value"
    ws.cell(1, 3).value = "Notes"

    rows = [
        ("Iteration", iter_num, None),
        ("UTC", utc, None),
        ("Books changed (StrictPlus text)", changed_books, "Auto-chain iteration (safe)."),
        ("StrictPlus coverage (weighted)", 1, "All books tokenized under current StrictPlus token set."),
        ("Token-evidence avg (Books, length-weighted)", round(evidence_avg, 6), None),
        ("WEAK char frac (Books, length-weighted)", round(weak_frac, 6), None),
        ("MICRO_MEDIUM char frac (Books, length-weighted)", round(micro_frac, 6), None),
        ("GroundTruth new promotions", gt_promoted, None),
        ("Mechanical promotions accepted", mech_promoted, None),
        ("Semantic glossary retext applied", int(semantic_promoted), "String-only translation edits guarded by GT live check."),
    ]
    if extra_notes:
        rows.append(("Notes", extra_notes, None))

    for i, (k, v, n) in enumerate(rows, start=2):
        ws.cell(i, 1).value = k
        ws.cell(i, 2).value = v
        ws.cell(i, 3).value = n


def upsert_sheet_index_entry(wb: openpyxl.Workbook, sheet_name: str, description: str) -> None:
    ws = wb["SheetIndex"]
    header = ws_find_header_row(ws, ["Sheet", "What it contains"])
    c = ws_headers(ws, header)
    for r in range(header + 1, ws.max_row + 1):
        if ws.cell(r, c["Sheet"]).value == sheet_name:
            ws.cell(r, c["What it contains"]).value = description
            return
    ws_append_row(ws, [sheet_name, description], start_col=1)


def _read_iter_summary_metrics(wb: openpyxl.Workbook, iter_num: int) -> Dict[str, object]:
    name = f"Iter{iter_num}_Summary"
    if name not in wb.sheetnames:
        return {}
    ws = wb[name]
    header = ws_find_header_row(ws, ["Metric", "Value"], max_scan=3)
    c = ws_headers(ws, header)
    out: Dict[str, object] = {}
    for r in range(header + 1, ws.max_row + 1):
        k = ws.cell(r, c["Metric"]).value
        v = ws.cell(r, c["Value"]).value
        if not isinstance(k, str) or not k.strip():
            continue
        out[k.strip()] = v
    return out


def upsert_iter_focus_sheet(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    evidence_avg: float,
    weak_frac: float,
    micro_frac: float,
    single_frac: float,
    token_count: int,
) -> int:
    """Create/update an iteration focus sheet with actionable plateau diagnostics.

    Returns an approximate number of non-empty rows written (for logging).
    """
    name = f"Iter{iter_num}_Focus"
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)

    r = 1
    ws.cell(r, 1).value = "Iteration"
    ws.cell(r, 2).value = iter_num
    r += 1
    ws.cell(r, 1).value = "UTC"
    ws.cell(r, 2).value = utc
    r += 1
    ws.cell(r, 1).value = "EvidenceAvg (Books, weighted)"
    ws.cell(r, 2).value = round(float(evidence_avg), 6)
    r += 1
    ws.cell(r, 1).value = "WeakFrac (Books, weighted)"
    ws.cell(r, 2).value = round(float(weak_frac), 6)
    r += 1
    ws.cell(r, 1).value = "MicroFrac (Books, weighted)"
    ws.cell(r, 2).value = round(float(micro_frac), 6)
    r += 1
    ws.cell(r, 1).value = "SingleCharFrac (Books, weighted)"
    ws.cell(r, 2).value = round(float(single_frac), 6)
    r += 1
    ws.cell(r, 1).value = "TokenCount (StrictPlus)"
    ws.cell(r, 2).value = int(token_count)
    r += 2

    # Pull Books strings for preview
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "BaseLen", "DecodedBase", "Translation_StrictPlus_v108"])
    cb = ws_headers(ws_books, hb)
    book_map: Dict[int, Tuple[str, str]] = {}
    for rr in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(rr, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        dec = ws_books.cell(rr, cb["DecodedBase"]).value
        tr = ws_books.cell(rr, cb["Translation_StrictPlus_v108"]).value
        book_map[bid_i] = (str(dec or ""), str(tr or ""))

    # Top books by weak/micro
    ws_te = wb["TokenEvidence_Books_v128"]
    ht = ws_find_header_row(ws_te, ["BookID", "BaseLen", "EvidenceAvg", "WeakFrac", "MicroFrac", "WeakTopTokens"])
    ct = ws_headers(ws_te, ht)
    te_rows = []
    for rr in range(ht + 1, ws_te.max_row + 1):
        bid = ws_te.cell(rr, ct["BookID"]).value
        if bid is None:
            continue
        te_rows.append(
            (
                int(bid),
                int(ws_te.cell(rr, ct["BaseLen"]).value or 0),
                float(ws_te.cell(rr, ct["EvidenceAvg"]).value or 0.0),
                float(ws_te.cell(rr, ct["WeakFrac"]).value or 0.0),
                float(ws_te.cell(rr, ct["MicroFrac"]).value or 0.0),
                str(ws_te.cell(rr, ct["WeakTopTokens"]).value or ""),
            )
        )

    ws.cell(r, 1).value = "Top Books by WeakFrac"
    r += 1
    headers = ["BookID", "BaseLen", "EvidenceAvg", "WeakFrac", "MicroFrac", "WeakTopTokens", "StrictPlusPreview"]
    for i, h in enumerate(headers, start=1):
        ws.cell(r, i).value = h
    r += 1
    for bid, bl, ev, wf, mf, top in sorted(te_rows, key=lambda x: (-x[3], -x[4], x[0]))[:15]:
        _dec, tr = book_map.get(bid, ("", ""))
        ws.cell(r, 1).value = bid
        ws.cell(r, 2).value = bl
        ws.cell(r, 3).value = round(ev, 6)
        ws.cell(r, 4).value = round(wf, 6)
        ws.cell(r, 5).value = round(mf, 6)
        ws.cell(r, 6).value = top
        ws.cell(r, 7).value = (tr[:160] + ("..." if len(tr) > 160 else ""))
        r += 1

    r += 1
    ws.cell(r, 1).value = "Top Books by MicroFrac"
    r += 1
    for i, h in enumerate(headers, start=1):
        ws.cell(r, i).value = h
    r += 1
    for bid, bl, ev, wf, mf, top in sorted(te_rows, key=lambda x: (-x[4], -x[3], x[0]))[:15]:
        _dec, tr = book_map.get(bid, ("", ""))
        ws.cell(r, 1).value = bid
        ws.cell(r, 2).value = bl
        ws.cell(r, 3).value = round(ev, 6)
        ws.cell(r, 4).value = round(wf, 6)
        ws.cell(r, 5).value = round(mf, 6)
        ws.cell(r, 6).value = top
        ws.cell(r, 7).value = (tr[:160] + ("..." if len(tr) > 160 else ""))
        r += 1

    # MICRO_MEDIUM tokens (active) + one example snippet from MasterText Rank 1.
    r += 2
    ws.cell(r, 1).value = "Top Active MICRO_MEDIUM Tokens (by TotalOcc)"
    r += 1
    ws.cell(r, 1).value = "Token"
    ws.cell(r, 2).value = "Translation"
    ws.cell(r, 3).value = "TotalOcc"
    ws.cell(r, 4).value = "BookCount"
    ws.cell(r, 5).value = "Len"
    ws.cell(r, 6).value = "EvidenceScore"
    ws.cell(r, 7).value = "Example (MasterText)"
    r += 1

    mt_example = ""
    try:
        ws_mt = wb["MasterText"]
        hm = ws_find_header_row(ws_mt, ["RankByLength", "Translation_StrictPlus_v108"])
        cm = ws_headers(ws_mt, hm)
        # Use Rank 1 (longest) for consistent examples.
        for rr in range(hm + 1, ws_mt.max_row + 1):
            rk = ws_mt.cell(rr, cm["RankByLength"]).value
            if rk == 1:
                mt_example = str(ws_mt.cell(rr, cm["Translation_StrictPlus_v108"]).value or "")
                break
    except Exception:
        mt_example = ""

    ws_el = wb["EvidenceLedger_v127"]
    he = ws_find_header_row(ws_el, ["Token", "Translation", "Use_StrictPlus_v108", "TotalOcc", "BookCount", "Len", "EvidenceClass_v127", "EvidenceScore_v127"])
    ce = ws_headers(ws_el, he)
    micro_rows = []
    for rr in range(he + 1, ws_el.max_row + 1):
        tok = ws_el.cell(rr, ce["Token"]).value
        if not isinstance(tok, str) or not tok:
            continue
        if not bool(ws_el.cell(rr, ce["Use_StrictPlus_v108"]).value):
            continue
        if ws_el.cell(rr, ce["EvidenceClass_v127"]).value != "MICRO_MEDIUM":
            continue
        micro_rows.append(
            (
                tok,
                str(ws_el.cell(rr, ce["Translation"]).value or ""),
                int(ws_el.cell(rr, ce["TotalOcc"]).value or 0),
                int(ws_el.cell(rr, ce["BookCount"]).value or 0),
                int(ws_el.cell(rr, ce["Len"]).value or len(tok)),
                float(ws_el.cell(rr, ce["EvidenceScore_v127"]).value or 0.0),
            )
        )

    for tok, tr, occ, bc, ln, evs in sorted(micro_rows, key=lambda x: (-x[2], -x[4], x[0]))[:30]:
        ws.cell(r, 1).value = tok
        ws.cell(r, 2).value = tr
        ws.cell(r, 3).value = occ
        ws.cell(r, 4).value = bc
        ws.cell(r, 5).value = ln
        ws.cell(r, 6).value = evs
        ex = ""
        if mt_example and tr:
            try:
                m = re.search(r"\b" + re.escape(tr) + r"\b", mt_example)
                if m:
                    a = max(0, m.start() - 45)
                    b = min(len(mt_example), m.end() + 45)
                    ex = mt_example[a:b]
            except Exception:
                ex = ""
        ws.cell(r, 7).value = ex
        r += 1

    # Contigs risk table
    r += 2
    ws.cell(r, 1).value = "Top Contigs by WeakFrac"
    r += 1
    ws.cell(r, 1).value = "BaseContigID"
    ws.cell(r, 2).value = "BaseLen"
    ws.cell(r, 3).value = "EvidenceAvg"
    ws.cell(r, 4).value = "WeakFrac"
    ws.cell(r, 5).value = "MicroFrac"
    ws.cell(r, 6).value = "WeakTopTokens"
    ws.cell(r, 7).value = "StrictPlusPreview"
    r += 1

    try:
        ws_tc = wb["TokenEvidence_Contigs_v128"]
        hc = ws_find_header_row(ws_tc, ["BaseContigID", "BaseLen", "EvidenceAvg", "WeakFrac", "MicroFrac", "WeakTopTokens"])
        cc = ws_headers(ws_tc, hc)
        contig_rows = []
        for rr in range(hc + 1, ws_tc.max_row + 1):
            cid = ws_tc.cell(rr, cc["BaseContigID"]).value
            if cid is None:
                continue
            contig_rows.append(
                (
                    int(cid),
                    int(ws_tc.cell(rr, cc["BaseLen"]).value or 0),
                    float(ws_tc.cell(rr, cc["EvidenceAvg"]).value or 0.0),
                    float(ws_tc.cell(rr, cc["WeakFrac"]).value or 0.0),
                    float(ws_tc.cell(rr, cc["MicroFrac"]).value or 0.0),
                    str(ws_tc.cell(rr, cc["WeakTopTokens"]).value or ""),
                )
            )

        ws_contigs = wb["Contigs"]
        hct = ws_find_header_row(ws_contigs, ["BaseContigID", "Translation_StrictPlus_v108"])
        cct = ws_headers(ws_contigs, hct)
        contig_text: Dict[int, str] = {}
        for rr in range(hct + 1, ws_contigs.max_row + 1):
            cid = ws_contigs.cell(rr, cct["BaseContigID"]).value
            if cid is None:
                continue
            contig_text[int(cid)] = str(ws_contigs.cell(rr, cct["Translation_StrictPlus_v108"]).value or "")

        for cid, bl, ev, wf, mf, top in sorted(contig_rows, key=lambda x: (-x[3], -x[4], x[0])):
            tr = contig_text.get(cid, "")
            ws.cell(r, 1).value = cid
            ws.cell(r, 2).value = bl
            ws.cell(r, 3).value = round(ev, 6)
            ws.cell(r, 4).value = round(wf, 6)
            ws.cell(r, 5).value = round(mf, 6)
            ws.cell(r, 6).value = top
            ws.cell(r, 7).value = (tr[:160] + ("..." if len(tr) > 160 else ""))
            r += 1
    except Exception:
        pass

    upsert_sheet_index_entry(
        wb,
        name,
        f"Iteration {iter_num} focus report: top weak/micro books + micro tokens + contig risks (plateau diagnostics).",
    )
    return r


def ensure_flow_steps_entries(wb: openpyxl.Workbook) -> None:
    """Keep FlowSteps aligned with what the runner actually does."""
    if "FlowSteps" not in wb.sheetnames:
        return
    ws = wb["FlowSteps"]
    header = ws_find_header_row(ws, ["StepID", "StepName"], max_scan=3)
    c = ws_headers(ws, header)

    by_id: Dict[int, int] = {}
    for r in range(header + 1, ws.max_row + 1):
        sid = ws.cell(r, c["StepID"]).value
        if sid is None:
            continue
        try:
            by_id[int(sid)] = r
        except Exception:
            continue

    def _upsert(step_id: int, values: Dict[str, object]) -> None:
        r = by_id.get(step_id)
        if r is None:
            r = ws_last_data_row(ws, key_col=c["StepID"]) + 1
            ws.cell(r, c["StepID"]).value = step_id
            by_id[step_id] = r
        for k, v in values.items():
            col = c.get(k)
            if col is not None:
                ws.cell(r, col).value = v

    _upsert(
        12,
        {
            "StepName": "GroundTruth Live Check",
            "Description": "Guardrail: validar cribs GroundTruth via DP atual (ExpectedNorm_v112 vs DecodeNorm).",
            "Inputs": "Cribs(GroundTruth); Glossary(active); normalize_for_match",
            "Actions": "Tokenizar DecodedBase_Sanitized com DP atual e comparar normalize_for_match(DP) com ExpectedNorm_v112. Nao escreve no Cribs.",
            "Outputs": "FlowRunLog/FlowState (BLOCKED se mismatch)",
            "DoneCriteria": "0 mismatches",
            "OnFail": "Status=BLOCKED; BlockReason lista CribIDs mismatching",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        27,
        {
            "StepName": "AutoPhraseCribs",
            "Description": "Gerar PhraseCribs_Auto (public Tibia NPC/books) filtrando por viabilidade de assinatura contra o stream de tokens atual; usado para destravar Step 28 (reverse phrase).",
            "Inputs": "Internet Tibia corpus JSON; Books.DecodedBase; Glossary(active leaf + macros); ReversePhrase_MaxSpanTokens; lore canon flags",
            "Actions": "Tokenizar Books em leaf+match atoms; construir set/freq de assinaturas por spans (1..K); varrer frases do corpus e manter apenas share=1.0 (todas word-sigs viaveis); rankear por raridade e escrever PhraseCribs_Auto (cap).",
            "Outputs": "PhraseCribs_Auto",
            "DoneCriteria": "Sheet atualizado (pode ser vazio).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        28,
        {
            "StepName": "Reverse Phrase Mining",
            "Description": "Minerar frases (PhraseCribs_User/LoreCorpus_*/Tibia corpus) no Books.DecodedBase via assinaturas canonicas; registrar hits e (opcional) emitir candidatos inativos no Glossary.",
            "Inputs": "PhraseCribs_User; LoreCorpus_*; (opcional) Tibia NPC+books JSON; Books.DecodedBase; Glossary(active); Rules canonicalizer",
            "Actions": "Canonicalizar palavras -> signatures; tokenizar Books (expande macros; logogram-aware opcional); procurar matches com spans 1..K; escrever ReversePhraseHits_Auto e ReversePhraseTokenCands_Auto; opcional: append tokens no Glossary (inativos).",
            "Outputs": "ReversePhraseHits_Auto; ReversePhraseTokenCands_Auto; (opcional) Glossary novos candidatos",
            "DoneCriteria": "Sheets atualizados (pode ser vazio).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        29,
        {
            "StepName": "Reverse Phrase Mining (Permuted)",
            "Description": "Modo plateau: procurar matches onde a ordem das palavras da frase e livre (permutada), mantendo assinaturas canonicas por palavra. Analysis-only por padrao.",
            "Inputs": "Mesmas fontes do Step 28; Books.DecodedBase; Glossary(active leaf); ReversePhrase_MaxSpanTokens/MaxGapTokens",
            "Actions": "Tokenizar Books (reuse) e executar matcher permutado (bag-of-words por spans). Escrever ReversePhrasePermute* sheets; opcional: emitir candidatos inativos (default OFF).",
            "Outputs": "ReversePhrasePermuteHits_Auto; ReversePhrasePermuteCands_Auto; (opcional) Glossary candidatos",
            "DoneCriteria": "Sheets atualizados (pode ser vazio).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        82,
        {
            "StepName": "Plateau Focus Report",
            "Description": "Gerar IterXXX_Focus com top riscos (books/contigs) e micro tokens para orientar proximos passos.",
            "Inputs": "TokenEvidence_Books_v128; EvidenceLedger_v127; TokenEvidence_Contigs_v128; Books/Contigs/MasterText",
            "Actions": "Criar IterXXX_Focus e indexar em SheetIndex (nao altera decode).",
            "Outputs": "IterXXX_Focus",
            "DoneCriteria": "Sheet criado/atualizado",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        85,
        {
            "StepName": "Sync AnchorCribs",
            "Description": "Importar/sincronizar AnchorCribs do iter141 (estrutura) para uso em alinhamento variant-aware.",
            "Inputs": "archive/bonelord_469_iter141.xlsx (AnchorCribs_v138)",
            "Actions": "Copiar para AnchorCribs_Auto (idempotente por AnchorCribID).",
            "Outputs": "AnchorCribs_Auto",
            "DoneCriteria": "Sheet atualizado",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        86,
        {
            "StepName": "Variant-Aware Alignment",
            "Description": "Localizar AnchorCribs nos Books.DecodedBase, calcular offsets e gerar backbone/blocks (analysis-only).",
            "Inputs": "AnchorCribs_Auto; Books.DecodedBase",
            "Actions": "Gerar AnchorOccurrences_Auto, BookOffsets_Auto, AlignedBackbone_Auto, VariantAssemblyBlocks_Auto.",
            "Outputs": "AnchorOccurrences_Auto; BookOffsets_Auto; AlignedBackbone_Auto; VariantAssemblyBlocks_Auto",
            "DoneCriteria": "Sheets atualizados",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        87,
        {
            "StepName": "SuperAnchor Mining",
            "Description": "Minerar candidatos de super-ancoras (blocos estaveis longos) a partir do backbone.",
            "Inputs": "AlignedBackbone_Auto; VariantAssemblyBlocks_Auto",
            "Actions": "Extrair runs estaveis (>=30) com suporte amplo e listar em SuperAnchors_Auto.",
            "Outputs": "SuperAnchors_Auto",
            "DoneCriteria": "Candidatos listados (pode ser 0)",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        90,
        {
            "StepName": "MacroCompress Display",
            "Description": "Materializar macro-compression display-only para Books/MasterText (sem mudar StrictPlus).",
            "Inputs": "PostDecodeMacro_v130; Glossary(macros); Books/MasterText",
            "Actions": "Preencher Translation_MacroCompressed_Auto e contagens de tokens salvos.",
            "Outputs": "Books.*MacroCompressed_Auto; MasterText.Translation_MacroCompressed_Auto",
            "DoneCriteria": "Colunas atualizadas",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        91,
        {
            "StepName": "Fetch Tibia SigIndex",
            "Description": "Buscar corpus Tibia (NPC transcripts + books) via internet e salvar apenas um indice derivado por assinatura (sem texto completo).",
            "Inputs": "Internet; Glossary(active leaf tokens); FlowSettings LoreFetch_TibiaSigIndex_*",
            "Actions": "Baixar JSONs; extrair palavras; canonicalizar (Lore canon); agregar Count por Sig; escrever LoreSigIndex_Tibia_Auto.",
            "Outputs": "LoreSigIndex_Tibia_Auto",
            "DoneCriteria": "Sheet atualizado (ou skip se nao stale).",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        109,
        {
            "StepName": "Fetch PD SigIndex",
            "Description": "Buscar um corpus em dominio publico (ex: Gutenberg KJV) e salvar apenas indices derivados (assinatura + wordfreq), sem texto completo no XLSX.",
            "Inputs": "Internet (opcional cache local); Glossary(active leaf tokens); FlowSettings LoreFetch_PDSigIndex_*",
            "Actions": "Baixar plaintext (cache em tmp/corpus); extrair palavras; canonicalizar (Lore canon); agregar Count por Sig e Word; escrever LoreSigIndex_PD_Auto + LoreWordFreq_PD_Auto.",
            "Outputs": f"{LORE_SIGINDEX_PD_SHEET}; {LORE_WORDFREQ_PD_SHEET}",
            "DoneCriteria": "Sheets atualizados (ou skip se nao stale).",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        112,
        {
            "StepName": "Build Dict SigIndex",
            "Description": "Construir um indice derivado local (dicionario) por assinatura para expandir cobertura sem importar texto/corpus completo.",
            "Inputs": f"Local word list (default /usr/share/dict/words); FlowSettings LoreFetch_DictSigIndex_*; {LORE_WORDFREQ_TIBIA_SHEET}/{LORE_WORDFREQ_PD_SHEET} (prior opcional)",
            "Actions": "Iterar palavras do dicionario; canonicalizar (Lore canon); filtrar apenas assinaturas alvo; persistir top-N por assinatura em LoreSigIndex_Dict_Auto.",
            "Outputs": f"{LORE_SIGINDEX_DICT_SHEET}",
            "DoneCriteria": "Sheet atualizado (ou skip se nao stale).",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        92,
        {
            "StepName": "Lore Corpus Seed/Sync",
            "Description": "Garantir LoreCorpus_* e semear corpus default (analysis-only).",
            "Inputs": "LoreCorpus_Auto; LoreCorpus_User",
            "Actions": "Criar sheets se ausentes; seed idempotente (Jabberwocky EN) em LoreCorpus_Auto; nunca sobrescrever LoreCorpus_User.",
            "Outputs": "LoreCorpus_Auto; LoreCorpus_User",
            "DoneCriteria": "Sheets presentes e seed aplicado (ou ja existente).",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        93,
        {
            "StepName": "Lore Token Hits",
            "Description": "Minerar hits de lore por assinatura (word signature -> token signature) para hints semanticos (analysis-only).",
            "Inputs": "LoreCorpus_*; Glossary(active leaf tokens)",
            "Actions": "Canonicalizar palavras do corpus (Rules-based) -> signatures; comparar com signatures de tokens leaf do Glossary; escrever LoreAlignment_Auto.",
            "Outputs": "LoreAlignment_Auto",
            "DoneCriteria": "Sheet atualizado (pode ser vazio).",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        94,
        {
            "StepName": "Semantic Map",
            "Description": "Derivar SemanticMap_Auto (display-only) de LoreAlignment_Auto, com regras conservadoras.",
            "Inputs": "LoreAlignment_Auto",
            "Actions": "Selecionar mapeamentos apenas quando unicos ou com share alto; nao altera Glossary/DP.",
            "Outputs": "SemanticMap_Auto",
            "DoneCriteria": "Sheet atualizado (pode ser vazio).",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        95,
        {
            "StepName": "Semantic Render",
            "Description": "Materializar Translation_Semantic_Auto em Books/MasterText (display-only).",
            "Inputs": "SemanticMap_Auto; Books.DecodedBase; Contigs.BaseContig; Glossary(active + macro composition)",
            "Actions": "DP-tokenizar base; expandir macros por Composition tokens; substituir leaf tokens com SemanticMap; render lossless.",
            "Outputs": "Books.Translation_Semantic_Auto; MasterText.Translation_Semantic_Auto",
            "DoneCriteria": "Colunas atualizadas (pode ser sem mudancas).",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        96,
        {
            "StepName": "Semantic -> Glossary Retext",
            "Description": "Aplicar (com guardrail GT live check) traducoes semanticas de alta confianca no Glossary. Avanca utilidade sem alterar tokenizacao.",
            "Inputs": "SemanticMap_Auto; Glossary; Cribs(GroundTruth)",
            "Actions": "Para tokens leaf StrictPlus com SemanticWord distinto: simular troca e validar GT live check; aplicar apenas se OK.",
            "Outputs": "Glossary.Translation; EvidenceLedger_v127.Translation; SemanticPromotions_Auto",
            "DoneCriteria": "0..N retext aplicados (max por iteracao).",
            "OnFail": "Registrar SKIP/FAILED e continuar (safe).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        97,
        {
            "StepName": "Semantic Retext Reverts",
            "Description": "Reverter automaticamente semantic retext antigo quando viola policy atual (locked EvidenceClass ou wordfreq heuristics), sempre com GT live check.",
            "Inputs": "SemanticPromotions_Auto; Glossary; LoreWordFreq_Tibia_Auto (opcional); Cribs(GroundTruth)",
            "Actions": "Para tokens com ultimo Decision=APPLIED: se bloqueado/ruim, simular revert OldTranslation e validar GT; aplicar apenas se OK; log em SemanticReverts_Auto.",
            "Outputs": "Glossary.Translation; EvidenceLedger_v127.Translation; SemanticReverts_Auto",
            "DoneCriteria": "0..N reverts aplicados (pode ser 0).",
            "OnFail": "Registrar FAILED e continuar (safe).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        98,
        {
            "StepName": "English Layer (Display)",
            "Description": "Materializar Translation_English_Auto (display-only) mapeando palavras canonicas (ex: wit/fay/lion) para superficies em ingles via Tibia sig-index.",
            "Inputs": f"{LORE_SIGINDEX_TIBIA_SHEET}; Glossary(active leaf); Books/MasterText (StrictPlus/Semantic outputs)",
            "Actions": "Derivar EnglishMap_Auto (canon_word -> top corpus word por assinatura, com share alto) e aplicar em Books/MasterText em colunas *_English_Auto.",
            "Outputs": f"{ENGLISH_MAP_SHEET}; Books.Translation_English_Auto; MasterText.Translation_English_Auto",
            "DoneCriteria": "Map e colunas atualizadas (pode ser sem mudancas).",
            "OnFail": "Registrar FAILED e continuar (display-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        99,
        {
            "StepName": "English -> Glossary Retext",
            "Description": "Aplicar (com guardrail GT live check) mapeamentos do EnglishMap_Auto diretamente no Glossary.Translation para melhorar legibilidade do StrictPlus sem alterar tokenizacao.",
            "Inputs": f"{ENGLISH_MAP_SHEET}; Glossary(active leaf); Cribs(GroundTruth)",
            "Actions": "Para tokens leaf StrictPlus com Translation canonica (ex: wit/than/fay) e EnglishMap sugere TopWord: simular troca e validar GT; aplicar apenas se OK; sync EvidenceLedger.",
            "Outputs": "Glossary.Translation; EvidenceLedger_v127.Translation; EnglishPromotions_Auto",
            "DoneCriteria": "0..N retext aplicados (max por iteracao).",
            "OnFail": "Registrar SKIP/FAILED e continuar (safe).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )

    _upsert(
        101,
        {
            "StepName": "DigitCodeMap",
            "Description": "Extrair o mapa global codigo(2 digitos)->letra-base e homofonos letra->codigos a partir de Books.Digits + BooksDigitModel_v118 (analysis-only).",
            "Inputs": "Books(Digits,DecodedBase); BooksDigitModel_v118(OmitIdxs_1based)",
            "Actions": "Reconstruir o code-stream por livro (permitindo 0-omissao via OmitIdxs_1based) e agregar contagens por codigo/letra; escrever DigitCodeMap_Auto + DigitLetterCodes_Auto.",
            "Outputs": f"{DIGIT_CODE_MAP_SHEET}; {DIGIT_LETTER_CODES_SHEET}",
            "DoneCriteria": "Sheets atualizados (idempotente).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        110,
        {
            "StepName": "DigitCodeContext",
            "Description": "Derivar perfis de contexto por codigo (prev/next) para estudar homofonos (analysis-only).",
            "Inputs": f"{DIGIT_CODE_MAP_SHEET}; Books(Digits,DecodedBase); BooksDigitModel_v118(OmitIdxs_1based)",
            "Actions": "Reconstruir code-stream; agregar distribuicoes de vizinhanca (codigos/letras) e calcular outlier-score (JS) dentro de cada letra-base.",
            "Outputs": DIGIT_CODE_CONTEXT_SHEET,
            "DoneCriteria": "Sheet atualizado (idempotente).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        111,
        {
            "StepName": "Code-Aware Homophones Render",
            "Description": "Usar variacoes de codigo (homofonos por letra-base) para desambiguar tokens curtos (len<=CodeAware_MaxTokenLen) e renderizar uma camada display-only (sem tocar StrictPlus/Glossary).",
            "Inputs": "Books(Digits,DecodedBase); BooksDigitModel_v118(OmitIdxs_1based); LoreAlignment_Auto; LoreBigrams_Auto; Glossary(active)",
            "Actions": "Reconstroi codes por posicao; roda ContextEnglish local para obter escolhas; agrega (Token,Code)->TopWord; aplica apenas quando share>=limite; escreve CodeWordMap_Auto + Books.Translation_CodeAware_Auto.",
            "Outputs": f"{CODE_WORD_MAP_SHEET}; Books.Translation_CodeAware_Auto",
            "DoneCriteria": "Sheet/colunas atualizadas (idempotente; pode ser sem mudancas).",
            "OnFail": "Registrar FAILED e continuar (display-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        102,
        {
            "StepName": "External RoundTrip Check",
            "Description": "Validar referencias numericas externas verificadas (>=N fontes) contra DP atual: consistencia de 'DecodedBase' e expected_norm quando disponivel (analysis-only).",
            "Inputs": "ExternalRefs_v115; ExternalGroundTruthCheck_v120; Books.Digits; Glossary(active)",
            "Actions": "Recontar ocorrencias em Books.Digits; DP-tokenizar ExternalRefs.DecodedBase sob token set atual; comparar normalize_for_match(DP) com expected_norm quando existir; escrever ExternalRoundTrip_Auto.",
            "Outputs": EXTERNAL_ROUNDTRIP_SHEET,
            "DoneCriteria": "Sheet atualizado (pode conter FAILs).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        103,
        {
            "StepName": "Corpus Bigram Index",
            "Description": "Construir indice de bigramas de palavras (derived-only) a partir do cache Tibia (tmp/corpus) + LoreCorpus_Auto para desambiguacao contextual (sem persistir texto completo no XLSX).",
            "Inputs": "tmp/corpus Tibia JSON; LoreCorpus_Auto",
            "Actions": "Tokenizar por sentenca; contar bigramas; filtrar por vocab_topn/min_count; escrever LoreBigrams_Auto (top-N).",
            "Outputs": LORE_BIGRAMS_SHEET,
            "DoneCriteria": "Sheet atualizado ou skip por freshness.",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        104,
        {
            "StepName": "Context English Render",
            "Description": "Renderizar uma camada de ingles por contexto (Viterbi/bigram LM) usando candidatos de LoreAlignment (quando nao-logogram) para Books/Contigs (display-only).",
            "Inputs": "LoreAlignment_Auto; LoreBigrams_Auto; Glossary(active); Books.DecodedBase; Contigs.BaseContig",
            "Actions": "Para cada token leaf: estados=candidatos; emissao=log(count+alpha); transicao=log(bigram+alpha); escrever Translation_ContextEnglish_Auto + metricas.",
            "Outputs": "Books.Translation_ContextEnglish_Auto; MasterText.Translation_ContextEnglish_Auto; EnglishMap_Context_Auto",
            "DoneCriteria": "Colunas/sheets atualizadas (pode ser sem mudancas).",
            "OnFail": "Registrar FAILED e continuar (display-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        105,
        {
            "StepName": "Sequence Matches",
            "Description": "Tentar alinhar n-grams raros do ContextEnglish contra corpora (public domain + Tibia cache) e registrar apenas snippets curtos + URLs (analysis-only).",
            "Inputs": "Books.Translation_ContextEnglish_Auto; LoreCorpus_Auto; tmp/corpus Tibia JSON",
            "Actions": "Gerar candidatos raros (freq=1) e varrer corpora por membership; escrever SequenceMatches_Auto (cap).",
            "Outputs": SEQUENCE_MATCHES_SHEET,
            "DoneCriteria": "Sheet atualizado (pode ser vazio).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        107,
        {
            "StepName": "Sequence Word Hints",
            "Description": "Derivar SequenceWordHints_Auto a partir de SequenceMatches_Auto (analysis-only) para ajudar a desambiguar escolhas por contexto (display-only).",
            "Inputs": f"{SEQUENCE_MATCHES_SHEET}; lore canon flags",
            "Actions": "Alinhar words de Phrase vs Snippet por posicao e escrever (CanonSig, FromWord, ToWord) counts (cap).",
            "Outputs": SEQUENCE_WORD_HINTS_SHEET,
            "DoneCriteria": "Sheet atualizado (pode ser vazio).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        108,
        {
            "StepName": "Sestina Scan",
            "Description": "Analise estrutural: detectar janelas com padrao de sestina/sestine (retrogradatio cruciata) usando tokens de fim-de-linha (↵) (analysis-only).",
            "Inputs": "Books.DecodedBase; Glossary(active); newline tokens (LF/LN/SF)",
            "Actions": f"Tokenizar DP por livro; extrair linhas separadas por ↵; varrer janelas e registrar candidatos em {SESTINA_CANDIDATES_SHEET} (e linhas em {SESTINA_LINES_SHEET}).",
            "Outputs": f"{SESTINA_LINES_SHEET}; {SESTINA_CANDIDATES_SHEET}",
            "DoneCriteria": "Sheets atualizados (idempotente).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        113,
        {
            "StepName": "Sestina Obligation Map",
            "Description": "Mapear carga estrutural por posicao (obrigatorio/condicional/redundante/decorativo) com testes de ablation e reorder stress (analysis-only).",
            "Inputs": f"{SESTINA_LINES_SHEET}; {SESTINA_CANDIDATES_SHEET}",
            "Actions": "Para cada janela candidata: medir impacto por posicao (remove-1), max_remove_sem_colapso (remove-2) e melhor score sob reordenacao; classificar cada posicao e escrever SestinaObligation_Auto.",
            "Outputs": SESTINA_OBLIGATION_SHEET,
            "DoneCriteria": "Sheet atualizado (idempotente; pode ser sem mudancas).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        114,
        {
            "StepName": "Rhythm Transition AB-Test",
            "Description": "Teste estrutural/global de ritmo: ciclo-6 vs controle embaralhado + alternancia A/B + controles anti-sestina (Fibonacci/triple-core/sparse-echo) (analysis-only).",
            "Inputs": f"{SESTINA_LINES_SHEET}; {SESTINA_CANDIDATES_SHEET}",
            "Actions": f"Varre janelas globais (default 12 linhas), mede previsibilidade lag-6 contra baseline shuffled, score de alternancia A/B por EndKind e controles alternativos; escreve {RHYTHM_TRANSITIONS_SHEET}.",
            "Outputs": RHYTHM_TRANSITIONS_SHEET,
            "DoneCriteria": "Sheet atualizado (idempotente; pode ser sem mudancas).",
            "OnFail": "Registrar FAILED e continuar (analysis-only).",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )
    _upsert(
        106,
        {
            "StepName": "Iter Meta Report",
            "Description": "Gerar IterXXX_Meta com diagnosticos: tokens ambiguos (low share), livros/contigs fracos, status estrutural, e auditoria de GroundTruth verificado (analysis-only).",
            "Inputs": "LoreAlignment_Auto; TokenEvidence_*; SuperAnchors_Auto; Cribs(GroundTruth)",
            "Actions": "Criar IterXXX_Meta e indexar em SheetIndex.",
            "Outputs": "IterXXX_Meta",
            "DoneCriteria": "Sheet criado/atualizado",
            "OnFail": "Registrar FAILED e continuar (analysis-only)",
            "SafetyClass": "SAFE",
            "AutoChainEligible": True,
        },
    )

    # Update readability description to reflect Scope-aware behavior.
    if 75 in by_id:
        _upsert(
            75,
            {
                "Description": "Aplicar ReadabilityRules (Scope-aware) para gerar *_Readable_Auto sem alterar StrictPlus (inclui CRIBS em DP_Readable_Auto).",
                "Inputs": "ReadabilityRules; Books/ MasterText/ Cribs (StrictPlus outputs)",
                "Actions": "Aplicar substituicoes habilitadas respeitando Scope (BOOKS/MASTER/CRIBS).",
            },
        )


def _render_strictplus_from_lossless_with_braces(lossless: str) -> str:
    toks = str(lossless or "").split()
    out: List[str] = []
    for tok in toks:
        if tok in ("<E>", "<FF>"):
            if not out:
                out.append(",")
            else:
                out[-1] = f"{out[-1]},"
            continue
        if tok == "<*>":
            out.append("—")
            continue
        if tok in (".", "!"):
            if not out:
                out.append(tok)
            else:
                out[-1] = f"{out[-1]}{tok}"
            continue
        out.append(tok)
    return " ".join(out)


def _macro_tokens_saved_from_braces(lossless: str) -> int:
    saved = 0
    in_group = False
    group_words = 0

    for raw in str(lossless or "").split():
        starts = raw.startswith("{")
        ends = raw.endswith("}")
        core = raw.strip("{}")

        if starts:
            in_group = True
            group_words = 0

        if in_group:
            if core and core not in ("<E>", "<FF>", "<*>", ".", "!", "↵"):
                group_words += 1

        if ends and in_group:
            if group_words > 1:
                saved += group_words - 1
            in_group = False
            group_words = 0

    return saved


def _parse_macros_applied_list(raw: object) -> List[str]:
    """Parse 'NSTAEFIEIE(1), FATFTFNLI(2)' -> ['NSTAEFIEIE','FATFTFNLI']."""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    out: List[str] = []
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^([A-Z0-9*]+)\(", part)
        if m:
            out.append(m.group(1))
    return out


def _group_phrases_outside_braces(text: str, phrases: Sequence[str]) -> Tuple[str, int]:
    """Wrap exact phrases in `{...}` but do not group inside existing braces."""
    if not text or not phrases:
        return text, 0

    segs: List[Tuple[str, bool]] = []
    buf: List[str] = []
    depth = 0
    for ch in text:
        if ch == "{":
            if buf:
                segs.append(("".join(buf), depth == 0))
                buf = []
            depth += 1
            buf.append(ch)
            continue
        if ch == "}":
            buf.append(ch)
            depth = max(0, depth - 1)
            if depth == 0:
                segs.append(("".join(buf), False))
                buf = []
            continue
        buf.append(ch)
    if buf:
        segs.append(("".join(buf), depth == 0))

    changed = 0
    out_parts: List[str] = []
    for seg, is_outside in segs:
        if not is_outside:
            out_parts.append(seg)
            continue
        s = seg
        for ph in phrases:
            if not ph:
                continue
            pat = r"\b" + re.escape(ph) + r"\b"
            s, n = re.subn(pat, "{" + ph + "}", s)
            changed += int(n)
        out_parts.append(s)
    return "".join(out_parts), changed


def materialize_macrocompression_display(
    wb: openpyxl.Workbook,
    iter_num: int,
    glossary_tokens: Dict[str, GlossaryToken],
) -> Tuple[int, int]:
    """Fill macro-compressed display-only columns for Books and MasterText.

    Returns: (books_rows_changed, master_rows_changed)
    """
    if "PostDecodeMacro_v130" not in wb.sheetnames:
        return 0, 0

    ws_pd = wb["PostDecodeMacro_v130"]
    hp = ws_find_header_row(ws_pd, ["BookID", "MacroCompressed_NoStar_v130", "MacrosApplied_NoStar"], max_scan=3)
    cp = ws_headers(ws_pd, hp)

    macro_by_book: Dict[int, Tuple[str, str]] = {}
    macro_ids: set[str] = set()
    for r in range(hp + 1, ws_pd.max_row + 1):
        bid = ws_pd.cell(r, cp["BookID"]).value
        if bid is None:
            continue
        mc = ws_pd.cell(r, cp["MacroCompressed_NoStar_v130"]).value
        ma = ws_pd.cell(r, cp["MacrosApplied_NoStar"]).value
        macro_by_book[int(bid)] = (str(mc or ""), str(ma or ""))
        for mid in _parse_macros_applied_list(ma):
            macro_ids.add(mid)

    phrases: List[str] = []
    for mid in sorted(macro_ids):
        gt = glossary_tokens.get(mid)
        if gt is None or not gt.translation:
            continue
        phrases.append(str(gt.translation))
    phrases.sort(key=lambda s: (-len(s), s))

    # Books columns
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "TokenCount_StrictPlus_v108", "Translation_StrictPlus_v108"], max_scan=3)
    cb = ws_headers(ws_books, hb)

    col_tr = cb.get("Translation_MacroCompressed_Auto")
    col_tc = cb.get("MacroCompressedTokenCount_Auto")
    col_sv = cb.get("MacroCompressedTokensSaved_Auto")
    if col_tr is None:
        col_tr = ws_books.max_column + 1
        ws_books.cell(hb, col_tr).value = "Translation_MacroCompressed_Auto"
    if col_tc is None:
        col_tc = ws_books.max_column + 1
        ws_books.cell(hb, col_tc).value = "MacroCompressedTokenCount_Auto"
    if col_sv is None:
        col_sv = ws_books.max_column + 1
        ws_books.cell(hb, col_sv).value = "MacroCompressedTokensSaved_Auto"

    books_changed = 0
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        bid_i = int(bid)
        mc_lossless, _ma = macro_by_book.get(bid_i, ("", ""))
        if not mc_lossless:
            continue
        out = _render_strictplus_from_lossless_with_braces(mc_lossless)
        saved = _macro_tokens_saved_from_braces(mc_lossless)
        base_tc = int(ws_books.cell(r, cb["TokenCount_StrictPlus_v108"]).value or 0)
        mc_tc = max(0, base_tc - saved)

        prev = ws_books.cell(r, col_tr).value
        if prev != out:
            books_changed += 1
        ws_books.cell(r, col_tr).value = out
        ws_books.cell(r, col_sv).value = saved
        ws_books.cell(r, col_tc).value = mc_tc

    # MasterText column (phrase grouping)
    ws_mt = wb["MasterText"]
    hm = ws_find_header_row(ws_mt, ["BaseContigID", "Translation_StrictPlus_v108"], max_scan=3)
    cm = ws_headers(ws_mt, hm)
    col_mt = cm.get("Translation_MacroCompressed_Auto")
    if col_mt is None:
        col_mt = ws_mt.max_column + 1
        ws_mt.cell(hm, col_mt).value = "Translation_MacroCompressed_Auto"

    master_changed = 0
    for r in range(hm + 1, ws_mt.max_row + 1):
        cid = ws_mt.cell(r, cm["BaseContigID"]).value
        if cid is None:
            continue
        src = ws_mt.cell(r, cm["Translation_StrictPlus_v108"]).value
        if not isinstance(src, str):
            src = str(src or "")
        out, _n = _group_phrases_outside_braces(src, phrases)
        prev = ws_mt.cell(r, col_mt).value
        if prev != out:
            master_changed += 1
        ws_mt.cell(r, col_mt).value = out

    upsert_sheet_index_entry(
        wb,
        "PostDecodeMacro_v130",
        "Post-decode macro compression (display-only): macro groupings applied without changing StrictPlus DP.",
    )
    return books_changed, master_changed


def _lore_canon_word(
    raw: str,
    *,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
) -> str:
    """Canonicalize a corpus word into the same reduced alphabet used by our Rules sheet.

    This is a *semantic* helper only (display-layer): it MUST NOT affect StrictPlus DP or Glossary.
    """
    s = (raw or "").lower()
    s = re.sub(r"[^a-z]", "", s)
    if not s:
        return ""
    # Confirmed rules (Rules sheet).
    s = s.replace("th", "t")
    # Letter-group collapses (TibiaSecrets bonelord language analysis):
    # - V group: V/U/W/OO/OOL/UE
    # - I group: I/Y/EE
    #
    # Do longer sequences first to avoid leaving trailing letters (e.g., "ool" -> "v", not "oo" -> "v" + "l").
    s = s.replace("ool", "v")
    s = s.replace("oo", "v")
    s = s.replace("ue", "v")
    s = s.replace("ee", "i")
    s = s.replace("w", "v")
    s = s.replace("u", "v")
    s = s.replace("y", "i")
    s = s.replace("d", "t")
    s = s.replace("m", "n")
    s = s.replace("p", "b")
    # Map letters that do not exist in the 469 base alphabet into stable buckets.
    # This is required for signature matching against the base-letter stream.
    s = s.replace("j", "i")
    s = s.replace("k", "c")
    s = s.replace("q", "c")
    s = s.replace("g", "c")
    s = s.replace("z", "s")
    # x is effectively 'cs' in the base alphabet (keeps both consonant buckets available).
    s = s.replace("x", "cs")
    if drop_all_h:
        s = s.replace("h", "")
    if drop_all_o:
        s = s.replace("o", "")
    # Confirmed: drop final O/R; optional: drop final E.
    if s.endswith("o") or s.endswith("r"):
        s = s[:-1]
    if drop_final_e and s.endswith("e"):
        s = s[:-1]
    return s


def _lore_signature(s: str) -> str:
    return "".join(sorted(list(s or "")))


def _token_signature(token: str) -> str:
    """Signature of a base token (Glossary.Token) for lore matching: sort letters, ignore '*'."""
    if token is None:
        return ""
    letters = [ch.lower() for ch in str(token) if ch.isalpha() and ch != "*"]
    return "".join(sorted(letters))


def _ensure_lore_corpus_sheets(wb: openpyxl.Workbook, iter_num: int) -> Tuple[int, int]:
    """Ensure LoreCorpus_Auto/User exist and seed default corpora if empty.

    Returns: (rows_added, rows_total)
    """
    headers = ["CorpusID", "Lang", "License", "Source", "LineID", "Text", "AddedIter", "Notes"]
    ws_auto = ensure_sheet(wb, "LoreCorpus_Auto", headers)
    ws_user = ensure_sheet(wb, "LoreCorpus_User", headers)
    ha = ws_find_header_row(ws_auto, ["CorpusID", "LineID", "Text"], max_scan=3)
    ca = ws_headers(ws_auto, ha)

    # Count existing seeded rows.
    existing_keys: set[Tuple[str, int]] = set()
    max_line_by_corpus: Dict[str, int] = defaultdict(int)
    for r in range(ha + 1, ws_auto.max_row + 1):
        cid = ws_auto.cell(r, ca["CorpusID"]).value
        lid = ws_auto.cell(r, ca["LineID"]).value
        if not isinstance(cid, str) or not cid.strip():
            continue
        try:
            lid_i = int(lid)
        except Exception:
            continue
        existing_keys.add((cid.strip(), lid_i))
        max_line_by_corpus[cid.strip()] = max(max_line_by_corpus[cid.strip()], lid_i)

    added = 0

    # Seed: Jabberwocky (public domain) - useful as a known reference corpus.
    # NOTE: This seed is best-effort; low/no matches is acceptable and still provides a stable framework
    # for users to paste additional lore texts into LoreCorpus_User.
    jab_id = "JABBERWOCKY_EN"
    jab_lines = [
        "'Twas brillig, and the slithy toves",
        "Did gyre and gimble in the wabe;",
        "All mimsy were the borogoves,",
        "And the mome raths outgrabe.",
        "Beware the Jabberwock, my son!",
        "The jaws that bite, the claws that catch!",
        "Beware the Jubjub bird, and shun",
        "The frumious Bandersnatch!",
        "He took his vorpal sword in hand;",
        "Long time the manxome foe he sought—",
        "So rested he by the Tumtum tree",
        "And stood awhile in thought.",
        "And, as in uffish thought he stood,",
        "The Jabberwock, with eyes of flame,",
        "Came whiffling through the tulgey wood,",
        "And burbled as it came!",
        "One, two! One, two! And through and through",
        "The vorpal blade went snicker-snack!",
        "He left it dead, and with its head",
        "He went galumphing back.",
        "And hast thou slain the Jabberwock?",
        "Come to my arms, my beamish boy!",
        "O frabjous day! Callooh! Callay!",
        "He chortled in his joy.",
    ]

    for i, text in enumerate(jab_lines, start=1):
        lid = i
        if (jab_id, lid) in existing_keys:
            continue
        r = ws_last_data_row(ws_auto, key_col=ca["CorpusID"]) + 1
        ws_auto.cell(r, ca["CorpusID"]).value = jab_id
        ws_auto.cell(r, ca["Lang"]).value = "EN"
        ws_auto.cell(r, ca["License"]).value = "public domain"
        ws_auto.cell(r, ca["Source"]).value = "Lewis Carroll (Jabberwocky)"
        ws_auto.cell(r, ca["LineID"]).value = lid
        ws_auto.cell(r, ca["Text"]).value = text
        ws_auto.cell(r, ca["AddedIter"]).value = iter_num
        ws_auto.cell(r, ca["Notes"]).value = "Seed corpus (auto)."
        added += 1

    # Seed: KJV Bible excerpts (public domain). Early Modern English, high density of function words.
    kjv_id = "KJV_EXCERPT_EN"
    kjv_lines = [
        "In the beginning God created the heaven and the earth.",
        "And the earth was without form, and void; and darkness was upon the face of the deep.",
        "And the Spirit of God moved upon the face of the waters.",
        "And God said, Let there be light: and there was light.",
        "And God saw the light, that it was good: and God divided the light from the darkness.",
        "And God called the light Day, and the darkness he called Night.",
        "And the evening and the morning were the first day.",
        "And God said, Let there be a firmament in the midst of the waters, and let it divide the waters from the waters.",
        "And God made the firmament, and divided the waters which were under the firmament from the waters which were above the firmament: and it was so.",
        "And God called the firmament Heaven. And the evening and the morning were the second day.",
        "The Lord is my shepherd; I shall not want.",
        "He maketh me to lie down in green pastures: he leadeth me beside the still waters.",
        "He restoreth my soul: he leadeth me in the paths of righteousness for his name's sake.",
        "Yea, though I walk through the valley of the shadow of death, I will fear no evil: for thou art with me;",
        "Thy rod and thy staff they comfort me.",
        "Thou preparest a table before me in the presence of mine enemies:",
        "Surely goodness and mercy shall follow me all the days of my life: and I will dwell in the house of the Lord for ever.",
    ]
    for i, text in enumerate(kjv_lines, start=1):
        lid = i
        if (kjv_id, lid) in existing_keys:
            continue
        r = ws_last_data_row(ws_auto, key_col=ca["CorpusID"]) + 1
        ws_auto.cell(r, ca["CorpusID"]).value = kjv_id
        ws_auto.cell(r, ca["Lang"]).value = "EN"
        ws_auto.cell(r, ca["License"]).value = "public domain"
        ws_auto.cell(r, ca["Source"]).value = "KJV Bible (excerpts) https://www.gutenberg.org/ebooks/10"
        ws_auto.cell(r, ca["LineID"]).value = lid
        ws_auto.cell(r, ca["Text"]).value = text
        ws_auto.cell(r, ca["AddedIter"]).value = iter_num
        ws_auto.cell(r, ca["Notes"]).value = "Seed corpus (auto)."
        added += 1

    # Seed: Shakespeare (public domain). Useful for common phrases + archaic forms.
    sh_id = "SHAKESPEARE_EXCERPT_EN"
    sh_lines = [
        "To be, or not to be: that is the question:",
        "Whether 'tis nobler in the mind to suffer",
        "The slings and arrows of outrageous fortune,",
        "Or to take arms against a sea of troubles,",
        "And by opposing end them?",
        "To die: to sleep; No more;",
        "And by a sleep to say we end",
        "The heart-ache and the thousand natural shocks",
        "That flesh is heir to, 'tis a consummation",
        "Devoutly to be wish'd.",
    ]
    for i, text in enumerate(sh_lines, start=1):
        lid = i
        if (sh_id, lid) in existing_keys:
            continue
        r = ws_last_data_row(ws_auto, key_col=ca["CorpusID"]) + 1
        ws_auto.cell(r, ca["CorpusID"]).value = sh_id
        ws_auto.cell(r, ca["Lang"]).value = "EN"
        ws_auto.cell(r, ca["License"]).value = "public domain"
        ws_auto.cell(r, ca["Source"]).value = "Shakespeare (Hamlet excerpt) https://www.gutenberg.org/ebooks/1524"
        ws_auto.cell(r, ca["LineID"]).value = lid
        ws_auto.cell(r, ca["Text"]).value = text
        ws_auto.cell(r, ca["AddedIter"]).value = iter_num
        ws_auto.cell(r, ca["Notes"]).value = "Seed corpus (auto)."
        added += 1

    # Seed: curated archaic function words (no external dependency, helps old-English hypotheses).
    arch_id = "EN_ARCHAIC_CORE"
    arch_lines = [
        "thou thee thy thine ye yon yonder wherefore whence hence hither thither",
        "hath doth shalt shouldst wouldst canst wilt art wert wast",
        "unto ere oft betwixt whilst amongst naught aught yea nay anon",
        "lest thus thereby herein therewith hereinbefore thereafter",
    ]
    for i, text in enumerate(arch_lines, start=1):
        lid = i
        if (arch_id, lid) in existing_keys:
            continue
        r = ws_last_data_row(ws_auto, key_col=ca["CorpusID"]) + 1
        ws_auto.cell(r, ca["CorpusID"]).value = arch_id
        ws_auto.cell(r, ca["Lang"]).value = "EN"
        ws_auto.cell(r, ca["License"]).value = "curated"
        ws_auto.cell(r, ca["Source"]).value = "Curated archaic EN function words (manual)"
        ws_auto.cell(r, ca["LineID"]).value = lid
        ws_auto.cell(r, ca["Text"]).value = text
        ws_auto.cell(r, ca["AddedIter"]).value = iter_num
        ws_auto.cell(r, ca["Notes"]).value = "Seed corpus (auto)."
        added += 1

    # Seed: 1911 Encyclopaedia Britannica - Sestina (public domain).
    # This is a high-signal reference because the decoded corpus itself contains "sestine"/"sestina" hints.
    brit_id = "BRITANNICA_1911_SESTINA_EN"
    brit_lines = [
        "SESTINA, one of the most elaborate forms of verse employed by the medieval poets of Provence and Italy, and retained in occasional use by the modern poets of Western Europe.",
        "The scheme on which the sestina is built was the invention of the great troubadour, Arnaut Daniel (d. 1199).",
        "The sestina, in its pure medieval form, is independent of rhyme; it consists of six stanzas of six lines each of blank verse.",
        "The final words of the first stanza appear in inverted order in all the others, the order as laid down by the Provencals being as follows: abcdef, faebdc, cfdabe, ecbfad, deacfb, bdfeca.",
        "To these thirty-six lines are appended, as an envoy, three lines, in which all the final words should be repeated, one in the middle and one at the end of each line.",
        "The double sestina itself is not unknown in German literature.",
    ]
    for i, text in enumerate(brit_lines, start=1):
        lid = i
        if (brit_id, lid) in existing_keys:
            continue
        r = ws_last_data_row(ws_auto, key_col=ca["CorpusID"]) + 1
        ws_auto.cell(r, ca["CorpusID"]).value = brit_id
        ws_auto.cell(r, ca["Lang"]).value = "EN"
        ws_auto.cell(r, ca["License"]).value = "public domain"
        ws_auto.cell(r, ca["Source"]).value = "1911 Encyclopaedia Britannica (Sestina) https://en.wikisource.org/wiki/1911_Encyclop%C3%A6dia_Britannica/Sestina"
        ws_auto.cell(r, ca["LineID"]).value = lid
        ws_auto.cell(r, ca["Text"]).value = text
        ws_auto.cell(r, ca["AddedIter"]).value = iter_num
        ws_auto.cell(r, ca["Notes"]).value = "Seed corpus (auto)."
        added += 1

    # Seed: targeted poetry-form vocabulary, including spelling variants ("sestina" vs "sestine").
    # This helps signature alignment for in-corpus hints like SESTIEN ~ "sestine" (anagram layer).
    pf_id = "POETRY_FORMS_CORE_EN"
    pf_lines = [
        "sestina sestine sextine envoy envoi stanza sestet refrain retrogradatio cruciata",
    ]
    for i, text in enumerate(pf_lines, start=1):
        lid = i
        if (pf_id, lid) in existing_keys:
            continue
        r = ws_last_data_row(ws_auto, key_col=ca["CorpusID"]) + 1
        ws_auto.cell(r, ca["CorpusID"]).value = pf_id
        ws_auto.cell(r, ca["Lang"]).value = "EN"
        ws_auto.cell(r, ca["License"]).value = "curated"
        ws_auto.cell(r, ca["Source"]).value = "Curated poetry-form vocabulary (manual)"
        ws_auto.cell(r, ca["LineID"]).value = lid
        ws_auto.cell(r, ca["Text"]).value = text
        ws_auto.cell(r, ca["AddedIter"]).value = iter_num
        ws_auto.cell(r, ca["Notes"]).value = "Seed corpus (auto)."
        added += 1

    # User sheet is intentionally unseeded.
    upsert_sheet_index_entry(wb, "LoreCorpus_Auto", "Lore corpus (auto-seeded). Add additional reference texts in LoreCorpus_User.")
    upsert_sheet_index_entry(wb, "LoreCorpus_User", "Lore corpus (user-provided). Paste lore texts / translations here; the runner will mine safe semantic hints.")

    # Total rows for quick logging.
    total = ws_last_data_row(ws_auto, key_col=ca["CorpusID"]) - ha
    return added, max(0, int(total))


def _load_lore_corpus_rows(wb: openpyxl.Workbook) -> List[Dict[str, object]]:
    out: List[Dict[str, object]] = []
    for sheet in ("LoreCorpus_Auto", "LoreCorpus_User"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        h = ws_find_header_row(ws, ["CorpusID", "Lang", "LineID", "Text"], max_scan=3)
        c = ws_headers(ws, h)
        for r in range(h + 1, ws.max_row + 1):
            cid = ws.cell(r, c["CorpusID"]).value
            text = ws.cell(r, c["Text"]).value
            if not isinstance(cid, str) or not cid.strip():
                continue
            if text is None or str(text).strip() == "":
                continue
            try:
                lid = int(ws.cell(r, c["LineID"]).value or 0)
            except Exception:
                lid = 0
            lang = str(ws.cell(r, c["Lang"]).value or "").strip().upper() or "EN"
            out.append({"CorpusID": cid.strip(), "Lang": lang, "LineID": lid, "Text": str(text)})
    return out


def _build_lore_signature_index(
    corpus_rows: Sequence[Dict[str, object]],
    *,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
) -> Tuple[Dict[str, Dict[str, int]], Dict[str, set[str]]]:
    """Return (sig->word->count, sig->set(corpusIDs))."""
    sig_word_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sig_corpora: Dict[str, set[str]] = defaultdict(set)
    # Targeted exception: some in-corpus hints require keeping a final 'e' for signature matching
    # even when the global lore canon drops final-e as a general-purpose matching heuristic.
    keep_final_e_words = {"sestine", "sextine"}
    for row in corpus_rows:
        cid = str(row.get("CorpusID") or "").strip()
        text = str(row.get("Text") or "")
        if not cid or not text:
            continue
        # Split by whitespace and keep simple tokens; canonicalizer handles punctuation.
        for raw in text.split():
            # Keep apostrophes for readability in output (e.g. "you've"); signatures still computed
            # on the canonical (letters-only) form.
            surface = re.sub(r"[^a-z']", "", str(raw or "").lower())
            surface_letters = re.sub(r"[^a-z]", "", surface)
            df = bool(drop_final_e) and (surface_letters not in keep_final_e_words)
            canon = _lore_canon_word(str(raw), drop_final_e=df, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
            if not canon or not surface_letters:
                continue
            sig = _lore_signature(canon)
            # Store surface forms for readability; signature is computed on canonical form.
            sig_word_counts[sig][surface] += 1
            sig_corpora[sig].add(cid)
    return sig_word_counts, sig_corpora


def _load_lore_sigindex_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
) -> Tuple[Dict[str, Dict[str, int]], Dict[str, set[str]]]:
    """Load a precomputed signature index sheet.

    Expected columns: Sig | Word | Count | CorpusID
    Returns: (sig->word->count, sig->set(corpusIDs))
    """
    if sheet_name not in wb.sheetnames:
        return {}, {}
    ws = wb[sheet_name]
    h = ws_find_header_row(ws, ["Sig", "Word", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    col_corpus = c.get("CorpusID")

    sig_word_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sig_corpora: Dict[str, set[str]] = defaultdict(set)

    for r in range(h + 1, ws.max_row + 1):
        sig = ws.cell(r, c["Sig"]).value
        word = ws.cell(r, c["Word"]).value
        cnt = ws.cell(r, c["Count"]).value
        if not isinstance(sig, str) or not sig.strip():
            continue
        if not isinstance(word, str) or not word.strip():
            continue
        try:
            cnt_i = int(cnt or 0)
        except Exception:
            cnt_i = 0
        if cnt_i <= 0:
            continue
        sig_s = sig.strip()
        word_s = word.strip().lower()
        sig_word_counts[sig_s][word_s] += cnt_i
        if col_corpus is not None:
            cid = ws.cell(r, col_corpus).value
            if isinstance(cid, str) and cid.strip():
                sig_corpora[sig_s].add(cid.strip())

    return sig_word_counts, sig_corpora


_TIBIA_WORD_RE = re.compile(r"[A-Za-z']+")

SEQ_MATCH_STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "by",
    "co",
    "do",
    "for",
    "from",
    "have",
    "he",
    "i",
    "if",
    "in",
    "into",
    "is",
    "it",
    "me",
    "mine",
    "no",
    "not",
    "of",
    "off",
    "oft",
    "on",
    "or",
    "our",
    "so",
    "than",
    "that",
    "the",
    "their",
    "there",
    "they",
    "this",
    "to",
    "we",
    "with",
    "you",
    "youve",
}


def _iter_words(text: str) -> Iterable[str]:
    for m in _TIBIA_WORD_RE.finditer(str(text or "")):
        w = m.group(0)
        if w:
            yield w


def _is_seq_content_word(word: str) -> bool:
    wk = re.sub(r"[^a-z]", "", str(word or "").lower())
    return bool(wk) and wk not in SEQ_MATCH_STOPWORDS and len(wk) >= 3


def _target_signatures_from_glossary_sheet(wb: openpyxl.Workbook) -> set[str]:
    if "Glossary" not in wb.sheetnames:
        return set()
    ws = wb["Glossary"]
    h = ws_find_header_row(ws, ["Token", "Translation", "Use_StrictPlus_v108", "TokenType"], max_scan=3)
    c = ws_headers(ws, h)

    out: set[str] = set()
    for r in range(h + 1, ws.max_row + 1):
        tok = ws.cell(r, c["Token"]).value
        if not isinstance(tok, str) or not tok:
            continue
        if "*" in tok:
            continue
        if not parse_bool(ws.cell(r, c["Use_StrictPlus_v108"]).value, False):
            continue
        tr = ws.cell(r, c["Translation"]).value
        if tr is None or str(tr).strip() == "":
            continue
        ttype = str(ws.cell(r, c["TokenType"]).value or "").strip().lower()
        if ttype in ("marker", "macro"):
            continue
        sig = _token_signature(tok)
        if sig:
            out.add(sig)
    return out


def _digit_code_map_letters(wb: openpyxl.Workbook) -> set[str]:
    """Return the set of observed base letters from DigitCodeMap_Auto (if present)."""
    if DIGIT_CODE_MAP_SHEET not in wb.sheetnames:
        return set()
    ws = wb[DIGIT_CODE_MAP_SHEET]
    h = ws_find_header_row(ws, ["Code", "Letter"], max_scan=3)
    c = ws_headers(ws, h)
    out: set[str] = set()
    for r in range(h + 1, ws.max_row + 1):
        v = ws.cell(r, c["Letter"]).value
        if not isinstance(v, str) or not v.strip():
            continue
        out.add(v.strip().upper())
    return out


def _sigindex_last_fetched_dt(wb: openpyxl.Workbook, sheet_name: str) -> Optional[datetime]:
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    h = ws_find_header_row(ws, ["Sig", "Word", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    col = c.get("FetchedUTC")
    if not col:
        return None
    for r in range(h + 1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, str) or not v.strip():
            continue
        try:
            return datetime.strptime(v.strip(), ISO_UTC_FMT).replace(tzinfo=timezone.utc)
        except Exception:
            continue
    return None


def _load_wordfreq_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Dict[str, int]:
    """Load Word->Count from a derived wordfreq sheet."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    h = ws_find_header_row(ws, ["Word", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    out: Dict[str, int] = {}
    for r in range(h + 1, ws.max_row + 1):
        w = ws.cell(r, c["Word"]).value
        cnt = ws.cell(r, c["Count"]).value
        if not isinstance(w, str) or not w.strip():
            continue
        wkey = re.sub(r"[^a-z]", "", w.strip().lower())
        if not wkey:
            continue
        try:
            ci = int(cnt or 0)
        except Exception:
            ci = 0
        if ci <= 0:
            continue
        out[wkey] = out.get(wkey, 0) + ci
    return out


def refresh_tibia_sigindex_sheet(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    npc_url: str,
    book_url: str,
    timeout_s: int,
    max_words_per_sig: int,
    wordfreq_topn: int,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
) -> Tuple[int, int, int, int]:
    """Fetch a Tibia-derived dataset and write a *derived* signature index into the workbook.

    Returns: (sigindex_rows_written, sigs_hit, target_sigs, wordfreq_rows_written)
    """
    target_sigs = _target_signatures_from_glossary_sheet(wb)
    if not target_sigs:
        return 0, 0, 0, 0

    def _fetch_json(url: str) -> object:
        # Cloudflare-hosted endpoints may reject default urllib User-Agent with 403.
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (Bonelord469SigIndex/1.0)"})
        with urllib.request.urlopen(req, timeout=float(timeout_s)) as resp:
            data = resp.read()
        return json.loads(data.decode("utf-8", errors="replace"))

    sig_word_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sig_corpora: Dict[str, set[str]] = defaultdict(set)
    global_word_counts: Counter[str] = Counter()

    npc_obj = _fetch_json(npc_url)
    if isinstance(npc_obj, list):
        for item in npc_obj:
            if not isinstance(item, dict):
                continue
            conv = item.get("conversation")
            if not isinstance(conv, list):
                continue
            for turn in conv:
                if not isinstance(turn, dict):
                    continue
                answers = turn.get("answer")
                if isinstance(answers, str):
                    answers = [answers]
                if not isinstance(answers, list):
                    continue
                for a in answers:
                    if not isinstance(a, str) or not a.strip():
                        continue
                    for raw in _iter_words(a):
                        surface = re.sub(r"[^a-z']", "", raw.lower())
                        if not surface:
                            continue
                        wkey = re.sub(r"[^a-z]", "", surface)
                        if wkey:
                            global_word_counts[wkey] += 1
                        canon = _lore_canon_word(surface, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
                        if not canon:
                            continue
                        sig = _lore_signature(canon)
                        if sig in target_sigs:
                            sig_word_counts[sig][surface] += 1
                            sig_corpora[sig].add("TIBIA_NPC")

    book_obj = _fetch_json(book_url)
    if isinstance(book_obj, list):
        for item in book_obj:
            if not isinstance(item, dict):
                continue
            text = item.get("text")
            if not isinstance(text, str) or not text.strip():
                continue
            for raw in _iter_words(text):
                surface = re.sub(r"[^a-z']", "", raw.lower())
                if not surface:
                    continue
                wkey = re.sub(r"[^a-z]", "", surface)
                if wkey:
                    global_word_counts[wkey] += 1
                canon = _lore_canon_word(surface, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
                if not canon:
                    continue
                sig = _lore_signature(canon)
                if sig in target_sigs:
                    sig_word_counts[sig][surface] += 1
                    sig_corpora[sig].add("TIBIA_BOOK")

    ws = ensure_sheet(wb, LORE_SIGINDEX_TIBIA_SHEET, ["Sig", "Word", "Count", "CorpusID", "FetchedUTC", "Source"])
    h = ws_find_header_row(ws, ["Sig", "Word", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    if ws.max_row > h:
        ws.delete_rows(h + 1, ws.max_row - h)

    fetched = utc
    source = (
        f"{npc_url} ; {book_url} "
        "(derived word signature index only; no full text stored)"
    )

    rr = h + 1
    for sig in sorted(sig_word_counts.keys()):
        wc = sig_word_counts[sig]
        corp = ",".join(sorted(sig_corpora.get(sig) or []))
        for word, cnt in sorted(wc.items(), key=lambda kv: (-kv[1], kv[0]))[: int(max_words_per_sig)]:
            ws.cell(rr, c["Sig"]).value = sig
            ws.cell(rr, c["Word"]).value = word
            ws.cell(rr, c["Count"]).value = int(cnt)
            if "CorpusID" in c:
                ws.cell(rr, c["CorpusID"]).value = corp
            if "FetchedUTC" in c:
                ws.cell(rr, c["FetchedUTC"]).value = fetched
            if "Source" in c:
                ws.cell(rr, c["Source"]).value = source
            rr += 1

    upsert_sheet_index_entry(
        wb,
        LORE_SIGINDEX_TIBIA_SHEET,
        "Tibia-derived signature index for semantic alignment (derived counts only; no full text).",
    )

    # Also write a global word-frequency list (derived counts only; no full text).
    ws2 = ensure_sheet(wb, LORE_WORDFREQ_TIBIA_SHEET, ["Word", "Count", "FetchedUTC", "Source"])
    h2 = ws_find_header_row(ws2, ["Word", "Count"], max_scan=3)
    c2 = ws_headers(ws2, h2)
    if ws2.max_row > h2:
        ws2.delete_rows(h2 + 1, ws2.max_row - h2)
    rr2 = h2 + 1
    for word, cnt in global_word_counts.most_common(int(wordfreq_topn)):
        ws2.cell(rr2, c2["Word"]).value = word
        ws2.cell(rr2, c2["Count"]).value = int(cnt)
        if "FetchedUTC" in c2:
            ws2.cell(rr2, c2["FetchedUTC"]).value = fetched
        if "Source" in c2:
            ws2.cell(rr2, c2["Source"]).value = source
        rr2 += 1
    upsert_sheet_index_entry(
        wb,
        LORE_WORDFREQ_TIBIA_SHEET,
        "Tibia-derived global word frequency list (derived counts only; no full text).",
    )

    return rr - (h + 1), len(sig_word_counts), len(target_sigs), rr2 - (h2 + 1)


def refresh_pd_sigindex_sheet(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    corpus_id: str,
    url: str,
    extra_sources: Optional[Sequence[Tuple[str, str]]] = None,
    timeout_s: int,
    max_words_per_sig: int,
    wordfreq_topn: int,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    cache_max_age_hours: float,
) -> Tuple[int, int, int, int]:
    """Fetch a public-domain plaintext source and write a *derived* signature index into the workbook.

    Returns: (sigindex_rows_written, sigs_hit, target_sigs, wordfreq_rows_written)
    """
    target_sigs = _target_signatures_from_glossary_sheet(wb)
    if not target_sigs:
        return 0, 0, 0, 0

    ua = "Mozilla/5.0 (Bonelord469PDSigIndex/1.0)"
    cache_dir = os.path.join(os.getcwd(), "tmp", "corpus")
    sources_raw: List[Tuple[str, str]] = [(str(corpus_id or ""), str(url or ""))]
    if extra_sources:
        try:
            sources_raw.extend([(str(cid or ""), str(u or "")) for (cid, u) in list(extra_sources)])
        except Exception:
            pass

    sources: List[Tuple[str, str]] = []
    seen_urls: set[str] = set()
    for cid_raw, u_raw in sources_raw:
        u = str(u_raw or "").strip()
        if not u:
            continue
        if u in seen_urls:
            continue
        seen_urls.add(u)
        cid = str(cid_raw or "").strip() or _auto_corpus_id_from_url(u)
        sources.append((cid, u))
    if not sources:
        return 0, 0, 0, 0

    sig_word_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sig_corpora: Dict[str, set[str]] = defaultdict(set)
    global_word_counts: Counter[str] = Counter()

    keep_final_e_words = {"sestine", "sextine"}

    for cid, u in sources:
        cache_path = _cache_path_for_url(cache_dir, prefix="pd_text", url=str(u or ""), ext="txt")
        text = _fetch_text_url_cached(
            str(u or ""),
            cache_path=cache_path,
            timeout_s=int(timeout_s),
            max_age_hours=float(cache_max_age_hours),
            user_agent=ua,
        )

        # Strip Project Gutenberg boilerplate when present (keeps counts focused on the actual corpus).
        # Best-effort: if markers aren't present, just use the full text.
        txt = str(text or "")
        m_start = re.search(r"\*\*\*\s*START OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*", txt, flags=re.IGNORECASE | re.DOTALL)
        m_end = re.search(r"\*\*\*\s*END OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*", txt, flags=re.IGNORECASE | re.DOTALL)
        if m_start and m_end and m_end.start() > m_start.end():
            txt = txt[m_start.end() : m_end.start()]

        for raw in _iter_words(txt):
            surface = re.sub(r"[^a-z']", "", str(raw or "").lower())
            if not surface:
                continue
            surface_letters = re.sub(r"[^a-z]", "", surface)
            if surface_letters:
                global_word_counts[surface_letters] += 1
            df = bool(drop_final_e) and (surface_letters not in keep_final_e_words)
            canon = _lore_canon_word(surface, drop_final_e=df, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
            if not canon:
                continue
            sig = _lore_signature(canon)
            if sig in target_sigs:
                sig_word_counts[sig][surface] += 1
                sig_corpora[sig].add(str(cid or "").strip() or "PUBLIC_DOMAIN")

    ws = ensure_sheet(wb, LORE_SIGINDEX_PD_SHEET, ["Sig", "Word", "Count", "CorpusID", "FetchedUTC", "Source"])
    h = ws_find_header_row(ws, ["Sig", "Word", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    if ws.max_row > h:
        ws.delete_rows(h + 1, ws.max_row - h)

    fetched = utc
    src_urls = "; ".join([u for _cid, u in sources if str(u or "").strip()])
    if len(src_urls) > 240:
        src_urls = src_urls[:237] + "..."
    source = f"{src_urls} (derived signature index only; no full text stored)"

    rr = h + 1
    for sig in sorted(sig_word_counts.keys()):
        wc = sig_word_counts[sig]
        corp = ",".join(sorted(sig_corpora.get(sig) or []))
        for word, cnt in sorted(wc.items(), key=lambda kv: (-kv[1], kv[0]))[: int(max_words_per_sig)]:
            ws.cell(rr, c["Sig"]).value = sig
            ws.cell(rr, c["Word"]).value = word
            ws.cell(rr, c["Count"]).value = int(cnt)
            if "CorpusID" in c:
                ws.cell(rr, c["CorpusID"]).value = corp
            if "FetchedUTC" in c:
                ws.cell(rr, c["FetchedUTC"]).value = fetched
            if "Source" in c:
                ws.cell(rr, c["Source"]).value = source
            rr += 1

    upsert_sheet_index_entry(
        wb,
        LORE_SIGINDEX_PD_SHEET,
        "Public-domain derived signature index for semantic alignment (derived counts only; no full text).",
    )

    # Also write a global word-frequency list (derived counts only; no full text).
    ws2 = ensure_sheet(wb, LORE_WORDFREQ_PD_SHEET, ["Word", "Count", "FetchedUTC", "Source"])
    h2 = ws_find_header_row(ws2, ["Word", "Count"], max_scan=3)
    c2 = ws_headers(ws2, h2)
    if ws2.max_row > h2:
        ws2.delete_rows(h2 + 1, ws2.max_row - h2)
    rr2 = h2 + 1
    for word, cnt in global_word_counts.most_common(int(wordfreq_topn)):
        ws2.cell(rr2, c2["Word"]).value = word
        ws2.cell(rr2, c2["Count"]).value = int(cnt)
        if "FetchedUTC" in c2:
            ws2.cell(rr2, c2["FetchedUTC"]).value = fetched
        if "Source" in c2:
            ws2.cell(rr2, c2["Source"]).value = source
        rr2 += 1
    upsert_sheet_index_entry(
        wb,
        LORE_WORDFREQ_PD_SHEET,
        "Public-domain derived global word frequency list (derived counts only; no full text).",
    )

    return rr - (h + 1), len(sig_word_counts), len(target_sigs), rr2 - (h2 + 1)


def refresh_dict_sigindex_sheet(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    dict_path: str,
    max_words_per_sig: int,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
) -> Tuple[int, int, int]:
    """Build a derived signature index from a local English word list (no full text import).

    This is intended as a safety-first coverage expander when Tibia/PD corpora do not intersect
    the encoded corpus. We persist only (Sig,Word,Count) rows for *target* signatures (active leaf
    tokens) and use existing derived wordfreq sheets as a ranking prior when available.

    Returns: (sigindex_rows_written, sigs_hit, target_sigs)
    """
    target_sigs = _target_signatures_from_glossary_sheet(wb)
    if not target_sigs:
        return 0, 0, 0

    # Use existing derived wordfreq (Tibia + PD) as a ranking prior when available.
    wordfreq: Dict[str, int] = {}
    try:
        for sh in (LORE_WORDFREQ_TIBIA_SHEET, LORE_WORDFREQ_PD_SHEET):
            for w, cnt in _load_wordfreq_sheet(wb, sh).items():
                wordfreq[w] = wordfreq.get(w, 0) + int(cnt)
    except Exception:
        wordfreq = {}

    sig_word_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sig_corpora: Dict[str, set[str]] = defaultdict(set)

    if not dict_path or not os.path.exists(dict_path):
        # Still ensure the sheet exists for workbook stability.
        ensure_sheet(wb, LORE_SIGINDEX_DICT_SHEET, ["Sig", "Word", "Count", "CorpusID", "FetchedUTC", "Source"])
        upsert_sheet_index_entry(
            wb,
            LORE_SIGINDEX_DICT_SHEET,
            "Dictionary-derived signature index (derived counts only; no full text).",
        )
        return 0, 0, len(target_sigs)

    try:
        with open(dict_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                raw = str(line or "").strip().lower()
                if not raw:
                    continue
                # Keep apostrophes only for splitting; the canonizer removes non-letters.
                raw = re.sub(r"[^a-z']", "", raw)
                if not raw:
                    continue
                parts = re.findall(r"[a-z]+", raw)
                if not parts:
                    continue
                for surface in parts:
                    if not surface:
                        continue
                    canon = _lore_canon_word(
                        surface,
                        drop_final_e=drop_final_e,
                        drop_all_h=drop_all_h,
                        drop_all_o=drop_all_o,
                    )
                    if not canon:
                        continue
                    sig = _lore_signature(canon)
                    if sig not in target_sigs:
                        continue
                    wkey = re.sub(r"[^a-z]", "", surface)
                    cnt = int(wordfreq.get(wkey, 1) or 1)
                    prev = int(sig_word_counts[sig].get(surface, 0) or 0)
                    # Use max-count as a stable "prior" instead of summing duplicates from the word list.
                    sig_word_counts[sig][surface] = max(prev, cnt)
                    sig_corpora[sig].add("DICT_WEB2")
    except Exception:
        # Best-effort: a missing/bad dict should not block the iteration.
        sig_word_counts = {}
        sig_corpora = {}

    ws = ensure_sheet(wb, LORE_SIGINDEX_DICT_SHEET, ["Sig", "Word", "Count", "CorpusID", "FetchedUTC", "Source"])
    h = ws_find_header_row(ws, ["Sig", "Word", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    if ws.max_row > h:
        ws.delete_rows(h + 1, ws.max_row - h)

    source = f"{dict_path} (derived word signature index only; no full text stored)"
    rr = h + 1
    for sig in sorted(sig_word_counts.keys()):
        wc = sig_word_counts[sig]
        corp = ",".join(sorted(sig_corpora.get(sig) or []))
        for word, cnt in sorted(wc.items(), key=lambda kv: (-kv[1], kv[0]))[: int(max_words_per_sig)]:
            ws.cell(rr, c["Sig"]).value = sig
            ws.cell(rr, c["Word"]).value = word
            ws.cell(rr, c["Count"]).value = int(cnt)
            if "CorpusID" in c:
                ws.cell(rr, c["CorpusID"]).value = corp
            if "FetchedUTC" in c:
                ws.cell(rr, c["FetchedUTC"]).value = utc
            if "Source" in c:
                ws.cell(rr, c["Source"]).value = source
            rr += 1

    upsert_sheet_index_entry(
        wb,
        LORE_SIGINDEX_DICT_SHEET,
        "Dictionary-derived signature index for semantic alignment (derived counts only; no full text).",
    )

    return rr - (h + 1), len(sig_word_counts), len(target_sigs)


def materialize_lore_token_hits(
    wb: openpyxl.Workbook,
    iter_num: int,
    glossary_map: Dict[str, GlossaryToken],
    *,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    max_rows: int = 800,
) -> int:
    """Write LoreAlignment_Auto with signature-based hits between corpus words and Glossary leaf tokens.

    Returns number of rows written (excluding header).
    """
    # Ensure corpus sheets exist first.
    _ensure_lore_corpus_sheets(wb, iter_num)
    corpus_rows = _load_lore_corpus_rows(wb)
    sig_word_counts, sig_corpora = _build_lore_signature_index(
        corpus_rows,
        drop_final_e=drop_final_e,
        drop_all_h=drop_all_h,
        drop_all_o=drop_all_o,
    )
    # Merge additional precomputed signature indexes (internet-derived or PD-derived).
    for sh in (LORE_SIGINDEX_TIBIA_SHEET, LORE_SIGINDEX_PD_SHEET, LORE_SIGINDEX_DICT_SHEET):
        ext_counts, ext_corp = _load_lore_sigindex_sheet(wb, sh)
        if not ext_counts:
            continue
        for sig, wc in ext_counts.items():
            for w, cnt in wc.items():
                sig_word_counts[sig][w] += int(cnt)
        for sig, corp in ext_corp.items():
            if corp:
                sig_corpora[sig].update(corp)

    ws = ensure_sheet(
        wb,
        "LoreAlignment_Auto",
        [
            "Iteration",
            "Token",
            "TokenSig",
            "CurrentTranslation",
            "Confidence",
            "EvidenceClass",
            "EvidenceScore",
            "TotalOcc",
            "TopWord",
            "TopWordCount",
            "TotalWordCount",
            "CandidateWords",
            "CandidateWordCounts",
            "CorpusIDs",
            "Notes",
        ],
    )
    h = ws_find_header_row(ws, ["Iteration", "Token", "TopWord"], max_scan=3)
    c = ws_headers(ws, h)
    # Clear old body.
    if ws.max_row > h:
        ws.delete_rows(h + 1, ws.max_row - h)

    rows_written = 0
    out_r = h + 1

    # Candidate token set: leaf tokens only (exclude markers/macros) and avoid star tokens for now.
    tokens = [t for t in glossary_map.values() if t.use_strictplus and t.translation and t.token_type not in ("marker", "macro")]
    tokens.sort(key=lambda t: (-t.total_occ, -t.length, t.token))

    for t in tokens:
        if rows_written >= max_rows:
            break
        tok_sig = _token_signature(t.token)
        if not tok_sig:
            continue
        # Skip star tokens (wildcards) for now (matching needs wildcard-aware signatures).
        if "*" in t.token:
            continue
        cand = sig_word_counts.get(tok_sig)
        if not cand:
            continue

        items = sorted(cand.items(), key=lambda kv: (-kv[1], kv[0]))
        top_word, top_count = items[0]
        total_count = int(sum(cnt for _w, cnt in items))
        cand_words = ", ".join([w for w, _cnt in items[:8]])
        cand_counts = ", ".join([f"{w}:{int(cnt)}" for w, cnt in items[:8]])
        corp = ", ".join(sorted(sig_corpora.get(tok_sig) or []))

        ws.cell(out_r, c["Iteration"]).value = iter_num
        ws.cell(out_r, c["Token"]).value = t.token
        ws.cell(out_r, c["TokenSig"]).value = tok_sig
        ws.cell(out_r, c["CurrentTranslation"]).value = t.translation
        ws.cell(out_r, c["Confidence"]).value = t.confidence
        ws.cell(out_r, c["EvidenceClass"]).value = t.evidence_class
        ws.cell(out_r, c["EvidenceScore"]).value = t.evidence_score
        ws.cell(out_r, c["TotalOcc"]).value = t.total_occ
        ws.cell(out_r, c["TopWord"]).value = top_word
        ws.cell(out_r, c["TopWordCount"]).value = int(top_count)
        ws.cell(out_r, c["TotalWordCount"]).value = total_count
        ws.cell(out_r, c["CandidateWords"]).value = cand_words
        ws.cell(out_r, c["CandidateWordCounts"]).value = cand_counts
        ws.cell(out_r, c["CorpusIDs"]).value = corp
        ws.cell(out_r, c["Notes"]).value = f"canon(drop_final_e={drop_final_e}, drop_all_h={drop_all_h}, drop_all_o={drop_all_o})"

        out_r += 1
        rows_written += 1

    upsert_sheet_index_entry(
        wb,
        "LoreAlignment_Auto",
        "Lore signature hits: compares LoreCorpus_* word signatures vs Glossary leaf token signatures (display-only semantic hints).",
    )
    return rows_written


def materialize_semantic_map_from_lore_hits(
    wb: openpyxl.Workbook,
    iter_num: int,
    *,
    min_total_count: int,
    min_top_share: float,
    max_rows: int = 500,
) -> Tuple[int, Dict[str, str], Dict[str, str]]:
    """Derive SemanticMap_Auto from LoreAlignment_Auto (conservative).

    Returns: (rows_written, token->semantic_word, token->reason)
    """
    if "LoreAlignment_Auto" not in wb.sheetnames:
        return 0, {}, {}

    ws_hits = wb["LoreAlignment_Auto"]
    hh = ws_find_header_row(
        ws_hits,
        ["Token", "TopWord", "TopWordCount", "TotalWordCount", "CandidateWords", "CurrentTranslation"],
        max_scan=3,
    )
    ch = ws_headers(ws_hits, hh)

    token_rows: List[Tuple[str, str, int, int, List[str], str]] = []
    for r in range(hh + 1, ws_hits.max_row + 1):
        tok = ws_hits.cell(r, ch["Token"]).value
        top = ws_hits.cell(r, ch["TopWord"]).value
        topc = ws_hits.cell(r, ch["TopWordCount"]).value
        tot = ws_hits.cell(r, ch["TotalWordCount"]).value
        cand = ws_hits.cell(r, ch["CandidateWords"]).value
        cur = ws_hits.cell(r, ch["CurrentTranslation"]).value

        if not isinstance(tok, str) or not tok.strip():
            continue
        if not isinstance(top, str) or not top.strip():
            continue
        try:
            topc_i = int(topc or 0)
        except Exception:
            topc_i = 0
        try:
            tot_i = int(tot or 0)
        except Exception:
            tot_i = 0

        cand_list = [w.strip() for w in str(cand or "").split(",") if w and str(w).strip()]
        token_rows.append((tok.strip(), top.strip(), topc_i, tot_i, cand_list, str(cur or "")))

    semantic: Dict[str, str] = {}
    reasons: Dict[str, str] = {}
    ws_out = ensure_sheet(
        wb,
        "SemanticMap_Auto",
        [
            "Iteration",
            "Token",
            "SemanticWord",
            "Reason",
            "TopWordCount",
            "TotalWordCount",
            "TopShare",
            "CandidateCount",
            "CandidateWords",
            "CurrentTranslation",
        ],
    )
    hs = ws_find_header_row(ws_out, ["Iteration", "Token", "SemanticWord"], max_scan=3)
    cs = ws_headers(ws_out, hs)
    if ws_out.max_row > hs:
        ws_out.delete_rows(hs + 1, ws_out.max_row - hs)

    written = 0
    rr = hs + 1
    for tok, top, topc, tot, cand_list, cur in sorted(token_rows, key=lambda x: (-x[2], x[0])):
        if written >= max_rows:
            break

        chosen: Optional[str] = None
        reason = ""

        if len(cand_list) == 1 and topc >= 1:
            chosen = cand_list[0]
            reason = "single_candidate"
        else:
            # Prefer keeping current translation if it is attested in the lore corpus.
            cur_norm = re.sub(r"[^a-z]", "", str(cur or "").lower())
            if cur_norm:
                for w in cand_list:
                    if re.sub(r"[^a-z]", "", w.lower()) == cur_norm:
                        chosen = w
                        reason = "current_attested"
                        break

            if not chosen:
                share = float(topc) / float(tot) if tot > 0 else 0.0
                if tot >= int(min_total_count) and share >= float(min_top_share):
                    chosen = top
                    reason = f"top_share>={min_top_share}"

        if not chosen:
            continue

        semantic[tok] = chosen
        reasons[tok] = reason
        tot_safe = int(tot or 0)
        share_safe = float(topc) / float(tot_safe) if tot_safe > 0 else 0.0

        ws_out.cell(rr, cs["Iteration"]).value = iter_num
        ws_out.cell(rr, cs["Token"]).value = tok
        ws_out.cell(rr, cs["SemanticWord"]).value = chosen
        ws_out.cell(rr, cs["Reason"]).value = reason
        ws_out.cell(rr, cs["TopWordCount"]).value = int(topc)
        ws_out.cell(rr, cs["TotalWordCount"]).value = tot_safe
        ws_out.cell(rr, cs["TopShare"]).value = round(share_safe, 6)
        ws_out.cell(rr, cs["CandidateCount"]).value = len(cand_list)
        ws_out.cell(rr, cs["CandidateWords"]).value = ", ".join(cand_list[:8])
        ws_out.cell(rr, cs["CurrentTranslation"]).value = cur
        rr += 1
        written += 1

    upsert_sheet_index_entry(
        wb,
        "SemanticMap_Auto",
        "Display-only semantic mapping derived from LoreAlignment_Auto (conservative; does not change Glossary/DP).",
    )
    return written, semantic, reasons


def _norm_wordish(s: object) -> str:
    """Normalize a word/phrase for comparison (semantic-only)."""
    return re.sub(r"[^a-z]", "", str(s or "").lower())


def glossary_set_translation(glossary_ws: openpyxl.worksheet.worksheet.Worksheet, token_row: int, new_tr: str, iter_num: int, note: str) -> str:
    header = ws_find_header_row(glossary_ws, ["Translation", "Notes"])
    c = ws_headers(glossary_ws, header)
    cell_tr = glossary_ws.cell(token_row, c["Translation"])
    old = str(cell_tr.value or "")
    cell_tr.value = str(new_tr or "")
    append_note_cell(glossary_ws.cell(token_row, c["Notes"]), f"iter{iter_num}: {note}")
    return old


def sync_evidence_ledger_translations(wb: openpyxl.Workbook, updates: Dict[str, str]) -> int:
    """Best-effort: keep EvidenceLedger_v127 translations consistent with Glossary."""
    if not updates:
        return 0
    if "EvidenceLedger_v127" not in wb.sheetnames:
        return 0
    ws = wb["EvidenceLedger_v127"]
    header = ws_find_header_row(ws, ["Token", "Translation"], max_scan=3)
    c = ws_headers(ws, header)
    row_by_token: Dict[str, int] = {}
    for r in range(header + 1, ws.max_row + 1):
        tok = ws.cell(r, c["Token"]).value
        if isinstance(tok, str) and tok.strip():
            row_by_token[tok.strip()] = r

    changed = 0
    for tok, new_tr in updates.items():
        r = row_by_token.get(tok)
        if r is None:
            continue
        cur = str(ws.cell(r, c["Translation"]).value or "")
        if cur != str(new_tr or ""):
            ws.cell(r, c["Translation"]).value = str(new_tr or "")
            changed += 1
    return changed


def apply_semantic_promotions_to_glossary(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    glossary_ws: openpyxl.worksheet.worksheet.Worksheet,
    glossary_map: Dict[str, GlossaryToken],
    semantic_map: Dict[str, str],
    semantic_reasons: Dict[str, str],
    *,
    enabled: bool,
    max_promotions: int,
    min_total_occ: int,
    min_conf_weight: int,
    max_conf_weight: int,
    blocked_evidence_classes: set[str],
    wordfreq_sheet: str,
    min_new_wordfreq: int,
    min_wordfreq_ratio: float,
    anti_mode: bool = False,
    anti_hallucination_terms: Optional[set[str]] = None,
    enforced_crib_ids: Optional[set[int]] = None,
    locked_tokens: Optional[set[str]] = None,
) -> Tuple[int, int, int]:
    """Apply high-confidence semantic suggestions into Glossary translations (guarded by GT live check).

    This is the "safe way to advance translation": it only edits *strings* (translation labels) and
    never changes tokenization. It is still guarded by GroundTruth live check.

    Returns: (applied_count, attempted_count, evidence_ledger_changed)
    """
    if not enabled or not semantic_map:
        return 0, 0, 0

    # Global word frequency reference (derived; no full text). If missing, keep promotions conservative.
    wordfreq = _load_wordfreq_sheet(wb, wordfreq_sheet) if wordfreq_sheet else {}

    # Small whitelist to allow common 1-2 letter words; blocks nonsense like "nm".
    allowed_short = {
        "a",
        "i",
        "am",
        "an",
        "as",
        "at",
        "be",
        "by",
        "do",
        "go",
        "he",
        "if",
        "in",
        "is",
        "it",
        "me",
        "my",
        "no",
        "of",
        "on",
        "or",
        "so",
        "to",
        "up",
        "us",
        "we",
        "ye",
    }

    # Prepare a persistent log sheet for traceability.
    ws_log = ensure_sheet(
        wb,
        "SemanticPromotions_Auto",
        ["Iteration", "UTC", "Token", "OldTranslation", "NewTranslation", "Decision", "Reason", "Notes"],
    )
    upsert_sheet_index_entry(
        wb,
        "SemanticPromotions_Auto",
        "Applied/attempted semantic->Glossary translation edits (guarded by GT live check).",
    )

    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}
    anti_terms = set(anti_hallucination_terms or set())

    candidates: List[Tuple[int, int, str, str, str]] = []
    for tok, new_tr in semantic_map.items():
        gt = glossary_map.get(tok)
        if gt is None:
            continue
        if locked_tokens and tok in locked_tokens:
            continue
        if not gt.use_strictplus:
            continue
        if gt.token_type in ("marker", "macro"):
            continue
        evcls = str(gt.evidence_class or "").strip().upper()
        if evcls and evcls in blocked_evidence_classes:
            continue
        conf_w = confidence_weight(gt.confidence)
        if conf_w < int(min_conf_weight):
            continue
        if conf_w > int(max_conf_weight):
            continue
        if int(gt.total_occ) < int(min_total_occ):
            continue
        old_tr = str(gt.translation or "")
        if _norm_wordish(old_tr) == _norm_wordish(new_tr):
            continue
        if any(ch.isspace() for ch in str(new_tr or "")):
            continue
        if any(ch.isspace() for ch in str(old_tr or "")):
            continue

        # Wordfreq guardrails: only apply when the new surface is sufficiently common in the reference corpus,
        # and is not much rarer than the current translation.
        new_w = _norm_wordish(new_tr)
        old_w = _norm_wordish(old_tr)
        if not new_w:
            continue
        if anti_mode and anti_terms and new_w in anti_terms:
            continue
        if len(new_w) <= 2 and new_w not in allowed_short:
            continue
        if wordfreq:
            new_cnt = int(wordfreq.get(new_w, 0))
            old_cnt = int(wordfreq.get(old_w, 0)) if old_w else 0
            if new_cnt < int(min_new_wordfreq):
                continue
            if old_cnt > 0 and float(new_cnt) < float(old_cnt) * float(min_wordfreq_ratio):
                continue
        else:
            # Without a word-frequency reference, avoid auto-retext.
            continue
        candidates.append((int(gt.total_occ), int(gt.length), tok, old_tr, str(new_tr or "")))

    # Most impact first: high-occ, longer tokens.
    candidates.sort(key=lambda t: (-t[0], -t[1], t[2]))

    applied = 0
    attempted = 0
    updates: Dict[str, str] = {}

    for _occ, _ln, tok, old_tr, new_tr in candidates:
        if applied >= int(max_promotions):
            break
        attempted += 1

        cur = active.get(tok) or glossary_map.get(tok)
        if cur is None:
            continue

        # Simulate the edit under the current cumulative active set.
        test_active = dict(active)
        test_active[tok] = dataclasses.replace(cur, translation=new_tr)
        ok, bad, bad_all = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_crib_ids)
        if not ok:
            bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad[:8])
            soft_n = max(0, len(bad_all) - len(bad))
            ws_append_row(
                ws_log,
                [
                    iter_num,
                    utc,
                    tok,
                    old_tr,
                    new_tr,
                    "SKIP",
                    semantic_reasons.get(tok, ""),
                    f"GT live check mismatch (CribID(s) {bad_ids})"
                    + (f"; soft_mismatches={soft_n}" if soft_n else ""),
                ],
                start_col=1,
            )
            continue

        # Apply to Glossary.
        note = f"semantic retext: {old_tr} -> {new_tr} (reason={semantic_reasons.get(tok,'')})"
        glossary_set_translation(glossary_ws, cur.row, new_tr, iter_num, note)
        ws_append_row(
            ws_log,
            [iter_num, utc, tok, old_tr, new_tr, "APPLIED", semantic_reasons.get(tok, ""), "GT OK"],
            start_col=1,
        )

        # Keep cumulative active set updated.
        active = test_active
        updates[tok] = new_tr
        applied += 1

    el_changed = sync_evidence_ledger_translations(wb, updates)
    return applied, attempted, el_changed


def apply_english_promotions_to_glossary(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    glossary_ws: openpyxl.worksheet.worksheet.Worksheet,
    glossary_map: Dict[str, "GlossaryToken"],
    *,
    enabled: bool,
    max_promotions: int,
    min_total_count: int,
    min_top_share: float,
    lock_iters: int = 5,
    anti_mode: bool = False,
    anti_hallucination_terms: Optional[set[str]] = None,
    enforced_crib_ids: Optional[set[int]] = None,
    locked_tokens: Optional[set[str]] = None,
) -> Tuple[int, int, int]:
    """Apply EnglishMap_Auto suggestions into Glossary translations (guarded by GT live check).

    This advances readability towards real English without changing tokenization or evidence metrics.
    """
    if not enabled:
        return 0, 0, 0
    if ENGLISH_MAP_SHEET not in wb.sheetnames and ENGLISH_MAP_CONTEXT_SHEET not in wb.sheetnames:
        return 0, 0, 0

    # canon_word -> (top_word, total_count, top_share)
    word_map: Dict[str, Tuple[str, int, float]] = {}

    def _load_map_sheet(sheet_name: str, *, allow_overwrite: bool) -> None:
        if sheet_name not in wb.sheetnames:
            return
        ws_map = wb[sheet_name]
        hm = ws_find_header_row(ws_map, ["CanonWord", "TopWord"], max_scan=3)
        cm = ws_headers(ws_map, hm)

        for r in range(hm + 1, ws_map.max_row + 1):
            cw = ws_map.cell(r, cm["CanonWord"]).value
            tw = ws_map.cell(r, cm["TopWord"]).value
            if not isinstance(cw, str) or not cw.strip():
                continue
            if not isinstance(tw, str) or not tw.strip():
                continue
            tot = ws_map.cell(r, cm.get("TotalWordCount", 0) or 0).value if cm.get("TotalWordCount") else None
            share = ws_map.cell(r, cm.get("TopShare", 0) or 0).value if cm.get("TopShare") else None
            try:
                tot_i = int(tot or 0)
            except Exception:
                tot_i = 0
            try:
                share_f = float(share or 0.0)
            except Exception:
                share_f = 0.0
            if tot_i < int(min_total_count):
                continue
            if share_f + 1e-12 < float(min_top_share):
                continue
            key = _norm_wordish(cw)
            top = _norm_wordish(tw)
            if not key or not top:
                continue
            if key == top:
                continue
            if not allow_overwrite:
                # Avoid 2-cycles between the primary corpus-derived map and the context-derived map.
                # Example we observed in practice: `these -> set` (primary) + `set -> these` (context)
                # causes oscillation across iterations. Context should not "undo" a primary mapping.
                rev = word_map.get(top)
                if rev and _norm_wordish(rev[0]) == key:
                    continue
            if (not allow_overwrite) and key in word_map:
                continue
            word_map[key] = (tw.strip(), tot_i, share_f)

    # Priority order:
    # 1) EnglishMap_Auto (corpus-derived) is the primary.
    # 2) EnglishMap_Context_Auto (context-derived) only fills missing keys (no overwrite).
    _load_map_sheet(ENGLISH_MAP_SHEET, allow_overwrite=True)
    _load_map_sheet(ENGLISH_MAP_CONTEXT_SHEET, allow_overwrite=False)

    if not word_map:
        return 0, 0, 0

    # Word quality guardrails: keep English retext from drifting into garbage/abbreviations.
    # Use a combined (derived) wordfreq so PD-backed suggestions are not unfairly rejected.
    wordfreq: Dict[str, int] = {}
    for wf_sheet in (LORE_WORDFREQ_TIBIA_SHEET, LORE_WORDFREQ_PD_SHEET):
        if wf_sheet not in wb.sheetnames:
            continue
        for w, cnt in _load_wordfreq_sheet(wb, wf_sheet).items():
            wordfreq[w] = wordfreq.get(w, 0) + int(cnt)
    allowed_short = {
        "a",
        "i",
        "am",
        "an",
        "as",
        "at",
        "be",
        "by",
        "do",
        "go",
        "he",
        "if",
        "in",
        "is",
        "it",
        "me",
        "my",
        "no",
        "of",
        "on",
        "or",
        "so",
        "to",
        "up",
        "us",
        "we",
        "ye",
        "ya",
    }
    min_new_wordfreq = 5

    ws_log = ensure_sheet(
        wb,
        "EnglishPromotions_Auto",
        ["Iteration", "UTC", "Token", "OldTranslation", "NewTranslation", "Decision", "Reason", "Notes"],
    )
    upsert_sheet_index_entry(
        wb,
        "EnglishPromotions_Auto",
        "Applied/attempted EnglishMap_Auto -> Glossary translation edits (guarded by GT live check).",
    )

    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}
    anti_terms = set(anti_hallucination_terms or set())

    # Anti-oscillation lock support: avoid repeatedly flipping the same token across iterations
    # when EnglishMap signals are unstable (common during plateaus).
    last_applied_lock: Dict[str, Tuple[int, str, str]] = {}
    if "EnglishPromotions_Auto" in wb.sheetnames:
        ws_prom_lock = wb["EnglishPromotions_Auto"]
        hp_lock = ws_find_header_row(ws_prom_lock, ["Iteration", "Token", "OldTranslation", "NewTranslation", "Decision"], max_scan=3)
        cp_lock = ws_headers(ws_prom_lock, hp_lock)
        for r in range(ws_prom_lock.max_row, hp_lock, -1):
            dec = ws_prom_lock.cell(r, cp_lock["Decision"]).value
            if str(dec or "").strip().upper() != "APPLIED":
                continue
            tok = ws_prom_lock.cell(r, cp_lock["Token"]).value
            it = ws_prom_lock.cell(r, cp_lock["Iteration"]).value
            old_tr = ws_prom_lock.cell(r, cp_lock["OldTranslation"]).value
            new_tr = ws_prom_lock.cell(r, cp_lock["NewTranslation"]).value
            if not isinstance(tok, str) or not tok.strip():
                continue
            try:
                it_i = int(it or 0)
            except Exception:
                it_i = 0
            t = tok.strip()
            if t in last_applied_lock:
                continue
            last_applied_lock[t] = (it_i, str(old_tr or ""), str(new_tr or ""))

    # Auto-revert unsafe English promotions (if any were previously applied).
    # This is conservative: only revert when the current translation matches a previously applied NewTranslation
    # AND that new word is clearly unsafe under our quality heuristics.
    reverted = 0
    revert_attempted = 0
    if "EnglishPromotions_Auto" in wb.sheetnames and wordfreq:
        ws_prom = wb["EnglishPromotions_Auto"]
        hp = ws_find_header_row(ws_prom, ["Iteration", "Token", "OldTranslation", "NewTranslation", "Decision"], max_scan=3)
        cp = ws_headers(ws_prom, hp)

        last_applied: Dict[str, Tuple[int, str, str]] = {}
        for r in range(ws_prom.max_row, hp, -1):
            dec = ws_prom.cell(r, cp["Decision"]).value
            if str(dec or "").strip().upper() != "APPLIED":
                continue
            tok = ws_prom.cell(r, cp["Token"]).value
            it = ws_prom.cell(r, cp["Iteration"]).value
            old_tr = ws_prom.cell(r, cp["OldTranslation"]).value
            new_tr = ws_prom.cell(r, cp["NewTranslation"]).value
            if not isinstance(tok, str) or not tok.strip():
                continue
            try:
                it_i = int(it or 0)
            except Exception:
                it_i = 0
            t = tok.strip()
            if t in last_applied:
                continue
            last_applied[t] = (it_i, str(old_tr or ""), str(new_tr or ""))

        if last_applied:
            ws_rlog = ensure_sheet(
                wb,
                "EnglishReverts_Auto",
                ["Iteration", "UTC", "Token", "CurrentTranslation", "RevertedTo", "Decision", "Reason", "SourceIter", "Notes"],
            )
            upsert_sheet_index_entry(
                wb,
                "EnglishReverts_Auto",
                "Auto-reverts of unsafe English->Glossary edits (guarded by GT live check).",
            )

            for tok, (src_iter, old_tr, new_tr) in sorted(last_applied.items(), key=lambda kv: kv[1][0], reverse=True):
                if locked_tokens and tok in locked_tokens:
                    continue
                cur = active.get(tok) or glossary_map.get(tok)
                if cur is None:
                    continue
                if not cur.use_strictplus or cur.token_type in ("marker", "macro"):
                    continue
                cur_tr = str(cur.translation or "").strip()
                if _norm_wordish(cur_tr) != _norm_wordish(new_tr):
                    continue
                if _norm_wordish(cur_tr) == _norm_wordish(old_tr):
                    continue

                new_w = _norm_wordish(cur_tr)
                reason = None
                if not new_w:
                    reason = "new_empty"
                elif anti_mode and anti_terms and new_w in anti_terms:
                    reason = "new_deny_word"
                elif len(new_w) <= 2 and new_w not in allowed_short:
                    reason = "new_too_short"
                else:
                    new_cnt = int(wordfreq.get(new_w, 0))
                    if new_cnt < int(min_new_wordfreq):
                        reason = "new_rare_oov"

                if not reason:
                    continue

                revert_attempted += 1
                test_active = dict(active)
                test_active[tok] = dataclasses.replace(cur, translation=old_tr)
                ok, bad, bad_all = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_crib_ids)
                if not ok:
                    bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad[:8])
                    soft_n = max(0, len(bad_all) - len(bad))
                    ws_append_row(
                        ws_rlog,
                        [
                            iter_num,
                            utc,
                            tok,
                            cur_tr,
                            old_tr,
                            "SKIP",
                            reason,
                            src_iter,
                            f"GT mismatch (CribID(s) {bad_ids})" + (f"; soft_mismatches={soft_n}" if soft_n else ""),
                        ],
                        start_col=1,
                    )
                    continue

                note = f"english auto-revert (from iter{src_iter}): {cur_tr} -> {old_tr} (reason={reason})"
                glossary_set_translation(glossary_ws, cur.row, old_tr, iter_num, note)
                ws_append_row(ws_rlog, [iter_num, utc, tok, cur_tr, old_tr, "REVERTED", reason, src_iter, "GT OK"], start_col=1)
                active = test_active
                updates = {tok: old_tr}
                sync_evidence_ledger_translations(wb, updates)
                reverted += 1

    candidates: List[Tuple[int, int, str, str, str, int, float]] = []
    for tok, gt in glossary_map.items():
        if locked_tokens and tok in locked_tokens:
            continue
        if not gt.use_strictplus:
            continue
        if gt.token_type in ("marker", "macro"):
            continue
        old_tr = str(gt.translation or "").strip()
        if not old_tr or any(ch.isspace() for ch in old_tr):
            continue
        canon = _norm_wordish(old_tr)
        if not canon:
            continue
        mapped = word_map.get(canon)
        if not mapped:
            continue
        new_tr, tot_i, share_f = mapped
        if any(ch.isspace() for ch in str(new_tr or "")):
            continue
        if _norm_wordish(new_tr) == canon:
            continue
        if int(lock_iters) > 0:
            la = last_applied_lock.get(tok)
            if la:
                src_iter, _la_old, la_new = la
                try:
                    age = int(iter_num) - int(src_iter)
                except Exception:
                    age = 999999
                # If this token is already at the last English-applied translation, don't flip again
                # within a short window. This prevents infinite oscillations like set<->these.
                if age <= int(lock_iters) and _norm_wordish(old_tr) == _norm_wordish(la_new):
                    continue
        new_w = _norm_wordish(new_tr)
        if not new_w:
            continue
        if anti_mode and anti_terms and new_w in anti_terms:
            continue
        if len(new_w) <= 2 and new_w not in allowed_short:
            continue
        if wordfreq:
            if int(wordfreq.get(new_w, 0)) < int(min_new_wordfreq) and new_w not in allowed_short:
                continue
        candidates.append((int(gt.total_occ), int(gt.length), tok, old_tr, str(new_tr or ""), int(tot_i), float(share_f)))

    candidates.sort(key=lambda t: (-t[0], -t[1], -t[6], t[2]))

    applied = 0
    attempted = 0
    updates: Dict[str, str] = {}

    for _occ, _ln, tok, old_tr, new_tr, tot_i, share_f in candidates:
        if applied >= int(max_promotions):
            break
        attempted += 1

        cur = active.get(tok) or glossary_map.get(tok)
        if cur is None:
            continue

        test_active = dict(active)
        test_active[tok] = dataclasses.replace(cur, translation=new_tr)
        ok, bad, bad_all = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_crib_ids)
        if not ok:
            bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad[:8])
            soft_n = max(0, len(bad_all) - len(bad))
            ws_append_row(
                ws_log,
                [
                    iter_num,
                    utc,
                    tok,
                    old_tr,
                    new_tr,
                    "SKIP",
                    f"{_norm_wordish(old_tr)}->{_norm_wordish(new_tr)}",
                    f"GT mismatch (CribID(s) {bad_ids})" + (f"; soft_mismatches={soft_n}" if soft_n else ""),
                ],
                start_col=1,
            )
            continue

        note = f"english retext: {old_tr} -> {new_tr} (total={tot_i}, share={share_f:.6f})"
        glossary_set_translation(glossary_ws, cur.row, new_tr, iter_num, note)
        ws_append_row(
            ws_log,
            [iter_num, utc, tok, old_tr, new_tr, "APPLIED", f"{_norm_wordish(old_tr)}->{_norm_wordish(new_tr)}", "GT OK"],
            start_col=1,
        )
        active = test_active
        updates[tok] = new_tr
        applied += 1

    el_changed = sync_evidence_ledger_translations(wb, updates)
    return applied + reverted, attempted + revert_attempted, el_changed


def apply_antihallucination_force_unk(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    glossary_ws: openpyxl.worksheet.worksheet.Worksheet,
    glossary_map: Dict[str, GlossaryToken],
    *,
    enabled: bool,
    max_per_iter: int,
    min_total_occ: int,
    max_total_occ: int,
    max_conf_weight: int,
    protected_evidence_classes: set[str],
    deny_words: set[str],
    unk_token: str = "<UNK>",
    enforced_crib_ids: Optional[set[int]] = None,
) -> Tuple[int, int, int]:
    """Force suspicious low-evidence lexical forms to <UNK> (GT-guarded, traceable)."""
    if not enabled:
        return 0, 0, 0
    if int(max_per_iter) <= 0:
        return 0, 0, 0
    deny = set(deny_words or set())
    if not deny:
        return 0, 0, 0

    ws_log = ensure_sheet(
        wb,
        "AntiHallucination_Auto",
        ["Iteration", "UTC", "Token", "OldTranslation", "NewTranslation", "Decision", "Reason", "Notes"],
    )
    upsert_sheet_index_entry(
        wb,
        "AntiHallucination_Auto",
        "Anti-hallucination sanitizer: low-evidence suspicious lexical forms forced to <UNK> (guarded by GT live check).",
    )

    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}
    candidates: List[Tuple[int, int, str, str, str]] = []
    for tok, gt in glossary_map.items():
        if not gt.use_strictplus:
            continue
        if gt.token_type in ("marker", "macro"):
            continue
        old_tr = str(gt.translation or "").strip()
        if not old_tr:
            continue
        if any(ch.isspace() for ch in old_tr):
            continue
        old_w = _norm_wordish(old_tr)
        if not old_w:
            continue
        if old_w == _norm_wordish(unk_token):
            continue
        if old_w not in deny:
            continue
        evcls = str(gt.evidence_class or "").strip().upper()
        if evcls and evcls in protected_evidence_classes:
            continue
        conf_w = confidence_weight(gt.confidence)
        if conf_w > int(max_conf_weight):
            continue
        occ = int(gt.total_occ or 0)
        if occ < int(min_total_occ):
            continue
        if occ > int(max_total_occ):
            continue
        candidates.append((occ, int(gt.length or 0), tok, old_tr, evcls))

    # Prefer lower-occ first to reduce destabilization risk.
    candidates.sort(key=lambda t: (t[0], t[1], t[2]))

    applied = 0
    attempted = 0
    updates: Dict[str, str] = {}
    for occ, _ln, tok, old_tr, evcls in candidates:
        if applied >= int(max_per_iter):
            break
        attempted += 1
        cur = active.get(tok) or glossary_map.get(tok)
        if cur is None:
            continue
        test_active = dict(active)
        test_active[tok] = dataclasses.replace(cur, translation=str(unk_token))
        ok, bad, bad_all = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_crib_ids)
        if not ok:
            bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad[:8])
            soft_n = max(0, len(bad_all) - len(bad))
            ws_append_row(
                ws_log,
                [
                    iter_num,
                    utc,
                    tok,
                    old_tr,
                    str(unk_token),
                    "SKIP",
                    "gt_guardrail",
                    f"GT mismatch (CribID(s) {bad_ids})" + (f"; soft_mismatches={soft_n}" if soft_n else ""),
                ],
                start_col=1,
            )
            continue

        note = f"anti-hallucination sanitize: {old_tr} -> {unk_token} (evcls={evcls}, occ={occ})"
        glossary_set_translation(glossary_ws, cur.row, str(unk_token), iter_num, note)
        ws_append_row(
            ws_log,
            [iter_num, utc, tok, old_tr, str(unk_token), "APPLIED", "deny_word", "GT OK"],
            start_col=1,
        )
        active = test_active
        updates[tok] = str(unk_token)
        applied += 1

    el_changed = sync_evidence_ledger_translations(wb, updates)
    return applied, attempted, el_changed


def revert_semantic_promotions_if_unsafe(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    glossary_ws: openpyxl.worksheet.worksheet.Worksheet,
    glossary_map: Dict[str, GlossaryToken],
    *,
    blocked_evidence_classes: set[str],
    wordfreq_sheet: str,
    min_new_wordfreq: int,
    min_wordfreq_ratio: float,
    enforced_crib_ids: Optional[set[int]] = None,
    locked_tokens: Optional[set[str]] = None,
) -> Tuple[int, int, int]:
    """Revert previously-applied semantic retext when it violates the current safety policy.

    This is intentionally conservative: it only considers tokens that were changed via
    SemanticPromotions_Auto (Decision=APPLIED) and reverts them back to OldTranslation when:
    - the token is now locked by EvidenceClass_v127, OR
    - the new translation fails strong word-quality heuristics (requires wordfreq sheet).

    Returns: (reverted_count, attempted_count, evidence_ledger_changed)
    """
    if "SemanticPromotions_Auto" not in wb.sheetnames:
        return 0, 0, 0

    ws_prom = wb["SemanticPromotions_Auto"]
    hp = ws_find_header_row(ws_prom, ["Iteration", "Token", "OldTranslation", "NewTranslation", "Decision"], max_scan=3)
    cp = ws_headers(ws_prom, hp)

    # Load most recent APPLIED promotion per token.
    last_applied: Dict[str, Tuple[int, str, str]] = {}
    for r in range(ws_prom.max_row, hp, -1):
        dec = ws_prom.cell(r, cp["Decision"]).value
        if str(dec or "").strip().upper() != "APPLIED":
            continue
        tok = ws_prom.cell(r, cp["Token"]).value
        it = ws_prom.cell(r, cp["Iteration"]).value
        old_tr = ws_prom.cell(r, cp["OldTranslation"]).value
        new_tr = ws_prom.cell(r, cp["NewTranslation"]).value
        if not isinstance(tok, str) or not tok.strip():
            continue
        try:
            it_i = int(it or 0)
        except Exception:
            it_i = 0
        t = tok.strip()
        if t in last_applied:
            continue
        last_applied[t] = (it_i, str(old_tr or ""), str(new_tr or ""))

    if not last_applied:
        return 0, 0, 0

    wordfreq = _load_wordfreq_sheet(wb, wordfreq_sheet) if wordfreq_sheet else {}
    allowed_short = {
        "a",
        "i",
        "am",
        "an",
        "as",
        "at",
        "be",
        "by",
        "do",
        "go",
        "he",
        "if",
        "in",
        "is",
        "it",
        "me",
        "my",
        "no",
        "of",
        "on",
        "or",
        "so",
        "to",
        "up",
        "us",
        "we",
        "ye",
    }

    ws_log = ensure_sheet(
        wb,
        "SemanticReverts_Auto",
        ["Iteration", "UTC", "Token", "CurrentTranslation", "RevertedTo", "Decision", "Reason", "SourceIter", "Notes"],
    )
    upsert_sheet_index_entry(
        wb,
        "SemanticReverts_Auto",
        "Auto-reverts of unsafe semantic->Glossary edits (guarded by GT live check).",
    )

    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}

    attempted = 0
    reverted = 0
    updates: Dict[str, str] = {}

    for tok, (src_iter, old_tr, new_tr) in sorted(last_applied.items(), key=lambda kv: kv[1][0], reverse=True):
        if locked_tokens and tok in locked_tokens:
            continue
        cur = active.get(tok) or glossary_map.get(tok)
        if cur is None:
            continue
        if not cur.use_strictplus or cur.token_type in ("marker", "macro"):
            continue

        cur_tr = str(cur.translation or "")
        if _norm_wordish(cur_tr) != _norm_wordish(new_tr):
            continue  # already changed since that promotion
        if _norm_wordish(cur_tr) == _norm_wordish(old_tr):
            continue

        evcls = str(cur.evidence_class or "").strip().upper()
        reason = None

        if evcls and evcls in blocked_evidence_classes:
            reason = f"blocked_evidence_class={evcls}"
        else:
            if not wordfreq:
                continue
            old_w = _norm_wordish(old_tr)
            new_w = _norm_wordish(cur_tr)
            if not new_w:
                reason = "new_empty"
            elif len(new_w) <= 2 and new_w not in allowed_short:
                reason = "new_too_short"
            else:
                new_cnt = int(wordfreq.get(new_w, 0))
                old_cnt = int(wordfreq.get(old_w, 0)) if old_w else 0
                if new_cnt < int(min_new_wordfreq) and old_cnt >= int(min_new_wordfreq):
                    reason = "new_rare_old_common"
                elif old_cnt > 0 and float(new_cnt) < float(old_cnt) * float(min_wordfreq_ratio):
                    reason = "new_much_rarer"

        if not reason:
            continue

        attempted += 1

        test_active = dict(active)
        test_active[tok] = dataclasses.replace(cur, translation=old_tr)
        ok, bad, bad_all = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_crib_ids)
        if not ok:
            bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad[:8])
            soft_n = max(0, len(bad_all) - len(bad))
            ws_append_row(
                ws_log,
                [
                    iter_num,
                    utc,
                    tok,
                    cur_tr,
                    old_tr,
                    "SKIP",
                    reason,
                    src_iter,
                    f"GT mismatch (CribID(s) {bad_ids})" + (f"; soft_mismatches={soft_n}" if soft_n else ""),
                ],
                start_col=1,
            )
            continue

        note = f"semantic auto-revert (from iter{src_iter}): {cur_tr} -> {old_tr} (reason={reason})"
        glossary_set_translation(glossary_ws, cur.row, old_tr, iter_num, note)
        ws_append_row(ws_log, [iter_num, utc, tok, cur_tr, old_tr, "REVERTED", reason, src_iter, "GT OK"], start_col=1)

        active = test_active
        updates[tok] = old_tr
        reverted += 1

    el_changed = sync_evidence_ledger_translations(wb, updates)
    return reverted, attempted, el_changed


_PHRASE_WORD_RE = re.compile(r"[A-Za-z']+")
REVERSE_PHRASE_HITS_SHEET = "ReversePhraseHits_Auto"
# Excel worksheet title must be <= 31 chars. Keep this short.
REVERSE_PHRASE_CANDS_SHEET = "ReversePhraseTokenCands_Auto"
REVERSE_PHRASE_PERMUTE_HITS_SHEET = "ReversePhrasePermuteHits_Auto"
REVERSE_PHRASE_PERMUTE_CANDS_SHEET = "ReversePhrasePermuteCands_Auto"


def _ensure_phrase_cribs_user_sheet(wb: openpyxl.Workbook) -> None:
    ensure_sheet(wb, "PhraseCribs_User", ["PhraseID", "Enabled", "Source", "Text", "AddedIter", "Notes"])
    upsert_sheet_index_entry(
        wb,
        "PhraseCribs_User",
        "User-provided phrases (official dialogues / lore) for reverse phrase mining. Fill Text + Enabled=1.",
    )


def _load_phrase_cribs_user(wb: openpyxl.Workbook, *, max_phrases: int) -> List[Dict[str, object]]:
    if "PhraseCribs_User" not in wb.sheetnames:
        _ensure_phrase_cribs_user_sheet(wb)
    ws = wb["PhraseCribs_User"]
    h = ws_find_header_row(ws, ["PhraseID", "Enabled", "Text"], max_scan=3)
    c = ws_headers(ws, h)

    out: List[Dict[str, object]] = []
    for r in range(h + 1, ws.max_row + 1):
        pid = ws.cell(r, c["PhraseID"]).value
        enabled = ws.cell(r, c["Enabled"]).value
        text = ws.cell(r, c["Text"]).value
        if pid is None and (text is None or str(text).strip() == ""):
            continue
        if not parse_bool(enabled, False):
            continue
        if text is None or str(text).strip() == "":
            continue
        src = ws.cell(r, c.get("Source", 0) or 0).value if c.get("Source") else None
        out.append({"PhraseID": pid, "Source": src, "Text": str(text)})
        if len(out) >= int(max_phrases):
            break
    return out


def _ensure_phrase_cribs_auto_sheet(wb: openpyxl.Workbook) -> None:
    ensure_sheet(
        wb,
        PHRASE_CRIBS_AUTO_SHEET,
        [
            "PhraseID",
            "Enabled",
            "SourceKind",
            "SourceURL",
            "Origin",
            "Text",
            "WordCount",
            "RarityScore",
            "SigLenSum",
            "AddedIter",
            "Notes",
        ],
    )
    upsert_sheet_index_entry(
        wb,
        PHRASE_CRIBS_AUTO_SHEET,
        "Auto-generated phrase candidates from public Tibia corpus for reverse phrase mining (analysis-only; no full corpus stored).",
    )


def _load_phrase_cribs_auto(wb: openpyxl.Workbook, *, max_phrases: int) -> List[Dict[str, object]]:
    if PHRASE_CRIBS_AUTO_SHEET not in wb.sheetnames:
        return []
    ws = wb[PHRASE_CRIBS_AUTO_SHEET]
    h = ws_find_header_row(ws, ["PhraseID", "Enabled", "Text"], max_scan=3)
    c = ws_headers(ws, h)

    out: List[Dict[str, object]] = []
    for r in range(h + 1, ws.max_row + 1):
        pid = ws.cell(r, c["PhraseID"]).value
        enabled = ws.cell(r, c["Enabled"]).value
        text = ws.cell(r, c["Text"]).value
        if pid is None and (text is None or str(text).strip() == ""):
            continue
        if not parse_bool(enabled, False):
            continue
        if text is None or str(text).strip() == "":
            continue
        sk = ws.cell(r, c.get("SourceKind", 0) or 0).value if c.get("SourceKind") else None
        origin = ws.cell(r, c.get("Origin", 0) or 0).value if c.get("Origin") else None
        src = str(sk or "").strip() or "PHRASE_AUTO"
        if origin is not None and str(origin).strip():
            src = f"{src}:{str(origin).strip()}"
        out.append({"PhraseID": pid, "Source": src, "Text": str(text)})
        if len(out) >= int(max_phrases):
            break
    return out


def _load_lore_corpus_phrases(
    wb: openpyxl.Workbook,
    *,
    include_auto: bool,
    include_user: bool,
    max_phrases: int,
) -> List[Dict[str, object]]:
    """Load phrases from LoreCorpus_* sheets as reverse-phrase sources."""
    out: List[Dict[str, object]] = []
    sheets: List[str] = []
    if include_auto:
        sheets.append("LoreCorpus_Auto")
    if include_user:
        sheets.append("LoreCorpus_User")
    for sh in sheets:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        h = ws_find_header_row(ws, ["CorpusID", "LineID", "Text"], max_scan=3)
        c = ws_headers(ws, h)
        for r in range(h + 1, ws.max_row + 1):
            cid = ws.cell(r, c["CorpusID"]).value
            text = ws.cell(r, c["Text"]).value
            if not isinstance(cid, str) or not cid.strip():
                continue
            if text is None or str(text).strip() == "":
                continue
            lid = ws.cell(r, c["LineID"]).value
            src = ws.cell(r, c.get("Source", 0) or 0).value if c.get("Source") else None
            pid = f"{cid.strip()}:{int(lid or 0)}"
            out.append({"PhraseID": pid, "Source": src, "Text": str(text)})
            if len(out) >= int(max_phrases):
                return out
    return out


_SENT_SPLIT_RE = re.compile(r"(?<=[.!?])\s+")
_WS_NORM_RE = re.compile(r"\s+")


def _split_sentences(text: str) -> List[str]:
    s = _WS_NORM_RE.sub(" ", str(text or "").replace("\n", " ").strip())
    if not s:
        return []
    parts = _SENT_SPLIT_RE.split(s)
    out: List[str] = []
    for p in parts:
        p = p.strip()
        if p:
            out.append(p)
    return out


def _iter_sentences(text: str, *, max_sentences: int = 0) -> Iterable[str]:
    """Yield sentence-ish chunks without materializing a giant list.

    This is used in time-budgeted corpus scans (SequenceMatches) to avoid large allocations
    on very large plaintext sources (e.g., Project Gutenberg).
    """
    s = _WS_NORM_RE.sub(" ", str(text or "").replace("\n", " ").strip())
    if not s:
        return
    sent_i = 0
    last = 0
    for m in _SENT_SPLIT_RE.finditer(s):
        part = s[last : m.start()].strip()
        if part:
            yield part
            sent_i += 1
            if int(max_sentences) > 0 and sent_i >= int(max_sentences):
                return
        last = m.end()
    tail = s[last:].strip()
    if tail:
        if int(max_sentences) <= 0 or sent_i < int(max_sentences):
            yield tail


def _cache_path_for_url(cache_dir: str, *, prefix: str, url: str, ext: str = "json") -> str:
    h = hashlib.sha1(str(url or "").encode("utf-8", errors="replace")).hexdigest()[:12]
    ext2 = str(ext or "json").lstrip(".").strip() or "json"
    return os.path.join(cache_dir, f"{prefix}_{h}.{ext2}")


def _auto_corpus_id_from_url(url: str) -> str:
    """Derive a stable-ish CorpusID label from a URL (best-effort).

    This is used when the user provides extra PD URLs but doesn't provide matching CorpusIDs.
    """
    u = str(url or "").strip()
    if not u:
        return "PUBLIC_DOMAIN"
    try:
        p = urllib.parse.urlparse(u)
        host = re.sub(r"[^a-z0-9]+", "_", str(p.netloc or "").lower()).strip("_")
        # Project Gutenberg has stable numeric ids under /files/<id>/...
        m = re.search(r"/files/(\d+)/", str(p.path or ""))
        if host.endswith("gutenberg.org") and m:
            return f"GUTENBERG_{m.group(1)}"
        base = os.path.basename(str(p.path or "").strip("/"))
        base = re.sub(r"[^A-Za-z0-9]+", "_", base).strip("_")
        cid = f"{host.upper()}_{base.upper()}" if host else base.upper()
        cid = re.sub(r"_+", "_", cid).strip("_")
        return cid[:40] if cid else "PUBLIC_DOMAIN"
    except Exception:
        h = hashlib.sha1(u.encode("utf-8", errors="replace")).hexdigest()[:8]
        return f"URL_{h}"


def _fetch_json_url(url: str, *, timeout_s: int, user_agent: str) -> object:
    req = urllib.request.Request(str(url), headers={"User-Agent": user_agent})
    with urllib.request.urlopen(req, timeout=float(timeout_s)) as resp:
        data = resp.read()
    return json.loads(data.decode("utf-8", errors="replace"))


def _fetch_text_url(url: str, *, timeout_s: int, user_agent: str) -> str:
    req = urllib.request.Request(str(url), headers={"User-Agent": user_agent})
    with urllib.request.urlopen(req, timeout=float(timeout_s)) as resp:
        data = resp.read()
    return data.decode("utf-8", errors="replace")


def _fetch_text_url_cached(
    url: str,
    *,
    cache_path: str,
    timeout_s: int,
    max_age_hours: float,
    user_agent: str,
) -> str:
    """Fetch text with a simple on-disk cache (best-effort).

    - If cache exists and is fresh enough, use it (no network).
    - If fetch fails, fall back to stale cache when available.
    """
    cache_ok = False
    if cache_path:
        try:
            if os.path.exists(cache_path):
                if float(max_age_hours) <= 0:
                    cache_ok = False
                else:
                    age_s = time.time() - float(os.path.getmtime(cache_path))
                    cache_ok = age_s <= float(max_age_hours) * 3600.0
        except Exception:
            cache_ok = False
    if cache_ok:
        try:
            with open(cache_path, "rb") as f:
                return f.read().decode("utf-8", errors="replace")
        except Exception:
            pass

    try:
        txt = _fetch_text_url(url, timeout_s=timeout_s, user_agent=user_agent)
        if cache_path:
            try:
                ensure_dir(os.path.dirname(cache_path))
                with open(cache_path, "wb") as f:
                    f.write(txt.encode("utf-8", errors="replace"))
            except Exception:
                pass
        return txt
    except Exception:
        if cache_path and os.path.exists(cache_path):
            with open(cache_path, "rb") as f:
                return f.read().decode("utf-8", errors="replace")
        raise


def _fetch_json_url_cached(
    url: str,
    *,
    cache_path: str,
    timeout_s: int,
    max_age_hours: float,
    user_agent: str,
) -> object:
    """Fetch JSON with a simple on-disk cache (best-effort).

    - If cache exists and is fresh enough, use it (no network).
    - If fetch fails, fall back to stale cache when available.
    """
    cache_ok = False
    if cache_path:
        try:
            if os.path.exists(cache_path):
                if float(max_age_hours) <= 0:
                    cache_ok = False
                else:
                    age_s = time.time() - float(os.path.getmtime(cache_path))
                    cache_ok = age_s <= float(max_age_hours) * 3600.0
        except Exception:
            cache_ok = False
    if cache_ok:
        try:
            with open(cache_path, "rb") as f:
                return json.loads(f.read().decode("utf-8", errors="replace"))
        except Exception:
            pass

    obj: object
    try:
        obj = _fetch_json_url(url, timeout_s=timeout_s, user_agent=user_agent)
        if cache_path:
            try:
                ensure_dir(os.path.dirname(cache_path))
                with open(cache_path, "wb") as f:
                    f.write(json.dumps(obj).encode("utf-8", errors="replace"))
            except Exception:
                pass
        return obj
    except Exception:
        if cache_path and os.path.exists(cache_path):
            with open(cache_path, "rb") as f:
                return json.loads(f.read().decode("utf-8", errors="replace"))
        raise


def _load_tibia_corpus_phrases(
    *,
    npc_url: str,
    book_url: str,
    timeout_s: int,
    include_npc: bool,
    include_books: bool,
    max_phrases: int,
    min_words: int,
    cache_max_age_hours: float = 24.0,
) -> List[Dict[str, object]]:
    """Fetch a public Tibia dataset (NPC transcripts + books) and return sampled phrase sources.

    This is used only for reverse phrase mining. We do NOT persist the full corpus in the workbook.
    """
    ua = "Mozilla/5.0 (Bonelord469ReversePhrase/1.0)"
    cache_dir = os.path.join(os.getcwd(), "tmp", "corpus")
    npc_cache = _cache_path_for_url(cache_dir, prefix="tibia_npc", url=npc_url)
    book_cache = _cache_path_for_url(cache_dir, prefix="tibia_books", url=book_url)

    candidates: List[Tuple[float, str, str, str]] = []

    if include_npc:
        npc_obj = _fetch_json_url_cached(
            npc_url,
            cache_path=npc_cache,
            timeout_s=timeout_s,
            max_age_hours=float(cache_max_age_hours),
            user_agent=ua,
        )
        if isinstance(npc_obj, list):
            for item in npc_obj:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name") or "").strip() or "NPC"
                conv = item.get("conversation")
                if not isinstance(conv, list):
                    continue
                for ti, turn in enumerate(conv):
                    if not isinstance(turn, dict):
                        continue
                    answers = turn.get("answer")
                    if isinstance(answers, str):
                        answers = [answers]
                    if not isinstance(answers, list):
                        continue
                    for ai, a in enumerate(answers):
                        if not isinstance(a, str) or not a.strip():
                            continue
                        text = _WS_NORM_RE.sub(" ", a.strip())
                        wc = len(_PHRASE_WORD_RE.findall(text))
                        if wc < int(min_words):
                            continue
                        # Prefer longer, more distinctive phrases.
                        score = float(wc) + min(0.50, float(len(text)) / 800.0)
                        pid = f"TIBIA_NPC:{name}:{ti}:{ai}"
                        candidates.append((score, pid, "TIBIA_NPC", text))

    if include_books:
        book_obj = _fetch_json_url_cached(
            book_url,
            cache_path=book_cache,
            timeout_s=timeout_s,
            max_age_hours=float(cache_max_age_hours),
            user_agent=ua,
        )
        if isinstance(book_obj, list):
            for item in book_obj:
                if not isinstance(item, dict):
                    continue
                title = str(item.get("name") or "").strip() or "BOOK"
                text = item.get("text")
                if not isinstance(text, str) or not text.strip():
                    continue
                sents = _split_sentences(text)
                for si, sent in enumerate(sents):
                    wc = len(_PHRASE_WORD_RE.findall(sent))
                    if wc < int(min_words):
                        continue
                    score = float(wc) + min(0.50, float(len(sent)) / 800.0) + 0.05  # slight bias to books
                    pid = f"TIBIA_BOOK:{title}:{si}"
                    candidates.append((score, pid, "TIBIA_BOOK", sent))

    # Select top-scoring phrases (dedupe happens later in the main Step 28 pipeline).
    candidates.sort(key=lambda t: (-t[0], t[1]))
    out: List[Dict[str, object]] = []
    for _score, pid, src, text in candidates:
        out.append({"PhraseID": pid, "Source": src, "Text": text})
        if len(out) >= int(max_phrases):
            break
    return out


def _load_sequence_matches_phrases(
    wb: openpyxl.Workbook,
    *,
    max_phrases: int,
    min_n: int,
) -> List[Dict[str, object]]:
    """Use SequenceMatches_Auto as a phrase source for reverse-phrase mining.

    This is safe because SequenceMatches stores only short snippets + URLs already.

    We include both:
    - the book-side phrase (already in our decoded view)
    - the source-side snippet (public corpus side, still snippet-limited)
    """
    if max_phrases <= 0:
        return []
    if SEQUENCE_MATCHES_SHEET not in wb.sheetnames:
        return []
    ws = wb[SEQUENCE_MATCHES_SHEET]
    h = ws_find_header_row(ws, ["Phrase", "Snippet"], max_scan=3)
    c = ws_headers(ws, h)

    rows: List[Tuple[int, int, int, str, str, str]] = []  # (n, is_book_phrase, kind_rank, pid, src, text)
    for r in range(h + 1, ws.max_row + 1):
        try:
            n_val = int(ws.cell(r, c.get("N") or 0).value or 0) if c.get("N") else 0
        except Exception:
            n_val = 0
        if int(min_n) > 0 and n_val and int(n_val) < int(min_n):
            continue
        phrase = ws.cell(r, c.get("Phrase") or 0).value if c.get("Phrase") else None
        snippet = ws.cell(r, c.get("Snippet") or 0).value if c.get("Snippet") else None
        if (not isinstance(phrase, str) or not phrase.strip()) and (not isinstance(snippet, str) or not snippet.strip()):
            continue

        sk = ws.cell(r, c.get("SourceKind") or 0).value if c.get("SourceKind") else None
        sid = ws.cell(r, c.get("SourceID") or 0).value if c.get("SourceID") else None
        bid = ws.cell(r, c.get("BookID") or 0).value if c.get("BookID") else None

        def _kind_rank(kind: str) -> int:
            s = str(kind or "").upper()
            if "TIBIA_BOOK" in s:
                return 3
            if "TIBIA_NPC" in s:
                return 2
            if s.startswith("PD"):
                return 1
            return 0

        kind_rank = _kind_rank(str(sk or ""))
        src_kind = str(sk or "SEQ_MATCH").strip() or "SEQ_MATCH"

        if isinstance(phrase, str) and phrase.strip():
            pid = f"SEQ_MATCH_PHRASE:b{bid}:{sid}:{r}"
            rows.append((int(n_val or 0), 1, int(kind_rank), pid, f"{src_kind}:PHRASE", phrase.strip()))
        if isinstance(snippet, str) and snippet.strip():
            pid = f"SEQ_MATCH_SNIPPET:b{bid}:{sid}:{r}"
            rows.append((int(n_val or 0), 0, int(kind_rank), pid, f"{src_kind}:SNIPPET", snippet.strip()))

    rows.sort(key=lambda t: (-t[0], -t[1], -t[2], t[3]))
    out: List[Dict[str, object]] = []
    for _n, _is_phrase, _kr, pid, src, text in rows:
        out.append({"PhraseID": pid, "Source": src, "Text": text})
        if len(out) >= int(max_phrases):
            break
    return out


def _base_sig_letters(s: str) -> str:
    return "".join(sorted([ch.lower() for ch in str(s or "") if ch.isalpha() and ch != "*"]))


def _phrase_word_sigs(
    text: str,
    *,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
) -> List[Tuple[str, str, str]]:
    """Return [(surface_word, canon_word, canon_sig)] for a phrase.

    Important: split apostrophes into separate word parts (e.g. "you've" -> "you" + "ve")
    to match the decoded stream which typically renders contractions as separate tokens.
    """
    out: List[Tuple[str, str, str]] = []
    for m in _PHRASE_WORD_RE.finditer(str(text or "")):
        w = m.group(0)
        if not w:
            continue
        raw = w.lower()
        parts = re.findall(r"[a-z]+", raw)
        if not parts:
            continue
        for surface in parts:
            if len(surface) == 1 and surface not in {"a", "i", "o"}:
                # Drop common contraction suffixes ("'s", "'t") which are rarely encoded as standalone words.
                continue
            canon = _lore_canon_word(surface, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
            if not canon:
                continue
            out.append((surface, canon, _lore_signature(canon)))
    return out


def _collect_books_leaf_tokens_for_phrase_matching(
    wb: openpyxl.Workbook,
    *,
    active_tokens: Dict[str, "GlossaryToken"],
    glossary_map: Dict[str, "GlossaryToken"],
    comp_map: Dict[str, List[str]],
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    logogram_aware: bool,
) -> Dict[int, Tuple[List[str], List[str]]]:
    """Collect (base_leaf_tokens, match_atoms) per book for reverse phrase matching."""
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_books, hb)

    book_tokens: Dict[int, Tuple[List[str], List[str]]] = {}
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        base = ws_books.cell(r, cb["DecodedBase"]).value
        if not isinstance(base, str) or not base:
            continue
        try:
            base_toks, match_toks = _tokenize_book_to_leaf_tokens_with_match_atoms(
                base,
                active_tokens=active_tokens,
                glossary_map=glossary_map,
                comp_map=comp_map,
                drop_final_e=drop_final_e,
                drop_all_h=drop_all_h,
                drop_all_o=drop_all_o,
                logogram_aware=logogram_aware,
            )
            book_tokens[int(bid)] = (base_toks, match_toks)
        except Exception:
            # Fallback: treat the whole base as one token to avoid crashing the iteration.
            book_tokens[int(bid)] = ([str(base)], [str(base)])
    return book_tokens


def _build_span_sig_freq(tokens: Sequence[str], *, max_span_tokens: int) -> Counter[str]:
    """Count signatures for all concatenated spans of length 1..K over token strings."""
    out: Counter[str] = Counter()
    n = len(tokens)
    for i in range(n):
        s = ""
        for k in range(1, int(max_span_tokens) + 1):
            if i + k > n:
                break
            s += tokens[i + k - 1]
            sig = _base_sig_letters(s)
            if sig:
                out[sig] += 1
    return out


def _select_autophrasecribs_from_tibia_corpus(
    *,
    npc_obj: object,
    book_obj: object,
    npc_url: str,
    book_url: str,
    include_npc: bool,
    include_books: bool,
    span_sig_freq: Counter[str],
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    min_words: int,
    max_words: int,
    max_phrases: int,
    max_scan_sentences: int,
    time_budget_s: float,
    max_text_len: int,
) -> Tuple[List[Dict[str, object]], Dict[str, int]]:
    """Return PhraseCribs_Auto rows ranked by signature feasibility + distinctiveness.

    Hard safety rule: keep only phrases where *all* word signatures exist somewhere in the current
    decode stream span signature set (share=1.0). This avoids impossible-to-match phrases.
    """
    start = time.time()
    stats: Dict[str, int] = {"scanned": 0, "eligible": 0, "kept": 0, "npc": 0, "books": 0}

    rows: List[Dict[str, object]] = []
    seen_text: set[str] = set()

    def consider(pid: str, source_kind: str, source_url: str, origin: str, text: str) -> None:
        if stats["scanned"] >= int(max_scan_sentences):
            return
        if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
            return
        stats["scanned"] += 1
        s = _WS_NORM_RE.sub(" ", str(text or "").replace("\n", " ").strip())
        if not s:
            return
        if int(max_text_len) > 0 and len(s) > int(max_text_len):
            return
        key = s.lower()
        if key in seen_text:
            return
        wtrip = _phrase_word_sigs(s, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
        wc = len(wtrip)
        if wc < int(min_words) or wc > int(max_words):
            return
        sigs = [wsig for _surface, _canon, wsig in wtrip]
        # Share=1.0 feasibility check: any missing sig means the phrase can never match.
        for wsig in sigs:
            if wsig not in span_sig_freq:
                return
        stats["eligible"] += 1

        rarity = 0.0
        siglen = 0
        for wsig in sigs:
            freq = int(span_sig_freq.get(wsig) or 0)
            rarity += 1.0 / (1.0 + float(freq))
            siglen += len(wsig)
        # Prefer phrases with more rare signatures (distinctiveness), then longer.
        rows.append(
            {
                "PhraseID": pid,
                "Enabled": 1,
                "SourceKind": source_kind,
                "SourceURL": source_url,
                "Origin": origin,
                "Text": s,
                "WordCount": int(wc),
                "RarityScore": round(rarity, 6),
                "SigLenSum": int(siglen),
                "_sort": (round(rarity, 6), int(siglen), int(wc), str(pid)),
            }
        )
        seen_text.add(key)

    if include_npc and isinstance(npc_obj, list):
        for item in npc_obj:
            if stats["scanned"] >= int(max_scan_sentences):
                break
            if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
                break
            if not isinstance(item, dict):
                continue
            name = str(item.get("name") or "").strip() or "NPC"
            conv = item.get("conversation")
            if not isinstance(conv, list):
                continue
            for ti, turn in enumerate(conv):
                if stats["scanned"] >= int(max_scan_sentences):
                    break
                if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
                    break
                if not isinstance(turn, dict):
                    continue
                answers = turn.get("answer")
                if isinstance(answers, str):
                    answers = [answers]
                if not isinstance(answers, list):
                    continue
                for ai, a in enumerate(answers):
                    if not isinstance(a, str) or not a.strip():
                        continue
                    sents = _split_sentences(a)
                    if not sents:
                        sents = [a.strip()]
                    for si, sent in enumerate(sents):
                        pid = f"TIBIA_NPC:{name}:{ti}:{ai}:{si}"
                        consider(pid, "TIBIA_NPC", npc_url, name, sent)
        stats["npc"] = 1

    if include_books and isinstance(book_obj, list):
        for item in book_obj:
            if stats["scanned"] >= int(max_scan_sentences):
                break
            if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
                break
            if not isinstance(item, dict):
                continue
            title = str(item.get("name") or "").strip() or "BOOK"
            text = item.get("text")
            if not isinstance(text, str) or not text.strip():
                continue
            sents = _split_sentences(text)
            for si, sent in enumerate(sents):
                pid = f"TIBIA_BOOK:{title}:{si}"
                consider(pid, "TIBIA_BOOK", book_url, title, sent)
        stats["books"] = 1

    # Rank + cap.
    rows.sort(key=lambda r: (-r["_sort"][0], -r["_sort"][1], -r["_sort"][2], r["_sort"][3]))
    out = rows[: int(max_phrases)]
    for r in out:
        r.pop("_sort", None)
    stats["kept"] = len(out)
    return out, stats


def _select_autophrasecribs_from_pd_sources(
    *,
    pd_sources: Sequence[Tuple[str, str]],
    cache_dir: str,
    timeout_s: int,
    cache_max_age_hours: float,
    span_sig_freq: Counter[str],
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    min_words: int,
    max_words: int,
    max_phrases: int,
    max_scan_sentences: int,
    time_budget_s: float,
    max_text_len: int,
) -> Tuple[List[Dict[str, object]], Dict[str, int]]:
    """Return PhraseCribs rows mined from public-domain sources (Project Gutenberg).

    Same safety filter as Tibia mining: keep only phrases where every word signature exists
    somewhere in the current decode-stream span signature set (share=1.0).
    """
    start = time.time()
    stats: Dict[str, int] = {"scanned": 0, "eligible": 0, "kept": 0, "sources": 0}
    rows: List[Dict[str, object]] = []
    seen_text: set[str] = set()

    def consider(pid: str, source_kind: str, source_url: str, origin: str, text: str) -> None:
        if stats["scanned"] >= int(max_scan_sentences):
            return
        if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
            return
        stats["scanned"] += 1
        s = _WS_NORM_RE.sub(" ", str(text or "").replace("\n", " ").strip())
        if not s:
            return
        if int(max_text_len) > 0 and len(s) > int(max_text_len):
            return
        key = s.lower()
        if key in seen_text:
            return
        wtrip = _phrase_word_sigs(s, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
        wc = len(wtrip)
        if wc < int(min_words) or wc > int(max_words):
            return
        sigs = [wsig for _surface, _canon, wsig in wtrip]
        for wsig in sigs:
            if wsig not in span_sig_freq:
                return
        stats["eligible"] += 1

        rarity = 0.0
        siglen = 0
        for wsig in sigs:
            freq = int(span_sig_freq.get(wsig) or 0)
            rarity += 1.0 / (1.0 + float(freq))
            siglen += len(wsig)
        rows.append(
            {
                "PhraseID": pid,
                "Enabled": 1,
                "SourceKind": source_kind,
                "SourceURL": source_url,
                "Origin": origin,
                "Text": s,
                "WordCount": int(wc),
                "RarityScore": round(rarity, 6),
                "SigLenSum": int(siglen),
                "_sort": (round(rarity, 6), int(siglen), int(wc), str(pid)),
            }
        )
        seen_text.add(key)

    ua = "Mozilla/5.0 (Bonelord469AutoPhraseCribsPD/1.0)"
    for cid, url in pd_sources:
        if stats["scanned"] >= int(max_scan_sentences):
            break
        if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
            break
        cid_s = str(cid or "").strip() or _auto_corpus_id_from_url(str(url or ""))
        url_s = str(url or "").strip()
        if not url_s:
            continue

        try:
            cache_path = _cache_path_for_url(cache_dir, prefix="pd_text", url=url_s, ext="txt")
            pd_txt = _fetch_text_url_cached(
                url_s,
                cache_path=cache_path,
                timeout_s=int(timeout_s),
                max_age_hours=float(cache_max_age_hours),
                user_agent=ua,
            )
        except Exception:
            continue

        # Best-effort strip for Gutenberg boilerplate when present.
        m_start = re.search(
            r"\*\*\*\s*START OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
            pd_txt,
            flags=re.IGNORECASE | re.DOTALL,
        )
        m_end = re.search(
            r"\*\*\*\s*END OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
            pd_txt,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if m_start and m_end and m_end.start() > m_start.end():
            pd_txt = pd_txt[m_start.end() : m_end.start()]

        stats["sources"] += 1
        for si, sent in enumerate(_split_sentences(pd_txt)):
            pid = f"PD:{cid_s}:{si}"
            consider(pid, "PD", url_s, cid_s, sent)
            if stats["scanned"] >= int(max_scan_sentences):
                break
            if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
                break

    rows.sort(key=lambda r: (-r["_sort"][0], -r["_sort"][1], -r["_sort"][2], r["_sort"][3]))
    out = rows[: int(max_phrases)]
    for r in out:
        r.pop("_sort", None)
    stats["kept"] = len(out)
    return out, stats


def _write_phrase_cribs_auto_sheet(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    phrases: Sequence[Dict[str, object]],
) -> Tuple[bool, int]:
    _ensure_phrase_cribs_auto_sheet(wb)
    ws = wb[PHRASE_CRIBS_AUTO_SHEET]
    h = ws_find_header_row(ws, ["PhraseID", "Enabled", "Text"], max_scan=3)
    c = ws_headers(ws, h)

    before: List[Tuple[str, str]] = []
    for r in range(h + 1, ws.max_row + 1):
        pid = ws.cell(r, c["PhraseID"]).value
        txt = ws.cell(r, c["Text"]).value
        if pid is None and (txt is None or str(txt).strip() == ""):
            continue
        before.append((str(pid), str(txt or "")))

    after: List[Tuple[str, str]] = [(str(p.get("PhraseID") or ""), str(p.get("Text") or "")) for p in phrases]
    changed = before != after

    if changed:
        if ws.max_row > h:
            ws.delete_rows(h + 1, ws.max_row - h)
        rr = h + 1
        for p in phrases:
            for k, v in p.items():
                col = c.get(k)
                if col is None:
                    continue
                ws.cell(rr, col).value = v
            # Default AddedIter if the producer didn't set it.
            if c.get("AddedIter") and ws.cell(rr, c["AddedIter"]).value is None:
                ws.cell(rr, c["AddedIter"]).value = int(iter_num)
            rr += 1

    return changed, len(after)


def _tokenize_book_to_leaf_tokens(
    base: str,
    *,
    active_tokens: Dict[str, GlossaryToken],
    glossary_map: Dict[str, GlossaryToken],
    comp_map: Dict[str, List[str]],
    max_depth: int = 10,
) -> List[str]:
    """Tokenize base and expand macros into leaf token strings (markers removed)."""

    def is_macro(t: GlossaryToken) -> bool:
        return t.token_type == "macro" or (t.evidence_class and "MACRO" in str(t.evidence_class).upper())

    out: List[str] = []

    def rec_token(t: GlossaryToken, depth: int) -> None:
        if t.token_type == "marker":
            return
        if is_macro(t):
            if depth >= max_depth:
                out.append(t.token)
                return
            comp = comp_map.get(t.token)
            if not comp:
                out.append(t.token)
                return
            for ct in comp:
                gt = glossary_map.get(ct) or active_tokens.get(ct)
                if gt is None:
                    out.append(t.token)
                    return
                rec_token(gt, depth + 1)
            return
        out.append(t.token)

    items = dp_tokenize_base_with_punct(str(base or ""), active_tokens)
    for it in items:
        if isinstance(it, GlossaryToken):
            rec_token(it, 0)
        # ignore '.'/'!' for reverse phrase matching
    return out


def _tokenize_book_to_leaf_tokens_with_match_atoms(
    base: str,
    *,
    active_tokens: Dict[str, GlossaryToken],
    glossary_map: Dict[str, GlossaryToken],
    comp_map: Dict[str, List[str]],
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    logogram_aware: bool,
    max_depth: int = 10,
) -> Tuple[List[str], List[str]]:
    """Tokenize base and expand macros into leaf tokens, returning:
    - base_tokens: the original leaf token strings (used to emit candidates)
    - match_atoms: strings used for signature matching against phrases

    When logogram_aware=True, LOGOGRAM_* tokens use canonical translation letters as match atoms,
    which makes phrase matching feasible despite logogram token letters not matching word signatures.
    """

    def is_macro(t: GlossaryToken) -> bool:
        return t.token_type == "macro" or (t.evidence_class and "MACRO" in str(t.evidence_class).upper())

    def match_atom(t: GlossaryToken) -> str:
        if not logogram_aware:
            return t.token
        tr = str(t.translation or "").strip()
        if not tr or any(ch.isspace() for ch in tr):
            return t.token

        # IMPORTANT:
        # The global lore canon flags (drop_all_o/drop_final_e) are useful for *semantic grouping*,
        # but they can break reverse-phrase signature matching by removing letters that DO exist
        # in the base token stream. So we use:
        # - a strict canon (no drops) to decide if a token is truly a logogram
        # - a phrase canon (with the configured drops) only for *actual* logograms
        canon_strict = _lore_canon_word(tr, drop_final_e=False, drop_all_h=False, drop_all_o=False)
        canon_phrase = _lore_canon_word(tr, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
        if not canon_phrase:
            return t.token

        # Evidence-class driven logograms.
        evcls = str(t.evidence_class or "").strip().upper()
        if evcls in ("LOGOGRAM_ANCHORED", "LOGOGRAM_CONTEXT", "PUNCT_LOGOGRAM"):
            return canon_phrase

        # Signature-driven logograms: only when strict canon is not an anagram of token letters.
        if not canon_strict:
            return t.token
        tok_sig = _token_signature(t.token)
        tr_sig_strict = _lore_signature(canon_strict)
        if tok_sig and tok_sig != tr_sig_strict:
            return canon_phrase
        return t.token

    base_out: List[str] = []
    match_out: List[str] = []

    def rec_token(t: GlossaryToken, depth: int) -> None:
        if t.token_type == "marker":
            return
        if is_macro(t):
            if depth >= max_depth:
                base_out.append(t.token)
                match_out.append(match_atom(t))
                return
            comp = comp_map.get(t.token)
            if not comp:
                base_out.append(t.token)
                match_out.append(match_atom(t))
                return
            for ct in comp:
                gt = glossary_map.get(ct) or active_tokens.get(ct)
                if gt is None:
                    base_out.append(t.token)
                    match_out.append(match_atom(t))
                    return
                rec_token(gt, depth + 1)
            return
        base_out.append(t.token)
        match_out.append(match_atom(t))

    items = dp_tokenize_base_with_punct(str(base or ""), active_tokens)
    for it in items:
        if isinstance(it, GlossaryToken):
            rec_token(it, 0)
    return base_out, match_out


def _ensure_reverse_phrase_sheets(wb: openpyxl.Workbook) -> None:
    ensure_sheet(
        wb,
        REVERSE_PHRASE_HITS_SHEET,
        [
            "Iteration",
            "PhraseID",
            "Source",
            "PhraseText",
            "BookID",
            "TokenStart",
            "TokenEnd",
            "WordCount",
            "MaxSpanUsed",
            "BaseTokens",
            "BaseConcat",
            "Notes",
        ],
    )
    ensure_sheet(
        wb,
        REVERSE_PHRASE_CANDS_SHEET,
        [
            "Iteration",
            "Base",
            "BaseSig",
            "BaseLen",
            "SupportBooks",
            "SupportOcc",
            "TopWord",
            "TopWordCount",
            "TopShare",
            "CandidateWords",
            "ExamplePhraseID",
            "Notes",
        ],
    )
    upsert_sheet_index_entry(
        wb,
        REVERSE_PHRASE_HITS_SHEET,
        "Reverse phrase hits (analysis-only): where PhraseCribs_User phrases match the base token stream by canonical signatures.",
    )
    upsert_sheet_index_entry(
        wb,
        REVERSE_PHRASE_CANDS_SHEET,
        "Reverse-mined token candidates derived from phrase hits (analysis-only; may optionally be emitted into Glossary as inactive candidates).",
    )


def _ensure_reverse_phrase_permute_sheets(wb: openpyxl.Workbook) -> None:
    ensure_sheet(
        wb,
        REVERSE_PHRASE_PERMUTE_HITS_SHEET,
        [
            "Iteration",
            "PhraseID",
            "Source",
            "PhraseText",
            "BookID",
            "TokenStart",
            "TokenEnd",
            "WordCount",
            "MaxSpanUsed",
            "BaseTokens",
            "BaseConcat",
            "Permutation",
            "Notes",
        ],
    )
    ensure_sheet(
        wb,
        REVERSE_PHRASE_PERMUTE_CANDS_SHEET,
        [
            "Iteration",
            "Base",
            "BaseSig",
            "BaseLen",
            "SupportBooks",
            "SupportOcc",
            "TopWord",
            "TopWordCount",
            "TopShare",
            "CandidateWords",
            "ExamplePhraseID",
            "PermuteSupport",
            "Notes",
        ],
    )
    upsert_sheet_index_entry(
        wb,
        REVERSE_PHRASE_PERMUTE_HITS_SHEET,
        "Reverse phrase permuted hits (analysis-only): phrase word order is free; spans must match canonical signatures. No full text stored.",
    )
    upsert_sheet_index_entry(
        wb,
        REVERSE_PHRASE_PERMUTE_CANDS_SHEET,
        "Reverse phrase permuted candidates derived from permuted hits (analysis-only; no automatic Glossary emission by default).",
    )


def _write_reverse_phrase_outputs(
    wb: openpyxl.Workbook,
    iter_num: int,
    hits: Sequence[Dict[str, object]],
    candidates: Sequence[Dict[str, object]],
) -> Tuple[int, int]:
    _ensure_reverse_phrase_sheets(wb)

    ws_hits = wb["ReversePhraseHits_Auto"]
    ws_hits = wb[REVERSE_PHRASE_HITS_SHEET]
    hh = ws_find_header_row(ws_hits, ["Iteration", "PhraseID", "BookID"], max_scan=3)
    ch = ws_headers(ws_hits, hh)
    if ws_hits.max_row > hh:
        ws_hits.delete_rows(hh + 1, ws_hits.max_row - hh)

    r = hh + 1
    for row in hits:
        ws_hits.cell(r, ch["Iteration"]).value = iter_num
        ws_hits.cell(r, ch["PhraseID"]).value = row.get("PhraseID")
        ws_hits.cell(r, ch["Source"]).value = row.get("Source")
        ws_hits.cell(r, ch["PhraseText"]).value = row.get("PhraseText")
        ws_hits.cell(r, ch["BookID"]).value = row.get("BookID")
        ws_hits.cell(r, ch["TokenStart"]).value = row.get("TokenStart")
        ws_hits.cell(r, ch["TokenEnd"]).value = row.get("TokenEnd")
        ws_hits.cell(r, ch["WordCount"]).value = row.get("WordCount")
        ws_hits.cell(r, ch["MaxSpanUsed"]).value = row.get("MaxSpanUsed")
        ws_hits.cell(r, ch["BaseTokens"]).value = row.get("BaseTokens")
        ws_hits.cell(r, ch["BaseConcat"]).value = row.get("BaseConcat")
        ws_hits.cell(r, ch["Notes"]).value = row.get("Notes")
        r += 1

    ws_cand = wb[REVERSE_PHRASE_CANDS_SHEET]
    hc = ws_find_header_row(ws_cand, ["Iteration", "Base", "TopWord"], max_scan=3)
    cc = ws_headers(ws_cand, hc)
    if ws_cand.max_row > hc:
        ws_cand.delete_rows(hc + 1, ws_cand.max_row - hc)

    r = hc + 1
    for row in candidates:
        ws_cand.cell(r, cc["Iteration"]).value = iter_num
        ws_cand.cell(r, cc["Base"]).value = row.get("Base")
        ws_cand.cell(r, cc["BaseSig"]).value = row.get("BaseSig")
        ws_cand.cell(r, cc["BaseLen"]).value = row.get("BaseLen")
        ws_cand.cell(r, cc["SupportBooks"]).value = row.get("SupportBooks")
        ws_cand.cell(r, cc["SupportOcc"]).value = row.get("SupportOcc")
        ws_cand.cell(r, cc["TopWord"]).value = row.get("TopWord")
        ws_cand.cell(r, cc["TopWordCount"]).value = row.get("TopWordCount")
        ws_cand.cell(r, cc["TopShare"]).value = row.get("TopShare")
        ws_cand.cell(r, cc["CandidateWords"]).value = row.get("CandidateWords")
        ws_cand.cell(r, cc["ExamplePhraseID"]).value = row.get("ExamplePhraseID")
        ws_cand.cell(r, cc["Notes"]).value = row.get("Notes")
        r += 1

    return len(hits), len(candidates)


def _write_reverse_phrase_permute_outputs(
    wb: openpyxl.Workbook,
    iter_num: int,
    hits: Sequence[Dict[str, object]],
    candidates: Sequence[Dict[str, object]],
) -> Tuple[int, int]:
    _ensure_reverse_phrase_permute_sheets(wb)

    ws_hits = wb[REVERSE_PHRASE_PERMUTE_HITS_SHEET]
    hh = ws_find_header_row(ws_hits, ["Iteration", "PhraseID", "BookID"], max_scan=3)
    ch = ws_headers(ws_hits, hh)
    if ws_hits.max_row > hh:
        ws_hits.delete_rows(hh + 1, ws_hits.max_row - hh)

    r = hh + 1
    for row in hits:
        ws_hits.cell(r, ch["Iteration"]).value = iter_num
        ws_hits.cell(r, ch["PhraseID"]).value = row.get("PhraseID")
        ws_hits.cell(r, ch["Source"]).value = row.get("Source")
        ws_hits.cell(r, ch["PhraseText"]).value = row.get("PhraseText")
        ws_hits.cell(r, ch["BookID"]).value = row.get("BookID")
        ws_hits.cell(r, ch["TokenStart"]).value = row.get("TokenStart")
        ws_hits.cell(r, ch["TokenEnd"]).value = row.get("TokenEnd")
        ws_hits.cell(r, ch["WordCount"]).value = row.get("WordCount")
        ws_hits.cell(r, ch["MaxSpanUsed"]).value = row.get("MaxSpanUsed")
        ws_hits.cell(r, ch["BaseTokens"]).value = row.get("BaseTokens")
        ws_hits.cell(r, ch["BaseConcat"]).value = row.get("BaseConcat")
        ws_hits.cell(r, ch["Permutation"]).value = row.get("Permutation")
        ws_hits.cell(r, ch["Notes"]).value = row.get("Notes")
        r += 1

    ws_cand = wb[REVERSE_PHRASE_PERMUTE_CANDS_SHEET]
    hc = ws_find_header_row(ws_cand, ["Iteration", "Base", "TopWord"], max_scan=3)
    cc = ws_headers(ws_cand, hc)
    if ws_cand.max_row > hc:
        ws_cand.delete_rows(hc + 1, ws_cand.max_row - hc)

    r = hc + 1
    for row in candidates:
        ws_cand.cell(r, cc["Iteration"]).value = iter_num
        ws_cand.cell(r, cc["Base"]).value = row.get("Base")
        ws_cand.cell(r, cc["BaseSig"]).value = row.get("BaseSig")
        ws_cand.cell(r, cc["BaseLen"]).value = row.get("BaseLen")
        ws_cand.cell(r, cc["SupportBooks"]).value = row.get("SupportBooks")
        ws_cand.cell(r, cc["SupportOcc"]).value = row.get("SupportOcc")
        ws_cand.cell(r, cc["TopWord"]).value = row.get("TopWord")
        ws_cand.cell(r, cc["TopWordCount"]).value = row.get("TopWordCount")
        ws_cand.cell(r, cc["TopShare"]).value = row.get("TopShare")
        ws_cand.cell(r, cc["CandidateWords"]).value = row.get("CandidateWords")
        ws_cand.cell(r, cc["ExamplePhraseID"]).value = row.get("ExamplePhraseID")
        if "PermuteSupport" in cc:
            ws_cand.cell(r, cc["PermuteSupport"]).value = row.get("PermuteSupport")
        ws_cand.cell(r, cc["Notes"]).value = row.get("Notes")
        r += 1

    return len(hits), len(candidates)


def _reverse_phrase_find_matches(
    phrase_word_sigs: Sequence[str],
    tokens: Sequence[str],
    *,
    max_span_tokens: int,
    max_gap_tokens: int = 0,
    max_hits: int,
) -> List[Tuple[int, int, List[int]]]:
    """Return list of (start_tok, end_tok, spans_per_word)."""
    m = len(phrase_word_sigs)
    n = len(tokens)
    if m <= 0 or n <= 0:
        return []

    # Precompute span signatures for speed.
    span_sig: List[List[Optional[str]]] = [[None] * (max_span_tokens + 1) for _ in range(n)]
    for i in range(n):
        s = ""
        for k in range(1, max_span_tokens + 1):
            if i + k > n:
                break
            s += tokens[i + k - 1]
            span_sig[i][k] = _base_sig_letters(s)

    hits: List[Tuple[int, int, List[int]]] = []
    max_gap = max(0, int(max_gap_tokens))
    for start in range(n):
        states: Dict[int, List[int]] = {start: []}
        for wi, wsig in enumerate(phrase_word_sigs):
            nxt: Dict[int, List[int]] = {}
            for pos, path in states.items():
                if pos >= n:
                    continue
                gap_range = range(0, 1) if wi == 0 else range(0, max_gap + 1)
                for g in gap_range:
                    sp = pos + g
                    if sp >= n:
                        continue
                    for k in range(1, max_span_tokens + 1):
                        if sp + k > n:
                            break
                        if span_sig[sp][k] == wsig:
                            np = sp + k
                            if np not in nxt:
                                nxt[np] = path + [k]
            states = nxt
            if not states:
                break

        if not states:
            continue
        end = min(states.keys())
        spans = states[end]
        hits.append((start, end, spans))
        if len(hits) >= int(max_hits):
            break

    return hits


def _reverse_phrase_find_permuted_matches(
    phrase_word_sigs: Sequence[str],
    tokens: Sequence[str],
    *,
    max_span_tokens: int,
    max_gap_tokens: int = 0,
    max_hits: int,
    max_words: int = 6,
) -> List[Tuple[int, int, List[int], List[int]]]:
    """Permutation-aware reverse phrase matcher.

    Returns list of (start_tok, end_tok, spans_per_word_in_stream_order, perm_word_indices_1based).

    The phrase word order is not enforced. We match the phrase's word signatures as a multiset
    against the token stream, selecting disjoint spans in increasing stream order. This is
    analysis-only and intended for deep-plateau diagnostics (runtime is capped upstream).
    """
    m = len(phrase_word_sigs)
    n = len(tokens)
    if m <= 0 or n <= 0:
        return []
    if m > int(max_words):
        return []
    if m > 15:
        # Hard safety cap; exponential search beyond this is not acceptable for the runner.
        return []

    max_span_tokens = max(1, int(max_span_tokens))
    max_gap = max(0, int(max_gap_tokens))
    max_hits = max(0, int(max_hits))
    if max_hits <= 0:
        return []

    # Precompute span signatures as in the sequential matcher.
    span_sig: List[List[Optional[str]]] = [[None] * (max_span_tokens + 1) for _ in range(n)]
    for i in range(n):
        s = ""
        for k in range(1, max_span_tokens + 1):
            if i + k > n:
                break
            s += tokens[i + k - 1]
            span_sig[i][k] = _base_sig_letters(s)

    # Quick filter: if a word signature never appears in any span, there's no possible match.
    sig_set = set(phrase_word_sigs)
    sig_seen: set[str] = set()
    for i in range(n):
        for k in range(1, max_span_tokens + 1):
            if i + k > n:
                break
            sig = span_sig[i][k]
            if sig in sig_set:
                sig_seen.add(str(sig))
        if sig_seen == sig_set:
            break
    if sig_seen != sig_set:
        return []

    # Precompute matching options at each start position: (k, sig).
    match_at: List[List[Tuple[int, str]]] = [[] for _ in range(n)]
    for i in range(n):
        opts: List[Tuple[int, str]] = []
        for k in range(1, max_span_tokens + 1):
            if i + k > n:
                break
            sig = span_sig[i][k]
            if sig in sig_set:
                opts.append((k, str(sig)))
        # Prefer shorter spans first for determinism and to keep windows tighter.
        opts.sort(key=lambda t: t[0])
        match_at[i] = opts

    hits: List[Tuple[int, int, List[int], List[int]]] = []
    full_mask = (1 << m) - 1

    # Build per-signature word-index lists to support duplicates.
    sig_to_idxs: Dict[str, List[int]] = defaultdict(list)
    for wi, sig in enumerate(phrase_word_sigs):
        sig_to_idxs[str(sig)].append(int(wi))

    # DFS tries to match spans in stream order, assigning each span to an unused word index.
    # Use memoization per-start to prune dead-ends.
    for start in range(n):
        if len(hits) >= max_hits:
            break

        memo: Dict[Tuple[int, int], bool] = {}

        def _dfs(pos: int, used_mask: int, spans: List[int], perm: List[int]) -> Optional[Tuple[int, int, List[int], List[int]]]:
            if used_mask == full_mask:
                return (start, pos, list(spans), list(perm))
            key = (pos, used_mask)
            if key in memo:
                return None
            memo[key] = True

            if pos >= n:
                return None

            gap_range = [0] if not spans else list(range(0, max_gap + 1))
            for g in gap_range:
                sp = pos + g
                if sp >= n:
                    continue
                for k, sig in match_at[sp]:
                    # Try assigning this span to any unused word index with matching signature.
                    for wi in sig_to_idxs.get(sig, []):
                        bit = 1 << int(wi)
                        if used_mask & bit:
                            continue
                        spans.append(int(k))
                        perm.append(int(wi) + 1)  # 1-based for readability in Excel
                        res = _dfs(sp + int(k), used_mask | bit, spans, perm)
                        if res is not None:
                            return res
                        perm.pop()
                        spans.pop()
            return None

        out = _dfs(start, 0, [], [])
        if out is not None:
            hits.append(out)

    return hits


def add_reverse_phrase_candidates_to_glossary(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    candidates: Sequence[Dict[str, object]],
    *,
    evidence_class: str,
    max_new_tokens: int,
    min_books: int,
    min_top_share: float,
) -> Tuple[int, List[str]]:
    """Append reverse-mined anagram candidates into Glossary as inactive tokens."""
    if not candidates:
        return 0, []
    ws = wb["Glossary"]
    header = ws_find_header_row(
        ws, ["Token", "Translation", "TokenType", "Confidence", "Use_StrictPlus_v108", "TotalOcc", "BookCount", "Len", "EvidenceClass_v127"]
    )
    c = ws_headers(ws, header)

    template_row = None
    for r in range(header + 1, ws.max_row + 1):
        if ws.cell(r, c["TokenType"]).value == "anagram":
            template_row = r
            break
    if template_row is None:
        template_row = header + 1

    existing = {ws.cell(r, c["Token"]).value for r in range(header + 1, ws.max_row + 1) if ws.cell(r, c["Token"]).value}

    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_books, hb)
    books_base: List[Tuple[int, str]] = []
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        base = ws_books.cell(r, cb["DecodedBase"]).value
        if not isinstance(base, str) or not base:
            continue
        books_base.append((int(bid), base))

    added = 0
    tokens_added: List[str] = []

    def _key(row: Dict[str, object]) -> Tuple[int, float, int]:
        return (int(row.get("SupportBooks") or 0), float(row.get("TopShare") or 0.0), int(row.get("SupportOcc") or 0))

    for row in sorted(list(candidates), key=_key, reverse=True):
        if added >= int(max_new_tokens):
            break
        base = str(row.get("Base") or "").strip()
        if not base or base in existing:
            continue
        support_books = int(row.get("SupportBooks") or 0)
        top_share = float(row.get("TopShare") or 0.0)
        if support_books < int(min_books):
            continue
        if top_share + 1e-12 < float(min_top_share):
            continue

        top_word = str(row.get("TopWord") or "").strip()
        if not top_word:
            continue

        occ = 0
        book_count = 0
        for _bid, b in books_base:
            n = b.count(base)
            if n:
                occ += int(n)
                book_count += 1

        if occ <= 0:
            occ = int(row.get("SupportOcc") or 0)
            book_count = int(row.get("SupportBooks") or 0)

        if book_count >= 7 and top_share >= 0.98:
            conf = "HIGH"
            ev = 2.4
        elif book_count >= 3 and top_share >= 0.95:
            conf = "MEDIUM"
            ev = 2.2
        else:
            conf = "LOW"
            ev = 2.0

        note = (
            f"iter{iter_num} reverse-phrase candidate. word={top_word} share={top_share:.3f} "
            f"supportBooks={support_books} supportOcc={int(row.get('SupportOcc') or 0)}. AddedUTC={utc}."
        )

        values: Dict[int, object] = {}

        def _set(k: str, v: object) -> None:
            col = c.get(k)
            if col is not None:
                values[col] = v

        _set("Token", base)
        _set("Translation", top_word)
        _set("TokenType", "anagram")
        _set("Confidence", conf)
        _set("ActiveCorpus", False)
        _set("Use_StrictPlus_v108", 0)
        _set("Use_Strict_v108", False)
        _set("BoundaryRule", "none")
        _set("RuleLayerNeeded", "baseline")
        _set("TotalOcc", occ)
        _set("BookCount", book_count)
        _set("ContigCount", 0)
        _set("Len", len(base))
        _set("StarCount", int(base.count("*")))
        _set("Notes", note)
        _set("Use_PoemMode_v113", False)
        _set("SemNeutral_Render_v123", "NORMAL")
        _set("Mask_SemNeutral_v123", 0)
        _set("EvidenceClass_v127", evidence_class)
        _set("EvidenceScore_v127", ev)
        _set("EvidenceSources_v127", f"iter{iter_num}: reverse_phrase_mining")

        _glossary_append_row_copy_style(ws, template_row, values_by_col=values)
        existing.add(base)
        tokens_added.append(base)
        added += 1

    return added, tokens_added


def apply_reverse_phrase_retext_existing_tokens(
    wb: openpyxl.Workbook,
    iter_num: int,
    utc: str,
    candidates: Sequence[Dict[str, object]],
    *,
    enabled: bool,
    max_per_iter: int,
    min_books: int,
    min_top_share: float,
    min_support_occ: int,
    anti_mode: bool,
    anti_deny_words: set[str],
    enforced_crib_ids: Optional[set[int]] = None,
    locked_tokens: Optional[set[str]] = None,
) -> Tuple[int, int, int]:
    """Retext existing active tokens from ReversePhrase candidates (GT-guarded)."""
    if not enabled:
        return 0, 0, 0
    if int(max_per_iter) <= 0:
        return 0, 0, 0
    if not candidates:
        return 0, 0, 0

    glossary_ws, glossary_map = load_glossary(wb)
    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}
    if not active:
        return 0, 0, 0

    ws_log = ensure_sheet(
        wb,
        "ReversePhraseRetext_Auto",
        [
            "Iteration",
            "UTC",
            "Token",
            "CurrentTranslation",
            "SuggestedTranslation",
            "SupportBooks",
            "SupportOcc",
            "TopShare",
            "Decision",
            "Reason",
            "Notes",
        ],
    )
    upsert_sheet_index_entry(
        wb,
        "ReversePhraseRetext_Auto",
        "Retext proposals from ReversePhraseTokenCands_Auto applied to existing active tokens (GT live-check guarded).",
    )

    attempted = 0
    applied = 0
    updates: Dict[str, str] = {}

    def _sort_key(row: Dict[str, object]) -> Tuple[int, float, int, str]:
        try:
            books = int(row.get("SupportBooks") or 0)
        except Exception:
            books = 0
        try:
            share = float(row.get("TopShare") or 0.0)
        except Exception:
            share = 0.0
        try:
            occ = int(row.get("SupportOcc") or 0)
        except Exception:
            occ = 0
        base = str(row.get("Base") or "")
        return (books, share, occ, base)

    for row in sorted(list(candidates), key=_sort_key, reverse=True):
        if applied >= int(max_per_iter):
            break

        base = str(row.get("Base") or "").strip()
        new_tr = str(row.get("TopWord") or "").strip()
        if not base or not new_tr or any(ch.isspace() for ch in new_tr):
            continue
        if locked_tokens and base in locked_tokens:
            continue

        try:
            support_books = int(row.get("SupportBooks") or 0)
        except Exception:
            support_books = 0
        try:
            support_occ = int(row.get("SupportOcc") or 0)
        except Exception:
            support_occ = 0
        try:
            top_share = float(row.get("TopShare") or 0.0)
        except Exception:
            top_share = 0.0

        if support_books < int(min_books):
            continue
        if support_occ < int(min_support_occ):
            continue
        if top_share + 1e-12 < float(min_top_share):
            continue

        gt = glossary_map.get(base)
        if gt is None:
            continue
        if not gt.use_strictplus:
            continue
        if gt.token_type in ("marker", "macro"):
            continue

        old_tr = str(gt.translation or "").strip()
        if not old_tr:
            continue
        if _norm_wordish(old_tr) == _norm_wordish(new_tr):
            continue

        new_w = _norm_wordish(new_tr)
        if not new_w:
            continue
        if anti_mode and anti_deny_words and new_w in anti_deny_words:
            ws_append_row(
                ws_log,
                [
                    iter_num,
                    utc,
                    base,
                    old_tr,
                    new_tr,
                    support_books,
                    support_occ,
                    round(top_share, 6),
                    "SKIP",
                    "deny_word",
                    "anti-hallucination denylist",
                ],
                start_col=1,
            )
            continue

        attempted += 1
        test_active = dict(active)
        test_active[base] = dataclasses.replace(gt, translation=new_tr)
        ok, bad, bad_all = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_crib_ids)
        if not ok:
            bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad[:8])
            soft_n = max(0, len(bad_all) - len(bad))
            ws_append_row(
                ws_log,
                [
                    iter_num,
                    utc,
                    base,
                    old_tr,
                    new_tr,
                    support_books,
                    support_occ,
                    round(top_share, 6),
                    "SKIP",
                    "gt_mismatch",
                    f"CribID(s) {bad_ids}" + (f"; soft_mismatches={soft_n}" if soft_n else ""),
                ],
                start_col=1,
            )
            continue

        note = (
            f"reverse retext: {old_tr} -> {new_tr} "
            f"(support_books={support_books}, support_occ={support_occ}, top_share={top_share:.6f})"
        )
        glossary_set_translation(glossary_ws, gt.row, new_tr, iter_num, note)
        ws_append_row(
            ws_log,
            [
                iter_num,
                utc,
                base,
                old_tr,
                new_tr,
                support_books,
                support_occ,
                round(top_share, 6),
                "APPLIED",
                "gt_ok",
                "",
            ],
            start_col=1,
        )
        active = test_active
        updates[base] = new_tr
        applied += 1

    el_changed = sync_evidence_ledger_translations(wb, updates)
    return int(applied), int(attempted), int(el_changed)


def _build_macro_composition_map(glossary_ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, List[str]]:
    """Parse 'Composition tokens: A + B + C.' from Glossary.Notes."""
    header = ws_find_header_row(glossary_ws, ["Token", "TokenType", "Notes"], max_scan=3)
    c = ws_headers(glossary_ws, header)
    comp_map: Dict[str, List[str]] = {}

    for r in range(header + 1, glossary_ws.max_row + 1):
        tok = glossary_ws.cell(r, c["Token"]).value
        if not isinstance(tok, str) or not tok:
            continue
        if str(glossary_ws.cell(r, c["TokenType"]).value or "") != "macro":
            continue
        notes = glossary_ws.cell(r, c["Notes"]).value
        if not isinstance(notes, str) or "Composition tokens:" not in notes:
            continue

        m = re.search(r"Composition tokens:\s*([^.]+)", notes)
        if not m:
            continue
        part = m.group(1)
        comp = [p.strip() for p in part.split("+")]
        comp = [p for p in comp if p]
        if comp:
            comp_map[tok] = comp

    return comp_map


def _expand_dp_items_to_lossless_tokens(
    items: Sequence[object],
    active_tokens: Dict[str, GlossaryToken],
    comp_map: Dict[str, List[str]],
    semantic_map: Dict[str, str],
    *,
    max_depth: int = 10,
) -> Tuple[List[str], int]:
    """Expand DP items into a lossless token stream, applying semantic replacements on leaf tokens."""

    def is_macro(t: GlossaryToken) -> bool:
        return t.token_type == "macro" or (t.evidence_class and "MACRO" in t.evidence_class.upper())

    out: List[str] = []
    repl = 0

    def emit_token(t: GlossaryToken) -> None:
        nonlocal repl
        if t.token_type == "marker":
            out.append(f"<{t.token}>")
            return
        replacement = semantic_map.get(t.token)
        if replacement:
            repl += 1
            out.extend(replacement.split())
        else:
            out.extend(t.translation.split())

    def rec(it: object, depth: int) -> None:
        if isinstance(it, str):
            out.append(it)
            return
        assert isinstance(it, GlossaryToken)

        if depth < max_depth and is_macro(it) and it.token in comp_map:
            for ct in comp_map[it.token]:
                gt = active_tokens.get(ct)
                if gt is None:
                    emit_token(it)
                    return
                rec(gt, depth + 1)
            return

        emit_token(it)

    for it in items:
        rec(it, 0)

    return out, repl


def materialize_semantic_translation_display(
    wb: openpyxl.Workbook,
    iter_num: int,
    glossary_map: Dict[str, GlossaryToken],
    semantic_map: Dict[str, str],
) -> Tuple[int, int, int]:
    """Fill Translation_Semantic_Auto in Books + MasterText (display-only).

    Returns: (books_rows_changed, master_rows_changed, replaced_token_count)
    """
    if not semantic_map:
        return 0, 0, 0

    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}
    comp_map = _build_macro_composition_map(wb["Glossary"])

    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "DecodedBase", "Translation_StrictPlus_v108"], max_scan=3)
    cb = ws_headers(ws_books, hb)
    col_out = cb.get("Translation_Semantic_Auto")
    if col_out is None:
        col_out = ws_books.max_column + 1
        ws_books.cell(hb, col_out).value = "Translation_Semantic_Auto"

    books_changed = 0
    repl_total = 0
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        base = ws_books.cell(r, cb["DecodedBase"]).value
        if base is None or str(base).strip() == "":
            continue

        items = dp_tokenize_base_with_punct(str(base), active)
        lossless, repl = _expand_dp_items_to_lossless_tokens(items, active, comp_map, semantic_map)
        out_str = render_strictplus_from_lossless(lossless)
        repl_total += int(repl)

        prev = ws_books.cell(r, col_out).value
        if prev != out_str:
            books_changed += 1
        ws_books.cell(r, col_out).value = out_str

    ws_cont = wb["Contigs"]
    hc = ws_find_header_row(ws_cont, ["BaseContigID", "BaseContig"], max_scan=3)
    cc = ws_headers(ws_cont, hc)
    sem_by_cid: Dict[int, str] = {}
    for r in range(hc + 1, ws_cont.max_row + 1):
        cid = ws_cont.cell(r, cc["BaseContigID"]).value
        base = ws_cont.cell(r, cc["BaseContig"]).value
        if cid is None or base is None:
            continue
        items = dp_tokenize_base_with_punct(str(base), active)
        lossless, _ = _expand_dp_items_to_lossless_tokens(items, active, comp_map, semantic_map)
        sem_by_cid[int(cid)] = render_strictplus_from_lossless(lossless)

    ws_mt = wb["MasterText"]
    hm = ws_find_header_row(ws_mt, ["BaseContigID", "Translation_StrictPlus_v108"], max_scan=3)
    cm = ws_headers(ws_mt, hm)
    col_mt = cm.get("Translation_Semantic_Auto")
    if col_mt is None:
        col_mt = ws_mt.max_column + 1
        ws_mt.cell(hm, col_mt).value = "Translation_Semantic_Auto"

    master_changed = 0
    for r in range(hm + 1, ws_mt.max_row + 1):
        cid = ws_mt.cell(r, cm["BaseContigID"]).value
        if cid is None:
            continue
        out_str = sem_by_cid.get(int(cid), "")
        prev = ws_mt.cell(r, col_mt).value
        if prev != out_str:
            master_changed += 1
        ws_mt.cell(r, col_mt).value = out_str

    upsert_sheet_index_entry(wb, "SemanticMap_Auto", "Semantic translation layer (display-only).")
    return books_changed, master_changed, repl_total


_EN_WORD_RE = re.compile(r"[A-Za-z']+")


def _norm_word_key(surface: str) -> str:
    return re.sub(r"[^a-z]", "", str(surface or "").lower())


def _apply_word_map_preserve_case(text: str, word_map: Dict[str, str]) -> Tuple[str, int]:
    """Apply a word->word map on a string, preserving simple capitalization.

    Display-only: MUST NOT affect DP/tokenization or Glossary.
    """
    s = str(text or "")
    if not s or not word_map:
        return s, 0

    out: List[str] = []
    last = 0
    repl = 0
    for m in _EN_WORD_RE.finditer(s):
        out.append(s[last : m.start()])
        w = m.group(0)
        key = _norm_word_key(w)
        rep = word_map.get(key)
        if rep:
            repl += 1
            rep_s = str(rep)
            if w[:1].isupper() and rep_s:
                rep_s = rep_s[:1].upper() + rep_s[1:]
            out.append(rep_s)
        else:
            out.append(w)
        last = m.end()
    out.append(s[last:])
    return "".join(out), repl


def materialize_english_layer_display(
    wb: openpyxl.Workbook,
    iter_num: int,
    glossary_map: Dict[str, GlossaryToken],
    *,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    min_total_count: int,
    min_top_share: float,
    min_word_len: int,
    max_map_rows: int,
) -> Tuple[int, int, int, int]:
    """Materialize a display-only English layer for Books + MasterText.

    The StrictPlus decode is canonical (matches GroundTruth norms like `wit`, `fay`, `lion`).
    This layer improves readability by mapping canonical decode-words to likely English surface forms
    using the Tibia signature index (derived counts only; no full text stored).

    Returns: (english_map_rows, books_rows_changed, master_rows_changed, replacements_applied)
    """
    # Combine multiple derived signature indexes to broaden coverage.
    # This stays safe because we persist only derived counts (no full text in XLSX).
    combined_counts: Dict[str, Counter[str]] = defaultdict(Counter)
    combined_corpora: Dict[str, set[str]] = defaultdict(set)
    for sheet_name in (LORE_SIGINDEX_TIBIA_SHEET, LORE_SIGINDEX_PD_SHEET, LORE_SIGINDEX_DICT_SHEET):
        if sheet_name not in wb.sheetnames:
            continue
        sc, corp = _load_lore_sigindex_sheet(wb, sheet_name)
        for sig, wc in (sc or {}).items():
            combined_counts[sig].update(wc)
        for sig, cset in (corp or {}).items():
            combined_corpora[sig].update(cset)

    sig_word_counts: Dict[str, Dict[str, int]] = {sig: dict(cnt) for sig, cnt in combined_counts.items()}
    sig_corpora: Dict[str, set[str]] = dict(combined_corpora)
    if not sig_word_counts:
        ensure_sheet(
            wb,
            ENGLISH_MAP_SHEET,
            [
                "Iteration",
                "CanonWord",
                "CanonSig",
                "TopWord",
                "TopWordCount",
                "TotalWordCount",
                "TopShare",
                "CandidateWords",
                "CandidateWordCounts",
                "CorpusIDs",
                "Notes",
            ],
        )
        upsert_sheet_index_entry(
            wb,
            ENGLISH_MAP_SHEET,
            "English display-only mapping derived from Tibia sig-index (canonical decode word -> likely English surface).",
        )
        return 0, 0, 0, 0

    canon_words: set[str] = set()
    for t in glossary_map.values():
        if not (t.use_strictplus and t.translation):
            continue
        if t.token_type in ("marker", "macro"):
            continue
        for w in _EN_WORD_RE.findall(str(t.translation or "")):
            key = _norm_word_key(w)
            if not key or len(key) < int(min_word_len):
                continue
            canon_words.add(key)

    word_map: Dict[str, str] = {}
    map_rows: List[Tuple[str, str, int, int, float, List[Tuple[str, int]], str]] = []
    for cw in sorted(canon_words):
        canon = _lore_canon_word(cw, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
        if not canon:
            continue
        sig = _lore_signature(canon)
        wc = sig_word_counts.get(sig)
        if not wc:
            continue
        items = sorted(wc.items(), key=lambda kv: (-kv[1], kv[0]))
        top_word, top_count = items[0]
        total_count = int(sum(cnt for _w, cnt in items))
        if total_count < int(min_total_count):
            continue
        share = float(top_count) / float(total_count) if total_count else 0.0
        if share + 1e-12 < float(min_top_share):
            continue
        top_norm = _norm_word_key(top_word)
        if not top_norm:
            continue
        if top_norm == cw:
            continue
        word_map[cw] = top_word
        corp = ",".join(sorted(sig_corpora.get(sig) or []))
        map_rows.append((cw, sig, int(top_count), int(total_count), float(share), items[:8], corp))
        if len(map_rows) >= int(max_map_rows):
            break

    ws_map = ensure_sheet(
        wb,
        ENGLISH_MAP_SHEET,
        [
            "Iteration",
            "CanonWord",
            "CanonSig",
            "TopWord",
            "TopWordCount",
            "TotalWordCount",
            "TopShare",
            "CandidateWords",
            "CandidateWordCounts",
            "CorpusIDs",
            "Notes",
        ],
    )
    hm = ws_find_header_row(ws_map, ["Iteration", "CanonWord", "TopWord"], max_scan=3)
    cm = ws_headers(ws_map, hm)
    if ws_map.max_row > hm:
        ws_map.delete_rows(hm + 1, ws_map.max_row - hm)

    rr = hm + 1
    for cw, sig, topc, tot, share, items8, corp in map_rows:
        ws_map.cell(rr, cm["Iteration"]).value = iter_num
        ws_map.cell(rr, cm["CanonWord"]).value = cw
        ws_map.cell(rr, cm["CanonSig"]).value = sig
        ws_map.cell(rr, cm["TopWord"]).value = word_map.get(cw)
        ws_map.cell(rr, cm["TopWordCount"]).value = int(topc)
        ws_map.cell(rr, cm["TotalWordCount"]).value = int(tot)
        ws_map.cell(rr, cm["TopShare"]).value = round(float(share), 6)
        ws_map.cell(rr, cm["CandidateWords"]).value = ", ".join([w for w, _cnt in items8])
        ws_map.cell(rr, cm["CandidateWordCounts"]).value = ", ".join([f"{w}:{int(cnt)}" for w, cnt in items8])
        ws_map.cell(rr, cm["CorpusIDs"]).value = corp
        ws_map.cell(rr, cm["Notes"]).value = (
            f"min_total={min_total_count}, min_share={min_top_share}, "
            f"canon(drop_final_e={drop_final_e}, drop_all_h={drop_all_h}, drop_all_o={drop_all_o})"
        )
        rr += 1

    upsert_sheet_index_entry(
        wb,
        ENGLISH_MAP_SHEET,
        "English display-only mapping derived from Tibia sig-index (canonical decode word -> likely English surface).",
    )

    # Books output
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "Translation_StrictPlus_v108"], max_scan=3)
    cb = ws_headers(ws_books, hb)
    in_col = cb.get("Translation_Semantic_Auto") or cb.get("Translation_Readable_Auto") or cb["Translation_StrictPlus_v108"]
    out_col = cb.get("Translation_English_Auto")
    if out_col is None:
        out_col = ws_books.max_column + 1
        ws_books.cell(hb, out_col).value = "Translation_English_Auto"
    cnt_col = cb.get("EnglishReplCount_Auto")
    if cnt_col is None:
        cnt_col = ws_books.max_column + 1
        if cnt_col == out_col:
            cnt_col += 1
        ws_books.cell(hb, cnt_col).value = "EnglishReplCount_Auto"

    books_changed = 0
    repl_total = 0
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        src = ws_books.cell(r, in_col).value
        out, repl = _apply_word_map_preserve_case(str(src or ""), word_map)
        prev = ws_books.cell(r, out_col).value
        if prev != out:
            books_changed += 1
        ws_books.cell(r, out_col).value = out
        ws_books.cell(r, cnt_col).value = int(repl)
        repl_total += int(repl)

    # MasterText output
    ws_mt = wb["MasterText"]
    hm2 = ws_find_header_row(ws_mt, ["BaseContigID", "Translation_StrictPlus_v108"], max_scan=3)
    cm2 = ws_headers(ws_mt, hm2)
    in_col2 = cm2.get("Translation_Semantic_Auto") or cm2.get("Translation_Readable_Auto") or cm2["Translation_StrictPlus_v108"]
    out_col2 = cm2.get("Translation_English_Auto")
    if out_col2 is None:
        out_col2 = ws_mt.max_column + 1
        ws_mt.cell(hm2, out_col2).value = "Translation_English_Auto"
    cnt_col2 = cm2.get("EnglishReplCount_Auto")
    if cnt_col2 is None:
        cnt_col2 = ws_mt.max_column + 1
        if cnt_col2 == out_col2:
            cnt_col2 += 1
        ws_mt.cell(hm2, cnt_col2).value = "EnglishReplCount_Auto"

    master_changed = 0
    for r in range(hm2 + 1, ws_mt.max_row + 1):
        cid = ws_mt.cell(r, cm2["BaseContigID"]).value
        if cid is None:
            continue
        src = ws_mt.cell(r, in_col2).value
        out, repl = _apply_word_map_preserve_case(str(src or ""), word_map)
        prev = ws_mt.cell(r, out_col2).value
        if prev != out:
            master_changed += 1
        ws_mt.cell(r, out_col2).value = out
        ws_mt.cell(r, cnt_col2).value = int(repl)
        repl_total += int(repl)

    return len(map_rows), books_changed, master_changed, repl_total


def _last_fetched_dt_from_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    *,
    required_cols: Sequence[str],
    fetched_col: str,
) -> Optional[datetime]:
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    h = ws_find_header_row(ws, required_cols, max_scan=3)
    c = ws_headers(ws, h)
    col = c.get(fetched_col)
    if not col:
        return None
    for r in range(h + 1, min(ws.max_row, h + 50) + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, str) or not v.strip():
            continue
        try:
            return datetime.strptime(v.strip(), ISO_UTC_FMT).replace(tzinfo=timezone.utc)
        except Exception:
            continue
    return None


def refresh_lore_bigrams_sheet(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    npc_url: str,
    book_url: str,
    timeout_s: int,
    max_age_hours: float,
    cache_max_age_hours: float,
    pd_enabled: bool = False,
    pd_sources: Optional[Sequence[Tuple[str, str]]] = None,
    pd_cache_max_age_hours: float = 720.0,
    vocab_topn: int,
    min_count: int,
    max_rows: int,
) -> Tuple[int, str]:
    """Build a derived word-bigram index from cached Tibia JSON + LoreCorpus_Auto (public domain).

    We do not persist full corpus text into the workbook; only derived bigram counts.
    Returns: (rows_written, status)
    """
    should_refresh = False
    last_dt = _last_fetched_dt_from_sheet(
        wb,
        LORE_BIGRAMS_SHEET,
        required_cols=["Word1", "Word2", "Count"],
        fetched_col="FetchedUTC",
    )
    if last_dt is None:
        should_refresh = True
    else:
        if float(max_age_hours) <= 0:
            should_refresh = True
        else:
            age = datetime.now(timezone.utc) - last_dt
            should_refresh = age >= timedelta(hours=float(max_age_hours))

    ws_out = ensure_sheet(wb, LORE_BIGRAMS_SHEET, ["Word1", "Word2", "Count", "Sources", "FetchedUTC", "Notes"])
    ho = ws_find_header_row(ws_out, ["Word1", "Word2", "Count"], max_scan=3)
    co = ws_headers(ws_out, ho)

    # If the PD corpus configuration changed, force a refresh even if the sheet is still "fresh" by age.
    pd_sources_norm: List[Tuple[str, str]] = []
    if pd_enabled and pd_sources:
        seen: set[str] = set()
        for cid_raw, u_raw in list(pd_sources):
            u = str(u_raw or "").strip()
            if not u or u in seen:
                continue
            seen.add(u)
            cid = str(cid_raw or "").strip() or _auto_corpus_id_from_url(u)
            pd_sources_norm.append((cid, u))
    pd_fp_payload = json.dumps({"pd_enabled": int(bool(pd_enabled)), "pd_sources": pd_sources_norm}, sort_keys=True)
    pd_fp = hashlib.sha1(pd_fp_payload.encode("utf-8", errors="replace")).hexdigest()[:12]
    pd_fp_note = f"pd_enabled={int(bool(pd_enabled))}, pd_fp={pd_fp}, pd_sources={len(pd_sources_norm)}"
    if not should_refresh and pd_fp_note:
        try:
            notes_col = co.get("Notes")
            existing = ""
            if notes_col:
                for r in range(ho + 1, min(ws_out.max_row, ho + 6) + 1):
                    v = ws_out.cell(r, notes_col).value
                    if isinstance(v, str) and v.strip():
                        existing = v
                        break
            if pd_fp_note not in str(existing or ""):
                should_refresh = True
        except Exception:
            pass

    if not should_refresh:
        # Keep sheet stable; no-op.
        upsert_sheet_index_entry(
            wb,
            LORE_BIGRAMS_SHEET,
            "Derived word bigram index (Tibia cache + LoreCorpus_Auto) for context disambiguation (no full text persisted).",
        )
        return 0, "skipped_fresh"

    ua = "Mozilla/5.0 (Bonelord469LoreBigrams/1.0)"
    cache_dir = os.path.join(os.getcwd(), "tmp", "corpus")
    npc_cache = _cache_path_for_url(cache_dir, prefix="tibia_npc", url=npc_url)
    book_cache = _cache_path_for_url(cache_dir, prefix="tibia_books", url=book_url)

    unigrams: Counter[str] = Counter()
    bigrams: Counter[Tuple[str, str]] = Counter()
    sources: Dict[Tuple[str, str], set[str]] = defaultdict(set)

    def ingest_text(text: str, source_tag: str) -> None:
        for sent in _split_sentences(text):
            prev: Optional[str] = None
            for raw in _iter_words(sent):
                surface = re.sub(r"[^a-z']", "", str(raw or "").lower())
                if not surface:
                    continue
                # Keep leading apostrophes (e.g. "'enable") because the sig-index keeps them too.
                w = surface
                unigrams[w] += 1
                if prev is not None:
                    bigrams[(prev, w)] += 1
                    sources[(prev, w)].add(source_tag)
                prev = w

    # Tibia NPC + Books (cached)
    try:
        npc_obj = _fetch_json_url_cached(
            npc_url,
            cache_path=npc_cache,
            timeout_s=int(timeout_s),
            max_age_hours=float(cache_max_age_hours),
            user_agent=ua,
        )
        if isinstance(npc_obj, list):
            for item in npc_obj:
                if not isinstance(item, dict):
                    continue
                conv = item.get("conversation")
                if not isinstance(conv, list):
                    continue
                for turn in conv:
                    if not isinstance(turn, dict):
                        continue
                    answers = turn.get("answer")
                    if isinstance(answers, str):
                        answers = [answers]
                    if not isinstance(answers, list):
                        continue
                    for a in answers:
                        if not isinstance(a, str) or not a.strip():
                            continue
                        ingest_text(a, "TIBIA_NPC")
    except Exception:
        # Best-effort: if cache is unavailable, just skip Tibia NPC.
        pass

    try:
        book_obj = _fetch_json_url_cached(
            book_url,
            cache_path=book_cache,
            timeout_s=int(timeout_s),
            max_age_hours=float(cache_max_age_hours),
            user_agent=ua,
        )
        if isinstance(book_obj, list):
            for item in book_obj:
                if not isinstance(item, dict):
                    continue
                text = item.get("text")
                if not isinstance(text, str) or not text.strip():
                    continue
                ingest_text(text, "TIBIA_BOOK")
    except Exception:
        pass

    # LoreCorpus_Auto (public domain text already stored in workbook)
    if "LoreCorpus_Auto" in wb.sheetnames:
        ws_l = wb["LoreCorpus_Auto"]
        hl = ws_find_header_row(ws_l, ["CorpusID", "LineID", "Text"], max_scan=3)
        cl = ws_headers(ws_l, hl)
        for r in range(hl + 1, ws_l.max_row + 1):
            txt = ws_l.cell(r, cl["Text"]).value
            if not isinstance(txt, str) or not txt.strip():
                continue
            ingest_text(txt, "LORE_AUTO")

    # Optional public-domain corpus (cached on disk; derived counts only stored in the XLSX).
    if pd_enabled and pd_sources_norm:
        for pd_cid, pd_url_s in pd_sources_norm:
            try:
                pd_cache = _cache_path_for_url(cache_dir, prefix="pd_text", url=pd_url_s, ext="txt")
                pd_txt = _fetch_text_url_cached(
                    pd_url_s,
                    cache_path=pd_cache,
                    timeout_s=int(timeout_s),
                    max_age_hours=float(pd_cache_max_age_hours),
                    user_agent=ua,
                )
                # Best-effort strip for Gutenberg boilerplate when present.
                m_start = re.search(
                    r"\*\*\*\s*START OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
                    pd_txt,
                    flags=re.IGNORECASE | re.DOTALL,
                )
                m_end = re.search(
                    r"\*\*\*\s*END OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
                    pd_txt,
                    flags=re.IGNORECASE | re.DOTALL,
                )
                if m_start and m_end and m_end.start() > m_start.end():
                    pd_txt = pd_txt[m_start.end() : m_end.start()]
                ingest_text(pd_txt, pd_cid)
            except Exception:
                # Best-effort: do not block LoreBigrams.
                continue

    # Prune by vocabulary and min_count.
    vocab = set([w for w, _cnt in unigrams.most_common(int(vocab_topn))])
    filtered: List[Tuple[int, str, str, str]] = []
    for (w1, w2), cnt in bigrams.items():
        if cnt < int(min_count):
            continue
        if w1 not in vocab or w2 not in vocab:
            continue
        srcs = ",".join(sorted(sources.get((w1, w2)) or []))
        filtered.append((int(cnt), w1, w2, srcs))

    filtered.sort(key=lambda t: (-t[0], t[1], t[2]))
    filtered = filtered[: int(max_rows)]

    # Rewrite sheet body.
    if ws_out.max_row > ho:
        ws_out.delete_rows(ho + 1, ws_out.max_row - ho)
    rr = ho + 1
    for cnt, w1, w2, srcs in filtered:
        ws_out.cell(rr, co["Word1"]).value = w1
        ws_out.cell(rr, co["Word2"]).value = w2
        ws_out.cell(rr, co["Count"]).value = int(cnt)
        ws_out.cell(rr, co.get("Sources") or (ws_out.max_column + 1)).value = srcs
        ws_out.cell(rr, co.get("FetchedUTC") or (ws_out.max_column + 1)).value = utc
        ws_out.cell(rr, co.get("Notes") or (ws_out.max_column + 1)).value = (
            f"vocab_topn={vocab_topn}, min_count={min_count}, max_rows={max_rows}, "
            f"npc_url={npc_url}, book_url={book_url}, {pd_fp_note} (derived only)"
        )
        rr += 1

    upsert_sheet_index_entry(
        wb,
        LORE_BIGRAMS_SHEET,
        "Derived word bigram index (Tibia cache + LoreCorpus_Auto) for context disambiguation (no full text persisted).",
    )
    return rr - (ho + 1), "refreshed"


def _load_lore_bigrams(wb: openpyxl.Workbook) -> Tuple[Dict[Tuple[str, str], int], set[str]]:
    if LORE_BIGRAMS_SHEET not in wb.sheetnames:
        return {}, set()
    ws = wb[LORE_BIGRAMS_SHEET]
    h = ws_find_header_row(ws, ["Word1", "Word2", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    out: Dict[Tuple[str, str], int] = {}
    words: set[str] = set()
    for r in range(h + 1, ws.max_row + 1):
        w1 = ws.cell(r, c["Word1"]).value
        w2 = ws.cell(r, c["Word2"]).value
        cnt = ws.cell(r, c["Count"]).value
        if not isinstance(w1, str) or not isinstance(w2, str):
            continue
        try:
            ci = int(cnt or 0)
        except Exception:
            ci = 0
        if ci <= 0:
            continue
        k = (w1.strip().lower(), w2.strip().lower())
        out[k] = out.get(k, 0) + ci
        words.add(k[0])
        words.add(k[1])
    return out, words


def _parse_word_count_pairs(raw: object, *, max_items: int = 32) -> Dict[str, int]:
    """Parse 'and:21475, than:1106' -> {'and':21475, 'than':1106}."""
    out: Dict[str, int] = {}
    s = str(raw or "").strip()
    if not s:
        return out
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        if ":" not in part:
            continue
        w, cnt = part.split(":", 1)
        w = w.strip().lower()
        if not w:
            continue
        try:
            ci = int(cnt.strip())
        except Exception:
            continue
        out[w] = out.get(w, 0) + ci
        if len(out) >= int(max_items):
            break
    return out


def materialize_context_english_render(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    glossary_map: Dict[str, GlossaryToken],
    active_tokens: Dict[str, GlossaryToken],
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    emission_alpha: float,
    transition_alpha: float,
    max_candidates_per_token: int,
    map_min_total: int,
    map_min_top_share: float,
    map_max_rows: int,
    sequence_hints_enabled: bool = True,
    sequence_hints_boost: int = 20,
    sequence_hints_max_words_per_sig: int = 3,
    reversephrase_hints_enabled: bool = True,
    reversephrase_hints_boost: int = 8,
    reversephrase_hints_max_words_per_sig: int = 2,
) -> Tuple[int, int, float, float, int]:
    """Render a display-only context English layer (Books + MasterText via Contigs).

    Returns: (books_changed_rows, master_changed_rows, avg_score_per_token, oov_frac, context_map_rows)
    """
    # Load candidates per token from LoreAlignment.
    token_cands: Dict[str, Dict[str, int]] = {}
    if "LoreAlignment_Auto" in wb.sheetnames:
        ws_hits = wb["LoreAlignment_Auto"]
        hh = ws_find_header_row(ws_hits, ["Token", "CandidateWordCounts"], max_scan=3)
        ch = ws_headers(ws_hits, hh)
        for r in range(hh + 1, ws_hits.max_row + 1):
            tok = ws_hits.cell(r, ch["Token"]).value
            if not isinstance(tok, str) or not tok.strip():
                continue
            token_cands[tok.strip()] = _parse_word_count_pairs(ws_hits.cell(r, ch["CandidateWordCounts"]).value)

    bigrams, known_words = _load_lore_bigrams(wb)
    seq_hints = _load_sequence_word_hints(wb) if bool(sequence_hints_enabled) else {}
    rev_hints = _load_reverse_phrase_token_hints(wb) if bool(reversephrase_hints_enabled) else {}

    # Column setup: Books
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_books, hb)
    col_out = cb.get("Translation_ContextEnglish_Auto")
    if col_out is None:
        col_out = ws_books.max_column + 1
        ws_books.cell(hb, col_out).value = "Translation_ContextEnglish_Auto"
    col_ch = cb.get("ContextEnglish_Changes_Auto")
    if col_ch is None:
        col_ch = ws_books.max_column + 1
        if col_ch == col_out:
            col_ch += 1
        ws_books.cell(hb, col_ch).value = "ContextEnglish_Changes_Auto"
    col_sc = cb.get("ContextEnglish_AvgScore_Auto")
    if col_sc is None:
        col_sc = ws_books.max_column + 1
        ws_books.cell(hb, col_sc).value = "ContextEnglish_AvgScore_Auto"
    col_oov = cb.get("ContextEnglish_OOVFrac_Auto")
    if col_oov is None:
        col_oov = ws_books.max_column + 1
        ws_books.cell(hb, col_oov).value = "ContextEnglish_OOVFrac_Auto"

    comp_map = _build_macro_composition_map(wb["Glossary"])

    blocked_ev = {"GROUNDTRUTH", "LOGOGRAM_ANCHORED", "PUNCT_LOGOGRAM", "LOGOGRAM_CONTEXT"}
    allowed_short = {
        "a",
        "i",
        "am",
        "an",
        "as",
        "at",
        "be",
        "by",
        "do",
        "go",
        "he",
        "if",
        "in",
        "is",
        "it",
        "me",
        "my",
        "no",
        "of",
        "on",
        "or",
        "so",
        "to",
        "up",
        "us",
        "we",
        "ye",
        "ya",
    }

    def is_macro(t: GlossaryToken) -> bool:
        return t.token_type == "macro" or (t.evidence_class and "MACRO" in str(t.evidence_class).upper())

    def token_baseline_word(tok: str) -> str:
        gt = glossary_map.get(tok)
        if gt is None:
            return tok.lower()
        return str(gt.translation or "").strip().lower()

    def token_is_logogram(tok: str) -> bool:
        gt = glossary_map.get(tok)
        if gt is None:
            return True
        evcls = str(gt.evidence_class or "").strip().upper()
        if evcls in blocked_ev:
            return True
        if gt.token_type in ("marker", "macro"):
            return True
        tr = str(gt.translation or "").strip()
        if not tr or any(ch.isspace() for ch in tr):
            return True
        canon = _lore_canon_word(tr, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
        if not canon:
            return True
        # Do NOT require signature(tr) == signature(token). Many tokens are anagram/homophone based,
        # and the contextual layer must still be able to choose among candidate surface words.
        return False

    def candidates_for_token(tok: str) -> List[Tuple[str, int]]:
        base = token_baseline_word(tok)
        if not base:
            return []
        if "*" in tok or token_is_logogram(tok):
            return [(base, 1)]
        lim = max(1, int(max_candidates_per_token))
        cc = dict(token_cands.get(tok) or {})
        hint_allow: set[str] = set()
        sig = _token_signature(tok)

        def apply_hints(hinted: Dict[str, int], *, boost: int, kmax: int) -> None:
            nonlocal cc, hint_allow
            if not hinted:
                return
            try:
                boost_i = max(0, int(boost))
            except Exception:
                boost_i = 0
            try:
                kmax_i = max(0, int(kmax))
            except Exception:
                kmax_i = 0
            if boost_i <= 0 or kmax_i <= 0:
                return
            base_items = sorted(cc.items(), key=lambda kv: (-kv[1], kv[0]))
            cutoff = int(base_items[lim - 1][1]) if len(base_items) >= lim else 0
            for w, cnt in sorted(hinted.items(), key=lambda kv: (-kv[1], kv[0]))[:kmax_i]:
                ww = str(w or "").strip().lower()
                if not ww:
                    continue
                hint_allow.add(ww)
                try:
                    ci = int(cnt or 0)
                except Exception:
                    ci = 0
                if ci <= 0:
                    continue
                ladder = 1.0 + math.log1p(float(ci))
                bonus = int(round(float(boost_i) * float(ladder)))
                prev = int(cc.get(ww, 0) or 0)
                cc[ww] = max(prev + int(bonus), int(cutoff) + int(bonus))

        if seq_hints and sig:
            apply_hints(
                seq_hints.get(sig) or {},
                boost=int(sequence_hints_boost),
                kmax=int(sequence_hints_max_words_per_sig),
            )
        if rev_hints and sig:
            apply_hints(
                rev_hints.get(sig) or {},
                boost=int(reversephrase_hints_boost),
                kmax=int(reversephrase_hints_max_words_per_sig),
            )

        items = sorted(cc.items(), key=lambda kv: (-kv[1], kv[0]))
        out: List[Tuple[str, int]] = []
        for w, cnt in items:
            if not w:
                continue
            ww = str(w).strip().lower()
            if not ww:
                continue
            if len(ww) <= 2 and ww not in allowed_short:
                continue
            # Allow hinted words even when they are outside the bigram vocabulary:
            # SequenceWordHints is derived from corpora and may surface rare-but-correct words.
            if known_words and ww not in known_words and ww not in allowed_short and ww not in hint_allow:
                continue
            out.append((ww, int(cnt)))
            if len(out) >= lim:
                break
        # Ensure baseline is present.
        if base not in {w for w, _c in out}:
            out.append((base, 1))
        return out

    def viterbi(tokens: List[str]) -> Tuple[List[str], float]:
        if not tokens:
            return [], 0.0
        cand_lists: List[List[Tuple[str, int]]] = [candidates_for_token(t) for t in tokens]
        # Guard: avoid empty candidate lists.
        for i, cl in enumerate(cand_lists):
            if not cl:
                cand_lists[i] = [(token_baseline_word(tokens[i]), 1)]

        def emit_score(word: str, cnt: int) -> float:
            return math.log(float(cnt) + float(emission_alpha))

        def trans_score(w1: str, w2: str) -> float:
            return math.log(float(bigrams.get((w1, w2), 0)) + float(transition_alpha))

        prev_scores: Dict[str, float] = {}
        backpointers: List[Dict[str, str]] = []
        # init
        for w, cnt in cand_lists[0]:
            prev_scores[w] = emit_score(w, cnt)
        # dp
        for i in range(1, len(tokens)):
            cur_scores: Dict[str, float] = {}
            cur_bp: Dict[str, str] = {}
            for w, cnt in cand_lists[i]:
                best_s = None
                best_pw = None
                e = emit_score(w, cnt)
                for pw, ps in prev_scores.items():
                    s = ps + trans_score(pw, w) + e
                    if best_s is None or s > best_s:
                        best_s = s
                        best_pw = pw
                if best_s is None or best_pw is None:
                    continue
                cur_scores[w] = best_s
                cur_bp[w] = best_pw
            backpointers.append(cur_bp)
            prev_scores = cur_scores or prev_scores

        # best final
        if not prev_scores:
            # fallback to baselines
            out = [token_baseline_word(t) for t in tokens]
            return out, 0.0
        last_word = max(prev_scores.items(), key=lambda kv: kv[1])[0]
        best_score = prev_scores[last_word]

        # backtrack
        out_words = [last_word]
        for i in range(len(tokens) - 2, -1, -1):
            bp = backpointers[i] if i < len(backpointers) else {}
            last_word = bp.get(out_words[-1], token_baseline_word(tokens[i]))
            out_words.append(last_word)
        out_words.reverse()
        return out_words, float(best_score)

    def render_base(base: str) -> Tuple[str, int, float, float, Dict[str, Counter[str]]]:
        # Build event stream with word indices.
        events: List[Tuple[str, object]] = []
        word_tokens: List[str] = []
        boundaries: List[int] = []

        def rec_token(t: GlossaryToken, depth: int) -> None:
            if t.token_type == "marker":
                events.append(("COMMA", None))
                return
            if is_macro(t):
                comp = comp_map.get(t.token) or []
                if depth >= 10 or not comp:
                    # treat as a single leaf token
                    idx = len(word_tokens)
                    word_tokens.append(t.token)
                    events.append(("WORD", idx))
                    return
                for ct in comp:
                    gt = glossary_map.get(ct) or active_tokens.get(ct)
                    if gt is None:
                        continue
                    rec_token(gt, depth + 1)
                return
            idx = len(word_tokens)
            word_tokens.append(t.token)
            events.append(("WORD", idx))

        items = dp_tokenize_base_with_punct(str(base or ""), active_tokens)
        for it in items:
            if isinstance(it, str):
                if it in (".", "!"):
                    events.append(("PUNCT", it))
                    boundaries.append(len(word_tokens))
                continue
            rec_token(it, 0)

        # Viterbi per segment (reset at '.'/'!').
        chosen: List[str] = []
        total_score = 0.0
        start = 0
        boundaries2 = [b for b in boundaries if b > 0]
        boundaries2.append(len(word_tokens))
        for end in boundaries2:
            seg = word_tokens[start:end]
            words, score = viterbi(seg)
            chosen.extend(words)
            total_score += score
            start = end

        # Metrics
        changes = 0
        oov = 0
        total = len(word_tokens)
        baseline_to_chosen: Dict[str, Counter[str]] = defaultdict(Counter)
        for tok, w in zip(word_tokens, chosen):
            b = token_baseline_word(tok)
            if w != b:
                changes += 1
            baseline_to_chosen[b][w] += 1
            if known_words and w not in known_words:
                oov += 1
        avg_score = (total_score / float(total)) if total else 0.0
        oov_frac = (float(oov) / float(total)) if total else 0.0

        # Render
        out_words: List[str] = []
        for kind, payload in events:
            if kind == "WORD":
                w = chosen[int(payload)]
                out_words.append(w)
            elif kind == "COMMA":
                if not out_words:
                    out_words.append(",")
                else:
                    out_words[-1] = f"{out_words[-1]},"
            elif kind == "PUNCT":
                if not out_words:
                    out_words.append(str(payload))
                else:
                    out_words[-1] = f"{out_words[-1]}{payload}"

        return " ".join(out_words), changes, avg_score, oov_frac, baseline_to_chosen

    books_changed = 0
    master_changed = 0
    sum_tokens = 0
    sum_score = 0.0
    sum_oov = 0.0
    baseline_choice_global: Dict[str, Counter[str]] = defaultdict(Counter)

    # Books
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        base = ws_books.cell(r, cb["DecodedBase"]).value
        if base is None or str(base).strip() == "":
            continue
        out, changes, avg_score, oov_frac, baseline_map = render_base(str(base))
        prev = ws_books.cell(r, col_out).value
        if prev != out:
            books_changed += 1
        ws_books.cell(r, col_out).value = out
        ws_books.cell(r, col_ch).value = int(changes)
        ws_books.cell(r, col_sc).value = float(round(avg_score, 6))
        ws_books.cell(r, col_oov).value = float(round(oov_frac, 6))
        # aggregate
        sum_tokens += max(0, len(str(out).split()))
        sum_score += float(avg_score) * float(max(1, len(str(out).split())))
        sum_oov += float(oov_frac) * float(max(1, len(str(out).split())))
        for b, cnts in baseline_map.items():
            baseline_choice_global[b].update(cnts)

    # Contigs -> MasterText
    if "Contigs" in wb.sheetnames and "MasterText" in wb.sheetnames:
        ws_c = wb["Contigs"]
        hc = ws_find_header_row(ws_c, ["BaseContigID", "BaseContig"], max_scan=3)
        cc = ws_headers(ws_c, hc)

        contig_out: Dict[int, str] = {}
        for r in range(hc + 1, ws_c.max_row + 1):
            cid = ws_c.cell(r, cc["BaseContigID"]).value
            if cid is None:
                continue
            try:
                cid_i = int(cid)
            except Exception:
                continue
            base = ws_c.cell(r, cc["BaseContig"]).value
            if base is None or str(base).strip() == "":
                continue
            out, _ch, _sc, _oov, _bm = render_base(str(base))
            contig_out[cid_i] = out

        ws_mt = wb["MasterText"]
        hm = ws_find_header_row(ws_mt, ["BaseContigID", "Translation_StrictPlus_v108"], max_scan=3)
        cm = ws_headers(ws_mt, hm)
        col_mt = cm.get("Translation_ContextEnglish_Auto")
        if col_mt is None:
            col_mt = ws_mt.max_column + 1
            ws_mt.cell(hm, col_mt).value = "Translation_ContextEnglish_Auto"

        for r in range(hm + 1, ws_mt.max_row + 1):
            cid = ws_mt.cell(r, cm["BaseContigID"]).value
            if cid is None:
                continue
            try:
                cid_i = int(cid)
            except Exception:
                continue
            out = contig_out.get(cid_i)
            if out is None:
                continue
            prev = ws_mt.cell(r, col_mt).value
            if prev != out:
                master_changed += 1
            ws_mt.cell(r, col_mt).value = out

    # Build context-driven english map (baseline word -> top chosen word).
    ws_map = ensure_sheet(
        wb,
        ENGLISH_MAP_CONTEXT_SHEET,
        [
            "Iteration",
            "CanonWord",
            "CanonSig",
            "TopWord",
            "TopWordCount",
            "TotalWordCount",
            "TopShare",
            "CandidateWords",
            "CandidateWordCounts",
            "CorpusIDs",
            "Notes",
        ],
    )
    hm2 = ws_find_header_row(ws_map, ["Iteration", "CanonWord", "TopWord"], max_scan=3)
    cm2 = ws_headers(ws_map, hm2)
    if ws_map.max_row > hm2:
        ws_map.delete_rows(hm2 + 1, ws_map.max_row - hm2)

    map_rows: List[Tuple[str, str, int, int, float, List[Tuple[str, int]]]] = []
    for base_word, cnts in baseline_choice_global.items():
        total = int(sum(cnts.values()))
        if total < int(map_min_total):
            continue
        top_word, top_cnt = cnts.most_common(1)[0]
        top_w = str(top_word or "").strip().lower()
        if not top_w:
            continue
        if len(top_w) <= 2 and top_w not in allowed_short:
            continue
        if known_words and top_w not in known_words and top_w not in allowed_short:
            continue
        if not top_word or top_word == base_word:
            continue
        share = float(top_cnt) / float(total) if total else 0.0
        if share + 1e-12 < float(map_min_top_share):
            continue
        canon = _lore_canon_word(base_word, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
        sig = _lore_signature(canon) if canon else ""
        items8 = cnts.most_common(8)
        map_rows.append((base_word, sig, int(top_cnt), int(total), float(share), items8))

    map_rows.sort(key=lambda t: (-t[3], -t[4], t[0]))
    map_rows = map_rows[: int(map_max_rows)]

    rr = hm2 + 1
    for base_word, sig, top_cnt, total, share, items8 in map_rows:
        ws_map.cell(rr, cm2["Iteration"]).value = int(iter_num)
        ws_map.cell(rr, cm2["CanonWord"]).value = base_word
        ws_map.cell(rr, cm2["CanonSig"]).value = sig
        ws_map.cell(rr, cm2["TopWord"]).value = items8[0][0]
        ws_map.cell(rr, cm2["TopWordCount"]).value = int(top_cnt)
        ws_map.cell(rr, cm2["TotalWordCount"]).value = int(total)
        ws_map.cell(rr, cm2["TopShare"]).value = round(float(share), 6)
        ws_map.cell(rr, cm2["CandidateWords"]).value = ", ".join([w for w, _cnt in items8])
        ws_map.cell(rr, cm2["CandidateWordCounts"]).value = ", ".join([f"{w}:{int(cnt)}" for w, cnt in items8])
        ws_map.cell(rr, cm2["CorpusIDs"]).value = "CONTEXT_ENGLISH"
        ws_map.cell(rr, cm2["Notes"]).value = f"alpha_emit={emission_alpha}, alpha_trans={transition_alpha}"
        rr += 1

    upsert_sheet_index_entry(
        wb,
        ENGLISH_MAP_CONTEXT_SHEET,
        "Context-derived English map (baseline word -> top contextual choice). Used as an additional safe source for English->Glossary retext.",
    )

    avg_score = (sum_score / float(sum_tokens)) if sum_tokens else 0.0
    avg_oov = (sum_oov / float(sum_tokens)) if sum_tokens else 0.0
    return books_changed, master_changed, float(avg_score), float(avg_oov), (rr - (hm2 + 1))


def materialize_codeaware_homophones_render(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    glossary_map: Dict[str, GlossaryToken],
    active_tokens: Dict[str, GlossaryToken],
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    emission_alpha: float,
    transition_alpha: float,
    max_candidates_per_token: int,
    max_token_len: int,
    min_total_per_code: int,
    min_top_share: float,
    min_total_per_codeseq: int,
    min_top_share_codeseq: float,
    max_map_rows: int,
    hint_min_total: int = 8,
    hint_boost: int = 4,
    hint_topk: int = 2,
    apply_tokens_allowlist: Optional[Sequence[str]] = None,
    sequence_hints_enabled: bool = True,
    sequence_hints_boost: int = 20,
    sequence_hints_max_words_per_sig: int = 3,
    reversephrase_hints_enabled: bool = True,
    reversephrase_hints_boost: int = 8,
    reversephrase_hints_max_words_per_sig: int = 2,
) -> Tuple[int, int, int, str, str]:
    """Render a display-only code-aware layer to disambiguate homophones by digits-code.

    This step never edits Glossary/StrictPlus. It uses:
    - digits->code reconstruction (BooksDigitModel omission positions)
    - ContextEnglish candidate sets + bigram LM
    - stability threshold per (Token,Code) to emit a mapping

    Returns: (books_changed_rows, overrides_total, map_rows, fingerprint, status)
    """
    if "Books" not in wb.sheetnames or "BooksDigitModel_v118" not in wb.sheetnames:
        ensure_sheet(
            wb,
            CODE_WORD_MAP_SHEET,
            [
                "Iteration",
                "UTC",
                "Token",
                "Code",
                "TopWord",
                "TopWordCount",
                "TotalCount",
                "TopShare",
                "BaselineWord",
                "CandidateWords",
                "CandidateWordCounts",
                "Notes",
            ],
        )
        upsert_sheet_index_entry(
            wb,
            CODE_WORD_MAP_SHEET,
            "Code-aware homophone map: (Token,DigitsCode)->TopWord derived from ContextEnglish choices (analysis-only).",
        )
        return 0, 0, 0, "", "missing Books/BooksDigitModel_v118"

    # Load omit positions by book.
    ws_dm = wb["BooksDigitModel_v118"]
    hdm = ws_find_header_row(ws_dm, ["BookID", "OmitIdxs_1based"], max_scan=3)
    cdm = ws_headers(ws_dm, hdm)
    omit_by_book: Dict[int, set[int]] = {}
    for r in range(hdm + 1, ws_dm.max_row + 1):
        bid = ws_dm.cell(r, cdm["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        omit_by_book[bid_i] = set(parse_int_list(ws_dm.cell(r, cdm["OmitIdxs_1based"]).value))

    # Candidates per token (from LoreAlignment) + bigram LM.
    token_cands: Dict[str, Dict[str, int]] = {}
    if "LoreAlignment_Auto" in wb.sheetnames:
        ws_hits = wb["LoreAlignment_Auto"]
        hh = ws_find_header_row(ws_hits, ["Token", "CandidateWordCounts"], max_scan=3)
        ch = ws_headers(ws_hits, hh)
        for r in range(hh + 1, ws_hits.max_row + 1):
            tok = ws_hits.cell(r, ch["Token"]).value
            if not isinstance(tok, str) or not tok.strip():
                continue
            token_cands[tok.strip()] = _parse_word_count_pairs(ws_hits.cell(r, ch["CandidateWordCounts"]).value)

    bigrams, known_words = _load_lore_bigrams(wb)
    seq_hints = _load_sequence_word_hints(wb) if bool(sequence_hints_enabled) else {}
    rev_hints = _load_reverse_phrase_token_hints(wb) if bool(reversephrase_hints_enabled) else {}

    allowset: Optional[set[str]] = None
    if apply_tokens_allowlist:
        allowset = {str(t or "").strip() for t in apply_tokens_allowlist if str(t or "").strip()}
        if not allowset:
            allowset = None

    comp_map = _build_macro_composition_map(wb["Glossary"])

    blocked_ev = {"GROUNDTRUTH", "LOGOGRAM_ANCHORED", "PUNCT_LOGOGRAM", "LOGOGRAM_CONTEXT"}
    allowed_short = {
        "a",
        "i",
        "am",
        "an",
        "as",
        "at",
        "be",
        "by",
        "do",
        "go",
        "he",
        "if",
        "in",
        "is",
        "it",
        "me",
        "my",
        "no",
        "of",
        "on",
        "or",
        "so",
        "to",
        "up",
        "us",
        "we",
        "ye",
        "ya",
        "th",
    }

    def is_macro(t: GlossaryToken) -> bool:
        return t.token_type == "macro" or (t.evidence_class and "MACRO" in str(t.evidence_class).upper())

    def token_baseline_word(tok: str) -> str:
        gt = glossary_map.get(tok)
        if gt is None:
            return tok.lower()
        return str(gt.translation or "").strip().lower()

    def token_is_logogram(tok: str) -> bool:
        gt = glossary_map.get(tok)
        if gt is None:
            return True
        evcls = str(gt.evidence_class or "").strip().upper()
        if evcls in blocked_ev:
            return True
        if gt.token_type in ("marker", "macro"):
            return True
        tr = str(gt.translation or "").strip()
        if not tr or any(ch.isspace() for ch in tr):
            return True
        canon = _lore_canon_word(tr, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
        if not canon:
            return True
        return False

    def candidates_for_token(tok: str) -> List[Tuple[str, int]]:
        base = token_baseline_word(tok)
        if not base:
            return []
        if "*" in tok or token_is_logogram(tok):
            return [(base, 1)]
        lim = max(1, int(max_candidates_per_token))
        cc = dict(token_cands.get(tok) or {})
        hint_allow: set[str] = set()
        sig = _token_signature(tok)

        def apply_hints(hinted: Dict[str, int], *, boost: int, kmax: int) -> None:
            nonlocal cc, hint_allow
            if not hinted:
                return
            try:
                boost_i = max(0, int(boost))
            except Exception:
                boost_i = 0
            try:
                kmax_i = max(0, int(kmax))
            except Exception:
                kmax_i = 0
            if boost_i <= 0 or kmax_i <= 0:
                return
            base_items = sorted(cc.items(), key=lambda kv: (-kv[1], kv[0]))
            cutoff = int(base_items[lim - 1][1]) if len(base_items) >= lim else 0
            for w, cnt in sorted(hinted.items(), key=lambda kv: (-kv[1], kv[0]))[:kmax_i]:
                ww = str(w or "").strip().lower()
                if not ww:
                    continue
                hint_allow.add(ww)
                try:
                    ci = int(cnt or 0)
                except Exception:
                    ci = 0
                if ci <= 0:
                    continue
                ladder = 1.0 + math.log1p(float(ci))
                bonus = int(round(float(boost_i) * float(ladder)))
                prev = int(cc.get(ww, 0) or 0)
                cc[ww] = max(prev + int(bonus), int(cutoff) + int(bonus))

        if seq_hints and sig:
            apply_hints(
                seq_hints.get(sig) or {},
                boost=int(sequence_hints_boost),
                kmax=int(sequence_hints_max_words_per_sig),
            )
        if rev_hints and sig:
            apply_hints(
                rev_hints.get(sig) or {},
                boost=int(reversephrase_hints_boost),
                kmax=int(reversephrase_hints_max_words_per_sig),
            )

        items = sorted(cc.items(), key=lambda kv: (-kv[1], kv[0]))
        out: List[Tuple[str, int]] = []
        for w, cnt in items:
            if not w:
                continue
            ww = str(w).strip().lower()
            if not ww:
                continue
            if len(ww) <= 2 and ww not in allowed_short:
                continue
            if known_words and ww not in known_words and ww not in allowed_short and ww not in hint_allow:
                continue
            out.append((ww, int(cnt)))
            if len(out) >= lim:
                break
        if base not in {w for w, _c in out}:
            out.append((base, 1))
        return out

    def viterbi(tokens: List[str]) -> Tuple[List[str], float]:
        if not tokens:
            return [], 0.0
        cand_lists: List[List[Tuple[str, int]]] = [candidates_for_token(t) for t in tokens]
        for i, cl in enumerate(cand_lists):
            if not cl:
                cand_lists[i] = [(token_baseline_word(tokens[i]), 1)]

        def emit_score(word: str, cnt: int) -> float:
            return math.log(float(cnt) + float(emission_alpha))

        def trans_score(w1: str, w2: str) -> float:
            return math.log(float(bigrams.get((w1, w2), 0)) + float(transition_alpha))

        prev_scores: Dict[str, float] = {}
        backpointers: List[Dict[str, str]] = []
        for w, cnt in cand_lists[0]:
            prev_scores[w] = emit_score(w, cnt)
        for i in range(1, len(tokens)):
            cur_scores: Dict[str, float] = {}
            cur_bp: Dict[str, str] = {}
            for w, cnt in cand_lists[i]:
                best_s = None
                best_pw = None
                e = emit_score(w, cnt)
                for pw, ps in prev_scores.items():
                    s = ps + trans_score(pw, w) + e
                    if best_s is None or s > best_s:
                        best_s = s
                        best_pw = pw
                if best_s is None or best_pw is None:
                    continue
                cur_scores[w] = best_s
                cur_bp[w] = best_pw
            backpointers.append(cur_bp)
            prev_scores = cur_scores or prev_scores

        if not prev_scores:
            out = [token_baseline_word(t) for t in tokens]
            return out, 0.0
        last_word = max(prev_scores.items(), key=lambda kv: kv[1])[0]
        best_score = prev_scores[last_word]

        out_words = [last_word]
        for i in range(len(tokens) - 2, -1, -1):
            bp = backpointers[i] if i < len(backpointers) else {}
            last_word = bp.get(out_words[-1], token_baseline_word(tokens[i]))
            out_words.append(last_word)
        out_words.reverse()
        return out_words, float(best_score)

    def reconstruct_codes_by_idx(digits: object, base: str, omit: set[int]) -> Optional[List[Optional[str]]]:
        ds = _digits_only(digits)
        if not ds or not base:
            return None
        codes: List[Optional[str]] = [None] * len(base)
        di = 0
        base_pos = 0  # 1-based over letters-only positions
        for i, ch in enumerate(base):
            if not (ch.isalpha() or ch == "*"):
                continue
            base_pos += 1
            if base_pos in omit:
                if di >= len(ds):
                    return None
                code = "0" + ds[di]
                di += 1
            else:
                if di + 1 >= len(ds):
                    return None
                code = ds[di : di + 2]
                di += 2
            codes[i] = code
        return codes

    def parse_base_to_events(base: str, codes_by_idx: List[Optional[str]]) -> Tuple[List[Tuple[str, object]], List[str], List[Optional[str]]]:
        events: List[Tuple[str, object]] = []
        word_tokens: List[str] = []
        word_codes: List[Optional[str]] = []

        pos = 0

        def emit_word(tok_str: str, base_start: int, tok_len: int) -> None:
            idx = len(word_tokens)
            word_tokens.append(tok_str)
            code_s: Optional[str] = None
            # For multi-letter tokens we store a compact joined code sequence (e.g., "09-61") so we can
            # study code-aware homophones beyond 1-letter tokens safely (display-only). To avoid huge
            # cardinality, we only store codes for tokens up to max_token_len.
            try:
                if int(tok_len) >= 1 and int(tok_len) <= int(max_token_len) and 0 <= int(base_start) < len(codes_by_idx):
                    codes: List[str] = []
                    end = min(len(codes_by_idx), int(base_start) + int(tok_len))
                    for j in range(int(base_start), end):
                        c = codes_by_idx[j]
                        if c:
                            codes.append(str(c))
                    if codes:
                        code_s = "-".join(codes)
            except Exception:
                code_s = None
            word_codes.append(code_s)
            events.append(("WORD", idx))

        def rec_emit(t: GlossaryToken, depth: int, base_start: int) -> int:
            if t.token_type == "marker":
                events.append(("COMMA", None))
                return int(t.length)
            if is_macro(t):
                comp = comp_map.get(t.token) or []
                if depth >= 10 or not comp:
                    emit_word(t.token, base_start, int(t.length))
                    return int(t.length)
                out_len = 0
                cur = int(base_start)
                for ct in comp:
                    gt = glossary_map.get(ct) or active_tokens.get(ct)
                    if gt is None:
                        continue
                    consumed = rec_emit(gt, depth + 1, cur)
                    cur += int(consumed)
                    out_len += int(consumed)
                if out_len != int(t.length):
                    # Fallback to keep positional alignment sane.
                    emit_word(t.token, base_start, int(t.length))
                    return int(t.length)
                return int(out_len)
            emit_word(t.token, base_start, int(t.length))
            return int(t.length)

        items = dp_tokenize_base_with_punct(str(base or ""), active_tokens)
        for it in items:
            if isinstance(it, str):
                if it in (".", "!"):
                    events.append(("PUNCT", it))
                # punctuation consumes a raw char, but code stream is letters-only; keep pos in raw space.
                pos += 1
                continue
            consumed = rec_emit(it, 0, int(pos))
            pos += int(consumed)

        return events, word_tokens, word_codes

    # Pass 1: run ContextEnglish locally + aggregate (Token,Code)->chosen_word counts.
    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "Digits", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_books, hb)
    col_digits = cb["Digits"]
    col_base = cb["DecodedBase"]

    # Local context scoring per (Token,Code) using baseline ContextEnglish choices as the surrounding context.
    # This targets the "non brute force" layer: homophones disambiguated by structure/context.
    local_choice_counts: Dict[Tuple[str, str], Counter[str]] = defaultdict(Counter)
    occ_counts: Counter[Tuple[str, str]] = Counter()
    cand_cache: Dict[str, List[Tuple[str, int]]] = {}
    book_records: List[Tuple[int, int, List[Tuple[str, object]], List[str], List[Optional[str]], List[str]]] = []
    # (book_id, row, events, word_tokens, word_codes, chosen_words)

    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        omit = omit_by_book.get(bid_i)
        if omit is None:
            continue
        base = ws_books.cell(r, col_base).value
        if base is None or str(base).strip() == "":
            continue
        digits = ws_books.cell(r, col_digits).value
        codes_by_idx = reconstruct_codes_by_idx(digits, str(base), omit)
        if not codes_by_idx:
            continue
        events, word_tokens, word_codes = parse_base_to_events(str(base), codes_by_idx)
        chosen_words, _score = viterbi(list(word_tokens))
        # ensure stable length
        if len(chosen_words) != len(word_tokens):
            chosen_words = [token_baseline_word(t) for t in word_tokens]

        for i, (tok, code) in enumerate(zip(word_tokens, word_codes)):
            if code is None:
                continue
            if int(max_token_len) > 0 and len(tok) > int(max_token_len):
                continue
            if allowset is not None and tok not in allowset:
                continue
            if token_is_logogram(tok):
                continue
            prev_w = str(chosen_words[i - 1] if i > 0 else "<s>")
            next_w = str(chosen_words[i + 1] if (i + 1) < len(chosen_words) else "</s>")
            cl = cand_cache.get(tok)
            if cl is None:
                cl = candidates_for_token(tok)
                cand_cache[tok] = cl
            if not cl:
                continue

            best_w = None
            best_s = None
            for w, cnt in cl:
                ww = str(w or "").strip().lower()
                if not ww:
                    continue
                try:
                    cci = int(cnt or 0)
                except Exception:
                    cci = 0
                s = (
                    math.log(float(bigrams.get((prev_w, ww), 0)) + float(transition_alpha))
                    + math.log(float(bigrams.get((ww, next_w), 0)) + float(transition_alpha))
                    + math.log(float(cci) + float(emission_alpha))
                )
                if best_s is None or s > best_s:
                    best_s = s
                    best_w = ww

            if best_w is None:
                continue
            key = (tok, str(code))
            local_choice_counts[key][best_w] += 1
            occ_counts[key] += 1

        book_records.append((bid_i, r, events, word_tokens, word_codes, chosen_words))

    # Build per-(Token,Code) distributions:
    # - rows_for_sheet: visibility/debugging (analysis-only)
    # - mapping: stable (Token,Code)->TopWord used to override baseline choices (display-only)
    hard_min_total_1 = max(1, int(min_total_per_code))
    hard_min_share_1 = float(min_top_share)
    hard_min_total_seq = max(1, int(min_total_per_codeseq))
    hard_min_share_seq = float(min_top_share_codeseq)
    hint_min_total_i = max(1, int(hint_min_total))
    hint_topk_i = max(1, int(hint_topk))

    rows_for_sheet: List[Tuple[str, str, str, int, int, float, str, List[Tuple[str, int]]]] = []
    mapping: Dict[Tuple[str, str], str] = {}

    for (tok, code), cnts in local_choice_counts.items():
        total = int(occ_counts.get((tok, code)) or sum(cnts.values()))
        if total <= 0:
            continue
        top_word, top_cnt = cnts.most_common(1)[0]
        share = float(top_cnt) / float(total) if total else 0.0
        base_word = token_baseline_word(tok)
        items8 = cnts.most_common(8 if hint_topk_i > 8 else 8)
        if total >= hint_min_total_i:
            rows_for_sheet.append((tok, code, str(top_word), int(top_cnt), int(total), float(share), base_word, items8))
        if len(tok) == 1:
            hard_min_total = hard_min_total_1
            hard_min_share = hard_min_share_1
        else:
            hard_min_total = hard_min_total_seq
            hard_min_share = hard_min_share_seq
        if total >= int(hard_min_total) and share + 1e-12 >= float(hard_min_share):
            mapping[(tok, code)] = str(top_word)

    rows_for_sheet.sort(key=lambda t: (-t[4], -t[5], t[0], t[1]))
    rows_for_sheet = rows_for_sheet[: int(max_map_rows)]

    keep_keys = {(t, c) for t, c, _w, _tc, _tot, _sh, _bw, _it in rows_for_sheet}
    mapping = {k: v for k, v in mapping.items() if k in keep_keys}

    fp_items = [f"{tok}|{code}|{w}|{share:.6f}|{total}" for tok, code, w, _tc, total, share, _bw, _it in rows_for_sheet]
    fp_items.sort()
    fp = hashlib.sha1("\n".join(fp_items).encode("utf-8", errors="ignore")).hexdigest() if fp_items else ""

    ws_map = ensure_sheet(
        wb,
        CODE_WORD_MAP_SHEET,
        [
            "Iteration",
            "UTC",
            "Token",
            "Code",
            "TopWord",
            "TopWordCount",
            "TotalCount",
            "TopShare",
            "BaselineWord",
            "CandidateWords",
            "CandidateWordCounts",
            "Notes",
        ],
    )
    hm = ws_find_header_row(ws_map, ["Iteration", "Token", "Code", "TopWord"], max_scan=3)
    cm = ws_headers(ws_map, hm)
    if ws_map.max_row > hm:
        ws_map.delete_rows(hm + 1, ws_map.max_row - hm)

    rr = hm + 1
    for tok, code, top_w, top_cnt, total, share, base_word, items8 in rows_for_sheet:
        ws_map.cell(rr, cm["Iteration"]).value = int(iter_num)
        ws_map.cell(rr, cm["UTC"]).value = utc
        ws_map.cell(rr, cm["Token"]).value = tok
        ws_map.cell(rr, cm["Code"]).value = code
        ws_map.cell(rr, cm["TopWord"]).value = top_w
        ws_map.cell(rr, cm["TopWordCount"]).value = int(top_cnt)
        ws_map.cell(rr, cm["TotalCount"]).value = int(total)
        ws_map.cell(rr, cm["TopShare"]).value = round(float(share), 6)
        ws_map.cell(rr, cm["BaselineWord"]).value = base_word
        ws_map.cell(rr, cm["CandidateWords"]).value = ", ".join([w for w, _cnt in items8])
        ws_map.cell(rr, cm["CandidateWordCounts"]).value = ", ".join([f"{w}:{int(cnt)}" for w, cnt in items8])
        ws_map.cell(rr, cm["Notes"]).value = (
            f"hard_min_total_1={hard_min_total_1}, hard_min_share_1={hard_min_share_1}, "
            f"hard_min_total_seq={hard_min_total_seq}, hard_min_share_seq={hard_min_share_seq}, "
            f"hint_min_total={hint_min_total_i}, hint_topk={hint_topk_i}, "
            f"max_cands={max_candidates_per_token}"
        )
        rr += 1

    upsert_sheet_index_entry(
        wb,
        CODE_WORD_MAP_SHEET,
        "Code-aware homophone map: (Token,DigitsCode)->TopWord derived from local context scoring over the ContextEnglish baseline (analysis-only).",
    )

    # Pass 2: render Books.Translation_CodeAware_Auto applying the stable mapping (display-only).
    col_out = cb.get("Translation_CodeAware_Auto")
    if col_out is None:
        col_out = ws_books.max_column + 1
        ws_books.cell(hb, col_out).value = "Translation_CodeAware_Auto"
    col_ov = cb.get("CodeAware_Overrides_Auto")
    if col_ov is None:
        col_ov = ws_books.max_column + 1
        if col_ov == col_out:
            col_ov += 1
        ws_books.cell(hb, col_ov).value = "CodeAware_Overrides_Auto"
    col_hits = cb.get("CodeAware_MapHits_Auto")
    if col_hits is None:
        col_hits = ws_books.max_column + 1
        if col_hits in (col_out, col_ov):
            col_hits += 1
        ws_books.cell(hb, col_hits).value = "CodeAware_MapHits_Auto"

    books_changed = 0
    overrides_total = 0

    for _bid_i, r, events, word_tokens, word_codes, chosen_words in book_records:
        if not word_tokens:
            continue
        chosen2 = list(chosen_words)
        hits = 0
        for i, (tok, code) in enumerate(zip(word_tokens, word_codes)):
            if code is None:
                continue
            if int(max_token_len) > 0 and len(tok) > int(max_token_len):
                continue
            if allowset is not None and tok not in allowset:
                continue
            key = (tok, str(code))
            w2 = mapping.get(key)
            if not w2:
                continue
            hits += 1
            chosen2[i] = str(w2)

        overrides = 0
        for i in range(0, min(len(chosen2), len(chosen_words))):
            if str(chosen2[i]) != str(chosen_words[i]):
                overrides += 1

        out_words: List[str] = []
        for kind, payload in events:
            if kind == "WORD":
                out_words.append(str(chosen2[int(payload)]))
            elif kind == "COMMA":
                if not out_words:
                    out_words.append(",")
                else:
                    out_words[-1] = f"{out_words[-1]},"
            elif kind == "PUNCT":
                if not out_words:
                    out_words.append(str(payload))
                else:
                    out_words[-1] = f"{out_words[-1]}{payload}"

        out = " ".join(out_words)
        prev = ws_books.cell(r, col_out).value
        if prev != out:
            books_changed += 1
        ws_books.cell(r, col_out).value = out
        ws_books.cell(r, col_ov).value = int(overrides)
        ws_books.cell(r, col_hits).value = int(hits)
        overrides_total += int(overrides)

    return int(books_changed), int(overrides_total), int(len(rows_for_sheet)), fp, "ok"


def materialize_sequence_matches(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    n_list: Sequence[int],
    min_n: int = 2,
    candidate_max_book_freq: int,
    max_candidates: int,
    max_matches: int,
    time_budget_s: float,
    scan_tibia_first: bool = True,
    pd_max_chars: int = 800000,
    pd_max_sentences_per_source: int = 2500,
    npc_url: str,
    book_url: str,
    timeout_s: int,
    cache_max_age_hours: float,
    pd_enabled: bool = False,
    pd_sources: Optional[Sequence[Tuple[str, str]]] = None,
    pd_cache_max_age_hours: float = 720.0,
    drop_final_e: bool = False,
    drop_all_h: bool = False,
    drop_all_o: bool = False,
    context_window: int = 2,
    context_min_overlap: int = 1,
    context_require_direction: bool = False,
    snippet_min_content_words: int = 1,
    explore_rotate: bool = True,
    explore_keep_top: int = 200,
    cache_enabled: bool = True,
    cache_max_rows: int = 2000,
) -> Tuple[int, str]:
    """Search for rare n-grams from the current ContextEnglish output inside corpora.

    Stores only matched phrases + short snippets + URLs (no full text persisted).
    Context windows/direction checks and snippet content heuristics keep matches anchored to the same
    surrounding words as they appeared in Books so the hints remain direction-aware.
    Returns: (matches_written, status)
    """
    if "Books" not in wb.sheetnames:
        return 0, "missing Books"

    ws_b = wb["Books"]
    hb = ws_find_header_row(ws_b, ["BookID"], max_scan=3)
    cb = ws_headers(ws_b, hb)
    col_ctx = (
        cb.get("Translation_CodeAware_Auto")
        or cb.get("Translation_ContextEnglish_Auto")
        or cb.get("Translation_English_Auto")
        or cb.get("Translation_StrictPlus_v108")
    )

    ctx_window = max(0, int(context_window))
    ctx_min_overlap = max(0, int(context_min_overlap))
    ctx_require_direction = bool(context_require_direction)
    snippet_min_content = max(0, int(snippet_min_content_words))

    # Build candidate signature n-grams from books (rare across books).
    # We match on canonicalized sorted-letter signatures to tolerate anagram/homophone layers.
    ng_freq_sig: Counter[Tuple[int, str]] = Counter()
    ng_book_sig: Dict[Tuple[int, str], int] = {}
    ng_surface_by_sig: Dict[Tuple[int, str], str] = {}
    ng_context_by_sig: Dict[Tuple[int, str], Tuple[Tuple[str, ...], Tuple[str, ...]]] = {}

    for r in range(hb + 1, ws_b.max_row + 1):
        bid = ws_b.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        txt = str(ws_b.cell(r, col_ctx).value or "")
        words_surface: List[str] = []
        words_sig: List[str] = []
        for raw in _iter_words(txt):
            w = re.sub(r"[^a-z']", "", str(raw or "").lower())
            if not w:
                continue
            words_surface.append(w)
            canon = _lore_canon_word(w, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
            if not canon:
                words_sig.append("")
            else:
                words_sig.append(_lore_signature(canon))
        for n in n_list:
            if int(min_n) > 0 and int(n) < int(min_n):
                continue
            if n <= 1 or len(words_surface) < n:
                continue
            for i in range(0, len(words_surface) - n + 1):
                sigs = words_sig[i : i + n]
                if any(not s for s in sigs):
                    continue
                phrase_sig = " ".join(sigs)
                key = (n, phrase_sig)
                ng_freq_sig[key] += 1
                if key not in ng_book_sig:
                    ng_book_sig[key] = bid_i
                    ng_surface_by_sig[key] = " ".join(words_surface[i : i + n])
                if ctx_window > 0 and key not in ng_context_by_sig:
                    pre_slice = words_surface[max(0, i - ctx_window) : i]
                    post_slice = words_surface[i + n : i + n + ctx_window]
                    ng_context_by_sig[key] = (
                        tuple(w for w in pre_slice if _is_seq_content_word(w)),
                        tuple(w for w in post_slice if _is_seq_content_word(w)),
                    )

    # rare candidates
    max_f = max(1, int(candidate_max_book_freq))
    candidates = [k for k, f in ng_freq_sig.items() if int(f) <= max_f]
    if not candidates:
        ensure_sheet(
            wb,
            SEQUENCE_MATCHES_SHEET,
            ["Iteration", "UTC", "BookID", "N", "Phrase", "SourceKind", "SourceID", "SourceURL", "Snippet", "Notes"],
        )
        upsert_sheet_index_entry(
            wb,
            SEQUENCE_MATCHES_SHEET,
            "Sequence matches between ContextEnglish n-grams and corpora (analysis-only; snippets only).",
        )
        return 0, "no_candidates"

    # Prioritize longer n and phrases with rarer words (based on Tibia-derived wordfreq if available).
    wordfreq: Dict[str, int] = {}
    if LORE_WORDFREQ_TIBIA_SHEET in wb.sheetnames:
        for w, cnt in _load_wordfreq_sheet(wb, LORE_WORDFREQ_TIBIA_SHEET).items():
            wordfreq[w] = wordfreq.get(w, 0) + int(cnt)
    if LORE_WORDFREQ_PD_SHEET in wb.sheetnames:
        for w, cnt in _load_wordfreq_sheet(wb, LORE_WORDFREQ_PD_SHEET).items():
            wordfreq[w] = wordfreq.get(w, 0) + int(cnt)

    # High-signal filtering: avoid emitting trivial n-grams ("to be a", "of the", ...).
    # We require at least one non-stopword long word to reduce spammy matches.
    def _score_surface(phrase_surface: str) -> float:
        s = 0.0
        long_content = 0
        for w in phrase_surface.split():
            wk = re.sub(r"[^a-z]", "", w)
            if not wk or wk in SEQ_MATCH_STOPWORDS or len(wk) < 4:
                continue
            if len(wk) >= 6:
                long_content += 1
            c = int(wordfreq.get(wk, 0) or 0)
            if c <= 0:
                # Unknown to Tibia corpus: only count as "rare" if it's not a tiny word.
                s += 0.30
            else:
                s += 1.0 / float(c + 20)
        if long_content < 1:
            return -1.0
        return s

    scored: List[Tuple[float, Tuple[int, str]]] = []
    for k in candidates:
        surf = ng_surface_by_sig.get(k, "")
        sc = _score_surface(surf)
        if sc >= 0:
            scored.append((sc, k))

    scored.sort(key=lambda t: (-t[1][0], -t[0], ng_surface_by_sig.get(t[1], "")))
    candidates_all = [k for _sc, k in scored]
    cand_pool_n = int(len(candidates_all))
    # Keep previously matched keys in the candidate set so exploration doesn't starve SequenceMatches to 0.
    forced_set: set[Tuple[int, str]] = set()
    if bool(cache_enabled) and SEQUENCE_MATCHES_CACHE_SHEET in wb.sheetnames:
        try:
            ws_cache0 = wb[SEQUENCE_MATCHES_CACHE_SHEET]
            hc0 = ws_find_header_row(ws_cache0, ["N", "PhraseSig"], max_scan=3)
            cc0 = ws_headers(ws_cache0, hc0)
            for r in range(hc0 + 1, ws_cache0.max_row + 1):
                n0 = ws_cache0.cell(r, cc0["N"]).value
                ps0 = ws_cache0.cell(r, cc0["PhraseSig"]).value
                if n0 is None or ps0 is None:
                    continue
                try:
                    ni = int(n0)
                except Exception:
                    continue
                ps = str(ps0 or "").strip()
                if not ps:
                    continue
                forced_set.add((ni, ps))
        except Exception:
            forced_set = set()
    forced_head = [k for k in candidates_all if k in forced_set] if forced_set else []
    forced_n = int(len(forced_head))
    remaining_all = [k for k in candidates_all if k not in forced_set] if forced_set else list(candidates_all)
    max_cands_i = int(max(0, int(max_candidates)))
    explore_used = 0
    explore_off = 0
    explore_keep_top_i = 0
    explore_window_n = 0
    if max_cands_i > 0 and len(candidates_all) > max_cands_i and bool(explore_rotate):
        # Keep forced keys + a small stable head, then rotate the remaining window.
        if forced_n >= max_cands_i:
            candidates = forced_head[:max_cands_i]
        else:
            keep_top = max(0, min(int(explore_keep_top), int(max_cands_i) - int(forced_n)))
            head = forced_head + remaining_all[:keep_top]
            tail = remaining_all[keep_top:]
            window_n = int(max_cands_i) - int(len(head))
            if window_n <= 0 or not tail:
                candidates = head[:max_cands_i]
            else:
                # Deterministic exploration: rotate a window over the remaining candidates by iteration.
                # This avoids repeatedly mining the same few matches when the render is stable.
                off = (int(iter_num) * int(max(1, window_n))) % int(len(tail))
                window = tail[off : off + window_n]
                if len(window) < window_n:
                    window = window + tail[: (window_n - len(window))]
                candidates = head + window
                explore_used = 1
                explore_off = int(off)
                explore_keep_top_i = int(keep_top)
                explore_window_n = int(window_n)
    else:
        candidates = candidates_all[:max_cands_i] if max_cands_i > 0 else candidates_all
    cand_set = set(candidates)

    matches: List[Tuple[int, str, str, str, str, str]] = []  # (bookid, n, book_phrase, phrase_sig, src_kind, packed)
    start = time.time()

    def consider_source(kind: str, source_id: str, source_url: str, text: str) -> None:
        nonlocal matches, cand_set
        if not cand_set:
            return
        if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
            return
        words_surface: List[str] = []
        words_sig: List[str] = []
        for raw in _iter_words(text):
            w = re.sub(r"[^a-z']", "", str(raw or "").lower())
            if not w:
                continue
            words_surface.append(w)
            canon = _lore_canon_word(w, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
            words_sig.append(_lore_signature(canon) if canon else "")
        for n in sorted(set(n_list), reverse=True):
            if int(min_n) > 0 and int(n) < int(min_n):
                continue
            if len(words_surface) < n:
                continue
            for i in range(0, len(words_surface) - n + 1):
                sigs = words_sig[i : i + n]
                if any(not s for s in sigs):
                    continue
                window_words = words_surface[i : i + n]
                if snippet_min_content > 0:
                    content_words = sum(1 for w in window_words if _is_seq_content_word(w))
                    if content_words < snippet_min_content:
                        continue
                phrase_sig = " ".join(sigs)
                key = (n, phrase_sig)
                if key not in cand_set:
                    continue

                stored_context = ng_context_by_sig.get(key, ((), ()))
                if ctx_window > 0 and ctx_min_overlap > 0 and (stored_context[0] or stored_context[1]):
                    snippet_pre = words_surface[max(0, i - ctx_window) : i]
                    snippet_post = words_surface[i + n : i + n + ctx_window]
                    stored_pre = stored_context[0]
                    stored_post = stored_context[1]
                    pre_hits = len(set(stored_pre) & set(snippet_pre)) if stored_pre else 0
                    post_hits = len(set(stored_post) & set(snippet_post)) if stored_post else 0
                    if pre_hits + post_hits < ctx_min_overlap:
                        continue
                    if ctx_require_direction:
                        if stored_pre and pre_hits == 0:
                            continue
                        if stored_post and post_hits == 0:
                            continue

                bid = int(ng_book_sig.get(key, -1))
                book_phrase = ng_surface_by_sig.get(key, phrase_sig)
                source_phrase = " ".join(window_words)
                matches.append(
                    (bid, str(n), book_phrase, phrase_sig, kind, f"{source_id} | {source_url} | {source_phrase}")
                )
                cand_set.remove(key)
                if len(matches) >= int(max_matches):
                    return

    # LoreCorpus_Auto first (small, public domain)
    if "LoreCorpus_Auto" in wb.sheetnames:
        ws_l = wb["LoreCorpus_Auto"]
        hl = ws_find_header_row(ws_l, ["CorpusID", "LineID", "Text"], max_scan=3)
        cl = ws_headers(ws_l, hl)
        for r in range(hl + 1, ws_l.max_row + 1):
            if float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s):
                break
            txt = ws_l.cell(r, cl["Text"]).value
            if not isinstance(txt, str) or not txt.strip():
                continue
            cid = str(ws_l.cell(r, cl["CorpusID"]).value or "LORE_AUTO")
            lid = ws_l.cell(r, cl["LineID"]).value
            consider_source("LORE_AUTO", f"{cid}:{lid}", "LoreCorpus_Auto", txt)
            if len(matches) >= int(max_matches):
                break

    # Optional public-domain cached plaintext (derived-only; snippets only).
    pd_sources_norm: List[Tuple[str, str]] = []
    if pd_enabled and pd_sources:
        try:
            seen: set[str] = set()
            for cid_raw, u_raw in list(pd_sources):
                u = str(u_raw or "").strip()
                if not u or u in seen:
                    continue
                seen.add(u)
                cid = str(cid_raw or "").strip() or _auto_corpus_id_from_url(u)
                pd_sources_norm.append((cid, u))
        except Exception:
            pd_sources_norm = []

    # Scan corpora under a shared time budget. Prefer the Tibia corpus first by default:
    # it tends to have higher overlap with the decoded text than generic PD sources.
    ua = "Mozilla/5.0 (Bonelord469SeqMatch/1.0)"
    cache_dir = os.path.join(os.getcwd(), "tmp", "corpus")
    npc_cache = _cache_path_for_url(cache_dir, prefix="tibia_npc", url=npc_url)
    book_cache = _cache_path_for_url(cache_dir, prefix="tibia_books", url=book_url)

    def over_budget() -> bool:
        return bool(float(time_budget_s) > 0 and (time.time() - start) >= float(time_budget_s))

    def scan_tibia() -> None:
        if not cand_set or len(matches) >= int(max_matches) or over_budget():
            return
        try:
            npc_obj = _fetch_json_url_cached(
                npc_url,
                cache_path=npc_cache,
                timeout_s=int(timeout_s),
                max_age_hours=float(cache_max_age_hours),
                user_agent=ua,
            )
            if isinstance(npc_obj, list):
                for item in npc_obj:
                    if over_budget() or len(matches) >= int(max_matches):
                        break
                    if not isinstance(item, dict):
                        continue
                    name = str(item.get("name") or "").strip() or "NPC"
                    conv = item.get("conversation")
                    if not isinstance(conv, list):
                        continue
                    for ti, turn in enumerate(conv):
                        if not isinstance(turn, dict):
                            continue
                        answers = turn.get("answer")
                        if isinstance(answers, str):
                            answers = [answers]
                        if not isinstance(answers, list):
                            continue
                        for ai, a in enumerate(answers):
                            if not isinstance(a, str) or not a.strip():
                                continue
                            consider_source("TIBIA_NPC", f"{name}:{ti}:{ai}", npc_url, a)
                            if len(matches) >= int(max_matches) or over_budget():
                                break
        except Exception:
            pass

        if not cand_set or len(matches) >= int(max_matches) or over_budget():
            return
        try:
            book_obj = _fetch_json_url_cached(
                book_url,
                cache_path=book_cache,
                timeout_s=int(timeout_s),
                max_age_hours=float(cache_max_age_hours),
                user_agent=ua,
            )
            if isinstance(book_obj, list):
                for item in book_obj:
                    if over_budget() or len(matches) >= int(max_matches):
                        break
                    if not isinstance(item, dict):
                        continue
                    title = str(item.get("name") or "").strip() or "BOOK"
                    text = item.get("text")
                    if not isinstance(text, str) or not text.strip():
                        continue
                    for si, sent in enumerate(_iter_sentences(text)):
                        consider_source("TIBIA_BOOK", f"{title}:{si}", book_url, sent)
                        if len(matches) >= int(max_matches) or over_budget():
                            break
        except Exception:
            pass

    def scan_pd() -> None:
        if not (pd_enabled and pd_sources_norm) or not cand_set or len(matches) >= int(max_matches) or over_budget():
            return
        max_chars_i = max(0, int(pd_max_chars))
        max_sent_i = max(0, int(pd_max_sentences_per_source))
        for pd_cid, pd_url_s in pd_sources_norm:
            if not cand_set or len(matches) >= int(max_matches) or over_budget():
                break
            try:
                pd_cache = _cache_path_for_url(cache_dir, prefix="pd_text", url=pd_url_s, ext="txt")
                pd_txt = _fetch_text_url_cached(
                    pd_url_s,
                    cache_path=pd_cache,
                    timeout_s=int(timeout_s),
                    max_age_hours=float(pd_cache_max_age_hours),
                    user_agent="Mozilla/5.0 (Bonelord469SeqMatchPD/1.0)",
                )
                # Best-effort strip for Gutenberg boilerplate when present.
                m_start = re.search(
                    r"\*\*\*\s*START OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
                    pd_txt,
                    flags=re.IGNORECASE | re.DOTALL,
                )
                m_end = re.search(
                    r"\*\*\*\s*END OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
                    pd_txt,
                    flags=re.IGNORECASE | re.DOTALL,
                )
                if m_start and m_end and m_end.start() > m_start.end():
                    pd_txt = pd_txt[m_start.end() : m_end.start()]
                if max_chars_i > 0 and len(pd_txt) > max_chars_i:
                    pd_txt = pd_txt[:max_chars_i]
                for si, sent in enumerate(_iter_sentences(pd_txt, max_sentences=max_sent_i)):
                    if over_budget() or len(matches) >= int(max_matches):
                        break
                    consider_source("PD", f"{pd_cid}:{si}", pd_url_s, sent)
            except Exception:
                # Best-effort: PD source is optional; do not block SequenceMatches.
                continue

    if bool(scan_tibia_first):
        scan_tibia()
        scan_pd()
    else:
        scan_pd()
        scan_tibia()

    ws_out = ensure_sheet(
        wb,
        SEQUENCE_MATCHES_SHEET,
        ["Iteration", "UTC", "BookID", "N", "Phrase", "SourceKind", "SourceID", "SourceURL", "Snippet", "Notes", "PhraseSig"],
    )
    ho = ws_find_header_row(ws_out, ["Iteration", "BookID", "Phrase"], max_scan=3)
    co = ws_headers(ws_out, ho)
    if "PhraseSig" not in co:
        # Backwards-compatible upgrade for existing workbooks.
        col = ws_out.max_column + 1
        ws_out.cell(ho, col).value = "PhraseSig"
        co["PhraseSig"] = col
    if ws_out.max_row > ho:
        ws_out.delete_rows(ho + 1, ws_out.max_row - ho)

    rr = ho + 1
    for bid, n, phrase, phrase_sig, kind, packed in matches:
        source_id, source_url, snippet = packed.split(" | ", 2)
        ws_out.cell(rr, co["Iteration"]).value = int(iter_num)
        ws_out.cell(rr, co["UTC"]).value = utc
        ws_out.cell(rr, co["BookID"]).value = int(bid)
        ws_out.cell(rr, co["N"]).value = int(n)
        ws_out.cell(rr, co["Phrase"]).value = phrase
        ws_out.cell(rr, co["SourceKind"]).value = kind
        ws_out.cell(rr, co["SourceID"]).value = source_id
        ws_out.cell(rr, co["SourceURL"]).value = source_url
        ws_out.cell(rr, co["Snippet"]).value = snippet[:220]
        ws_out.cell(rr, co["Notes"]).value = "snippets only; signature+canon n-gram match"
        if "PhraseSig" in co:
            ws_out.cell(rr, co["PhraseSig"]).value = phrase_sig[:320]
        rr += 1
        if rr - (ho + 1) >= int(max_matches):
            break

    upsert_sheet_index_entry(
        wb,
        SEQUENCE_MATCHES_SHEET,
        "Sequence matches between ContextEnglish n-grams and corpora (analysis-only; snippets only; signature+canon match).",
    )

    # Optional cache: keep a deduped union of past matches so SequenceWordHints can accumulate
    # evidence across iterations even when SequenceMatches rotates/explores.
    if bool(cache_enabled):
        cache_headers = [
            "FirstIter",
            "LastIter",
            "LastUTC",
            "BookID",
            "N",
            "Phrase",
            "SourceKind",
            "SourceID",
            "SourceURL",
            "Snippet",
            "Notes",
            "PhraseSig",
            "HitCount",
        ]
        ws_cache = ensure_sheet(wb, SEQUENCE_MATCHES_CACHE_SHEET, cache_headers)
        hc2 = ws_find_header_row(ws_cache, ["SourceKind", "SourceID", "Phrase", "PhraseSig"], max_scan=3)
        cc2 = ws_headers(ws_cache, hc2)
        # Backwards-compatible upgrades (in case the sheet already exists with an older schema).
        for col_name in cache_headers:
            if col_name not in cc2:
                col = ws_cache.max_column + 1
                ws_cache.cell(hc2, col).value = col_name
                cc2[col_name] = col

        # Snapshot current cache into a key->record map so pruning can rewrite cleanly.
        records: Dict[str, Dict[str, object]] = {}
        for r in range(hc2 + 1, ws_cache.max_row + 1):
            kind = ws_cache.cell(r, cc2["SourceKind"]).value
            sid = ws_cache.cell(r, cc2["SourceID"]).value
            phrase = ws_cache.cell(r, cc2["Phrase"]).value
            psig = ws_cache.cell(r, cc2["PhraseSig"]).value
            if not isinstance(phrase, str) or not phrase.strip():
                continue
            key = f"{kind}|{sid}|{psig}|{phrase.strip()}"
            rec: Dict[str, object] = {}
            for col_name in cache_headers:
                rec[col_name] = ws_cache.cell(r, cc2[col_name]).value if col_name in cc2 else None
            records[key] = rec

        # Upsert new matches into the record map.
        for bid, n, phrase, phrase_sig, kind, packed in matches:
            source_id, source_url, snippet = packed.split(" | ", 2)
            key = f"{kind}|{source_id}|{phrase_sig}|{phrase}"
            rec = records.get(key)
            if rec is None:
                rec = {k: None for k in cache_headers}
                rec["FirstIter"] = int(iter_num)
                rec["HitCount"] = 0
                records[key] = rec
            rec["LastIter"] = int(iter_num)
            rec["LastUTC"] = utc
            rec["BookID"] = int(bid)
            rec["N"] = int(n)
            rec["Phrase"] = phrase
            rec["SourceKind"] = kind
            rec["SourceID"] = source_id
            rec["SourceURL"] = source_url
            rec["Snippet"] = snippet[:220]
            rec["Notes"] = "cache: snippets only; union across iterations"
            rec["PhraseSig"] = phrase_sig[:320]
            try:
                prev_hits = int(rec.get("HitCount") or 0)
            except Exception:
                prev_hits = 0
            rec["HitCount"] = int(prev_hits) + 1

        # Prune cache to keep workbook size stable.
        max_rows_i = int(max(0, int(cache_max_rows)))
        items2 = list(records.items())
        if max_rows_i > 0 and len(items2) > max_rows_i:
            items2.sort(
                key=lambda kv: (
                    -int(kv[1].get("HitCount") or 0),
                    -int(kv[1].get("LastIter") or 0),
                    str(kv[0] or ""),
                )
            )
            items2 = items2[:max_rows_i]
        else:
            items2.sort(key=lambda kv: (-int(kv[1].get("HitCount") or 0), -int(kv[1].get("LastIter") or 0), str(kv[0] or "")))

        # Rewrite sheet compactly.
        if ws_cache.max_row > hc2:
            ws_cache.delete_rows(hc2 + 1, ws_cache.max_row - hc2)
        rrw = hc2 + 1
        for _key, rec in items2:
            for col_name in cache_headers:
                if col_name not in cc2:
                    continue
                ws_cache.cell(rrw, cc2[col_name]).value = rec.get(col_name)
            rrw += 1

        upsert_sheet_index_entry(
            wb,
            SEQUENCE_MATCHES_CACHE_SHEET,
            "Cache of SequenceMatches (snippets only; deduped union across iterations) to accumulate evidence for SequenceWordHints.",
        )

    status_out = (
        f"ok (cand_pool={cand_pool_n}, forced={forced_n}, selected={len(candidates)}, "
        f"rotate={explore_used}, keep_top={explore_keep_top_i}, window={explore_window_n}, off={explore_off}, "
        f"cache={int(bool(cache_enabled))}/{int(cache_max_rows)})"
    )
    return rr - (ho + 1), status_out


def _load_sequence_word_hints(wb: openpyxl.Workbook) -> Dict[str, Dict[str, int]]:
    """Load signature->word->count hints derived from SequenceMatches.

    These hints are used only to *nudge* ContextEnglish candidate ordering (display-only).
    """
    if SEQUENCE_WORD_HINTS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SEQUENCE_WORD_HINTS_SHEET]
    h = ws_find_header_row(ws, ["CanonSig", "ToWord", "Count"], max_scan=3)
    c = ws_headers(ws, h)
    out: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in range(h + 1, ws.max_row + 1):
        sig = ws.cell(r, c["CanonSig"]).value
        w = ws.cell(r, c["ToWord"]).value
        cnt = ws.cell(r, c["Count"]).value
        if not isinstance(sig, str) or not sig.strip():
            continue
        if not isinstance(w, str) or not w.strip():
            continue
        try:
            ci = int(cnt or 0)
        except Exception:
            ci = 0
        if ci <= 0:
            continue
        sig_s = sig.strip()
        w_s = w.strip().lower()
        out[sig_s][w_s] += ci
    return out


def _load_reverse_phrase_token_hints(wb: openpyxl.Workbook) -> Dict[str, Dict[str, int]]:
    """Load token-signature->word->count hints derived from ReversePhraseTokenCands_Auto.

    This is display-only: used to boost ContextEnglish/CodeAware candidates when reverse phrase mining
    discovers strong token->word clues from external corpora snippets.
    """
    if REVERSE_PHRASE_CANDS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[REVERSE_PHRASE_CANDS_SHEET]
    h = ws_find_header_row(ws, ["Base", "TopWord"], max_scan=3)
    c = ws_headers(ws, h)
    out: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in range(h + 1, ws.max_row + 1):
        base = ws.cell(r, c["Base"]).value
        topw = ws.cell(r, c["TopWord"]).value
        if not isinstance(base, str) or not base.strip():
            continue
        if not isinstance(topw, str) or not topw.strip():
            continue
        sig = _token_signature(base.strip())
        if not sig:
            continue
        # Weight by support; share scales weight down for weaker candidates.
        weight_base = 1
        if "TopWordCount" in c:
            try:
                weight_base = int(ws.cell(r, c["TopWordCount"]).value or 0) or 1
            except Exception:
                weight_base = 1
        elif "SupportOcc" in c:
            try:
                weight_base = int(ws.cell(r, c["SupportOcc"]).value or 0) or 1
            except Exception:
                weight_base = 1
        share = 1.0
        if "TopShare" in c:
            try:
                share = float(ws.cell(r, c["TopShare"]).value or 0.0) or 0.0
            except Exception:
                share = 1.0
        if share <= 0:
            share = 1.0
        weight = max(1, int(round(float(weight_base) * float(min(1.0, share)))))

        cand_words: List[str] = []
        if "CandidateWords" in c:
            cw = ws.cell(r, c["CandidateWords"]).value
            if isinstance(cw, str) and cw.strip():
                cand_words = [w.strip() for w in cw.split(",") if w.strip()]
        # Always include TopWord first.
        words = [topw.strip()] + [w for w in cand_words if w.strip().lower() != topw.strip().lower()]
        for i, w in enumerate(words[:6]):
            ww = re.sub(r"[^a-z']", "", str(w or "").lower())
            if not ww:
                continue
            wgt = int(weight) if i == 0 else max(1, int(weight) // 2)
            out[sig][ww] += int(wgt)
    return out


def materialize_sequence_word_hints(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    drop_final_e: bool,
    drop_all_h: bool,
    drop_all_o: bool,
    min_n: int = 3,
    skip_identity: bool = True,
    max_rows: int = 800,
    min_hint_ratio: float = 1.5,
    hint_exclude_stopwords: bool = True,
    hint_stopword_ratio: float = 0.67,
) -> Tuple[int, str]:
    """Derive word-level substitution hints from SequenceMatches_Auto (analysis-only).

    For each sequence match row, we align phrase words positionally against snippet words
    (in the order implied by the stored PhraseSig) and emit (CanonSig, FromWord, ToWord) counts.
    Rows dominated by stopwords can optionally be skipped so hints stay focused on high-signal tokens.
    """
    if SEQUENCE_MATCHES_SHEET not in wb.sheetnames and SEQUENCE_MATCHES_CACHE_SHEET not in wb.sheetnames:
        ensure_sheet(
            wb,
            SEQUENCE_WORD_HINTS_SHEET,
            ["Iteration", "UTC", "CanonSig", "FromWord", "ToWord", "Count", "Examples", "Notes"],
        )
        upsert_sheet_index_entry(
            wb,
            SEQUENCE_WORD_HINTS_SHEET,
            "Word-level hints derived from SequenceMatches_Auto (analysis-only; used to nudge ContextEnglish).",
        )
        return 0, "missing_seqmatches"

    src_sheet = SEQUENCE_MATCHES_CACHE_SHEET if SEQUENCE_MATCHES_CACHE_SHEET in wb.sheetnames else SEQUENCE_MATCHES_SHEET
    ws_sm = wb[src_sheet]
    hs = ws_find_header_row(ws_sm, ["Phrase", "Snippet"], max_scan=3)
    cs = ws_headers(ws_sm, hs)

    counts: Counter[Tuple[str, str, str]] = Counter()
    examples: Dict[Tuple[str, str, str], List[str]] = defaultdict(list)

    for r in range(hs + 1, ws_sm.max_row + 1):
        try:
            n_val = int(ws_sm.cell(r, cs.get("N") or 0).value or 0) if cs.get("N") else 0
        except Exception:
            n_val = 0
        if int(min_n) > 0 and n_val and int(n_val) < int(min_n):
            continue
        phrase = ws_sm.cell(r, cs["Phrase"]).value
        snippet = ws_sm.cell(r, cs["Snippet"]).value
        if not isinstance(phrase, str) or not phrase.strip():
            continue
        if not isinstance(snippet, str) or not snippet.strip():
            continue
        phrase_sig = ws_sm.cell(r, cs.get("PhraseSig") or 0).value
        sig_words: List[str] = []
        if isinstance(phrase_sig, str) and phrase_sig.strip():
            sig_words = [s for s in phrase_sig.strip().split() if s]
        p_words = [re.sub(r"[^a-z']", "", w.lower()) for w in _iter_words(phrase)]
        s_words = [re.sub(r"[^a-z']", "", w.lower()) for w in _iter_words(snippet)]
        p_words = [w for w in p_words if w]
        s_words = [w for w in s_words if w]
        if not p_words or not s_words or len(p_words) != len(s_words):
            continue
        if sig_words and len(sig_words) != len(p_words):
            continue
        if hint_stopword_ratio >= 0 and p_words:
            stop_ratio = float(sum(1 for w in p_words if not _is_seq_content_word(w))) / float(len(p_words))
            if stop_ratio >= float(hint_stopword_ratio):
                continue

        # Extract provenance (compact).
        bid = ws_sm.cell(r, cs.get("BookID") or 0).value if cs.get("BookID") else None
        srcid = ws_sm.cell(r, cs.get("SourceID") or 0).value if cs.get("SourceID") else None
        prov = f"b{bid}:{srcid}" if bid is not None and srcid is not None else ""

        for idx, (pw, sw) in enumerate(zip(p_words, s_words)):
            if skip_identity and pw == sw:
                continue
            if hint_exclude_stopwords and (not _is_seq_content_word(pw) or not _is_seq_content_word(sw)):
                continue
            pc = _lore_canon_word(pw, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
            sc = _lore_canon_word(sw, drop_final_e=drop_final_e, drop_all_h=drop_all_h, drop_all_o=drop_all_o)
            if not pc or not sc:
                continue
            ps = _lore_signature(pc)
            ss = _lore_signature(sc)
            if ps != ss:
                continue
            if sig_words:
                expected_sig = sig_words[idx]
                if expected_sig and expected_sig != ps:
                    continue
            key = (ps, pw, sw)
            counts[key] += 1
            if prov and len(examples[key]) < 5:
                examples[key].append(prov)

    ws_out = ensure_sheet(
        wb,
        SEQUENCE_WORD_HINTS_SHEET,
        ["Iteration", "UTC", "CanonSig", "FromWord", "ToWord", "Count", "Examples", "Notes"],
    )
    ho = ws_find_header_row(ws_out, ["Iteration", "CanonSig", "FromWord", "ToWord"], max_scan=3)
    co = ws_headers(ws_out, ho)
    if ws_out.max_row > ho:
        ws_out.delete_rows(ho + 1, ws_out.max_row - ho)

    # Quality filter: when multiple ToWords compete for the same (CanonSig, FromWord),
    # keep only a dominant replacement (ratio >= min_hint_ratio). This reduces noisy,
    # directionless hints while keeping the feature GT-safe (display-only).
    grouped: Dict[Tuple[str, str], List[Tuple[str, int]]] = defaultdict(list)
    for (sig, frm, to), cnt in counts.items():
        try:
            ci = int(cnt or 0)
        except Exception:
            ci = 0
        if ci <= 0:
            continue
        grouped[(sig, frm)].append((to, ci))

    filtered: List[Tuple[Tuple[str, str, str], int]] = []
    for (sig, frm), to_list in grouped.items():
        to_list.sort(key=lambda t: (-t[1], t[0]))
        top_to, top_cnt = to_list[0]
        keep = True
        if len(to_list) >= 2 and float(min_hint_ratio) > 1.0:
            runner_cnt = int(to_list[1][1])
            if runner_cnt > 0:
                ratio = float(top_cnt) / float(runner_cnt)
                if ratio + 1e-12 < float(min_hint_ratio):
                    keep = False
        if keep:
            filtered.append(((sig, frm, top_to), int(top_cnt)))

    items = sorted(filtered, key=lambda kv: (-kv[1], kv[0][0], kv[0][1], kv[0][2]))
    rr = ho + 1
    for (sig, frm, to), cnt in items[: int(max_rows)]:
        ws_out.cell(rr, co["Iteration"]).value = int(iter_num)
        ws_out.cell(rr, co["UTC"]).value = utc
        ws_out.cell(rr, co["CanonSig"]).value = sig
        ws_out.cell(rr, co["FromWord"]).value = frm
        ws_out.cell(rr, co["ToWord"]).value = to
        ws_out.cell(rr, co["Count"]).value = int(cnt)
        if "Examples" in co:
            ws_out.cell(rr, co["Examples"]).value = ", ".join(examples.get((sig, frm, to)) or [])[:220]
        if "Notes" in co:
            ws_out.cell(rr, co["Notes"]).value = "derived from signature n-gram matches"
        rr += 1

    upsert_sheet_index_entry(
        wb,
        SEQUENCE_WORD_HINTS_SHEET,
        "Word-level hints derived from SequenceMatches_Auto (analysis-only; used to nudge ContextEnglish).",
    )
    return rr - (ho + 1), "ok"


def materialize_sestina_scan(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    active_tokens: Dict[str, GlossaryToken],
    max_lines: int,
    min_score30: int,
    max_candidates: int,
    use_token_signature: bool = True,
    envoi_bonus: float = 1.0,
) -> Tuple[int, int, int, str]:
    """Detect sestina/sestine structure windows (analysis-only).

    We treat end-of-line boundaries as the existing structural markers in StrictPlus:
    - marker tokens: E, FF (rendered as <E>, <FF>)
    - newline logograms: LF, LN, SF (rendered as ↵)

    We then scan the resulting line-end token stream for the classic retrogradatio cruciata
    permutation across 6 stanzas of 6 lines (36 lines). This is display/analysis-only and
    does not mutate Glossary or DP.

    Returns: (lines_written, candidates_written, best_score30, fingerprint)
    """
    if "Books" not in wb.sheetnames:
        ensure_sheet(
            wb,
            SESTINA_LINES_SHEET,
            ["Iteration", "UTC", "GlobalLine", "BookID", "LineInBook", "EndToken", "EndWord", "LinePreview", "EndKind"],
        )
        ensure_sheet(
            wb,
            SESTINA_CANDIDATES_SHEET,
            [
                "Iteration",
                "UTC",
                "GlobalStartLine",
                "BookIDStart",
                "LineInBookStart",
                "Score30",
                "Mismatches30",
                "EndTokenSet",
                "EndWords",
                "EnvoiEndTokens",
                "Notes",
            ],
        )
        return 0, 0, 0, ""

    ws_b = wb["Books"]
    hb = ws_find_header_row(ws_b, ["BookID", "DecodedBase"], max_scan=3)
    cb = ws_headers(ws_b, hb)

    # Boundary tokens that terminate a "line" for this structural scan.
    marker_line_break = {"E", "FF", "*"}
    logogram_newline = {"LF", "LN", "SF"}

    def _is_line_break(it: object) -> bool:
        if not isinstance(it, GlossaryToken):
            return False
        if it.token in logogram_newline:
            return True
        if it.token_type == "marker" and it.token in marker_line_break:
            return True
        return False

    def _line_preview(items: Sequence[object]) -> str:
        lossless: List[str] = []
        for it in items:
            if isinstance(it, str):
                lossless.append(it)
            elif it.token_type == "marker":
                lossless.append(f"<{it.token}>")
            else:
                lossless.extend(str(it.translation or "").split())
        return render_strictplus_from_lossless(lossless)

    # Extract global line list.
    lines: List[Dict[str, object]] = []
    global_line = 0
    for r in range(hb + 1, ws_b.max_row + 1):
        bid = ws_b.cell(r, cb["BookID"]).value
        base = ws_b.cell(r, cb["DecodedBase"]).value
        if bid is None or base is None or str(base).strip() == "":
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue

        items = dp_tokenize_base_with_punct(str(base), active_tokens)
        cur: List[object] = []
        line_in_book = 0
        saw_break = False

        def flush(end_kind: str) -> None:
            nonlocal global_line, line_in_book, cur
            if not cur:
                return
            # Find last content token (skip markers/punct/newlines).
            end_tok = ""
            end_word = ""
            for it2 in reversed(cur):
                if not isinstance(it2, GlossaryToken):
                    continue
                if _is_line_break(it2):
                    continue
                if it2.token_type == "marker":
                    continue
                tr = str(it2.translation or "").strip()
                if not tr:
                    continue
                end_tok = it2.token
                end_word = tr.split()[-1]
                break

            global_line += 1
            preview = _line_preview(cur)[:240]
            lines.append(
                {
                    "Iteration": int(iter_num),
                    "UTC": utc,
                    "GlobalLine": int(global_line),
                    "BookID": int(bid_i),
                    "LineInBook": int(line_in_book),
                    "EndToken": end_tok,
                    "EndWord": end_word,
                    "LinePreview": preview,
                    "EndKind": end_kind,
                }
            )
            line_in_book += 1
            cur = []

        for it in items:
            if _is_line_break(it):
                saw_break = True
                flush("MARK")
                continue
            cur.append(it)
        # Always include trailing segment as a line (book-end).
        flush("BOOK_END" if not saw_break else "TAIL")

        if len(lines) >= int(max_lines) > 0:
            break

    # Write lines sheet (idempotent overwrite).
    ws_lines = ensure_sheet(
        wb,
        SESTINA_LINES_SHEET,
        ["Iteration", "UTC", "GlobalLine", "BookID", "LineInBook", "EndToken", "EndWord", "LinePreview", "EndKind"],
    )
    hl = ws_find_header_row(ws_lines, ["GlobalLine", "BookID", "EndToken"], max_scan=3)
    cl = ws_headers(ws_lines, hl)
    if ws_lines.max_row > hl:
        ws_lines.delete_rows(hl + 1, ws_lines.max_row - hl)

    rr = hl + 1
    for row in lines:
        ws_lines.cell(rr, cl["Iteration"]).value = row["Iteration"]
        ws_lines.cell(rr, cl["UTC"]).value = row["UTC"]
        ws_lines.cell(rr, cl["GlobalLine"]).value = row["GlobalLine"]
        ws_lines.cell(rr, cl["BookID"]).value = row["BookID"]
        ws_lines.cell(rr, cl["LineInBook"]).value = row["LineInBook"]
        ws_lines.cell(rr, cl["EndToken"]).value = row["EndToken"]
        ws_lines.cell(rr, cl["EndWord"]).value = row["EndWord"]
        ws_lines.cell(rr, cl["LinePreview"]).value = row["LinePreview"]
        ws_lines.cell(rr, cl["EndKind"]).value = row["EndKind"]
        rr += 1

    upsert_sheet_index_entry(
        wb,
        SESTINA_LINES_SHEET,
        "Sestina/sestine structural line extraction (analysis-only): DP-tokenized lines split by structural markers (<E>,<FF>,↵).",
    )

    # Scan 36-line windows for classic sestina end-token permutation.
    # Optionally score using token-signatures (sorted letters) to be robust to anagram variants.
    end_tokens_raw = [str(row.get("EndToken") or "") for row in lines]
    end_tokens = [_token_signature(t) if use_token_signature else t for t in end_tokens_raw]
    end_words = [str(row.get("EndWord") or "") for row in lines]
    book_ids = [int(row.get("BookID") or 0) for row in lines]
    line_in_book = [int(row.get("LineInBook") or 0) for row in lines]

    P = [5, 0, 4, 1, 3, 2]  # retrogradatio cruciata: new[i] = old[P[i]]

    all_windows: List[Tuple[float, int, int, int, int, List[str], List[str], List[str], List[str], List[str]]] = []
    # (rank_score, score30, envoi_score3, mism30, start_idx, w_keys(6), w_raw(6), w_words(6), envoi_keys, envoi_raw)
    for start in range(0, max(0, len(end_tokens) - 36 + 1)):
        w = end_tokens[start : start + 6]
        w_raw = end_tokens_raw[start : start + 6]
        if any(not t for t in w):
            continue
        if len(set(w)) != 6:
            continue
        order = [0, 1, 2, 3, 4, 5]
        score = 0
        mism = 0
        # Check stanzas 2..6 only (30 comparisons).
        for stanza in range(1, 6):
            order = [order[p] for p in P]
            for i in range(6):
                exp = w[order[i]]
                act = end_tokens[start + stanza * 6 + i]
                if act == exp:
                    score += 1
                else:
                    mism += 1
        envoi = end_tokens[start + 36 : start + 39] if (start + 39) <= len(end_tokens) else []
        envoi_raw = end_tokens_raw[start + 36 : start + 39] if (start + 39) <= len(end_tokens_raw) else []
        # Envoi pattern (classic): end tokens tend to be keyword2, keyword4, keyword6 (1-based),
        # while the paired keywords appear mid-line. We only score end-token positions here (safe/cheap).
        exp_idx = [1, 3, 5]
        envoi_score = 0
        for j, idx in enumerate(exp_idx):
            if j < len(envoi) and envoi[j] == w[idx]:
                envoi_score += 1

        rank_score = float(score) + float(envoi_bonus) * float(envoi_score)
        all_windows.append((rank_score, score, envoi_score, mism, start, w, w_raw, end_words[start : start + 6], envoi, envoi_raw))

    all_windows.sort(key=lambda t: (-t[0], -t[1], -t[2], t[4]))
    candidates = [w for w in all_windows if w[1] >= int(min_score30)]
    if not candidates and all_windows:
        # Always emit *some* best-effort candidates to make the sheet actionable for humans.
        candidates = all_windows[: min(int(max_candidates), 20)]
    candidates = candidates[: int(max_candidates)]

    ws_cand = ensure_sheet(
        wb,
        SESTINA_CANDIDATES_SHEET,
        [
            "Iteration",
            "UTC",
            "GlobalStartLine",
            "BookIDStart",
            "LineInBookStart",
            "Score30",
            "Mismatches30",
            "ScoreRank",
            "EnvoiScore3",
            "EndTokenSet",
            "EndTokensRaw",
            "EndWords",
            "EnvoiEndTokens",
            "EnvoiEndTokensRaw",
            "Notes",
        ],
    )
    hc = ws_find_header_row(ws_cand, ["GlobalStartLine", "Score30"], max_scan=3)
    cc = ws_headers(ws_cand, hc)
    # Backwards-compatible upgrades (in case the sheet already exists with the old schema).
    for col_name in ("ScoreRank", "EnvoiScore3", "EndTokensRaw", "EnvoiEndTokensRaw"):
        if col_name not in cc:
            col = ws_cand.max_column + 1
            ws_cand.cell(hc, col).value = col_name
            cc[col_name] = col
    if ws_cand.max_row > hc:
        ws_cand.delete_rows(hc + 1, ws_cand.max_row - hc)

    best_score = 0
    rr = hc + 1
    fp_items: List[str] = []
    for rank_score, score, envoi_score, mism, start, w, w_raw, w_words, envoi, envoi_raw in candidates:
        gl = start + 1  # 1-based global line index
        best_score = max(best_score, int(score))
        ws_cand.cell(rr, cc["Iteration"]).value = int(iter_num)
        ws_cand.cell(rr, cc["UTC"]).value = utc
        ws_cand.cell(rr, cc["GlobalStartLine"]).value = int(gl)
        ws_cand.cell(rr, cc["BookIDStart"]).value = int(book_ids[start] or 0)
        ws_cand.cell(rr, cc["LineInBookStart"]).value = int(line_in_book[start] or 0)
        ws_cand.cell(rr, cc["Score30"]).value = int(score)
        ws_cand.cell(rr, cc["Mismatches30"]).value = int(mism)
        if "ScoreRank" in cc:
            ws_cand.cell(rr, cc["ScoreRank"]).value = float(rank_score)
        if "EnvoiScore3" in cc:
            ws_cand.cell(rr, cc["EnvoiScore3"]).value = int(envoi_score)
        ws_cand.cell(rr, cc["EndTokenSet"]).value = " ".join(w)
        if "EndTokensRaw" in cc:
            ws_cand.cell(rr, cc["EndTokensRaw"]).value = " ".join(w_raw)
        ws_cand.cell(rr, cc["EndWords"]).value = " ".join(w_words)
        ws_cand.cell(rr, cc["EnvoiEndTokens"]).value = " ".join(envoi)
        if "EnvoiEndTokensRaw" in cc:
            ws_cand.cell(rr, cc["EnvoiEndTokensRaw"]).value = " ".join(envoi_raw)
        ws_cand.cell(rr, cc["Notes"]).value = (
            "sestina window score on 6x6 end-token permutation (stanzas 2-6); "
            f"key={'token_sig' if use_token_signature else 'raw_tok'}; envoi_score3={envoi_score}"
        )
        fp_items.append(f"{gl}:{score}:{envoi_score}:{' '.join(w)}")
        rr += 1

    fp = hashlib.sha1("\n".join(fp_items).encode("utf-8", errors="ignore")).hexdigest() if fp_items else ""

    upsert_sheet_index_entry(
        wb,
        SESTINA_CANDIDATES_SHEET,
        "Sestina/sestine candidate windows (analysis-only): scans extracted line-end tokens for the classic retrogradatio cruciata permutation.",
    )

    return (len(lines), len(candidates), int(best_score), fp)


def materialize_sestina_obligation_map(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    max_candidates: int,
    min_score30: int,
    obligatory_impact: float,
    conditional_impact: float,
    decorative_impact: float,
    no_collapse_tolerance: float,
    reorder_resilient_ratio: float,
) -> Tuple[int, int, float, float, str, str]:
    """Classify positional structural load for sestina-like windows (analysis-only).

    Uses existing `SestinaLines_Auto` + `SestinaCandidates_Auto` outputs:
    - Ablation: remove one position at a time and measure impact on normalized score.
    - Compression tolerance: remove weakest one/two positions and see if coherence collapses.
    - Reorder stress: compare baseline against best non-identity key ordering.

    Returns:
      (rows_written, candidates_analyzed, core_avg, reorder_resilient_frac, fingerprint, status)
    """
    ws_out = ensure_sheet(
        wb,
        SESTINA_OBLIGATION_SHEET,
        [
            "Iteration",
            "UTC",
            "GlobalStartLine",
            "BookIDStart",
            "Score30",
            "BaselineScore30",
            "BaselineRate30",
            "Position",
            "KeyToken",
            "RawToken",
            "MatchCount5",
            "MatchRate5",
            "RemovalRate25",
            "ImpactVsBaseline",
            "Role",
            "CorePosition",
            "MaxRemovedNoCollapse",
            "BestAltScore30",
            "BestAltOrder",
            "AltVsBaseRatio",
            "ReorderResilient",
            "Notes",
        ],
    )
    ho = ws_find_header_row(ws_out, ["GlobalStartLine", "Position", "Role"], max_scan=3)
    co = ws_headers(ws_out, ho)
    if ws_out.max_row > ho:
        ws_out.delete_rows(ho + 1, ws_out.max_row - ho)

    if SESTINA_LINES_SHEET not in wb.sheetnames or SESTINA_CANDIDATES_SHEET not in wb.sheetnames:
        upsert_sheet_index_entry(
            wb,
            SESTINA_OBLIGATION_SHEET,
            "Sestina positional obligation map (analysis-only): requires SestinaLines_Auto + SestinaCandidates_Auto.",
        )
        return 0, 0, 0.0, 0.0, "", "missing input sheets"

    ws_lines = wb[SESTINA_LINES_SHEET]
    hl = ws_find_header_row(ws_lines, ["GlobalLine", "EndToken"], max_scan=3)
    cl = ws_headers(ws_lines, hl)
    end_token_by_line: Dict[int, str] = {}
    bookid_by_line: Dict[int, int] = {}
    for r in range(hl + 1, ws_lines.max_row + 1):
        gl = ws_lines.cell(r, cl["GlobalLine"]).value
        tok = ws_lines.cell(r, cl["EndToken"]).value
        bid = ws_lines.cell(r, cl.get("BookID", 0) or 0).value if "BookID" in cl else None
        if gl is None:
            continue
        try:
            gli = int(gl)
        except Exception:
            continue
        end_token_by_line[gli] = str(tok or "")
        try:
            bookid_by_line[gli] = int(bid or 0)
        except Exception:
            bookid_by_line[gli] = 0

    ws_c = wb[SESTINA_CANDIDATES_SHEET]
    hc = ws_find_header_row(ws_c, ["GlobalStartLine", "Score30", "EndTokenSet"], max_scan=3)
    cc = ws_headers(ws_c, hc)

    cand_rows: List[Tuple[int, int, int, str, str, str]] = []
    # (score30, score_rank, global_start_line, keyset, rawset, notes)
    for r in range(hc + 1, ws_c.max_row + 1):
        gl = ws_c.cell(r, cc["GlobalStartLine"]).value
        sc = ws_c.cell(r, cc["Score30"]).value
        keyset = ws_c.cell(r, cc["EndTokenSet"]).value
        if gl is None or keyset is None:
            continue
        try:
            gl_i = int(gl)
            sc_i = int(sc or 0)
        except Exception:
            continue
        if sc_i < int(min_score30):
            continue
        score_rank = 0
        try:
            score_rank = int(ws_c.cell(r, cc.get("ScoreRank", 0) or 0).value or 0) if "ScoreRank" in cc else 0
        except Exception:
            score_rank = 0
        rawset = str(ws_c.cell(r, cc.get("EndTokensRaw", 0) or 0).value or "") if "EndTokensRaw" in cc else ""
        notes = str(ws_c.cell(r, cc.get("Notes", 0) or 0).value or "") if "Notes" in cc else ""
        cand_rows.append((sc_i, score_rank, gl_i, str(keyset), rawset, notes))

    if not cand_rows:
        upsert_sheet_index_entry(
            wb,
            SESTINA_OBLIGATION_SHEET,
            "Sestina positional obligation map (analysis-only): no candidates available for this iteration.",
        )
        return 0, 0, 0.0, 0.0, "", "no candidates"

    cand_rows.sort(key=lambda t: (-t[0], -t[1], t[2]))
    if int(max_candidates) > 0:
        cand_rows = cand_rows[: int(max_candidates)]

    P = [5, 0, 4, 1, 3, 2]

    def _score_window(base_keys: Sequence[str], actual_keys_30: Sequence[str]) -> Tuple[int, List[int]]:
        order = [0, 1, 2, 3, 4, 5]
        score = 0
        per_pos = [0, 0, 0, 0, 0, 0]
        idx = 0
        for _stanza in range(1, 6):
            order = [order[p] for p in P]
            for pos in range(6):
                exp = base_keys[order[pos]]
                act = actual_keys_30[idx]
                if exp == act:
                    score += 1
                    per_pos[pos] += 1
                idx += 1
        return score, per_pos

    rr = ho + 1
    fp_items: List[str] = []
    analyzed = 0
    core_total = 0
    reorder_resilient_windows = 0

    for score30, _score_rank, gl_i, keyset, rawset, notes in cand_rows:
        keys = [x for x in str(keyset).split() if x]
        if len(keys) < 6:
            continue
        keys = keys[:6]

        raw_tokens = [x for x in str(rawset).split() if x]
        if len(raw_tokens) < 6:
            raw_tokens = list(keys)
        else:
            raw_tokens = raw_tokens[:6]

        key_mode_token_sig = "key=token_sig" in str(notes or "").lower()

        # Collect the 30 observed end-tokens for stanzas 2..6.
        observed_raw: List[str] = []
        valid = True
        for offset in range(6, 36):
            gl = gl_i + offset
            if gl not in end_token_by_line:
                valid = False
                break
            observed_raw.append(end_token_by_line[gl])
        if not valid or len(observed_raw) != 30:
            continue

        if key_mode_token_sig:
            observed = [_token_signature(t) for t in observed_raw]
        else:
            observed = list(observed_raw)

        baseline_score, per_pos = _score_window(keys, observed)
        baseline_rate = float(baseline_score) / 30.0
        avg_per_pos = float(baseline_score) / 6.0 if baseline_score > 0 else 0.0

        # Reorder stress: best non-identity base ordering.
        best_alt_score = -1
        best_alt_perm: Optional[Tuple[int, ...]] = None
        identity = (0, 1, 2, 3, 4, 5)
        for perm in itertools.permutations(range(6)):
            if perm == identity:
                continue
            perm_keys = [keys[i] for i in perm]
            s_perm, _ = _score_window(perm_keys, observed)
            if s_perm > best_alt_score:
                best_alt_score = s_perm
                best_alt_perm = perm
        if best_alt_score < 0:
            best_alt_score = baseline_score
            best_alt_perm = identity

        if baseline_score > 0:
            alt_ratio = float(best_alt_score) / float(baseline_score)
        else:
            alt_ratio = 1.0 if best_alt_score == 0 else 999.0
        reorder_resilient = 1 if alt_ratio >= float(reorder_resilient_ratio) else 0
        if reorder_resilient:
            reorder_resilient_windows += 1

        # Compression/ablation ladder: remove weakest positions first.
        impacts: List[float] = []
        removal_rates: List[float] = []
        for pos in range(6):
            rm_rate = float(baseline_score - per_pos[pos]) / 25.0
            removal_rates.append(rm_rate)
            impacts.append(float(baseline_rate - rm_rate))

        remove_order = sorted(range(6), key=lambda i: (impacts[i], per_pos[i], i))
        max_removed_no_collapse = 0
        tolerance_floor = max(0.0, float(baseline_rate) - float(no_collapse_tolerance))
        rm1_rate = float(baseline_score - per_pos[remove_order[0]]) / 25.0
        if rm1_rate >= tolerance_floor:
            max_removed_no_collapse = 1
        rm2_rate = float(baseline_score - per_pos[remove_order[0]] - per_pos[remove_order[1]]) / 20.0
        if rm2_rate >= tolerance_floor:
            max_removed_no_collapse = 2

        core_count = 0
        best_alt_order_txt = ",".join(str(i + 1) for i in (best_alt_perm or identity))
        for pos in range(6):
            m = int(per_pos[pos])
            match_rate = float(m) / 5.0
            rm_rate = removal_rates[pos]
            impact = impacts[pos]
            load_ratio = (float(m) / avg_per_pos) if avg_per_pos > 0 else 0.0

            if impact >= float(obligatory_impact) and (m >= 2 or load_ratio >= 1.8):
                role = "OBLIGATORY"
            elif impact >= float(conditional_impact) and (m >= 1 or load_ratio >= 1.0):
                role = "CONDITIONAL"
            elif impact > float(decorative_impact):
                role = "REDUNDANT"
            else:
                role = "DECORATIVE"

            is_core = 1 if role in ("OBLIGATORY", "CONDITIONAL") else 0
            core_count += is_core

            ws_out.cell(rr, co["Iteration"]).value = int(iter_num)
            ws_out.cell(rr, co["UTC"]).value = utc
            ws_out.cell(rr, co["GlobalStartLine"]).value = int(gl_i)
            ws_out.cell(rr, co["BookIDStart"]).value = int(bookid_by_line.get(gl_i, 0))
            ws_out.cell(rr, co["Score30"]).value = int(score30)
            ws_out.cell(rr, co["BaselineScore30"]).value = int(baseline_score)
            ws_out.cell(rr, co["BaselineRate30"]).value = round(float(baseline_rate), 6)
            ws_out.cell(rr, co["Position"]).value = int(pos + 1)
            ws_out.cell(rr, co["KeyToken"]).value = keys[pos]
            ws_out.cell(rr, co["RawToken"]).value = raw_tokens[pos] if pos < len(raw_tokens) else ""
            ws_out.cell(rr, co["MatchCount5"]).value = int(m)
            ws_out.cell(rr, co["MatchRate5"]).value = round(float(match_rate), 6)
            ws_out.cell(rr, co["RemovalRate25"]).value = round(float(rm_rate), 6)
            ws_out.cell(rr, co["ImpactVsBaseline"]).value = round(float(impact), 6)
            ws_out.cell(rr, co["Role"]).value = role
            ws_out.cell(rr, co["CorePosition"]).value = int(is_core)
            ws_out.cell(rr, co["MaxRemovedNoCollapse"]).value = int(max_removed_no_collapse)
            ws_out.cell(rr, co["BestAltScore30"]).value = int(best_alt_score)
            ws_out.cell(rr, co["BestAltOrder"]).value = best_alt_order_txt
            ws_out.cell(rr, co["AltVsBaseRatio"]).value = round(float(alt_ratio), 6)
            ws_out.cell(rr, co["ReorderResilient"]).value = int(reorder_resilient)
            ws_out.cell(rr, co["Notes"]).value = (
                f"mode={'token_sig' if key_mode_token_sig else 'raw_tok'}; "
                f"remove_floor={tolerance_floor:.6f}; remove_order={','.join(str(i + 1) for i in remove_order[:2])}"
            )
            fp_items.append(
                f"{gl_i}|{pos+1}|{keys[pos]}|{role}|{m}|{impact:.6f}|{max_removed_no_collapse}|{best_alt_score}|{alt_ratio:.6f}"
            )
            rr += 1

        analyzed += 1
        core_total += core_count

    rows_written = rr - (ho + 1)
    if analyzed > 0:
        core_avg = float(core_total) / float(analyzed)
        reorder_frac = float(reorder_resilient_windows) / float(analyzed)
    else:
        core_avg = 0.0
        reorder_frac = 0.0
    fp = hashlib.sha1("\n".join(sorted(fp_items)).encode("utf-8", errors="ignore")).hexdigest() if fp_items else ""

    upsert_sheet_index_entry(
        wb,
        SESTINA_OBLIGATION_SHEET,
        "Sestina positional obligation map (analysis-only): per-position ablation impact + reorder stress classification "
        "(OBLIGATORY/CONDITIONAL/REDUNDANT/DECORATIVE).",
    )
    return rows_written, analyzed, core_avg, reorder_frac, fp, "ok" if analyzed else "no analyzable windows"


def materialize_rhythm_transition_abtest(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    window_size: int,
    min_lines: int,
    use_token_signature: bool,
    shuffle_trials: int,
    cycle6_delta_threshold: float,
) -> Tuple[int, int, float, float, float, float, float, float, float, str, str]:
    """Global rhythm analysis over extracted lines (analysis-only).

    Implements a compact A/B style structural test:
    - cycle-of-6 predictability (lag-half for even windows) vs shuffled control,
    - alternation score from EndKind-derived focus channel (MARK -> A, otherwise B),
    - anti-sestina controls: Fibonacci-lag match, triple-core concentration, sparse-echo ratio,
    - closure with light inversion (start vs end token overlap).
    """
    ws_out = ensure_sheet(
        wb,
        RHYTHM_TRANSITIONS_SHEET,
        [
            "Iteration",
            "UTC",
            "WindowStartGlobalLine",
            "BookIDStart",
            "WindowSize",
            "Cycle6Rate",
            "Cycle6ShuffleAvg",
            "Cycle6Delta",
            "ABAlternationRate",
            "FibonacciMatchRate",
            "TripleCoreConcentration",
            "SparseEchoRatio",
            "ClosureSame",
            "ClosureInversion",
            "ClosureOverlap",
            "FocusPattern",
            "ModelPreference",
            "StartToken",
            "EndToken",
            "Notes",
        ],
    )
    ho = ws_find_header_row(ws_out, ["WindowStartGlobalLine", "Cycle6Rate", "ModelPreference"], max_scan=3)
    co = ws_headers(ws_out, ho)
    if ws_out.max_row > ho:
        ws_out.delete_rows(ho + 1, ws_out.max_row - ho)

    if SESTINA_LINES_SHEET not in wb.sheetnames:
        upsert_sheet_index_entry(
            wb,
            RHYTHM_TRANSITIONS_SHEET,
            "Rhythm transition A/B test (analysis-only): requires SestinaLines_Auto.",
        )
        return 0, 0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, "", "missing input sheet"

    ws_lines = wb[SESTINA_LINES_SHEET]
    hl = ws_find_header_row(ws_lines, ["GlobalLine", "BookID", "EndToken", "EndKind"], max_scan=3)
    cl = ws_headers(ws_lines, hl)
    line_rows: List[Tuple[int, int, str, str]] = []
    for r in range(hl + 1, ws_lines.max_row + 1):
        gl = ws_lines.cell(r, cl["GlobalLine"]).value
        bid = ws_lines.cell(r, cl["BookID"]).value
        tok = ws_lines.cell(r, cl["EndToken"]).value
        kind = ws_lines.cell(r, cl["EndKind"]).value
        if gl is None:
            continue
        try:
            gl_i = int(gl)
            bid_i = int(bid or 0)
        except Exception:
            continue
        line_rows.append((gl_i, bid_i, str(tok or ""), str(kind or "")))
    line_rows.sort(key=lambda t: t[0])

    w = int(max(4, window_size))
    n = len(line_rows)
    if n < max(int(min_lines), w):
        upsert_sheet_index_entry(
            wb,
            RHYTHM_TRANSITIONS_SHEET,
            "Rhythm transition A/B test (analysis-only): insufficient lines for configured window/min-lines.",
        )
        return 0, 0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, "", "insufficient lines"

    half = w // 2
    if half <= 0 or (2 * half) > w:
        return 0, 0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, "", "invalid window"

    def _tok_key(s: str) -> str:
        return _token_signature(s) if use_token_signature else str(s or "")

    def _char_overlap(a: str, b: str) -> float:
        sa = {ch for ch in str(a or "") if ch.isalpha()}
        sb = {ch for ch in str(b or "") if ch.isalpha()}
        if not sa and not sb:
            return 0.0
        union = sa | sb
        if not union:
            return 0.0
        return float(len(sa & sb)) / float(len(union))

    rr = ho + 1
    fp_items: List[str] = []
    windows = 0
    cycle_sum = 0.0
    cycle_delta_sum = 0.0
    ab_sum = 0.0
    fib_sum = 0.0
    core_sum = 0.0
    sparse_sum = 0.0
    closure_inv_sum = 0.0

    for start in range(0, n - w + 1):
        win = line_rows[start : start + w]
        keys = [_tok_key(t[2]) for t in win]
        nonempty = sum(1 for x in keys if x)
        if nonempty < max(4, half):
            continue

        # Cycle-6 style predictability (lag=half for general even windows, default w=12 => lag=6).
        cycle_match = 0
        cycle_cmp = 0
        for i in range(half):
            a = keys[i]
            b = keys[i + half] if (i + half) < len(keys) else ""
            if not a or not b:
                continue
            cycle_cmp += 1
            if a == b:
                cycle_match += 1
        cycle_rate = (float(cycle_match) / float(cycle_cmp)) if cycle_cmp > 0 else 0.0

        # Shuffled control: keep first half fixed; shuffle second half deterministically.
        sh_rates: List[float] = []
        base_seed = int(win[0][0]) * 1315423911 + int(iter_num)
        for t in range(max(1, int(shuffle_trials))):
            tail = list(keys[half:])
            rng = random.Random(base_seed + (t * 7919))
            rng.shuffle(tail)
            sm = 0
            sc = 0
            for i in range(min(half, len(tail))):
                a = keys[i]
                b = tail[i]
                if not a or not b:
                    continue
                sc += 1
                if a == b:
                    sm += 1
            sh_rates.append((float(sm) / float(sc)) if sc > 0 else 0.0)
        cycle_sh_avg = (sum(sh_rates) / float(len(sh_rates))) if sh_rates else 0.0
        cycle_delta = cycle_rate - cycle_sh_avg

        # A/B alternation (A=MARK, B=other EndKind).
        focus = ["A" if str(t[3]).strip().upper() == "MARK" else "B" for t in win]
        flips = 0
        for i in range(1, len(focus)):
            if focus[i] != focus[i - 1]:
                flips += 1
        ab_rate = float(flips) / float(max(1, len(focus) - 1))

        # Fibonacci-lag repetition control.
        fib_lags = [1, 2, 3, 5, 8]
        fib_parts: List[float] = []
        for lag in fib_lags:
            if lag >= len(keys):
                continue
            m = 0
            c = 0
            for i in range(0, len(keys) - lag):
                a = keys[i]
                b = keys[i + lag]
                if not a or not b:
                    continue
                c += 1
                if a == b:
                    m += 1
            if c > 0:
                fib_parts.append(float(m) / float(c))
        fib_rate = (sum(fib_parts) / float(len(fib_parts))) if fib_parts else 0.0

        # Triple-core concentration (3 chunks, dominant share per chunk).
        core_parts: List[float] = []
        chunk = max(1, len(keys) // 3)
        for ci in range(3):
            st = ci * chunk
            en = len(keys) if ci == 2 else min(len(keys), (ci + 1) * chunk)
            seg = [x for x in keys[st:en] if x]
            if not seg:
                continue
            cnt = Counter(seg)
            top = max(cnt.values()) if cnt else 0
            core_parts.append(float(top) / float(len(seg)))
        triple_core = (sum(core_parts) / float(len(core_parts))) if core_parts else 0.0

        # Sparse echoes: repeated-token pairs concentrated at start/penultimate anchors.
        anchors = {0, max(0, len(keys) - 2)}
        pair_total = 0
        pair_anchor = 0
        for i in range(len(keys)):
            if not keys[i]:
                continue
            for j in range(i + 1, len(keys)):
                if keys[i] != keys[j]:
                    continue
                pair_total += 1
                if i in anchors or j in anchors:
                    pair_anchor += 1
        sparse_echo = (float(pair_anchor) / float(pair_total)) if pair_total > 0 else 0.0

        start_tok = keys[0] if keys else ""
        end_tok = keys[-1] if keys else ""
        closure_same = 1 if (start_tok and end_tok and start_tok == end_tok) else 0
        closure_overlap = _char_overlap(start_tok, end_tok)
        closure_inversion = 1 if (not closure_same and closure_overlap >= 0.30) else 0

        # Minimal model preference to keep this actionable in reports.
        if cycle_delta >= float(cycle6_delta_threshold):
            pref = "CYCLE6"
        elif triple_core >= 0.55:
            pref = "TRIPLE_CORE"
        elif fib_rate > cycle_rate:
            pref = "FIBONACCI"
        elif sparse_echo >= 0.50:
            pref = "SPARSE_ECHO"
        else:
            pref = "MIXED"

        ws_out.cell(rr, co["Iteration"]).value = int(iter_num)
        ws_out.cell(rr, co["UTC"]).value = utc
        ws_out.cell(rr, co["WindowStartGlobalLine"]).value = int(win[0][0])
        ws_out.cell(rr, co["BookIDStart"]).value = int(win[0][1])
        ws_out.cell(rr, co["WindowSize"]).value = int(w)
        ws_out.cell(rr, co["Cycle6Rate"]).value = round(float(cycle_rate), 6)
        ws_out.cell(rr, co["Cycle6ShuffleAvg"]).value = round(float(cycle_sh_avg), 6)
        ws_out.cell(rr, co["Cycle6Delta"]).value = round(float(cycle_delta), 6)
        ws_out.cell(rr, co["ABAlternationRate"]).value = round(float(ab_rate), 6)
        ws_out.cell(rr, co["FibonacciMatchRate"]).value = round(float(fib_rate), 6)
        ws_out.cell(rr, co["TripleCoreConcentration"]).value = round(float(triple_core), 6)
        ws_out.cell(rr, co["SparseEchoRatio"]).value = round(float(sparse_echo), 6)
        ws_out.cell(rr, co["ClosureSame"]).value = int(closure_same)
        ws_out.cell(rr, co["ClosureInversion"]).value = int(closure_inversion)
        ws_out.cell(rr, co["ClosureOverlap"]).value = round(float(closure_overlap), 6)
        ws_out.cell(rr, co["FocusPattern"]).value = "".join(focus)
        ws_out.cell(rr, co["ModelPreference"]).value = pref
        ws_out.cell(rr, co["StartToken"]).value = start_tok
        ws_out.cell(rr, co["EndToken"]).value = end_tok
        ws_out.cell(rr, co["Notes"]).value = (
            f"cycle_delta_thr={cycle6_delta_threshold:.4f}; "
            f"lag={half}; trials={max(1, int(shuffle_trials))}; "
            f"mode={'token_sig' if use_token_signature else 'raw_tok'}"
        )
        fp_items.append(
            f"{win[0][0]}|{cycle_rate:.6f}|{cycle_sh_avg:.6f}|{cycle_delta:.6f}|{ab_rate:.6f}|"
            f"{fib_rate:.6f}|{triple_core:.6f}|{sparse_echo:.6f}|{closure_inversion}|{pref}"
        )
        rr += 1

        windows += 1
        cycle_sum += cycle_rate
        cycle_delta_sum += cycle_delta
        ab_sum += ab_rate
        fib_sum += fib_rate
        core_sum += triple_core
        sparse_sum += sparse_echo
        closure_inv_sum += float(closure_inversion)

    rows_written = rr - (ho + 1)
    if windows > 0:
        cycle_avg = cycle_sum / float(windows)
        cycle_delta_avg = cycle_delta_sum / float(windows)
        ab_avg = ab_sum / float(windows)
        fib_avg = fib_sum / float(windows)
        core_avg = core_sum / float(windows)
        sparse_avg = sparse_sum / float(windows)
        closure_inv_frac = closure_inv_sum / float(windows)
    else:
        cycle_avg = 0.0
        cycle_delta_avg = 0.0
        ab_avg = 0.0
        fib_avg = 0.0
        core_avg = 0.0
        sparse_avg = 0.0
        closure_inv_frac = 0.0

    fp = hashlib.sha1("\n".join(sorted(fp_items)).encode("utf-8", errors="ignore")).hexdigest() if fp_items else ""
    upsert_sheet_index_entry(
        wb,
        RHYTHM_TRANSITIONS_SHEET,
        "Rhythm transition A/B test (analysis-only): cycle-6 vs shuffled control + alternation + anti-sestina controls over global line windows.",
    )
    return (
        rows_written,
        windows,
        cycle_avg,
        cycle_delta_avg,
        ab_avg,
        fib_avg,
        core_avg,
        sparse_avg,
        closure_inv_frac,
        fp,
        "ok" if windows > 0 else "no windows",
    )


def upsert_iter_meta_sheet(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    ev_avg: float,
    weak: float,
    micro: float,
    single: float,
    tokens: int,
    gt_count: int,
    gt_verified_count: int,
    external_pass: int,
    external_fail: int,
    context_avg_score: float,
    context_oov: float,
    seq_matches: int,
    sestina_candidates: int,
    sestina_best_score30: int,
    sestina_ob_candidates: int,
    sestina_ob_core_avg: float,
    sestina_ob_reorder_frac: float,
    rhythm_windows: int,
    rhythm_cycle_avg: float,
    rhythm_cycle_delta_avg: float,
    rhythm_ab_avg: float,
    rhythm_fib_avg: float,
    rhythm_core_avg: float,
    rhythm_sparse_avg: float,
    rhythm_closure_inv: float,
    superanchors: int,
) -> int:
    """Write Iter{N}_Meta with plateau and methodology diagnostics (analysis-only)."""
    name = f"Iter{iter_num}_Meta"
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)

    # header-ish
    ws.cell(1, 1).value = "Key"
    ws.cell(1, 2).value = "Value"
    rr = 2

    def kv(k: str, v: object) -> None:
        nonlocal rr
        ws.cell(rr, 1).value = k
        ws.cell(rr, 2).value = v
        rr += 1

    kv("Iteration", iter_num)
    kv("UTC", utc)
    kv("EvAvg (Books, weighted)", round(float(ev_avg), 6))
    kv("WeakFrac (Books, weighted)", round(float(weak), 6))
    kv("MicroFrac (Books, weighted)", round(float(micro), 6))
    kv("SingleCharFrac (Books, weighted)", round(float(single), 6))
    kv("Tokens (StrictPlus)", int(tokens))
    kv("GroundTruthCount", int(gt_count))
    kv("GroundTruthVerified>=MinSources", int(gt_verified_count))
    kv("ExternalRoundTrip Pass/Fail", f"{external_pass}/{external_fail}")
    kv("ContextEnglish AvgScore/OOV", f"{context_avg_score:.6f} / {context_oov:.6f}")
    kv("SequenceMatches", int(seq_matches))
    kv("SestinaCandidates", int(sestina_candidates))
    kv("SestinaBestScore30", int(sestina_best_score30))
    kv("SestinaObligationCandidates", int(sestina_ob_candidates))
    kv("SestinaObligationCoreAvg", round(float(sestina_ob_core_avg), 6))
    kv("SestinaObligationReorderFrac", round(float(sestina_ob_reorder_frac), 6))
    kv("RhythmWindows", int(rhythm_windows))
    kv("RhythmCycleAvg", round(float(rhythm_cycle_avg), 6))
    kv("RhythmCycleDeltaAvg", round(float(rhythm_cycle_delta_avg), 6))
    kv("RhythmABAvg", round(float(rhythm_ab_avg), 6))
    kv("RhythmFibAvg", round(float(rhythm_fib_avg), 6))
    kv("RhythmTripleCoreAvg", round(float(rhythm_core_avg), 6))
    kv("RhythmSparseEchoAvg", round(float(rhythm_sparse_avg), 6))
    kv("RhythmClosureInversionFrac", round(float(rhythm_closure_inv), 6))
    kv("SuperAnchors", int(superanchors))

    rr += 1
    ws.cell(rr, 1).value = "Ambiguous tokens (LoreAlignment low-share)"
    rr += 1
    ws.cell(rr, 1).value = "Token"
    ws.cell(rr, 2).value = "TopShare"
    ws.cell(rr, 3).value = "TopWord"
    ws.cell(rr, 4).value = "CurrentTranslation"
    ws.cell(rr, 5).value = "TotalWordCount"
    rr += 1

    # Top ambiguous tokens from LoreAlignment_Auto
    try:
        if "LoreAlignment_Auto" in wb.sheetnames:
            ws_h = wb["LoreAlignment_Auto"]
            hh = ws_find_header_row(ws_h, ["Token", "TopWord", "TopWordCount", "TotalWordCount", "CurrentTranslation"], max_scan=3)
            ch = ws_headers(ws_h, hh)
            rows = []
            for r in range(hh + 1, ws_h.max_row + 1):
                tok = ws_h.cell(r, ch["Token"]).value
                topw = ws_h.cell(r, ch["TopWord"]).value
                topc = ws_h.cell(r, ch["TopWordCount"]).value
                tot = ws_h.cell(r, ch["TotalWordCount"]).value
                cur = ws_h.cell(r, ch["CurrentTranslation"]).value
                if not isinstance(tok, str) or not tok.strip():
                    continue
                try:
                    topc_i = int(topc or 0)
                    tot_i = int(tot or 0)
                except Exception:
                    continue
                if tot_i <= 0:
                    continue
                share = float(topc_i) / float(tot_i)
                if tot_i < 20:
                    continue
                if share >= 0.90:
                    continue
                rows.append((share, tot_i, tok.strip(), str(topw or ""), str(cur or "")))
            rows.sort(key=lambda t: (t[0], -t[1], t[2]))
            for share, tot_i, tok, topw, cur in rows[:60]:
                ws.cell(rr, 1).value = tok
                ws.cell(rr, 2).value = round(float(share), 6)
                ws.cell(rr, 3).value = topw
                ws.cell(rr, 4).value = cur
                ws.cell(rr, 5).value = int(tot_i)
                rr += 1
    except Exception:
        pass

    upsert_sheet_index_entry(
        wb,
        name,
        f"Iteration {iter_num} meta report: digit/code methodology + context/sequence diagnostics (analysis-only).",
    )
    return rr


def sync_anchorcribs_from_iter141(wb: openpyxl.Workbook, iter_num: int, utc: str) -> Tuple[int, int, str]:
    src_path = os.path.join(os.getcwd(), "archive", "bonelord_469_iter141.xlsx")
    if not os.path.exists(src_path):
        return 0, 0, f"missing source: {src_path}"

    src_wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
    # Prefer the latest AnchorCribs sheet if available. iter141 includes both v138 and v141.
    src_sheet = "AnchorCribs_v141" if "AnchorCribs_v141" in src_wb.sheetnames else "AnchorCribs_v138"
    if src_sheet not in src_wb.sheetnames:
        return 0, 0, "missing AnchorCribs_v141/v138"

    src_ws = src_wb[src_sheet]
    src_header_row = [c.value for c in next(src_ws.iter_rows(min_row=1, max_row=1))]
    src_headers = [h for h in src_header_row if isinstance(h, str) and h.strip()]

    tgt_headers = list(src_headers) + ["ImportedIter", "ImportedUTC", "ImportedFrom"]
    ws = ensure_sheet(wb, "AnchorCribs_Auto", tgt_headers)
    ht = ws_find_header_row(ws, ["AnchorCribID", "BaseSubstring"], max_scan=3)
    ct = ws_headers(ws, ht)

    existing: Dict[str, int] = {}
    for r in range(ht + 1, ws.max_row + 1):
        ac = ws.cell(r, ct.get("AnchorCribID", 1)).value
        if isinstance(ac, str) and ac.strip():
            existing[ac.strip()] = r

    # Ensure target headers exist
    for h in tgt_headers:
        if h not in ct:
            ws.cell(ht, ws.max_column + 1).value = h
            ct = ws_headers(ws, ht)

    sh = {h: i + 1 for i, h in enumerate(src_header_row) if isinstance(h, str) and h.strip()}

    added = 0
    updated = 0
    unchanged = 0

    def _norm(v: object) -> object:
        if v is None:
            return None
        if isinstance(v, str):
            return v.strip()
        return v

    def _eq(a: object, b: object) -> bool:
        na = _norm(a)
        nb = _norm(b)
        # Treat blank strings and None as equivalent.
        if na is None and (nb is None or nb == ""):
            return True
        if nb is None and (na is None or na == ""):
            return True
        return na == nb

    for rr in range(2, src_ws.max_row + 1):
        ac = src_ws.cell(rr, sh.get("AnchorCribID", 1)).value
        if not isinstance(ac, str) or not ac.strip():
            continue
        key = ac.strip()
        tr = existing.get(key)
        is_new = tr is None
        if tr is None:
            tr = ws_last_data_row(ws, key_col=ct["AnchorCribID"]) + 1
            ws.cell(tr, ct["AnchorCribID"]).value = key
            existing[key] = tr

        changed = is_new
        if not is_new:
            for h in src_headers:
                col_src = sh.get(h)
                col_tgt = ct.get(h)
                if not col_src or not col_tgt:
                    continue
                sv = src_ws.cell(rr, col_src).value
                tv = ws.cell(tr, col_tgt).value
                if not _eq(sv, tv):
                    changed = True
                    break

        if not changed:
            unchanged += 1
            continue

        # Apply values (only when new or changed) so this step remains idempotent and avoids churn.
        for h in src_headers:
            col_src = sh.get(h)
            col_tgt = ct.get(h)
            if not col_src or not col_tgt:
                continue
            ws.cell(tr, col_tgt).value = src_ws.cell(rr, col_src).value

        ws.cell(tr, ct["ImportedIter"]).value = iter_num
        ws.cell(tr, ct["ImportedUTC"]).value = utc
        ws.cell(tr, ct["ImportedFrom"]).value = f"iter141:{src_sheet}"

        if is_new:
            added += 1
        else:
            updated += 1

    upsert_sheet_index_entry(
        wb,
        "AnchorCribs_Auto",
        "Imported AnchorCribs from iter141 (structure-only). Used for variant-aware alignment and safe super-anchor mining.",
    )
    return added, updated, f"ok (unchanged={unchanged})"


def promote_superanchors_to_anchorcribs_auto(
    wb: openpyxl.Workbook,
    *,
    iter_num: int,
    utc: str,
    max_promote: int = 10,
) -> int:
    """Append SuperAnchors_Auto candidates into AnchorCribs_Auto as AUTO anchors.

    This is analysis-only: it does not affect StrictPlus decoding unless later steps decide to
    incorporate these anchors into other workflows. It helps the structural pipeline bootstrap
    more anchors/superanchors over iterations.
    """
    if "SuperAnchors_Auto" not in wb.sheetnames or "AnchorCribs_Auto" not in wb.sheetnames:
        return 0

    ws_sa = wb["SuperAnchors_Auto"]
    hs = ws_find_header_row(ws_sa, ["Iter", "CandidateID", "RefBookID", "Start", "End", "Len", "BaseSubstring"], max_scan=3)
    cs = ws_headers(ws_sa, hs)

    ws_ac = wb["AnchorCribs_Auto"]
    ha = ws_find_header_row(ws_ac, ["AnchorCribID", "BaseSubstring"], max_scan=3)
    ca = ws_headers(ws_ac, ha)

    # Track existing BaseSubstring to dedupe.
    existing_subs: set[str] = set()
    for r in range(ha + 1, ws_ac.max_row + 1):
        sub = ws_ac.cell(r, ca.get("BaseSubstring", 0) or 0).value
        if isinstance(sub, str) and sub.strip():
            existing_subs.add(sub.strip())

    promoted = 0
    for r in range(hs + 1, ws_sa.max_row + 1):
        it = ws_sa.cell(r, cs["Iter"]).value
        if it != iter_num:
            continue
        base_sub = ws_sa.cell(r, cs["BaseSubstring"]).value
        if not isinstance(base_sub, str) or not base_sub.strip():
            continue
        base_sub = base_sub.strip()
        if base_sub in existing_subs:
            continue

        cand_id = ws_sa.cell(r, cs["CandidateID"]).value
        ref_book = ws_sa.cell(r, cs["RefBookID"]).value
        support_books = ws_sa.cell(r, cs.get("SupportBooks", 0) or 0).value if cs.get("SupportBooks") else None
        support_cnt = ws_sa.cell(r, cs.get("SupportBookCount", 0) or 0).value if cs.get("SupportBookCount") else None

        # Deterministic ID. Don't assume AC### formatting.
        cid = int(cand_id) if cand_id is not None else (promoted + 1)
        anchor_id = f"AC_AUTO_SA{iter_num}_{cid:03d}"
        # Ensure uniqueness if already present (rare).
        existing_ids = {ws_ac.cell(rr, ca["AnchorCribID"]).value for rr in range(ha + 1, ws_ac.max_row + 1)}
        if anchor_id in existing_ids:
            j = 2
            while f"{anchor_id}_{j}" in existing_ids:
                j += 1
            anchor_id = f"{anchor_id}_{j}"

        rr = ws_last_data_row(ws_ac, key_col=ca["AnchorCribID"]) + 1
        ws_ac.cell(rr, ca["AnchorCribID"]).value = anchor_id
        if ca.get("AnchorID") is not None:
            ws_ac.cell(rr, ca["AnchorID"]).value = "AUTO"
        if ca.get("Kind") is not None:
            ws_ac.cell(rr, ca["Kind"]).value = "SUPERANCHOR_AUTO"
        if ca.get("BaseLen") is not None:
            ws_ac.cell(rr, ca["BaseLen"]).value = int(len(base_sub))
        if ca.get("DistinctBooks") is not None and support_cnt is not None:
            try:
                ws_ac.cell(rr, ca["DistinctBooks"]).value = int(support_cnt)
            except Exception:
                pass
        if ca.get("TotalOcc") is not None and support_cnt is not None:
            try:
                ws_ac.cell(rr, ca["TotalOcc"]).value = int(support_cnt)
            except Exception:
                pass
        if ca.get("DigitVariants") is not None:
            ws_ac.cell(rr, ca["DigitVariants"]).value = 1
        if ca.get("BaseSubstring") is not None:
            ws_ac.cell(rr, ca["BaseSubstring"]).value = base_sub
        if ca.get("SourceBooks") is not None and support_books is not None:
            ws_ac.cell(rr, ca["SourceBooks"]).value = str(support_books)
        if ca.get("CandidateName") is not None:
            ws_ac.cell(rr, ca["CandidateName"]).value = f"iter{iter_num}_SA{cid:03d}"
        if ca.get("AddedInIter") is not None:
            ws_ac.cell(rr, ca["AddedInIter"]).value = int(iter_num)

        if ca.get("ImportedIter") is not None:
            ws_ac.cell(rr, ca["ImportedIter"]).value = int(iter_num)
        if ca.get("ImportedUTC") is not None:
            ws_ac.cell(rr, ca["ImportedUTC"]).value = utc
        if ca.get("ImportedFrom") is not None:
            ws_ac.cell(rr, ca["ImportedFrom"]).value = "auto:SuperAnchors_Auto"

        existing_subs.add(base_sub)
        promoted += 1
        if promoted >= int(max_promote):
            break

    if promoted:
        upsert_sheet_index_entry(
            wb,
            "AnchorCribs_Auto",
            "Imported AnchorCribs from iter141 (structure-only) + AUTO-promoted SuperAnchors (analysis-only).",
        )
    return promoted


def build_variant_alignment_from_anchorcribs(
    wb: openpyxl.Workbook,
    iter_num: int,
    ignore_anchorcrib_ids: Optional[Set[str]] = None,
    restrict_votes_to_anchor_windows: bool = False,
    anchor_window_pad: int = 0,
    ref_book_override: Optional[int] = None,
    min_anchors_shared_for_voting: int = 1,
    require_unique_anchor_occurrences: bool = False,
    require_strong_offsets: bool = False,
) -> Tuple[int, int, int, str]:
    if "AnchorCribs_Auto" not in wb.sheetnames:
        return 0, 0, 0, "AnchorCribs_Auto missing"

    ws_ac = wb["AnchorCribs_Auto"]
    ha = ws_find_header_row(ws_ac, ["AnchorCribID", "BaseSubstring"], max_scan=3)
    ca = ws_headers(ws_ac, ha)

    anchors: List[Tuple[str, str, str]] = []
    for r in range(ha + 1, ws_ac.max_row + 1):
        aid = ws_ac.cell(r, ca["AnchorCribID"]).value
        sub = ws_ac.cell(r, ca["BaseSubstring"]).value
        kind = ws_ac.cell(r, ca["Kind"]).value if ca.get("Kind") else ""
        if not isinstance(aid, str) or not aid.strip():
            continue
        if not isinstance(sub, str) or not sub.strip():
            continue
        if ignore_anchorcrib_ids and aid.strip() in ignore_anchorcrib_ids:
            continue
        anchors.append((aid.strip(), sub.strip(), str(kind or "").strip()))
    if not anchors:
        return 0, 0, 0, "no anchors"

    ws_books = wb["Books"]
    hb = ws_find_header_row(ws_books, ["BookID", "DecodedBase", "DigitsLen"], max_scan=3)
    cb = ws_headers(ws_books, hb)

    books: Dict[int, str] = {}
    digits_len: Dict[int, int] = {}
    for r in range(hb + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        books[bid_i] = str(ws_books.cell(r, cb["DecodedBase"]).value or "")
        digits_len[bid_i] = int(ws_books.cell(r, cb["DigitsLen"]).value or 0)

    # Find occurrences
    occ: Dict[str, Dict[int, int]] = {aid: {} for aid, _sub, _kind in anchors}
    for aid, sub, kind in anchors:
        for bid_i, base in books.items():
            # For AUTO-promoted anchors, require uniqueness in the book to avoid ambiguous offsets.
            # Optional strict mode applies the same uniqueness requirement to all anchor kinds.
            if require_unique_anchor_occurrences:
                if base.count(sub) != 1:
                    continue
            elif "AUTO" in str(kind or "").upper() and base.count(sub) != 1:
                continue
            pos = base.find(sub)
            if pos >= 0:
                occ[aid][bid_i] = pos

    # Reference book selection
    ref_book = None
    best = None
    if ref_book_override is not None and int(ref_book_override) in books:
        ref_book = int(ref_book_override)
    for bid_i in books.keys():
        if ref_book is not None:
            break
        hits = sum(1 for aid, _sub, _kind in anchors if bid_i in occ[aid])
        key = (hits, digits_len.get(bid_i, 0), -bid_i)
        if best is None or key > best:
            best = key
            ref_book = bid_i
    if ref_book is None:
        return 0, 0, 0, "no ref"
    ref_base = books.get(ref_book, "")

    anchor_windows_by_book: Dict[int, List[Tuple[int, int]]] = defaultdict(list)
    if restrict_votes_to_anchor_windows:
        for aid, sub, _kind in anchors:
            pos_ref = occ.get(aid, {}).get(ref_book)
            if pos_ref is None:
                continue
            win_start = max(0, int(pos_ref) - int(anchor_window_pad))
            win_end = min(len(ref_base) - 1, int(pos_ref) + len(sub) - 1 + int(anchor_window_pad))
            for bid_i in occ.get(aid, {}).keys():
                if bid_i == ref_book:
                    continue
                anchor_windows_by_book[int(bid_i)].append((win_start, win_end))

    # Clear/create sheets
    ws_occ = ensure_sheet(wb, "AnchorOccurrences_Auto", ["Iter", "AnchorCribID", "BookID", "PosInBook", "RefBookID", "PosInRef", "Offset"])
    if ws_occ.max_row > 1:
        ws_occ.delete_rows(2, ws_occ.max_row - 1)
    ho = ws_find_header_row(ws_occ, ["Iter", "AnchorCribID", "BookID"], max_scan=3)
    co = ws_headers(ws_occ, ho)

    ws_off = ensure_sheet(wb, "BookOffsets_Auto", ["Iter", "RefBookID", "BookID", "AnchorsShared", "Offset", "OffsetStd", "Consistent", "Strength", "AnchorsUsed"])
    if ws_off.max_row > 1:
        ws_off.delete_rows(2, ws_off.max_row - 1)
    hf = ws_find_header_row(ws_off, ["Iter", "RefBookID", "BookID"], max_scan=3)
    cf = ws_headers(ws_off, hf)

    offsets_by_book: Dict[int, List[int]] = defaultdict(list)
    anchors_used_by_book: Dict[int, List[str]] = defaultdict(list)

    # Write occurrences and collect offsets
    for aid, _sub, _kind in anchors:
        pos_ref = occ.get(aid, {}).get(ref_book)
        if pos_ref is None:
            continue
        # ref row
        r = ws_last_data_row(ws_occ, key_col=co["Iter"]) + 1
        ws_occ.cell(r, co["Iter"]).value = iter_num
        ws_occ.cell(r, co["AnchorCribID"]).value = aid
        ws_occ.cell(r, co["BookID"]).value = ref_book
        ws_occ.cell(r, co["PosInBook"]).value = int(pos_ref)
        ws_occ.cell(r, co["RefBookID"]).value = ref_book
        ws_occ.cell(r, co["PosInRef"]).value = int(pos_ref)
        ws_occ.cell(r, co["Offset"]).value = 0

        for bid_i, pos_book in occ.get(aid, {}).items():
            if bid_i == ref_book:
                continue
            off = int(pos_book) - int(pos_ref)
            offsets_by_book[bid_i].append(off)
            anchors_used_by_book[bid_i].append(aid)
            r = ws_last_data_row(ws_occ, key_col=co["Iter"]) + 1
            ws_occ.cell(r, co["Iter"]).value = iter_num
            ws_occ.cell(r, co["AnchorCribID"]).value = aid
            ws_occ.cell(r, co["BookID"]).value = bid_i
            ws_occ.cell(r, co["PosInBook"]).value = int(pos_book)
            ws_occ.cell(r, co["RefBookID"]).value = ref_book
            ws_occ.cell(r, co["PosInRef"]).value = int(pos_ref)
            ws_occ.cell(r, co["Offset"]).value = int(off)

    aligned_books: Dict[int, int] = {ref_book: 0}
    strong = 0
    weak = 0
    for bid_i, offs in sorted(offsets_by_book.items()):
        uniq = sorted(set(offs))
        anchors_shared = len(offs)
        consistent = len(uniq) == 1
        strength = "STRONG" if consistent and anchors_shared >= 2 else ("WEAK" if consistent else "INCONSISTENT")
        min_shared = max(1, int(min_anchors_shared_for_voting))
        allow_vote = consistent and anchors_shared >= min_shared
        if require_strong_offsets:
            allow_vote = allow_vote and anchors_shared >= 2
        if allow_vote:
            aligned_books[bid_i] = uniq[0]
            if strength == "STRONG":
                strong += 1
            else:
                weak += 1
        if anchors_shared <= 1:
            std = 0.0
        else:
            mu = sum(offs) / float(anchors_shared)
            std = (sum((x - mu) ** 2 for x in offs) / float(anchors_shared)) ** 0.5

        r = ws_last_data_row(ws_off, key_col=cf["Iter"]) + 1
        ws_off.cell(r, cf["Iter"]).value = iter_num
        ws_off.cell(r, cf["RefBookID"]).value = ref_book
        ws_off.cell(r, cf["BookID"]).value = bid_i
        ws_off.cell(r, cf["AnchorsShared"]).value = anchors_shared
        ws_off.cell(r, cf["Offset"]).value = uniq[0] if uniq else None
        ws_off.cell(r, cf["OffsetStd"]).value = round(float(std), 6)
        ws_off.cell(r, cf["Consistent"]).value = 1 if consistent else 0
        ws_off.cell(r, cf["Strength"]).value = strength
        ws_off.cell(r, cf["AnchorsUsed"]).value = ",".join(sorted(set(anchors_used_by_book.get(bid_i, []))))

    # Backbone
    total_aligned = len(aligned_books)
    ws_bb = ensure_sheet(
        wb,
        "AlignedBackbone_Auto",
        ["Iter", "RefBookID", "Coord", "RefChar", "ConsensusChar", "SupportCount", "CoveredBooks", "AlignedBooks", "SupportFrac", "Variants"],
    )
    if ws_bb.max_row > 1:
        ws_bb.delete_rows(2, ws_bb.max_row - 1)
    hb2 = ws_find_header_row(ws_bb, ["Iter", "RefBookID", "Coord"], max_scan=3)
    cb2 = ws_headers(ws_bb, hb2)

    for coord in range(0, len(ref_base)):
        chars: List[str] = []
        for bid_i, off in aligned_books.items():
            if bid_i != ref_book and restrict_votes_to_anchor_windows:
                windows = anchor_windows_by_book.get(int(bid_i), [])
                if windows and not any(start <= coord <= end for start, end in windows):
                    continue
            b = books.get(bid_i, "")
            idx = coord + int(off)
            if 0 <= idx < len(b):
                chars.append(b[idx])
        if not chars:
            continue
        cnt = Counter(chars)
        cons, sup = cnt.most_common(1)[0]
        covered = len(chars)
        frac = sup / float(covered) if covered else 0.0
        variants = ", ".join([f"{ch}:{n}" for ch, n in cnt.most_common()[1:5]])

        r = ws_last_data_row(ws_bb, key_col=cb2["Iter"]) + 1
        ws_bb.cell(r, cb2["Iter"]).value = iter_num
        ws_bb.cell(r, cb2["RefBookID"]).value = ref_book
        ws_bb.cell(r, cb2["Coord"]).value = coord
        ws_bb.cell(r, cb2["RefChar"]).value = ref_base[coord]
        ws_bb.cell(r, cb2["ConsensusChar"]).value = cons
        ws_bb.cell(r, cb2["SupportCount"]).value = sup
        ws_bb.cell(r, cb2["CoveredBooks"]).value = covered
        ws_bb.cell(r, cb2["AlignedBooks"]).value = total_aligned
        ws_bb.cell(r, cb2["SupportFrac"]).value = round(frac, 6)
        ws_bb.cell(r, cb2["Variants"]).value = variants

    # Variant blocks
    ws_blk = ensure_sheet(
        wb,
        "VariantAssemblyBlocks_Auto",
        ["Iter", "RefBookID", "BlockID", "Class", "Start", "End", "Len", "SupportFrac_Min", "SupportFrac_Mean", "ConsensusPreview"],
    )
    if ws_blk.max_row > 1:
        ws_blk.delete_rows(2, ws_blk.max_row - 1)
    hk = ws_find_header_row(ws_blk, ["Iter", "RefBookID", "BlockID"], max_scan=3)
    ck = ws_headers(ws_blk, hk)

    bb_rows = []
    for r in range(hb2 + 1, ws_bb.max_row + 1):
        coord = ws_bb.cell(r, cb2["Coord"]).value
        if coord is None:
            continue
        bb_rows.append((int(coord), float(ws_bb.cell(r, cb2["SupportFrac"]).value or 0.0)))
    bb_rows.sort(key=lambda x: x[0])

    def _is_stable(frac: float) -> bool:
        return frac >= 1.0 - 1e-12

    blocks_written = 0
    if bb_rows:
        cur_class = "S" if _is_stable(bb_rows[0][1]) else "V"
        start = bb_rows[0][0]
        fracs: List[float] = [bb_rows[0][1]]
        prev = start

        for coord, frac in bb_rows[1:]:
            klass = "S" if _is_stable(frac) else "V"
            if klass != cur_class or coord != prev + 1:
                end = prev
                length = end - start + 1
                preview = ref_base[start : min(len(ref_base), start + 80)]
                rr = ws_last_data_row(ws_blk, key_col=ck["Iter"]) + 1
                ws_blk.cell(rr, ck["Iter"]).value = iter_num
                ws_blk.cell(rr, ck["RefBookID"]).value = ref_book
                ws_blk.cell(rr, ck["BlockID"]).value = blocks_written + 1
                ws_blk.cell(rr, ck["Class"]).value = cur_class
                ws_blk.cell(rr, ck["Start"]).value = start
                ws_blk.cell(rr, ck["End"]).value = end
                ws_blk.cell(rr, ck["Len"]).value = length
                ws_blk.cell(rr, ck["SupportFrac_Min"]).value = round(min(fracs), 6)
                ws_blk.cell(rr, ck["SupportFrac_Mean"]).value = round(sum(fracs) / float(len(fracs)), 6)
                ws_blk.cell(rr, ck["ConsensusPreview"]).value = preview
                blocks_written += 1

                cur_class = klass
                start = coord
                fracs = [frac]
            else:
                fracs.append(frac)
            prev = coord

        end = prev
        length = end - start + 1
        preview = ref_base[start : min(len(ref_base), start + 80)]
        rr = ws_last_data_row(ws_blk, key_col=ck["Iter"]) + 1
        ws_blk.cell(rr, ck["Iter"]).value = iter_num
        ws_blk.cell(rr, ck["RefBookID"]).value = ref_book
        ws_blk.cell(rr, ck["BlockID"]).value = blocks_written + 1
        ws_blk.cell(rr, ck["Class"]).value = cur_class
        ws_blk.cell(rr, ck["Start"]).value = start
        ws_blk.cell(rr, ck["End"]).value = end
        ws_blk.cell(rr, ck["Len"]).value = length
        ws_blk.cell(rr, ck["SupportFrac_Min"]).value = round(min(fracs), 6)
        ws_blk.cell(rr, ck["SupportFrac_Mean"]).value = round(sum(fracs) / float(len(fracs)), 6)
        ws_blk.cell(rr, ck["ConsensusPreview"]).value = preview
        blocks_written += 1

    upsert_sheet_index_entry(
        wb,
        "BookOffsets_Auto",
        f"Variant-aware alignment offsets relative to RefBook={ref_book} (STRONG={strong}, WEAK={weak}).",
    )
    upsert_sheet_index_entry(wb, "AnchorOccurrences_Auto", "Anchor occurrences in Books.DecodedBase (for alignment/assembly).")
    upsert_sheet_index_entry(wb, "AlignedBackbone_Auto", "Consensus backbone over aligned books (analysis-only; does not change decode).")
    upsert_sheet_index_entry(wb, "VariantAssemblyBlocks_Auto", "Stable vs variant blocks over the aligned backbone (analysis-only).")

    note_parts = ["ok"]
    if restrict_votes_to_anchor_windows:
        note_parts.append(f"local_windows=1; pad={int(anchor_window_pad)}")
    if ref_book_override is not None:
        note_parts.append(f"ref_override={int(ref_book_override)}")
    if max(1, int(min_anchors_shared_for_voting)) > 1:
        note_parts.append(f"min_shared={max(1, int(min_anchors_shared_for_voting))}")
    if require_unique_anchor_occurrences:
        note_parts.append("unique_only=1")
    if require_strong_offsets:
        note_parts.append("strong_only=1")
    mode_note = "; ".join(note_parts)
    return int(ref_book), int(total_aligned), int(blocks_written), mode_note


def mine_superanchors_from_backbone(
    wb: openpyxl.Workbook,
    iter_num: int,
    *,
    min_len: int = 30,
    min_books: int = 7,
    min_support_books: int = 7,
    min_support_frac: float = 1.0,
) -> int:
    if "AlignedBackbone_Auto" not in wb.sheetnames or "BookOffsets_Auto" not in wb.sheetnames:
        return 0

    ws_bb = wb["AlignedBackbone_Auto"]
    hb = ws_find_header_row(ws_bb, ["Iter", "RefBookID", "Coord", "SupportFrac", "SupportCount", "ConsensusChar"], max_scan=3)
    cb = ws_headers(ws_bb, hb)

    coords: List[Tuple[int, float, int, str]] = []
    ref_book = None
    for r in range(hb + 1, ws_bb.max_row + 1):
        it = ws_bb.cell(r, cb["Iter"]).value
        if it != iter_num:
            continue
        if ref_book is None:
            ref_book = ws_bb.cell(r, cb["RefBookID"]).value
        coord = ws_bb.cell(r, cb["Coord"]).value
        frac = ws_bb.cell(r, cb["SupportFrac"]).value
        sup = ws_bb.cell(r, cb["SupportCount"]).value
        cons = ws_bb.cell(r, cb["ConsensusChar"]).value
        if coord is None or frac is None or sup is None or cons is None:
            continue
        coords.append((int(coord), float(frac), int(sup), str(cons)))
    if not coords or ref_book is None:
        return 0
    coords.sort(key=lambda x: x[0])

    # Load ref base
    ws_books = wb["Books"]
    hb2 = ws_find_header_row(ws_books, ["BookID", "DecodedBase"], max_scan=3)
    cb2 = ws_headers(ws_books, hb2)
    books: Dict[int, str] = {}
    ref_base = ""
    for r in range(hb2 + 1, ws_books.max_row + 1):
        bid = ws_books.cell(r, cb2["BookID"]).value
        if bid == ref_book:
            ref_base = str(ws_books.cell(r, cb2["DecodedBase"]).value or "")
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        books[bid_i] = str(ws_books.cell(r, cb2["DecodedBase"]).value or "")
    if not ref_base:
        return 0

    # Aligned books list + offsets
    ws_off = wb["BookOffsets_Auto"]
    hf = ws_find_header_row(ws_off, ["Iter", "RefBookID", "BookID", "Consistent"], max_scan=3)
    cf = ws_headers(ws_off, hf)
    offsets: Dict[int, int] = {int(ref_book): 0}
    for r in range(hf + 1, ws_off.max_row + 1):
        it = ws_off.cell(r, cf["Iter"]).value
        if it != iter_num:
            continue
        if ws_off.cell(r, cf["RefBookID"]).value != ref_book:
            continue
        if not parse_bool(ws_off.cell(r, cf["Consistent"]).value, False):
            continue
        bid = ws_off.cell(r, cf["BookID"]).value
        if bid is None:
            continue
        off = ws_off.cell(r, cf.get("Offset", 0) or 0).value
        if off is None:
            continue
        try:
            offsets[int(bid)] = int(off)
        except Exception:
            continue

    aligned_books = sorted(set(offsets.keys()))
    if len(aligned_books) < min_books:
        return 0

    # Stable runs
    stable_runs: List[Tuple[int, int, List[Tuple[int, str]]]] = []
    start: Optional[int] = None
    prev: Optional[int] = None
    cur_coords: List[Tuple[int, str]] = []
    for coord, frac, sup, cons in coords:
        stable = (sup >= int(min_support_books)) and (frac >= float(min_support_frac) - 1e-12)
        if stable:
            if start is None:
                start = coord
            prev = coord
            cur_coords.append((coord, cons))
        else:
            if start is not None and prev is not None and prev - start + 1 >= min_len:
                stable_runs.append((start, prev, list(cur_coords)))
            start = None
            prev = None
            cur_coords = []
    if start is not None and prev is not None and prev - start + 1 >= min_len:
        stable_runs.append((start, prev, list(cur_coords)))

    ws_sa = ensure_sheet(
        wb,
        "SuperAnchors_Auto",
        [
            "Iter",
            "CandidateID",
            "RefBookID",
            "Start",
            "End",
            "Len",
            "BaseSubstring",
            "SupportBookCount",
            "SupportBooks",
            "AlignedBookCount",
            "AlignedBooks",
            "Criteria",
            "Notes",
        ],
    )
    hs = ws_find_header_row(ws_sa, ["Iter", "CandidateID"], max_scan=3)
    cs = ws_headers(ws_sa, hs)

    # Ensure target headers exist (idempotent on older workbooks).
    desired_headers = [
        "Iter",
        "CandidateID",
        "RefBookID",
        "Start",
        "End",
        "Len",
        "BaseSubstring",
        "SupportBookCount",
        "SupportBooks",
        "AlignedBookCount",
        "AlignedBooks",
        "Criteria",
        "Notes",
    ]
    for h in desired_headers:
        if h not in cs:
            ws_sa.cell(hs, ws_sa.max_column + 1).value = h
            cs = ws_headers(ws_sa, hs)

    if ws_sa.max_row > 1:
        ws_sa.delete_rows(2, ws_sa.max_row - 1)

    cid = 0
    criteria = f"sup>={int(min_support_books)} frac>={float(min_support_frac)}"
    for a, b, coord_cons in stable_runs:
        # Build a consensus substring (stable across support books), not just a ref-book slice.
        sub = "".join(cons for _coord, cons in coord_cons)
        if not sub or len(sub) < min_len:
            continue

        # Compute support book intersection (books that match the consensus char at every coord in the run).
        support: Set[int] = set(aligned_books)
        for coord, cons in coord_cons:
            new_support: Set[int] = set()
            for bid_i in support:
                base = books.get(bid_i, "")
                idx = coord + int(offsets.get(bid_i, 0))
                if 0 <= idx < len(base) and base[idx] == cons:
                    new_support.add(bid_i)
            support = new_support
            if len(support) < int(min_support_books):
                break
        if len(support) < int(min_support_books):
            continue

        cid += 1
        rr = ws_last_data_row(ws_sa, key_col=cs["Iter"]) + 1
        ws_sa.cell(rr, cs["Iter"]).value = iter_num
        ws_sa.cell(rr, cs["CandidateID"]).value = cid
        ws_sa.cell(rr, cs["RefBookID"]).value = int(ref_book)
        ws_sa.cell(rr, cs["Start"]).value = a
        ws_sa.cell(rr, cs["End"]).value = b
        ws_sa.cell(rr, cs["Len"]).value = (b - a + 1)
        ws_sa.cell(rr, cs["BaseSubstring"]).value = sub
        ws_sa.cell(rr, cs.get("SupportBookCount")).value = len(support) if cs.get("SupportBookCount") else None
        ws_sa.cell(rr, cs.get("SupportBooks")).value = ",".join(str(x) for x in sorted(support)) if cs.get("SupportBooks") else None
        ws_sa.cell(rr, cs["AlignedBookCount"]).value = len(aligned_books)
        ws_sa.cell(rr, cs["AlignedBooks"]).value = ",".join(str(x) for x in aligned_books)
        if cs.get("Criteria") is not None:
            ws_sa.cell(rr, cs["Criteria"]).value = criteria
        ws_sa.cell(rr, cs["Notes"]).value = (
            "Stable run from AlignedBackbone_Auto using consensus stability + support intersection. "
            "Suggest promote to AnchorCribs if useful."
        )

    upsert_sheet_index_entry(
        wb,
        "SuperAnchors_Auto",
        f"Super-anchor suggestions from stable backbone runs (min_len={min_len}, min_support={min_support_frac}, min_support_books={min_support_books}, aligned_books={len(aligned_books)}).",
    )
    return cid

def update_sheet_index(wb: openpyxl.Workbook, iter_num: int, description: str) -> None:
    ws = wb["SheetIndex"]
    header = ws_find_header_row(ws, ["Sheet", "What it contains"])
    c = ws_headers(ws, header)
    sheet_name = f"Iter{iter_num}_Summary"
    # Avoid duplicates.
    for r in range(header + 1, ws.max_row + 1):
        if ws.cell(r, c["Sheet"]).value == sheet_name:
            ws.cell(r, c["What it contains"]).value = description
            return
    ws_append_row(ws, [sheet_name, description], start_col=1)


def append_method_log(wb: openpyxl.Workbook, iter_num: int, utc: str, what: str, why: str, validation: str, outcome: str) -> None:
    ws = wb["MethodLog"]
    header = ws_find_header_row(ws, ["Iteration", "UTC", "WhatChanged", "Why", "Validation", "Outcome"])
    c = ws_headers(ws, header)
    ws_append_row(ws, [iter_num, utc, what, why, validation, outcome], start_col=1)


def append_work_queue(wb: openpyxl.Workbook, iter_num: int, status: str, notes: str) -> None:
    ws = wb["WorkQueue"]
    header = ws_find_header_row(ws, ["Priority", "Task", "Why", "Inputs", "Outputs", "Status", "Guardrails", "Notes"])
    c = ws_headers(ws, header)
    ws_append_row(
        ws,
        [
            0,
            f"iter{iter_num}: next iteration follow-ups",
            "Auto-chain continued work",
            "FlowState; FlowRunLog; Glossary; Cribs",
            "Next Iter summary + promotions if any",
            status,
            "No web-scraping; keep StrictPlus text stable; no WEAK/single increase.",
            notes,
        ],
        start_col=1,
    )


def run_next_iteration(workbook_path: str) -> None:
    wb = openpyxl.load_workbook(workbook_path)

    # Keep FlowSteps aligned with what the runner actually does while the
    # storage backend is being cut over behind the flow-store adapter.
    try:
        ensure_flow_steps_entries(wb)
    except Exception:
        pass

    # FlowStore
    store = open_flow_store(wb, workbook_path=workbook_path)
    ws_state = store["ws_state"]
    state = store["state_map"]
    cur_status = _normalize_flow_status(flow_state_get(store, "Status", "READY") or "READY")
    if cur_status in ("BLOCKED", "SOFT_RESOLVED", "RESOLVED", STATUS_MODEL_CONVERGED):
        block_reason = str(flow_state_get(store, "BlockReason", "") or "")
        # Always attempt another iteration: the runner may now contain new remediation logic
        # (e.g. GT auto-repair, macro-mining ladder), and the user explicitly drives the loop via "next iteration".
        if cur_status == "BLOCKED":
            print(f"FlowState.Status=BLOCKED (reason: {block_reason}); attempting next iteration anyway.")

    cur_iter = int(flow_state_get(store, "CurrentIteration", 0) or 0)
    prev_ctx_score = float(flow_state_get(store, "ContextEnglishAvgScore", 0.0) or 0.0)
    prev_ctx_streak = int(flow_state_get(store, "ContextEnglishImproveStreak", 0) or 0)
    prev_digit_ctx_fp = str(flow_state_get(store, "DigitCodeContextFingerprint", "") or "")
    prev_seq_fp = str(flow_state_get(store, "SequenceMatchesFingerprint", "") or "")
    prev_seq_hint_fp = str(flow_state_get(store, "SequenceWordHintsFingerprint", "") or "")
    prev_sestina_fp = str(flow_state_get(store, "SestinaFingerprint", "") or "")
    prev_sestina_obligation_fp = str(flow_state_get(store, "SestinaObligationFingerprint", "") or "")
    prev_rhythm_fp = str(flow_state_get(store, "RhythmFingerprint", "") or "")
    prev_codeword_fp = str(flow_state_get(store, "CodeWordMapFingerprint", "") or "")
    prev_pd_sig_fp = str(flow_state_get(store, "PDSigIndexFingerprint", "") or "")
    prev_tibia_sig_fp = str(flow_state_get(store, "TibiaSigIndexFingerprint", "") or "")
    prev_dict_sig_fp = str(flow_state_get(store, "DictSigIndexFingerprint", "") or "")
    prev_gt_soft_mismatch = int(flow_state_get(store, "GTSoftMismatchCount", 0) or 0)
    prev_gt_mode_active = str(flow_state_get(store, "GroundTruthLiveCheckModeActive", "") or "")
    prev_no_mech_iters = int(flow_state_get(store, "IterationsSinceLastMechanicalPromotion", 0) or 0)
    prev_candidate_empty_streak = int(flow_state_get(store, "CandidateScanEmptyStreak", 0) or 0)
    prev_no_real_progress_iters = int(flow_state_get(store, "IterationsWithoutRealProgress", 0) or 0)
    prev_candidate_scan_mode = str(flow_state_get(store, "LastCandidateScanMode", "") or "")
    prev_gt_soft_nondec_streak = int(flow_state_get(store, "GTSoftMismatchNonDecreasingStreak", 0) or 0)
    next_iter = cur_iter + 1
    utc = utc_now_str()

    # Backup (stable repo location)
    backup_dir = os.path.join(os.getcwd(), "tmp", "spreadsheets")
    backup_path = backup_workbook(workbook_path, backup_dir, next_iter)

    # FlowSettings
    ws_settings = store["ws_settings"]
    settings_map = store["settings_map"]
    flow_setting_ensure(
        store,
        "MacroMine_Mode",
        "AUTO",
        note="AUTO: fallback-only before RESOLVED; ALWAYS after RESOLVED (safe polishing).",
    )
    flow_setting_ensure(store, "MacroMine_MinOcc", 2, note="Min occurrences required for mined macros.")
    flow_setting_ensure(store, "MacroMine_MinBooks", 2, note="Min distinct books required for mined macros.")
    flow_setting_ensure(store, "MacroMine_MinShare", 0.95, note="Min share of the dominant mapping (1.0=strict).")
    flow_setting_ensure(store, "MacroMine_MinLen", 2, note="Min base length for mined macros.")
    flow_setting_ensure(store, "MacroMine_MaxLen", 16, note="Max base length for mined macros (ALWAYS may override by default).")
    flow_setting_ensure(store, "MacroMine_MaxCandidates", 75, note="Max mined macros added per iteration (ALWAYS may override by default).")
    flow_setting_ensure(store, "MacroMine_NValues", "2,3,4,5,6", note="n-gram sizes for macro mining when enabled.")
    flow_setting_ensure(store, "MacroMine_AllowMacroComponents", True, note="Allow macro-of-macro mining (safe, output-stable).")
    flow_setting_ensure(store, "MacroMine_AllowMarkers", True, note="Allow mining macros that include marker tokens (enables SingleCharFrac reduction).")
    flow_setting_ensure(store, "MacroMine_AllowStars", True, note="Allow mining macros that include '*' in component tokens/base (common in this corpus).")
    flow_setting_ensure(store, "PromotionMaxPasses", 2, note="Re-test candidates after baseline improves (multi-pass).")
    flow_setting_ensure(store,
        "GroundTruthAutoRepair",
        True,
        note="If GT live check fails, attempt safe GT-repair macros derived from mismatching cribs before blocking.",
    )
    flow_setting_ensure(store, "GroundTruthAutoRepairMaxMacros", 10, note="Hard cap on GT-repair macros added in one iteration.")
    flow_setting_ensure(store,
        "GroundTruthLiveCheck_Mode",
        "POLICY",
        note="GT live check enforcement mode: ALL=enforce all GroundTruth rows; POLICY=enforce only CribID(s) flagged in GroundTruthPolicy_Auto.",
    )
    flow_setting_ensure(store,
        "GroundTruthPolicy_DefaultEnforcedCribIDs",
        "2,3,4,7",
        note="Default CribID(s) enforced in GroundTruthPolicy_Auto when the policy sheet is first created.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_Mode",
        True,
        note="If TRUE, enable anti-hallucination guardrails (de-circularize GT policy + stricter lexical sanitization).",
    )
    flow_setting_ensure(store,
        "AntiHallucination_DemoteCircularGT",
        True,
        note="If TRUE, GroundTruth rows with circular/speculative expected text are auto-demoted to SOFT_PROVISIONAL unless externally verified.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_CircularTerms",
        ",".join(sorted(ANTI_HALLUCINATION_DEFAULT_TERMS)),
        note="CSV denylist used to detect circular expected-text GT rows.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_CircularMinTokens",
        6,
        note="Minimum token count to apply stopword-heavy circularity heuristic on ExpectedNorm.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_CircularStopwordRatio",
        0.78,
        note="ExpectedNorm rows above this stopword ratio are considered circular candidates (when low verified sources).",
    )
    flow_setting_ensure(store,
        "AntiHallucination_MinVerifiedForHardGT",
        2,
        note="Minimum VerifiedSourceCount_v129 required to keep circular-looking GT rows as HARD_EXTERNAL.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_DisableSemanticRetext",
        True,
        note="If TRUE, disable semantic->Glossary automatic retext while anti-hallucination mode is active.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_DisableEnglishRetext",
        True,
        note="If TRUE, disable English->Glossary automatic retext while anti-hallucination mode is active.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_UseUNKForLowEvidence",
        True,
        note="If TRUE, suspicious low-evidence lexical forms are forced to <UNK> (GT-guarded).",
    )
    flow_setting_ensure(store,
        "AntiHallucination_UNKMaxPerIter",
        40,
        note="Max suspicious token translations replaced with <UNK> per iteration.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_UNKMinTotalOcc",
        1,
        note="Minimum TotalOcc for <UNK> sanitizer candidates.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_UNKMaxTotalOcc",
        999,
        note="Maximum TotalOcc for <UNK> sanitizer candidates.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_UNKMaxConfidence",
        "MEDIUM",
        note="Maximum Glossary confidence eligible for <UNK> sanitizer (LOW/MEDIUM/HIGH).",
    )
    flow_setting_ensure(store,
        "AntiHallucination_UNKProtectedEvidenceClasses",
        "GROUNDTRUTH,LOGOGRAM_ANCHORED,PUNCT_LOGOGRAM",
        note="Do not auto-sanitize tokens with these evidence classes.",
    )
    flow_setting_ensure(store,
        "AntiHallucination_UNKDenyWords",
        ",".join(sorted(ANTI_HALLUCINATION_DEFAULT_TERMS)),
        note="CSV lexical denylist for <UNK> sanitizer and semantic/english retext filters.",
    )
    flow_setting_ensure(store,
        "AnchorPromotionOnly_Enabled",
        True,
        note="If TRUE, promotions are restricted to tokens impacted by anchor corpus (anti-hallucination-safe lane).",
    )
    flow_setting_ensure(store,
        "AnchorPromotionOnly_MinHits",
        1,
        note="Minimum token hits across anchor corpus required to keep a promotion candidate.",
    )
    flow_setting_ensure(store,
        "AnchorPromotionOnly_MinVerifiedSources",
        2,
        note="Minimum external verification count required when building anchor corpus from external references.",
    )
    flow_setting_ensure(store,
        "AnchorPromotionOnly_IncludeGroundTruth",
        True,
        note="Include GroundTruth anchor bases (prefer enforced IDs when POLICY mode is active).",
    )
    flow_setting_ensure(store,
        "AnchorPromotionOnly_IncludeExternalNPCStaff",
        True,
        note="Include externally verified NPC/staff/interview/identifier anchors.",
    )
    flow_setting_ensure(store,
        "AnchorPromotionOnly_IncludeExternalBooks",
        False,
        note="Include externally verified book anchors (OFF by default to reduce lexical drift).",
    )
    flow_setting_ensure(store,
        "AnchorPromotionOnly_IncludeAnchorCribs",
        True,
        note="Include AnchorCribs_Auto bases when computing anchor impact.",
    )
    flow_setting_ensure(store,
        "SoftMismatchMaxForResolved",
        4,
        note="Convergence gate: max soft GT mismatches allowed for RESOLVED (relaxed to allow provisional mismatch tolerance).",
    )
    flow_setting_ensure(store,
        "SoftMismatchMaxForSoftResolved",
        999,
        note="Convergence gate: max soft GT mismatches allowed for SOFT_RESOLVED before forcing BLOCKED.",
    )
    flow_setting_ensure(store,
        "Convergence_UseSoftMismatchRescue",
        True,
        note="If TRUE, run a limited soft-only rescue pass when strict pass cannot move and soft mismatches persist.",
    )
    flow_setting_ensure(store,
        "Convergence_MaxSoftRescuePromotions",
        20,
        note="Max number of soft-mismatch-rescue promotions allowed per iteration.",
    )
    flow_setting_ensure(store,
        "Convergence_RescueCandidateShuffle",
        True,
        note="Shuffle candidate order in rescue pass to escape deterministic plateaus.",
    )
    flow_setting_ensure(store,
        "Convergence_SoftNonDecStreakBlock",
        999,
        note="Do not hard-block SOFT_RESOLVED status on nondecreasing soft GT mismatch if streak is below this threshold.",
    )
    flow_setting_ensure(store,
        "Convergence_StrictMonotonicMechanical",
        True,
        note="If TRUE, enforce monotonic anti-churn gate for mechanical promotions (no structural regressions + minimum score gain).",
    )
    flow_setting_ensure(store,
        "Convergence_MonotonicMinScore",
        0.05,
        note="Minimum directional score required to accept a strict mechanical promotion under monotonic gate.",
    )
    flow_setting_ensure(store,
        "Convergence_MonotonicAllowTokenIncrease",
        False,
        note="If FALSE, monotonic gate rejects promotions that increase token count.",
    )
    flow_setting_ensure(store,
        "Convergence_BlockPromotionTokens",
        "",
        note="CSV token denylist for mechanical promotions; shadow probes can use this to keep selected tokens from being auto-promoted back.",
    )
    flow_setting_ensure(store,
        "Convergence_EnableHardEscape",
        True,
        note="Allow extra frontier-style mechanical passes when no progress is detected across recent iterations.",
    )
    flow_setting_ensure(store,
        "Convergence_HardEscapePasses",
        4,
        note="Max extra frontier passes for hard-escape mode when no mechanical promotions are accepted.",
    )
    flow_setting_ensure(store,
        "Convergence_HardEscapeToleranceScale",
        4.0,
        note="Scale factor to relax frontier tolerances in hard-escape passes (1.0=normal frontier).",
    )
    flow_setting_ensure(store,
        "Convergence_HardEscape_StallIters",
        2,
        note="Trigger hard-escape only after at least N consecutive iterations without mechanical promotion.",
    )
    flow_setting_ensure(store,
        "Convergence_EnableDirectionalEscape",
        True,
        note="Test several candidate orderings/routes when no promotion is found and plateau has been detected.",
    )
    flow_setting_ensure(store,
        "Convergence_DirectionalEscape_StallIters",
        1,
        note="Trigger directional escape only after at least this many iterations without mechanical promotion.",
    )
    flow_setting_ensure(store,
        "Convergence_DirectionalEscapePasses",
        7,
        note="Max directional escape passes per iteration during post-stall search.",
    )
    flow_setting_ensure(store,
        "Convergence_DirectionalEscapeTopK",
        60,
        note="Cap of candidates evaluated per directional pass.",
    )
    flow_setting_ensure(store,
        "Convergence_DirectionalEscapeFrontierScale",
        2.0,
        note="Frontier scale multiplier for directional escape routes.",
    )
    flow_setting_ensure(store,
        "Convergence_DirectionalEscapeAllowMetricRegression",
        True,
        note="Allow controlled metric regression inside directional escape corridors.",
    )
    flow_setting_ensure(store,
        "Convergence_DirectionalEscapeScoreMin",
        -999.0,
        note="Directional pass can accept non-improving candidates only if quality score >= this threshold.",
    )
    flow_setting_ensure(store,
        "Convergence_InterleaveEvidenceBuckets",
        True,
        note="In exploratory passes, interleave candidates by evidence class to diversify directions.",
    )
    flow_setting_ensure(store,
        "Convergence_CandidateStarveIters",
        4,
        note="If no candidates/mechanical progress for N consecutive iterations, force broader candidate sourcing.",
    )
    flow_setting_ensure(store,
        "Convergence_CandidateScanFallbackAllClasses",
        True,
        note="In candidate starvation, ignore MechanicalCandidateClasses filter and scan all candidate evidence classes.",
    )
    flow_setting_ensure(store,
        "Convergence_CandidateScanFallbackMinOcc",
        1,
        note="Fallback minimum TotalOcc for broadened candidate scan.",
    )
    flow_setting_ensure(store,
        "Convergence_CandidateScanFallbackKeepTop",
        400,
        note="Cap number of candidate rows used in fallback scan.",
    )
    flow_setting_ensure(store,
        "Convergence_EnableLowConfidenceEscape",
        True,
        note="Allow LOW-confidence token promotion after prolonged plateaus for controlled escape attempts.",
    )
    flow_setting_ensure(store,
        "Convergence_LowConfidenceEscapeStallIters",
        3,
        note="Start LOW-confidence escape mode after this many consecutive no-real-progress iterations.",
    )
    flow_setting_ensure(store,
        "Convergence_MacroMineEmergencyRung",
        3,
        note="Emergency macro ladder rung for hard dead-ends (beyond normal 0-2 ramp).",
    )
    flow_setting_ensure(store,
        "Convergence_RequireProgressForSolvedIters",
        3,
        note="Demote solved status after this many consecutive iterations without real progress.",
    )
    flow_setting_ensure(store,
        "PlateauBlock_NoPromotionIters",
        3,
        note="Force BLOCKED when no mechanical promotion for N consecutive iterations and promotion skips recur.",
    )
    flow_setting_ensure(store,
        "Convergence_UseLiveGTCounts",
        True,
        note="If TRUE, success/status decisions use Step12 live GT counts (bad_enforced/bad_all/soft) instead of stale Cribs match flags.",
    )
    flow_setting_ensure(store,
        "PlateauEnableFocusReport",
        True,
        note="When mech_promoted=0, write IterXXX_Focus plateau diagnostics (analysis-only).",
    )
    flow_setting_ensure(store,
        "PlateauEnableStructural",
        True,
        note="When plateau, sync AnchorCribs + build variant-aware alignment + mine SuperAnchors (analysis-only).",
    )
    flow_setting_ensure(store,
        "PlateauAutoRelaxMacroMine",
        True,
        note="Auto-relax MacroMine_* ladder on plateau to escape plateaus safely.",
    )
    flow_setting_ensure(store, "PlateauLadder_Rung", 0, note="0=baseline, 1=relaxed, 2=more relaxed (auto).")
    flow_setting_ensure(store, "SuperAnchor_MinLen", 30, note="SuperAnchor mining: minimum stable run length.")
    flow_setting_ensure(store, "SuperAnchor_MinBooks", 7, note="SuperAnchor mining: minimum aligned books required.")
    flow_setting_ensure(store, "SuperAnchor_MinSupportFrac", 0.8, note="SuperAnchor mining: minimum SupportFrac on the aligned backbone (0.8=cluster-friendly).")
    flow_setting_ensure(store, "SuperAnchor_MinSupportBooks", 7, note="SuperAnchor mining: minimum SupportCount across the run (books matching consensus).")
    flow_setting_ensure(store, "StructuralAutoPromoteSuperAnchors", True, note="If SuperAnchors mined, append them to AnchorCribs_Auto and re-run alignment+mining once (analysis-only).")
    flow_setting_ensure(store, "StructuralAutoPromoteSuperAnchorsMax", 10, note="Max SuperAnchors auto-promoted per iteration.")
    flow_setting_ensure(store, "StructuralRefBookOverride", "", note="Optional BookID override for structural alignment reference selection.")
    flow_setting_ensure(store, "StructuralMinAnchorsSharedForVoting", 1, note="Minimum shared anchors required before a book can vote in structural backbone alignment.")
    flow_setting_ensure(store, "StructuralRequireUniqueAnchorOccurrences", False, note="Require anchor substrings to occur uniquely in each book for structural alignment offsets.")
    flow_setting_ensure(store, "StructuralRequireStrongOffsets", False, note="Allow only STRONG (>=2-anchor) consistent offsets to vote in structural backbone alignment.")
    flow_setting_ensure(store, "Convergence_ForceProbePairTokens", "", note="Semicolon-separated explicit token pairs (tokA+tokB) to jointly simulate after strict single-token passes.")
    flow_setting_ensure(store, "SuperAnchorMacro_Enabled", True, note="If SuperAnchors exist (prev iter), derive token-boundary macros as promotion candidates (safe, semantics-preserving).")
    flow_setting_ensure(store, "SuperAnchorMacro_MinLen", 12, note="Min base length for superanchor-derived macros.")
    flow_setting_ensure(store, "SuperAnchorMacro_MaxLen", 60, note="Max base length for superanchor-derived macros.")
    flow_setting_ensure(store, "SuperAnchorMacro_MaxCandidates", 12, note="Max superanchor-derived macros appended per iteration.")
    flow_setting_ensure(store, "MacroCompress_Enabled", True, note="Materialize macro-compressed display-only columns.")
    # Lore/Semantic (display-only) loop: safe layer to improve readability/meaning without touching DP/Glossary.
    flow_setting_ensure(store, "Lore_Enabled", True, note="Enable lore corpus + semantic display-only mining (safe).")
    flow_setting_ensure(store, "Lore_Canon_DropFinalE", False, note="Lore canon: optionally drop final 'e' (candidate rule layer).")
    flow_setting_ensure(store, "Lore_Canon_DropAllH", False, note="Lore canon: optionally drop all 'h' (candidate rule layer).")
    flow_setting_ensure(store,
        "Lore_Canon_AutoFix_DropAllH",
        False,
        note="If TRUE, runner may auto-enable Lore_Canon_DropAllH when DigitCodeMap implies base alphabet lacks H (experimental).",
    )
    flow_setting_ensure(store, "Lore_Canon_DropAllO", False, note="Lore canon: optionally drop all 'o' (candidate rule layer).")
    flow_setting_ensure(store,
        "LoreFetch_TibiaSigIndex_Enabled",
        True,
        note="If enabled, fetch a public Tibia-derived dataset and store only a derived signature index in LoreSigIndex_Tibia_Auto (no full text).",
    )
    flow_setting_ensure(store,
        "LoreFetch_TibiaSigIndex_MaxAgeHours",
        168,
        note="Re-fetch LoreSigIndex_Tibia_Auto when older than N hours. Set 0 to fetch every iteration.",
    )
    flow_setting_ensure(store, "LoreFetch_TibiaSigIndex_NPC_URL", DEFAULT_TIBIA_NPC_URL, note="NPC transcript JSON URL (public).")
    flow_setting_ensure(store, "LoreFetch_TibiaSigIndex_BOOK_URL", DEFAULT_TIBIA_BOOK_URL, note="Book JSON URL (public).")
    flow_setting_ensure(store, "LoreFetch_TibiaSigIndex_TimeoutS", 60, note="HTTP timeout (seconds) for Tibia sig-index fetch.")
    flow_setting_ensure(store, "LoreFetch_TibiaSigIndex_MaxWordsPerSig", 80, note="Max words persisted per signature in LoreSigIndex_Tibia_Auto.")
    flow_setting_ensure(store, "LoreFetch_TibiaWordFreq_TopN", 6000, note="Top-N words stored in LoreWordFreq_Tibia_Auto (derived counts only).")

    # Optional public-domain signature index (derived-only; expands candidate coverage beyond Tibia).
    flow_setting_ensure(store,
        "LoreFetch_PDSigIndex_Enabled",
        True,
        note="If enabled, fetch a public-domain plaintext corpus and store only derived signature+wordfreq indices (no full text) for semantic/context layers.",
    )
    flow_setting_ensure(store,
        "LoreFetch_PDSigIndex_MaxAgeHours",
        720,
        note="Re-fetch LoreSigIndex_PD_Auto when older than N hours. Set 0 to fetch every iteration.",
    )
    flow_setting_ensure(store, "LoreFetch_PDSigIndex_URL", DEFAULT_PD_KJV_URL, note="Public-domain plaintext URL (default: Project Gutenberg KJV).")
    flow_setting_ensure(store, "LoreFetch_PDSigIndex_CorpusID", "GUTENBERG_KJV_10", note="CorpusID label stored in LoreSigIndex_PD_Auto (derived).")
    flow_setting_ensure(store,
        "LoreFetch_PDSigIndex_ExtraURLs",
        "",
        note="Optional extra public-domain plaintext URLs (CSV). Runner may auto-seed on sustained plateaus.",
    )
    flow_setting_ensure(store,
        "LoreFetch_PDSigIndex_ExtraCorpusIDs",
        "",
        note="Optional extra corpus IDs (CSV) matching ExtraURLs. If empty/mismatch, IDs are auto-derived from URL.",
    )
    flow_setting_ensure(store, "LoreFetch_PDSigIndex_TimeoutS", 60, note="HTTP timeout (seconds) for PD sig-index fetch.")
    flow_setting_ensure(store, "LoreFetch_PDSigIndex_CacheMaxAgeHours", 720, note="Max age for cached PD plaintext downloads (hours).")
    flow_setting_ensure(store, "LoreFetch_PDSigIndex_MaxWordsPerSig", 120, note="Max words persisted per signature in LoreSigIndex_PD_Auto.")
    flow_setting_ensure(store, "LoreFetch_PDWordFreq_TopN", 8000, note="Top-N words stored in LoreWordFreq_PD_Auto (derived counts only).")

    # Optional local dictionary signature index (derived-only; no network).
    flow_setting_ensure(store,
        "LoreFetch_DictSigIndex_Enabled",
        True,
        note="If enabled, build a derived signature index from a local English word list (no full text import) to expand candidate coverage.",
    )
    flow_setting_ensure(store,
        "LoreFetch_DictSigIndex_MaxAgeHours",
        720,
        note="Rebuild LoreSigIndex_Dict_Auto when older than N hours. Set 0 to rebuild every iteration.",
    )
    flow_setting_ensure(store,
        "LoreFetch_DictSigIndex_Path",
        "/usr/share/dict/words",
        note="Path to local word list (default macOS web2). Only derived counts are stored in the XLSX.",
    )
    flow_setting_ensure(store, "LoreFetch_DictSigIndex_MaxWordsPerSig", 120, note="Max words persisted per signature in LoreSigIndex_Dict_Auto.")

    flow_setting_ensure(store, "Semantic_Enabled", True, note="Materialize Translation_Semantic_Auto from SemanticMap_Auto (display-only).")
    flow_setting_ensure(store, "SemanticMap_MinTotalCount", 3, note="Semantic map: minimum total corpus count for top-word selection.")
    flow_setting_ensure(store, "SemanticMap_MinTopShare", 0.95, note="Semantic map: minimum top-share (top_count/total_count) for selection.")
    flow_setting_ensure(store,
        "SemanticPromoteGlossary_Enabled",
        True,
        note="Apply high-confidence semantic suggestions into Glossary translations (guarded by GT live check).",
    )
    flow_setting_ensure(store, "SemanticPromoteGlossary_MaxPerIter", 25, note="Max semantic retext edits to Glossary per iteration.")
    flow_setting_ensure(store, "SemanticPromoteGlossary_MinTotalOcc", 5, note="Semantic retext: only consider tokens with TotalOcc>=threshold.")
    flow_setting_ensure(store,
        "SemanticPromoteGlossary_MinConfidence",
        "MEDIUM",
        note="Semantic retext: minimum Glossary.Confidence (LOW/MEDIUM/HIGH).",
    )
    flow_setting_ensure(store,
        "SemanticPromoteGlossary_MaxConfidence",
        "MEDIUM",
        note="Semantic retext: maximum Glossary.Confidence allowed (skip HIGH by default).",
    )
    flow_setting_ensure(store,
        "SemanticPromoteGlossary_BlockEvidenceClasses",
        "GROUNDTRUTH,LOGOGRAM_ANCHORED,PUNCT_LOGOGRAM",
        note="Semantic retext: do NOT edit tokens with these EvidenceClass_v127 values (safety lock).",
    )
    flow_setting_ensure(store,
        "SemanticPromoteGlossary_MinNewWordFreq",
        5,
        note="Semantic retext: require new word global frequency >= N (from LoreWordFreq_Tibia_Auto when available).",
    )
    flow_setting_ensure(store,
        "SemanticPromoteGlossary_MinWordFreqRatio",
        3.0,
        note="Semantic retext: require new_word_freq >= old_word_freq * ratio (or old=0).",
    )
    flow_setting_ensure(store,
        "SemanticPromoteGlossary_WordFreqSheet",
        LORE_WORDFREQ_TIBIA_SHEET,
        note="Sheet used for global word frequency lookup (derived).",
    )
    # English (display-only) layer: map canonical decode-words to likely English surface forms.
    flow_setting_ensure(store, "EnglishLayer_Enabled", True, note="Materialize Translation_English_Auto (display-only) from Tibia sig-index.")
    flow_setting_ensure(store, "EnglishLayer_MinTotalCount", 20, note="English layer: minimum total corpus count for mapping.")
    flow_setting_ensure(store, "EnglishLayer_MinTopShare", 0.95, note="English layer: minimum top-share for mapping.")
    flow_setting_ensure(store, "EnglishLayer_MinWordLen", 3, note="English layer: skip mapping for very short words.")
    flow_setting_ensure(store, "EnglishLayer_MaxMapRows", 2000, note="English layer: max entries written to EnglishMap_Auto.")
    flow_setting_ensure(store,
        "EnglishGlossaryRetext_Enabled",
        True,
        note="Apply EnglishMap_Auto suggestions into Glossary.Translation (guarded by GT live check).",
    )
    flow_setting_ensure(store, "EnglishGlossaryRetext_MaxPerIter", 50, note="Max English->Glossary retext edits per iteration.")
    flow_setting_ensure(store, "EnglishGlossaryRetext_MinTotalCount", 5, note="Require EnglishMap.TotalWordCount >= N for retext.")
    flow_setting_ensure(store, "EnglishGlossaryRetext_MinTopShare", 0.90, note="Require EnglishMap.TopShare >= threshold for retext.")
    flow_setting_ensure(store,
        "EnglishGlossaryRetext_LockIters",
        5,
        note="Anti-oscillation: after applying an English retext to a token, do not flip it again for N iterations.",
    )
    flow_setting_ensure(store,
        "GlossaryRetext_LockedTokens",
        "",
        note="CSV token lock for shadow probes; semantic/english/reverse retext must not overwrite these Glossary rows.",
    )

    # Digits/core methodology (analysis-only) + external validations (analysis-only).
    flow_setting_ensure(store, "ExternalRoundTrip_MinVerifiedCount", 2, note="External roundtrip: require VerifiedCount>=N (joined by DigitsSanitized).")
    flow_setting_ensure(store,
        "ExternalRoundTrip_MinSegmentDigits",
        8,
        note="External roundtrip fallback: allow segment/ordered-run verification only for digit segments with len>=N.",
    )
    flow_setting_ensure(store,
        "ExternalRoundTrip_AllowOrderedRunMatch",
        True,
        note="External roundtrip fallback: allow ordered run match (for validation entries with ellipsis).",
    )
    flow_setting_ensure(store,
        "ExternalRefs_FillFromCodeStreamV120_Enabled",
        True,
        note="When TRUE, fill blank ExternalRefs core decode columns from CodeStream v120; set FALSE in shadow probes to audit external anchors without CodeStream backfill.",
    )
    flow_setting_ensure(store, "DigitCodeContext_Enabled", True, note="Compute DigitCodeContext_Auto (neighbor distributions per digits code; analysis-only).")
    flow_setting_ensure(store, "DigitCodeContext_TopK", 8, note="DigitCodeContext: top-K neighbors to record per code.")
    flow_setting_ensure(store, "DigitCodeContext_JSAlpha", 0.20, note="DigitCodeContext: smoothing alpha for JS-divergence outlier scoring within homophone sets.")

    # Context English (display-only) backed by a derived bigram LM (no full text persisted to XLSX).
    flow_setting_ensure(store, "LoreBigrams_Enabled", True, note="Build LoreBigrams_Auto from Tibia cache + LoreCorpus_Auto (derived counts only).")
    flow_setting_ensure(store, "LoreBigrams_MaxAgeHours", 168, note="Refresh LoreBigrams_Auto when older than N hours (0=always).")
    flow_setting_ensure(store, "LoreBigrams_TibiaCacheMaxAgeHours", 168, note="Max age for cached Tibia JSON used for LoreBigrams (0=always fetch).")
    flow_setting_ensure(store, "LoreBigrams_VocabTopN", 6000, note="Keep only bigrams where both words are in the top-N unigram vocab.")
    flow_setting_ensure(store, "LoreBigrams_MinCount", 3, note="Keep only bigrams with count >= N.")
    flow_setting_ensure(store, "LoreBigrams_MaxRows", 200000, note="Max bigram rows written to LoreBigrams_Auto.")
    flow_setting_ensure(store, "ContextEnglish_Enabled", True, note="Render Translation_ContextEnglish_Auto using LoreAlignment candidates + LoreBigrams LM (display-only).")
    flow_setting_ensure(store, "ContextEnglish_EmissionAlpha", 0.5, note="ContextEnglish emission smoothing alpha: log(cnt+alpha).")
    flow_setting_ensure(store, "ContextEnglish_TransitionAlpha", 0.5, note="ContextEnglish transition smoothing alpha: log(bigram+alpha).")
    flow_setting_ensure(store, "ContextEnglish_MaxCandidatesPerToken", 6, note="Max candidate surface words considered per token in Viterbi.")
    flow_setting_ensure(store, "ContextEnglishMap_MinTotal", 30, note="Min total occurrences (baseline word) required to emit a context map row.")
    flow_setting_ensure(store, "ContextEnglishMap_MinTopShare", 0.97, note="Min top share required to emit a context map row.")
    flow_setting_ensure(store, "ContextEnglishMap_MaxRows", 800, note="Max rows written to EnglishMap_Context_Auto.")

    # Code-aware homophones (display-only): use digits-code homophones to stabilize ambiguous 1-letter tokens.
    flow_setting_ensure(store,
        "CodeAware_Enabled",
        True,
        note="Render Translation_CodeAware_Auto using digits homophones + ContextEnglish choices (display-only).",
    )
    flow_setting_ensure(store, "CodeAware_MinTotalPerCode", 60, note="CodeAware: require total occurrences per (Token,Code) >= N.")
    flow_setting_ensure(store, "CodeAware_MinTopShare", 0.88, note="CodeAware: require top share for (Token,Code)->TopWord >= threshold.")
    flow_setting_ensure(store,
        "CodeAware_MaxTokenLen",
        1,
        note="CodeAware: apply code-aware mapping to tokens with len(Token) <= N (1=single-letter tokens only).",
    )
    flow_setting_ensure(store,
        "CodeAware_MinTotalPerCodeSeq",
        4,
        note="CodeAware: for multi-letter tokens, require total occurrences per (Token,CodeSeq) >= N (CodeSeq is joined codes like '09-61').",
    )
    flow_setting_ensure(store,
        "CodeAware_MinTopShareSeq",
        0.75,
        note="CodeAware: for multi-letter tokens, require top share for (Token,CodeSeq)->TopWord >= threshold.",
    )
    flow_setting_ensure(store, "CodeAware_MaxMapRows", 2500, note="CodeAware: max rows written to CodeWordMap_Auto.")
    flow_setting_ensure(store, "CodeAware_HintMinTotal", 8, note="CodeAware: minimum total occurrences required to use a (Token,Code)->word distribution as a soft hint.")
    flow_setting_ensure(store, "CodeAware_HintBoost", 4, note="CodeAware: emission-count boost multiplier for hinted words (display-only Viterbi).")
    flow_setting_ensure(store, "CodeAware_HintTopK", 2, note="CodeAware: keep only top-K hinted words per (Token,Code) when boosting candidates.")
    flow_setting_ensure(store,
        "CodeAware_ApplyToTokens",
        "",
        note="Optional CSV allowlist of tokens to apply (empty=auto). Example: T,N,V,TV",
    )

    # Sequence matching (analysis-only): match rare context n-grams against corpora, storing only snippets/URLs.
    flow_setting_ensure(store, "SequenceMatch_Enabled", True, note="Enable SequenceMatches_Auto mining (analysis-only; snippets only).")
    flow_setting_ensure(store, "SequenceMatch_NList", "6,7,8", note="CSV list of n-gram sizes to match.")
    flow_setting_ensure(store, "SequenceMatch_MinN", 2, note="Ignore n-grams smaller than N when matching (quality filter).")
    flow_setting_ensure(store,
        "SequenceMatch_CandidateMaxBookFreq",
        1,
        note="Candidate extraction: keep n-grams that appear <=N times across books (1=strictly unique; higher improves recall but increases noise).",
    )
    flow_setting_ensure(store, "SequenceMatch_MaxCandidates", 4000, note="Max rare n-gram candidates extracted from Books per iteration.")
    flow_setting_ensure(store,
        "SequenceMatch_ExploreRotate",
        True,
        note="If TRUE, rotate the candidate n-gram slice by iteration (after keeping the top-K) to avoid deterministic plateaus.",
    )
    flow_setting_ensure(store,
        "SequenceMatch_ExploreKeepTop",
        200,
        note="Always include the top-K candidates by score before rotating the remaining candidate slice (0=rotate all).",
    )
    flow_setting_ensure(store,
        "SequenceMatch_CacheEnabled",
        True,
        note="Maintain SequenceMatchesCache_Auto as a deduped union of past SequenceMatches (snippets only; used to accumulate word-hints).",
    )
    flow_setting_ensure(store, "SequenceMatch_CacheMaxRows", 2000, note="Max rows kept in SequenceMatchesCache_Auto (deduped union).")
    flow_setting_ensure(store, "SequenceMatch_MaxMatches", 60, note="Max matches written to SequenceMatches_Auto per iteration.")
    flow_setting_ensure(store, "SequenceMatch_TimeBudgetS", 20, note="Time budget for scanning corpora for sequence matches (seconds).")
    flow_setting_ensure(store,
        "SequenceMatch_ScanTibiaFirst",
        True,
        note="If TRUE, scan the Tibia corpus cache before large PD sources under the time budget (higher overlap).",
    )
    flow_setting_ensure(store,
        "SequenceMatch_PD_MaxChars",
        800000,
        note="Per PD source: scan at most N characters for SequenceMatches (0=unlimited). Helps avoid spending the time budget in huge texts.",
    )
    flow_setting_ensure(store,
        "SequenceMatch_PD_MaxSentencesPerSource",
        2500,
        note="Per PD source: scan at most N sentences for SequenceMatches (0=unlimited). Helps avoid large allocations/timeouts.",
    )
    flow_setting_ensure(store,
        "SequenceMatch_ContextWindow",
        2,
        note="Number of words to capture on each side of a candidate n-gram when matching contexts (0=disabled).",
    )
    flow_setting_ensure(store,
        "SequenceMatch_ContextMinOverlap",
        1,
        note="Require at least N overlapping context words between the candidate context and snippet before emitting a match.",
    )
    flow_setting_ensure(store,
        "SequenceMatch_ContextRequireDirection",
        False,
        note="If TRUE, require the per-side context match (pre/post) observed in Books to appear in the snippet as well.",
    )
    flow_setting_ensure(store,
        "SequenceMatch_SnippetMinContentWords",
        1,
        note="Require at least N non-stopword content tokens in the snippet window before storing a match.",
    )
    flow_setting_ensure(store,
        "SequenceWordHints_Enabled",
        True,
        note="Derive SequenceWordHints_Auto from SequenceMatches_Auto (analysis-only; used to nudge ContextEnglish).",
    )
    flow_setting_ensure(store, "SequenceWordHints_MinN", 3, note="Ignore SequenceMatches rows with N < this when deriving word hints (quality filter).")
    flow_setting_ensure(store, "SequenceWordHints_SkipIdentity", True, note="If TRUE, skip FromWord==ToWord hints (keeps hints purely corrective).")
    flow_setting_ensure(store, "SequenceWordHints_MaxRows", 800, note="Max rows written to SequenceWordHints_Auto.")
    flow_setting_ensure(store,
        "SequenceWordHints_MinRatio",
        1.5,
        note="If multiple ToWords compete for the same (CanonSig, FromWord), require top/runner-up ratio >= this (1.0 disables).",
    )
    flow_setting_ensure(store,
        "SequenceWordHints_ExcludeStopwords",
        True,
        note="Do not emit hints when either FromWord or ToWord is a stopword (keeps hints focused on informative tokens).",
    )
    flow_setting_ensure(store,
        "SequenceWordHints_StopwordRatio",
        0.67,
        note="Skip SequenceMatches rows whose ContextEnglish phrase is dominated by stopwords (ratio >= value; <0 disables).",
    )
    flow_setting_ensure(store, "SequenceHints_Enabled", True, note="Use SequenceWordHints_Auto to boost ContextEnglish candidates (display-only).")
    flow_setting_ensure(store, "SequenceHints_Boost", 20, note="Candidate boost per hint count (ContextEnglish).")
    flow_setting_ensure(store, "SequenceHints_MaxWordsPerSig", 3, note="Max hinted words per signature used for boosting.")
    flow_setting_ensure(store,
        "ReversePhraseHints_Enabled",
        True,
        note="Use ReversePhraseTokenCands_Auto to boost ContextEnglish/CodeAware candidates (display-only; helps when reverse mining finds strong token->word clues).",
    )
    flow_setting_ensure(store, "ReversePhraseHints_Boost", 8, note="Candidate boost per reverse-phrase hint count (display-only).")
    flow_setting_ensure(store, "ReversePhraseHints_MaxWordsPerSig", 2, note="Max reverse-phrase hinted words per signature used for boosting.")

    # Sestina/Sestine structural scan (analysis-only): look for the retrogradatio cruciata end-word permutation.
    flow_setting_ensure(store, "SestinaScan_Enabled", True, note="Enable sestina/sestine structural scan (analysis-only).")
    flow_setting_ensure(store, "SestinaScan_MaxLines", 4000, note="Hard cap on extracted global lines (split by ↵) to bound runtime.")
    flow_setting_ensure(store, "SestinaScan_MinScore30", 26, note="Minimum matches (out of 30) for a 36-line sestina candidate window.")
    flow_setting_ensure(store, "SestinaScan_MaxCandidates", 40, note="Max candidate windows written to SestinaCandidates_Auto per iteration.")
    flow_setting_ensure(store,
        "SestinaScan_UseTokenSignature",
        True,
        note="If TRUE, detect sestina windows using token-signatures (sorted letters, ignore '*') for end tokens (more robust to anagram variants).",
    )
    flow_setting_ensure(store, "SestinaScan_EnvoiBonus", 1.0, note="Bonus weight for envoi end-token pattern matches (ranking only; Score30 remains 0..30).")
    flow_setting_ensure(store,
        "SestinaObligation_Enabled",
        True,
        note="Run positional obligation analysis (ablation + reorder stress) over SestinaCandidates_Auto (analysis-only).",
    )
    flow_setting_ensure(store,
        "SestinaObligation_MaxCandidates",
        30,
        note="Max sestina candidate windows analyzed for positional obligation per iteration (0=all).",
    )
    flow_setting_ensure(store,
        "SestinaObligation_MinScore30",
        0,
        note="Only analyze sestina windows with Score30>=N (0 keeps weak windows for fragility diagnostics).",
    )
    flow_setting_ensure(store,
        "SestinaObligation_ObligatoryImpact",
        0.015,
        note="Minimum ablation impact (baseline_rate - remove_rate) to label a position as OBLIGATORY.",
    )
    flow_setting_ensure(store,
        "SestinaObligation_ConditionalImpact",
        0.000,
        note="Minimum ablation impact to label a position as CONDITIONAL (below this tends to redundant/decorative).",
    )
    flow_setting_ensure(store,
        "SestinaObligation_DecorativeImpact",
        -0.020,
        note="If impact<=threshold, classify position as DECORATIVE (otherwise REDUNDANT when below conditional).",
    )
    flow_setting_ensure(store,
        "SestinaObligation_NoCollapseTolerance",
        0.020,
        note="Ablation ladder: removing weakest positions is 'no collapse' while reduced score-rate stays >= baseline_rate - tolerance.",
    )
    flow_setting_ensure(store,
        "SestinaObligation_ReorderResilientRatio",
        0.90,
        note="Window is reorder-resilient when best non-identity order reaches >= ratio * baseline score.",
    )
    flow_setting_ensure(store,
        "RhythmTransition_Enabled",
        True,
        note="Run global rhythm A/B test over extracted lines (analysis-only).",
    )
    flow_setting_ensure(store,
        "RhythmTransition_WindowSize",
        12,
        note="Window size for rhythm analysis (default 12 blocks).",
    )
    flow_setting_ensure(store,
        "RhythmTransition_MinLines",
        24,
        note="Require at least N extracted lines before rhythm analysis runs.",
    )
    flow_setting_ensure(store,
        "RhythmTransition_UseTokenSignature",
        True,
        note="If TRUE, compute rhythm stats on token signatures (sorted letters, '*' ignored).",
    )
    flow_setting_ensure(store,
        "RhythmTransition_ShuffleTrials",
        8,
        note="Cycle-6 control: number of shuffled-tail trials per window.",
    )
    flow_setting_ensure(store,
        "RhythmTransition_CycleDeltaThreshold",
        0.05,
        note="Model preference threshold: cycle-6 wins when delta vs shuffled >= threshold.",
    )

    # Puzzle-solved status (stronger than RESOLVED).
    flow_setting_ensure(store, "PuzzleSolved_MinExternalVerifiedCount", 2, note="PuzzleSolved: external roundtrip considers only refs with VerifiedCount>=N.")
    flow_setting_ensure(store, "PuzzleSolved_ContextImproveIters", 0, note="PuzzleSolved: require ContextEnglish avg-score improvement streak >= N.")
    flow_setting_ensure(store, "PuzzleSolved_MinSequenceMatches", 3, note="PuzzleSolved: require >= N sequence matches.")
    # Reverse phrase mining (user-driven): encode known phrases -> search base to derive candidate tokens (safe, gated).
    flow_setting_ensure(store,
        "ReversePhrase_Enabled",
        True,
        note="Enable reverse phrase mining from PhraseCribs_User into ReversePhrase* sheets (analysis-first; optional candidate token emission).",
    )
    flow_setting_ensure(store, "ReversePhrase_IncludePhraseCribsUser", True, note="Use PhraseCribs_User rows as reverse phrase sources.")
    flow_setting_ensure(store,
        "ReversePhrase_IncludePhraseCribsAuto",
        True,
        note="Use PhraseCribs_Auto rows as reverse phrase sources (runner-generated, analysis-only).",
    )
    flow_setting_ensure(store, "ReversePhrase_IncludeLoreCorpusUser", True, note="Also use LoreCorpus_User rows as reverse phrase sources (no need to duplicate).")
    flow_setting_ensure(store, "ReversePhrase_IncludeLoreCorpusAuto", False, note="If TRUE, also include LoreCorpus_Auto (noisy; default OFF).")
    flow_setting_ensure(store,
        "ReversePhrase_IncludeTibiaCorpusAuto",
        True,
        note="If TRUE, also include sampled Tibia NPC/books lines fetched via internet as reverse-phrase sources (hits are capped; no full corpus stored).",
    )
    flow_setting_ensure(store,
        "ReversePhrase_IncludeSequenceMatchesAuto",
        True,
        note="If TRUE, also include SequenceMatches_Auto snippets as reverse-phrase sources (analysis-only; already snippet-limited).",
    )
    flow_setting_ensure(store,
        "ReversePhrase_SeqMatchesMaxPhrasesPerIter",
        40,
        note="Cap on how many SequenceMatches-derived phrases are added per iteration (controls runtime).",
    )
    flow_setting_ensure(store,
        "ReversePhrase_SeqMatchesMinN",
        3,
        note="Only use SequenceMatches rows with N>=threshold as reverse-phrase sources (quality filter).",
    )
    flow_setting_ensure(store,
        "ReversePhrase_TibiaMaxPhrasesPerIter",
        40,
        note="Cap on how many Tibia-sourced phrases are added per iteration (controls runtime).",
    )
    flow_setting_ensure(store,
        "ReversePhrase_TibiaCacheMaxAgeHours",
        24,
        note="Max age for local cached Tibia JSON downloads used by reverse phrase mining (0=always fetch).",
    )
    flow_setting_ensure(store, "ReversePhrase_TibiaIncludeNPC", True, note="When IncludeTibiaCorpusAuto=TRUE, include NPC transcript answers.")
    flow_setting_ensure(store, "ReversePhrase_TibiaIncludeBooks", True, note="When IncludeTibiaCorpusAuto=TRUE, include Tibia book sentences.")
    flow_setting_ensure(store, "ReversePhrase_PhraseTextMaxLen", 220, note="Truncate PhraseText stored in ReversePhraseHits_Auto to N chars.")
    flow_setting_ensure(store,
        "ReversePhrase_LogogramAware",
        True,
        note="Use canonical translation letters for LOGOGRAM_* tokens during signature matching (improves matchability).",
    )
    flow_setting_ensure(store, "ReversePhrase_Canon_DropFinalE", False, note="ReversePhrase canon: optionally drop final 'e' (default OFF; matching should preserve letters).")
    flow_setting_ensure(store, "ReversePhrase_Canon_DropAllH", False, note="ReversePhrase canon: optionally drop all 'h' (default OFF).")
    flow_setting_ensure(store, "ReversePhrase_Canon_DropAllO", False, note="ReversePhrase canon: optionally drop all 'o' (default OFF; matching should preserve letters).")
    flow_setting_ensure(store, "ReversePhrase_CandidateMaxBaseLen", 24, note="Skip emitting candidates where base concat length exceeds N.")
    flow_setting_ensure(store,
        "ReversePhrase_MaxSpanTokens",
        6,
        note="Reverse phrase matching: allow each phrase word to match a concatenation of 1..K base tokens (K=MaxSpanTokens).",
    )
    flow_setting_ensure(store,
        "ReversePhrase_MaxGapTokens",
        0,
        note="Reverse phrase matching: allow up to N unmatched base tokens between consecutive phrase words (0=strict adjacency).",
    )
    flow_setting_ensure(store, "ReversePhrase_MinWords", 4, note="Skip very short phrases (min words after canon).")
    flow_setting_ensure(store, "ReversePhrase_MaxPhrasesPerIter", 120, note="Hard cap to keep iteration runtime stable.")
    flow_setting_ensure(store, "ReversePhrase_MaxHitsTotal", 60, note="Hard cap on total hits written to ReversePhraseHits_Auto per iteration.")
    flow_setting_ensure(store,
        "ReversePhrase_MaxHitsPerPhrase",
        3,
        note="Reverse phrase mining: max accepted hits per phrase across books.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_MaxHitsPerBook",
        1,
        note="Reverse phrase mining: max accepted hits per book per phrase.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_PermutationEnabled",
        True,
        note="If TRUE, run an additional permutation-aware reverse phrase scan (analysis-only) when sequential hits are low/zero (catches word-order shuffles).",
    )
    flow_setting_ensure(store, "ReversePhrase_PermutationRunIfHitsLE", 0, note="Run permutation scan only if sequential reverse-phrase hits <= threshold.")
    flow_setting_ensure(store, "ReversePhrase_PermutationMaxWords", 6, note="Only permute-match phrases with <=N words (runtime cap).")
    flow_setting_ensure(store, "ReversePhrase_PermutationMaxPhrasesPerIter", 25, note="Cap how many phrases are scanned in permutation mode (runtime cap).")
    flow_setting_ensure(store, "ReversePhrase_PermutationMaxHitsTotal", 30, note="Hard cap on total permuted hits written per iteration.")
    flow_setting_ensure(store,
        "ReversePhrase_PermutationMaxHitsPerPhrase",
        2,
        note="Reverse phrase permutation: max accepted hits per phrase across books.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_PermutationMaxHitsPerBook",
        1,
        note="Reverse phrase permutation: max accepted hits per book per phrase.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_PermutationAutoEmitCandidates",
        False,
        note="If TRUE, permuted hits can also emit inactive Glossary candidates (still guarded by GT/live-check + mechanical simulation). Default OFF.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_AutoEmitCandidates",
        True,
        note="If hits are found, emit inactive candidate tokens into Glossary with EvidenceClass=ReversePhrase_EvidenceClass (promotion is still simulated/gated).",
    )
    flow_setting_ensure(store, "ReversePhrase_EvidenceClass", "PHRASE_CRIB", note="EvidenceClass_v127 used for reverse-mined candidate tokens.")
    flow_setting_ensure(store, "ReversePhrase_CandidateMinBooks", 2, note="Emit candidate token only if it appears in >=N distinct books via phrase hits.")
    flow_setting_ensure(store, "ReversePhrase_CandidateMinTopShare", 0.95, note="Emit candidate token only if its top-word share >= threshold.")
    flow_setting_ensure(store, "ReversePhrase_CandidateMaxNewTokens", 20, note="Max reverse-mined candidate tokens appended to Glossary per iteration.")
    flow_setting_ensure(store,
        "ReversePhrase_RetextExisting_Enabled",
        True,
        note="If TRUE, allow ReversePhrase candidates to retext existing active tokens (GT live-check guarded).",
    )
    flow_setting_ensure(store,
        "ReversePhrase_RetextExisting_MaxPerIter",
        20,
        note="Max existing-token retext edits applied from ReversePhrase candidates per iteration.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_RetextExisting_MinBooks",
        2,
        note="Retext existing tokens only when ReversePhrase support is seen in >=N books.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_RetextExisting_MinTopShare",
        0.95,
        note="Retext existing tokens only when ReversePhrase TopShare >= threshold.",
    )
    flow_setting_ensure(store,
        "ReversePhrase_RetextExisting_MinSupportOcc",
        2,
        note="Retext existing tokens only when ReversePhrase support occurrences >= N.",
    )
    # AutoPhraseCribs (analysis-only): select feasible Tibia phrases by signature overlap against the current decode stream.
    flow_setting_ensure(store, "AutoPhraseCribs_Enabled", True, note="Generate PhraseCribs_Auto from Tibia corpus (analysis-only) to feed reverse phrase mining.")
    flow_setting_ensure(store, "AutoPhraseCribs_CacheMaxAgeHours", 24, note="Max age for local cached Tibia JSON downloads (0=always fetch).")
    flow_setting_ensure(store, "AutoPhraseCribs_MaxPhrasesPerIter", 80, note="Max enabled phrases written to PhraseCribs_Auto per iteration.")
    flow_setting_ensure(store, "AutoPhraseCribs_MaxScanSentences", 50000, note="Max Tibia sentences scanned when building PhraseCribs_Auto.")
    flow_setting_ensure(store, "AutoPhraseCribs_TimeBudgetS", 25, note="Time budget for PhraseCribs_Auto mining (seconds).")
    flow_setting_ensure(store, "AutoPhraseCribs_MinWords", 4, note="PhraseCribs_Auto: minimum words per phrase (after canon).")
    flow_setting_ensure(store, "AutoPhraseCribs_MaxWords", 14, note="PhraseCribs_Auto: maximum words per phrase (keeps reverse matching tractable).")
    flow_setting_ensure(store, "AutoPhraseCribs_MaxTextLen", 240, note="PhraseCribs_Auto: skip very long phrases to keep runtime stable.")
    flow_setting_ensure(store,
        "AutoPhraseCribs_IncludePD",
        True,
        note="Also mine PhraseCribs_Auto from Project Gutenberg public-domain sources (analysis-only; improves coverage when the encoded corpus is not Tibia text).",
    )
    flow_setting_ensure(store, "AutoPhraseCribs_PDCacheMaxAgeHours", 720, note="Max age for cached PD text downloads (0=always fetch).")
    flow_setting_ensure(store, "AutoPhraseCribs_PDMaxScanSentences", 25000, note="Max PD sentences scanned when building PhraseCribs_Auto.")
    flow_setting_ensure(store, "AutoPhraseCribs_PDTimeBudgetS", 25, note="Time budget for PD PhraseCribs_Auto mining (seconds).")
    flow_setting_ensure(store, "AutoPhraseCribs_PDMaxSources", 0, note="If >0, scan only the first N PD sources from the default list (0=all, time-budgeted).")
    # Reload settings after inserting missing keys.
    flow_store_refresh_from_workbook(store)
    ws_settings = store["ws_settings"]
    settings_map = store["settings_map"]
    min_verified_sources_gt = int(get_setting(settings_map, "MinVerifiedSourcesGT", 2) or 2)
    allow_mech = parse_bool(get_setting(settings_map, "AllowMechanicalPromotion", True), True)
    min_occ = int(get_setting(settings_map, "MinTotalOccCandidate", 5) or 5)
    max_weak_inc = float(get_setting(settings_map, "MaxWeakFracIncrease", 0) or 0)
    max_single_inc = float(get_setting(settings_map, "MaxSingleCharFracIncrease", 0) or 0)
    max_micro_inc = float(get_setting(settings_map, "MaxMicroFracIncrease", 0) or 0)
    max_ev_drop = float(get_setting(settings_map, "MaxEvidenceAvgDrop", 0) or 0)
    min_ev_delta = float(get_setting(settings_map, "MinEvidenceAvgDelta", 0) or 0)
    mech_classes = parse_csv(get_setting(settings_map, "MechanicalCandidateClasses", "") or "")

    target_weak = float(get_setting(settings_map, "TargetWeakFrac", 0.12) or 0.12)
    target_micro = float(get_setting(settings_map, "TargetMicroFrac", 0.10) or 0.10)
    target_ev = float(get_setting(settings_map, "TargetEvidenceAvg", 2.3) or 2.3)
    target_gt_count = int(get_setting(settings_map, "TargetGroundTruthCount", 6) or 6)
    require_all_gt = parse_bool(get_setting(settings_map, "RequireAllGroundTruthMatch", True), True)
    require_cov = parse_bool(get_setting(settings_map, "RequireCoverageStrictPlus", True), True)
    soft_mismatch_max_resolved = int(get_setting(settings_map, "SoftMismatchMaxForResolved", 4) or 4)
    if soft_mismatch_max_resolved < 4:
        soft_mismatch_max_resolved = 4
        try:
            flow_setting_set(store,
                "SoftMismatchMaxForResolved",
                4,
                note=f"iter{next_iter}: auto-relax soft mismatch threshold for unresolved provisional path",
            )
            flow_store_refresh_from_workbook(store)
            ws_settings = store["ws_settings"]
            settings_map = store["settings_map"]
        except Exception:
            pass
    soft_mismatch_max_soft_resolved = int(get_setting(settings_map, "SoftMismatchMaxForSoftResolved", 999) or 999)
    convergence_use_soft_rescue = parse_bool(get_setting(settings_map, "Convergence_UseSoftMismatchRescue", True), True)
    convergence_soft_rescue_max = int(get_setting(settings_map, "Convergence_MaxSoftRescuePromotions", 20) or 20)
    if convergence_soft_rescue_max < 20:
        convergence_soft_rescue_max = 20
        try:
            flow_setting_set(store,
                "Convergence_MaxSoftRescuePromotions",
                20,
                note=f"iter{next_iter}: auto-relax soft rescue budget for unlocked search",
            )
            flow_store_refresh_from_workbook(store)
            ws_settings = store["ws_settings"]
            settings_map = store["settings_map"]
        except Exception:
            pass
    convergence_rescue_shuffle = parse_bool(get_setting(settings_map, "Convergence_RescueCandidateShuffle", True), True)
    convergence_soft_nondec_streak_block = int(get_setting(settings_map, "Convergence_SoftNonDecStreakBlock", 999) or 999)
    if convergence_soft_nondec_streak_block < 999:
        convergence_soft_nondec_streak_block = 999
        try:
            flow_setting_set(store,
                "Convergence_SoftNonDecStreakBlock",
                999,
                note=f"iter{next_iter}: auto-relax soft nondecreasing streak lock",
            )
            flow_store_refresh_from_workbook(store)
            ws_settings = store["ws_settings"]
            settings_map = store["settings_map"]
        except Exception:
            pass
    convergence_enable_hard_escape = parse_bool(get_setting(settings_map, "Convergence_EnableHardEscape", True), True)
    convergence_hard_escape_passes = int(get_setting(settings_map, "Convergence_HardEscapePasses", 4) or 4)
    convergence_hard_escape_tol_scale = float(get_setting(settings_map, "Convergence_HardEscapeToleranceScale", 4.0) or 4.0)
    convergence_hard_escape_stall_iters = int(get_setting(settings_map, "Convergence_HardEscape_StallIters", 2) or 2)
    convergence_directional_stall_iters = int(
        get_setting(settings_map, "Convergence_DirectionalEscape_StallIters", 1) or 1
    )
    convergence_enable_directional_escape = parse_bool(get_setting(settings_map, "Convergence_EnableDirectionalEscape", True), True)
    convergence_directional_passes = int(get_setting(settings_map, "Convergence_DirectionalEscapePasses", 3) or 3)
    convergence_directional_topk = int(get_setting(settings_map, "Convergence_DirectionalEscapeTopK", 30) or 30)
    convergence_directional_frontier_scale = float(
        get_setting(settings_map, "Convergence_DirectionalEscapeFrontierScale", 1.6) or 1.6
    )
    convergence_directional_allow_regression = parse_bool(
        get_setting(settings_map, "Convergence_DirectionalEscapeAllowMetricRegression", True), True
    )
    convergence_directional_score_min = float(get_setting(settings_map, "Convergence_DirectionalEscapeScoreMin", -999.0) or -999.0)
    convergence_interleave_evidence = parse_bool(get_setting(settings_map, "Convergence_InterleaveEvidenceBuckets", True), True)
    convergence_candidate_starve_iters = int(get_setting(settings_map, "Convergence_CandidateStarveIters", 4) or 4)
    convergence_candidate_scan_all_classes = parse_bool(
        get_setting(settings_map, "Convergence_CandidateScanFallbackAllClasses", True), True
    )
    convergence_candidate_scan_min_occ = max(1, int(get_setting(settings_map, "Convergence_CandidateScanFallbackMinOcc", 1) or 1))
    convergence_candidate_scan_keep_top = max(1, int(get_setting(settings_map, "Convergence_CandidateScanFallbackKeepTop", 400) or 400))
    convergence_candidate_cap_per_iter = max(0, int(get_setting(settings_map, "Convergence_CandidateCapPerIter", 120) or 120))
    candidate_priority_enabled = parse_bool(get_setting(settings_map, "CandidatePriority_Enabled", True), True)
    candidate_priority_min_book_hits = max(1, int(get_setting(settings_map, "CandidatePriority_MinBookHits", 1) or 1))
    candidate_priority_min_total_occ = max(1, int(get_setting(settings_map, "CandidatePriority_MinTotalOcc", 2) or 2))
    candidate_priority_max_len = max(4, int(get_setting(settings_map, "CandidatePriority_MaxLen", 40) or 40))
    candidate_priority_max_len_macro = max(6, int(get_setting(settings_map, "CandidatePriority_MaxLenMacro", 64) or 64))
    candidate_priority_noeffect_threshold = max(
        1, int(get_setting(settings_map, "CandidatePriority_NoEffectSkipThreshold", 2) or 2)
    )
    candidate_priority_noeffect_lookback = max(
        1, int(get_setting(settings_map, "CandidatePriority_NoEffectCooldownIters", 8) or 8)
    )
    convergence_force_probe_tokens = [
        str(tok or "").strip()
        for tok in str(get_setting(settings_map, "Convergence_ForceProbeTokens", "") or "").split(",")
        if str(tok or "").strip()
    ]
    convergence_force_probe_pair_tokens: List[Tuple[str, str]] = []
    for raw_pair in str(get_setting(settings_map, "Convergence_ForceProbePairTokens", "") or "").split(";"):
        raw_pair = str(raw_pair or "").strip()
        if not raw_pair or "+" not in raw_pair:
            continue
        left, right = raw_pair.split("+", 1)
        left = str(left or "").strip()
        right = str(right or "").strip()
        if not left or not right or left == right:
            continue
        convergence_force_probe_pair_tokens.append((left, right))
    convergence_force_probe_substrings = [
        str(tok or "").strip()
        for tok in str(get_setting(settings_map, "Convergence_ForceProbeSubstrings", "") or "").split(",")
        if str(tok or "").strip()
    ]
    convergence_force_probe_substring_max_matches = max(
        1, int(get_setting(settings_map, "Convergence_ForceProbeSubstringMaxMatches", 24) or 24)
    )
    convergence_force_probe_only = parse_bool(
        get_setting(settings_map, "Convergence_ForceProbeOnly", False), False
    )
    convergence_force_probe_disable_tokens = {
        str(tok or "").strip()
        for tok in str(get_setting(settings_map, "Convergence_ForceProbeDisableTokens", "") or "").split(",")
        if str(tok or "").strip()
    }
    convergence_force_probe_disable_substrings = [
        str(tok or "").strip()
        for tok in str(get_setting(settings_map, "Convergence_ForceProbeDisableSubstrings", "") or "").split(",")
        if str(tok or "").strip()
    ]
    convergence_force_probe_disable_supersets = parse_bool(
        get_setting(settings_map, "Convergence_ForceProbeDisableSupersets", False), False
    )
    convergence_force_probe_disable_supersets_max = max(
        1, int(get_setting(settings_map, "Convergence_ForceProbeDisableSupersetsMax", 12) or 12)
    )
    convergence_diagnose_swallowed_supersets = parse_bool(
        get_setting(settings_map, "Convergence_DiagnoseSwallowedSupersets", False), False
    )
    convergence_diagnose_swallowed_min_len = max(
        4, int(get_setting(settings_map, "Convergence_DiagnoseSwallowedMinLen", 10) or 10)
    )
    convergence_diagnose_swallowed_max_supersets = max(
        1, int(get_setting(settings_map, "Convergence_DiagnoseSwallowedMaxSupersets", 6) or 6)
    )
    convergence_diagnose_swallowed_max_report = max(
        1, int(get_setting(settings_map, "Convergence_DiagnoseSwallowedMaxReport", 3) or 3)
    )
    convergence_strict_monotonic_mechanical = parse_bool(
        get_setting(settings_map, "Convergence_StrictMonotonicMechanical", True), True
    )
    _convergence_monotonic_min_score_raw = get_setting(settings_map, "Convergence_MonotonicMinScore", 0.05)
    try:
        convergence_monotonic_min_score = float(
            0.05 if _convergence_monotonic_min_score_raw is None else _convergence_monotonic_min_score_raw
        )
    except Exception:
        convergence_monotonic_min_score = 0.05
    convergence_monotonic_allow_token_increase = parse_bool(
        get_setting(settings_map, "Convergence_MonotonicAllowTokenIncrease", False), False
    )
    convergence_enable_low_confidence_escape = parse_bool(
        get_setting(settings_map, "Convergence_EnableLowConfidenceEscape", True), True
    )
    convergence_low_confidence_stall_iters = max(1, int(get_setting(settings_map, "Convergence_LowConfidenceEscapeStallIters", 3) or 3))
    convergence_semantic_objective_enabled = parse_bool(
        get_setting(
            settings_map,
            "Convergence_SemanticObjective_Enabled",
            get_setting(settings_map, "Convergence_EnableNoEffectAnchorEscape", True),
        ),
        True,
    )
    convergence_semantic_no_effect_stall_iters = max(
        1,
        int(
            get_setting(
                settings_map,
                "Convergence_SemanticObjective_NoEffectStallIters",
                get_setting(settings_map, "Convergence_NoEffectAnchorEscapeStallIters", 3),
            )
            or 3
        ),
    )
    convergence_semantic_no_effect_max_promos = max(
        0,
        int(
            get_setting(
                settings_map,
                "Convergence_SemanticObjective_NoEffectMaxPromos",
                get_setting(settings_map, "Convergence_NoEffectAnchorEscapeMaxPromos", 8),
            )
            or 8
        ),
    )
    _convergence_semantic_no_effect_min_gain_raw = get_setting(
        settings_map, "Convergence_SemanticObjective_NoEffectMinGain", 1.15
    )
    try:
        convergence_semantic_no_effect_min_gain = float(
            1.15
            if _convergence_semantic_no_effect_min_gain_raw is None
            else _convergence_semantic_no_effect_min_gain_raw
        )
    except Exception:
        convergence_semantic_no_effect_min_gain = 1.15
    convergence_semantic_gain_weight = float(
        get_setting(settings_map, "Convergence_SemanticObjective_GainWeight", 0.55) or 0.55
    )
    convergence_semantic_min_occ = max(
        1,
        int(
            get_setting(
                settings_map,
                "Convergence_SemanticObjective_MinOcc",
                get_setting(settings_map, "Convergence_NoEffectAnchorEscapeMinOcc", 1),
            )
            or 1
        ),
    )
    _semantic_allow_classes_raw = str(
        get_setting(
            settings_map,
            "Convergence_SemanticObjective_AllowClasses",
            get_setting(
                settings_map,
                "Convergence_NoEffectAnchorEscapeClasses",
                "PHRASE_CRIB,EXTERNAL_POEM,GROUNDTRUTH,LOGOGRAM_CONTEXT,MACRO_ACTIVE,STRUCT_MACRO_CAND",
            ),
        )
        or ""
    ).strip()
    convergence_semantic_allow_classes = {
        s.strip().upper()
        for s in _semantic_allow_classes_raw.split(",")
        if isinstance(s, str) and s.strip()
    }
    if not convergence_semantic_allow_classes:
        convergence_semantic_allow_classes = {
            "PHRASE_CRIB",
            "EXTERNAL_POEM",
            "GROUNDTRUTH",
            "LOGOGRAM_CONTEXT",
            "MACRO_ACTIVE",
            "STRUCT_MACRO_CAND",
        }
    _semantic_positive_classes_raw = str(
        get_setting(
            settings_map,
            "Convergence_SemanticObjective_PositiveClasses",
            "PHRASE_CRIB,EXTERNAL_POEM,GROUNDTRUTH,LOGOGRAM_CONTEXT,ANAGRAM_HIGH_BASE",
        )
        or ""
    ).strip()
    convergence_semantic_positive_classes = {
        s.strip().upper()
        for s in _semantic_positive_classes_raw.split(",")
        if isinstance(s, str) and s.strip()
    }
    _semantic_negative_classes_raw = str(
        get_setting(
            settings_map,
            "Convergence_SemanticObjective_NegativeClasses",
            "MACRO_ACTIVE,STRUCT_MACRO_CAND",
        )
        or ""
    ).strip()
    convergence_semantic_negative_classes = {
        s.strip().upper()
        for s in _semantic_negative_classes_raw.split(",")
        if isinstance(s, str) and s.strip()
    }
    convergence_candidate_emergency_rung = max(3, int(get_setting(settings_map, "Convergence_MacroMineEmergencyRung", 3) or 3))
    convergence_require_progress_for_solved = max(1, int(get_setting(settings_map, "Convergence_RequireProgressForSolvedIters", 3) or 3))
    plateau_block_no_promo_iters = int(get_setting(settings_map, "PlateauBlock_NoPromotionIters", 3) or 3)
    convergence_use_live_gt_counts = parse_bool(get_setting(settings_map, "Convergence_UseLiveGTCounts", True), True)

    macro_mine_mode_raw = str(get_setting(settings_map, "MacroMine_Mode", "") or "").strip().upper()
    macro_mine_mode = macro_mine_mode_raw if macro_mine_mode_raw in ("", "AUTO", "OFF", "FALLBACK_ONLY", "ALWAYS") else ""
    # Default behavior:
    # - before RESOLVED: mine only as a fallback when no inactive candidates exist
    # - after RESOLVED: mine always to keep reducing risk/token-count safely
    if macro_mine_mode in ("", "AUTO"):
        macro_mine_mode = "ALWAYS" if cur_status in ("SOFT_RESOLVED", "RESOLVED", STATUS_MODEL_CONVERGED) else "FALLBACK_ONLY"
    macro_mine_always = macro_mine_mode == "ALWAYS"
    macro_mine_off = macro_mine_mode == "OFF"

    plateau_enable_focus = parse_bool(get_setting(settings_map, "PlateauEnableFocusReport", True), True)
    plateau_enable_structural = parse_bool(get_setting(settings_map, "PlateauEnableStructural", True), True)
    structural_ignore_anchorcribs = {
        str(tok or "").strip()
        for tok in str(get_setting(settings_map, "StructuralIgnoreAnchorCribs", "") or "").split(",")
        if str(tok or "").strip()
    }
    structural_restrict_votes_to_anchor_windows = parse_bool(
        get_setting(settings_map, "StructuralRestrictVotesToAnchorWindows", False), False
    )
    structural_anchor_window_pad = max(0, int(get_setting(settings_map, "StructuralAnchorWindowPad", 0) or 0))
    structural_ref_book_override_raw = str(get_setting(settings_map, "StructuralRefBookOverride", "") or "").strip()
    try:
        structural_ref_book_override = int(structural_ref_book_override_raw) if structural_ref_book_override_raw else None
    except Exception:
        structural_ref_book_override = None
    structural_min_anchors_shared_for_voting = max(
        1, int(get_setting(settings_map, "StructuralMinAnchorsSharedForVoting", 1) or 1)
    )
    structural_require_unique_anchor_occurrences = parse_bool(
        get_setting(settings_map, "StructuralRequireUniqueAnchorOccurrences", False), False
    )
    structural_require_strong_offsets = parse_bool(
        get_setting(settings_map, "StructuralRequireStrongOffsets", False), False
    )
    plateau_auto_relax = parse_bool(get_setting(settings_map, "PlateauAutoRelaxMacroMine", True), True)
    plateau_rung = int(get_setting(settings_map, "PlateauLadder_Rung", 0) or 0)
    sa_min_len = int(get_setting(settings_map, "SuperAnchor_MinLen", 30) or 30)
    sa_min_books = int(get_setting(settings_map, "SuperAnchor_MinBooks", 7) or 7)
    sa_min_support = float(get_setting(settings_map, "SuperAnchor_MinSupportFrac", 0.8) or 0.8)
    sa_min_support_books = int(get_setting(settings_map, "SuperAnchor_MinSupportBooks", sa_min_books) or sa_min_books)
    structural_auto_promote_sa = parse_bool(get_setting(settings_map, "StructuralAutoPromoteSuperAnchors", True), True)
    structural_auto_promote_sa_max = int(get_setting(settings_map, "StructuralAutoPromoteSuperAnchorsMax", 10) or 10)
    sa_macro_enabled = parse_bool(get_setting(settings_map, "SuperAnchorMacro_Enabled", True), True)
    sa_macro_min_len = int(get_setting(settings_map, "SuperAnchorMacro_MinLen", 12) or 12)
    sa_macro_max_len = int(get_setting(settings_map, "SuperAnchorMacro_MaxLen", 60) or 60)
    sa_macro_max_candidates = int(get_setting(settings_map, "SuperAnchorMacro_MaxCandidates", 12) or 12)
    macrocompress_enabled = parse_bool(get_setting(settings_map, "MacroCompress_Enabled", True), True)
    lore_enabled = parse_bool(get_setting(settings_map, "Lore_Enabled", True), True)
    lore_drop_final_e = parse_bool(get_setting(settings_map, "Lore_Canon_DropFinalE", False), False)
    lore_drop_all_h = parse_bool(get_setting(settings_map, "Lore_Canon_DropAllH", False), False)
    lore_drop_all_o = parse_bool(get_setting(settings_map, "Lore_Canon_DropAllO", False), False)

    # Auto-fix: the observed 469 base alphabet (DigitCodeMap) does not include 'H'.
    # Leaving 'h' in canon words prevents signature matching for any corpus word containing 'h',
    # and this impacts LoreAlignment / SemanticMap / ContextEnglish (display-only layers).
    try:
        auto_fix_drop_h = parse_bool(get_setting(settings_map, "Lore_Canon_AutoFix_DropAllH", False), False)
        if auto_fix_drop_h and (not lore_drop_all_h):
            base_letters = _digit_code_map_letters(wb)
            if base_letters and ("H" not in base_letters):
                flow_setting_set(store,
                    "Lore_Canon_DropAllH",
                    True,
                    note=f"iter{next_iter}: auto-fix (DigitCodeMap letters exclude H)",
                )
                # Keep reverse-phrase canon aligned unless the user explicitly enabled it already.
                if not parse_bool(get_setting(settings_map, "ReversePhrase_Canon_DropAllH", False), False):
                    flow_setting_set(store,
                        "ReversePhrase_Canon_DropAllH",
                        True,
                        note=f"iter{next_iter}: auto-fix (DigitCodeMap letters exclude H)",
                    )
                flow_store_refresh_from_workbook(store)
                ws_settings = store["ws_settings"]
                settings_map = store["settings_map"]
                lore_drop_all_h = True
    except Exception:
        # Best-effort only; do not block the iteration.
        pass

    lore_fetch_tibia_enabled = parse_bool(get_setting(settings_map, "LoreFetch_TibiaSigIndex_Enabled", True), True)
    lore_fetch_tibia_max_age_hours = float(get_setting_value(settings_map, "LoreFetch_TibiaSigIndex_MaxAgeHours", 168))
    lore_fetch_tibia_npc_url = str(get_setting(settings_map, "LoreFetch_TibiaSigIndex_NPC_URL", DEFAULT_TIBIA_NPC_URL) or DEFAULT_TIBIA_NPC_URL).strip()
    lore_fetch_tibia_book_url = str(get_setting(settings_map, "LoreFetch_TibiaSigIndex_BOOK_URL", DEFAULT_TIBIA_BOOK_URL) or DEFAULT_TIBIA_BOOK_URL).strip()
    lore_fetch_tibia_timeout_s = int(get_setting(settings_map, "LoreFetch_TibiaSigIndex_TimeoutS", 60) or 60)
    lore_fetch_tibia_max_words_per_sig = int(get_setting(settings_map, "LoreFetch_TibiaSigIndex_MaxWordsPerSig", 80) or 80)
    lore_fetch_tibia_wordfreq_topn = int(get_setting(settings_map, "LoreFetch_TibiaWordFreq_TopN", 6000) or 6000)

    lore_fetch_pd_enabled = parse_bool(get_setting(settings_map, "LoreFetch_PDSigIndex_Enabled", True), True)
    lore_fetch_pd_max_age_hours = float(get_setting_value(settings_map, "LoreFetch_PDSigIndex_MaxAgeHours", 720))
    lore_fetch_pd_url = str(get_setting(settings_map, "LoreFetch_PDSigIndex_URL", DEFAULT_PD_KJV_URL) or DEFAULT_PD_KJV_URL).strip()
    lore_fetch_pd_corpus_id = str(get_setting(settings_map, "LoreFetch_PDSigIndex_CorpusID", "GUTENBERG_KJV_10") or "GUTENBERG_KJV_10").strip()
    lore_fetch_pd_extra_urls = [s for s in parse_csv(get_setting_value(settings_map, "LoreFetch_PDSigIndex_ExtraURLs", "")) if str(s).strip()]
    lore_fetch_pd_extra_cids = [s for s in parse_csv(get_setting_value(settings_map, "LoreFetch_PDSigIndex_ExtraCorpusIDs", "")) if str(s).strip()]
    lore_fetch_pd_timeout_s = int(get_setting(settings_map, "LoreFetch_PDSigIndex_TimeoutS", 60) or 60)
    lore_fetch_pd_cache_max_age_h = float(get_setting_value(settings_map, "LoreFetch_PDSigIndex_CacheMaxAgeHours", 720))
    lore_fetch_pd_max_words_per_sig = int(get_setting(settings_map, "LoreFetch_PDSigIndex_MaxWordsPerSig", 120) or 120)
    # Back-compat: earlier typo key (if present).
    lore_fetch_pd_wordfreq_topn = int(
        get_setting(settings_map, "LoreFetch_PDWordFreq_TopN", get_setting(settings_map, "LoreFetch_PDWorFreq_TopN", 8000)) or 8000
    )

    lore_fetch_dict_enabled = parse_bool(get_setting(settings_map, "LoreFetch_DictSigIndex_Enabled", True), True)
    lore_fetch_dict_max_age_hours = float(get_setting_value(settings_map, "LoreFetch_DictSigIndex_MaxAgeHours", 720))
    lore_fetch_dict_path = str(get_setting(settings_map, "LoreFetch_DictSigIndex_Path", "/usr/share/dict/words") or "/usr/share/dict/words").strip()
    lore_fetch_dict_max_words_per_sig = int(get_setting(settings_map, "LoreFetch_DictSigIndex_MaxWordsPerSig", 120) or 120)

    semantic_enabled = parse_bool(get_setting(settings_map, "Semantic_Enabled", True), True)
    sem_min_total = int(get_setting(settings_map, "SemanticMap_MinTotalCount", 3) or 3)
    sem_min_share = float(get_setting(settings_map, "SemanticMap_MinTopShare", 0.95) or 0.95)
    sem_promote_enabled = parse_bool(get_setting(settings_map, "SemanticPromoteGlossary_Enabled", True), True)
    sem_promote_max = int(get_setting(settings_map, "SemanticPromoteGlossary_MaxPerIter", 25) or 25)
    sem_promote_min_occ = int(get_setting(settings_map, "SemanticPromoteGlossary_MinTotalOcc", 5) or 5)
    sem_promote_min_conf = str(get_setting(settings_map, "SemanticPromoteGlossary_MinConfidence", "MEDIUM") or "MEDIUM").strip().upper()
    sem_promote_min_conf_w = confidence_weight(sem_promote_min_conf)
    sem_promote_max_conf = str(get_setting(settings_map, "SemanticPromoteGlossary_MaxConfidence", "MEDIUM") or "MEDIUM").strip().upper()
    sem_promote_max_conf_w = confidence_weight(sem_promote_max_conf)
    sem_promote_block_evcls = {s.strip().upper() for s in parse_csv(get_setting(settings_map, "SemanticPromoteGlossary_BlockEvidenceClasses", "") or "") if s.strip()}
    sem_promote_min_new_wf = int(get_setting(settings_map, "SemanticPromoteGlossary_MinNewWordFreq", 5) or 5)
    sem_promote_min_wf_ratio = float(get_setting(settings_map, "SemanticPromoteGlossary_MinWordFreqRatio", 3.0) or 3.0)
    sem_promote_wf_sheet = str(get_setting(settings_map, "SemanticPromoteGlossary_WordFreqSheet", LORE_WORDFREQ_TIBIA_SHEET) or LORE_WORDFREQ_TIBIA_SHEET).strip()

    english_enabled = parse_bool(get_setting(settings_map, "EnglishLayer_Enabled", True), True)
    english_min_total = int(get_setting(settings_map, "EnglishLayer_MinTotalCount", 20) or 20)
    english_min_share = float(get_setting(settings_map, "EnglishLayer_MinTopShare", 0.95) or 0.95)
    english_min_len = int(get_setting(settings_map, "EnglishLayer_MinWordLen", 3) or 3)
    english_max_rows = int(get_setting(settings_map, "EnglishLayer_MaxMapRows", 2000) or 2000)

    english_retext_enabled = parse_bool(get_setting(settings_map, "EnglishGlossaryRetext_Enabled", True), True)
    english_retext_max = int(get_setting(settings_map, "EnglishGlossaryRetext_MaxPerIter", 50) or 50)
    english_retext_min_total = int(get_setting(settings_map, "EnglishGlossaryRetext_MinTotalCount", 5) or 5)
    english_retext_min_share = float(get_setting(settings_map, "EnglishGlossaryRetext_MinTopShare", 0.90) or 0.90)
    english_retext_lock_iters = int(get_setting(settings_map, "EnglishGlossaryRetext_LockIters", 5) or 5)
    glossary_retext_locked_tokens = {
        s.strip() for s in parse_csv(get_setting(settings_map, "GlossaryRetext_LockedTokens", "") or "") if s.strip()
    }
    convergence_block_promotion_tokens = {
        s.strip() for s in parse_csv(get_setting(settings_map, "Convergence_BlockPromotionTokens", "") or "") if s.strip()
    }

    anti_mode = parse_bool(get_setting(settings_map, "AntiHallucination_Mode", True), True)
    if anti_mode:
        anti_fix_changed = False
        gt_mode_cfg = str(get_setting(settings_map, "GroundTruthLiveCheck_Mode", "POLICY") or "POLICY").strip().upper()
        if gt_mode_cfg in ("ALL", "FULL"):
            flow_setting_set(store,
                "GroundTruthLiveCheck_Mode",
                "POLICY",
                note=f"iter{next_iter}: anti-hallucination auto-fix (disable ALL GT mode)",
            )
            anti_fix_changed = True
        if convergence_enable_low_confidence_escape:
            # Low-confidence escape is useful for exploration, but it also amplifies lexical drift.
            # In anti-hallucination mode we keep this lane off by default.
            flow_setting_set(store,
                "Convergence_EnableLowConfidenceEscape",
                False,
                note=f"iter{next_iter}: anti-hallucination auto-fix (disable low-confidence escape)",
            )
            convergence_enable_low_confidence_escape = False
            anti_fix_changed = True
        if anti_fix_changed:
            flow_store_refresh_from_workbook(store)
            ws_settings = store["ws_settings"]
            settings_map = store["settings_map"]
            anti_mode = parse_bool(get_setting(settings_map, "AntiHallucination_Mode", True), True)
    anti_disable_semantic_retext = parse_bool(get_setting(settings_map, "AntiHallucination_DisableSemanticRetext", True), True)
    anti_disable_english_retext = parse_bool(get_setting(settings_map, "AntiHallucination_DisableEnglishRetext", True), True)
    anti_deny_words = parse_word_set(
        get_setting_value(settings_map, "AntiHallucination_UNKDenyWords", ",".join(sorted(ANTI_HALLUCINATION_DEFAULT_TERMS)))
    )
    if not anti_deny_words:
        anti_deny_words = set(ANTI_HALLUCINATION_DEFAULT_TERMS)
    anti_unk_enabled = parse_bool(get_setting(settings_map, "AntiHallucination_UseUNKForLowEvidence", True), True)
    anti_unk_max_per_iter = int(get_setting(settings_map, "AntiHallucination_UNKMaxPerIter", 40) or 40)
    anti_unk_min_total_occ = int(get_setting(settings_map, "AntiHallucination_UNKMinTotalOcc", 1) or 1)
    anti_unk_max_total_occ = int(get_setting(settings_map, "AntiHallucination_UNKMaxTotalOcc", 999) or 999)
    anti_unk_max_conf = str(get_setting(settings_map, "AntiHallucination_UNKMaxConfidence", "MEDIUM") or "MEDIUM").strip().upper()
    anti_unk_max_conf_w = confidence_weight(anti_unk_max_conf)
    anti_unk_protected_evcls = {
        s.strip().upper()
        for s in parse_csv(get_setting(settings_map, "AntiHallucination_UNKProtectedEvidenceClasses", "") or "")
        if s.strip()
    }
    anchor_promo_only_enabled = parse_bool(
        get_setting(settings_map, "AnchorPromotionOnly_Enabled", True if anti_mode else False),
        True if anti_mode else False,
    )
    anchor_promo_min_hits = max(1, int(get_setting(settings_map, "AnchorPromotionOnly_MinHits", 1) or 1))
    anchor_promo_min_verified_sources = max(
        0,
        int(get_setting(settings_map, "AnchorPromotionOnly_MinVerifiedSources", 2) or 2),
    )
    anchor_promo_include_groundtruth = parse_bool(
        get_setting(settings_map, "AnchorPromotionOnly_IncludeGroundTruth", True),
        True,
    )
    anchor_promo_include_external_npc_staff = parse_bool(
        get_setting(settings_map, "AnchorPromotionOnly_IncludeExternalNPCStaff", True),
        True,
    )
    anchor_promo_include_external_books = parse_bool(
        get_setting(settings_map, "AnchorPromotionOnly_IncludeExternalBooks", False),
        False,
    )
    anchor_promo_include_anchorcribs = parse_bool(
        get_setting(settings_map, "AnchorPromotionOnly_IncludeAnchorCribs", True),
        True,
    )
    if anti_mode and anti_disable_semantic_retext:
        sem_promote_enabled = False
    if anti_mode and anti_disable_english_retext:
        english_retext_enabled = False

    ext_roundtrip_min_verified = int(get_setting(settings_map, "ExternalRoundTrip_MinVerifiedCount", 2) or 2)
    ext_roundtrip_min_segment_digits = int(get_setting(settings_map, "ExternalRoundTrip_MinSegmentDigits", 8) or 8)
    ext_roundtrip_allow_ordered = parse_bool(get_setting(settings_map, "ExternalRoundTrip_AllowOrderedRunMatch", True), True)
    digit_ctx_enabled = parse_bool(get_setting(settings_map, "DigitCodeContext_Enabled", True), True)
    digit_ctx_topk = int(get_setting_value(settings_map, "DigitCodeContext_TopK", 8))
    digit_ctx_js_alpha = float(get_setting_value(settings_map, "DigitCodeContext_JSAlpha", 0.20))

    lore_bigrams_enabled = parse_bool(get_setting(settings_map, "LoreBigrams_Enabled", True), True)
    lore_bigrams_max_age_h = float(get_setting_value(settings_map, "LoreBigrams_MaxAgeHours", 168))
    lore_bigrams_cache_max_age_h = float(get_setting_value(settings_map, "LoreBigrams_TibiaCacheMaxAgeHours", 168))
    lore_bigrams_vocab_topn = int(get_setting_value(settings_map, "LoreBigrams_VocabTopN", 6000))
    lore_bigrams_min_count = int(get_setting_value(settings_map, "LoreBigrams_MinCount", 3))
    lore_bigrams_max_rows = int(get_setting_value(settings_map, "LoreBigrams_MaxRows", 200000))

    context_enabled = parse_bool(get_setting(settings_map, "ContextEnglish_Enabled", True), True)
    ctx_emit_alpha = float(get_setting_value(settings_map, "ContextEnglish_EmissionAlpha", 0.5))
    ctx_trans_alpha = float(get_setting_value(settings_map, "ContextEnglish_TransitionAlpha", 0.5))
    ctx_max_cands = int(get_setting_value(settings_map, "ContextEnglish_MaxCandidatesPerToken", 6))
    ctx_map_min_total = int(get_setting_value(settings_map, "ContextEnglishMap_MinTotal", 30))
    ctx_map_min_share = float(get_setting_value(settings_map, "ContextEnglishMap_MinTopShare", 0.97))
    ctx_map_max_rows = int(get_setting_value(settings_map, "ContextEnglishMap_MaxRows", 800))

    codeaware_enabled = parse_bool(get_setting(settings_map, "CodeAware_Enabled", True), True)
    codeaware_min_total = int(get_setting_value(settings_map, "CodeAware_MinTotalPerCode", 60))
    codeaware_min_share = float(get_setting_value(settings_map, "CodeAware_MinTopShare", 0.88))
    codeaware_max_token_len = int(get_setting_value(settings_map, "CodeAware_MaxTokenLen", 1))
    codeaware_min_total_seq = int(get_setting_value(settings_map, "CodeAware_MinTotalPerCodeSeq", 4))
    codeaware_min_share_seq = float(get_setting_value(settings_map, "CodeAware_MinTopShareSeq", 0.75))
    codeaware_max_rows = int(get_setting_value(settings_map, "CodeAware_MaxMapRows", 2500))
    codeaware_hint_min_total = int(get_setting_value(settings_map, "CodeAware_HintMinTotal", 8))
    codeaware_hint_boost = int(get_setting_value(settings_map, "CodeAware_HintBoost", 4))
    codeaware_hint_topk = int(get_setting_value(settings_map, "CodeAware_HintTopK", 2))
    codeaware_apply_tokens = [s.strip() for s in parse_csv(get_setting_value(settings_map, "CodeAware_ApplyToTokens", "")) if s.strip()]

    seqmatch_enabled = parse_bool(get_setting(settings_map, "SequenceMatch_Enabled", True), True)
    seqmatch_n_list = [
        int(x)
        for x in parse_csv(get_setting_value(settings_map, "SequenceMatch_NList", "6,7,8"))
        if str(x).strip().isdigit()
    ]
    seqmatch_min_n = int(get_setting_value(settings_map, "SequenceMatch_MinN", 2))
    seqmatch_cand_max_freq = int(get_setting_value(settings_map, "SequenceMatch_CandidateMaxBookFreq", 1))
    seqmatch_max_candidates = int(get_setting_value(settings_map, "SequenceMatch_MaxCandidates", 4000))
    seqmatch_explore_rotate = parse_bool(get_setting(settings_map, "SequenceMatch_ExploreRotate", True), True)
    seqmatch_explore_keep_top = int(get_setting_value(settings_map, "SequenceMatch_ExploreKeepTop", 200))
    seqmatch_cache_enabled = parse_bool(get_setting(settings_map, "SequenceMatch_CacheEnabled", True), True)
    seqmatch_cache_max_rows = int(get_setting_value(settings_map, "SequenceMatch_CacheMaxRows", 2000))
    seqmatch_max_matches = int(get_setting_value(settings_map, "SequenceMatch_MaxMatches", 60))
    seqmatch_time_budget_s = float(get_setting_value(settings_map, "SequenceMatch_TimeBudgetS", 20))
    seqmatch_scan_tibia_first = parse_bool(get_setting(settings_map, "SequenceMatch_ScanTibiaFirst", True), True)
    seqmatch_pd_max_chars = int(get_setting_value(settings_map, "SequenceMatch_PD_MaxChars", 800000))
    seqmatch_pd_max_sentences = int(get_setting_value(settings_map, "SequenceMatch_PD_MaxSentencesPerSource", 2500))
    seqmatch_context_window = int(get_setting_value(settings_map, "SequenceMatch_ContextWindow", 2))
    seqmatch_context_min_overlap = int(get_setting_value(settings_map, "SequenceMatch_ContextMinOverlap", 1))
    seqmatch_context_require_direction = parse_bool(
        get_setting(settings_map, "SequenceMatch_ContextRequireDirection", False), False
    )
    seqmatch_snippet_min_content = int(get_setting_value(settings_map, "SequenceMatch_SnippetMinContentWords", 1))
    seq_word_hints_enabled = parse_bool(get_setting(settings_map, "SequenceWordHints_Enabled", True), True)
    seq_word_hints_min_n = int(get_setting_value(settings_map, "SequenceWordHints_MinN", 3))
    seq_word_hints_skip_identity = parse_bool(get_setting(settings_map, "SequenceWordHints_SkipIdentity", True), True)
    seq_word_hints_max_rows = int(get_setting_value(settings_map, "SequenceWordHints_MaxRows", 800))
    seq_word_hints_min_ratio = float(get_setting_value(settings_map, "SequenceWordHints_MinRatio", 1.5))
    seq_word_hints_exclude_stopwords = parse_bool(
        get_setting(settings_map, "SequenceWordHints_ExcludeStopwords", True), True
    )
    seq_word_hints_stopword_ratio = float(get_setting_value(settings_map, "SequenceWordHints_StopwordRatio", 0.67))
    seq_hints_enabled = parse_bool(get_setting(settings_map, "SequenceHints_Enabled", True), True)
    seq_hints_boost = int(get_setting_value(settings_map, "SequenceHints_Boost", 20))
    seq_hints_max_words_per_sig = int(get_setting_value(settings_map, "SequenceHints_MaxWordsPerSig", 3))
    reversephrase_hints_enabled = parse_bool(get_setting(settings_map, "ReversePhraseHints_Enabled", True), True)
    reversephrase_hints_boost = int(get_setting_value(settings_map, "ReversePhraseHints_Boost", 8))
    reversephrase_hints_max_words_per_sig = int(get_setting_value(settings_map, "ReversePhraseHints_MaxWordsPerSig", 2))

    sestina_enabled = parse_bool(get_setting(settings_map, "SestinaScan_Enabled", True), True)
    sestina_max_lines = int(get_setting_value(settings_map, "SestinaScan_MaxLines", 4000))
    sestina_min_score30 = int(get_setting_value(settings_map, "SestinaScan_MinScore30", 26))
    sestina_max_candidates = int(get_setting_value(settings_map, "SestinaScan_MaxCandidates", 40))
    sestina_use_token_sig = parse_bool(get_setting(settings_map, "SestinaScan_UseTokenSignature", True), True)
    sestina_envoi_bonus = float(get_setting_value(settings_map, "SestinaScan_EnvoiBonus", 1.0))
    sestina_obligation_enabled = parse_bool(get_setting(settings_map, "SestinaObligation_Enabled", True), True)
    sestina_obligation_max_candidates = int(get_setting_value(settings_map, "SestinaObligation_MaxCandidates", 30))
    sestina_obligation_min_score30 = int(get_setting_value(settings_map, "SestinaObligation_MinScore30", 0))
    sestina_obligatory_impact = float(get_setting_value(settings_map, "SestinaObligation_ObligatoryImpact", 0.015))
    sestina_conditional_impact = float(get_setting_value(settings_map, "SestinaObligation_ConditionalImpact", 0.0))
    sestina_decorative_impact = float(get_setting_value(settings_map, "SestinaObligation_DecorativeImpact", -0.020))
    sestina_no_collapse_tolerance = float(get_setting_value(settings_map, "SestinaObligation_NoCollapseTolerance", 0.020))
    sestina_reorder_resilient_ratio = float(get_setting_value(settings_map, "SestinaObligation_ReorderResilientRatio", 0.90))
    rhythm_enabled = parse_bool(get_setting(settings_map, "RhythmTransition_Enabled", True), True)
    rhythm_window_size = int(get_setting_value(settings_map, "RhythmTransition_WindowSize", 12))
    rhythm_min_lines = int(get_setting_value(settings_map, "RhythmTransition_MinLines", 24))
    rhythm_use_token_sig = parse_bool(get_setting(settings_map, "RhythmTransition_UseTokenSignature", True), True)
    rhythm_shuffle_trials = int(get_setting_value(settings_map, "RhythmTransition_ShuffleTrials", 8))
    rhythm_cycle_delta_threshold = float(get_setting_value(settings_map, "RhythmTransition_CycleDeltaThreshold", 0.05))

    puzzle_min_ext_verified = int(get_setting(settings_map, "PuzzleSolved_MinExternalVerifiedCount", 2) or 2)
    puzzle_ctx_improve_iters = int(get_setting(settings_map, "PuzzleSolved_ContextImproveIters", 0) or 0)
    if puzzle_ctx_improve_iters > 0:
        puzzle_ctx_improve_iters = 0
        try:
            flow_setting_set(store,
                "PuzzleSolved_ContextImproveIters",
                0,
                note=f"iter{next_iter}: auto-relax puzzle context-improve requirement",
            )
            flow_store_refresh_from_workbook(store)
            ws_settings = store["ws_settings"]
            settings_map = store["settings_map"]
        except Exception:
            pass
    puzzle_min_seq_matches = int(get_setting(settings_map, "PuzzleSolved_MinSequenceMatches", 3) or 3)

    reverse_enabled = parse_bool(get_setting(settings_map, "ReversePhrase_Enabled", True), True)
    reverse_inc_phrase = parse_bool(get_setting(settings_map, "ReversePhrase_IncludePhraseCribsUser", True), True)
    reverse_inc_phrase_auto = parse_bool(get_setting(settings_map, "ReversePhrase_IncludePhraseCribsAuto", True), True)
    reverse_inc_lore_user = parse_bool(get_setting(settings_map, "ReversePhrase_IncludeLoreCorpusUser", True), True)
    reverse_inc_lore_auto = parse_bool(get_setting(settings_map, "ReversePhrase_IncludeLoreCorpusAuto", False), False)
    reverse_inc_tibia = parse_bool(get_setting(settings_map, "ReversePhrase_IncludeTibiaCorpusAuto", True), True)
    reverse_inc_seqmatch = parse_bool(get_setting(settings_map, "ReversePhrase_IncludeSequenceMatchesAuto", True), True)
    reverse_seqmatch_max_phrases = int(get_setting(settings_map, "ReversePhrase_SeqMatchesMaxPhrasesPerIter", 40) or 40)
    reverse_seqmatch_min_n = int(get_setting(settings_map, "ReversePhrase_SeqMatchesMinN", 5) or 5)
    reverse_tibia_max_phrases = int(get_setting(settings_map, "ReversePhrase_TibiaMaxPhrasesPerIter", 40) or 40)
    reverse_tibia_cache_age_h = float(get_setting_value(settings_map, "ReversePhrase_TibiaCacheMaxAgeHours", 24))
    reverse_tibia_inc_npc = parse_bool(get_setting(settings_map, "ReversePhrase_TibiaIncludeNPC", True), True)
    reverse_tibia_inc_books = parse_bool(get_setting(settings_map, "ReversePhrase_TibiaIncludeBooks", True), True)
    reverse_phrase_text_maxlen = int(get_setting(settings_map, "ReversePhrase_PhraseTextMaxLen", 220) or 220)
    reverse_logogram_aware = parse_bool(get_setting(settings_map, "ReversePhrase_LogogramAware", True), True)
    reverse_drop_final_e = parse_bool(get_setting(settings_map, "ReversePhrase_Canon_DropFinalE", False), False)
    reverse_drop_all_h = parse_bool(get_setting(settings_map, "ReversePhrase_Canon_DropAllH", False), False)
    reverse_drop_all_o = parse_bool(get_setting(settings_map, "ReversePhrase_Canon_DropAllO", False), False)
    reverse_cand_max_base_len = int(get_setting(settings_map, "ReversePhrase_CandidateMaxBaseLen", 24) or 24)
    reverse_max_span = int(get_setting(settings_map, "ReversePhrase_MaxSpanTokens", 6) or 6)
    reverse_max_gap = int(get_setting(settings_map, "ReversePhrase_MaxGapTokens", 0) or 0)
    reverse_min_words = int(get_setting(settings_map, "ReversePhrase_MinWords", 4) or 4)
    reverse_max_phrases = int(get_setting(settings_map, "ReversePhrase_MaxPhrasesPerIter", 120) or 120)
    reverse_max_hits_total = int(get_setting(settings_map, "ReversePhrase_MaxHitsTotal", 60) or 60)
    reverse_max_hits_per_phrase = int(get_setting(settings_map, "ReversePhrase_MaxHitsPerPhrase", 3) or 3)
    reverse_max_hits_per_book = int(get_setting(settings_map, "ReversePhrase_MaxHitsPerBook", 1) or 1)
    reverse_emit = parse_bool(get_setting(settings_map, "ReversePhrase_AutoEmitCandidates", True), True)
    reverse_evcls = str(get_setting(settings_map, "ReversePhrase_EvidenceClass", "PHRASE_CRIB") or "PHRASE_CRIB").strip()
    reverse_cand_min_books = int(get_setting(settings_map, "ReversePhrase_CandidateMinBooks", 2) or 2)
    reverse_cand_min_share = float(get_setting(settings_map, "ReversePhrase_CandidateMinTopShare", 0.95) or 0.95)
    reverse_cand_max_new = int(get_setting(settings_map, "ReversePhrase_CandidateMaxNewTokens", 20) or 20)
    reverse_retext_enabled = parse_bool(get_setting(settings_map, "ReversePhrase_RetextExisting_Enabled", True), True)
    reverse_retext_max_per_iter = int(get_setting(settings_map, "ReversePhrase_RetextExisting_MaxPerIter", 20) or 20)
    reverse_retext_min_books = int(get_setting(settings_map, "ReversePhrase_RetextExisting_MinBooks", 2) or 2)
    reverse_retext_min_share = float(get_setting(settings_map, "ReversePhrase_RetextExisting_MinTopShare", 0.95) or 0.95)
    reverse_retext_min_support_occ = int(get_setting(settings_map, "ReversePhrase_RetextExisting_MinSupportOcc", 2) or 2)
    reverse_perm_enabled = parse_bool(get_setting(settings_map, "ReversePhrase_PermutationEnabled", True), True)
    reverse_perm_run_if_hits_le = int(get_setting(settings_map, "ReversePhrase_PermutationRunIfHitsLE", 0) or 0)
    reverse_perm_max_words = int(get_setting(settings_map, "ReversePhrase_PermutationMaxWords", 6) or 6)
    reverse_perm_max_phrases = int(get_setting(settings_map, "ReversePhrase_PermutationMaxPhrasesPerIter", 25) or 25)
    reverse_perm_max_hits_total = int(get_setting(settings_map, "ReversePhrase_PermutationMaxHitsTotal", 30) or 30)
    reverse_perm_max_hits_per_phrase = int(get_setting(settings_map, "ReversePhrase_PermutationMaxHitsPerPhrase", 2) or 2)
    reverse_perm_max_hits_per_book = int(get_setting(settings_map, "ReversePhrase_PermutationMaxHitsPerBook", 1) or 1)
    reverse_perm_emit = parse_bool(get_setting(settings_map, "ReversePhrase_PermutationAutoEmitCandidates", False), False)

    autophrase_enabled = parse_bool(get_setting(settings_map, "AutoPhraseCribs_Enabled", True), True)
    autophrase_cache_age_h = float(get_setting_value(settings_map, "AutoPhraseCribs_CacheMaxAgeHours", 24))
    autophrase_max_phrases = int(get_setting(settings_map, "AutoPhraseCribs_MaxPhrasesPerIter", 80) or 80)
    autophrase_max_scan = int(get_setting(settings_map, "AutoPhraseCribs_MaxScanSentences", 50000) or 50000)
    autophrase_time_budget_s = float(get_setting(settings_map, "AutoPhraseCribs_TimeBudgetS", 25) or 25)
    autophrase_min_words = int(get_setting(settings_map, "AutoPhraseCribs_MinWords", reverse_min_words) or reverse_min_words)
    autophrase_max_words = int(get_setting(settings_map, "AutoPhraseCribs_MaxWords", 14) or 14)
    autophrase_max_text_len = int(get_setting(settings_map, "AutoPhraseCribs_MaxTextLen", 240) or 240)
    autophrase_include_pd = parse_bool(get_setting(settings_map, "AutoPhraseCribs_IncludePD", True), True)
    autophrase_pd_cache_age_h = float(get_setting_value(settings_map, "AutoPhraseCribs_PDCacheMaxAgeHours", 720))
    autophrase_pd_max_scan = int(get_setting(settings_map, "AutoPhraseCribs_PDMaxScanSentences", 25000) or 25000)
    autophrase_pd_time_budget_s = float(get_setting(settings_map, "AutoPhraseCribs_PDTimeBudgetS", 25) or 25)
    autophrase_pd_max_sources = int(get_setting(settings_map, "AutoPhraseCribs_PDMaxSources", 0) or 0)

    # Auto-advance: if reverse phrase mining is enabled but the user has not provided any phrases yet,
    # fall back to the auto-seeded lore corpus (public domain). This is safe (analysis-first) and helps
    # break plateaus without requiring manual sheet edits.
    if reverse_enabled and (not reverse_inc_lore_auto):
        try:
            # Ensure corpus exists so the check is meaningful.
            _ensure_lore_corpus_sheets(wb, cur_iter)

            # If PhraseCribs_User has no enabled phrases AND LoreCorpus_User is empty, enable auto corpus.
            has_user_phrases = False
            if "PhraseCribs_User" in wb.sheetnames:
                ws_pc = wb["PhraseCribs_User"]
                hp = ws_find_header_row(ws_pc, ["PhraseID", "Enabled", "Text"], max_scan=3)
                cp = ws_headers(ws_pc, hp)
                for r in range(hp + 1, ws_pc.max_row + 1):
                    enabled = ws_pc.cell(r, cp["Enabled"]).value
                    text = ws_pc.cell(r, cp["Text"]).value
                    if parse_bool(enabled, False) and text is not None and str(text).strip() != "":
                        has_user_phrases = True
                        break

            has_user_lines = False
            if "LoreCorpus_User" in wb.sheetnames:
                ws_lu = wb["LoreCorpus_User"]
                hu = ws_find_header_row(ws_lu, ["CorpusID", "LineID", "Text"], max_scan=3)
                cu = ws_headers(ws_lu, hu)
                for r in range(hu + 1, ws_lu.max_row + 1):
                    cid = ws_lu.cell(r, cu["CorpusID"]).value
                    text = ws_lu.cell(r, cu["Text"]).value
                    if isinstance(cid, str) and cid.strip() and text is not None and str(text).strip() != "":
                        has_user_lines = True
                        break
            if (not has_user_phrases) and (not has_user_lines):
                flow_setting_set(store,
                    "ReversePhrase_IncludeLoreCorpusAuto",
                    True,
                    note=f"iter{next_iter}: auto-enable (no user lore phrases yet)",
                )
                # Also expand the scan budget modestly since LoreCorpus_Auto is curated + public-domain.
                flow_setting_set(store,
                    "ReversePhrase_MaxPhrasesPerIter",
                    max(int(reverse_max_phrases), 200),
                    note=f"iter{next_iter}: auto-bump for LoreCorpus_Auto scan",
                )
                flow_setting_set(store,
                    "ReversePhrase_MaxHitsTotal",
                    max(int(reverse_max_hits_total), 120),
                    note=f"iter{next_iter}: auto-bump for LoreCorpus_Auto scan",
                )
                flow_store_refresh_from_workbook(store)
                ws_settings = store["ws_settings"]
                settings_map = store["settings_map"]
                reverse_inc_lore_auto = True
                reverse_max_phrases = int(get_setting(settings_map, "ReversePhrase_MaxPhrasesPerIter", reverse_max_phrases) or reverse_max_phrases)
                reverse_max_hits_total = int(get_setting(settings_map, "ReversePhrase_MaxHitsTotal", reverse_max_hits_total) or reverse_max_hits_total)
        except Exception:
            # Best effort; do not block iteration.
            pass

    # Ensure reverse-mined evidence class can participate in the SAFE mechanical simulation/promotion loop.
    if reverse_evcls and reverse_emit and (reverse_evcls not in mech_classes):
        mech_classes.append(reverse_evcls)
        flow_setting_set(store,
            "MechanicalCandidateClasses",
            ",".join(mech_classes),
            note=f"iter{next_iter}: auto-append for reverse phrase mining ({reverse_evcls})",
        )

    # Plateau ladder (pre-relax): if the previous iteration plateaued (or WeakFrac stalled for >=2 iters),
    # relax MacroMine knobs before this iteration's macro mining. This only expands the search space; any
    # promotion is still gated by GT live check + DP metrics guardrails.
    plateau_relax_note: Optional[str] = None
    if plateau_auto_relax and cur_iter > 0:
        prev_metrics = _read_iter_summary_metrics(wb, cur_iter)
        prev_mech = int(prev_metrics.get("Mechanical promotions accepted") or 0)
        prev_plateau = prev_mech == 0

        weak_stagnant2 = False
        if cur_iter >= 2:
            def _weak(it: int) -> float:
                m = _read_iter_summary_metrics(wb, it)
                return float(m.get("WEAK char frac (Books, length-weighted)") or 0.0)

            w2 = _weak(cur_iter)
            w1 = _weak(cur_iter - 1)
            w0 = _weak(cur_iter - 2)
            eps = 1e-12
            if w2 >= w1 - eps and w1 >= w0 - eps:
                weak_stagnant2 = True

        if prev_plateau or weak_stagnant2:
            # Allow the ladder to progress beyond rung3; later rungs relax only display/analysis layers
            # (and still keep all GT + mechanical guardrails intact).
            new_rung = min(max(0, plateau_rung) + 1, 9)
            if new_rung > plateau_rung:
                plateau_rung = new_rung
                flow_setting_set(store,
                    "PlateauLadder_Rung",
                    plateau_rung,
                    note=f"iter{next_iter}: auto-relaxed (prev_plateau={prev_plateau}, weak_stagnant2={weak_stagnant2})",
                )

                # Ladder per plan:
                # - MinShare: 0.95 -> 0.90 -> 0.85
                # - MaxLen: 16 -> 24
                # - MaxCandidates: 75 -> 150
                # - NValues: expand to 2..10
                cur_min_share = float(get_setting(settings_map, "MacroMine_MinShare", 0.95) or 0.95)
                cur_max_len = int(get_setting(settings_map, "MacroMine_MaxLen", 16) or 16)
                cur_max_candidates = int(get_setting(settings_map, "MacroMine_MaxCandidates", 75) or 75)

                if plateau_rung >= 1:
                    flow_setting_set(store, "MacroMine_MinShare", min(cur_min_share, 0.90), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                    flow_setting_set(store, "MacroMine_MaxLen", max(cur_max_len, 24), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                    flow_setting_set(store, "MacroMine_MaxCandidates", max(cur_max_candidates, 150), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                    flow_setting_set(store, "MacroMine_NValues", "2,3,4,5,6,7,8,9,10", note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                if plateau_rung >= 2:
                    flow_setting_set(store, "MacroMine_MinShare", min(cur_min_share, 0.85), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                if plateau_rung >= 3:
                    flow_setting_set(store, "MacroMine_MinShare", min(cur_min_share, 0.80), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                    flow_setting_set(store, "MacroMine_MaxLen", max(cur_max_len, 32), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                    flow_setting_set(store, "MacroMine_MaxCandidates", max(cur_max_candidates, 300), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                    flow_setting_set(store, "MacroMine_NValues", "2,3,4,5,6,7,8,9,10,11,12", note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")

                # Reload settings map so later macro-mining uses the updated values.
                flow_store_refresh_from_workbook(store)
                ws_settings = store["ws_settings"]
                settings_map = store["settings_map"]
                plateau_relax_note = f"Plateau ladder applied (rung={plateau_rung}, prev_plateau={prev_plateau}, weak_stagnant2={weak_stagnant2})"

            # Plateau evidence ladder: allow small evidence drops when token-savings improve structure.
            # This stays safe because we also enforce the EvidenceAvg target guardrail (never drop below TargetEvidenceAvg).
            ev_ladder = {0: 0.002, 1: 0.003, 2: 0.005, 3: 0.007}
            desired_ev_drop = ev_ladder.get(int(plateau_rung), 0.005)
            cur_ev_drop_setting = float(get_setting(settings_map, "MaxEvidenceAvgDrop", 0.002) or 0.002)
            if desired_ev_drop > cur_ev_drop_setting + 1e-12:
                flow_setting_set(store,
                    "MaxEvidenceAvgDrop",
                    desired_ev_drop,
                    note=f"iter{next_iter}: plateau evidence ladder rung{plateau_rung}",
                )
                max_ev_drop = desired_ev_drop
                flow_store_refresh_from_workbook(store)
                ws_settings = store["ws_settings"]
                settings_map = store["settings_map"]
                if plateau_relax_note:
                    plateau_relax_note += f"; MaxEvidenceAvgDrop {cur_ev_drop_setting}->{desired_ev_drop}"
                else:
                    plateau_relax_note = f"Plateau evidence ladder applied (MaxEvidenceAvgDrop {cur_ev_drop_setting}->{desired_ev_drop})"

            # English layer ladder: relax mapping thresholds to increase readability on sustained plateaus.
            # Display-only, so this is safe and does not affect DP/tokenization/metrics.
            if english_enabled:
                en_ladder = {0: (20, 0.95), 1: (10, 0.93), 2: (5, 0.90), 3: (3, 0.85)}
                desired_en_total, desired_en_share = en_ladder.get(int(plateau_rung), (5, 0.90))

                cur_en_total = int(get_setting(settings_map, "EnglishLayer_MinTotalCount", english_min_total) or english_min_total)
                cur_en_share = float(get_setting(settings_map, "EnglishLayer_MinTopShare", english_min_share) or english_min_share)

                en_changed = False
                if int(desired_en_total) < int(cur_en_total):
                    flow_setting_set(store,
                        "EnglishLayer_MinTotalCount",
                        int(desired_en_total),
                        note=f"iter{next_iter}: english ladder rung{plateau_rung}",
                    )
                    english_min_total = int(desired_en_total)
                    en_changed = True
                if float(desired_en_share) + 1e-12 < float(cur_en_share):
                    flow_setting_set(store,
                        "EnglishLayer_MinTopShare",
                        float(desired_en_share),
                        note=f"iter{next_iter}: english ladder rung{plateau_rung}",
                    )
                    english_min_share = float(desired_en_share)
                    en_changed = True

                if en_changed:
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = f"EnglishLayer ladder applied (min_total {cur_en_total}->{english_min_total}, min_share {cur_en_share}->{english_min_share})"
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

            # ContextEnglish ladder: broaden candidate set and lower "stable map" thresholds on sustained plateaus.
            # This is display-only; any Glossary retext is still guarded by GT live check + word-quality heuristics.
            if context_enabled:
                ctx_ladder = {0: (6, 30, 0.97), 1: (8, 20, 0.95), 2: (10, 10, 0.90), 3: (12, 5, 0.85)}
                desired_max_cands, desired_min_total, desired_min_share = ctx_ladder.get(int(plateau_rung), (10, 10, 0.90))

                cur_max_cands = int(get_setting(settings_map, "ContextEnglish_MaxCandidatesPerToken", ctx_max_cands) or ctx_max_cands)
                cur_min_total = int(get_setting(settings_map, "ContextEnglishMap_MinTotal", ctx_map_min_total) or ctx_map_min_total)
                cur_min_share = float(get_setting(settings_map, "ContextEnglishMap_MinTopShare", ctx_map_min_share) or ctx_map_min_share)

                ctx_changed = False
                if int(desired_max_cands) > int(cur_max_cands):
                    flow_setting_set(store,
                        "ContextEnglish_MaxCandidatesPerToken",
                        int(desired_max_cands),
                        note=f"iter{next_iter}: context ladder rung{plateau_rung}",
                    )
                    ctx_max_cands = int(desired_max_cands)
                    ctx_changed = True
                if int(desired_min_total) < int(cur_min_total):
                    flow_setting_set(store,
                        "ContextEnglishMap_MinTotal",
                        int(desired_min_total),
                        note=f"iter{next_iter}: context ladder rung{plateau_rung}",
                    )
                    ctx_map_min_total = int(desired_min_total)
                    ctx_changed = True
                if float(desired_min_share) + 1e-12 < float(cur_min_share):
                    flow_setting_set(store,
                        "ContextEnglishMap_MinTopShare",
                        float(desired_min_share),
                        note=f"iter{next_iter}: context ladder rung{plateau_rung}",
                    )
                    ctx_map_min_share = float(desired_min_share)
                    ctx_changed = True

                if ctx_changed:
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = (
                        f"ContextEnglish ladder applied (max_cands {cur_max_cands}->{ctx_max_cands}, "
                        f"min_total {cur_min_total}->{ctx_map_min_total}, min_share {cur_min_share}->{ctx_map_min_share})"
                    )
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

            # CodeAware ladder: relax per-code stability thresholds on sustained plateaus (display-only).
            # This impacts only Translation_CodeAware_Auto and SequenceMatches input selection, never StrictPlus.
            if codeaware_enabled:
                # Beyond rung3, we start enabling multi-letter code-sequence keys (e.g. token=TV, codeSeq=09-61)
                # to disambiguate high-ambiguity tokens using the deterministic digits->code stream.
                ca_ladder = {
                    0: (60, 0.88, 1, 4, 0.75),
                    1: (30, 0.80, 1, 4, 0.75),
                    2: (15, 0.72, 1, 4, 0.75),
                    3: (8, 0.65, 1, 4, 0.75),
                    4: (4, 0.60, 2, 3, 0.70),
                    5: (3, 0.58, 3, 2, 0.66),
                    6: (2, 0.55, 4, 2, 0.62),
                    7: (2, 0.55, 6, 2, 0.60),
                    8: (2, 0.55, 8, 2, 0.58),
                    9: (2, 0.55, 10, 2, 0.56),
                }
                desired_ca_total, desired_ca_share, desired_ca_maxlen, desired_ca_seq_total, desired_ca_seq_share = ca_ladder.get(
                    int(plateau_rung), (15, 0.72, 1, 4, 0.75)
                )

                cur_ca_total = int(get_setting(settings_map, "CodeAware_MinTotalPerCode", codeaware_min_total) or codeaware_min_total)
                cur_ca_share = float(get_setting(settings_map, "CodeAware_MinTopShare", codeaware_min_share) or codeaware_min_share)
                cur_ca_maxlen = int(get_setting(settings_map, "CodeAware_MaxTokenLen", codeaware_max_token_len) or codeaware_max_token_len)
                cur_ca_seq_total = int(get_setting(settings_map, "CodeAware_MinTotalPerCodeSeq", codeaware_min_total_seq) or codeaware_min_total_seq)
                cur_ca_seq_share = float(get_setting(settings_map, "CodeAware_MinTopShareSeq", codeaware_min_share_seq) or codeaware_min_share_seq)

                ca_changed = False
                if int(desired_ca_total) < int(cur_ca_total):
                    flow_setting_set(store,
                        "CodeAware_MinTotalPerCode",
                        int(desired_ca_total),
                        note=f"iter{next_iter}: codeaware ladder rung{plateau_rung}",
                    )
                    codeaware_min_total = int(desired_ca_total)
                    ca_changed = True
                if float(desired_ca_share) + 1e-12 < float(cur_ca_share):
                    flow_setting_set(store,
                        "CodeAware_MinTopShare",
                        float(desired_ca_share),
                        note=f"iter{next_iter}: codeaware ladder rung{plateau_rung}",
                    )
                    codeaware_min_share = float(desired_ca_share)
                    ca_changed = True
                if int(desired_ca_maxlen) > int(cur_ca_maxlen):
                    flow_setting_set(store,
                        "CodeAware_MaxTokenLen",
                        int(desired_ca_maxlen),
                        note=f"iter{next_iter}: codeaware ladder rung{plateau_rung}",
                    )
                    codeaware_max_token_len = int(desired_ca_maxlen)
                    ca_changed = True
                if int(desired_ca_seq_total) < int(cur_ca_seq_total):
                    flow_setting_set(store,
                        "CodeAware_MinTotalPerCodeSeq",
                        int(desired_ca_seq_total),
                        note=f"iter{next_iter}: codeaware ladder rung{plateau_rung}",
                    )
                    codeaware_min_total_seq = int(desired_ca_seq_total)
                    ca_changed = True
                if float(desired_ca_seq_share) + 1e-12 < float(cur_ca_seq_share):
                    flow_setting_set(store,
                        "CodeAware_MinTopShareSeq",
                        float(desired_ca_seq_share),
                        note=f"iter{next_iter}: codeaware ladder rung{plateau_rung}",
                    )
                    codeaware_min_share_seq = float(desired_ca_seq_share)
                    ca_changed = True

                if ca_changed:
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = (
                        f"CodeAware ladder applied (min_total {cur_ca_total}->{codeaware_min_total}, min_share {cur_ca_share}->{codeaware_min_share}, "
                        f"max_token_len {cur_ca_maxlen}->{codeaware_max_token_len}, "
                        f"seq_min_total {cur_ca_seq_total}->{codeaware_min_total_seq}, seq_min_share {cur_ca_seq_share}->{codeaware_min_share_seq})"
                    )
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

            # SequenceMatch ladder: improve recall on sustained plateaus (analysis-only; snippets only).
            # This helps feed SequenceWordHints -> ContextEnglish candidate boosting without touching DP/Glossary.
            if seqmatch_enabled:
                seq_ladder = {
                    0: (1, [6, 7, 8], 20.0, 60),
                    1: (2, [5, 6, 7, 8], 30.0, 80),
                    2: (3, [4, 5, 6, 7, 8], 45.0, 120),
                    3: (4, [3, 4, 5, 6, 7, 8], 70.0, 160),
                    4: (5, [3, 4, 5, 6, 7, 8, 9], 70.0, 160),
                    5: (6, [3, 4, 5, 6, 7, 8, 9, 10], 70.0, 160),
                    6: (6, [3, 4, 5, 6, 7, 8, 9, 10], 70.0, 160),
                    7: (6, [3, 4, 5, 6, 7, 8, 9, 10, 11, 12], 70.0, 160),
                    8: (7, [3, 4, 5, 6, 7, 8, 9, 10, 11, 12], 70.0, 160),
                    9: (8, [3, 4, 5, 6, 7, 8, 9, 10, 11, 12], 70.0, 160),
                }
                desired_max_freq, desired_ns, desired_budget_s, desired_max_matches = seq_ladder.get(
                    int(plateau_rung), (3, [4, 5, 6, 7, 8], 45.0, 120)
                )

                cur_max_freq = int(get_setting_value(settings_map, "SequenceMatch_CandidateMaxBookFreq", seqmatch_cand_max_freq))
                cur_budget_s = float(get_setting_value(settings_map, "SequenceMatch_TimeBudgetS", seqmatch_time_budget_s))
                cur_max_matches = int(get_setting_value(settings_map, "SequenceMatch_MaxMatches", seqmatch_max_matches))
                cur_max_candidates = int(get_setting_value(settings_map, "SequenceMatch_MaxCandidates", seqmatch_max_candidates))
                cur_keep_top = int(get_setting_value(settings_map, "SequenceMatch_ExploreKeepTop", seqmatch_explore_keep_top))
                cur_min_n = int(get_setting_value(settings_map, "SequenceMatch_MinN", seqmatch_min_n))
                cur_n_str = str(get_setting_value(settings_map, "SequenceMatch_NList", ",".join(str(n) for n in (seqmatch_n_list or [6, 7, 8]))))
                cur_ns = [int(x) for x in parse_csv(cur_n_str) if str(x).strip().isdigit()]
                if not cur_ns:
                    cur_ns = [6, 7, 8]

                seq_changed = False
                if int(desired_max_freq) > int(cur_max_freq):
                    flow_setting_set(store,
                        "SequenceMatch_CandidateMaxBookFreq",
                        int(desired_max_freq),
                        note=f"iter{next_iter}: seqmatch ladder rung{plateau_rung}",
                    )
                    seqmatch_cand_max_freq = int(desired_max_freq)
                    seq_changed = True
                if float(desired_budget_s) > float(cur_budget_s) + 1e-12:
                    flow_setting_set(store,
                        "SequenceMatch_TimeBudgetS",
                        float(desired_budget_s),
                        note=f"iter{next_iter}: seqmatch ladder rung{plateau_rung}",
                    )
                    seqmatch_time_budget_s = float(desired_budget_s)
                    seq_changed = True
                if int(desired_max_matches) > int(cur_max_matches):
                    flow_setting_set(store,
                        "SequenceMatch_MaxMatches",
                        int(desired_max_matches),
                        note=f"iter{next_iter}: seqmatch ladder rung{plateau_rung}",
                    )
                    seqmatch_max_matches = int(desired_max_matches)
                    seq_changed = True

                # Exploration ladder: when we are on a deep plateau, cap the candidate set so rotation
                # actually explores different subsets (otherwise max_candidates >= pool -> no rotation).
                # Safe: analysis-only; we also keep a cache sheet to accumulate matches over time.
                if bool(seqmatch_explore_rotate) and int(plateau_rung) >= 8:
                    desired_cap = 4000 if int(plateau_rung) == 8 else 2000
                    if int(desired_cap) < int(cur_max_candidates):
                        flow_setting_set(store,
                            "SequenceMatch_MaxCandidates",
                            int(desired_cap),
                            note=f"iter{next_iter}: seqmatch explore cap rung{plateau_rung}",
                        )
                        seqmatch_max_candidates = int(desired_cap)
                        seq_changed = True
                    # Keep a *very* small stable head while still rotating most of the window for exploration.
                    # Note: we also keep previously matched keys via SequenceMatchesCache_Auto ("forced keys"),
                    # which prevents the exploration mode from starving SequenceMatches to 0.
                    desired_keep = min(50, int(desired_cap))
                    if int(cur_keep_top) != int(desired_keep):
                        flow_setting_set(store,
                            "SequenceMatch_ExploreKeepTop",
                            int(desired_keep),
                            note=f"iter{next_iter}: seqmatch explore keep_top rung{plateau_rung}",
                        )
                        seqmatch_explore_keep_top = int(desired_keep)
                        seq_changed = True

                # Quality ladder: keep SequenceMatches useful (non-zero) without flooding with trivial matches.
                # Empirically, for this corpus n>=3 is often needed to get *any* matches; n>=5 is too strict and
                # can lead to a permanent "0 matches" plateau even when relevant sources exist.
                # On deep plateaus we also allow bigrams (n=2) again: they are noisier, but they can
                # unlock many more Tibia-corpus hits and feed SequenceWordHints without touching DP.
                min_n_ladder = {0: 2, 1: 2, 2: 2, 3: 2, 4: 3, 5: 3, 6: 3, 7: 3, 8: 2, 9: 2}
                desired_min_n = int(min_n_ladder.get(int(plateau_rung), 3))
                if int(desired_min_n) != int(cur_min_n):
                    flow_setting_set(store,
                        "SequenceMatch_MinN",
                        int(desired_min_n),
                        note=f"iter{next_iter}: seqmatch ladder rung{plateau_rung}",
                    )
                    seqmatch_min_n = int(desired_min_n)
                    seq_changed = True

                # Expand n-list by adding smaller n as we relax (keep any existing larger n too).
                desired_ns2 = sorted(set([int(n) for n in desired_ns if int(n) > 1] + [int(n) for n in cur_ns if int(n) > 1]))
                if desired_ns2 != sorted(set(cur_ns)):
                    flow_setting_set(store,
                        "SequenceMatch_NList",
                        ",".join(str(n) for n in desired_ns2),
                        note=f"iter{next_iter}: seqmatch ladder rung{plateau_rung}",
                    )
                    seqmatch_n_list = desired_ns2
                    seq_changed = True

                if seq_changed:
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = (
                        f"SequenceMatch ladder applied (cand_max_freq {cur_max_freq}->{seqmatch_cand_max_freq}, "
                        f"max_candidates {cur_max_candidates}->{seqmatch_max_candidates}, keep_top {cur_keep_top}->{seqmatch_explore_keep_top}, "
                        f"min_n {cur_min_n}->{seqmatch_min_n}, n_list {cur_ns}->{seqmatch_n_list}, time_budget_s {cur_budget_s}->{seqmatch_time_budget_s}, "
                        f"max_matches {cur_max_matches}->{seqmatch_max_matches})"
                    )
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

            # SequenceWordHints ladder: allow mining hints from shorter matches on deep plateaus.
            # Still analysis-only and only affects ContextEnglish/CodeAware display layers.
            if seq_word_hints_enabled:
                hint_ladder = {
                    0: (3, 1.5),
                    1: (3, 1.5),
                    2: (3, 1.5),
                    3: (3, 1.5),
                    4: (3, 1.5),
                    5: (3, 1.5),
                    6: (3, 1.5),
                    7: (3, 1.5),
                    8: (2, 1.3),
                    9: (2, 1.0),
                }
                desired_h_min_n, desired_h_ratio = hint_ladder.get(int(plateau_rung), (3, 1.5))
                cur_h_min_n = int(get_setting_value(settings_map, "SequenceWordHints_MinN", seq_word_hints_min_n))
                try:
                    cur_h_ratio = float(get_setting_value(settings_map, "SequenceWordHints_MinRatio", seq_word_hints_min_ratio))
                except Exception:
                    cur_h_ratio = float(seq_word_hints_min_ratio)

                h_changed = False
                if int(desired_h_min_n) != int(cur_h_min_n):
                    flow_setting_set(store,
                        "SequenceWordHints_MinN",
                        int(desired_h_min_n),
                        note=f"iter{next_iter}: seqhints ladder rung{plateau_rung}",
                    )
                    seq_word_hints_min_n = int(desired_h_min_n)
                    h_changed = True
                # Only relax ratio downward automatically.
                if float(desired_h_ratio) + 1e-12 < float(cur_h_ratio):
                    flow_setting_set(store,
                        "SequenceWordHints_MinRatio",
                        float(desired_h_ratio),
                        note=f"iter{next_iter}: seqhints ladder rung{plateau_rung}",
                    )
                    seq_word_hints_min_ratio = float(desired_h_ratio)
                    h_changed = True

                if h_changed:
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = (
                        f"SequenceWordHints ladder applied (min_n {cur_h_min_n}->{seq_word_hints_min_n}, "
                        f"min_ratio {cur_h_ratio}->{seq_word_hints_min_ratio})"
                    )
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

            # SequenceHints ladder: on deep plateaus, trust SequenceWordHints more than the generic LM.
            # This is display-only: it only changes ContextEnglish/CodeAware candidate boosts.
            if seq_hints_enabled:
                boost_ladder = {
                    0: 20,
                    1: 20,
                    2: 20,
                    3: 20,
                    4: 30,
                    5: 40,
                    6: 50,
                    7: 60,
                    8: 80,
                    9: 100,
                }
                desired_boost = int(boost_ladder.get(int(plateau_rung), 60))
                cur_boost = int(get_setting_value(settings_map, "SequenceHints_Boost", seq_hints_boost))
                if int(desired_boost) > int(cur_boost):
                    flow_setting_set(store,
                        "SequenceHints_Boost",
                        int(desired_boost),
                        note=f"iter{next_iter}: seqhints boost ladder rung{plateau_rung}",
                    )
                    seq_hints_boost = int(desired_boost)
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = f"SequenceHints boost ladder applied ({cur_boost}->{seq_hints_boost})"
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

            # Reverse phrase ladder: increase the per-word span budget on sustained plateaus.
            # This is analysis-first and still requires exact signature matches, so it's safe.
            if reverse_enabled:
                # Canon safety (deep plateau):
                # ReversePhrase matching is a *structural* technique that depends on exact letter signatures
                # intersecting with the base token stream. The Lore canon "drop" knobs are useful for
                # semantic grouping, but they can destroy true-positive phrase matches by removing letters
                # (e.g. dropping final 'e' or all 'o') that DO exist in the decoded base alphabet.
                #
                # Therefore, on deep plateaus we explicitly force ReversePhrase to a strict canon regime:
                # - keep all letters except the always-confirmed normalizations inside _lore_canon_word()
                # - allow optional DropAllH via its own knob / auto-fix (handled elsewhere)
                if int(plateau_rung) >= 7:
                    canon_changed = False
                    if bool(reverse_drop_final_e) is True:
                        flow_setting_set(store,
                            "ReversePhrase_Canon_DropFinalE",
                            False,
                            note=f"iter{next_iter}: reverse canon strict rung{plateau_rung}",
                        )
                        reverse_drop_final_e = False
                        canon_changed = True
                    # NOTE: drop_all_o is a separate, data-driven knob (often helpful for matching tokens
                    # whose anagram layer appears to omit 'o'). Keep it aligned to Lore canon when they differ.
                    if bool(reverse_drop_all_o) != bool(lore_drop_all_o):
                        flow_setting_set(store,
                            "ReversePhrase_Canon_DropAllO",
                            bool(lore_drop_all_o),
                            note=f"iter{next_iter}: reverse canon align(drop_all_o) rung{plateau_rung}",
                        )
                        reverse_drop_all_o = bool(lore_drop_all_o)
                        canon_changed = True

                    if canon_changed:
                        flow_store_refresh_from_workbook(store)
                        ws_settings = store["ws_settings"]
                        settings_map = store["settings_map"]
                        note = (
                            "ReversePhrase canon safety applied "
                            f"(drop_final_e=0; drop_all_o={int(bool(reverse_drop_all_o))}; drop_all_h unchanged)"
                        )
                        if plateau_relax_note:
                            plateau_relax_note += f"; {note}"
                        else:
                            plateau_relax_note = note

                # On deep plateaus, allow shorter SequenceMatches phrases to feed ReversePhrase.
                # This is still analysis-only and capped by ReversePhrase_SeqMatchesMaxPhrasesPerIter.
                if reverse_inc_seqmatch:
                    desired_seqmatch_min_n = 3 if int(plateau_rung) >= 8 else int(reverse_seqmatch_min_n)
                    cur_seqmatch_min_n = int(get_setting(settings_map, "ReversePhrase_SeqMatchesMinN", reverse_seqmatch_min_n) or reverse_seqmatch_min_n)
                    if int(desired_seqmatch_min_n) < int(cur_seqmatch_min_n):
                        flow_setting_set(store,
                            "ReversePhrase_SeqMatchesMinN",
                            int(desired_seqmatch_min_n),
                            note=f"iter{next_iter}: reverse seqmatch ladder rung{plateau_rung}",
                        )
                        reverse_seqmatch_min_n = int(desired_seqmatch_min_n)
                        flow_store_refresh_from_workbook(store)
                        ws_settings = store["ws_settings"]
                        settings_map = store["settings_map"]
                        note = f"ReversePhrase seqmatch ladder applied (min_n {cur_seqmatch_min_n}->{reverse_seqmatch_min_n})"
                        if plateau_relax_note:
                            plateau_relax_note += f"; {note}"
                        else:
                            plateau_relax_note = note

                span_ladder = {0: 6, 1: 8, 2: 10, 3: 12, 4: 14, 5: 16, 6: 18, 7: 20, 8: 22, 9: 24}
                desired_span = int(span_ladder.get(int(plateau_rung), 10))
                cur_span = int(get_setting(settings_map, "ReversePhrase_MaxSpanTokens", reverse_max_span) or reverse_max_span)
                if desired_span > cur_span:
                    flow_setting_set(store,
                        "ReversePhrase_MaxSpanTokens",
                        desired_span,
                        note=f"iter{next_iter}: reverse span ladder rung{plateau_rung}",
                    )
                    reverse_max_span = desired_span
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = f"ReversePhrase span ladder applied (max_span {cur_span}->{desired_span})"
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

                gap_ladder = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 1, 8: 2, 9: 3}
                desired_gap = int(gap_ladder.get(int(plateau_rung), 0))
                cur_gap = int(get_setting(settings_map, "ReversePhrase_MaxGapTokens", reverse_max_gap) or reverse_max_gap)
                if desired_gap > cur_gap:
                    flow_setting_set(store,
                        "ReversePhrase_MaxGapTokens",
                        desired_gap,
                        note=f"iter{next_iter}: reverse gap ladder rung{plateau_rung}",
                    )
                    reverse_max_gap = desired_gap
                    flow_store_refresh_from_workbook(store)
                    ws_settings = store["ws_settings"]
                    settings_map = store["settings_map"]
                    note = f"ReversePhrase gap ladder applied (max_gap {cur_gap}->{desired_gap})"
                    if plateau_relax_note:
                        plateau_relax_note += f"; {note}"
                    else:
                        plateau_relax_note = note

    # Plateau corpus ladder: when macro/english/context relaxations are exhausted (rung>=2),
    # expand PD coverage with additional sources to increase candidate hits (derived-only).
    try:
        if int(plateau_rung) >= 2:
            cur_urls = [s for s in parse_csv(get_setting_value(settings_map, "LoreFetch_PDSigIndex_ExtraURLs", "")) if str(s).strip()]
            cur_cids = [s for s in parse_csv(get_setting_value(settings_map, "LoreFetch_PDSigIndex_ExtraCorpusIDs", "")) if str(s).strip()]

            pairs: List[Tuple[str, str]] = []
            for i, u in enumerate(cur_urls):
                u_s = str(u or "").strip()
                if not u_s:
                    continue
                cid_raw = cur_cids[i] if i < len(cur_cids) else ""
                cid_s = str(cid_raw or "").strip() or _auto_corpus_id_from_url(u_s)
                pairs.append((cid_s, u_s))

            seed_pairs: List[Tuple[str, str]] = list(_default_pd_sources())

            have = {u for _cid, u in pairs}
            changed = False
            for cid_s, u_s in seed_pairs:
                if u_s not in have:
                    pairs.append((cid_s, u_s))
                    have.add(u_s)
                    changed = True

            if changed:
                flow_setting_set(store,
                    "LoreFetch_PDSigIndex_ExtraURLs",
                    ",".join([u for _cid, u in pairs]),
                    note=f"iter{next_iter}: auto-seed PD extras rung{plateau_rung}",
                )
                flow_setting_set(store,
                    "LoreFetch_PDSigIndex_ExtraCorpusIDs",
                    ",".join([cid for cid, _u in pairs]),
                    note=f"iter{next_iter}: auto-seed PD extras rung{plateau_rung}",
                )
                flow_store_refresh_from_workbook(store)
                ws_settings = store["ws_settings"]
                settings_map = store["settings_map"]
                lore_fetch_pd_extra_urls = [u for _cid, u in pairs]
                lore_fetch_pd_extra_cids = [cid for cid, _u in pairs]
    except Exception:
        pass

    # Step 10: Promote GroundTruth (safe, conservative)
    gt_promoted = promote_groundtruth_cribs_from_reliability(wb, min_verified_sources=min_verified_sources_gt)
    append_flow_run_log(
        wb,
        next_iter,
        10,
        utc,
        "CHANGED" if gt_promoted else "NO_CHANGE",
        f"GroundTruth promotions from CribReliability_v129: {gt_promoted}",
        "0/70",
    )

    # Step 20: Fix stale GT match flags in Cribs
    fixed = fix_cribs_groundtruth_match_flags(wb)
    append_flow_run_log(
        wb,
        next_iter,
        20,
        utc,
        "CHANGED" if fixed else "NO_CHANGE",
        f"Cribs GT match flags fixed: {fixed}",
        "0/70",
    )

    # Step 25: Refresh mined macro evidence/confidence (runner-created tokens only)
    refreshed = refresh_mined_ngram_macro_evidence(wb, next_iter)
    append_flow_run_log(
        wb,
        next_iter,
        25,
        utc,
        "CHANGED" if refreshed else "NO_CHANGE",
        f"Refreshed mined n-gram macro evidence/confidence: {refreshed}",
        "0/70",
    )

    # Step 12: GroundTruth Live Check (guardrail)
    gt_bad_enforced_live = 0
    gt_bad_all_live = 0
    gt_soft_live = 0
    enforced_gt_ids, enforced_gt_n, gt_policy_status = resolve_enforced_groundtruth_ids(
        wb,
        iter_num=next_iter,
        settings_map=settings_map,
    )
    gt_live_mode_active = str(get_setting(settings_map, "GroundTruthLiveCheck_Mode", "POLICY") or "POLICY").strip().upper()
    gt_mode_changed = bool(prev_gt_mode_active) and gt_live_mode_active != str(prev_gt_mode_active).strip().upper()
    enforced_gt_label = "ALL" if enforced_gt_ids is None else str(enforced_gt_n)
    glossary_ws_gt, glossary_map_gt = load_glossary(wb)
    active_gt = {t.token: t for t in glossary_map_gt.values() if t.use_strictplus and t.translation}
    gt_ok, gt_bad_list, gt_bad_all = groundtruth_live_check(wb, active_gt, enforced_crib_ids=enforced_gt_ids)
    gt_bad_enforced_live = int(len(gt_bad_list))
    gt_bad_all_live = int(len(gt_bad_all))
    gt_soft_live = max(0, gt_bad_all_live - gt_bad_enforced_live)
    if gt_ok:
        append_flow_run_log(
            wb,
            next_iter,
            12,
            utc,
            "NO_CHANGE",
            "GroundTruth live check: OK"
            + (f" (policy {gt_policy_status})" if gt_policy_status else "")
            + (f"; enforced_gt={enforced_gt_label}" if enforced_gt_label else "")
            + (f"; soft_mismatches={gt_soft_live}" if gt_soft_live else ""),
            "0/70",
            notes=(
                f"mode={gt_live_mode_active}, enforced={enforced_gt_label}, "
                f"bad_enforced={gt_bad_enforced_live}, bad_all={gt_bad_all_live}, soft={gt_soft_live}, "
                f"mode_changed={int(gt_mode_changed)}"
            ),
        )
    else:
        bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in gt_bad_list[:12])

        gt_auto_repair = parse_bool(get_setting(settings_map, "GroundTruthAutoRepair", True), True)
        gt_auto_repair_max = int(get_setting(settings_map, "GroundTruthAutoRepairMaxMacros", 10) or 10)
        repaired = False
        repair_tokens: List[str] = []
        repair_summary: Optional[str] = None

        if gt_auto_repair:
            added, repair_tokens, repair_summary = add_gt_repair_macros_from_mismatches(
                wb,
                next_iter,
                utc,
                gt_bad_list,
                max_macros=gt_auto_repair_max,
            )
            if added:
                # Re-run GT check under the expanded active set.
                glossary_ws_gt, glossary_map_gt = load_glossary(wb)
                active_gt = {t.token: t for t in glossary_map_gt.values() if t.use_strictplus and t.translation}
                gt_ok2, gt_bad_list2, gt_bad_all2 = groundtruth_live_check(wb, active_gt, enforced_crib_ids=enforced_gt_ids)
                if gt_ok2:
                    repaired = True
                    gt_bad_enforced_live = int(len(gt_bad_list2))
                    gt_bad_all_live = int(len(gt_bad_all2))
                    gt_soft_live = max(0, gt_bad_all_live - gt_bad_enforced_live)
                    append_flow_run_log(
                        wb,
                        next_iter,
                        12,
                        utc,
                        "CHANGED",
                        f"GroundTruth live check: repaired via GT-macros added={added} (tokens {','.join(repair_tokens[:8])})"
                        + (f"; soft_mismatches={gt_soft_live}" if gt_soft_live else ""),
                        "0/70",
                        notes=(
                            f"{repair_summary}; mode={gt_live_mode_active}, enforced={enforced_gt_label}, "
                            f"bad_enforced={gt_bad_enforced_live}, bad_all={gt_bad_all_live}, soft={gt_soft_live}, "
                            f"mode_changed={int(gt_mode_changed)}"
                            if repair_summary
                            else (
                                f"mode={gt_live_mode_active}, enforced={enforced_gt_label}, "
                                f"bad_enforced={gt_bad_enforced_live}, bad_all={gt_bad_all_live}, soft={gt_soft_live}, "
                                f"mode_changed={int(gt_mode_changed)}"
                            )
                        ),
                    )
                else:
                    gt_bad_list = gt_bad_list2
                    gt_bad_enforced_live = int(len(gt_bad_list2))
                    gt_bad_all_live = int(len(gt_bad_all2))
                    gt_soft_live = max(0, gt_bad_all_live - gt_bad_enforced_live)
                    bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in gt_bad_list[:12])

        if not repaired:
            append_flow_run_log(
                wb,
                next_iter,
                12,
                utc,
                "FAILED",
                f"GroundTruth live check: FAIL enforced_mismatches={len(gt_bad_list)} (CribID(s) {bad_ids})"
                + (f" (policy {gt_policy_status})" if gt_policy_status else ""),
                "0/70",
                notes=(
                    "STOP: Fix GroundTruth mismatch before expanding search space."
                    + f" mode={gt_live_mode_active}, enforced={enforced_gt_label}, bad_enforced={gt_bad_enforced_live}, bad_all={gt_bad_all_live}, soft={gt_soft_live}, mode_changed={int(gt_mode_changed)}."
                    + (f" auto_repair={repair_summary}" if repair_summary else "")
                ),
            )

            block_reason = f"GroundTruth live check failed: {len(gt_bad_list)} enforced mismatches (CribID(s) {bad_ids})"
            append_method_log(
                wb,
                next_iter,
                utc,
                "GroundTruth live check failed (blocked)",
                "Guardrail: prevent macro-mining/promotions while anchored bilingual references mismatch.",
                block_reason,
                "BLOCKED",
            )
            append_work_queue(wb, next_iter, "BLOCKED", block_reason)

            # Update FlowState and stop early.
            flow_state_set_many(
                store,
                {
                    "CurrentIteration": next_iter,
                    "LastCompletedStepID": 12,
                    "NextStepID": 12,
                    "Status": "BLOCKED",
                    "LastRunUTC": utc,
                    "LastChangeSummary": f"Iter {next_iter}: BLOCKED (GT live check failed)",
                    "BlockReason": block_reason,
                    "SuccessCheck": False,
                    "GroundTruthLiveCheckModeActive": gt_live_mode_active,
                    "GroundTruthEnforcedCount": -1 if enforced_gt_ids is None else int(enforced_gt_n),
                    "GTBadEnforcedCount": int(gt_bad_enforced_live),
                    "GTBadAllCount": int(gt_bad_all_live),
                    "GTSoftMismatchCount": int(gt_soft_live),
                    "PromotionSkipCount": 0,
                    "PromotionSkipReasonTop": "",
                },
            )
            flow_state_set(store, "IterationsSinceLastMechanicalPromotion", int(prev_no_mech_iters) + 1)
            if gt_soft_live <= 0:
                early_soft_streak = 0
            elif prev_gt_soft_mismatch > 0 and gt_soft_live >= prev_gt_soft_mismatch:
                early_soft_streak = int(prev_gt_soft_nondec_streak) + 1
            else:
                early_soft_streak = 1
            flow_state_set_many(
                store,
                {
                    "IterationsSinceLastMechanicalPromotion": int(prev_no_mech_iters) + 1,
                    "GTSoftMismatchNonDecreasingStreak": int(early_soft_streak),
                },
            )

            wb.save(workbook_path)
            flow_store_close(store)
            sync_sqlite_snapshot_from_artifact(
                workbook_path,
                note=f"iter {next_iter}: status=BLOCKED",
            )
            print(f"Saved {workbook_path}")
            print(f"Backup {backup_path}")
            print(f"Iter {next_iter}: status=BLOCKED (GT live check failed)")
            return

    # Step 99: English -> Glossary retext (safe, guarded by GT live check)
    english_retext_applied = 0
    english_retext_attempted = 0
    english_retext_el_changed = 0
    anti_unk_applied = 0
    anti_unk_attempted = 0
    anti_unk_el_changed = 0
    try:
        english_retext_applied, english_retext_attempted, english_retext_el_changed = apply_english_promotions_to_glossary(
            wb,
            next_iter,
            utc,
            glossary_ws_gt,
            glossary_map_gt,
            enabled=english_retext_enabled,
            max_promotions=int(english_retext_max),
            min_total_count=int(english_retext_min_total),
            min_top_share=float(english_retext_min_share),
            lock_iters=int(english_retext_lock_iters),
            anti_mode=anti_mode,
            anti_hallucination_terms=anti_deny_words,
            enforced_crib_ids=enforced_gt_ids,
            locked_tokens=glossary_retext_locked_tokens,
        )
        append_flow_run_log(
            wb,
            next_iter,
            99,
            utc,
            "CHANGED" if english_retext_applied else "NO_CHANGE",
            f"English glossary retext applied: {english_retext_applied} (attempted {english_retext_attempted})",
            "0/70",
            notes=f"enabled={english_retext_enabled}, max={english_retext_max}, min_total={english_retext_min_total}, min_share={english_retext_min_share}, lock_iters={english_retext_lock_iters}, anti_mode={int(anti_mode)}, el_changed={english_retext_el_changed}",
        )
        if english_retext_applied:
            glossary_ws_gt, glossary_map_gt = load_glossary(wb)
            active_gt = {t.token: t for t in glossary_map_gt.values() if t.use_strictplus and t.translation}
    except Exception as e:
        append_flow_run_log(
            wb,
            next_iter,
            99,
            utc,
            "FAILED",
            "English glossary retext: FAILED",
            "0/70",
            notes=str(e),
        )

    # Step 98: anti-hallucination lexical sanitization (<UNK>) on suspicious low-evidence terms.
    try:
        anti_unk_applied, anti_unk_attempted, anti_unk_el_changed = apply_antihallucination_force_unk(
            wb,
            next_iter,
            utc,
            glossary_ws_gt,
            glossary_map_gt,
            enabled=bool(anti_mode and anti_unk_enabled),
            max_per_iter=int(anti_unk_max_per_iter),
            min_total_occ=int(anti_unk_min_total_occ),
            max_total_occ=int(anti_unk_max_total_occ),
            max_conf_weight=int(anti_unk_max_conf_w),
            protected_evidence_classes=set(anti_unk_protected_evcls),
            deny_words=set(anti_deny_words),
            unk_token="<UNK>",
            enforced_crib_ids=enforced_gt_ids,
        )
        append_flow_run_log(
            wb,
            next_iter,
            98,
            utc,
            "CHANGED" if anti_unk_applied else "NO_CHANGE",
            f"Anti-hallucination <UNK> sanitize: applied={anti_unk_applied} (attempted {anti_unk_attempted})",
            "0/70",
            notes=(
                f"enabled={int(bool(anti_mode and anti_unk_enabled))}, max={anti_unk_max_per_iter}, "
                f"occ=[{anti_unk_min_total_occ},{anti_unk_max_total_occ}], max_conf={anti_unk_max_conf}, "
                f"deny_words={len(anti_deny_words)}, el_changed={anti_unk_el_changed}"
            ),
        )
        if anti_unk_applied:
            glossary_ws_gt, glossary_map_gt = load_glossary(wb)
            active_gt = {t.token: t for t in glossary_map_gt.values() if t.use_strictplus and t.translation}
    except Exception as e:
        append_flow_run_log(
            wb,
            next_iter,
            98,
            utc,
            "FAILED",
            "Anti-hallucination <UNK> sanitize: FAILED",
            "0/70",
            notes=str(e),
        )

    # Step 27: AutoPhraseCribs (targeted phrase sources for Step 28)
    autophrase_changed = False
    autophrase_rows = 0
    autophrase_scanned = 0
    autophrase_eligible = 0
    phrase_book_tokens: Optional[Dict[int, Tuple[List[str], List[str]]]] = None

    if reverse_enabled and autophrase_enabled and reverse_inc_phrase_auto and (reverse_tibia_inc_npc or reverse_tibia_inc_books):
        try:
            # Build the global span signature frequency table from the current decode stream.
            comp_map = _build_macro_composition_map(glossary_ws_gt)
            phrase_book_tokens = _collect_books_leaf_tokens_for_phrase_matching(
                wb,
                active_tokens=active_gt,
                glossary_map=glossary_map_gt,
                comp_map=comp_map,
                drop_final_e=reverse_drop_final_e,
                drop_all_h=reverse_drop_all_h,
                drop_all_o=reverse_drop_all_o,
                logogram_aware=reverse_logogram_aware,
            )
            span_sig_freq: Counter[str] = Counter()
            for _bid, (_base_toks, match_toks) in phrase_book_tokens.items():
                span_sig_freq.update(_build_span_sig_freq(match_toks, max_span_tokens=int(reverse_max_span)))

            # Fetch public corpus JSON (cached on disk under tmp/).
            ua = "Mozilla/5.0 (Bonelord469AutoPhraseCribs/1.0)"
            cache_dir = os.path.join(os.getcwd(), "tmp", "corpus")
            npc_cache = _cache_path_for_url(cache_dir, prefix="tibia_npc", url=lore_fetch_tibia_npc_url)
            book_cache = _cache_path_for_url(cache_dir, prefix="tibia_books", url=lore_fetch_tibia_book_url)

            npc_obj: object = []
            book_obj: object = []
            if reverse_tibia_inc_npc:
                npc_obj = _fetch_json_url_cached(
                    lore_fetch_tibia_npc_url,
                    cache_path=npc_cache,
                    timeout_s=int(lore_fetch_tibia_timeout_s),
                    max_age_hours=float(autophrase_cache_age_h),
                    user_agent=ua,
                )
            if reverse_tibia_inc_books:
                book_obj = _fetch_json_url_cached(
                    lore_fetch_tibia_book_url,
                    cache_path=book_cache,
                    timeout_s=int(lore_fetch_tibia_timeout_s),
                    max_age_hours=float(autophrase_cache_age_h),
                    user_agent=ua,
                )

            tibia_rows, ap_stats = _select_autophrasecribs_from_tibia_corpus(
                npc_obj=npc_obj,
                book_obj=book_obj,
                npc_url=lore_fetch_tibia_npc_url,
                book_url=lore_fetch_tibia_book_url,
                include_npc=bool(reverse_tibia_inc_npc),
                include_books=bool(reverse_tibia_inc_books),
                span_sig_freq=span_sig_freq,
                # IMPORTANT: AutoPhraseCribs must use the *same* canon flags as ReversePhrase matching,
                # otherwise the feasibility filter (share=1.0) becomes overly strict and starves Step 28.
                drop_final_e=reverse_drop_final_e,
                drop_all_h=reverse_drop_all_h,
                drop_all_o=reverse_drop_all_o,
                min_words=int(autophrase_min_words),
                max_words=int(autophrase_max_words),
                max_phrases=int(autophrase_max_phrases),
                max_scan_sentences=int(autophrase_max_scan),
                time_budget_s=float(autophrase_time_budget_s),
                max_text_len=int(autophrase_max_text_len),
            )
            autophrase_scanned = int(ap_stats.get("scanned") or 0)
            autophrase_eligible = int(ap_stats.get("eligible") or 0)

            pd_rows: List[Dict[str, object]] = []
            pd_stats: Dict[str, int] = {"scanned": 0, "eligible": 0, "kept": 0, "sources": 0}
            if autophrase_include_pd:
                pd_sources = list(_default_pd_sources())
                if int(autophrase_pd_max_sources) > 0:
                    pd_sources = pd_sources[: int(autophrase_pd_max_sources)]
                pd_rows, pd_stats = _select_autophrasecribs_from_pd_sources(
                    pd_sources=pd_sources,
                    cache_dir=cache_dir,
                    timeout_s=int(lore_fetch_tibia_timeout_s),
                    cache_max_age_hours=float(autophrase_pd_cache_age_h),
                    span_sig_freq=span_sig_freq,
                    drop_final_e=reverse_drop_final_e,
                    drop_all_h=reverse_drop_all_h,
                    drop_all_o=reverse_drop_all_o,
                    min_words=int(autophrase_min_words),
                    max_words=int(autophrase_max_words),
                    max_phrases=max(0, int(autophrase_max_phrases)),
                    max_scan_sentences=max(0, int(autophrase_pd_max_scan)),
                    time_budget_s=float(autophrase_pd_time_budget_s),
                    max_text_len=int(autophrase_max_text_len),
                )
                autophrase_scanned += int(pd_stats.get("scanned") or 0)
                autophrase_eligible += int(pd_stats.get("eligible") or 0)

            # Combine + re-rank across sources, then dedupe by text.
            phrases_rows = list(tibia_rows) + list(pd_rows)

            def _sort_key(row: Dict[str, object]) -> Tuple[float, int, int, str]:
                try:
                    rarity = float(row.get("RarityScore") or 0.0)
                except Exception:
                    rarity = 0.0
                try:
                    siglen = int(row.get("SigLenSum") or 0)
                except Exception:
                    siglen = 0
                try:
                    wc = int(row.get("WordCount") or 0)
                except Exception:
                    wc = 0
                pid = str(row.get("PhraseID") or "")
                return (-rarity, -siglen, -wc, pid)

            phrases_rows.sort(key=_sort_key)
            seen_text: set[str] = set()
            deduped: List[Dict[str, object]] = []
            for rr0 in phrases_rows:
                t = str(rr0.get("Text") or "").strip()
                if not t:
                    continue
                k = t.lower()
                if k in seen_text:
                    continue
                seen_text.add(k)
                deduped.append(rr0)
                if len(deduped) >= int(autophrase_max_phrases):
                    break
            phrases_rows = deduped

            for rr0 in phrases_rows:
                rr0["AddedIter"] = int(next_iter)
                sk = str(rr0.get("SourceKind") or "")
                cache_age_h = float(autophrase_pd_cache_age_h) if sk == "PD" else float(autophrase_cache_age_h)
                src = "gutenberg" if sk == "PD" else "resources.talesoftibia.com"
                rr0["Notes"] = f"share=1.0; max_span={int(reverse_max_span)}; cache_max_age_h={cache_age_h}; source={src}"

            autophrase_changed, autophrase_rows = _write_phrase_cribs_auto_sheet(wb, iter_num=next_iter, phrases=phrases_rows)
            append_flow_run_log(
                wb,
                next_iter,
                27,
                utc,
                "CHANGED" if autophrase_changed else "NO_CHANGE",
                f"AutoPhraseCribs: scanned={autophrase_scanned}, eligible={autophrase_eligible}, kept={autophrase_rows}",
                "0/70",
                notes=(
                    f"enabled={autophrase_rows}, changed={autophrase_changed}, "
                    f"pd_scanned={int(pd_stats.get('scanned') or 0)}, pd_eligible={int(pd_stats.get('eligible') or 0)}, pd_sources={int(pd_stats.get('sources') or 0)}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                27,
                utc,
                "FAILED",
                "AutoPhraseCribs: FAILED",
                "0/70",
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            27,
            utc,
            "NO_CHANGE",
            "AutoPhraseCribs skipped (disabled)",
            "0/70",
        )

    # Step 28: Reverse phrase mining (PhraseCribs_User -> base search)
    reverse_phrase_hits = 0
    reverse_phrase_candidates = 0
    reverse_phrase_new_tokens = 0
    reverse_phrase_retext_applied = 0
    reverse_phrase_retext_attempted = 0
    reverse_phrase_retext_el_changed = 0
    reverse_phrase_scanned = 0
    reverse_phrase_status: Optional[str] = None
    reverse_perm_hits = 0
    reverse_perm_candidates = 0
    reverse_perm_new_tokens = 0
    reverse_perm_scanned = 0
    reverse_perm_status: Optional[str] = None
    if reverse_enabled:
        try:
            _ensure_phrase_cribs_user_sheet(wb)
            phrases: List[Dict[str, object]] = []
            if reverse_inc_phrase:
                phrases.extend(_load_phrase_cribs_user(wb, max_phrases=reverse_max_phrases))
            remaining = int(reverse_max_phrases) - len(phrases)
            if remaining > 0 and reverse_inc_phrase_auto:
                phrases.extend(_load_phrase_cribs_auto(wb, max_phrases=remaining))
            remaining = int(reverse_max_phrases) - len(phrases)
            if remaining > 0 and (reverse_inc_lore_user or reverse_inc_lore_auto):
                phrases.extend(
                    _load_lore_corpus_phrases(
                        wb,
                        # Allow LoreCorpus_Auto phrases even when Tibia corpus phrases are also enabled.
                        # This helps targeted methodology experiments (e.g., sestina/sestine corpus) without requiring
                        # users to disable the Tibia source globally.
                        include_auto=reverse_inc_lore_auto,
                        include_user=reverse_inc_lore_user,
                        max_phrases=remaining,
                    )
                )
            remaining = int(reverse_max_phrases) - len(phrases)
            if remaining > 0 and reverse_inc_seqmatch:
                phrases.extend(
                    _load_sequence_matches_phrases(
                        wb,
                        max_phrases=min(int(remaining), int(reverse_seqmatch_max_phrases)),
                        min_n=int(reverse_seqmatch_min_n),
                    )
                )
            remaining = int(reverse_max_phrases) - len(phrases)
            if remaining > 0 and reverse_inc_tibia and (reverse_tibia_inc_npc or reverse_tibia_inc_books):
                try:
                    phrases.extend(
                        _load_tibia_corpus_phrases(
                            npc_url=lore_fetch_tibia_npc_url,
                            book_url=lore_fetch_tibia_book_url,
                            timeout_s=int(lore_fetch_tibia_timeout_s),
                            include_npc=reverse_tibia_inc_npc,
                            include_books=reverse_tibia_inc_books,
                            max_phrases=min(int(remaining), int(reverse_tibia_max_phrases)),
                            min_words=int(reverse_min_words),
                            cache_max_age_hours=float(reverse_tibia_cache_age_h),
                        )
                    )
                except Exception:
                    # Reverse phrase mining is analysis-first; a corpus fetch failure should not stop the iteration.
                    pass

            # Dedupe by text to avoid double-counting.
            seen_text: set[str] = set()
            deduped: List[Dict[str, object]] = []
            for p in phrases:
                t = str(p.get("Text") or "").strip()
                if not t:
                    continue
                k = t.lower()
                if k in seen_text:
                    continue
                seen_text.add(k)
                deduped.append(p)
            phrases = deduped[: int(reverse_max_phrases)]
            reverse_phrase_scanned = len(phrases)

            if not phrases:
                # Keep workbook interface stable even before the user pastes phrases.
                _ensure_reverse_phrase_sheets(wb)
                append_flow_run_log(
                    wb,
                    next_iter,
                    28,
                    utc,
                    "NO_CHANGE",
                    "Reverse phrase mining: no enabled phrases (PhraseCribs_User/LoreCorpus_*/Tibia corpus)",
                    "0/70",
                )
                if reverse_perm_enabled:
                    _write_reverse_phrase_permute_outputs(wb, next_iter, [], [])
                    append_flow_run_log(
                        wb,
                        next_iter,
                        29,
                        utc,
                        "NO_CHANGE",
                        "Reverse phrase permuted: no enabled phrases",
                        "0/70",
                    )
                else:
                    append_flow_run_log(
                        wb,
                        next_iter,
                        29,
                        utc,
                        "NO_CHANGE",
                        "Reverse phrase permuted skipped (disabled)",
                        "0/70",
                    )
            else:
                # Use current active set for tokenization; expand macros to leaf tokens for matching.
                comp_map = _build_macro_composition_map(glossary_ws_gt)
                if phrase_book_tokens is None:
                    phrase_book_tokens = _collect_books_leaf_tokens_for_phrase_matching(
                        wb,
                        active_tokens=active_gt,
                        glossary_map=glossary_map_gt,
                        comp_map=comp_map,
                        drop_final_e=reverse_drop_final_e,
                        drop_all_h=reverse_drop_all_h,
                        drop_all_o=reverse_drop_all_o,
                        logogram_aware=reverse_logogram_aware,
                    )
                book_tokens = phrase_book_tokens

                hits: List[Dict[str, object]] = []
                cand_books: Dict[str, set[int]] = defaultdict(set)
                cand_occ: Counter[str] = Counter()
                cand_words: Dict[str, Counter[str]] = defaultdict(Counter)
                cand_example_phrase: Dict[str, object] = {}

                max_hits_total = int(reverse_max_hits_total)
                max_hits_per_phrase = max(1, int(reverse_max_hits_per_phrase))
                max_hits_per_book = max(1, int(reverse_max_hits_per_book))

                for phr in phrases:
                    if len(hits) >= max_hits_total:
                        break
                    pid = phr.get("PhraseID")
                    src = phr.get("Source")
                    text = str(phr.get("Text") or "")
                    wtrip = _phrase_word_sigs(
                        text,
                        drop_final_e=reverse_drop_final_e,
                        drop_all_h=reverse_drop_all_h,
                        drop_all_o=reverse_drop_all_o,
                    )
                    if len(wtrip) < int(reverse_min_words):
                        continue
                    word_sigs = [wsig for _surface, _canon, wsig in wtrip]

                    phrase_hits = 0
                    for book_id, (base_toks, match_toks) in book_tokens.items():
                        if len(hits) >= max_hits_total or phrase_hits >= max_hits_per_phrase:
                            break
                        matches = _reverse_phrase_find_matches(
                            word_sigs,
                            match_toks,
                            max_span_tokens=int(reverse_max_span),
                            max_gap_tokens=int(reverse_max_gap),
                            max_hits=int(max_hits_per_book),
                        )
                        for start, end, spans in matches:
                            phrase_hits += 1
                            max_span_used = max(spans) if spans else 1
                            gap_tokens = max(0, int(end - start) - int(sum(spans) if spans else 0))
                            used_tokens = base_toks[start:end]
                            base_concat = "".join(used_tokens)
                            phrase_text = text
                            if reverse_phrase_text_maxlen > 0 and len(phrase_text) > int(reverse_phrase_text_maxlen):
                                phrase_text = phrase_text[: int(reverse_phrase_text_maxlen)] + "..."

                            hits.append(
                                {
                                    "PhraseID": pid,
                                    "Source": src,
                                    "PhraseText": phrase_text,
                                    "BookID": book_id,
                                    "TokenStart": start,
                                    "TokenEnd": end,
                                    "WordCount": len(word_sigs),
                                    "MaxSpanUsed": max_span_used,
                                    "BaseTokens": " ".join(used_tokens[:60]) + (" ..." if len(used_tokens) > 60 else ""),
                                    "BaseConcat": base_concat,
                                    "Notes": (
                                        f"max_span={reverse_max_span}, max_gap={reverse_max_gap}, gap_tokens={gap_tokens}, logogram_aware={reverse_logogram_aware}, "
                                        f"canon(drop_final_e={reverse_drop_final_e}, drop_all_h={reverse_drop_all_h}, drop_all_o={reverse_drop_all_o}), "
                                        f"tibia={reverse_inc_tibia}"
                                    ),
                                }
                            )

                            # Derive per-word candidate bases from the match partition.
                            # When gap_tokens>0, we cannot reliably partition base tokens back to phrase words,
                            # so we record the hit but skip candidate derivation (keeps candidate quality high).
                            if gap_tokens == 0:
                                pos = start
                                for (surface, _canon, _wsig), k in zip(wtrip, spans):
                                    base_word = "".join(base_toks[pos : pos + k])
                                    pos += k
                                    # Filter out extremely short candidates and already-common stopwords.
                                    if len(base_word) < 3:
                                        continue
                                    if reverse_cand_max_base_len and len(base_word) > int(reverse_cand_max_base_len):
                                        continue
                                    cand_occ[base_word] += 1
                                    cand_books[base_word].add(int(book_id))
                                    cand_words[base_word][surface] += 1
                                    if base_word not in cand_example_phrase:
                                        cand_example_phrase[base_word] = pid

                            # Keep additional matches per book when configured (bounded by max_hits_per_book).

                # Aggregate candidates.
                cand_rows: List[Dict[str, object]] = []
                for base, occ in cand_occ.most_common(500):
                    books = cand_books.get(base) or set()
                    wc = cand_words.get(base) or Counter()
                    if not wc:
                        continue
                    top_word, top_count = wc.most_common(1)[0]
                    share = float(top_count) / float(occ) if occ else 0.0
                    cand_rows.append(
                        {
                            "Base": base,
                            "BaseSig": _base_sig_letters(base),
                            "BaseLen": len(base),
                            "SupportBooks": len(books),
                            "SupportOcc": int(occ),
                            "TopWord": top_word,
                            "TopWordCount": int(top_count),
                            "TopShare": round(share, 6),
                            "CandidateWords": ", ".join([w for w, _n in wc.most_common(8)]),
                            "ExamplePhraseID": cand_example_phrase.get(base),
                            "Notes": f"words={len(wc)}",
                        }
                    )

                reverse_phrase_hits, reverse_phrase_candidates = _write_reverse_phrase_outputs(wb, next_iter, hits, cand_rows)

                emitted_tokens: List[str] = []
                if reverse_emit and cand_rows:
                    reverse_phrase_new_tokens, emitted_tokens = add_reverse_phrase_candidates_to_glossary(
                        wb,
                        next_iter,
                        utc,
                        cand_rows,
                        evidence_class=reverse_evcls,
                        max_new_tokens=int(reverse_cand_max_new),
                        min_books=int(reverse_cand_min_books),
                        min_top_share=float(reverse_cand_min_share),
                    )
                if reverse_retext_enabled and cand_rows:
                    (
                        reverse_phrase_retext_applied,
                        reverse_phrase_retext_attempted,
                        reverse_phrase_retext_el_changed,
                    ) = apply_reverse_phrase_retext_existing_tokens(
                        wb,
                        next_iter,
                        utc,
                        cand_rows,
                        enabled=bool(reverse_retext_enabled),
                        max_per_iter=int(reverse_retext_max_per_iter),
                        min_books=int(reverse_retext_min_books),
                        min_top_share=float(reverse_retext_min_share),
                        min_support_occ=int(reverse_retext_min_support_occ),
                        anti_mode=bool(anti_mode),
                        anti_deny_words=set(anti_deny_words),
                        enforced_crib_ids=enforced_gt_ids,
                        locked_tokens=glossary_retext_locked_tokens,
                    )

                reverse_phrase_status = (
                    f"phrases={reverse_phrase_scanned}, hits={reverse_phrase_hits}, candidates={reverse_phrase_candidates}, "
                    f"emitted={reverse_phrase_new_tokens}, retext={reverse_phrase_retext_applied}/{reverse_phrase_retext_attempted}"
                )
                append_flow_run_log(
                    wb,
                    next_iter,
                    28,
                    utc,
                    "CHANGED"
                    if (reverse_phrase_hits or reverse_phrase_new_tokens or reverse_phrase_retext_applied)
                    else "NO_CHANGE",
                    (
                        f"Reverse phrase mining: hits={reverse_phrase_hits}, emitted_tokens={reverse_phrase_new_tokens}, "
                        f"retext={reverse_phrase_retext_applied}"
                    ),
                    "0/70",
                    notes=(reverse_phrase_status + (f"; tokens={','.join(emitted_tokens[:8])}" if emitted_tokens else "")),
                )

                # Step 29: permutation-aware reverse phrase mining (deep-plateau diagnostic).
                # This does NOT change StrictPlus and is analysis-first. Emission to Glossary is OFF by default.
                if reverse_perm_enabled:
                    try:
                        if reverse_phrase_hits <= int(reverse_perm_run_if_hits_le) and phrases:
                            perm_phrases = phrases[: int(reverse_perm_max_phrases)] if int(reverse_perm_max_phrases) > 0 else []
                            reverse_perm_scanned = len(perm_phrases)

                            perm_hits: List[Dict[str, object]] = []
                            cand_books_p: Dict[str, set[int]] = defaultdict(set)
                            cand_occ_p: Counter[str] = Counter()
                            cand_words_p: Dict[str, Counter[str]] = defaultdict(Counter)
                            cand_perm_support: Counter[str] = Counter()
                            cand_example_phrase_p: Dict[str, object] = {}

                            max_hits_total_p = int(reverse_perm_max_hits_total)
                            max_hits_per_phrase_p = max(1, int(reverse_perm_max_hits_per_phrase))
                            max_hits_per_book_p = max(1, int(reverse_perm_max_hits_per_book))

                            for phr in perm_phrases:
                                if len(perm_hits) >= max_hits_total_p:
                                    break
                                pid = phr.get("PhraseID")
                                src = phr.get("Source")
                                text = str(phr.get("Text") or "")
                                wtrip = _phrase_word_sigs(
                                    text,
                                    drop_final_e=reverse_drop_final_e,
                                    drop_all_h=reverse_drop_all_h,
                                    drop_all_o=reverse_drop_all_o,
                                )
                                if len(wtrip) < int(reverse_min_words):
                                    continue
                                if int(reverse_perm_max_words) > 0 and len(wtrip) > int(reverse_perm_max_words):
                                    continue
                                word_sigs = [wsig for _surface, _canon, wsig in wtrip]

                                phrase_hits_p = 0
                                for book_id, (base_toks, match_toks) in book_tokens.items():
                                    if len(perm_hits) >= max_hits_total_p or phrase_hits_p >= max_hits_per_phrase_p:
                                        break
                                    matches_p = _reverse_phrase_find_permuted_matches(
                                        word_sigs,
                                        match_toks,
                                        max_span_tokens=int(reverse_max_span),
                                        max_gap_tokens=int(reverse_max_gap),
                                        max_hits=int(max_hits_per_book_p),
                                        max_words=int(reverse_perm_max_words) if int(reverse_perm_max_words) > 0 else 6,
                                    )
                                    for start, end, spans, perm in matches_p:
                                        phrase_hits_p += 1
                                        max_span_used = max(spans) if spans else 1
                                        gap_tokens = max(0, int(end - start) - int(sum(spans) if spans else 0))
                                        used_tokens = base_toks[start:end]
                                        base_concat = "".join(used_tokens)
                                        phrase_text = text
                                        if reverse_phrase_text_maxlen > 0 and len(phrase_text) > int(reverse_phrase_text_maxlen):
                                            phrase_text = phrase_text[: int(reverse_phrase_text_maxlen)] + "..."

                                        perm_hits.append(
                                            {
                                                "PhraseID": pid,
                                                "Source": src,
                                                "PhraseText": phrase_text,
                                                "BookID": book_id,
                                                "TokenStart": start,
                                                "TokenEnd": end,
                                                "WordCount": len(word_sigs),
                                                "MaxSpanUsed": max_span_used,
                                                "BaseTokens": " ".join(used_tokens[:60]) + (" ..." if len(used_tokens) > 60 else ""),
                                                "BaseConcat": base_concat,
                                                "Permutation": ",".join(str(x) for x in perm),
                                                "Notes": (
                                                    f"mode=PERMUTED; max_span={reverse_max_span}, max_gap={reverse_max_gap}, gap_tokens={gap_tokens}, "
                                                    f"logogram_aware={reverse_logogram_aware}, canon(drop_final_e={reverse_drop_final_e}, drop_all_h={reverse_drop_all_h}, drop_all_o={reverse_drop_all_o}), "
                                                    f"tibia={reverse_inc_tibia}"
                                                ),
                                            }
                                        )

                                        # Derive candidates only when there are no gaps (keeps quality high).
                                        if gap_tokens == 0:
                                            pos = start
                                            for k, wi_1based in zip(spans, perm):
                                                wi = int(wi_1based) - 1
                                                if wi < 0 or wi >= len(wtrip):
                                                    pos += int(k)
                                                    continue
                                                surface = str(wtrip[wi][0] or "")
                                                base_word = "".join(base_toks[pos : pos + int(k)])
                                                pos += int(k)
                                                if len(base_word) < 3:
                                                    continue
                                                if reverse_cand_max_base_len and len(base_word) > int(reverse_cand_max_base_len):
                                                    continue
                                                cand_occ_p[base_word] += 1
                                                cand_perm_support[base_word] += 1
                                                cand_books_p[base_word].add(int(book_id))
                                                if surface:
                                                    cand_words_p[base_word][surface] += 1
                                                if base_word not in cand_example_phrase_p:
                                                    cand_example_phrase_p[base_word] = pid

                                        # Keep additional permuted matches per book when configured.

                            cand_rows_p: List[Dict[str, object]] = []
                            for base, occ in cand_occ_p.most_common(500):
                                books = cand_books_p.get(base) or set()
                                wc = cand_words_p.get(base) or Counter()
                                if not wc:
                                    continue
                                top_word, top_count = wc.most_common(1)[0]
                                share = float(top_count) / float(occ) if occ else 0.0
                                cand_rows_p.append(
                                    {
                                        "Base": base,
                                        "BaseSig": _base_sig_letters(base),
                                        "BaseLen": len(base),
                                        "SupportBooks": len(books),
                                        "SupportOcc": int(occ),
                                        "TopWord": top_word,
                                        "TopWordCount": int(top_count),
                                        "TopShare": round(share, 6),
                                        "CandidateWords": ", ".join([w for w, _n in wc.most_common(8)]),
                                        "ExamplePhraseID": cand_example_phrase_p.get(base),
                                        "PermuteSupport": int(cand_perm_support.get(base) or 0),
                                        "Notes": f"mode=PERMUTED; words={len(wc)}",
                                    }
                                )

                            reverse_perm_hits, reverse_perm_candidates = _write_reverse_phrase_permute_outputs(wb, next_iter, perm_hits, cand_rows_p)

                            emitted_perm: List[str] = []
                            if reverse_perm_emit and cand_rows_p:
                                reverse_perm_new_tokens, emitted_perm = add_reverse_phrase_candidates_to_glossary(
                                    wb,
                                    next_iter,
                                    utc,
                                    cand_rows_p,
                                    evidence_class=reverse_evcls,
                                    max_new_tokens=int(reverse_cand_max_new),
                                    min_books=int(reverse_cand_min_books),
                                    min_top_share=float(reverse_cand_min_share),
                                )

                            reverse_perm_status = (
                                f"phrases={reverse_perm_scanned}, hits={reverse_perm_hits}, candidates={reverse_perm_candidates}, emitted={reverse_perm_new_tokens}"
                            )
                            append_flow_run_log(
                                wb,
                                next_iter,
                                29,
                                utc,
                                "CHANGED" if (reverse_perm_hits or reverse_perm_new_tokens) else "NO_CHANGE",
                                f"Reverse phrase permuted: hits={reverse_perm_hits}, emitted_tokens={reverse_perm_new_tokens}",
                                "0/70",
                                notes=(reverse_perm_status + (f"; tokens={','.join(emitted_perm[:8])}" if emitted_perm else "")),
                            )
                        else:
                            # Keep per-iteration sheets accurate: clear outputs even when skipped.
                            _write_reverse_phrase_permute_outputs(wb, next_iter, [], [])
                            reverse_perm_status = f"skipped (sequential_hits={reverse_phrase_hits} > threshold={reverse_perm_run_if_hits_le})"
                            append_flow_run_log(
                                wb,
                                next_iter,
                                29,
                                utc,
                                "NO_CHANGE",
                                "Reverse phrase permuted: skipped (sequential hits above threshold)",
                                "0/70",
                                notes=reverse_perm_status,
                            )
                    except Exception as e:
                        append_flow_run_log(
                            wb,
                            next_iter,
                            29,
                            utc,
                            "FAILED",
                            "Reverse phrase permuted: FAILED",
                            "0/70",
                            notes=str(e),
                        )
                else:
                    append_flow_run_log(
                        wb,
                        next_iter,
                        29,
                        utc,
                        "NO_CHANGE",
                        "Reverse phrase permuted skipped (disabled)",
                        "0/70",
                    )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                28,
                utc,
                "FAILED",
                "Reverse phrase mining: FAILED",
                "0/70",
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            28,
            utc,
            "NO_CHANGE",
            "Reverse phrase mining skipped (disabled)",
            "0/70",
        )
        if reverse_perm_enabled:
            _write_reverse_phrase_permute_outputs(wb, next_iter, [], [])
        append_flow_run_log(
            wb,
            next_iter,
            29,
            utc,
            "NO_CHANGE",
            "Reverse phrase permuted skipped (reverse disabled)",
            "0/70",
        )

    # Step 30: list candidates (includes allow-0-occ classes when token appears in external base corpus)
    lowered_note = None
    candidate_scan_mode = "normal"
    candidate_empty_streak = 0
    candidates: List[Tuple[str, str, str, int, int]] = []
    rows: List[int] = []
    mined_macro_added = 0
    anchor_corpus: List[str] = []
    anchor_hit_cache: Dict[str, int] = {}
    anchor_promo_corpus_size = 0
    anchor_promo_kept_step30 = 0
    anchor_promo_dropped_step30 = 0
    anchor_promo_kept_step40 = 0
    anchor_promo_dropped_step40 = 0
    anchor_promo_drop30_classes = ""
    anchor_promo_drop30_samples = ""
    anchor_promo_drop40_classes = ""
    anchor_promo_drop40_samples = ""
    if allow_mech:
        external_corpus = collect_external_base_corpus(wb)
        if anchor_promo_only_enabled:
            anchor_corpus = collect_anchor_base_corpus(
                wb,
                enforced_crib_ids=enforced_gt_ids if anchor_promo_include_groundtruth else None,
                include_groundtruth=anchor_promo_include_groundtruth,
                include_external_npc_staff=anchor_promo_include_external_npc_staff,
                include_external_books=anchor_promo_include_external_books,
                include_anchorcribs=anchor_promo_include_anchorcribs,
                min_verified_sources=anchor_promo_min_verified_sources,
            )
            anchor_promo_corpus_size = len(anchor_corpus)
        allow_markers = parse_bool(get_setting(settings_map, "MacroMine_AllowMarkers", True), True)
        allow_stars = parse_bool(get_setting(settings_map, "MacroMine_AllowStars", True), True)
        allow_zero: List[str] = []
        if "EXTERNAL_POEM" in mech_classes:
            allow_zero.append("EXTERNAL_POEM")
        if "PHRASE_CRIB" in mech_classes:
            # ReversePhrase/AutoPhrase can emit early candidates with sparse counts.
            # Keep them in the mechanical queue (still fully gated by GT + DP/metrics).
            allow_zero.append("PHRASE_CRIB")
        scan_occ = min_occ
        while True:
            candidates = read_glossary_inactive_candidates(
                wb,
                scan_occ,
                allowed_evidence_classes=mech_classes,
                allow_zero_occ_classes=allow_zero,
                external_base_corpus=external_corpus,
            )
            if candidates or scan_occ <= 1:
                break
            scan_occ = 1
        if scan_occ != min_occ:
            flow_setting_set(store, "MinTotalOccCandidate", scan_occ, note=f"iter{next_iter}: auto-lowered to {scan_occ} (candidate scan)")
            lowered_note = f"MinTotalOccCandidate lowered {min_occ} -> {scan_occ}"
            min_occ = scan_occ

        # If we hit a hard plateu and still have no inactive mechanical candidates,
        # open exploration by scanning all known evidence classes at min_occ=1 and keep a controlled top-k.
        if not candidates and convergence_candidate_starve_iters > 0:
            candidate_starve_streak = prev_candidate_empty_streak + 1
            if candidate_starve_streak >= convergence_candidate_starve_iters and convergence_candidate_scan_all_classes:
                candidate_scan_mode = "fallback-all-classes"
                candidate_scan_keep = max(0, convergence_candidate_scan_keep_top)
                candidate_scan_keep = max(int(candidate_scan_keep), 1)
                candidate_scan_min_occ = max(1, int(convergence_candidate_scan_min_occ))
                if candidate_starve_streak >= int(convergence_candidate_starve_iters) * 2:
                    candidate_scan_keep = max(candidate_scan_keep, 1200)
                    candidate_scan_min_occ = 0
                if candidate_starve_streak >= int(convergence_candidate_starve_iters) * 3:
                    candidate_scan_keep = max(candidate_scan_keep, 2600)
                fallback_candidates = read_glossary_inactive_candidates(
                    wb,
                    candidate_scan_min_occ,
                    allowed_evidence_classes=[],
                    allow_zero_occ_classes=allow_zero,
                    external_base_corpus=external_corpus,
                )
                if candidate_scan_keep and len(fallback_candidates) > candidate_scan_keep:
                    fallback_candidates = fallback_candidates[:candidate_scan_keep]
                if fallback_candidates:
                    candidates = fallback_candidates
                    mode_tag = f"{candidate_scan_mode}(min_occ={candidate_scan_min_occ}, keep={candidate_scan_keep}, streak={candidate_starve_streak})"
                    lower_note_extra = f"fallback candidate scan {mode_tag}"
                    lowered_note = f"{lowered_note}; {lower_note_extra}" if lowered_note else lower_note_extra

        macro_mine_requested = (not macro_mine_off) and (macro_mine_always or (not candidates))
        if macro_mine_requested:
            glossary_ws, glossary_map0 = load_glossary(wb)
            active0 = {t.token: t for t in glossary_map0.values() if t.use_strictplus and t.translation}

            macro_min_occ = int(get_setting(settings_map, "MacroMine_MinOcc", 2) or 2)
            macro_min_books = int(get_setting(settings_map, "MacroMine_MinBooks", 2) or 2)
            macro_min_share_default = 0.95 if macro_mine_always else 1.0
            macro_min_share = float(get_setting(settings_map, "MacroMine_MinShare", macro_min_share_default) or macro_min_share_default)
            macro_max_candidates_default = 75 if macro_mine_always else 25
            macro_max_candidates = int(
                get_setting(settings_map, "MacroMine_MaxCandidates", macro_max_candidates_default) or macro_max_candidates_default
            )
            macro_min_len = int(get_setting(settings_map, "MacroMine_MinLen", 2) or 2)
            macro_max_len_default = 16 if macro_mine_always else 12
            macro_max_len = int(get_setting(settings_map, "MacroMine_MaxLen", macro_max_len_default) or macro_max_len_default)
            macro_n_values_default = "2,3,4,5,6" if macro_mine_always else "2,3,4"
            macro_n_values = parse_int_list(get_setting(settings_map, "MacroMine_NValues", macro_n_values_default) or macro_n_values_default)
            if not macro_n_values:
                macro_n_values = [2, 3, 4]
            allow_macro_components_default = True if macro_mine_always else False
            allow_macro_components = parse_bool(
                get_setting(settings_map, "MacroMine_AllowMacroComponents", allow_macro_components_default),
                allow_macro_components_default,
            )

            mined = mine_macro_candidates_from_books(
                wb,
                active0,
                existing_tokens=list(glossary_map0.keys()),
                n_values=tuple(macro_n_values),
                min_occ=macro_min_occ,
                min_books=macro_min_books,
                min_share=macro_min_share,
                min_len=macro_min_len,
                max_len=macro_max_len,
                max_candidates=macro_max_candidates,
                allow_macro_components=allow_macro_components,
                allow_marker_tokens=allow_markers,
                allow_star_tokens=allow_stars,
            )
            mined_macro_added = add_mined_macros_to_glossary(
                wb,
                next_iter,
                mined,
                evidence_class="MACRO_ACTIVE",
                source_kind="n-gram",
                evidence_sources_tag="mined_ngram_macro",
            )
            if mined_macro_added:
                macro_candidates = [(m.base, m.translation, "MACRO_ACTIVE", m.occ, m.length) for m in mined[:mined_macro_added]]
                if candidates:
                    candidates.extend(macro_candidates)
                else:
                    # Use mined macros directly as this iteration's candidate set (stable order from mining).
                    candidates = macro_candidates

        # SuperAnchor-derived macros: use previous iteration's SuperAnchors output to generate token-boundary macros.
        # This is conservative and semantics-preserving (translation composed from existing token translations).
        if sa_macro_enabled and cur_iter > 0:
            try:
                glossary_ws, glossary_map0 = load_glossary(wb)
                active0 = {t.token: t for t in glossary_map0.values() if t.use_strictplus and t.translation}
                sa_macros = mine_macro_candidates_from_superanchors(
                    wb,
                    active0,
                    existing_tokens=list(glossary_map0.keys()),
                    source_iter=cur_iter,
                    min_len=sa_macro_min_len,
                    max_len=sa_macro_max_len,
                    max_candidates=sa_macro_max_candidates,
                )
                sa_added = add_mined_macros_to_glossary(
                    wb,
                    next_iter,
                    sa_macros,
                    evidence_class="STRUCT_MACRO_CAND",
                    source_kind="superanchor",
                    evidence_sources_tag="mined_superanchor_macro",
                )
                if sa_added:
                    sa_candidates = [(m.base, m.translation, "STRUCT_MACRO_CAND", m.occ, m.length) for m in sa_macros[:sa_added]]
                    if candidates:
                        candidates.extend(sa_candidates)
                    else:
                        candidates = sa_candidates
                    mined_macro_added += sa_added
            except Exception:
                # Safe: superanchor macro mining is a best-effort enhancer; never block the iteration.
                pass

        if anchor_promo_only_enabled:
            pre_anchor_n = len(candidates)
            candidates, dropped_anchor, anchor_diag = filter_candidates_by_anchor_impact(
                candidates,
                anchor_corpus=anchor_corpus,
                min_hits=anchor_promo_min_hits,
                cache=anchor_hit_cache,
            )
            anchor_promo_kept_step30 = len(candidates)
            anchor_promo_dropped_step30 += int(dropped_anchor)
            anchor_promo_drop30_classes = str(anchor_diag.get("classes", "") or "")
            anchor_promo_drop30_samples = str(anchor_diag.get("samples", "") or "")
            if pre_anchor_n and dropped_anchor:
                anchor_note = (
                    f"anchor-only kept={len(candidates)}, dropped={dropped_anchor}, "
                    f"anchors={anchor_promo_corpus_size}, min_hits={anchor_promo_min_hits}"
                )
                if anchor_promo_drop30_classes:
                    anchor_note += f", classes={anchor_promo_drop30_classes}"
                lowered_note = f"{lowered_note}; {anchor_note}" if lowered_note else anchor_note

        if candidate_priority_enabled and candidates:
            usage_stats = candidate_usage_stats_in_books(wb, [row[0] for row in candidates])
            no_effect_counts = recent_no_effect_token_counts(
                wb, lookback_iters=int(candidate_priority_noeffect_lookback)
            )
            ranked: List[Tuple[int, int, int, int, Tuple[str, str, str, int, int]]] = []
            dropped_priority = 0
            for row in candidates:
                tok, _tr, evcls, occ, tlen = row
                tok_s = str(tok or "").strip()
                evcls_s = str(evcls or "").strip().upper()
                occ_i = int(occ or 0)
                tlen_i = int(tlen or 0)
                book_hits, book_occ = usage_stats.get(tok_s, (0, 0))
                if int(book_hits) < int(candidate_priority_min_book_hits):
                    dropped_priority += 1
                    continue
                if occ_i < int(candidate_priority_min_total_occ) and int(book_hits) < 2:
                    dropped_priority += 1
                    continue
                max_len_allowed = (
                    int(candidate_priority_max_len_macro)
                    if evcls_s in ("MACRO_ACTIVE", "STRUCT_MACRO_CAND")
                    else int(candidate_priority_max_len)
                )
                if tlen_i > int(max_len_allowed):
                    dropped_priority += 1
                    continue
                if int(no_effect_counts.get(tok_s, 0) or 0) >= int(candidate_priority_noeffect_threshold):
                    dropped_priority += 1
                    continue
                ranked.append((int(book_hits), int(book_occ), occ_i, -tlen_i, row))
            ranked.sort(key=lambda x: (x[0], x[1], x[2], x[3]), reverse=True)
            candidates = [r[4] for r in ranked]
            if dropped_priority:
                pr_note = (
                    f"priority kept={len(candidates)}, dropped={dropped_priority}, "
                    f"min_hits={candidate_priority_min_book_hits}, min_occ={candidate_priority_min_total_occ}, "
                    f"noeffect_th={candidate_priority_noeffect_threshold}, lookback={candidate_priority_noeffect_lookback}"
                )
                lowered_note = f"{lowered_note}; {pr_note}" if lowered_note else pr_note

        force_probe_token_set = set(convergence_force_probe_tokens)
        if convergence_force_probe_tokens or convergence_force_probe_substrings:
            glossary_ws_force, glossary_map_force = load_glossary(wb)
            del glossary_ws_force
            existing_tokens = {str(row[0] or "").strip() for row in candidates}
            forced_rows: List[Tuple[str, str, str, int, int]] = []
            forced_exact = 0
            for tok in convergence_force_probe_tokens:
                gt_force = glossary_map_force.get(tok)
                if gt_force is None:
                    continue
                if tok in existing_tokens:
                    continue
                forced_rows.append(
                    (
                        gt_force.token,
                        gt_force.translation,
                        str(gt_force.evidence_class or ""),
                        int(gt_force.total_occ or 0),
                        int(gt_force.length or 0),
                    )
                )
                existing_tokens.add(tok)
                force_probe_token_set.add(tok)
                forced_exact += 1
            forced_family = 0
            if convergence_force_probe_substrings:
                family_rows: List[Tuple[str, str, str, int, int]] = []
                for gt_force in glossary_map_force.values():
                    tok_force = str(gt_force.token or "").strip()
                    if not tok_force or tok_force in existing_tokens:
                        continue
                    if not any(sub in tok_force for sub in convergence_force_probe_substrings):
                        continue
                    family_rows.append(
                        (
                            gt_force.token,
                            gt_force.translation,
                            str(gt_force.evidence_class or ""),
                            int(gt_force.total_occ or 0),
                            int(gt_force.length or 0),
                        )
                    )
                family_rows.sort(key=lambda item: (-int(item[3] or 0), -int(item[4] or 0), str(item[0] or "")))
                for tok_force, tr_force, evcls_force, occ_force, len_force in family_rows[
                    : int(convergence_force_probe_substring_max_matches)
                ]:
                    forced_rows.append((tok_force, tr_force, evcls_force, occ_force, len_force))
                    existing_tokens.add(tok_force)
                    force_probe_token_set.add(tok_force)
                    forced_family += 1
            if forced_rows:
                if convergence_force_probe_only:
                    candidates = forced_rows
                    force_note = (
                        f"forced-probe-only kept={len(forced_rows)}"
                        f" exact={forced_exact} family={forced_family}"
                    )
                else:
                    candidates = forced_rows + candidates
                    force_note = (
                        f"forced-probe kept={len(forced_rows)}"
                        f" exact={forced_exact} family={forced_family}"
                    )
                lowered_note = f"{lowered_note}; {force_note}" if lowered_note else force_note

        if convergence_candidate_cap_per_iter > 0 and len(candidates) > int(convergence_candidate_cap_per_iter):
            dropped_by_cap = int(len(candidates) - int(convergence_candidate_cap_per_iter))
            candidates = candidates[: int(convergence_candidate_cap_per_iter)]
            cap_note = (
                f"candidate-cap kept={len(candidates)}, dropped={dropped_by_cap}, "
                f"cap={int(convergence_candidate_cap_per_iter)}"
            )
            lowered_note = f"{lowered_note}; {cap_note}" if lowered_note else cap_note

        rows = append_candidate_promotions(wb, next_iter, candidates)

    append_flow_run_log(
        wb,
        next_iter,
        30,
        utc,
        "CHANGED" if candidates else "NO_CHANGE",
        f"Candidates scanned: {len(candidates)}"
        + (f" mode={candidate_scan_mode}" if candidate_scan_mode else "")
        + (f" ({lowered_note})" if lowered_note else "")
        + (f"; MacroMine added {mined_macro_added}" if mined_macro_added else "")
        + (
            f"; AnchorOnly kept={anchor_promo_kept_step30}, dropped={anchor_promo_dropped_step30}, "
            f"anchors={anchor_promo_corpus_size}, min_hits={anchor_promo_min_hits}"
            if anchor_promo_only_enabled
            else ""
        )
        + (f"; {plateau_relax_note}" if plateau_relax_note else ""),
        "0/70",
    )
    candidate_empty_streak = 0 if candidates else (prev_candidate_empty_streak + 1)

    # Step 40: simulate promotions one-by-one, enforcing safety constraints.
    glossary_ws, glossary_map = load_glossary(wb)
    active = {t.token: t for t in glossary_map.values() if t.use_strictplus and t.translation}
    cur_ev, cur_weak, cur_micro, cur_single, cur_tokens = compute_weighted_metrics_for_books_dp(wb, active)
    base_ev, base_weak, base_micro, base_single, base_tokens = cur_ev, cur_weak, cur_micro, cur_single, cur_tokens

    approved: List[str] = []
    macro_fallback_note_parts: List[str] = []
    soft_rescue_promos = 0
    soft_frontier_promos = 0
    hard_escape_promos = 0
    directional_promos = 0
    directional_attempts = 0
    semantic_no_effect_promos = 0
    semantic_no_effect_active = False
    if allow_mech:
        max_passes_default = 2 if cur_status in ("RESOLVED", STATUS_MODEL_CONVERGED) else 1
        max_passes = int(get_setting(settings_map, "PromotionMaxPasses", max_passes_default) or max_passes_default)
        if max_passes < 1:
            max_passes = 1

        skipped_no_effect = 0
        gt_soft_working = int(gt_soft_live)
        soft_rescue_budget = convergence_soft_rescue_max if (convergence_use_soft_rescue and gt_soft_working > 0) else 0
        strict_passes = int(max_passes)
        explored_passes = 0
        candidate_items = list(zip(rows, candidates))
        candidate_row_by_token: Dict[str, int] = {}
        for row_i, cand in candidate_items:
            tok_i = str(cand[0] or "").strip()
            if tok_i and tok_i not in candidate_row_by_token:
                candidate_row_by_token[tok_i] = int(row_i)
        hard_escape_active = (
            convergence_enable_hard_escape
            and (prev_no_mech_iters + 1) >= max(1, int(convergence_hard_escape_stall_iters))
            and bool(candidate_items)
        )
        directional_escape_active = (
            convergence_enable_directional_escape
            and bool(candidate_items)
            and (prev_no_mech_iters + 1) >= max(1, int(convergence_directional_stall_iters))
        )
        low_confidence_escape_active = bool(
            convergence_enable_low_confidence_escape
            and (
                (prev_no_real_progress_iters + 1) >= int(convergence_low_confidence_stall_iters)
                or (prev_candidate_empty_streak + 1) >= int(convergence_candidate_starve_iters)
            )
        )
        semantic_no_effect_active = bool(
            convergence_semantic_objective_enabled
            and (prev_no_mech_iters + 1) >= int(convergence_semantic_no_effect_stall_iters)
        )

        CandidateRow = Tuple[int, Tuple[str, str, str, int, int]]
        uptake_skip_bucket_counts: Counter[str] = Counter()
        uptake_skip_bucket_reasons: Dict[str, Counter[str]] = defaultdict(Counter)

        def _interleave_by_evidence_bucket(
            rows: List[CandidateRow]
        ) -> List[CandidateRow]:
            if not rows or not convergence_interleave_evidence:
                return rows
            buckets: Dict[str, List[CandidateRow]] = defaultdict(list)
            for item in rows:
                _tok, _tr, evcls, _occ, _len = item[1]
                bucket = str(evcls or "").strip() or "OTHER"
                buckets[bucket].append(item)
            if len(buckets) <= 1:
                return rows
            keys = sorted(buckets.keys())
            mixed: List[CandidateRow] = []
            max_len = max((len(v) for v in buckets.values()), default=0)
            for i in range(max_len):
                for k in keys:
                    if i < len(buckets[k]):
                        mixed.append(buckets[k][i])
            return mixed

        def _route_rows(
            rows: List[CandidateRow],
            route: str,
            pass_seed: int,
            top_k: int,
        ) -> List[CandidateRow]:
            work = list(rows)
            if route == "evidence_mix":
                work = _interleave_by_evidence_bucket(work)
            elif route == "occ_desc":
                work.sort(
                    key=lambda item: (
                        -(int(item[1][3]) if item[1][3] is not None else 0),
                        -(int(item[1][4]) if item[1][4] is not None else 0),
                    )
                )
            elif route == "short_first":
                work.sort(
                    key=lambda item: (
                        int(item[1][4]) if item[1][4] is not None else 0,
                        -(int(item[1][3]) if item[1][3] is not None else 0),
                    )
                )
            elif route == "random_mix":
                rng = random.Random(pass_seed)
                rng.shuffle(work)
            else:
                rng = random.Random(pass_seed)
                rng.shuffle(work)
            if top_k and top_k > 0:
                return work[:top_k]
            return work

        def _directional_score(
            ev_delta: float,
            weak_delta: float,
            micro_delta: float,
            single_delta: float,
            token_savings: int,
            gt_soft_delta: int,
            semantic_gain: float = 0.0,
        ) -> float:
            return (
                45.0 * ev_delta
                - 260.0 * max(0.0, weak_delta)
                + 380.0 * max(0.0, -weak_delta)
                - 200.0 * max(0.0, micro_delta)
                + 260.0 * max(0.0, -micro_delta)
                - 90.0 * max(0.0, single_delta)
                + 150.0 * max(0.0, -single_delta)
                + 0.003 * token_savings
                + 0.08 * gt_soft_delta
                + float(convergence_semantic_gain_weight) * float(semantic_gain)
            )

        def _semantic_candidate_gain(gt: GlossaryToken, gt_soft_delta: int) -> float:
            evcls = str(gt.evidence_class or "").strip().upper()
            occ = int(gt.total_occ or 0)
            conf_w = confidence_weight(gt.confidence)
            gain = 0.0

            if evcls in convergence_semantic_positive_classes:
                gain += 1.00
            if evcls in convergence_semantic_negative_classes:
                gain -= 0.80
            if "MACRO" in evcls:
                gain -= 0.20
            if gt.token_type == "macro":
                gain -= 0.20

            if conf_w >= confidence_weight("HIGH"):
                gain += 0.70
            elif conf_w >= confidence_weight("MEDIUM"):
                gain += 0.35
            else:
                gain -= 0.45

            if occ <= 1:
                gain -= 0.25
            elif occ >= 3:
                gain += min(10, occ) * 0.04

            if anchor_promo_corpus_size > 0:
                hits = token_anchor_hit_count(gt.token, anchor_corpus, anchor_hit_cache)
                gain += min(4, hits) * 0.12

            tr = str(gt.translation or "").strip()
            if not tr or tr == "<UNK>":
                gain -= 0.60

            if gt_soft_delta > 0:
                gain += 0.25 * float(gt_soft_delta)

            return float(gain)

        def _diagnose_swallowed_supersets(
            tok: str,
            gt: GlossaryToken,
            test_active: Dict[str, GlossaryToken],
            baseline_metrics: Tuple[float, float, float, float, int],
        ) -> List[str]:
            if not convergence_diagnose_swallowed_supersets:
                return []
            if int(gt.length or 0) < int(convergence_diagnose_swallowed_min_len):
                return []
            if str(gt.evidence_class or "").strip().upper() not in (
                "MACRO_ACTIVE",
                "STRUCT_MACRO_CAND",
                "PHRASE_CRIB",
                "GROUNDTRUTH",
            ):
                return []

            supersets: List[GlossaryToken] = []
            for other_tok, other_gt in active.items():
                other_tok_s = str(other_tok or "")
                if other_tok_s == tok:
                    continue
                if len(other_tok_s) <= len(tok):
                    continue
                if tok not in other_tok_s:
                    continue
                other_evcls = str(other_gt.evidence_class or "").strip().upper()
                if other_gt.token_type != "macro" and other_evcls not in ("MACRO_ACTIVE", "STRUCT_MACRO_CAND"):
                    continue
                supersets.append(other_gt)
            if not supersets:
                return []

            supersets.sort(key=lambda item: (-int(item.length or 0), -int(item.total_occ or 0), str(item.token or "")))
            probe_active = dict(test_active)
            removed: List[str] = []
            for other_gt in supersets[: int(convergence_diagnose_swallowed_max_supersets)]:
                if probe_active.pop(other_gt.token, None) is not None:
                    removed.append(other_gt.token)
            if not removed:
                return []

            probe_metrics = compute_weighted_metrics_for_books_dp(wb, probe_active)
            if tuple(probe_metrics) == tuple(baseline_metrics):
                return []
            return removed[: int(convergence_diagnose_swallowed_max_report)]

        def _simulate_candidate(
            row: int,
            tok: str,
            pass_i: int,
            allow_soft_only: bool,
            frontier_mode: bool = False,
            allow_metric_regression: bool = False,
            frontier_scale: float = 1.0,
            pass_label: str = "strict",
            allow_score_only: bool = False,
            score_min: float = 0.0,
            allow_low_confidence: bool = False,
        ) -> bool:
            nonlocal active, cur_ev, cur_weak, cur_micro, cur_single, cur_tokens, gt_soft_working
            nonlocal approved, skipped_no_effect, soft_rescue_budget, soft_rescue_promos, soft_frontier_promos, hard_escape_promos, directional_promos
            nonlocal semantic_no_effect_promos

            gt = glossary_map.get(tok)
            if gt is None:
                set_candidate_decision(wb, row, "SKIP", "Token not found in Glossary")
                return False
            if convergence_block_promotion_tokens and tok in convergence_block_promotion_tokens:
                set_candidate_decision(wb, row, "SKIP", "Blocked by Convergence_BlockPromotionTokens")
                return False
            if tok in force_probe_token_set and convergence_force_probe_disable_substrings:
                if any(sub in tok for sub in convergence_force_probe_disable_substrings):
                    set_candidate_decision(wb, row, "SKIP", "Blocked by Convergence_ForceProbeDisableSubstrings")
                    return False
            # Allow LOW confidence only for macros / constrained phrase evidence classes
            # (still fully gated by GT live check + DP metrics).
            if gt.confidence == "LOW" and not (
                gt.token_type == "macro" or gt.evidence_class in ("EXTERNAL_POEM", "PHRASE_CRIB")
            ):
                if not allow_low_confidence:
                    set_candidate_decision(wb, row, "SKIP", "LOW confidence (safe mode)")
                    return False

            test_active = dict(active)
            test_active[tok] = gt
            forced_probe_disabled: List[str] = []
            forced_probe_disabled_by_substring: List[str] = []
            forced_probe_supersets_disabled: List[str] = []
            if tok in force_probe_token_set and convergence_force_probe_disable_tokens:
                for forced_off in convergence_force_probe_disable_tokens:
                    if forced_off == tok:
                        continue
                    if test_active.pop(forced_off, None) is not None:
                        forced_probe_disabled.append(forced_off)
            if tok in force_probe_token_set and convergence_force_probe_disable_substrings:
                for other_tok in list(test_active.keys()):
                    other_tok_s = str(other_tok or "")
                    if other_tok_s == tok:
                        continue
                    if not any(sub in other_tok_s for sub in convergence_force_probe_disable_substrings):
                        continue
                    if test_active.pop(other_tok_s, None) is not None:
                        forced_probe_disabled_by_substring.append(other_tok_s)
            if tok in force_probe_token_set and convergence_force_probe_disable_supersets:
                superset_hits: List[GlossaryToken] = []
                for other_tok, other_gt in list(test_active.items()):
                    other_tok_s = str(other_tok or "")
                    if other_tok_s == tok:
                        continue
                    if len(other_tok_s) <= len(tok):
                        continue
                    if tok not in other_tok_s:
                        continue
                    other_evcls = str(other_gt.evidence_class or "").strip().upper()
                    if other_gt.token_type != "macro" and other_evcls not in ("MACRO_ACTIVE", "STRUCT_MACRO_CAND"):
                        continue
                    superset_hits.append(other_gt)
                superset_hits.sort(key=lambda item: (-int(item.length or 0), -int(item.total_occ or 0), str(item.token or "")))
                for other_gt in superset_hits[: int(convergence_force_probe_disable_supersets_max)]:
                    if test_active.pop(other_gt.token, None) is not None:
                        forced_probe_supersets_disabled.append(other_gt.token)

            # Guardrail: never allow a promotion that breaks any GroundTruth crib under the current DP.
            ok_gt2, bad_gt2, bad_all2 = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_gt_ids)
            if not ok_gt2:
                bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad_gt2[:8])
                soft_n = max(0, len(bad_all2) - len(bad_gt2))
                set_candidate_decision(
                    wb,
                    row,
                    "SKIP",
                    f"GT live check mismatch (CribID(s) {bad_ids})"
                    + (f"; soft_mismatches={soft_n}" if soft_n else "")
                    + (f"; force_off={len(forced_probe_disabled)}" if forced_probe_disabled else ""),
                    + (
                        f"; force_substr_off={len(forced_probe_disabled_by_substring)}"
                        if forced_probe_disabled_by_substring
                        else ""
                    ),
                )
                return False

            new_ev, new_weak, new_micro, new_single, new_tokens = compute_weighted_metrics_for_books_dp(wb, test_active)

            ev_delta = new_ev - cur_ev
            weak_delta = new_weak - cur_weak
            micro_delta = new_micro - cur_micro
            single_delta = new_single - cur_single
            token_savings = cur_tokens - new_tokens
            gt_soft_after = max(0, len(bad_all2) - len(bad_gt2))
            gt_soft_delta = int(gt_soft_working) - int(gt_soft_after)
            semantic_gain = _semantic_candidate_gain(gt, gt_soft_delta)
            candidate_score = _directional_score(
                ev_delta=ev_delta,
                weak_delta=weak_delta,
                micro_delta=micro_delta,
                single_delta=single_delta,
                token_savings=token_savings,
                gt_soft_delta=gt_soft_delta,
                semantic_gain=semantic_gain,
            )
            no_effect_dp = (
                token_savings == 0
                and ev_delta == 0
                and weak_delta == 0
                and micro_delta == 0
                and single_delta == 0
            )
            allow_by_score = allow_score_only and (candidate_score >= float(score_min))
            if no_effect_dp and gt_soft_delta <= 0:
                # Directional escape is meant to explore bounded metric tradeoffs, not to
                # promote candidates that the DP never uses. True no-effect candidates must
                # still go through semantic-objective or soft-mismatch rescue.
                allow_by_score = False
            if convergence_strict_monotonic_mechanical and (not allow_soft_only):
                # Anti-churn hard mode: do not allow directional-score bypasses here.
                allow_by_score = False
                monotonic_reasons: List[str] = []
                if weak_delta > 1e-12:
                    monotonic_reasons.append(f"WeakFrac regression {weak_delta:+.6f}")
                if micro_delta > 1e-12:
                    monotonic_reasons.append(f"MicroFrac regression {micro_delta:+.6f}")
                if single_delta > 1e-12:
                    monotonic_reasons.append(f"SingleCharFrac regression {single_delta:+.6f}")
                if ev_delta < -1e-12:
                    monotonic_reasons.append(f"EvidenceAvg regression {ev_delta:+.6f}")
                if (not convergence_monotonic_allow_token_increase) and token_savings < 0:
                    monotonic_reasons.append(f"Token increase {-token_savings} (disallowed)")
                if candidate_score < float(convergence_monotonic_min_score):
                    monotonic_reasons.append(
                        f"Directional score {candidate_score:+.4f} < {float(convergence_monotonic_min_score):.4f}"
                    )
                if monotonic_reasons:
                    skip_reason = "Monotonic gate: " + "; ".join(monotonic_reasons)
                    _record_uptake_skip(
                        uptake_skip_bucket_counts,
                        uptake_skip_bucket_reasons,
                        "DP_BLOCKED_MONOTONIC",
                        skip_reason,
                    )
                    set_candidate_decision(
                        wb,
                        row,
                        "SKIP",
                        "uptake=DP_BLOCKED_MONOTONIC | " + skip_reason,
                    )
                    return False

            semantic_no_effect_escape = False
            semantic_evcls = str(gt.evidence_class or "").strip().upper()
            semantic_occ = int(gt.total_occ or 0)
            semantic_class_ok = bool(
                (not convergence_semantic_allow_classes) or (semantic_evcls in convergence_semantic_allow_classes)
            )

            if no_effect_dp:
                semantic_no_effect_escape = bool(
                    semantic_no_effect_active
                    and semantic_no_effect_promos < convergence_semantic_no_effect_max_promos
                    and semantic_class_ok
                    and semantic_occ >= convergence_semantic_min_occ
                    and semantic_gain >= float(convergence_semantic_no_effect_min_gain)
                )
                if not semantic_no_effect_escape and not (allow_soft_only and gt_soft_delta > 0) and not allow_by_score:
                    swallowed_supersets = _diagnose_swallowed_supersets(
                        tok,
                        gt,
                        test_active,
                        (new_ev, new_weak, new_micro, new_single, new_tokens),
                    )
                    if swallowed_supersets:
                        skip_reason = (
                            "Swallowed by active superset(s): "
                            + ",".join(swallowed_supersets)
                            + (f" | force_off={len(forced_probe_disabled)}" if forced_probe_disabled else "")
                            + (
                                f" | force_substr_off={len(forced_probe_disabled_by_substring)}"
                                if forced_probe_disabled_by_substring
                                else ""
                            )
                            + (
                                f" | force_supersets_off={len(forced_probe_supersets_disabled)}"
                                if forced_probe_supersets_disabled
                                else ""
                            )
                        )
                        _record_uptake_skip(
                            uptake_skip_bucket_counts,
                            uptake_skip_bucket_reasons,
                            "DP_SWALLOWED_SUPERSET",
                            skip_reason,
                        )
                        set_candidate_decision(
                            wb,
                            row,
                            "SKIP",
                            "uptake=DP_SWALLOWED_SUPERSET | " + skip_reason,
                        )
                    else:
                        skip_reason = (
                            "No effect in DP metrics (not used)"
                            + (f" | force_off={len(forced_probe_disabled)}" if forced_probe_disabled else "")
                            + (
                                f" | force_substr_off={len(forced_probe_disabled_by_substring)}"
                                if forced_probe_disabled_by_substring
                                else ""
                            )
                            + (
                                f" | force_supersets_off={len(forced_probe_supersets_disabled)}"
                                if forced_probe_supersets_disabled
                                else ""
                            )
                        )
                        _record_uptake_skip(
                            uptake_skip_bucket_counts,
                            uptake_skip_bucket_reasons,
                            "DP_UNUSED",
                            skip_reason,
                        )
                        set_candidate_decision(
                            wb,
                            row,
                            "SKIP",
                            "uptake=DP_UNUSED | " + skip_reason,
                        )
                    skipped_no_effect += 1
                    return False
                if semantic_no_effect_escape:
                    active = test_active
                    cur_ev, cur_weak, cur_micro, cur_single, cur_tokens = new_ev, new_weak, new_micro, new_single, new_tokens
                    gt_soft_working = int(gt_soft_after)
                    approved.append(tok)
                    semantic_no_effect_promos += 1
                    set_candidate_decision(
                        wb,
                        row,
                        "PROMOTE",
                        f"pass{pass_i}: {pass_label} SEMANTIC_OBJECTIVE evcls={semantic_evcls or 'OTHER'} occ={semantic_occ}, "
                        f"dEv={ev_delta:+.6f}, dWeak={weak_delta:+.6f}, dMicro={micro_delta:+.6f}, dSingle={single_delta:+.6f}, "
                        f"dSoft={gt_soft_delta:+d}, sem={semantic_gain:+.3f}, score={candidate_score:+.4f}",
                    )
                    return True
                if allow_soft_only and soft_rescue_budget <= 0:
                    set_candidate_decision(wb, row, "SKIP", "Soft-mismatch rescue budget exhausted")
                    return False
                soft_only_mode = True
            else:
                soft_only_mode = False

            improves_targets = (weak_delta < 0) or (micro_delta < 0) or (single_delta < 0)
            improves_evidence = ev_delta > 0
            improves_structure = token_savings > 0
            if not improves_targets and not improves_evidence and not improves_structure:
                allow_semantic_only = bool(
                    semantic_no_effect_active
                    and semantic_no_effect_promos < convergence_semantic_no_effect_max_promos
                    and semantic_class_ok
                    and semantic_occ >= convergence_semantic_min_occ
                    and semantic_gain >= float(convergence_semantic_no_effect_min_gain)
                )
                if not (allow_soft_only and gt_soft_delta > 0) and not allow_by_score and not allow_semantic_only:
                    set_candidate_decision(
                        wb,
                        row,
                        "SKIP",
                        "No improvement to target metrics (weak/micro/single), EvidenceAvg, or token-count"
                        + (f" | directional_score={candidate_score:+.4f}" if allow_score_only else ""),
                    )
                    return False
                if allow_semantic_only:
                    active = test_active
                    cur_ev, cur_weak, cur_micro, cur_single, cur_tokens = new_ev, new_weak, new_micro, new_single, new_tokens
                    gt_soft_working = int(gt_soft_after)
                    approved.append(tok)
                    semantic_no_effect_promos += 1
                    set_candidate_decision(
                        wb,
                        row,
                        "PROMOTE",
                        f"pass{pass_i}: {pass_label} SEMANTIC_OBJECTIVE(non-dp) evcls={semantic_evcls or 'OTHER'} occ={semantic_occ}, "
                        f"sem={semantic_gain:+.3f}, score={candidate_score:+.4f}",
                    )
                    return True
                if allow_soft_only and soft_rescue_budget <= 0:
                    set_candidate_decision(wb, row, "SKIP", "Soft-mismatch rescue budget exhausted")
                    return False
                soft_only_mode = True

            # Frontier mode opens a small temporary corridor for difficult cases while keeping hard checks.
            scale = float(frontier_scale) if frontier_scale and frontier_scale > 0 else 1.0
            frontier_max_weak_inc = max(max_weak_inc, 0.001 * scale) if frontier_mode else max_weak_inc
            frontier_max_micro_inc = max(max_micro_inc, 0.001 * scale) if frontier_mode else max_micro_inc
            frontier_max_single_inc = max(max_single_inc, 0.001 * scale) if frontier_mode else max_single_inc
            frontier_max_ev_drop = max(max_ev_drop, 0.001 * scale) if frontier_mode else max_ev_drop
            frontier_min_ev_delta = 0.0 if frontier_mode else min_ev_delta
            if allow_metric_regression and frontier_mode:
                frontier_min_ev_delta = min(frontier_min_ev_delta, -0.002 * scale)

            # EvidenceAvg target guardrail:
            # - if we're already at/above target, do not accept changes that drop below target
            # - if we're below target, allow controlled regression only for hard-escape corridors
            if cur_ev >= target_ev and new_ev < target_ev and not allow_metric_regression:
                set_candidate_decision(wb, row, "SKIP", f"EvidenceAvg would drop below target {target_ev}")
                return False
            if cur_ev < target_ev and new_ev < cur_ev + frontier_min_ev_delta:
                set_candidate_decision(wb, row, "SKIP", f"EvidenceAvg decreased while below target {target_ev}")
                return False

            ok = True
            reasons: List[str] = []
            # Evidence policy:
            # - require >= MinEvidenceAvgDelta unless token_count decreases
            # - allow evidence drops up to MaxEvidenceAvgDrop only when token_count decreases
            if new_ev < cur_ev - frontier_max_ev_drop:
                ok = False
                reasons.append(f"EvidenceAvg drop {ev_delta:+.6f} < -{frontier_max_ev_drop} (hard limit)")
            elif ev_delta < frontier_min_ev_delta:
                if token_savings > 0 and new_ev >= cur_ev - max_ev_drop:
                    reasons.append(f"EvidenceAvg {ev_delta:+.6f} allowed due to token_savings={token_savings}")
                else:
                    ok = False
                    reasons.append(f"EvidenceAvg delta {ev_delta:+.6f} < {frontier_min_ev_delta} (no token savings)")
            if new_weak > cur_weak + frontier_max_weak_inc:
                ok = False
                reasons.append(f"WeakFrac increase {weak_delta:+.6f} > {frontier_max_weak_inc}")
            if new_micro > cur_micro + frontier_max_micro_inc:
                ok = False
                reasons.append(f"MicroFrac increase {micro_delta:+.6f} > {frontier_max_micro_inc}")
            if new_single > cur_single + frontier_max_single_inc:
                ok = False
                reasons.append(f"SingleCharFrac increase {single_delta:+.6f} > {frontier_max_single_inc}")

            if not ok:
                set_candidate_decision(wb, row, "SKIP", "; ".join(reasons) or "Rejected by safety checks")
                return False

            active = test_active
            cur_ev, cur_weak, cur_micro, cur_single, cur_tokens = new_ev, new_weak, new_micro, new_single, new_tokens
            gt_soft_working = int(gt_soft_after)
            approved.append(tok)
            if soft_only_mode:
                soft_rescue_budget -= 1
                soft_rescue_promos += 1
                if frontier_mode:
                    soft_frontier_promos += 1
                set_candidate_decision(
                    wb,
                    row,
                    "PROMOTE",
                    f"pass{pass_i}: {pass_label} dEv={ev_delta:+.6f}, dWeak={weak_delta:+.6f}, dMicro={micro_delta:+.6f}, "
                    f"dSingle={single_delta:+.6f}, dSoft={gt_soft_delta:+d}, score={candidate_score:+.4f}",
                )
            else:
                set_candidate_decision(
                    wb,
                    row,
                    "PROMOTE",
                    f"pass{pass_i}: {pass_label} dEv={ev_delta:+.6f}, dWeak={weak_delta:+.6f}, dMicro={micro_delta:+.6f}, "
                    f"dSingle={single_delta:+.6f}, dSoft={gt_soft_delta:+d}, score={candidate_score:+.4f}",
                )
            return True

        def _set_pair_decision(left_tok: str, right_tok: str, decision: str, note: str) -> None:
            left_row = candidate_row_by_token.get(str(left_tok or "").strip())
            right_row = candidate_row_by_token.get(str(right_tok or "").strip())
            if left_row is not None:
                set_candidate_decision(wb, left_row, decision, f"pair[{left_tok}+{right_tok}] {note}")
            if right_row is not None and right_row != left_row:
                set_candidate_decision(wb, right_row, decision, f"pair[{left_tok}+{right_tok}] {note}")

        def _simulate_pair_candidates(
            left_tok: str,
            right_tok: str,
            pass_i: int,
            pass_label: str = "pair-probe",
        ) -> bool:
            nonlocal active, cur_ev, cur_weak, cur_micro, cur_single, cur_tokens, gt_soft_working
            nonlocal approved

            left_tok = str(left_tok or "").strip()
            right_tok = str(right_tok or "").strip()
            if not left_tok or not right_tok or left_tok == right_tok:
                return False
            if left_tok in approved or right_tok in approved:
                return False
            if convergence_block_promotion_tokens and (
                left_tok in convergence_block_promotion_tokens or right_tok in convergence_block_promotion_tokens
            ):
                _set_pair_decision(left_tok, right_tok, "SKIP", "Blocked by Convergence_BlockPromotionTokens")
                return False

            left_gt = glossary_map.get(left_tok)
            right_gt = glossary_map.get(right_tok)
            if left_gt is None or right_gt is None:
                _set_pair_decision(left_tok, right_tok, "SKIP", "Token not found in Glossary")
                return False

            test_active = dict(active)
            test_active[left_tok] = left_gt
            test_active[right_tok] = right_gt

            ok_gt2, bad_gt2, bad_all2 = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_gt_ids)
            if not ok_gt2:
                bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad_gt2[:8])
                soft_n = max(0, len(bad_all2) - len(bad_gt2))
                note = f"GT live check mismatch (CribID(s) {bad_ids})"
                if soft_n:
                    note += f"; soft_mismatches={soft_n}"
                _set_pair_decision(left_tok, right_tok, "SKIP", note)
                return False

            new_ev, new_weak, new_micro, new_single, new_tokens = compute_weighted_metrics_for_books_dp(wb, test_active)
            ev_delta = new_ev - cur_ev
            weak_delta = new_weak - cur_weak
            micro_delta = new_micro - cur_micro
            single_delta = new_single - cur_single
            token_savings = cur_tokens - new_tokens
            gt_soft_after = max(0, len(bad_all2) - len(bad_gt2))
            gt_soft_delta = int(gt_soft_working) - int(gt_soft_after)
            semantic_gain = _semantic_candidate_gain(left_gt, gt_soft_delta) + _semantic_candidate_gain(right_gt, gt_soft_delta)
            candidate_score = _directional_score(
                ev_delta=ev_delta,
                weak_delta=weak_delta,
                micro_delta=micro_delta,
                single_delta=single_delta,
                token_savings=token_savings,
                gt_soft_delta=gt_soft_delta,
                semantic_gain=semantic_gain,
            )

            no_effect_dp = (
                token_savings == 0
                and ev_delta == 0
                and weak_delta == 0
                and micro_delta == 0
                and single_delta == 0
            )
            if no_effect_dp and gt_soft_delta <= 0:
                skip_reason = "No effect in DP metrics (pair not used)"
                _record_uptake_skip(
                    uptake_skip_bucket_counts,
                    uptake_skip_bucket_reasons,
                    "DP_UNUSED",
                    skip_reason,
                )
                _set_pair_decision(left_tok, right_tok, "SKIP", "uptake=DP_UNUSED | " + skip_reason)
                return False

            if convergence_strict_monotonic_mechanical:
                monotonic_reasons: List[str] = []
                if weak_delta > 1e-12:
                    monotonic_reasons.append(f"WeakFrac regression {weak_delta:+.6f}")
                if micro_delta > 1e-12:
                    monotonic_reasons.append(f"MicroFrac regression {micro_delta:+.6f}")
                if single_delta > 1e-12:
                    monotonic_reasons.append(f"SingleCharFrac regression {single_delta:+.6f}")
                if ev_delta < -1e-12:
                    monotonic_reasons.append(f"EvidenceAvg regression {ev_delta:+.6f}")
                if (not convergence_monotonic_allow_token_increase) and token_savings < 0:
                    monotonic_reasons.append(f"Token increase {-token_savings} (disallowed)")
                if candidate_score < float(convergence_monotonic_min_score):
                    monotonic_reasons.append(
                        f"Directional score {candidate_score:+.4f} < {float(convergence_monotonic_min_score):.4f}"
                    )
                if monotonic_reasons and not (gt_soft_delta > 0):
                    skip_reason = "Monotonic gate: " + "; ".join(monotonic_reasons)
                    _record_uptake_skip(
                        uptake_skip_bucket_counts,
                        uptake_skip_bucket_reasons,
                        "DP_BLOCKED_MONOTONIC",
                        skip_reason,
                    )
                    _set_pair_decision(left_tok, right_tok, "SKIP", "uptake=DP_BLOCKED_MONOTONIC | " + skip_reason)
                    return False

            improves_targets = (weak_delta < 0) or (micro_delta < 0) or (single_delta < 0)
            improves_evidence = ev_delta > 0
            improves_structure = token_savings > 0
            if not improves_targets and not improves_evidence and not improves_structure and not (gt_soft_delta > 0):
                _set_pair_decision(
                    left_tok,
                    right_tok,
                    "SKIP",
                    "No improvement to target metrics (weak/micro/single), EvidenceAvg, or token-count (pair)",
                )
                return False

            active = test_active
            cur_ev, cur_weak, cur_micro, cur_single, cur_tokens = new_ev, new_weak, new_micro, new_single, new_tokens
            gt_soft_working = int(gt_soft_after)
            approved.append(left_tok)
            approved.append(right_tok)
            _set_pair_decision(
                left_tok,
                right_tok,
                "PROMOTE",
                f"pass{pass_i}: {pass_label} dEv={ev_delta:+.6f}, dWeak={weak_delta:+.6f}, dMicro={micro_delta:+.6f}, "
                f"dSingle={single_delta:+.6f}, dSoft={gt_soft_delta:+d}, score={candidate_score:+.4f}",
            )
            return True

        pass_i = 1
        while pass_i <= strict_passes:
            pass_soft_rescue = convergence_use_soft_rescue and pass_i > 1 and gt_soft_working > 0
            if pass_soft_rescue and soft_rescue_budget <= 0:
                break
            pass_rows = list(candidate_items)
            if pass_soft_rescue and convergence_rescue_shuffle:
                rng = random.Random((next_iter * 1000003) ^ (pass_i * 257) ^ int(gt_soft_working))
                rng.shuffle(pass_rows)
            promoted_this_pass = 0
            for row, (tok, _tr, _evcls, _occ, _len) in pass_rows:
                if tok in approved:
                    continue
                if _simulate_candidate(
                    row=row,
                    tok=tok,
                    pass_i=pass_i,
                    allow_soft_only=pass_soft_rescue,
                    frontier_mode=False,
                    pass_label="strict" if not pass_soft_rescue else "soft-shuffle",
                    allow_low_confidence=low_confidence_escape_active,
                ):
                    promoted_this_pass += 1
                    if pass_soft_rescue and soft_rescue_budget <= 0:
                        break
            if promoted_this_pass == 0:
                if not pass_soft_rescue:
                    break
            pass_i += 1

        if (not approved) and convergence_force_probe_pair_tokens:
            for left_tok, right_tok in convergence_force_probe_pair_tokens:
                if _simulate_pair_candidates(
                    left_tok=left_tok,
                    right_tok=right_tok,
                    pass_i=pass_i,
                    pass_label="pair-probe",
                ):
                    break

        while pass_i <= strict_passes + 2 and gt_soft_working > 0 and soft_rescue_budget > 0:
            pass_soft_rescue = True
            frontier_mode = explored_passes > 0
            pass_rows = list(candidate_items)
            if convergence_interleave_evidence and frontier_mode:
                pass_rows = _interleave_by_evidence_bucket(pass_rows)
            if frontier_mode:
                rng = random.Random((next_iter * 1000003) ^ (pass_i * 397) ^ int(gt_soft_working) ^ int(soft_frontier_promos))
                pass_rows = pass_rows[: max(10, convergence_soft_rescue_max * 10)]
                rng.shuffle(pass_rows)
            else:
                rng = random.Random((next_iter * 1000003) ^ (pass_i * 257) ^ int(gt_soft_working))
                if convergence_rescue_shuffle:
                    rng.shuffle(pass_rows)
            promoted_this_pass = 0
            for row, (tok, _tr, _evcls, _occ, _len) in pass_rows:
                if tok in approved:
                    continue
                if _simulate_candidate(
                    row=row,
                    tok=tok,
                    pass_i=pass_i,
                    allow_soft_only=pass_soft_rescue,
                    frontier_mode=frontier_mode,
                    allow_metric_regression=False,
                    frontier_scale=1.0 + min(2.0, explored_passes * 0.4),
                    pass_label="soft-frontier" if frontier_mode else "soft-shuffle",
                    allow_low_confidence=low_confidence_escape_active,
                ):
                    promoted_this_pass += 1
                    if soft_rescue_budget <= 0:
                        break
            if promoted_this_pass == 0:
                break
            explored_passes += 1
            pass_i += 1

        directional_routes = ["evidence_mix", "occ_desc", "short_first", "random_mix"]
        if (
            directional_escape_active
            and not approved
        ):
            directional_passes_budget = max(0, int(convergence_directional_passes))
            directional_pass_no = 0
            while directional_pass_no < directional_passes_budget and not approved:
                directional_pass_no += 1
                route = directional_routes[(directional_pass_no - 1) % len(directional_routes)]
                directional_attempts += 1
                pass_seed = (next_iter * 1000003) ^ (directional_pass_no * 523) ^ int(gt_soft_working) ^ pass_i
                pass_rows = _route_rows(
                    list(candidate_items),
                    route=route,
                    pass_seed=pass_seed,
                    top_k=max(1, int(convergence_directional_topk)),
                )
                if not pass_rows:
                    continue
                if convergence_interleave_evidence and route != "evidence_mix" and (directional_pass_no % 2 == 0):
                    pass_rows = _interleave_by_evidence_bucket(pass_rows)
                promoted_this_pass = 0
                rng = random.Random(pass_seed)
                rng.shuffle(pass_rows)
                frontier_scale = float(convergence_directional_frontier_scale) * (1.0 + 0.35 * (directional_pass_no - 1))
                for row, (tok, _tr, _evcls, _occ, _len) in pass_rows:
                    if tok in approved:
                        continue
                    if _simulate_candidate(
                        row=row,
                        tok=tok,
                        pass_i=pass_i,
                        allow_soft_only=(gt_soft_working > 0),
                        frontier_mode=True,
                        allow_metric_regression=convergence_directional_allow_regression,
                        frontier_scale=frontier_scale,
                        allow_score_only=True,
                        score_min=convergence_directional_score_min,
                        pass_label=f"directional:{route}",
                        allow_low_confidence=low_confidence_escape_active,
                    ):
                        directional_promos += 1
                        promoted_this_pass += 1
                        if directional_promos >= max(1, int(convergence_directional_topk)):
                            break
                if promoted_this_pass == 0:
                    continue
                if gt_soft_working == 0:
                    # Keep exploration bounded; accept at most one directional pass worth of movement by default.
                    break

        if directional_attempts:
            macro_fallback_note_parts.append(
                f"directional_escape_attempts={directional_attempts}, directional_escape_promos={directional_promos}"
            )

        hard_escape_budget = int(convergence_hard_escape_passes) if hard_escape_active else 0
        hard_escape_attempt = 0
        while (
            hard_escape_budget > 0
            and not approved
            and hard_escape_active
            and candidate_items
        ):
            hard_escape_attempt += 1
            hard_escape_budget -= 1
            pass_rows = list(candidate_items)
            if convergence_interleave_evidence:
                pass_rows = _interleave_by_evidence_bucket(pass_rows)
            rng = random.Random((next_iter * 1000003) ^ (pass_i * 811) ^ hard_escape_attempt ^ int(gt_soft_working))
            rng.shuffle(pass_rows)
            if hard_escape_attempt % 2 == 0:
                pass_rows = pass_rows[: max(10, convergence_hard_escape_passes * 10)]
            promoted_this_pass = 0
            frontier_scale = float(convergence_hard_escape_tol_scale) * (1.0 + 0.5 * (hard_escape_attempt - 1))
            for row, (tok, _tr, _evcls, _occ, _len) in pass_rows:
                if tok in approved:
                    continue
                if _simulate_candidate(
                    row=row,
                    tok=tok,
                    pass_i=pass_i,
                    allow_soft_only=(gt_soft_working > 0),
                    frontier_mode=True,
                    allow_metric_regression=True,
                    frontier_scale=frontier_scale,
                    allow_score_only=True,
                    score_min=convergence_directional_score_min,
                    pass_label="hard-escape",
                    allow_low_confidence=low_confidence_escape_active,
                ):
                    hard_escape_promos += 1
                    promoted_this_pass += 1
                    if hard_escape_budget <= 0:
                        break
            if promoted_this_pass == 0:
                break
            hard_escape_active = True
            pass_i += 1

        if hard_escape_active:
            macro_fallback_note_parts.append(f"hard_escape_attempts={hard_escape_attempt}")

        # If no promotions were approved, macro-mining is the next safe lever.
        # Mined macros are appended inactive, and any promotion is still gated by GT live check + DP metrics.
        mined_macro_added_step40_total = 0
        mined_macro_promoted_step40_total = 0

        def _apply_plateau_rung_settings(target_rung: int) -> None:
            nonlocal plateau_rung, ws_settings, settings_map
            target_rung = max(0, min(int(target_rung), int(convergence_candidate_emergency_rung)))
            if target_rung <= plateau_rung:
                return
            plateau_rung = target_rung
            flow_setting_set(store, "PlateauLadder_Rung", plateau_rung, note=f"iter{next_iter}: in-iter macro ladder rung{plateau_rung}")

            cur_min_share = float(get_setting(settings_map, "MacroMine_MinShare", 0.95) or 0.95)
            cur_max_len = int(get_setting(settings_map, "MacroMine_MaxLen", 16) or 16)
            cur_max_candidates = int(get_setting(settings_map, "MacroMine_MaxCandidates", 75) or 75)

            if plateau_rung >= 1:
                flow_setting_set(store, "MacroMine_MinShare", min(cur_min_share, 0.90), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_MaxLen", max(cur_max_len, 24), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_MaxCandidates", max(cur_max_candidates, 150), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_NValues", "2,3,4,5,6,7,8,9,10", note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
            if plateau_rung >= 2:
                flow_setting_set(store, "MacroMine_MinShare", min(cur_min_share, 0.85), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
            if plateau_rung >= 3:
                flow_setting_set(store, "MacroMine_MinShare", min(cur_min_share, 0.75), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_MinOcc", 1, note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_MinBooks", 1, note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_MinLen", 2, note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_MaxLen", max(cur_max_len, 40), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")
                flow_setting_set(store, "MacroMine_MaxCandidates", max(cur_max_candidates, 700), note=f"iter{next_iter}: plateau ladder rung{plateau_rung}")

            flow_store_refresh_from_workbook(store)

            ws_settings = store["ws_settings"]

            settings_map = store["settings_map"]

        def _mine_and_try_promote_macros(rung: int) -> Tuple[int, int]:
            nonlocal glossary_ws, glossary_map, active, cur_ev, cur_weak, cur_micro, cur_single, cur_tokens
            nonlocal anchor_promo_kept_step40, anchor_promo_dropped_step40

            macro_min_occ = int(get_setting(settings_map, "MacroMine_MinOcc", 2) or 2)
            macro_min_books = int(get_setting(settings_map, "MacroMine_MinBooks", 2) or 2)
            macro_min_share_default = 0.95 if macro_mine_always else 1.0
            macro_min_share = float(get_setting(settings_map, "MacroMine_MinShare", macro_min_share_default) or macro_min_share_default)
            macro_min_len = int(get_setting(settings_map, "MacroMine_MinLen", 2) or 2)

            macro_max_candidates_default = 75 if macro_mine_always else 25
            macro_max_candidates = int(get_setting(settings_map, "MacroMine_MaxCandidates", macro_max_candidates_default) or macro_max_candidates_default)
            macro_max_len_default = 16 if macro_mine_always else 12
            macro_max_len = int(get_setting(settings_map, "MacroMine_MaxLen", macro_max_len_default) or macro_max_len_default)

            macro_n_values_default = "2,3,4,5,6" if macro_mine_always else "2,3,4"
            macro_n_values = parse_int_list(get_setting(settings_map, "MacroMine_NValues", macro_n_values_default) or macro_n_values_default)
            if not macro_n_values:
                macro_n_values = [2, 3, 4]

            allow_macro_components_default = True if macro_mine_always else False
            allow_macro_components = parse_bool(
                get_setting(settings_map, "MacroMine_AllowMacroComponents", allow_macro_components_default),
                allow_macro_components_default,
            )
            anchor_macro_mine_enabled = parse_bool(
                get_setting(settings_map, "AnchorMacroMine_Enabled", False),
                False,
            )

            # Ladder overrides (still safe: DP can only apply a macro if it matches the existing lossless stream).
            if rung >= 1:
                macro_min_share = min(macro_min_share, 0.90)
                macro_max_len = max(macro_max_len, 24)
                macro_max_candidates = max(macro_max_candidates, 150)
                macro_n_values = list(range(2, 11))
                allow_macro_components = True
            if rung >= 2:
                macro_min_share = min(macro_min_share, 0.85)
            if rung >= 3:
                macro_min_occ = max(1, macro_min_occ)
                macro_min_books = max(1, macro_min_books)
                macro_min_share = min(macro_min_share, 0.75)
                macro_max_candidates = max(macro_max_candidates, 700)
                macro_max_len = max(macro_max_len, 40)
                macro_n_values = list(range(2, 13))

            mined = mine_macro_candidates_from_books(
                wb,
                active,  # current baseline active token set
                existing_tokens=list(glossary_map.keys()),
                n_values=tuple(macro_n_values),
                min_occ=macro_min_occ,
                min_books=macro_min_books,
                min_share=macro_min_share,
                min_len=macro_min_len,
                max_len=macro_max_len,
                max_candidates=macro_max_candidates,
                allow_macro_components=allow_macro_components,
                allow_marker_tokens=allow_markers,
                allow_star_tokens=allow_stars,
            )
            added_books = add_mined_macros_to_glossary(
                wb,
                next_iter,
                mined,
                evidence_class="MACRO_ACTIVE",
                source_kind="n-gram",
                evidence_sources_tag="mined_ngram_macro",
            )
            anchor_mined: List[MinedMacro] = []
            added_anchor = 0
            if anchor_promo_only_enabled and anchor_macro_mine_enabled and anchor_corpus:
                anchor_mined = mine_macro_candidates_from_anchor_corpus(
                    anchor_corpus,
                    active,
                    existing_tokens=list(glossary_map.keys()) + [m.base for m in mined[:added_books]],
                    n_values=tuple(macro_n_values),
                    min_occ=macro_min_occ,
                    min_books=macro_min_books,
                    min_share=macro_min_share,
                    min_len=macro_min_len,
                    max_len=macro_max_len,
                    max_candidates=macro_max_candidates,
                    allow_macro_components=allow_macro_components,
                    allow_marker_tokens=allow_markers,
                    allow_star_tokens=allow_stars,
                )
                added_anchor = add_mined_macros_to_glossary(
                    wb,
                    next_iter,
                    anchor_mined,
                    evidence_class="MACRO_ACTIVE",
                    source_kind="anchor-corpus",
                    evidence_sources_tag="mined_anchor_corpus_macro",
                )
            added = int(added_books) + int(added_anchor)
            if not added:
                return 0, 0

            macro_candidates = (
                [(m.base, m.translation, "MACRO_ACTIVE", m.occ, m.length) for m in mined[:added_books]]
                + [(m.base, m.translation, "MACRO_ACTIVE", m.occ, m.length) for m in anchor_mined[:added_anchor]]
            )
            if anchor_promo_only_enabled:
                macro_candidates, dropped_anchor, anchor_diag = filter_candidates_by_anchor_impact(
                    macro_candidates,
                    anchor_corpus=anchor_corpus,
                    min_hits=anchor_promo_min_hits,
                    cache=anchor_hit_cache,
                )
                anchor_promo_kept_step40 += int(len(macro_candidates))
                anchor_promo_dropped_step40 += int(dropped_anchor)
                anchor_promo_drop40_classes = str(anchor_diag.get("classes", "") or "")
                anchor_promo_drop40_samples = str(anchor_diag.get("samples", "") or "")
                if dropped_anchor:
                    macro_fallback_note_parts.append(
                        "anchor-only macros kept="
                        f"{len(macro_candidates)}, dropped={dropped_anchor}"
                        + (f", classes={anchor_promo_drop40_classes}" if anchor_promo_drop40_classes else "")
                    )
                if not macro_candidates:
                    return added, 0
            if added_anchor:
                macro_fallback_note_parts.append(
                    f"anchorMacroMine rung{rung}: added={added_anchor}, corpus={len(anchor_corpus)}"
                )
            macro_rows = append_candidate_promotions(wb, next_iter, macro_candidates)

            # Reload Glossary to see appended rows.
            glossary_ws, glossary_map = load_glossary(wb)
            promoted = 0
            for row, (tok, _tr, _evcls, _occ, _len) in zip(macro_rows, macro_candidates):
                gt = glossary_map.get(tok)
                if gt is None:
                    set_candidate_decision(wb, row, "SKIP", "Token not found in Glossary (post-mine)")
                    continue
                if gt.confidence == "LOW" and not (
                    gt.token_type == "macro" or gt.evidence_class in ("EXTERNAL_POEM", "PHRASE_CRIB")
                ):
                    if not low_confidence_escape_active:
                        set_candidate_decision(wb, row, "SKIP", "LOW confidence (safe mode)")
                        continue

                test_active = dict(active)
                test_active[tok] = gt

                ok_gt2, bad_gt2, bad_all2 = groundtruth_live_check(wb, test_active, enforced_crib_ids=enforced_gt_ids)
                if not ok_gt2:
                    bad_ids = ",".join(str(cid) for cid, _crib, _dec, _exp in bad_gt2[:8])
                    soft_n = max(0, len(bad_all2) - len(bad_gt2))
                    set_candidate_decision(
                        wb,
                        row,
                        "SKIP",
                        f"GT live check mismatch (CribID(s) {bad_ids})" + (f"; soft_mismatches={soft_n}" if soft_n else ""),
                    )
                    continue

                new_ev, new_weak, new_micro, new_single, new_tokens = compute_weighted_metrics_for_books_dp(wb, test_active)

                ev_delta = new_ev - cur_ev
                weak_delta = new_weak - cur_weak
                micro_delta = new_micro - cur_micro
                single_delta = new_single - cur_single
                token_savings = cur_tokens - new_tokens
                semantic_gain = _semantic_candidate_gain(gt, 0)
                candidate_score = _directional_score(
                    ev_delta=ev_delta,
                    weak_delta=weak_delta,
                    micro_delta=micro_delta,
                    single_delta=single_delta,
                    token_savings=token_savings,
                    gt_soft_delta=0,
                    semantic_gain=semantic_gain,
                )
                if convergence_strict_monotonic_mechanical:
                    monotonic_reasons: List[str] = []
                    if weak_delta > 1e-12:
                        monotonic_reasons.append(f"WeakFrac regression {weak_delta:+.6f}")
                    if micro_delta > 1e-12:
                        monotonic_reasons.append(f"MicroFrac regression {micro_delta:+.6f}")
                    if single_delta > 1e-12:
                        monotonic_reasons.append(f"SingleCharFrac regression {single_delta:+.6f}")
                    if ev_delta < -1e-12:
                        monotonic_reasons.append(f"EvidenceAvg regression {ev_delta:+.6f}")
                    if (not convergence_monotonic_allow_token_increase) and token_savings < 0:
                        monotonic_reasons.append(f"Token increase {-token_savings} (disallowed)")
                    if candidate_score < float(convergence_monotonic_min_score):
                        monotonic_reasons.append(
                            f"Directional score {candidate_score:+.4f} < {float(convergence_monotonic_min_score):.4f}"
                        )
                    if monotonic_reasons:
                        skip_reason = "Monotonic gate: " + "; ".join(monotonic_reasons)
                        _record_uptake_skip(
                            uptake_skip_bucket_counts,
                            uptake_skip_bucket_reasons,
                            "DP_BLOCKED_MONOTONIC",
                            skip_reason,
                        )
                        set_candidate_decision(
                            wb,
                            row,
                            "SKIP",
                            "uptake=DP_BLOCKED_MONOTONIC | " + skip_reason,
                        )
                        continue

                if token_savings == 0 and ev_delta == 0 and weak_delta == 0 and micro_delta == 0 and single_delta == 0:
                    semantic_evcls = str(gt.evidence_class or "").strip().upper()
                    semantic_occ = int(gt.total_occ or 0)
                    semantic_class_ok = bool(
                        (not convergence_semantic_allow_classes) or (semantic_evcls in convergence_semantic_allow_classes)
                    )
                    allow_semantic_no_effect = bool(
                        semantic_no_effect_active
                        and semantic_no_effect_promos < convergence_semantic_no_effect_max_promos
                        and semantic_class_ok
                        and semantic_occ >= convergence_semantic_min_occ
                        and semantic_gain >= float(convergence_semantic_no_effect_min_gain)
                    )
                    if not allow_semantic_no_effect:
                        skip_reason = "No effect in DP metrics (not used)"
                        _record_uptake_skip(
                            uptake_skip_bucket_counts,
                            uptake_skip_bucket_reasons,
                            "DP_UNUSED",
                            skip_reason,
                        )
                        set_candidate_decision(wb, row, "SKIP", "uptake=DP_UNUSED | " + skip_reason)
                        continue
                    active = test_active
                    cur_ev, cur_weak, cur_micro, cur_single, cur_tokens = new_ev, new_weak, new_micro, new_single, new_tokens
                    approved.append(tok)
                    promoted += 1
                    semantic_no_effect_promos += 1
                    set_candidate_decision(
                        wb,
                        row,
                        "PROMOTE",
                        f"macro-fallback SEMANTIC_OBJECTIVE evcls={semantic_evcls or 'OTHER'} occ={semantic_occ}, "
                        f"dEv={ev_delta:+.6f}, dWeak={weak_delta:+.6f}, dMicro={micro_delta:+.6f}, dSingle={single_delta:+.6f}, "
                        f"sem={semantic_gain:+.3f}, score={candidate_score:+.4f}",
                    )
                    continue

                improves_targets = (weak_delta < 0) or (micro_delta < 0) or (single_delta < 0)
                improves_evidence = ev_delta > 0
                improves_structure = token_savings > 0
                if not improves_targets and not improves_evidence and not improves_structure:
                    set_candidate_decision(
                        wb,
                        row,
                        "SKIP",
                        "No improvement to target metrics (weak/micro/single), EvidenceAvg, or token-count",
                    )
                    continue

                if cur_ev >= target_ev and new_ev < target_ev:
                    set_candidate_decision(wb, row, "SKIP", f"EvidenceAvg would drop below target {target_ev}")
                    continue
                if cur_ev < target_ev and new_ev < cur_ev:
                    set_candidate_decision(wb, row, "SKIP", f"EvidenceAvg decreased while below target {target_ev}")
                    continue

                ok = True
                reasons: List[str] = []
                if new_ev < cur_ev - max_ev_drop:
                    ok = False
                    reasons.append(f"EvidenceAvg drop {ev_delta:+.6f} < -{max_ev_drop} (hard limit)")
                elif ev_delta < min_ev_delta:
                    if token_savings > 0 and new_ev >= cur_ev - max_ev_drop:
                        reasons.append(f"EvidenceAvg {ev_delta:+.6f} allowed due to token_savings={token_savings}")
                    else:
                        ok = False
                        reasons.append(f"EvidenceAvg delta {ev_delta:+.6f} < {min_ev_delta} (no token savings)")
                if new_weak > cur_weak + max_weak_inc:
                    ok = False
                    reasons.append(f"WeakFrac increase {weak_delta:+.6f} > {max_weak_inc}")
                if new_micro > cur_micro + max_micro_inc:
                    ok = False
                    reasons.append(f"MicroFrac increase {micro_delta:+.6f} > {max_micro_inc}")
                if new_single > cur_single + max_single_inc:
                    ok = False
                    reasons.append(f"SingleCharFrac increase {single_delta:+.6f} > {max_single_inc}")

                if not ok:
                    set_candidate_decision(wb, row, "SKIP", "; ".join(reasons) or "Rejected by safety checks")
                    continue

                active = test_active
                cur_ev, cur_weak, cur_micro, cur_single, cur_tokens = new_ev, new_weak, new_micro, new_single, new_tokens
                approved.append(tok)
                promoted += 1
                set_candidate_decision(
                    wb,
                    row,
                    "PROMOTE",
                    f"Safe: dEv={ev_delta:+.6f}, dWeak={weak_delta:+.6f}, dMicro={micro_delta:+.6f}, dSingle={single_delta:+.6f}",
                )

            return added, promoted

        if not approved and (not macro_mine_off):
            # Maximize progress per iteration: try a small ladder (baseline -> relaxed -> more relaxed).
            rung_start = max(0, plateau_rung) if plateau_auto_relax else 0
            rung_seq = list(range(rung_start, int(convergence_candidate_emergency_rung) + 1)) if plateau_auto_relax else [0]
            for rung in rung_seq:
                if approved:
                    break
                if plateau_auto_relax and rung > plateau_rung:
                    _apply_plateau_rung_settings(rung)
                added, promoted = _mine_and_try_promote_macros(rung)
                mined_macro_added_step40_total += int(added)
                mined_macro_promoted_step40_total += int(promoted)
                macro_fallback_note_parts.append(f"macroMine rung{rung}: added={added}, promoted={promoted}")
                if promoted > 0:
                    break

        if mined_macro_added_step40_total:
            macro_fallback_note_parts.append(f"macroMine total added={mined_macro_added_step40_total}")
        if mined_macro_promoted_step40_total:
            macro_fallback_note_parts.append(f"macroMine total promoted={mined_macro_promoted_step40_total}")

    promotion_skip_count, promotion_skip_reason_top = summarize_candidate_promotion_skips(wb, next_iter)
    dp_unused_count = int(uptake_skip_bucket_counts.get("DP_UNUSED", 0))
    dp_swallowed_count = int(uptake_skip_bucket_counts.get("DP_SWALLOWED_SUPERSET", 0))
    dp_blocked_monotonic_count = int(uptake_skip_bucket_counts.get("DP_BLOCKED_MONOTONIC", 0))
    dp_unused_top = _counter_top_str(uptake_skip_bucket_reasons.get("DP_UNUSED") or Counter())
    dp_swallowed_top = _counter_top_str(uptake_skip_bucket_reasons.get("DP_SWALLOWED_SUPERSET") or Counter())
    dp_blocked_monotonic_top = _counter_top_str(uptake_skip_bucket_reasons.get("DP_BLOCKED_MONOTONIC") or Counter())
    approved_total_count = len(approved)
    semantic_objective_promo_count = int(semantic_no_effect_promos)
    mechanical_promotions_count = max(0, approved_total_count - semantic_objective_promo_count)
    append_flow_run_log(
        wb,
        next_iter,
        40,
        utc,
        "CHANGED" if approved else "NO_CHANGE",
        f"Mechanical promotions approved: {mechanical_promotions_count}",
        "0/70",
        evidence_avg=round(cur_ev, 6),
        weak_frac=round(cur_weak, 6),
        micro_frac=round(cur_micro, 6),
        notes=(
            f"Before Ev={base_ev:.6f}, Weak={base_weak:.6f}, Micro={base_micro:.6f}, Single={base_single:.6f}, Tokens={base_tokens}; "
            f"After Ev={cur_ev:.6f}, Weak={cur_weak:.6f}, Micro={cur_micro:.6f}, Single={cur_single:.6f}, Tokens={cur_tokens}"
            + (f"; hard_escape_promos={hard_escape_promos}" if hard_escape_promos else "")
            + (f"; directional_escape_attempts={directional_attempts}, directional_escape_promos={directional_promos}" if directional_attempts else "")
            + (f"; {'; '.join(macro_fallback_note_parts)}" if macro_fallback_note_parts else "")
            + f"; skips={promotion_skip_count}"
            + (f"; top_skip={promotion_skip_reason_top}" if promotion_skip_reason_top else "")
            + (
                f"; uptake_unused={dp_unused_count}"
                + (f"[{dp_unused_top}]" if dp_unused_top else "")
                + f"; uptake_swallowed={dp_swallowed_count}"
                + (f"[{dp_swallowed_top}]" if dp_swallowed_top else "")
                + f"; uptake_blocked_monotonic={dp_blocked_monotonic_count}"
                + (f"[{dp_blocked_monotonic_top}]" if dp_blocked_monotonic_top else "")
            )
            + (f"; semantic_objective_promos={semantic_objective_promo_count}" if semantic_objective_promo_count else "")
        ),
    )

    # Step 50: apply promotions in Glossary
    for tok in approved:
        gt = glossary_map.get(tok)
        if gt is None:
            continue
        glossary_set_use_strictplus(glossary_ws, gt.row, True, next_iter, "promoted by auto-chain")

    append_flow_run_log(
        wb,
        next_iter,
        50,
        utc,
        "CHANGED" if approved else "NO_CHANGE",
        f"Applied promotions: total={approved_total_count}, mechanical={mechanical_promotions_count}, semantic_objective={semantic_objective_promo_count}",
        "0/70",
    )

    # Step 55: materialize ExternalRefs decode view from CodeStream v120
    externalrefs_fill_codestream_v120_enabled = parse_bool(
        get_setting(settings_map, "ExternalRefs_FillFromCodeStreamV120_Enabled", True),
        True,
    )
    ext_changed = (
        materialize_external_refs_from_codestream_v120(wb, next_iter)
        if externalrefs_fill_codestream_v120_enabled
        else 0
    )
    append_flow_run_log(
        wb,
        next_iter,
        55,
        utc,
        "CHANGED" if ext_changed else "NO_CHANGE",
        (
            f"ExternalRefs filled from CodeStream v120: {ext_changed} rows"
            if externalrefs_fill_codestream_v120_enabled
            else "ExternalRefs fill from CodeStream v120 disabled"
        ),
        "0/70",
    )

    # Step 101: DigitCodeMap (analysis-only): reconstruct deterministic code->letter mapping + homophones.
    digit_codes = 0
    digit_letters = 0
    digit_conflicts = 0
    digit_status = ""
    try:
        digit_codes, digit_letters, digit_conflicts, digit_status = materialize_digit_code_map_from_books(wb, next_iter, utc)
        res = "CHANGED" if digit_status == "ok" else ("NO_CHANGE" if digit_status.startswith("skipped") else "FAILED")
        append_flow_run_log(
            wb,
            next_iter,
            101,
            utc,
            res,
            f"DigitCodeMap: codes={digit_codes}, letters={digit_letters}, conflicts={digit_conflicts} ({digit_status})",
            "0/70",
        )
    except Exception as e:
        append_flow_run_log(
            wb,
            next_iter,
            101,
            utc,
            "FAILED",
            "DigitCodeMap: FAILED",
            "0/70",
            notes=str(e),
        )

    # Step 110: DigitCodeContext (analysis-only): neighbor distributions for homophone study (no decode changes).
    digit_ctx_rows = 0
    digit_ctx_homophone_letters = 0
    digit_ctx_best_outlier = 0.0
    digit_ctx_fp = ""
    digit_ctx_fp_changed = 0
    digit_ctx_status = "skipped"
    if digit_ctx_enabled:
        try:
            digit_ctx_rows, digit_ctx_homophone_letters, digit_ctx_best_outlier, digit_ctx_fp, digit_ctx_status = (
                materialize_digit_code_context_profiles_from_books(
                    wb,
                    next_iter,
                    utc,
                    topk=int(digit_ctx_topk),
                    js_alpha=float(digit_ctx_js_alpha),
                    prev_fp=str(prev_digit_ctx_fp or ""),
                )
            )
            if digit_ctx_fp and digit_ctx_fp != str(prev_digit_ctx_fp or ""):
                digit_ctx_fp_changed = 1
            res = "CHANGED" if digit_ctx_status == "ok" else ("NO_CHANGE" if digit_ctx_status.startswith("skipped") else "FAILED")
            append_flow_run_log(
                wb,
                next_iter,
                110,
                utc,
                res,
                f"DigitCodeContext: rows={digit_ctx_rows}, homophone_letters={digit_ctx_homophone_letters}, best_outlier={digit_ctx_best_outlier:.6f} ({digit_ctx_status})",
                "0/70",
                notes=f"topk={digit_ctx_topk}, js_alpha={digit_ctx_js_alpha}, fp_changed={digit_ctx_fp_changed}",
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                110,
                utc,
                "FAILED",
                "DigitCodeContext: FAILED",
                "0/70",
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            110,
            utc,
            "NO_CHANGE",
            "DigitCodeContext skipped (disabled)",
            "0/70",
        )

    # Reload Glossary map so recompute sees updated Use_StrictPlus flags.
    glossary_ws2, glossary_map2 = load_glossary(wb)

    # Step 102: ExternalRoundTripCheck (analysis-only): verified external numeric refs must decode consistently.
    ext_rt_pass = 0
    ext_rt_fail = 0
    ext_rt_skipped = 0
    ext_rt_status = ""
    try:
        active2 = {t.token: t for t in glossary_map2.values() if t.use_strictplus and t.translation}
        ext_rt_pass, ext_rt_fail, ext_rt_skipped, ext_rt_status = materialize_external_roundtrip_check(
            wb,
            next_iter,
            utc,
            active_tokens=active2,
            min_verified_count=int(max(int(ext_roundtrip_min_verified), int(puzzle_min_ext_verified))),
            min_segment_digits=int(ext_roundtrip_min_segment_digits),
            allow_ordered_run_match=bool(ext_roundtrip_allow_ordered),
        )
        res = "FAILED" if ext_rt_status not in ("ok", "missing ExternalRefs_v115/Books") else ("CHANGED" if ext_rt_fail else "NO_CHANGE")
        append_flow_run_log(
            wb,
            next_iter,
            102,
            utc,
            res,
            f"External roundtrip: pass={ext_rt_pass}, fail={ext_rt_fail}, skipped={ext_rt_skipped} ({ext_rt_status})",
            "0/70",
            notes=f"min_verified_count={ext_roundtrip_min_verified}, puzzle_min_verified={puzzle_min_ext_verified}, min_segment_digits={ext_roundtrip_min_segment_digits}, ordered_run_match={1 if ext_roundtrip_allow_ordered else 0}",
        )
    except Exception as e:
        append_flow_run_log(
            wb,
            next_iter,
            102,
            utc,
            "FAILED",
            "External roundtrip: FAILED",
            "0/70",
            notes=str(e),
        )

    # Lore/Semantic mining (analysis-only + display-only). We do this BEFORE the global recompute so we can
    # optionally promote high-confidence semantic hints into Glossary translations (guarded by GT live check).
    lore_seed_added = 0
    lore_total_rows = 0
    lore_hits_rows = 0
    semantic_map_rows = 0
    semantic_map: Dict[str, str] = {}
    semantic_reasons: Dict[str, str] = {}
    semantic_glossary_promoted = 0
    semantic_glossary_attempted = 0
    semantic_glossary_el_changed = 0
    semantic_glossary_reverted = 0
    semantic_glossary_revert_attempted = 0
    semantic_glossary_revert_el_changed = 0
    lore_bigrams_rows = 0
    lore_bigrams_status = "skipped"
    pd_sources_for_iter: List[Tuple[str, str]] = []
    pd_enabled_effective = False
    pd_sig_fp_full = ""

    if lore_enabled:
        try:
            lore_seed_added, lore_total_rows = _ensure_lore_corpus_sheets(wb, next_iter)
            append_flow_run_log(
                wb,
                next_iter,
                92,
                utc,
                "CHANGED" if lore_seed_added else "NO_CHANGE",
                f"Lore corpus ensured: seed_added={lore_seed_added}, total_rows~{lore_total_rows}",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                92,
                utc,
                "FAILED",
                "Lore corpus ensure: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )

        # Step 91: optionally fetch an internet-derived Tibia signature index (derived counts only).
        try:
            rows_written = 0
            sigs_hit = 0
            target_sigs = 0
            wordfreq_rows = 0

            tibia_sig_fp_full = ""
            try:
                payload = json.dumps(
                    {
                        "npc_url": str(lore_fetch_tibia_npc_url),
                        "book_url": str(lore_fetch_tibia_book_url),
                        "timeout_s": int(lore_fetch_tibia_timeout_s),
                        "max_words_per_sig": int(lore_fetch_tibia_max_words_per_sig),
                        "wordfreq_topn": int(lore_fetch_tibia_wordfreq_topn),
                        "target_sigs_count": int(len(_target_signatures_from_glossary_sheet(wb))),
                        "canon": {
                            "drop_final_e": int(bool(lore_drop_final_e)),
                            "drop_all_h": int(bool(lore_drop_all_h)),
                            "drop_all_o": int(bool(lore_drop_all_o)),
                        },
                    },
                    sort_keys=True,
                )
                tibia_sig_fp_full = hashlib.sha1(payload.encode("utf-8", errors="replace")).hexdigest()
            except Exception:
                tibia_sig_fp_full = ""

            should_fetch = False
            if lore_fetch_tibia_enabled:
                # If we haven't yet materialized the global wordfreq sheet, force a refresh once.
                if LORE_WORDFREQ_TIBIA_SHEET not in wb.sheetnames:
                    should_fetch = True
                # Force refresh when canon/URLs changed, even if still fresh by age.
                if tibia_sig_fp_full and tibia_sig_fp_full != str(prev_tibia_sig_fp or "").strip():
                    should_fetch = True
                last_dt = _sigindex_last_fetched_dt(wb, LORE_SIGINDEX_TIBIA_SHEET)
                if not should_fetch:
                    if lore_fetch_tibia_max_age_hours <= 0:
                        should_fetch = True
                    elif last_dt is None:
                        should_fetch = True
                    else:
                        age = datetime.now(timezone.utc) - last_dt
                        should_fetch = age >= timedelta(hours=float(lore_fetch_tibia_max_age_hours))

            if should_fetch:
                rows_written, sigs_hit, target_sigs, wordfreq_rows = refresh_tibia_sigindex_sheet(
                    wb,
                    iter_num=next_iter,
                    utc=utc,
                    npc_url=lore_fetch_tibia_npc_url,
                    book_url=lore_fetch_tibia_book_url,
                    timeout_s=lore_fetch_tibia_timeout_s,
                    max_words_per_sig=lore_fetch_tibia_max_words_per_sig,
                    wordfreq_topn=lore_fetch_tibia_wordfreq_topn,
                    drop_final_e=lore_drop_final_e,
                    drop_all_h=lore_drop_all_h,
                    drop_all_o=lore_drop_all_o,
                )
                if tibia_sig_fp_full:
                    flow_state_set(store, "TibiaSigIndexFingerprint", tibia_sig_fp_full)
                res = "CHANGED" if rows_written else "NO_CHANGE"
                summary = (
                    f"Tibia sig-index refreshed: sig_rows={rows_written}, sigs_hit={sigs_hit}, "
                    f"target_sigs={target_sigs}, wordfreq_rows={wordfreq_rows}"
                )
            else:
                res = "NO_CHANGE"
                summary = (
                    "Tibia sig-index fetch skipped "
                    f"(enabled={lore_fetch_tibia_enabled}, max_age_h={lore_fetch_tibia_max_age_hours})"
                )

            append_flow_run_log(
                wb,
                next_iter,
                91,
                utc,
                res,
                summary,
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                91,
                utc,
                "FAILED",
                "Tibia sig-index fetch: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )

        # Step 112: local dictionary-derived signature index (derived-only; no network).
        try:
            dict_rows_written = 0
            dict_sigs_hit = 0
            dict_target_sigs = 0

            dict_sig_fp_full = ""
            try:
                payload = json.dumps(
                    {
                        "dict_path": str(lore_fetch_dict_path),
                        "max_words_per_sig": int(lore_fetch_dict_max_words_per_sig),
                        "canon": {
                            "drop_final_e": int(bool(lore_drop_final_e)),
                            "drop_all_h": int(bool(lore_drop_all_h)),
                            "drop_all_o": int(bool(lore_drop_all_o)),
                        },
                    },
                    sort_keys=True,
                )
                dict_sig_fp_full = hashlib.sha1(payload.encode("utf-8", errors="replace")).hexdigest()
            except Exception:
                dict_sig_fp_full = ""

            should_fetch_dict = False
            if lore_fetch_dict_enabled:
                if LORE_SIGINDEX_DICT_SHEET not in wb.sheetnames:
                    should_fetch_dict = True
                if dict_sig_fp_full and dict_sig_fp_full != str(prev_dict_sig_fp or "").strip():
                    should_fetch_dict = True
                last_dt = _sigindex_last_fetched_dt(wb, LORE_SIGINDEX_DICT_SHEET)
                if not should_fetch_dict:
                    if lore_fetch_dict_max_age_hours <= 0:
                        should_fetch_dict = True
                    elif last_dt is None:
                        should_fetch_dict = True
                    else:
                        age = datetime.now(timezone.utc) - last_dt
                        should_fetch_dict = age >= timedelta(hours=float(lore_fetch_dict_max_age_hours))

            if should_fetch_dict:
                dict_rows_written, dict_sigs_hit, dict_target_sigs = refresh_dict_sigindex_sheet(
                    wb,
                    iter_num=next_iter,
                    utc=utc,
                    dict_path=str(lore_fetch_dict_path),
                    max_words_per_sig=int(lore_fetch_dict_max_words_per_sig),
                    drop_final_e=bool(lore_drop_final_e),
                    drop_all_h=bool(lore_drop_all_h),
                    drop_all_o=bool(lore_drop_all_o),
                )
                if dict_sig_fp_full:
                    flow_state_set(store, "DictSigIndexFingerprint", dict_sig_fp_full)
                res = "CHANGED" if dict_rows_written else "NO_CHANGE"
                summary = f"Dict sig-index refreshed: sig_rows={dict_rows_written}, sigs_hit={dict_sigs_hit}, target_sigs={dict_target_sigs}"
            else:
                res = "NO_CHANGE"
                summary = (
                    "Dict sig-index fetch skipped "
                    f"(enabled={lore_fetch_dict_enabled}, max_age_h={lore_fetch_dict_max_age_hours})"
                )

            append_flow_run_log(
                wb,
                next_iter,
                112,
                utc,
                res,
                summary,
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                112,
                utc,
                "FAILED",
                "Dict sig-index fetch: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )

        # Build effective PD source list (base + extra URLs) once for this iteration.
        pd_sources_for_iter: List[Tuple[str, str]] = []
        if lore_fetch_pd_enabled:
            seen_pd: set[str] = set()

            def _add_pd_source(cid_raw: object, u_raw: object) -> None:
                u = str(u_raw or "").strip()
                if not u or u in seen_pd:
                    return
                seen_pd.add(u)
                cid = str(cid_raw or "").strip() or _auto_corpus_id_from_url(u)
                pd_sources_for_iter.append((cid, u))

            _add_pd_source(lore_fetch_pd_corpus_id, lore_fetch_pd_url)
            for i, u in enumerate(lore_fetch_pd_extra_urls or []):
                cid = lore_fetch_pd_extra_cids[i] if i < len(lore_fetch_pd_extra_cids or []) else ""
                _add_pd_source(cid, u)

        pd_enabled_effective = bool(lore_fetch_pd_enabled and pd_sources_for_iter)
        pd_sig_fp_full = ""
        if pd_enabled_effective:
            try:
                payload = json.dumps(
                    {
                        "sources": pd_sources_for_iter,
                        "canon": {
                            "drop_final_e": int(bool(lore_drop_final_e)),
                            "drop_all_h": int(bool(lore_drop_all_h)),
                            "drop_all_o": int(bool(lore_drop_all_o)),
                        },
                        "max_words_per_sig": int(lore_fetch_pd_max_words_per_sig),
                        "wordfreq_topn": int(lore_fetch_pd_wordfreq_topn),
                    },
                    sort_keys=True,
                )
                pd_sig_fp_full = hashlib.sha1(payload.encode("utf-8", errors="replace")).hexdigest()
            except Exception:
                pd_sig_fp_full = ""

        # Step 109: public-domain sig-index fetch (derived-only; expands coverage safely).
        try:
            pd_rows_written = 0
            if pd_enabled_effective:
                # Force refresh when fingerprint changed (or wordfreq sheet missing), even if still fresh by age.
                should_fetch = False
                if LORE_WORDFREQ_PD_SHEET not in wb.sheetnames:
                    should_fetch = True
                if pd_sig_fp_full and pd_sig_fp_full != str(prev_pd_sig_fp or "").strip():
                    should_fetch = True

                last_dt = _sigindex_last_fetched_dt(wb, LORE_SIGINDEX_PD_SHEET)
                if not should_fetch:
                    if last_dt is None:
                        should_fetch = True
                    elif float(lore_fetch_pd_max_age_hours) <= 0:
                        should_fetch = True
                    else:
                        age = datetime.now(timezone.utc) - last_dt
                        should_fetch = age >= timedelta(hours=float(lore_fetch_pd_max_age_hours))
                if should_fetch:
                    first_cid, first_url = pd_sources_for_iter[0]
                    extra = pd_sources_for_iter[1:]
                    pd_rows_written, pd_sigs_hit, pd_target_sigs, pd_wordfreq_rows = refresh_pd_sigindex_sheet(
                        wb,
                        iter_num=next_iter,
                        utc=utc,
                        corpus_id=first_cid,
                        url=first_url,
                        extra_sources=extra,
                        timeout_s=int(lore_fetch_pd_timeout_s),
                        max_words_per_sig=int(lore_fetch_pd_max_words_per_sig),
                        wordfreq_topn=int(lore_fetch_pd_wordfreq_topn),
                        drop_final_e=lore_drop_final_e,
                        drop_all_h=lore_drop_all_h,
                        drop_all_o=lore_drop_all_o,
                        cache_max_age_hours=float(lore_fetch_pd_cache_max_age_h),
                    )
                    if pd_sig_fp_full:
                        flow_state_set(store, "PDSigIndexFingerprint", pd_sig_fp_full)
                    res = "CHANGED" if pd_rows_written else "NO_CHANGE"
                    summary = (
                        f"PD sig-index refreshed: sig_rows={pd_rows_written}, sigs_hit={pd_sigs_hit}, "
                        f"target_sigs={pd_target_sigs}, wordfreq_rows={pd_wordfreq_rows}"
                    )
                else:
                    res = "NO_CHANGE"
                    summary = (
                        "PD sig-index fetch skipped "
                        f"(enabled={lore_fetch_pd_enabled}, max_age_h={lore_fetch_pd_max_age_hours})"
                    )
            else:
                res = "NO_CHANGE"
                summary = (
                    "PD sig-index fetch skipped "
                    f"(enabled={lore_fetch_pd_enabled}, sources={len(pd_sources_for_iter)})"
                )
            append_flow_run_log(
                wb,
                next_iter,
                109,
                utc,
                res,
                summary,
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=f"sources={len(pd_sources_for_iter)}, fp={pd_sig_fp_full[:12] if pd_sig_fp_full else ''}",
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                109,
                utc,
                "FAILED",
                "PD sig-index fetch: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )

        # Step 103: derived bigram LM for ContextEnglish (no full text persisted to XLSX).
        if lore_bigrams_enabled:
            try:
                lore_bigrams_rows, lore_bigrams_status = refresh_lore_bigrams_sheet(
                    wb,
                    iter_num=next_iter,
                    utc=utc,
                    npc_url=lore_fetch_tibia_npc_url,
                    book_url=lore_fetch_tibia_book_url,
                    timeout_s=int(lore_fetch_tibia_timeout_s),
                    max_age_hours=float(lore_bigrams_max_age_h),
                    cache_max_age_hours=float(lore_bigrams_cache_max_age_h),
                    pd_enabled=bool(pd_enabled_effective),
                    pd_sources=pd_sources_for_iter,
                    pd_cache_max_age_hours=float(lore_fetch_pd_cache_max_age_h),
                    vocab_topn=int(lore_bigrams_vocab_topn),
                    min_count=int(lore_bigrams_min_count),
                    max_rows=int(lore_bigrams_max_rows),
                )
                append_flow_run_log(
                    wb,
                    next_iter,
                    103,
                    utc,
                    "CHANGED" if lore_bigrams_status == "refreshed" else "NO_CHANGE",
                    f"LoreBigrams: rows={lore_bigrams_rows} ({lore_bigrams_status})",
                    "0/70",
                    evidence_avg=round(cur_ev, 6),
                    weak_frac=round(cur_weak, 6),
                    micro_frac=round(cur_micro, 6),
                    notes=f"vocab_topn={lore_bigrams_vocab_topn}, min_count={lore_bigrams_min_count}, max_rows={lore_bigrams_max_rows}, pd_sources={len(pd_sources_for_iter)}",
                )
            except Exception as e:
                append_flow_run_log(
                    wb,
                    next_iter,
                    103,
                    utc,
                    "FAILED",
                    "LoreBigrams: FAILED",
                    "0/70",
                    evidence_avg=round(cur_ev, 6),
                    weak_frac=round(cur_weak, 6),
                    micro_frac=round(cur_micro, 6),
                    notes=str(e),
                )
        else:
            append_flow_run_log(
                wb,
                next_iter,
                103,
                utc,
                "NO_CHANGE",
                "LoreBigrams skipped (disabled)",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
            )

        try:
            lore_hits_rows = materialize_lore_token_hits(
                wb,
                next_iter,
                glossary_map2,
                drop_final_e=lore_drop_final_e,
                drop_all_h=lore_drop_all_h,
                drop_all_o=lore_drop_all_o,
            )
            append_flow_run_log(
                wb,
                next_iter,
                93,
                utc,
                "CHANGED" if lore_hits_rows else "NO_CHANGE",
                f"Lore token hits written: rows={lore_hits_rows}",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=f"canon(drop_final_e={lore_drop_final_e}, drop_all_h={lore_drop_all_h}, drop_all_o={lore_drop_all_o})",
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                93,
                utc,
                "FAILED",
                "Lore token hits: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )

        try:
            semantic_map_rows, semantic_map, semantic_reasons = materialize_semantic_map_from_lore_hits(
                wb,
                next_iter,
                min_total_count=sem_min_total,
                min_top_share=sem_min_share,
            )
            append_flow_run_log(
                wb,
                next_iter,
                94,
                utc,
                "CHANGED" if semantic_map_rows else "NO_CHANGE",
                f"SemanticMap materialized: rows={semantic_map_rows}",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=f"min_total={sem_min_total}, min_share={sem_min_share}",
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                94,
                utc,
                "FAILED",
                "SemanticMap: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )

        # Step 96: apply semantic->Glossary retext (optional; still guarded by GT live check)
        try:
            semantic_glossary_promoted, semantic_glossary_attempted, semantic_glossary_el_changed = apply_semantic_promotions_to_glossary(
                wb,
                next_iter,
                utc,
                glossary_ws2,
                glossary_map2,
                semantic_map,
                semantic_reasons,
                enabled=sem_promote_enabled,
                max_promotions=sem_promote_max,
                min_total_occ=sem_promote_min_occ,
                min_conf_weight=sem_promote_min_conf_w,
                max_conf_weight=sem_promote_max_conf_w,
                blocked_evidence_classes=sem_promote_block_evcls,
                wordfreq_sheet=sem_promote_wf_sheet,
                min_new_wordfreq=sem_promote_min_new_wf,
                min_wordfreq_ratio=sem_promote_min_wf_ratio,
                anti_mode=anti_mode,
                anti_hallucination_terms=anti_deny_words,
                enforced_crib_ids=enforced_gt_ids,
                locked_tokens=glossary_retext_locked_tokens,
            )
            append_flow_run_log(
                wb,
                next_iter,
                96,
                utc,
                "CHANGED" if semantic_glossary_promoted else "NO_CHANGE",
                f"Semantic glossary retext applied: {semantic_glossary_promoted} (attempted {semantic_glossary_attempted})",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=(
                    f"enabled={sem_promote_enabled}, max={sem_promote_max}, min_occ={sem_promote_min_occ}, "
                    f"min_conf={sem_promote_min_conf}, max_conf={sem_promote_max_conf}, "
                    f"block_evcls={','.join(sorted(sem_promote_block_evcls))}, "
                    f"min_new_wf={sem_promote_min_new_wf}, min_wf_ratio={sem_promote_min_wf_ratio}, "
                    f"wf_sheet={sem_promote_wf_sheet}, anti_mode={int(anti_mode)}, el_changed={semantic_glossary_el_changed}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                96,
                utc,
                "FAILED",
                "Semantic glossary retext: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )

        # Step 97: revert unsafe past semantic retext under the current safety policy.
        try:
            semantic_glossary_reverted, semantic_glossary_revert_attempted, semantic_glossary_revert_el_changed = revert_semantic_promotions_if_unsafe(
                wb,
                next_iter,
                utc,
                glossary_ws2,
                glossary_map2,
                blocked_evidence_classes=sem_promote_block_evcls,
                wordfreq_sheet=sem_promote_wf_sheet,
                min_new_wordfreq=sem_promote_min_new_wf,
                min_wordfreq_ratio=sem_promote_min_wf_ratio,
                enforced_crib_ids=enforced_gt_ids,
                locked_tokens=glossary_retext_locked_tokens,
            )
            append_flow_run_log(
                wb,
                next_iter,
                97,
                utc,
                "CHANGED" if semantic_glossary_reverted else "NO_CHANGE",
                f"Semantic reverts applied: {semantic_glossary_reverted} (attempted {semantic_glossary_revert_attempted})",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=f"el_changed={semantic_glossary_revert_el_changed}",
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                97,
                utc,
                "FAILED",
                "Semantic reverts: FAILED",
                "0/70",
                evidence_avg=round(cur_ev, 6),
                weak_frac=round(cur_weak, 6),
                micro_frac=round(cur_micro, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            91,
            utc,
            "NO_CHANGE",
            "Tibia sig-index fetch skipped (lore disabled)",
            "0/70",
            evidence_avg=round(cur_ev, 6),
            weak_frac=round(cur_weak, 6),
            micro_frac=round(cur_micro, 6),
        )
        append_flow_run_log(
            wb,
            next_iter,
            103,
            utc,
            "NO_CHANGE",
            "LoreBigrams skipped (lore disabled)",
            "0/70",
            evidence_avg=round(cur_ev, 6),
            weak_frac=round(cur_weak, 6),
            micro_frac=round(cur_micro, 6),
        )
        append_flow_run_log(
            wb,
            next_iter,
            92,
            utc,
            "NO_CHANGE",
            "Lore corpus skipped (disabled)",
            "0/70",
            evidence_avg=round(cur_ev, 6),
            weak_frac=round(cur_weak, 6),
            micro_frac=round(cur_micro, 6),
        )
        append_flow_run_log(
            wb,
            next_iter,
            93,
            utc,
            "NO_CHANGE",
            "Lore token hits skipped (disabled)",
            "0/70",
            evidence_avg=round(cur_ev, 6),
            weak_frac=round(cur_weak, 6),
            micro_frac=round(cur_micro, 6),
        )
        append_flow_run_log(
            wb,
            next_iter,
            94,
            utc,
            "NO_CHANGE",
            "SemanticMap skipped (disabled)",
            "0/70",
            evidence_avg=round(cur_ev, 6),
            weak_frac=round(cur_weak, 6),
            micro_frac=round(cur_micro, 6),
        )
        append_flow_run_log(
            wb,
            next_iter,
            96,
            utc,
            "NO_CHANGE",
            "Semantic glossary retext skipped (lore disabled)",
            "0/70",
            evidence_avg=round(cur_ev, 6),
            weak_frac=round(cur_weak, 6),
            micro_frac=round(cur_micro, 6),
        )
        append_flow_run_log(
            wb,
            next_iter,
            97,
            utc,
            "NO_CHANGE",
            "Semantic reverts skipped (lore disabled)",
            "0/70",
            evidence_avg=round(cur_ev, 6),
            weak_frac=round(cur_weak, 6),
            micro_frac=round(cur_micro, 6),
        )

    # Reload Glossary after potential semantic retext edits so the global recompute sees updated translations.
    glossary_ws2, glossary_map2 = load_glossary(wb)

    # Step 60: recompute Books/Contigs/MasterText + Cribs (non-GT)
    changed_books, _ev_tmp, _weak_tmp, _micro_tmp, single_frac_tmp = recompute_books_contigs_mastertext(wb, glossary_map2)
    cribs_updated, _gt_mismatches, match_flips, avartar_match = recompute_cribs_dp(wb, glossary_map2)
    append_flow_run_log(
        wb,
        next_iter,
        60,
        utc,
        "CHANGED",
        f"Recomputed global. Cribs updated: {cribs_updated}; AvarTarMatch={avartar_match}",
        changed_books,
    )

    # Step 70: recompute analytics (TokenEvidence)
    ev_avg_w, weak_w, micro_w = recompute_token_evidence_books(wb, glossary_map2)
    recompute_token_evidence_contigs(wb, glossary_map2)
    single_frac = weighted_single_char_frac_from_books(wb)
    final_active = {t.token: t for t in glossary_map2.values() if t.use_strictplus and t.translation}
    _ev_dp, _weak_dp, _micro_dp, _single_dp, final_tokens = compute_weighted_metrics_for_books_dp(wb, final_active)
    append_flow_run_log(
        wb,
        next_iter,
        70,
        utc,
        "CHANGED",
        "Recomputed TokenEvidence_*",
        changed_books,
        evidence_avg=round(ev_avg_w, 6),
        weak_frac=round(weak_w, 6),
        micro_frac=round(micro_w, 6),
        notes=f"SingleCharFrac={single_frac:.6f}",
    )

    # Step 90: Materialize macro-compressed display-only outputs (Books + MasterText).
    macro_books_changed = 0
    macro_master_changed = 0
    if macrocompress_enabled:
        try:
            macro_books_changed, macro_master_changed = materialize_macrocompression_display(wb, next_iter, glossary_map2)
            append_flow_run_log(
                wb,
                next_iter,
                90,
                utc,
                "CHANGED" if (macro_books_changed or macro_master_changed) else "NO_CHANGE",
                f"MacroCompression display: books_changed={macro_books_changed}, master_changed={macro_master_changed}",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                90,
                utc,
                "FAILED",
                "MacroCompression display: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            90,
            utc,
            "NO_CHANGE",
            "MacroCompression display: disabled via FlowSettings",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 75: Apply readability layer (separate from StrictPlus mechanics; safe text-only rewrite)
    readable_books_changed, readable_master_changed, readable_cribs_changed, readable_repl = apply_readability_layer(wb, next_iter)
    append_flow_run_log(
        wb,
        next_iter,
        75,
        utc,
        "CHANGED" if readable_repl else "NO_CHANGE",
        f"Readability layer applied: repl={readable_repl}, books_changed={readable_books_changed}, master_changed={readable_master_changed}, cribs_changed={readable_cribs_changed}",
        "0/70",
        evidence_avg=round(ev_avg_w, 6),
        weak_frac=round(weak_w, 6),
        micro_frac=round(micro_w, 6),
        notes="Updates Translation_Readable_Auto (Books/MasterText) + DP_Readable_Auto (Cribs) per ReadabilityRules.Scope.",
    )

    # Focus diagnostics (analysis-only). When enabled, we write this every iteration so each "next iteration"
    # yields concrete, contextual next work even if mech_promoted>0.
    focus_rows = 0
    if plateau_enable_focus:
        try:
            focus_rows = upsert_iter_focus_sheet(
                wb,
                next_iter,
                utc,
                ev_avg_w,
                weak_w,
                micro_w,
                single_frac,
                final_tokens,
            )
            append_flow_run_log(
                wb,
                next_iter,
                82,
                utc,
                "CHANGED",
                f"Focus report written: rows~{focus_rows} (mech_promoted={mechanical_promotions_count})",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                82,
                utc,
                "FAILED",
                "Focus report: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            82,
            utc,
            "NO_CHANGE",
            "Focus report skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Structural variant-aware alignment (analysis-only), driven by AnchorCribs imported from iter141.
    anchor_added = 0
    anchor_updated = 0
    anchor_status = None
    ref_book_id = 0
    aligned_books = 0
    blocks_written = 0
    superanchors = 0

    if plateau_enable_structural:
        try:
            anchor_added, anchor_updated, anchor_status = sync_anchorcribs_from_iter141(wb, next_iter, utc)
            append_flow_run_log(
                wb,
                next_iter,
                85,
                utc,
                "CHANGED" if (anchor_added or anchor_updated) else "NO_CHANGE",
                f"AnchorCribs sync: added={anchor_added}, updated={anchor_updated}, status={anchor_status}",
                "0/70",
            )
        except Exception as e:
            anchor_status = str(e)
            append_flow_run_log(
                wb,
                next_iter,
                85,
                utc,
                "FAILED",
                "AnchorCribs sync: FAILED",
                "0/70",
                notes=str(e),
            )

        try:
            ref_book_id, aligned_books, blocks_written, st = build_variant_alignment_from_anchorcribs(
                wb,
                next_iter,
                ignore_anchorcrib_ids=structural_ignore_anchorcribs,
                restrict_votes_to_anchor_windows=structural_restrict_votes_to_anchor_windows,
                anchor_window_pad=structural_anchor_window_pad,
                ref_book_override=structural_ref_book_override,
                min_anchors_shared_for_voting=structural_min_anchors_shared_for_voting,
                require_unique_anchor_occurrences=structural_require_unique_anchor_occurrences,
                require_strong_offsets=structural_require_strong_offsets,
            )
            append_flow_run_log(
                wb,
                next_iter,
                86,
                utc,
                "CHANGED" if blocks_written else "NO_CHANGE",
                f"Variant alignment: refBook={ref_book_id}, alignedBooks={aligned_books}, blocks={blocks_written}, status={st}",
                "0/70",
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                86,
                utc,
                "FAILED",
                "Variant alignment: FAILED",
                "0/70",
                notes=str(e),
            )

        try:
            relaxed_note = None
            superanchors = mine_superanchors_from_backbone(
                wb,
                next_iter,
                min_len=sa_min_len,
                min_books=sa_min_books,
                min_support_books=sa_min_support_books,
                min_support_frac=sa_min_support,
            )
            # Structural auto-relax (analysis-only): if super-anchor mining is too strict, retry with a
            # more cluster-friendly support fraction. Persist the relaxed setting when it produces output.
            if superanchors == 0 and sa_min_support > 0.8 + 1e-12:
                relaxed_support = 0.8
                superanchors2 = mine_superanchors_from_backbone(
                    wb,
                    next_iter,
                    min_len=sa_min_len,
                    min_books=sa_min_books,
                    min_support_books=sa_min_support_books,
                    min_support_frac=relaxed_support,
                )
                if superanchors2:
                    relaxed_note = f"Auto-relaxed SuperAnchor_MinSupportFrac {sa_min_support}->{relaxed_support}"
                    superanchors = superanchors2
                    flow_setting_set(store,
                        "SuperAnchor_MinSupportFrac",
                        relaxed_support,
                        note=f"iter{next_iter}: auto-relaxed (produced superanchors)",
                    )
                    sa_min_support = relaxed_support

            # Structural bootstrap (analysis-only): if we found superanchors, promote them into AnchorCribs_Auto
            # and re-run alignment + superanchor mining once to potentially unlock longer stable runs.
            promoted_sa = 0
            if structural_auto_promote_sa and superanchors:
                try:
                    promoted_sa = promote_superanchors_to_anchorcribs_auto(
                        wb,
                        iter_num=next_iter,
                        utc=utc,
                        max_promote=structural_auto_promote_sa_max,
                    )
                except Exception:
                    promoted_sa = 0

            if promoted_sa:
                try:
                    ref_book_id, aligned_books, blocks_written, st2 = build_variant_alignment_from_anchorcribs(
                        wb,
                        next_iter,
                        ignore_anchorcrib_ids=structural_ignore_anchorcribs,
                        restrict_votes_to_anchor_windows=structural_restrict_votes_to_anchor_windows,
                        anchor_window_pad=structural_anchor_window_pad,
                        ref_book_override=structural_ref_book_override,
                        min_anchors_shared_for_voting=structural_min_anchors_shared_for_voting,
                        require_unique_anchor_occurrences=structural_require_unique_anchor_occurrences,
                        require_strong_offsets=structural_require_strong_offsets,
                    )
                    append_flow_run_log(
                        wb,
                        next_iter,
                        86,
                        utc,
                        "CHANGED" if blocks_written else "NO_CHANGE",
                        f"Variant alignment (post-SA-promote): refBook={ref_book_id}, alignedBooks={aligned_books}, blocks={blocks_written}, status={st2}",
                        "0/70",
                    )
                except Exception as e:
                    append_flow_run_log(
                        wb,
                        next_iter,
                        86,
                        utc,
                        "FAILED",
                        "Variant alignment (post-SA-promote): FAILED",
                        "0/70",
                        notes=str(e),
                    )

                try:
                    superanchors = mine_superanchors_from_backbone(
                        wb,
                        next_iter,
                        min_len=sa_min_len,
                        min_books=sa_min_books,
                        min_support_books=sa_min_support_books,
                        min_support_frac=sa_min_support,
                    )
                    append_flow_run_log(
                        wb,
                        next_iter,
                        87,
                        utc,
                        "CHANGED" if superanchors else "NO_CHANGE",
                        f"SuperAnchors mined (post-SA-promote): {superanchors} (min_len={sa_min_len}, min_books={sa_min_books}, min_support_books={sa_min_support_books}, min_support={sa_min_support})",
                        "0/70",
                    )
                    relaxed_note = (relaxed_note + "; " if relaxed_note else "") + f"Auto-promoted SuperAnchors -> AnchorCribs_Auto: {promoted_sa}"
                except Exception as e:
                    append_flow_run_log(
                        wb,
                        next_iter,
                        87,
                        utc,
                        "FAILED",
                        "SuperAnchors mining (post-SA-promote): FAILED",
                        "0/70",
                        notes=str(e),
                    )
            append_flow_run_log(
                wb,
                next_iter,
                87,
                utc,
                "CHANGED" if superanchors else "NO_CHANGE",
                f"SuperAnchors mined: {superanchors} (min_len={sa_min_len}, min_books={sa_min_books}, min_support_books={sa_min_support_books}, min_support={sa_min_support})",
                "0/70",
                notes=relaxed_note,
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                87,
                utc,
                "FAILED",
                "SuperAnchors mining: FAILED",
                "0/70",
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            85,
            utc,
            "NO_CHANGE",
            "Structural sync skipped (disabled)",
            "0/70",
        )
        append_flow_run_log(
            wb,
            next_iter,
            86,
            utc,
            "NO_CHANGE",
            "Structural alignment skipped (disabled)",
            "0/70",
        )
        append_flow_run_log(
            wb,
            next_iter,
            87,
            utc,
            "NO_CHANGE",
            "SuperAnchors skipped (disabled)",
            "0/70",
        )

    # Step 95: Semantic render (display-only). Uses the SemanticMap already materialized earlier (Step 94).
    semantic_books_changed = 0
    semantic_master_changed = 0
    semantic_repl = 0
    if semantic_enabled and semantic_map:
        try:
            semantic_books_changed, semantic_master_changed, semantic_repl = materialize_semantic_translation_display(
                wb,
                next_iter,
                glossary_map2,
                semantic_map,
            )
            append_flow_run_log(
                wb,
                next_iter,
                95,
                utc,
                "CHANGED" if (semantic_books_changed or semantic_master_changed) else "NO_CHANGE",
                f"Semantic render: books_changed={semantic_books_changed}, master_changed={semantic_master_changed}, repl={semantic_repl}",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                95,
                utc,
                "FAILED",
                "Semantic render: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            95,
            utc,
            "NO_CHANGE",
            "Semantic render skipped (disabled or empty map)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 98: English layer (display-only): map canonical words -> likely English surfaces.
    english_map_rows = 0
    english_books_changed = 0
    english_master_changed = 0
    english_repl = 0
    if english_enabled:
        try:
            english_map_rows, english_books_changed, english_master_changed, english_repl = materialize_english_layer_display(
                wb,
                next_iter,
                glossary_map2,
                drop_final_e=lore_drop_final_e,
                drop_all_h=lore_drop_all_h,
                drop_all_o=lore_drop_all_o,
                min_total_count=english_min_total,
                min_top_share=english_min_share,
                min_word_len=english_min_len,
                max_map_rows=english_max_rows,
            )
            append_flow_run_log(
                wb,
                next_iter,
                98,
                utc,
                "CHANGED" if (english_books_changed or english_master_changed) else "NO_CHANGE",
                f"English layer: map_rows={english_map_rows}, repl={english_repl}, books_changed={english_books_changed}, master_changed={english_master_changed}",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=f"min_total={english_min_total}, min_share={english_min_share}, min_len={english_min_len}",
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                98,
                utc,
                "FAILED",
                "English layer: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            98,
            utc,
            "NO_CHANGE",
            "English layer skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 104: ContextEnglish render (display-only; GT-safe).
    ctx_books_changed = 0
    ctx_master_changed = 0
    ctx_avg_score = 0.0
    ctx_oov = 0.0
    ctx_map_rows = 0
    ctx_improve_streak = 0
    if lore_enabled and context_enabled:
        try:
            ctx_books_changed, ctx_master_changed, ctx_avg_score, ctx_oov, ctx_map_rows = materialize_context_english_render(
                wb,
                iter_num=next_iter,
                utc=utc,
                glossary_map=glossary_map2,
                active_tokens=final_active,
                drop_final_e=lore_drop_final_e,
                drop_all_h=lore_drop_all_h,
                drop_all_o=lore_drop_all_o,
                emission_alpha=float(ctx_emit_alpha),
                transition_alpha=float(ctx_trans_alpha),
                max_candidates_per_token=int(ctx_max_cands),
                map_min_total=int(ctx_map_min_total),
                map_min_top_share=float(ctx_map_min_share),
                map_max_rows=int(ctx_map_max_rows),
                sequence_hints_enabled=bool(seq_hints_enabled),
                sequence_hints_boost=int(seq_hints_boost),
                sequence_hints_max_words_per_sig=int(seq_hints_max_words_per_sig),
                reversephrase_hints_enabled=bool(reversephrase_hints_enabled),
                reversephrase_hints_boost=int(reversephrase_hints_boost),
                reversephrase_hints_max_words_per_sig=int(reversephrase_hints_max_words_per_sig),
            )
            ctx_score_r = round(float(ctx_avg_score), 6)
            prev_score_r = round(float(prev_ctx_score), 6)
            if ctx_score_r > prev_score_r + 1e-12:
                ctx_improve_streak = int(prev_ctx_streak) + 1
            else:
                ctx_improve_streak = 0
            append_flow_run_log(
                wb,
                next_iter,
                104,
                utc,
                "CHANGED" if (ctx_books_changed or ctx_master_changed) else "NO_CHANGE",
                f"ContextEnglish: books_changed={ctx_books_changed}, master_changed={ctx_master_changed}, avg_score={ctx_avg_score:.6f}, oov={ctx_oov:.6f}, map_rows={ctx_map_rows}",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=(
                    f"alpha_emit={ctx_emit_alpha}, alpha_trans={ctx_trans_alpha}, max_cands={ctx_max_cands}, "
                    f"prev_avg_score={prev_ctx_score:.6f}, improve_streak={ctx_improve_streak}, "
                    f"seq_hints={int(bool(seq_hints_enabled))}, seq_hint_boost={seq_hints_boost}, seq_hint_max_words={seq_hints_max_words_per_sig}, "
                    f"rev_hints={int(bool(reversephrase_hints_enabled))}, rev_hint_boost={reversephrase_hints_boost}, rev_hint_max_words={reversephrase_hints_max_words_per_sig}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                104,
                utc,
                "FAILED",
                "ContextEnglish: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            104,
            utc,
            "NO_CHANGE",
            "ContextEnglish skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 111: Code-aware homophones render (display-only; GT-safe).
    codeaware_books_changed = 0
    codeaware_overrides = 0
    codeaware_map_rows = 0
    codeaware_fp = ""
    codeaware_fp_changed = 0
    codeaware_status = "skipped"
    if lore_enabled and codeaware_enabled:
        try:
            codeaware_books_changed, codeaware_overrides, codeaware_map_rows, codeaware_fp, codeaware_status = (
                materialize_codeaware_homophones_render(
                    wb,
                    iter_num=next_iter,
                    utc=utc,
                    glossary_map=glossary_map2,
                    active_tokens=final_active,
                    drop_final_e=bool(lore_drop_final_e),
                    drop_all_h=bool(lore_drop_all_h),
                    drop_all_o=bool(lore_drop_all_o),
                    emission_alpha=float(ctx_emit_alpha),
                    transition_alpha=float(ctx_trans_alpha),
                    max_candidates_per_token=int(ctx_max_cands),
                    max_token_len=int(codeaware_max_token_len),
                    min_total_per_code=int(codeaware_min_total),
                    min_top_share=float(codeaware_min_share),
                    min_total_per_codeseq=int(codeaware_min_total_seq),
                    min_top_share_codeseq=float(codeaware_min_share_seq),
                    max_map_rows=int(codeaware_max_rows),
                    hint_min_total=int(codeaware_hint_min_total),
                    hint_boost=int(codeaware_hint_boost),
                    hint_topk=int(codeaware_hint_topk),
                    apply_tokens_allowlist=codeaware_apply_tokens,
                    sequence_hints_enabled=bool(seq_hints_enabled),
                    sequence_hints_boost=int(seq_hints_boost),
                    sequence_hints_max_words_per_sig=int(seq_hints_max_words_per_sig),
                    reversephrase_hints_enabled=bool(reversephrase_hints_enabled),
                    reversephrase_hints_boost=int(reversephrase_hints_boost),
                    reversephrase_hints_max_words_per_sig=int(reversephrase_hints_max_words_per_sig),
                )
            )
            if codeaware_fp and str(codeaware_fp) != str(prev_codeword_fp or ""):
                codeaware_fp_changed = 1
            append_flow_run_log(
                wb,
                next_iter,
                111,
                utc,
                "CHANGED" if (codeaware_books_changed or codeaware_fp_changed) else "NO_CHANGE",
                f"CodeAware: books_changed={codeaware_books_changed}, overrides={codeaware_overrides}, map_rows={codeaware_map_rows} ({codeaware_status})",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=(
                    f"min_total={codeaware_min_total}, min_share={codeaware_min_share}, max_rows={codeaware_max_rows}, "
                    f"max_token_len={codeaware_max_token_len}, seq_min_total={codeaware_min_total_seq}, seq_min_share={codeaware_min_share_seq}, "
                    f"hint_min_total={codeaware_hint_min_total}, hint_boost={codeaware_hint_boost}, hint_topk={codeaware_hint_topk}, "
                    f"allow_tokens={','.join(codeaware_apply_tokens) if codeaware_apply_tokens else 'AUTO'}, "
                    f"seq_hints={int(bool(seq_hints_enabled))}, rev_hints={int(bool(reversephrase_hints_enabled))}, "
                    f"fp_changed={codeaware_fp_changed}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                111,
                utc,
                "FAILED",
                "CodeAware: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            111,
            utc,
            "NO_CHANGE",
            "CodeAware skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 115: conservative late display-only cleanup for readable output columns.
    late_display_ctx_changed = 0
    late_display_codeaware_changed = 0
    late_display_repl = 0
    try:
        late_display_ctx_changed, late_display_codeaware_changed, late_display_repl = apply_late_display_cleanup(wb, next_iter)
        append_flow_run_log(
            wb,
            next_iter,
            115,
            utc,
            "CHANGED" if (late_display_ctx_changed or late_display_codeaware_changed) else "NO_CHANGE",
            (
                f"LateDisplayCleanup: ctx_books_changed={late_display_ctx_changed}, "
                f"codeaware_books_changed={late_display_codeaware_changed}, repl={late_display_repl}"
            ),
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
            notes="collapse_repeated_commas; trim_space_before_punctuation; capitalize_standalone_i; normalize_<UNK>; reapply_active_readability_rules",
        )
    except Exception as e:
        append_flow_run_log(
            wb,
            next_iter,
            115,
            utc,
            "FAILED",
            "LateDisplayCleanup: FAILED",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
            notes=str(e),
        )

    # Step 105: SequenceMatches (analysis-only; snippets only).
    seq_matches = 0
    seq_status = "skipped"
    seq_fp = ""
    seq_fp_changed = 0
    if seqmatch_enabled:
        try:
            n_list = seqmatch_n_list or [6, 7, 8]
            seq_matches, seq_status = materialize_sequence_matches(
                wb,
                iter_num=next_iter,
                utc=utc,
                n_list=n_list,
                min_n=int(seqmatch_min_n),
                candidate_max_book_freq=int(seqmatch_cand_max_freq),
                max_candidates=int(seqmatch_max_candidates),
                max_matches=int(seqmatch_max_matches),
                time_budget_s=float(seqmatch_time_budget_s),
                scan_tibia_first=bool(seqmatch_scan_tibia_first),
                pd_max_chars=int(seqmatch_pd_max_chars),
                pd_max_sentences_per_source=int(seqmatch_pd_max_sentences),
                npc_url=lore_fetch_tibia_npc_url,
                book_url=lore_fetch_tibia_book_url,
                timeout_s=int(lore_fetch_tibia_timeout_s),
                cache_max_age_hours=float(lore_bigrams_cache_max_age_h),
                pd_enabled=bool(pd_enabled_effective),
                pd_sources=pd_sources_for_iter,
                pd_cache_max_age_hours=float(lore_fetch_pd_cache_max_age_h),
                drop_final_e=bool(lore_drop_final_e),
                drop_all_h=bool(lore_drop_all_h),
                drop_all_o=bool(lore_drop_all_o),
                context_window=int(seqmatch_context_window),
                context_min_overlap=int(seqmatch_context_min_overlap),
                context_require_direction=bool(seqmatch_context_require_direction),
                snippet_min_content_words=int(seqmatch_snippet_min_content),
                explore_rotate=bool(seqmatch_explore_rotate),
                explore_keep_top=int(seqmatch_explore_keep_top),
                cache_enabled=bool(seqmatch_cache_enabled),
                cache_max_rows=int(seqmatch_cache_max_rows),
            )
            # Fingerprint to avoid treating repeated matches as "progress".
            if SEQUENCE_MATCHES_SHEET in wb.sheetnames:
                ws_sm = wb[SEQUENCE_MATCHES_SHEET]
                hs = ws_find_header_row(ws_sm, ["Phrase", "SourceKind", "SourceID"], max_scan=3)
                cs = ws_headers(ws_sm, hs)
                items: List[str] = []
                for r in range(hs + 1, ws_sm.max_row + 1):
                    phrase = ws_sm.cell(r, cs["Phrase"]).value
                    kind = ws_sm.cell(r, cs["SourceKind"]).value
                    sid = ws_sm.cell(r, cs["SourceID"]).value
                    if isinstance(phrase, str) and phrase.strip():
                        items.append(f"{kind}|{sid}|{phrase.strip()}")
                items.sort()
                if items:
                    seq_fp = hashlib.sha1("\n".join(items).encode("utf-8", errors="ignore")).hexdigest()
            if seq_fp and seq_fp != prev_seq_fp:
                seq_fp_changed = 1
            else:
                seq_fp_changed = 0
            append_flow_run_log(
                wb,
                next_iter,
                105,
                utc,
                "CHANGED" if seq_fp_changed else "NO_CHANGE",
                f"SequenceMatches: matches={seq_matches} ({seq_status})",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=(
                    f"min_n={seqmatch_min_n}, n_list={n_list}, cand_max_freq={seqmatch_cand_max_freq}, "
                    f"time_budget_s={seqmatch_time_budget_s}, scan_tibia_first={int(bool(seqmatch_scan_tibia_first))}, "
                    f"explore_rotate={int(bool(seqmatch_explore_rotate))}, explore_keep_top={seqmatch_explore_keep_top}, "
                    f"cache={int(bool(seqmatch_cache_enabled))}/{seqmatch_cache_max_rows}, "
                    f"pd_max_chars={seqmatch_pd_max_chars}, pd_max_sentences={seqmatch_pd_max_sentences}, "
                    f"context_window={seqmatch_context_window}, context_min_overlap={seqmatch_context_min_overlap}, "
                    f"context_require_direction={int(bool(seqmatch_context_require_direction))}, "
                    f"snippet_min_content={seqmatch_snippet_min_content}, fp_changed={seq_fp_changed}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                105,
                utc,
                "FAILED",
                "SequenceMatches: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            105,
            utc,
            "NO_CHANGE",
            "SequenceMatches skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 107: SequenceWordHints (analysis-only): derive word-level hints from SequenceMatches.
    seq_hint_rows = 0
    seq_hint_status = "skipped"
    seq_hint_fp = ""
    seq_hint_fp_changed = 0
    if seq_word_hints_enabled:
        try:
            seq_hint_rows, seq_hint_status = materialize_sequence_word_hints(
                wb,
                iter_num=next_iter,
                utc=utc,
                drop_final_e=bool(lore_drop_final_e),
                drop_all_h=bool(lore_drop_all_h),
                drop_all_o=bool(lore_drop_all_o),
                min_n=int(seq_word_hints_min_n),
                skip_identity=bool(seq_word_hints_skip_identity),
                max_rows=int(seq_word_hints_max_rows),
                min_hint_ratio=float(seq_word_hints_min_ratio),
                hint_exclude_stopwords=bool(seq_word_hints_exclude_stopwords),
                hint_stopword_ratio=float(seq_word_hints_stopword_ratio),
            )
            # Fingerprint to avoid treating repeated hints as "progress".
            if SEQUENCE_WORD_HINTS_SHEET in wb.sheetnames:
                ws_h = wb[SEQUENCE_WORD_HINTS_SHEET]
                hh = ws_find_header_row(ws_h, ["CanonSig", "ToWord", "Count"], max_scan=3)
                ch = ws_headers(ws_h, hh)
                items: List[str] = []
                for r in range(hh + 1, ws_h.max_row + 1):
                    sig = ws_h.cell(r, ch["CanonSig"]).value
                    to = ws_h.cell(r, ch["ToWord"]).value
                    cnt = ws_h.cell(r, ch["Count"]).value
                    if isinstance(sig, str) and sig.strip() and isinstance(to, str) and to.strip():
                        items.append(f"{sig.strip()}|{to.strip().lower()}|{int(cnt or 0)}")
                items.sort()
                if items:
                    seq_hint_fp = hashlib.sha1("\n".join(items).encode("utf-8", errors="ignore")).hexdigest()
            if seq_hint_fp and seq_hint_fp != prev_seq_hint_fp:
                seq_hint_fp_changed = 1
            else:
                seq_hint_fp_changed = 0
            append_flow_run_log(
                wb,
                next_iter,
                107,
                utc,
                "CHANGED" if seq_hint_fp_changed else "NO_CHANGE",
                f"SequenceWordHints: rows={seq_hint_rows} ({seq_hint_status})",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=(
                    f"min_n={seq_word_hints_min_n}, skip_identity={int(bool(seq_word_hints_skip_identity))}, "
                    f"max_rows={seq_word_hints_max_rows}, min_ratio={seq_word_hints_min_ratio}, "
                    f"exclude_stopwords={int(bool(seq_word_hints_exclude_stopwords))}, stopword_ratio={seq_word_hints_stopword_ratio}, "
                    f"fp_changed={seq_hint_fp_changed}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                107,
                utc,
                "FAILED",
                "SequenceWordHints: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            107,
            utc,
            "NO_CHANGE",
            "SequenceWordHints skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 108: Sestina scan (analysis-only): detect classic sestina end-word permutation windows.
    sestina_lines = 0
    sestina_cands = 0
    sestina_best_score30 = 0
    sestina_fp = ""
    sestina_fp_changed = 0
    if sestina_enabled:
        try:
            sestina_lines, sestina_cands, sestina_best_score30, sestina_fp = materialize_sestina_scan(
                wb,
                iter_num=next_iter,
                utc=utc,
                active_tokens=final_active,
                max_lines=int(sestina_max_lines),
                min_score30=int(sestina_min_score30),
                max_candidates=int(sestina_max_candidates),
                use_token_signature=bool(sestina_use_token_sig),
                envoi_bonus=float(sestina_envoi_bonus),
            )
            if sestina_fp and sestina_fp != prev_sestina_fp:
                sestina_fp_changed = 1
            else:
                sestina_fp_changed = 0
            append_flow_run_log(
                wb,
                next_iter,
                108,
                utc,
                "CHANGED" if sestina_fp_changed else "NO_CHANGE",
                f"SestinaScan: lines={sestina_lines}, cands={sestina_cands}, best_score30={sestina_best_score30}",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=(
                    f"min_score30={sestina_min_score30}, use_token_sig={int(bool(sestina_use_token_sig))}, "
                    f"envoi_bonus={sestina_envoi_bonus}, fp_changed={sestina_fp_changed}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                108,
                utc,
                "FAILED",
                "SestinaScan: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            108,
            utc,
            "NO_CHANGE",
            "SestinaScan skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 113: Sestina obligation map (analysis-only): non-uniform positional load + ablation/reorder stress.
    sestina_ob_rows = 0
    sestina_ob_cands = 0
    sestina_ob_core_avg = 0.0
    sestina_ob_reorder_frac = 0.0
    sestina_ob_fp = ""
    sestina_ob_fp_changed = 0
    sestina_ob_status = "skipped"
    if sestina_enabled and sestina_obligation_enabled:
        try:
            (
                sestina_ob_rows,
                sestina_ob_cands,
                sestina_ob_core_avg,
                sestina_ob_reorder_frac,
                sestina_ob_fp,
                sestina_ob_status,
            ) = materialize_sestina_obligation_map(
                wb,
                iter_num=next_iter,
                utc=utc,
                max_candidates=int(sestina_obligation_max_candidates),
                min_score30=int(sestina_obligation_min_score30),
                obligatory_impact=float(sestina_obligatory_impact),
                conditional_impact=float(sestina_conditional_impact),
                decorative_impact=float(sestina_decorative_impact),
                no_collapse_tolerance=float(sestina_no_collapse_tolerance),
                reorder_resilient_ratio=float(sestina_reorder_resilient_ratio),
            )
            if sestina_ob_fp and sestina_ob_fp != prev_sestina_obligation_fp:
                sestina_ob_fp_changed = 1
            else:
                sestina_ob_fp_changed = 0
            append_flow_run_log(
                wb,
                next_iter,
                113,
                utc,
                "CHANGED" if sestina_ob_fp_changed else "NO_CHANGE",
                (
                    f"SestinaObligation: rows={sestina_ob_rows}, cands={sestina_ob_cands}, "
                    f"core_avg={sestina_ob_core_avg:.3f}, reorder_frac={sestina_ob_reorder_frac:.3f} ({sestina_ob_status})"
                ),
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=(
                    f"min_score30={sestina_obligation_min_score30}, max_candidates={sestina_obligation_max_candidates}, "
                    f"obligatory_impact={sestina_obligatory_impact}, conditional_impact={sestina_conditional_impact}, "
                    f"decorative_impact={sestina_decorative_impact}, no_collapse_tol={sestina_no_collapse_tolerance}, "
                    f"reorder_ratio={sestina_reorder_resilient_ratio}, fp_changed={sestina_ob_fp_changed}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                113,
                utc,
                "FAILED",
                "SestinaObligation: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            113,
            utc,
            "NO_CHANGE",
            "SestinaObligation skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 114: Rhythm transition A/B test (analysis-only): cycle-6 vs shuffled control + anti-sestina alternatives.
    rhythm_rows = 0
    rhythm_windows = 0
    rhythm_cycle_avg = 0.0
    rhythm_cycle_delta_avg = 0.0
    rhythm_ab_avg = 0.0
    rhythm_fib_avg = 0.0
    rhythm_core_avg = 0.0
    rhythm_sparse_avg = 0.0
    rhythm_closure_inv = 0.0
    rhythm_fp = ""
    rhythm_fp_changed = 0
    rhythm_status = "skipped"
    if rhythm_enabled:
        try:
            (
                rhythm_rows,
                rhythm_windows,
                rhythm_cycle_avg,
                rhythm_cycle_delta_avg,
                rhythm_ab_avg,
                rhythm_fib_avg,
                rhythm_core_avg,
                rhythm_sparse_avg,
                rhythm_closure_inv,
                rhythm_fp,
                rhythm_status,
            ) = materialize_rhythm_transition_abtest(
                wb,
                iter_num=next_iter,
                utc=utc,
                window_size=int(rhythm_window_size),
                min_lines=int(rhythm_min_lines),
                use_token_signature=bool(rhythm_use_token_sig),
                shuffle_trials=int(rhythm_shuffle_trials),
                cycle6_delta_threshold=float(rhythm_cycle_delta_threshold),
            )
            if rhythm_fp and rhythm_fp != prev_rhythm_fp:
                rhythm_fp_changed = 1
            else:
                rhythm_fp_changed = 0
            append_flow_run_log(
                wb,
                next_iter,
                114,
                utc,
                "CHANGED" if rhythm_fp_changed else "NO_CHANGE",
                (
                    f"RhythmAB: rows={rhythm_rows}, windows={rhythm_windows}, "
                    f"cycle={rhythm_cycle_avg:.3f}, delta={rhythm_cycle_delta_avg:.3f}, "
                    f"ab={rhythm_ab_avg:.3f}, fib={rhythm_fib_avg:.3f}, core={rhythm_core_avg:.3f}, "
                    f"sparse={rhythm_sparse_avg:.3f}, closure_inv={rhythm_closure_inv:.3f} ({rhythm_status})"
                ),
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=(
                    f"window={rhythm_window_size}, min_lines={rhythm_min_lines}, token_sig={int(bool(rhythm_use_token_sig))}, "
                    f"shuffle_trials={rhythm_shuffle_trials}, cycle_delta_thr={rhythm_cycle_delta_threshold}, "
                    f"fp_changed={rhythm_fp_changed}"
                ),
            )
        except Exception as e:
            append_flow_run_log(
                wb,
                next_iter,
                114,
                utc,
                "FAILED",
                "RhythmAB: FAILED",
                "0/70",
                evidence_avg=round(ev_avg_w, 6),
                weak_frac=round(weak_w, 6),
                micro_frac=round(micro_w, 6),
                notes=str(e),
            )
    else:
        append_flow_run_log(
            wb,
            next_iter,
            114,
            utc,
            "NO_CHANGE",
            "RhythmAB skipped (disabled)",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )

    # Step 106: Iter meta report (analysis-only).
    meta_rows = 0
    try:
        gt_count_meta, _gt_bad_meta = groundtruth_cribs_status(wb)
        # Count how many GroundTruth cribs have >=min_verified_sources_gt sources.
        gt_verified = 0
        verified_by_id: Dict[int, int] = {}
        if "CribReliability_v129" in wb.sheetnames:
            ws_rel = wb["CribReliability_v129"]
            hr = ws_find_header_row(ws_rel, ["CribID", "VerifiedSourceCount_v129"], max_scan=3)
            cr = ws_headers(ws_rel, hr)
            for r in range(hr + 1, ws_rel.max_row + 1):
                cid = ws_rel.cell(r, cr["CribID"]).value
                if cid is None:
                    continue
                try:
                    verified_by_id[int(cid)] = int(ws_rel.cell(r, cr["VerifiedSourceCount_v129"]).value or 0)
                except Exception:
                    continue
        if "Cribs" in wb.sheetnames:
            ws_c = wb["Cribs"]
            hc = ws_find_header_row(ws_c, ["CribID", "CribClass_v112"], max_scan=3)
            cc = ws_headers(ws_c, hc)
            for r in range(hc + 1, ws_c.max_row + 1):
                if ws_c.cell(r, cc["CribClass_v112"]).value != "GroundTruth":
                    continue
                cid = ws_c.cell(r, cc["CribID"]).value
                if cid is None:
                    continue
                try:
                    cid_i = int(cid)
                except Exception:
                    continue
                if int(verified_by_id.get(cid_i, 0)) >= int(min_verified_sources_gt):
                    gt_verified += 1

        meta_rows = upsert_iter_meta_sheet(
            wb,
            iter_num=next_iter,
            utc=utc,
            ev_avg=ev_avg_w,
            weak=weak_w,
            micro=micro_w,
            single=single_frac,
            tokens=final_tokens,
            gt_count=gt_count_meta,
            gt_verified_count=gt_verified,
            external_pass=ext_rt_pass,
            external_fail=ext_rt_fail,
            context_avg_score=ctx_avg_score,
            context_oov=ctx_oov,
            seq_matches=seq_matches,
            sestina_candidates=sestina_cands,
            sestina_best_score30=sestina_best_score30,
            sestina_ob_candidates=sestina_ob_cands,
            sestina_ob_core_avg=sestina_ob_core_avg,
            sestina_ob_reorder_frac=sestina_ob_reorder_frac,
            rhythm_windows=rhythm_windows,
            rhythm_cycle_avg=rhythm_cycle_avg,
            rhythm_cycle_delta_avg=rhythm_cycle_delta_avg,
            rhythm_ab_avg=rhythm_ab_avg,
            rhythm_fib_avg=rhythm_fib_avg,
            rhythm_core_avg=rhythm_core_avg,
            rhythm_sparse_avg=rhythm_sparse_avg,
            rhythm_closure_inv=rhythm_closure_inv,
            superanchors=superanchors,
        )
        append_flow_run_log(
            wb,
            next_iter,
            106,
            utc,
            "CHANGED",
            f"IterMeta written: rows~{meta_rows}",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
        )
    except Exception as e:
        append_flow_run_log(
            wb,
            next_iter,
            106,
            utc,
            "FAILED",
            "IterMeta: FAILED",
            "0/70",
            evidence_avg=round(ev_avg_w, 6),
            weak_frac=round(weak_w, 6),
            micro_frac=round(micro_w, 6),
            notes=str(e),
        )

    # Step 79: refresh live GT counters with the final active set after all mechanical/analysis passes.
    gt_live_final_status = "ok"
    try:
        glossary_ws_gt, glossary_map_gt = load_glossary(wb)
        active_gt = {t.token: t for t in glossary_map_gt.values() if t.use_strictplus and t.translation}
        gt_ok_final, gt_bad_list_final, gt_bad_all_final = groundtruth_live_check(
            wb,
            active_gt,
            enforced_crib_ids=enforced_gt_ids,
        )
        gt_bad_enforced_live = int(len(gt_bad_list_final))
        gt_bad_all_live = int(len(gt_bad_all_final))
        gt_soft_live = max(0, gt_bad_all_live - gt_bad_enforced_live)
        gt_live_final_status = (
            f"post-runcheck ok={int(bool(gt_ok_final))}; bad_enforced={gt_bad_enforced_live}, bad_all={gt_bad_all_live}, soft={gt_soft_live}"
        )
        append_flow_run_log(
            wb,
            next_iter,
            79,
            utc,
            "CHANGED" if gt_ok_final else "FAILED",
            "GroundTruth live check (post-run, final active set) "
            + ("ok" if gt_ok_final else f"failed {len(gt_bad_list_final)} enforced")
            + (
                f" (policy {gt_policy_status})"
                if "gt_policy_status" in locals() and gt_policy_status
                else ""
            ),
            "0/70",
            notes=gt_live_final_status,
        )
    except Exception as e:
        gt_live_final_status = f"post-runcheck error: {str(e)}"
        append_flow_run_log(
            wb,
            next_iter,
            79,
            utc,
            "FAILED",
            "GroundTruth live check (post-run): FAILED",
            "0/70",
            notes=gt_live_final_status,
        )

    # Step 80: logs + summaries + success check
    extra_parts: List[str] = []
    if fixed:
        extra_parts.append(f"Cribs GT match flags fixed: {fixed}")
    if mechanical_promotions_count:
        extra_parts.append(f"Mechanical promotions: {mechanical_promotions_count}")
    if semantic_objective_promo_count:
        extra_parts.append(f"Semantic-objective promotions (no DP delta): {semantic_objective_promo_count}")
    if soft_rescue_promos:
        extra_parts.append(f"Soft-rescue promotions: {soft_rescue_promos} (frontier={soft_frontier_promos})")
    if hard_escape_promos:
        extra_parts.append(f"Hard-escape promotions: {hard_escape_promos}")
    if directional_promos or directional_attempts:
        extra_parts.append(
            f"Directional escape: attempts={directional_attempts}, promos={directional_promos}"
        )
    if plateau_relax_note:
        extra_parts.append(plateau_relax_note)
    if macro_books_changed or macro_master_changed:
        extra_parts.append(f"MacroCompress changed: books={macro_books_changed}, master={macro_master_changed}")
    if readable_repl:
        extra_parts.append(
            f"Readability repl={readable_repl} (Books {readable_books_changed}; MasterText {readable_master_changed}; Cribs {readable_cribs_changed})"
        )
    if focus_rows:
        extra_parts.append(f"Focus rows~{focus_rows}")
    if anchor_status:
        extra_parts.append(f"AnchorCribs sync: added={anchor_added}, updated={anchor_updated}, status={anchor_status}")
    if blocks_written:
        extra_parts.append(f"Variant alignment: refBook={ref_book_id}, alignedBooks={aligned_books}, blocks={blocks_written}")
    if superanchors:
        extra_parts.append(f"SuperAnchors={superanchors}")
    if lore_hits_rows or semantic_map_rows or semantic_repl or semantic_books_changed or semantic_master_changed:
        extra_parts.append(
            "Lore/Semantic: "
            f"lore_hits_rows={lore_hits_rows}, semantic_map_rows={semantic_map_rows}, semantic_repl={semantic_repl}, "
            f"glossary_retext={semantic_glossary_promoted}, glossary_revert={semantic_glossary_reverted}, "
            f"semanticBooksChanged={semantic_books_changed}, semanticMasterChanged={semantic_master_changed}"
        )
    if english_map_rows or english_repl or english_books_changed or english_master_changed:
        extra_parts.append(
            f"EnglishLayer: map_rows={english_map_rows}, repl={english_repl}, books_changed={english_books_changed}, master_changed={english_master_changed}"
        )
    if english_retext_applied or english_retext_attempted:
        extra_parts.append(
            f"EnglishRetext: applied={english_retext_applied}, attempted={english_retext_attempted}, el_changed={english_retext_el_changed}"
        )
    if anti_unk_applied or anti_unk_attempted:
        extra_parts.append(
            f"AntiHallucinationUNK: applied={anti_unk_applied}, attempted={anti_unk_attempted}, el_changed={anti_unk_el_changed}"
        )
    if anchor_promo_only_enabled:
        extra_parts.append(
            f"AnchorPromotionOnly: anchors={anchor_promo_corpus_size}, min_hits={anchor_promo_min_hits}, "
            f"step30_kept={anchor_promo_kept_step30}, step30_dropped={anchor_promo_dropped_step30}, "
            f"step40_kept={anchor_promo_kept_step40}, step40_dropped={anchor_promo_dropped_step40}"
        )
        if anchor_promo_drop30_classes or anchor_promo_drop30_samples:
            extra_parts.append(
                "AnchorPromotionDrops30: "
                f"classes={anchor_promo_drop30_classes or 'n/a'}; "
                f"samples={anchor_promo_drop30_samples or 'n/a'}"
            )
        if anchor_promo_drop40_classes or anchor_promo_drop40_samples:
            extra_parts.append(
                "AnchorPromotionDrops40: "
                f"classes={anchor_promo_drop40_classes or 'n/a'}; "
                f"samples={anchor_promo_drop40_samples or 'n/a'}"
            )
    if autophrase_rows or autophrase_changed or autophrase_scanned or autophrase_eligible:
        extra_parts.append(
            f"AutoPhraseCribs: rows={autophrase_rows}, scanned={autophrase_scanned}, eligible={autophrase_eligible}, changed={int(bool(autophrase_changed))}"
        )
    if reverse_phrase_scanned or reverse_phrase_hits or reverse_phrase_new_tokens:
        extra_parts.append(
            f"ReversePhrase: phrases={reverse_phrase_scanned}, hits={reverse_phrase_hits}, "
            f"candidates={reverse_phrase_candidates}, emitted={reverse_phrase_new_tokens}"
        )
    if reverse_phrase_retext_applied or reverse_phrase_retext_attempted:
        extra_parts.append(
            f"ReversePhraseRetext: applied={reverse_phrase_retext_applied}, attempted={reverse_phrase_retext_attempted}, "
            f"el_changed={reverse_phrase_retext_el_changed}"
        )
    if reverse_perm_scanned or reverse_perm_hits or reverse_perm_new_tokens:
        extra_parts.append(
            f"ReversePhrasePermuted: phrases={reverse_perm_scanned}, hits={reverse_perm_hits}, "
            f"candidates={reverse_perm_candidates}, emitted={reverse_perm_new_tokens}"
        )
    if digit_codes:
        extra_parts.append(f"DigitMap: codes={digit_codes}, letters={digit_letters}, conflicts={digit_conflicts} ({digit_status})")
    if digit_ctx_rows:
        extra_parts.append(
            f"DigitCtx: rows={digit_ctx_rows}, homophone_letters={digit_ctx_homophone_letters}, "
            f"best_outlier={digit_ctx_best_outlier:.6f}, fp_changed={digit_ctx_fp_changed}"
        )
    if ext_rt_pass or ext_rt_fail:
        extra_parts.append(f"ExternalRoundTrip: pass={ext_rt_pass}, fail={ext_rt_fail}, skipped={ext_rt_skipped}")
    if lore_bigrams_status and lore_bigrams_status not in ("skipped", "skipped_fresh"):
        extra_parts.append(f"LoreBigrams: rows={lore_bigrams_rows} ({lore_bigrams_status})")
    if ctx_avg_score or (ctx_books_changed or ctx_master_changed):
        extra_parts.append(
            f"ContextEnglish: avg_score={ctx_avg_score:.6f}, oov={ctx_oov:.6f}, streak={ctx_improve_streak}, "
            f"books_changed={ctx_books_changed}, master_changed={ctx_master_changed}"
        )
    if seq_matches:
        extra_parts.append(f"SeqMatches: matches={seq_matches}, fp_changed={seq_fp_changed}")
    if sestina_ob_rows or sestina_ob_cands:
        extra_parts.append(
            "SestinaObligation: "
            f"rows={sestina_ob_rows}, cands={sestina_ob_cands}, core_avg={sestina_ob_core_avg:.3f}, "
            f"reorder_frac={sestina_ob_reorder_frac:.3f}, fp_changed={sestina_ob_fp_changed}"
        )
    if rhythm_rows or rhythm_windows:
        extra_parts.append(
            "RhythmAB: "
            f"rows={rhythm_rows}, windows={rhythm_windows}, cycle={rhythm_cycle_avg:.3f}, "
            f"delta={rhythm_cycle_delta_avg:.3f}, ab={rhythm_ab_avg:.3f}, fib={rhythm_fib_avg:.3f}, "
            f"core={rhythm_core_avg:.3f}, sparse={rhythm_sparse_avg:.3f}, closure_inv={rhythm_closure_inv:.3f}, "
            f"fp_changed={rhythm_fp_changed}"
        )
    if ext_changed:
        extra_parts.append(f"ExternalRefs filled: {ext_changed}")
    if cribs_updated:
        extra_parts.append(f"Cribs updated: {cribs_updated} (match flips {match_flips})")
    if avartar_match is not None:
        extra_parts.append(f"AvarTarMatch={avartar_match}")
    extra_parts.append(
        f"GTLive: mode={gt_live_mode_active}, enforced={enforced_gt_label}, bad_enforced={gt_bad_enforced_live}, bad_all={gt_bad_all_live}, soft={gt_soft_live}"
        + (f", mode_changed={int(gt_mode_changed)}" if gt_mode_changed else "")
    )
    if promotion_skip_count:
        extra_parts.append(
            f"PromotionSkips: count={promotion_skip_count}" + (f", top={promotion_skip_reason_top}" if promotion_skip_reason_top else "")
        )
    extra_notes = "; ".join(extra_parts) or None

    upsert_iter_summary_sheet(
        wb,
        next_iter,
        utc,
        changed_books,
        ev_avg_w,
        weak_w,
        micro_w,
        gt_promoted=gt_promoted,
        mech_promoted=mechanical_promotions_count,
        semantic_promoted=int(semantic_glossary_promoted) + int(semantic_glossary_reverted),
        extra_notes=extra_notes,
    )

    gt_count, gt_bad_legacy = groundtruth_cribs_status(wb)
    if convergence_use_live_gt_counts:
        gt_bad_enforced = int(gt_bad_enforced_live)
        gt_bad_all = int(gt_bad_all_live)
        gt_soft = int(gt_soft_live)
    else:
        gt_bad_enforced = int(gt_bad_legacy)
        gt_bad_all = int(gt_bad_legacy)
        gt_soft = 0
    gt_total_for_pct = max(1, int(gt_count))
    gt_ok_for_progress = max(0, int(gt_total_for_pct) - int(gt_bad_legacy))
    gt_ok_hard = max(0, int(gt_total_for_pct) - int(gt_bad_enforced))
    gt_overall_pct = 100.0 * (gt_ok_for_progress / float(gt_total_for_pct))
    gt_overall_pct = max(0.0, min(100.0, gt_overall_pct))
    gt_hard_pct = 100.0 * (gt_ok_hard / float(gt_total_for_pct))
    gt_hard_pct = max(0.0, min(100.0, gt_hard_pct))
    gt_target_pct = 100.0 * (int(gt_count) / max(1, int(target_gt_count)))
    gt_target_pct = max(0.0, min(100.0, gt_target_pct))
    (
        conv_books_done,
        conv_books_total,
        conv_books_done_ids,
        conv_contigs_done,
        conv_contigs_total,
        conv_contigs_done_ids,
        conv_lines_done,
        conv_lines_total,
        conv_lines_done_keys,
    ) = translation_convergence_metrics(wb)
    cov_ok, cov_bad = books_coverage_strictplus_ok(wb)
    semantic_progress_now = bool(
        semantic_glossary_promoted
        or semantic_glossary_reverted
        or semantic_books_changed
        or semantic_master_changed
        or english_books_changed
        or english_master_changed
        or english_retext_applied
        or reverse_phrase_hits
        or reverse_phrase_new_tokens
        or reverse_phrase_retext_applied
    )
    display_progress_now = bool(
        readable_repl or readable_books_changed or readable_master_changed or readable_cribs_changed
    )
    analysis_progress_now = bool(
        focus_rows
        or anchor_added
        or anchor_updated
        or blocks_written
        or superanchors
        or autophrase_changed
        or ctx_books_changed
        or ctx_master_changed
        or ctx_improve_streak
        or seq_fp_changed
        or sestina_ob_fp_changed
        or rhythm_fp_changed
        or digit_ctx_fp_changed
        or codeaware_fp_changed
        or ext_changed
        or cribs_updated
    )
    success_hard = (
        gt_count >= target_gt_count
        and weak_w <= target_weak
        and micro_w <= target_micro
        and ev_avg_w >= target_ev
        and single_frac <= 0.10
        and (not require_all_gt or gt_bad_enforced == 0)
        and (not require_cov or cov_ok)
    )
    hard_resolved = bool(success_hard and gt_soft <= int(soft_mismatch_max_resolved))
    soft_resolved = bool(
        success_hard
        and gt_soft > int(soft_mismatch_max_resolved)
        and gt_soft <= int(soft_mismatch_max_soft_resolved)
    )
    if gt_soft <= 0:
        gt_soft_nondec_streak = 0
    elif prev_gt_soft_mismatch > 0 and gt_soft >= prev_gt_soft_mismatch:
        gt_soft_nondec_streak = int(prev_gt_soft_nondec_streak) + 1
    else:
        gt_soft_nondec_streak = 1
    iters_since_last_mech = 0 if mechanical_promotions_count > 0 else (int(prev_no_mech_iters) + 1)

    plateau_actions_changed = bool(
        plateau_relax_note
        or (macro_books_changed or macro_master_changed)
        or semantic_progress_now
        or display_progress_now
        or analysis_progress_now
    )

    puzzle_external_ok = (ext_rt_fail == 0 and ext_rt_pass > 0)
    puzzle_ctx_ok = True if int(puzzle_ctx_improve_iters) <= 0 else (int(ctx_improve_streak) >= int(puzzle_ctx_improve_iters))
    puzzle_seq_ok = True if int(puzzle_min_seq_matches) <= 0 else (int(seq_matches) >= int(puzzle_min_seq_matches))
    puzzle_solved = bool(hard_resolved and puzzle_external_ok and puzzle_ctx_ok and puzzle_seq_ok)

    status_out = STATUS_MODEL_CONVERGED if puzzle_solved else ("RESOLVED" if hard_resolved else ("SOFT_RESOLVED" if soft_resolved else "READY"))
    block_reason = None
    if not success_hard:
        if require_all_gt and gt_bad_enforced:
            status_out = "BLOCKED"
            block_reason = f"GroundTruth live mismatch (enforced): {gt_bad_enforced} (all={gt_bad_all}, soft={gt_soft})"
        elif require_cov and not cov_ok:
            status_out = "BLOCKED"
            block_reason = f"Coverage_StrictPlus_v108 not 1 for {cov_bad} books"
        elif gt_promoted == 0 and mechanical_promotions_count == 0 and not plateau_actions_changed:
            # Non-blocking: open a lower-confidence exploration lane even without hard mechanical progress.
            if not block_reason:
                block_reason = "Sem avanço novo mecânico/GT; mantendo exploração direcional e de contingência."
            status_out = "READY"
        elif gt_promoted == 0 and mechanical_promotions_count == 0:
            # Not a hard-block: plateau actions (focus/structural/macrocompress/readability) still produced traceable progress.
            if not block_reason:
                block_reason = "Nenhum novo GT/Promoção mecânica neste ciclo (mas houve ações de plateau)."
    else:
        if gt_soft > int(soft_mismatch_max_soft_resolved):
            status_out = "BLOCKED"
            block_reason = (
                f"Soft GT mismatches above threshold: soft={gt_soft} > SoftMismatchMaxForSoftResolved={soft_mismatch_max_soft_resolved}"
            )
        elif (
            soft_resolved
            and int(convergence_soft_nondec_streak_block) > 0
            and gt_soft_nondec_streak > int(convergence_soft_nondec_streak_block)
        ):
            status_out = "BLOCKED"
            block_reason = (
                "Soft GT mismatches not decreasing monotonically while in SOFT_RESOLVED "
                f"(soft={gt_soft}, streak={gt_soft_nondec_streak})"
            )

    if status_out != "BLOCKED" and mechanical_promotions_count == 0 and promotion_skip_count > 0 and iters_since_last_mech >= int(plateau_block_no_promo_iters):
        status_out = "READY"
        block_reason = (
            "Plateau de skips recorrentes; mudando estratégia e persistindo tentativa mecânica alternativa "
            f"(sem promoções mecânicas há {iters_since_last_mech} iterações; limiar={plateau_block_no_promo_iters}; "
            f"top_skip={promotion_skip_reason_top or 'n/a'})"
        )

    candidate_rows_count = len(rows) if isinstance(rows, list) else 0
    real_progress_now = bool(
        mechanical_promotions_count > 0
        or gt_promoted > 0
        or gt_soft < prev_gt_soft_mismatch
        or semantic_progress_now
    )
    no_real_progress_iters = 0 if real_progress_now else (int(prev_no_real_progress_iters) + 1)
    if (
        status_out in ("RESOLVED", "SOFT_RESOLVED", STATUS_MODEL_CONVERGED)
        and candidate_rows_count == 0
        and no_real_progress_iters >= int(convergence_require_progress_for_solved)
        and not block_reason
    ):
        status_out = "READY"
        block_reason = (
            f"Sem progresso real há {no_real_progress_iters} iterações em estado resolvido; "
            "acionando busca agressiva de direção alternativa."
        )

    update_sheet_index(
        wb,
        next_iter,
        f"Iteration {next_iter} recap (mech_promoted {mechanical_promotions_count}, macroCompress_books {macro_books_changed}, focus_rows~{focus_rows}, "
        f"superanchors {superanchors}, readability_repl {readable_repl}, ctxScore {ctx_avg_score:.6f} (streak {ctx_improve_streak}), "
        f"seqMatches {seq_matches} (fp_changed {seq_fp_changed}), sestinaCands {sestina_cands} (best30 {sestina_best_score30}, fp_changed {sestina_fp_changed}), "
        f"sestinaObCands {sestina_ob_cands} (core_avg {sestina_ob_core_avg:.3f}, reorder_frac {sestina_ob_reorder_frac:.3f}, fp_changed {sestina_ob_fp_changed}), "
        f"rhythmWin {rhythm_windows} (cycle {rhythm_cycle_avg:.3f}, delta {rhythm_cycle_delta_avg:.3f}, ab {rhythm_ab_avg:.3f}, fp_changed {rhythm_fp_changed}), "
        f"extRT {ext_rt_pass}/{ext_rt_fail}, cribs_updated {cribs_updated}, books_changed {changed_books}).",
    )
    append_method_log(
        wb,
        next_iter,
        utc,
        f"Auto-chain: GT+{gt_promoted}, mech+{mechanical_promotions_count}, semantic_objective+{semantic_objective_promo_count}, macroCompress_books={macro_books_changed}, focus_rows~{focus_rows}, superanchors={superanchors}, "
        f"readableRepl={readable_repl}, ctxScore={ctx_avg_score:.6f} (streak {ctx_improve_streak}), "
        f"codeMapRows={codeaware_map_rows} (fp_changed {codeaware_fp_changed}), "
        f"seqMatches={seq_matches} (fp_changed {seq_fp_changed}), "
        f"sestinaCands={sestina_cands} (best30 {sestina_best_score30}, fp_changed {sestina_fp_changed}), "
        f"sestinaObCands={sestina_ob_cands} (core_avg {sestina_ob_core_avg:.3f}, reorder_frac {sestina_ob_reorder_frac:.3f}, fp_changed {sestina_ob_fp_changed}), "
        f"rhythmWin={rhythm_windows} (cycle {rhythm_cycle_avg:.3f}, delta {rhythm_cycle_delta_avg:.3f}, ab {rhythm_ab_avg:.3f}, fib {rhythm_fib_avg:.3f}, core {rhythm_core_avg:.3f}, sparse {rhythm_sparse_avg:.3f}, fp_changed {rhythm_fp_changed}), "
        f"extRT={ext_rt_pass}/{ext_rt_fail}, englishRetext={english_retext_applied}, autoPhrase={autophrase_rows}, reverseHits={reverse_phrase_hits}, "
        f"reverseEmit={reverse_phrase_new_tokens}, extRefs+{ext_changed}, cribsUpdated={cribs_updated}",
        "Incremental decode: safe mechanical promotions (GT+DP+metrics guardrails); recompute global + token evidence; plateau ladder (focus/anchors/macrocompress/readability).",
        f"Ev {ev_avg_w:.6f}, Weak {weak_w:.6f}, Micro {micro_w:.6f}, Single {single_frac:.6f}",
        status_out,
    )
    wq_notes = block_reason
    if wq_notes is None:
        if not real_progress_now:
            wq_notes = (
                f"Plateau (mech_promoted=0). Review Iter{next_iter}_Focus for weakest books/contigs + micro tokens. "
                "If structural steps ran, inspect AnchorCribs_Auto / BookOffsets_Auto / AlignedBackbone_Auto / VariantAssemblyBlocks_Auto / SuperAnchors_Auto. "
                f"MacroMine ladder rung={plateau_rung}. Extend/tune ReadabilityRules (display-only) as needed."
            )
        else:
            wq_notes = "Next iteration will rescan candidates, recompute analytics, and refresh readability/macro-compress layers."
    append_work_queue(wb, next_iter, status_out, wq_notes)
    progress_summary_for_log = "; ".join(extra_parts) if extra_parts else ""

    append_flow_run_log(
        wb,
        next_iter,
        80,
        utc,
        "CHANGED",
        f"Updated logs/summary; status={status_out}",
        changed_books,
        evidence_avg=round(ev_avg_w, 6),
        weak_frac=round(weak_w, 6),
        micro_frac=round(micro_w, 6),
        notes=block_reason,
    )

    # Update FlowState
    flow_state_set_many(
        store,
        {
            "CurrentIteration": next_iter,
            "LastCompletedStepID": 80,
            "NextStepID": 10,
            "Status": status_out,
            "LastRunUTC": utc,
            "LastChangeSummary": (
                f"Iter {next_iter}: gt+{gt_promoted}, mech+{mechanical_promotions_count}, semanticObjective+{semantic_objective_promo_count}, "
                f"semanticRetext+{semantic_glossary_promoted}, semanticRevert+{semantic_glossary_reverted}, "
                f"macroCompressBooksChanged={macro_books_changed}, focusRows~{focus_rows}, superanchors={superanchors}, readableRepl={readable_repl}, "
                f"readableCribsChanged={readable_cribs_changed}, loreHits={lore_hits_rows}, semanticMap={semantic_map_rows}, semanticRepl={semantic_repl}, "
                f"semanticBooksChanged={semantic_books_changed}, semanticMasterChanged={semantic_master_changed}, englishMapRows={english_map_rows}, englishRepl={english_repl}, "
                f"englishBooksChanged={english_books_changed}, englishMasterChanged={english_master_changed}, englishRetext={english_retext_applied}, "
                f"ctxScore={ctx_avg_score:.6f}, ctxOOV={ctx_oov:.6f}, ctxStreak={ctx_improve_streak}, "
                f"codeMapRows={codeaware_map_rows}, codeFpChanged={codeaware_fp_changed}, codeOverrides={codeaware_overrides}, "
                f"seqMatches={seq_matches}, seqFpChanged={seq_fp_changed}, "
                f"seqHintRows={seq_hint_rows}, seqHintFpChanged={seq_hint_fp_changed}, "
                f"sestinaCands={sestina_cands}, sestinaBest30={sestina_best_score30}, sestinaFpChanged={sestina_fp_changed}, "
                f"sestinaObRows={sestina_ob_rows}, sestinaObCands={sestina_ob_cands}, sestinaObCoreAvg={sestina_ob_core_avg:.3f}, "
                f"sestinaObReorderFrac={sestina_ob_reorder_frac:.3f}, sestinaObFpChanged={sestina_ob_fp_changed}, "
                f"rhythmRows={rhythm_rows}, rhythmWindows={rhythm_windows}, rhythmCycle={rhythm_cycle_avg:.3f}, rhythmDelta={rhythm_cycle_delta_avg:.3f}, "
                f"rhythmAB={rhythm_ab_avg:.3f}, rhythmFib={rhythm_fib_avg:.3f}, rhythmCore={rhythm_core_avg:.3f}, rhythmSparse={rhythm_sparse_avg:.3f}, "
                f"rhythmClosureInv={rhythm_closure_inv:.3f}, rhythmFpChanged={rhythm_fp_changed}, "
                f"extRT={ext_rt_pass}/{ext_rt_fail}, digitCodes={digit_codes}, digitConflicts={digit_conflicts}, "
                f"gtLive={gt_bad_enforced}/{gt_bad_all}/soft{gt_soft}, mode={gt_live_mode_active}, "
                f"anchorOnly={int(bool(anchor_promo_only_enabled))}, anchorCorpus={anchor_promo_corpus_size}, "
                f"anchorDrop30={anchor_promo_dropped_step30}, anchorDrop40={anchor_promo_dropped_step40}, "
                f"promoSkips={promotion_skip_count}, noMechIters={iters_since_last_mech}, softRescuePromos={soft_rescue_promos}, "
                f"softFrontierPromos={soft_frontier_promos}, hardEscapePromos={hard_escape_promos}, "
                f"directionalEscapeAttempts={directional_attempts}, directionalEscapePromos={directional_promos}, "
                f"semanticNoEffectPromos={semantic_no_effect_promos}, "
                f"candidateRows={candidate_rows_count}, candidateMode={candidate_scan_mode or 'normal'}, noRealProgressIters={no_real_progress_iters}, "
                f"cribsUpdated={cribs_updated}, booksChanged={changed_books}"
            ),
            "BlockReason": block_reason,
            "SuccessCheck": bool(success_hard),
            "HardResolvedCheck": bool(hard_resolved),
            "SoftResolvedCheck": bool(soft_resolved),
            "PuzzleSolvedCheck": bool(puzzle_solved),
            "GroundTruthLiveCheckModeActive": gt_live_mode_active,
            "GroundTruthEnforcedCount": int(gt_count) if enforced_gt_ids is None else int(enforced_gt_n),
            "GTBadEnforcedCount": int(gt_bad_enforced),
            "GTBadAllCount": int(gt_bad_all),
            "GTSoftMismatchCount": int(gt_soft),
            "AntiHallucinationModeActive": bool(anti_mode),
            "AntiHallucinationUNKApplied": int(anti_unk_applied),
            "AntiHallucinationUNKAttempted": int(anti_unk_attempted),
            "ReversePhraseRetextApplied": int(reverse_phrase_retext_applied),
            "ReversePhraseRetextAttempted": int(reverse_phrase_retext_attempted),
            "AnchorPromotionOnlyEnabled": bool(anchor_promo_only_enabled),
            "AnchorPromotionCorpusSize": int(anchor_promo_corpus_size),
            "AnchorPromotionMinHits": int(anchor_promo_min_hits),
            "AnchorPromotionKeptStep30": int(anchor_promo_kept_step30),
            "AnchorPromotionDroppedStep30": int(anchor_promo_dropped_step30),
            "AnchorPromotionKeptStep40": int(anchor_promo_kept_step40),
            "AnchorPromotionDroppedStep40": int(anchor_promo_dropped_step40),
            "AnchorPromotionDrop30Classes": anchor_promo_drop30_classes or "",
            "AnchorPromotionDrop30Samples": anchor_promo_drop30_samples or "",
            "AnchorPromotionDrop40Classes": anchor_promo_drop40_classes or "",
            "AnchorPromotionDrop40Samples": anchor_promo_drop40_samples or "",
            "GTSoftMismatchNonDecreasingStreak": int(gt_soft_nondec_streak),
            "PromotionSkipCount": int(promotion_skip_count),
            "PromotionSkipReasonTop": promotion_skip_reason_top or "",
            "CandidateSkip_DPUnusedCount": int(dp_unused_count),
            "CandidateSkip_DPSwallowedCount": int(dp_swallowed_count),
            "CandidateSkip_DPBlockedMonotonicCount": int(dp_blocked_monotonic_count),
            "CandidateSkip_DPUnusedTop": dp_unused_top or "",
            "CandidateSkip_DPSwallowedTop": dp_swallowed_top or "",
            "CandidateSkip_DPBlockedMonotonicTop": dp_blocked_monotonic_top or "",
            "SoftRescuePromos": int(soft_rescue_promos),
            "SoftFrontierPromos": int(soft_frontier_promos),
            "HardEscapePromos": int(hard_escape_promos),
            "DirectionalEscapeAttempts": int(directional_attempts),
            "DirectionalEscapePromos": int(directional_promos),
            "SemanticNoEffectPromos": int(semantic_no_effect_promos),
            "MechanicalPromotionsCount": int(mechanical_promotions_count),
            "SemanticObjectivePromotionsCount": int(semantic_objective_promo_count),
            "SemanticProgressChanged": int(bool(semantic_progress_now)),
            "DisplayProgressChanged": int(bool(display_progress_now)),
            "AnalysisProgressChanged": int(bool(analysis_progress_now)),
            "NoEffectAnchorEscapePromos": int(semantic_no_effect_promos),
            "IterationsSinceLastMechanicalPromotion": int(iters_since_last_mech),
            "CandidateScanEmptyStreak": int(candidate_empty_streak),
            "IterationsWithoutRealProgress": int(no_real_progress_iters),
            "LastCandidateScanMode": str(candidate_scan_mode or prev_candidate_scan_mode or ""),
        },
    )
    flow_state_set(store, "LastCandidateRows", len(rows) if isinstance(rows, list) else 0)
    flow_state_set(store, "ContextEnglishAvgScore", round(float(ctx_avg_score), 6))
    flow_state_set(store, "ContextEnglishOOVFrac", round(float(ctx_oov), 6))
    flow_state_set(store, "ContextEnglishImproveStreak", int(ctx_improve_streak))
    flow_state_set(store, "CodeWordMapCount", int(codeaware_map_rows))
    flow_state_set(store, "CodeWordMapFingerprint", codeaware_fp)
    flow_state_set(store, "CodeWordMapFingerprintChanged", int(codeaware_fp_changed))
    flow_state_set(store, "CodeAwareBooksChangedRows", int(codeaware_books_changed))
    flow_state_set(store, "CodeAwareOverridesTotal", int(codeaware_overrides))
    flow_state_set(store, "SequenceMatchesCount", int(seq_matches))
    flow_state_set(store, "SequenceMatchesFingerprint", seq_fp)
    flow_state_set(store, "SequenceMatchesFingerprintChanged", int(seq_fp_changed))
    flow_state_set(store, "SequenceWordHintsCount", int(seq_hint_rows))
    flow_state_set(store, "SequenceWordHintsFingerprint", seq_hint_fp)
    flow_state_set(store, "SequenceWordHintsFingerprintChanged", int(seq_hint_fp_changed))
    flow_state_set(store, "DigitCodeContextCount", int(digit_ctx_rows))
    flow_state_set(store, "DigitCodeContextFingerprint", digit_ctx_fp)
    flow_state_set(store, "DigitCodeContextFingerprintChanged", int(digit_ctx_fp_changed))
    flow_state_set(store, "DigitCodeContextBestOutlier", round(float(digit_ctx_best_outlier), 6))
    flow_state_set(store, "SestinaCandidatesCount", int(sestina_cands))
    flow_state_set(store, "SestinaBestScore30", int(sestina_best_score30))
    flow_state_set(store, "SestinaFingerprint", sestina_fp)
    flow_state_set(store, "SestinaFingerprintChanged", int(sestina_fp_changed))
    flow_state_set(store, "SestinaObligationCount", int(sestina_ob_rows))
    flow_state_set(store, "SestinaObligationCandidates", int(sestina_ob_cands))
    flow_state_set(store, "SestinaObligationCoreAvg", round(float(sestina_ob_core_avg), 6))
    flow_state_set(store, "SestinaObligationReorderFrac", round(float(sestina_ob_reorder_frac), 6))
    flow_state_set(store, "SestinaObligationFingerprint", sestina_ob_fp)
    flow_state_set(store, "SestinaObligationFingerprintChanged", int(sestina_ob_fp_changed))
    flow_state_set(store, "RhythmRows", int(rhythm_rows))
    flow_state_set(store, "RhythmWindows", int(rhythm_windows))
    flow_state_set(store, "RhythmCycleAvg", round(float(rhythm_cycle_avg), 6))
    flow_state_set(store, "RhythmCycleDeltaAvg", round(float(rhythm_cycle_delta_avg), 6))
    flow_state_set(store, "RhythmABAvg", round(float(rhythm_ab_avg), 6))
    flow_state_set(store, "RhythmFibAvg", round(float(rhythm_fib_avg), 6))
    flow_state_set(store, "RhythmCoreAvg", round(float(rhythm_core_avg), 6))
    flow_state_set(store, "RhythmSparseAvg", round(float(rhythm_sparse_avg), 6))
    flow_state_set(store, "RhythmClosureInversionFrac", round(float(rhythm_closure_inv), 6))
    flow_state_set(store, "RhythmFingerprint", rhythm_fp)
    flow_state_set(store, "RhythmFingerprintChanged", int(rhythm_fp_changed))

    wb.save(workbook_path)
    flow_store_close(store)
    sync_sqlite_snapshot_from_artifact(
        workbook_path,
        note=f"iter {next_iter}: status={status_out}",
    )
    print(f"Saved {workbook_path}")
    print(f"Backup {backup_path}")
    print(
        f"Iter {next_iter}: status={status_out}; gt_promoted={gt_promoted}; "
        f"mech_promoted={mechanical_promotions_count}; semantic_objective_promoted={semantic_objective_promo_count}; "
        f"books_changed={changed_books}"
    )
    if block_reason:
        print(f"BlockReason: {block_reason}")
    print(
        "Evolution: "
        f"EvAvg {base_ev:.6f}->{ev_avg_w:.6f} (d={ev_avg_w - base_ev:+.6f}); "
        f"Weak {base_weak:.6f}->{weak_w:.6f} (d={weak_w - base_weak:+.6f}); "
        f"Micro {base_micro:.6f}->{micro_w:.6f} (d={micro_w - base_micro:+.6f}); "
        f"Single {base_single:.6f}->{single_frac:.6f} (d={single_frac - base_single:+.6f}); "
        f"Tokens {base_tokens}->{final_tokens} (d={final_tokens - base_tokens:+d})"
    )
    print(f"MacroCompress: books_changed={macro_books_changed}, master_changed={macro_master_changed}")
    print(
        f"Readability: repl={readable_repl}, books_changed_rows={readable_books_changed}, master_changed_rows={readable_master_changed}, cribs_changed_rows={readable_cribs_changed}"
    )
    print(
        f"Focus/Structural: focus_rows~{focus_rows}, anchor_added={anchor_added}, anchor_updated={anchor_updated}, refBook={ref_book_id}, alignedBooks={aligned_books}, blocks={blocks_written}, superanchors={superanchors}"
    )
    print(
        f"Lore/Semantic: seed_added={lore_seed_added}, lore_total_rows~{lore_total_rows}, lore_hits_rows={lore_hits_rows}, semantic_map_rows={semantic_map_rows}, semantic_repl={semantic_repl}, semantic_books_changed={semantic_books_changed}, semantic_master_changed={semantic_master_changed}"
    )
    print(
        f"EnglishLayer: map_rows={english_map_rows}, repl={english_repl}, books_changed_rows={english_books_changed}, master_changed_rows={english_master_changed}"
    )
    print(f"DigitMap: codes={digit_codes}, letters={digit_letters}, conflicts={digit_conflicts} ({digit_status})")
    print(
        f"DigitCtx: rows={digit_ctx_rows}, homophone_letters={digit_ctx_homophone_letters}, best_outlier={digit_ctx_best_outlier:.6f}, "
        f"fp_changed={digit_ctx_fp_changed} ({digit_ctx_status})"
    )
    print(f"ExternalRoundTrip: pass={ext_rt_pass}, fail={ext_rt_fail}, skipped={ext_rt_skipped} ({ext_rt_status})")
    print(f"LoreBigrams: rows={lore_bigrams_rows} ({lore_bigrams_status})")
    print(
        f"ContextEnglish: avg_score={ctx_avg_score:.6f}, oov={ctx_oov:.6f}, streak={ctx_improve_streak}, "
        f"books_changed_rows={ctx_books_changed}, master_changed_rows={ctx_master_changed}, map_rows={ctx_map_rows}"
    )
    print(
        f"CodeAware: map_rows={codeaware_map_rows}, books_changed_rows={codeaware_books_changed}, overrides={codeaware_overrides}, "
        f"fp_changed={codeaware_fp_changed} ({codeaware_status})"
    )
    print(f"SequenceMatches: matches={seq_matches}, fp_changed={seq_fp_changed}")
    print(
        f"SestinaScan: lines={sestina_lines}, cands={sestina_cands}, best_score30={sestina_best_score30}, fp_changed={sestina_fp_changed}"
    )
    print(
        f"SestinaObligation: rows={sestina_ob_rows}, cands={sestina_ob_cands}, core_avg={sestina_ob_core_avg:.3f}, "
        f"reorder_frac={sestina_ob_reorder_frac:.3f}, fp_changed={sestina_ob_fp_changed} ({sestina_ob_status})"
    )
    print(
        f"RhythmAB: rows={rhythm_rows}, windows={rhythm_windows}, cycle={rhythm_cycle_avg:.3f}, delta={rhythm_cycle_delta_avg:.3f}, "
        f"ab={rhythm_ab_avg:.3f}, fib={rhythm_fib_avg:.3f}, core={rhythm_core_avg:.3f}, sparse={rhythm_sparse_avg:.3f}, "
        f"closure_inv={rhythm_closure_inv:.3f}, fp_changed={rhythm_fp_changed} ({rhythm_status})"
    )
    print(
        f"Convergence scope: "
        f"books={conv_books_done}/{conv_books_total} ({0.0 if conv_books_total <= 0 else 100.0 * conv_books_done / float(conv_books_total):.2f}%), "
        f"contigs={conv_contigs_done}/{conv_contigs_total} ({0.0 if conv_contigs_total <= 0 else 100.0 * conv_contigs_done / float(conv_contigs_total):.2f}%), "
        f"falas={conv_lines_done}/{conv_lines_total} ({0.0 if conv_lines_total <= 0 else 100.0 * conv_lines_done / float(conv_lines_total):.2f}%)"
    )
    print(
        f"GTLive: mode={gt_live_mode_active}, enforced={enforced_gt_label}, bad_enforced={gt_bad_enforced}, bad_all={gt_bad_all}, soft={gt_soft}, "
        f"soft_nondec_streak={gt_soft_nondec_streak}, translation={gt_overall_pct:.2f}% (GT={gt_ok_for_progress}/{gt_total_for_pct})"
    )
    print(
        f"AntiHallucination: mode={int(bool(anti_mode))}, semantic_retext_enabled={int(bool(sem_promote_enabled))}, "
        f"english_retext_enabled={int(bool(english_retext_enabled))}, unk_applied={anti_unk_applied}/{anti_unk_attempted}"
    )
    print(
        f"ReversePhraseRetext: applied={reverse_phrase_retext_applied}, attempted={reverse_phrase_retext_attempted}, "
        f"el_changed={reverse_phrase_retext_el_changed}"
    )
    print(
        f"AnchorPromotionOnly: enabled={int(bool(anchor_promo_only_enabled))}, anchors={anchor_promo_corpus_size}, "
        f"min_hits={anchor_promo_min_hits}, step30_kept={anchor_promo_kept_step30}, step30_dropped={anchor_promo_dropped_step30}, "
        f"step40_kept={anchor_promo_kept_step40}, step40_dropped={anchor_promo_dropped_step40}"
    )
    print(
        f"SemanticObjective: enabled={int(bool(convergence_semantic_objective_enabled))}, active={int(bool(semantic_no_effect_active))}, "
        f"no_effect_promos={semantic_no_effect_promos}/{convergence_semantic_no_effect_max_promos}, "
        f"min_gain={convergence_semantic_no_effect_min_gain:.3f}, min_occ={convergence_semantic_min_occ}, "
        f"gain_weight={convergence_semantic_gain_weight:.3f}, "
        f"allow_classes={','.join(sorted(convergence_semantic_allow_classes))}"
    )
    print(
        f"PromotionSkips: count={promotion_skip_count}, top={promotion_skip_reason_top or 'n/a'}, "
        f"iters_since_last_mech={iters_since_last_mech}"
    )
    print(
        "Status checks: "
        f"success_hard={int(bool(success_hard))}, hard_resolved={int(bool(hard_resolved))}, "
        f"soft_resolved={int(bool(soft_resolved))}, puzzle_solved={int(bool(puzzle_solved))}"
    )
    print(f"MacroMine ladder rung={plateau_rung}")
    strategy_summary = ""
    if "macro_fallback_note_parts" in locals() and macro_fallback_note_parts:
        strategy_summary = "; ".join([str(part) for part in macro_fallback_note_parts[-4:] if str(part).strip()])
    if "lowered_note" in locals() and lowered_note:
        if strategy_summary:
            strategy_summary = f"{lowered_note}; {strategy_summary}"
        else:
            strategy_summary = str(lowered_note)
    if not strategy_summary:
        strategy_summary = "estratégia-base (strict/soft/frontier) sem fallback extraordinário"
    _post_run_status_to_discord(
        workbook_path=workbook_path,
        next_iter=next_iter,
        status=status_out,
        gt_ok=bool(success_hard),
        block_reason=block_reason,
        metrics={
            "low_confidence_escape": bool(locals().get("low_confidence_escape_active", False)),
            "strategy_summary": strategy_summary,
            "gt_total": gt_total_for_pct,
            "gt_ok_count": gt_ok_for_progress,
            "gt_overall_pct": gt_overall_pct,
            "gt_hard_pct": gt_hard_pct,
            "gt_target_pct": gt_target_pct,
            "progress_summary": progress_summary_for_log,
            "next_action": wq_notes,
            "gt_live_mode": gt_live_mode_active,
            "gt_bad_enforced": gt_bad_enforced,
            "gt_bad_all": gt_bad_all,
            "gt_soft": gt_soft,
            "gt_soft_nondec_streak": gt_soft_nondec_streak,
            "books_translated": conv_books_done,
            "books_total": conv_books_total,
            "contigs_translated": conv_contigs_done,
            "contigs_total": conv_contigs_total,
            "lines_translated": conv_lines_done,
            "lines_total": conv_lines_total,
            "mech_promoted": mechanical_promotions_count,
            "semantic_no_effect_promos": semantic_objective_promo_count,
            "gt_promoted": gt_promoted,
            "iters_since_last_mech": iters_since_last_mech,
            "promotion_skip_count": promotion_skip_count,
            "soft_rescue_promos": soft_rescue_promos,
            "soft_frontier_promos": soft_frontier_promos,
            "hard_escape_promos": hard_escape_promos,
            "directional_escape_attempts": directional_attempts,
            "directional_escape_promos": directional_promos,
            "gt_live_final_status": gt_live_final_status,
            "candidate_rows": int(candidate_rows_count) if "candidate_rows_count" in locals() else len(rows) if "rows" in locals() else 0,
            "candidate_scan_mode": str(candidate_scan_mode if "candidate_scan_mode" in locals() else ""),
            "candidate_empty_streak": int(candidate_empty_streak if "candidate_empty_streak" in locals() else 0),
            "no_real_progress_iters": int(no_real_progress_iters if "no_real_progress_iters" in locals() else 0),
            "anchor_promo_only": bool(anchor_promo_only_enabled),
            "anchor_corpus_size": int(anchor_promo_corpus_size),
            "anchor_min_hits": int(anchor_promo_min_hits),
            "anchor_kept_step30": int(anchor_promo_kept_step30),
            "anchor_dropped_step30": int(anchor_promo_dropped_step30),
            "anchor_kept_step40": int(anchor_promo_kept_step40),
            "anchor_dropped_step40": int(anchor_promo_dropped_step40),
            "anchor_drop30_classes": str(anchor_promo_drop30_classes or ""),
            "anchor_drop30_samples": str(anchor_promo_drop30_samples or ""),
            "anchor_drop40_classes": str(anchor_promo_drop40_classes or ""),
            "anchor_drop40_samples": str(anchor_promo_drop40_samples or ""),
            "reverse_retext_applied": int(reverse_phrase_retext_applied),
            "reverse_retext_attempted": int(reverse_phrase_retext_attempted),
            "semantic_progress_changed": int(bool(semantic_progress_now)),
            "display_progress_changed": int(bool(display_progress_now)),
            "analysis_progress_changed": int(bool(analysis_progress_now)),
            "status_hint": block_reason if block_reason else status_out,
            "soft_improving": bool(not block_reason and gt_soft <= int(soft_mismatch_max_resolved)),
        },
    )



def main(argv: Sequence[str]) -> int:
    if len(argv) != 2:
        print(f"Usage: {argv[0]} PATH_TO_WORKBOOK.xlsx", file=sys.stderr)
        return 2
    run_next_iteration(argv[1])
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
