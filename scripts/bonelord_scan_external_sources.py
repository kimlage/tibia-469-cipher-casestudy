#!/usr/bin/env python3
"""
Scan community-source URLs listed in workbook and index numeric sequences.

Purpose:
- keep analysis-only trace of external community evidence
- extract candidate digit runs from discovered links
- cross-check if those runs already appear in Books.Digits
- flag exact matches against ExternalRefs_v115
"""

from __future__ import annotations

import html
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from typing import Dict, Iterable, List, Sequence, Tuple

import openpyxl


SRC_SHEET = "ExternalCommunitySources_v472"
OUT_SHEET = "ExternalSourceDigitHits_v472"
BOOKS_SHEET = "Books"
EXTREFS_SHEET = "ExternalRefs_v115"


def _digits_only(s: object) -> str:
    return "".join(ch for ch in str(s or "") if ch.isdigit())


def _fetch_text(url: str, timeout_s: int = 25) -> Tuple[bool, str]:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
            )
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout_s) as resp:
            raw = resp.read()
            ctype = str(resp.headers.get("Content-Type", ""))
            charset = "utf-8"
            m = re.search(r"charset=([\w\-]+)", ctype, flags=re.I)
            if m:
                charset = m.group(1)
            txt = raw.decode(charset, errors="replace")
            return True, txt
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError):
        return False, ""
    except Exception:
        return False, ""


def _html_to_text(s: str) -> str:
    s2 = re.sub(r"(?is)<script.*?>.*?</script>", " ", s)
    s2 = re.sub(r"(?is)<style.*?>.*?</style>", " ", s2)
    s2 = re.sub(r"(?is)<[^>]+>", " ", s2)
    s2 = html.unescape(s2)
    s2 = re.sub(r"\\s+", " ", s2).strip()
    return s2


def _ws_find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, required: Sequence[str], max_scan: int = 10) -> int:
    req = set(required)
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        got = {str(v).strip() for v in row_vals if isinstance(v, str)}
        if req.issubset(got):
            return r
    raise ValueError(f"header not found in sheet {ws.title}: {required}")


def _ws_headers(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            out[v.strip()] = c
    return out


def _ensure_sheet(wb: openpyxl.Workbook, name: str, headers: Sequence[str]) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in wb.sheetnames:
        ws = wb[name]
        hr = _ws_find_header_row(ws, list(headers), max_scan=3)
        if ws.max_row > hr:
            ws.delete_rows(hr + 1, ws.max_row - hr)
        return ws
    ws = wb.create_sheet(name)
    ws.append(list(headers))
    return ws


def _unique_runs(text: str, min_len: int = 8) -> List[str]:
    runs = re.findall(r"\d+", text or "")
    out: List[str] = []
    seen = set()
    for r in runs:
        if len(r) < min_len:
            continue
        if r in seen:
            continue
        seen.add(r)
        out.append(r)
    return out


def _load_books_digits(wb: openpyxl.Workbook) -> Dict[int, str]:
    ws = wb[BOOKS_SHEET]
    h = _ws_find_header_row(ws, ["BookID", "Digits"], max_scan=3)
    c = _ws_headers(ws, h)
    out: Dict[int, str] = {}
    for r in range(h + 1, ws.max_row + 1):
        bid = ws.cell(r, c["BookID"]).value
        if bid is None:
            continue
        try:
            bid_i = int(bid)
        except Exception:
            continue
        out[bid_i] = _digits_only(ws.cell(r, c["Digits"]).value)
    return out


def _load_extref_digits(wb: openpyxl.Workbook) -> Dict[str, str]:
    ws = wb[EXTREFS_SHEET]
    h = _ws_find_header_row(ws, ["RefName", "DigitsSanitized", "NumericText"], max_scan=3)
    c = _ws_headers(ws, h)
    out: Dict[str, str] = {}
    for r in range(h + 1, ws.max_row + 1):
        ref = ws.cell(r, c["RefName"]).value
        if not isinstance(ref, str) or not ref.strip():
            continue
        ds_col = c.get("DigitsSanitized")
        ds = ws.cell(r, ds_col).value if ds_col is not None else None
        digits = _digits_only(ds) if ds is not None else _digits_only(ws.cell(r, c["NumericText"]).value)
        if digits:
            out[ref.strip()] = digits
    return out


def scan(workbook_path: str) -> Tuple[int, int, int]:
    wb = openpyxl.load_workbook(workbook_path)
    if SRC_SHEET not in wb.sheetnames:
        raise SystemExit(f"Missing required sheet: {SRC_SHEET}")

    ws_src = wb[SRC_SHEET]
    hs = _ws_find_header_row(ws_src, ["SourceID", "URL", "RecommendedAction"], max_scan=3)
    cs = _ws_headers(ws_src, hs)

    books_digits = _load_books_digits(wb)
    extref_digits = _load_extref_digits(wb)
    ref_by_digits: Dict[str, List[str]] = defaultdict(list)
    for ref, ds in extref_digits.items():
        ref_by_digits[ds].append(ref)

    ws_out = _ensure_sheet(
        wb,
        OUT_SHEET,
        [
            "Iteration",
            "UTC",
            "SourceID",
            "URL",
            "FetchOK",
            "Culture",
            "Language",
            "Type",
            "RecommendedAction",
            "DigitsRun",
            "DigitsLen",
            "InBooksCount",
            "InBooksBookIDs",
            "ExactRefMatch",
            "HitKind",
            "Notes",
        ],
    )
    hr = _ws_find_header_row(ws_out, ["Iteration", "SourceID", "DigitsRun"], max_scan=3)
    co = _ws_headers(ws_out, hr)

    # Determine current iteration best-effort from FlowState.
    iter_num = 0
    if "FlowState" in wb.sheetnames:
        ws_state = wb["FlowState"]
        for r in range(2, ws_state.max_row + 1):
            k = ws_state.cell(r, 1).value
            if k == "CurrentIteration":
                try:
                    iter_num = int(ws_state.cell(r, 2).value or 0)
                except Exception:
                    iter_num = 0
                break
    iter_num += 1
    utc = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    rr = hr + 1
    source_count = 0
    fetch_ok_count = 0
    hit_rows = 0

    for r in range(hs + 1, ws_src.max_row + 1):
        sid = ws_src.cell(r, cs["SourceID"]).value
        url = ws_src.cell(r, cs["URL"]).value
        if not isinstance(sid, str) or not sid.strip() or not isinstance(url, str) or not url.strip():
            continue
        source_count += 1
        sid_s = sid.strip()
        url_s = url.strip()
        culture = ws_src.cell(r, cs.get("Culture", 0)).value if cs.get("Culture") else None
        lang = ws_src.cell(r, cs.get("Language", 0)).value if cs.get("Language") else None
        typ = ws_src.cell(r, cs.get("Type", 0)).value if cs.get("Type") else None
        rec = ws_src.cell(r, cs.get("RecommendedAction", 0)).value if cs.get("RecommendedAction") else None
        seq_hint = ws_src.cell(r, cs.get("SequencesMentioned", 0)).value if cs.get("SequencesMentioned") else None

        ok, payload = _fetch_text(url_s)
        if ok:
            fetch_ok_count += 1
        text = _html_to_text(payload) if ok else ""
        runs = _unique_runs(text, min_len=8)
        hint_runs = _unique_runs(str(seq_hint or ""), min_len=4)
        if not runs and hint_runs:
            runs = hint_runs
        if not runs:
            # keep an explicit row for traceability even when no run is extracted.
            ws_out.cell(rr, co["Iteration"]).value = iter_num
            ws_out.cell(rr, co["UTC"]).value = utc
            ws_out.cell(rr, co["SourceID"]).value = sid_s
            ws_out.cell(rr, co["URL"]).value = url_s
            ws_out.cell(rr, co["FetchOK"]).value = bool(ok)
            ws_out.cell(rr, co["Culture"]).value = culture
            ws_out.cell(rr, co["Language"]).value = lang
            ws_out.cell(rr, co["Type"]).value = typ
            ws_out.cell(rr, co["RecommendedAction"]).value = rec
            ws_out.cell(rr, co["DigitsRun"]).value = None
            ws_out.cell(rr, co["DigitsLen"]).value = 0
            ws_out.cell(rr, co["InBooksCount"]).value = 0
            ws_out.cell(rr, co["InBooksBookIDs"]).value = None
            ws_out.cell(rr, co["ExactRefMatch"]).value = None
            ws_out.cell(rr, co["HitKind"]).value = "NONE"
            ws_out.cell(rr, co["Notes"]).value = "no digit runs >=8 found (html + SequencesMentioned)"
            rr += 1
            continue

        for run in runs[:80]:
            hit_books = [bid for bid, ds in books_digits.items() if run in ds]
            exact_refs = ref_by_digits.get(run, [])
            kind = "EXACT_REF" if exact_refs else ("IN_BOOKS" if hit_books else "EXTERNAL_ONLY")
            note = "from_html"
            if run in hint_runs and run not in _unique_runs(text, min_len=8):
                note = "from_sequences_mentioned"

            ws_out.cell(rr, co["Iteration"]).value = iter_num
            ws_out.cell(rr, co["UTC"]).value = utc
            ws_out.cell(rr, co["SourceID"]).value = sid_s
            ws_out.cell(rr, co["URL"]).value = url_s
            ws_out.cell(rr, co["FetchOK"]).value = bool(ok)
            ws_out.cell(rr, co["Culture"]).value = culture
            ws_out.cell(rr, co["Language"]).value = lang
            ws_out.cell(rr, co["Type"]).value = typ
            ws_out.cell(rr, co["RecommendedAction"]).value = rec
            ws_out.cell(rr, co["DigitsRun"]).value = run
            ws_out.cell(rr, co["DigitsLen"]).value = len(run)
            ws_out.cell(rr, co["InBooksCount"]).value = len(hit_books)
            ws_out.cell(rr, co["InBooksBookIDs"]).value = ",".join(str(x) for x in sorted(hit_books)) if hit_books else None
            ws_out.cell(rr, co["ExactRefMatch"]).value = "|".join(sorted(exact_refs)) if exact_refs else None
            ws_out.cell(rr, co["HitKind"]).value = kind
            ws_out.cell(rr, co["Notes"]).value = note
            rr += 1
            hit_rows += 1

    wb.save(workbook_path)
    return source_count, fetch_ok_count, hit_rows


def main(argv: Sequence[str]) -> int:
    if len(argv) != 2:
        print(f"Usage: {argv[0]} <workbook.xlsx>")
        return 2
    src, ok, hits = scan(argv[1])
    print(f"Scanned sources={src}, fetch_ok={ok}, digit_hit_rows={hits}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
