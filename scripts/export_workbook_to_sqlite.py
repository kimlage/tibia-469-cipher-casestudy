#!/usr/bin/env python3
"""
Ingest a legacy spreadsheet snapshot into a local SQLite database.

Goals:
- Preserve the source artifact exactly enough to avoid information loss.
- Provide convenient SQL tables for analysis across sheets.
- Support repeated exports into the same DB as new snapshots.

This script creates:
- `exports`: one row per ingested snapshot
- `sheets`: sheet-level metadata
- `sheet_columns`: detected header/column metadata
- `cells`: lossless-ish cell dump with raw and cached values
- `row_json`: JSON row payloads for each non-empty row
- `sheet__<name>` convenience tables with one TEXT column per detected header
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sqlite3
from dataclasses import dataclass
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.utils import get_column_letter


HEADER_SCAN_ROWS = 10


@dataclass
class SheetHeader:
    header_row: int
    raw_names: List[str]
    normalized_names: List[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest a legacy spreadsheet snapshot into SQLite")
    parser.add_argument("artifact", help="Path to the legacy spreadsheet snapshot")
    parser.add_argument(
        "--db",
        default="./data/bonelord_workbook.sqlite",
        help="Output SQLite DB path",
    )
    parser.add_argument(
        "--replace-db",
        action="store_true",
        help="Delete the destination DB first",
    )
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def py_type_name(value: object) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int) and not isinstance(value, bool):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    return type(value).__name__


def to_text(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def cell_formula(cell: Cell | ReadOnlyCell) -> Optional[str]:
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def is_row_empty(values: Sequence[object]) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def detect_header(ws) -> SheetHeader:
    best_row = 1
    best_values: List[str] = []
    best_score = -1
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, HEADER_SCAN_ROWS), values_only=True),
        start=1,
    ):
        values = ["" if v is None else str(v).strip() for v in row]
        non_empty = [v for v in values if v]
        score = len(non_empty)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_values = values
    normalized = normalize_headers(best_values or [f"col_{i+1}" for i in range(ws.max_column)])
    return SheetHeader(best_row, best_values, normalized)


def normalize_identifier(text: str, fallback: str) -> str:
    value = (text or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        value = fallback
    if value[0].isdigit():
        value = f"c_{value}"
    return value


def normalize_headers(values: Sequence[object]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for idx, raw in enumerate(values, start=1):
        name = normalize_identifier("" if raw is None else str(raw), f"col_{idx}")
        seen[name] = seen.get(name, 0) + 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        out.append(name)
    return out


def ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;
        PRAGMA synchronous=NORMAL;

        CREATE TABLE IF NOT EXISTS exports (
            export_id INTEGER PRIMARY KEY AUTOINCREMENT,
            workbook_path TEXT NOT NULL,
            workbook_name TEXT NOT NULL,
            workbook_sha256 TEXT NOT NULL,
            workbook_size INTEGER NOT NULL,
            workbook_mtime REAL NOT NULL,
            artifact_path TEXT,
            artifact_name TEXT,
            artifact_sha256 TEXT,
            artifact_size INTEGER,
            artifact_mtime REAL,
            artifact_kind TEXT,
            exported_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS sheets (
            export_id INTEGER NOT NULL,
            sheet_index INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            max_row INTEGER NOT NULL,
            max_col INTEGER NOT NULL,
            header_row INTEGER NOT NULL,
            PRIMARY KEY (export_id, sheet_name)
        );

        CREATE TABLE IF NOT EXISTS sheet_columns (
            export_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            column_index INTEGER NOT NULL,
            column_letter TEXT NOT NULL,
            raw_name TEXT,
            normalized_name TEXT NOT NULL,
            PRIMARY KEY (export_id, sheet_name, column_index)
        );

        CREATE TABLE IF NOT EXISTS cells (
            export_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            column_index INTEGER NOT NULL,
            column_letter TEXT NOT NULL,
            coordinate TEXT NOT NULL,
            data_type TEXT,
            python_type TEXT,
            raw_value_text TEXT,
            cached_value_text TEXT,
            formula_text TEXT,
            number_format TEXT,
            PRIMARY KEY (export_id, sheet_name, row_index, column_index)
        );

        CREATE TABLE IF NOT EXISTS row_json (
            export_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            row_json TEXT NOT NULL,
            PRIMARY KEY (export_id, sheet_name, row_index)
        );

        CREATE INDEX IF NOT EXISTS idx_cells_sheet_row ON cells (sheet_name, row_index);
        CREATE INDEX IF NOT EXISTS idx_cells_sheet_col ON cells (sheet_name, column_index);
        CREATE INDEX IF NOT EXISTS idx_row_json_sheet_row ON row_json (sheet_name, row_index);
        """
    )
    ensure_export_alias_columns(conn)


def ensure_export_alias_columns(conn: sqlite3.Connection) -> None:
    cols = {row[1] for row in conn.execute("PRAGMA table_info(exports)").fetchall()}
    additions = []
    if "artifact_path" not in cols:
        additions.append(("artifact_path", "TEXT"))
    if "artifact_name" not in cols:
        additions.append(("artifact_name", "TEXT"))
    if "artifact_sha256" not in cols:
        additions.append(("artifact_sha256", "TEXT"))
    if "artifact_size" not in cols:
        additions.append(("artifact_size", "INTEGER"))
    if "artifact_mtime" not in cols:
        additions.append(("artifact_mtime", "REAL"))
    if "artifact_kind" not in cols:
        additions.append(("artifact_kind", "TEXT"))
    for name, dtype in additions:
        conn.execute(f'ALTER TABLE exports ADD COLUMN "{name}" {dtype}')
    if additions:
        conn.commit()
    conn.execute(
        """
        UPDATE exports
        SET artifact_path = COALESCE(artifact_path, workbook_path),
            artifact_name = COALESCE(artifact_name, workbook_name),
            artifact_sha256 = COALESCE(artifact_sha256, workbook_sha256),
            artifact_size = COALESCE(artifact_size, workbook_size),
            artifact_mtime = COALESCE(artifact_mtime, workbook_mtime),
            artifact_kind = COALESCE(artifact_kind, 'legacy_xlsx')
        """
    )
    conn.commit()


def ensure_sheet_table(conn: sqlite3.Connection, sheet_name: str, normalized_names: Sequence[str]) -> str:
    table_name = f"sheet__{normalize_identifier(sheet_name, 'sheet')}"
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS "{table_name}" (
            __export_id INTEGER NOT NULL,
            __row_index INTEGER NOT NULL,
            __sheet_name TEXT NOT NULL,
            PRIMARY KEY (__export_id, __row_index)
        )
        """
    )
    existing = {
        row[1]
        for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    }
    for col in normalized_names:
        if col not in existing:
            conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{col}" TEXT')
    return table_name


def insert_export(conn: sqlite3.Connection, artifact_path: Path) -> int:
    stat = artifact_path.stat()
    digest = sha256_file(artifact_path)
    cur = conn.execute(
        """
        INSERT INTO exports (
            workbook_path, workbook_name, workbook_sha256, workbook_size, workbook_mtime,
            artifact_path, artifact_name, artifact_sha256, artifact_size, artifact_mtime, artifact_kind,
            exported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(artifact_path),
            artifact_path.name,
            digest,
            stat.st_size,
            stat.st_mtime,
            str(artifact_path),
            artifact_path.name,
            digest,
            stat.st_size,
            stat.st_mtime,
            "legacy_xlsx",
            datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        ),
    )
    return int(cur.lastrowid)


def existing_export_id(conn: sqlite3.Connection, artifact_sha256: str) -> Optional[int]:
    row = conn.execute(
        """
        SELECT export_id
        FROM exports
        WHERE COALESCE(artifact_sha256, workbook_sha256) = ?
        ORDER BY export_id DESC
        LIMIT 1
        """,
        (artifact_sha256,),
    ).fetchone()
    if not row:
        return None
    return int(row[0])


def export_sheet(
    conn: sqlite3.Connection,
    export_id: int,
    sheet_index: int,
    raw_ws,
    cached_ws,
) -> Tuple[int, int]:
    header = detect_header(raw_ws)
    conn.execute(
        """
        INSERT INTO sheets (export_id, sheet_index, sheet_name, max_row, max_col, header_row)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (export_id, sheet_index, raw_ws.title, raw_ws.max_row, raw_ws.max_column, header.header_row),
    )

    for idx, (raw_name, normalized_name) in enumerate(zip(header.raw_names, header.normalized_names), start=1):
        conn.execute(
            """
            INSERT INTO sheet_columns (
                export_id, sheet_name, column_index, column_letter, raw_name, normalized_name
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (export_id, raw_ws.title, idx, get_column_letter(idx), raw_name, normalized_name),
        )

    table_name = ensure_sheet_table(conn, raw_ws.title, header.normalized_names)

    cell_rows = 0
    data_rows = 0
    raw_iter = raw_ws.iter_rows()
    cached_iter = cached_ws.iter_rows()
    for row_index, (raw_row, cached_row) in enumerate(zip(raw_iter, cached_iter), start=1):
        row_payload: Dict[str, Optional[str]] = {}
        convenience_values: Dict[str, Optional[str]] = {}
        row_has_value = False
        for column_index, (raw_cell, cached_cell) in enumerate(zip(raw_row, cached_row), start=1):
            raw_text = to_text(raw_cell.value)
            cached_text = to_text(cached_cell.value)
            formula_text = cell_formula(raw_cell)
            number_format = getattr(raw_cell, "number_format", None)
            if raw_text is not None or cached_text is not None or formula_text is not None:
                row_has_value = True
                conn.execute(
                    """
                    INSERT INTO cells (
                        export_id, sheet_name, row_index, column_index, column_letter, coordinate,
                        data_type, python_type, raw_value_text, cached_value_text, formula_text, number_format
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        export_id,
                        raw_ws.title,
                        row_index,
                        column_index,
                        get_column_letter(column_index),
                        raw_cell.coordinate,
                        getattr(raw_cell, "data_type", None),
                        py_type_name(raw_cell.value),
                        raw_text,
                        cached_text,
                        formula_text,
                        number_format,
                    ),
                )
                cell_rows += 1

            if column_index <= len(header.normalized_names):
                key = header.normalized_names[column_index - 1]
                value_text = cached_text if cached_text is not None else raw_text
                row_payload[key] = value_text
                convenience_values[key] = value_text

        if row_has_value:
            conn.execute(
                """
                INSERT INTO row_json (export_id, sheet_name, row_index, row_json)
                VALUES (?, ?, ?, ?)
                """,
                (
                    export_id,
                    raw_ws.title,
                    row_index,
                    json.dumps(row_payload, ensure_ascii=True, sort_keys=False),
                ),
            )

        if row_index > header.header_row and any(v is not None and str(v).strip() != "" for v in convenience_values.values()):
            cols = ['__export_id', '__row_index', '__sheet_name'] + [f'"{name}"' for name in header.normalized_names]
            vals = [export_id, row_index, raw_ws.title] + [convenience_values.get(name) for name in header.normalized_names]
            placeholders = ", ".join(["?"] * len(vals))
            conn.execute(
                f'INSERT INTO "{table_name}" ({", ".join(cols)}) VALUES ({placeholders})',
                vals,
            )
            data_rows += 1

    return cell_rows, data_rows


def main() -> int:
    args = parse_args()
    artifact_path = Path(args.artifact).resolve()
    db_path = Path(args.db).resolve()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if args.replace_db and db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(str(db_path))
    try:
        ensure_schema(conn)
        digest = sha256_file(artifact_path)
        already = existing_export_id(conn, digest)
        if already is not None:
            print(
                json.dumps(
                    {
                        "db_path": str(db_path),
                        "export_id": already,
                        "artifact_path": str(artifact_path),
                        "skipped": "duplicate_sha256",
                    },
                    ensure_ascii=True,
                )
            )
            return 0
        export_id = insert_export(conn, artifact_path)

        raw_wb = load_workbook(artifact_path, data_only=False, read_only=True)
        cached_wb = load_workbook(artifact_path, data_only=True, read_only=True)

        total_cells = 0
        total_rows = 0
        for sheet_index, sheet_name in enumerate(raw_wb.sheetnames, start=1):
            raw_ws = raw_wb[sheet_name]
            cached_ws = cached_wb[sheet_name]
            cell_rows, data_rows = export_sheet(conn, export_id, sheet_index, raw_ws, cached_ws)
            total_cells += cell_rows
            total_rows += data_rows
            conn.commit()

        print(
            json.dumps(
                {
                    "db_path": str(db_path),
                    "export_id": export_id,
                    "artifact_path": str(artifact_path),
                    "sheet_count": len(raw_wb.sheetnames),
                    "data_rows": total_rows,
                    "cell_rows": total_cells,
                },
                ensure_ascii=True,
            )
        )
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
